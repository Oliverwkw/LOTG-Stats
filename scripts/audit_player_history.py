"""Audit player-history continuity (roster lineage).

Every player's Sleeper-style history (the hover-comment built in lotg.py and
surfaced on the player_all_time / picks sheets) should be an unbroken chain of
roster events: a player can only be *dropped* or *traded away* by a team that
currently holds them, and can only be *added off free agency / waivers* when no
team holds them. A break in that chain means a step is missing — the player
"teleports" onto or off of a roster.

This script reconstructs each player's holder timeline from the history text
embedded as cell comments in `exports/LOTG_Stats.xlsx` and reports every break:

  MISSING_ARRIVAL_BEFORE_DROP  dropped by a team that wasn't holding the player
  MISSING_ARRIVAL_BEFORE_TRADE traded away a player the team didn't hold
  MISSING_DROP                 picked up off FA while another team still held them

Exit code is non-zero when any break is found, so it can gate CI.

Usage: python3 scripts/audit_player_history.py [path/to/LOTG_Stats.xlsx]
"""
from __future__ import annotations

import re
import sys
from collections import Counter
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent

_RE_ADD = re.compile(r"^(\d{4}-\d{2}-\d{2}): added by (\S+) \((?:free agent|waiver \$\d+)")
_RE_TRADE = re.compile(r"^(\d{4}-\d{2}-\d{2}): traded to (\S+) \(.*\)$")
_RE_DROP = re.compile(r"^(\d{4}-\d{2}-\d{2}): (?:dropped|released) by (\S+)")
# Draft-arrival line. Matches the plain "YYYY Draft:"/"YYYY draft:" form and any
# draft-descriptor prefix before "draft:" — e.g. "YYYY supplemental veteran draft:"
# (the 2021 vet draft) and the legacy "YYYY startup (vet) draft:" wording.
_RE_DRAFT = re.compile(r"^(\d{4}) (?:[\w() ]+ )?[Dd]raft: (\S+) ")
_RE_HDR = re.compile(r"^(\d{4}).* — originally (\S+)'s pick")
_RE_CMOVE = re.compile(r"^(\d{4}): Commissioner moved to (\S+)$")
_RE_PICKHOP = re.compile(r"^(\d{4}-\d{2}-\d{2}): pick traded to ")


def load_history_comments(xlsx_path: Path) -> dict[str, str]:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=False)
    out: dict[str, str] = {}
    for sheet in ("player_all_time", "picks"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for row in ws.iter_rows():
            for c in row:
                if c.comment and c.column == 1:
                    key = str(c.value)
                    # player_all_time keys by player name; keep the first seen.
                    out.setdefault(f"{sheet}:{key}", c.comment.text)
    return out


def audit_text(name: str, txt: str) -> list[tuple]:
    """Return a list of (name, date, kind, detail, line) breaks for one history."""
    breaks = []
    holder = None
    for line in txt.splitlines():
        line = line.strip()
        if not line:
            continue
        m = _RE_ADD.match(line)
        if m:
            d, team = m.group(1), m.group(2)
            if holder is not None and holder != team:
                breaks.append((name, d, "MISSING_DROP",
                               f"FA/waiver add by {team} while still held by {holder}", line))
            holder = team
            continue
        m = _RE_TRADE.match(line)
        if m:
            d, team = m.groups()
            if holder is None:
                breaks.append((name, d, "MISSING_ARRIVAL_BEFORE_TRADE",
                               f"traded to {team} but not on any roster", line))
            holder = team
            continue
        m = _RE_DROP.match(line)
        if m:
            d, team = m.group(1), m.group(2)
            if holder != team:
                breaks.append((name, d, "MISSING_ARRIVAL_BEFORE_DROP",
                               f"dropped by {team} but held by {holder}", line))
            holder = None
            continue
        m = _RE_DRAFT.match(line)
        if m:
            d, team = m.groups()
            holder = team
            continue
        if _RE_HDR.match(line) or _RE_CMOVE.match(line) or _RE_PICKHOP.match(line):
            continue
        # Unrecognized line — surface so the parser stays honest as text evolves.
        breaks.append((name, "", "UNPARSED", "history line not recognized", line))
    return breaks


def main() -> int:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else (REPO / "exports" / "LOTG_Stats.xlsx")
    if not xlsx.exists():
        print(f"no xlsx at {xlsx}", file=sys.stderr)
        return 2
    comments = load_history_comments(xlsx)
    # De-dup: the same player appears on both player_all_time and (as a drafted
    # pick) picks — audit the player_all_time copy, fall back to picks.
    seen_player: dict[str, str] = {}
    for k, v in comments.items():
        sheet, name = k.split(":", 1)
        if sheet == "player_all_time":
            seen_player[name] = v
    for k, v in comments.items():
        sheet, name = k.split(":", 1)
        if sheet == "picks":
            seen_player.setdefault(name, v)

    all_breaks = []
    for name, txt in seen_player.items():
        all_breaks.extend(audit_text(name, txt))

    real = [b for b in all_breaks if b[2] != "UNPARSED"]
    unparsed = [b for b in all_breaks if b[2] == "UNPARSED"]

    print(f"players audited: {len(seen_player)}")
    print(f"continuity breaks: {len(real)}  {dict(Counter(b[2] for b in real))}")
    if unparsed:
        print(f"unparsed lines: {len(unparsed)} (parser may need updating)")
    print("=" * 70)
    for b in sorted(real):
        print(f"{b[0]:26} {b[1]} {b[2]:30} | {b[4]}")

    return 1 if real else 0


if __name__ == "__main__":
    sys.exit(main())
