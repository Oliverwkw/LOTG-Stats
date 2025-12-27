
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Any, Tuple, Optional
from datetime import datetime, timezone, timedelta, date
from collections import Counter, deque, defaultdict
import json
import math
import re
import traceback
import logging
import warnings
import numpy as np

warnings.filterwarnings("ignore", category=FutureWarning, message="Downcasting object dtype arrays")

LOG = logging.getLogger("lotg")
if not LOG.handlers:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")


import pandas as pd


def _round_from_label(label: Any) -> str:
    s = str(label or "").strip().lower()
    if s in ("semifinal", "semifinals"):
        return "semifinals"
    if s in ("final", "finals"):
        return "Finals"
    if s in ("toilet semis", "toilet semifinal", "toilet semifinals"):
        return "Toilet Semis"
    if s in ("toilet final", "toilet finals"):
        return "toilet finals"
    if s in ("3rd place", "third place"):
        return "3rd place game"
    if s in ("toilet trash", "trash"):
        return "toilet trash"
    return "regular season"
# ----------------------------
# DataFrame safety helpers
# ----------------------------
def ensure_cols(df: pd.DataFrame, cols, default=None):
    """Ensure columns exist; if missing, create with default."""
    for c in cols:
        if c not in df.columns:
            df[c] = default
    return df

def to_num_series(s, default=0.0):
    """Robust numeric coercion for series/array-like; returns float series."""
    if s is None:
        return pd.Series([], dtype='float64')
    out = pd.to_numeric(s, errors='coerce')
    if isinstance(out, pd.Series):
        return out.fillna(default)
    # scalar
    return pd.Series([out if pd.notna(out) else default], dtype='float64')

def safe_to_numeric(df: pd.DataFrame, col: str, default=0.0):
    """Convert df[col] to numeric if present; otherwise create with default."""
    if col not in df.columns:
        df[col] = default
    else:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(default)
    return df

def as_bool(df: pd.DataFrame, col: str, default=False):
    """Ensure a boolean column exists, with pandas BooleanDtype."""
    if col not in df.columns:
        df[col] = default
    df[col] = df[col].fillna(default).astype('boolean')
    return df

def log_df(df: pd.DataFrame, name: str, sample_cols=None, n=3):
    """Log basic df shape and missingness for debugging."""
    try:
        LOG.info('%s: shape=%s', name, df.shape)
        if sample_cols:
            miss = {c: int(df[c].isna().sum()) for c in sample_cols if c in df.columns}
            LOG.info('%s: missing=%s', name, miss)
    except Exception as e:
        LOG.warning('log_df failed for %s: %s', name, e)
    return df


def log_missing_cols(df: pd.DataFrame, name: str, required: list[str]) -> None:
    """Log missing required columns; helps catch silent schema drift."""
    missing = [c for c in required if c not in df.columns]
    if missing:
        LOG.warning("%s: missing expected columns: %s", name, missing)


import yaml
from dateutil import parser as dateparser

from .utils import HttpConfig, safe_div, clean_name
from .sleeper import SleeperClient
from .external import (
    ExternalConfig,
    load_dynastyprocess_playerids,
    load_dynastyprocess_values_players,
    load_dynastyprocess_values_picks,
    load_nflverse_injuries,
    load_nflverse_player_week_stats,
)
from .lineup import compute_optimal_lineup
from .plan import load_plan_catalog, require_columns


# ============================================================
# Build philosophy (two-step internally):
#  1) Pull & cache ALL Sleeper + supporting NFL data we need.
#  2) Compute every sheet deterministically from those caches.
# ============================================================


@dataclass
class RunConfig:
    league_id: str
    min_season: int | None
    max_season: int | None
    season_type: str = "regular"



def _norm_team_name(name: Any) -> str:
    """Normalize owner/team names for consistent joins (case-insensitive, space-insensitive)."""
    s = str(name or "").strip().lower()
    s = re.sub(r"\s+", "", s)
    return s



# --------------------------
# Logging helpers
# --------------------------

def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()

def _log(path: Path, msg: str) -> None:
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(path.read_text() + msg + "\n" if path.exists() else msg + "\n")
    except Exception:
        pass

def _log_exc(path: Path, where: str, e: Exception) -> None:
    _log(path, f"[{_now_iso()}] ERROR at {where}: {type(e).__name__}: {e}\n{traceback.format_exc()}")

def _to_int(x: Any, default: Optional[int] = None) -> Optional[int]:
    try:
        return int(x)
    except Exception:
        return default

def _to_float(x: Any, default: Optional[float] = None) -> Optional[float]:
    try:
        return float(x)
    except Exception:
        return default

def _safe_df(obj: Any) -> pd.DataFrame:
    return obj if isinstance(obj, pd.DataFrame) else pd.DataFrame()

def _first_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    if not isinstance(df, pd.DataFrame) or df.empty:
        return None
    for c in candidates:
        if c in df.columns:
            return c
    return None

def _epoch_ms_to_dt(ms: Any) -> Optional[datetime]:
    try:
        ms_i = int(ms)
        if ms_i <= 0:
            return None
        return datetime.fromtimestamp(ms_i / 1000, tz=timezone.utc)
    except Exception:
        return None

def _epoch_ms_to_date(ms: Any) -> Optional[date]:
    dt = _epoch_ms_to_dt(ms)
    return dt.date() if dt else None

def _calc_age(birth_date_str: Optional[str], on_date: date) -> Optional[float]:
    if not birth_date_str:
        return None
    try:
        bd = dateparser.parse(str(birth_date_str)).date()
        return round((on_date - bd).days / 365.25, 2)
    except Exception:
        return None


# --------------------------
# Team handle mapping (HANDLE, not franchise name)
# --------------------------

def _team_handle_map(users: List[Dict[str, Any]]) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for u in users or []:
        uid = str(u.get("user_id"))
        handle = u.get("display_name") or u.get("username")
        if not handle:
            meta = u.get("metadata") or {}
            handle = meta.get("team_name") or uid
        out[uid] = str(handle)
    return out


# --------------------------
# NFL team normalization + bye schedule support
# --------------------------

_TEAM_NORMALIZE = {
    "LA": "LAR", "STL": "LAR",
    "SD": "LAC",
    "WSH": "WAS",
    "ARZ": "ARI", "AZ": "ARI",
    "NWE": "NE",
    "KCC": "KC",
    "NOR": "NO",
    "SFO": "SF",
    "TAM": "TB",
    "GNB": "GB",
    "LVR": "LV",
    # already-normalized common codes:
    "ARI": "ARI", "ATL": "ATL", "BAL": "BAL", "BUF": "BUF", "CAR": "CAR",
    "CHI": "CHI", "CIN": "CIN", "CLE": "CLE", "DAL": "DAL", "DEN": "DEN",
    "DET": "DET", "GB": "GB", "HOU": "HOU", "IND": "IND", "JAX": "JAX",
    "KC": "KC", "LAC": "LAC", "LAR": "LAR", "LV": "LV", "MIA": "MIA",
    "MIN": "MIN", "NE": "NE", "NO": "NO", "NYG": "NYG", "NYJ": "NYJ",
    "PHI": "PHI", "PIT": "PIT", "SEA": "SEA", "SF": "SF", "TB": "TB",
    "TEN": "TEN", "WAS": "WAS",
}

def _norm_team(t: Any) -> Optional[str]:
    if not t:
        return None
    s = str(t).strip().upper()
    return _TEAM_NORMALIZE.get(s, s)

def _download_csv_best_effort(urls: List[str], path: Path, timeout: int = 120, debug: Optional[Path]=None) -> pd.DataFrame:
    import requests
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists() and path.stat().st_size > 0:
        try:
            return pd.read_csv(path)
        except Exception:
            pass
    last_err = None
    session = requests.Session()
    session.trust_env = False
    for url in urls:
        try:
            r = session.get(url, timeout=timeout, proxies={"http": None, "https": None})
            if r.status_code == 200 and r.content:
                path.write_bytes(r.content)
                try:
                    return pd.read_csv(path)
                except Exception:
                    return pd.DataFrame()
            last_err = f"{r.status_code} {url}"
        except Exception as e:
            last_err = f"{type(e).__name__}: {e} {url}"
    if debug:
        _log(debug, f"[{_now_iso()}] WARN csv download failed: {last_err}")
    return pd.DataFrame()

def _played_teams_by_week(games: pd.DataFrame, season: int) -> Dict[int, set]:
    games = _safe_df(games)
    out: Dict[int, set] = {}
    if games.empty:
        return out
    if not {"season", "week", "home_team", "away_team"}.issubset(set(games.columns)):
        return out
    sub = games.copy()
    try:
        sub["season"] = pd.to_numeric(sub["season"], errors="coerce").astype("Int64")
        sub["week"] = pd.to_numeric(sub["week"], errors="coerce").astype("Int64")
    except Exception:
        return out
    sub = sub[sub["season"] == season]
    if sub.empty:
        return out
    for wk, g in sub.groupby("week"):
        if pd.isna(wk):
            continue
        home = g["home_team"].dropna().astype(str).map(_norm_team).tolist()
        away = g["away_team"].dropna().astype(str).map(_norm_team).tolist()
        out[int(wk)] = set([t for t in (home + away) if t])
    return out


# --------------------------
# Injury/Suspension flags (brand new approach)
# --------------------------

def _canon_name(s: Any) -> str:
    """Normalize player names for join keys (lowercase, remove punctuation/spaces)."""
    if s is None:
        return ""
    s = str(s)
    s = s.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s

def _status_to_flags(status: Any) -> Tuple[bool, bool]:
    """Convert a text status to (injury, suspension)."""
    s = str(status or "").strip().lower()
    if not s:
        return (False, False)
    if ("susp" in s) or ("sspd" in s):
        return (False, True)
    # only count definite misses
    if ("out" in s) or ("ir" in s) or ("inactive" in s) or ("pup" in s) or ("nfi" in s):
        return (True, False)
    return (False, False)


# --------------------------
# League chain
# --------------------------

def _walk_league_chain(sc: SleeperClient, start_league_id: str, min_season: int | None, max_season: int | None) -> List[Dict[str, Any]]:
    chain: List[Dict[str, Any]] = []
    lid = str(start_league_id)
    seen = set()
    while lid and lid not in seen:
        seen.add(lid)
        try:
            lg = sc.league(lid)
        except Exception:
            break
        if not isinstance(lg, dict):
            break
        season = _to_int(lg.get("season"), None)
        if season is not None and min_season is not None and season < min_season:
            break
        chain.append(lg)
        prev = lg.get("previous_league_id")
        lid = str(prev) if prev else ""
        if lid == "None":
            lid = ""
    chain = sorted(chain, key=lambda x: _to_int(x.get("season"), 0) or 0)
    if max_season is not None:
        chain = [x for x in chain if (_to_int(x.get("season"), 0) or 0) <= max_season]
    return chain


# --------------------------
# Plan column enforcement
# --------------------------

def _ensure_plan_columns(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    df = _safe_df(df).copy()
    for c in cols:
        if c not in df.columns:
            df[c] = None
    return df[cols]


def _is_text_column(col: str) -> bool:
    col_l = col.lower()
    text_markers = [
        "team",
        "player",
        "opponent",
        "record",
        "result",
        "link",
        "assets",
        "type of transaction",
        "date",
        "trade",
        "etc",
        "original",
        "number",
        "pick",
        "position",
        "starter/bench",
        "nfl team",
    ]
    numeric_markers = [
        "number of ",
        "points",
        "pf",
        "avg",
        "max",
        "min",
        "range",
        "margin",
        "win %",
        "win?",
        "loss",
        "efficiency",
        "hardship",
        "luck",
        "tanking",
        "difference",
        "change",
        "weeks",
        "week",
        "year",
        "age",
        "ppg",
        "faab",
        "value",
        "streak",
        "score",
    ]
    if any(m in col_l for m in numeric_markers):
        return False
    return any(m in col_l for m in text_markers)


def _fill_empty_columns(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    for col in cols:
        if col not in df.columns:
            continue
        if df[col].isna().all():
            df[col] = "N/A"
    return df


def _fill_missing_values(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    for col in cols:
        if col not in df.columns:
            continue
        df[col] = df[col].astype(object).replace("", "N/A").fillna("N/A")
    return df


# --------------------------
# Matchup naming for playoffs/toilet
# --------------------------

def _matchup_stage(week: int, playoff_start: Optional[int]) -> Optional[str]:
    if not playoff_start:
        return None
    if week < playoff_start:
        return None
    if week == playoff_start:
        return "SEMIS"
    if week == playoff_start + 1:
        return "FINALS"
    return None


# --------------------------
# Main build
# --------------------------

def build_all(repo_root: Path) -> None:
    debug = repo_root / "exports" / "raw" / "build_debug.log"
    _log(debug, f"\n[{_now_iso()}] ===== Build start =====")

    plan_csv = repo_root / "plan" / "LOTG Plan - Sheet1.csv"
    catalog = load_plan_catalog(plan_csv)

    cfg = yaml.safe_load((repo_root / "config/league.yaml").read_text())
    run_cfg = RunConfig(
        league_id=str(cfg["league_id"]),
        min_season=cfg.get("min_season"),
        max_season=cfg.get("max_season"),
        season_type=str(cfg.get("season_type", "regular")).lower(),
    )

    http = HttpConfig(timeout_seconds=30, max_retries=10, backoff_base_seconds=0.7)
    sc = SleeperClient(run_cfg.league_id, http)

    cache_dir = repo_root / ".cache"
    cache_dir.mkdir(exist_ok=True)
    ext = ExternalConfig(cache_dir=cache_dir, timeout_seconds=120)

    # ------------- External data -------------
    try:
        dp_ids = _safe_df(load_dynastyprocess_playerids(ext))
    except Exception as e:
        dp_ids = pd.DataFrame()
        _log_exc(debug, "load_dynastyprocess_playerids", e)

    for c in ["sleeper_id", "gsis_id", "name"]:
        if c in dp_ids.columns:
            dp_ids[c] = dp_ids[c].astype(str)

    # nflverse games for byes
    games = _download_csv_best_effort(
        urls=[
            "https://raw.githubusercontent.com/nflverse/nfldata/master/data/games.csv",
            "https://github.com/nflverse/nfldata/raw/master/data/games.csv",
        ],
        path=cache_dir / "nfldata_games.csv",
        timeout=120,
        debug=debug,
    )
    played_by_week_by_season: Dict[int, Dict[int, set]] = {}
    if not games.empty:
        try:
            games["season"] = pd.to_numeric(games["season"], errors="coerce").astype("Int64")
            games["week"] = pd.to_numeric(games["week"], errors="coerce").astype("Int64")
        except Exception:
            pass

    # Sleeper NFL players (meta)
    try:
        players_nfl = sc.players_nfl()
    except Exception as e:
        players_nfl = {}
        _log_exc(debug, "players_nfl", e)

    pid_meta: Dict[str, Dict[str, Any]] = {}
    pid_pos: Dict[str, str] = {}
    for pid, meta in (players_nfl or {}).items():
        if not isinstance(meta, dict):
            continue
        pid = str(pid)
        full = meta.get("full_name") or (f"{meta.get('first_name','')} {meta.get('last_name','')}".strip())
        pid_meta[pid] = {
            "full_name": full or pid,
            "pos": (meta.get("position") or ""),
            "team": _norm_team(meta.get("team")),
            "birth_date": meta.get("birth_date") or meta.get("birthdate"),
            "years_exp": meta.get("years_exp"),
            "status": meta.get("status"),
            "injury_status": meta.get("injury_status"),
            "gsis_id": meta.get("gsis_id"),
        }
        pid_pos[pid] = (pid_meta[pid]["pos"] or "").upper()

    # ------------- League chain -------------
    leagues = _walk_league_chain(sc, run_cfg.league_id, run_cfg.min_season, run_cfg.max_season)
    raw_dir = repo_root / "exports" / "raw"
    raw_dir.mkdir(parents=True, exist_ok=True)

    def write_outputs(tables: List[Tuple[str, pd.DataFrame, str]]) -> None:
        out_dir = repo_root / "exports"
        out_dir.mkdir(exist_ok=True)
        for fname, frame, plan_key in tables:
            cols = catalog.get(plan_key, [])
            frame = _safe_df(frame)
            out = _ensure_plan_columns(frame, cols)
            out = _fill_empty_columns(out, cols)
            out = _fill_missing_values(out, cols)
            try:
                require_columns(out, cols, fname.replace(".csv", ""))
            except Exception as e:
                _log_exc(debug, f"require_columns_{fname}", e)
            out.to_csv(out_dir / fname, index=False)

        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            wb.remove(wb.active)

            for csvf in sorted(out_dir.glob("*.csv")):
                sheet_name = csvf.stem[:31]
                ws = wb.create_sheet(title=sheet_name)

                try:
                    d = pd.read_csv(csvf)
                except Exception:
                    d = pd.DataFrame()

                ws.append(list(d.columns))
                for row in d.itertuples(index=False, name=None):
                    ws.append(list(row))
                ws.freeze_panes = "A2"

                if ws.max_column >= 1:
                    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{max(1, ws.max_row)}"

                try:
                    for j, col in enumerate(d.columns, 1):
                        max_len = max([len(str(col))] + [len(str(x)) for x in d[col].head(200).fillna("").astype(str).tolist()])
                        ws.column_dimensions[get_column_letter(j)].width = min(60, max(10, max_len + 2))
                except Exception:
                    pass

            wb.save(out_dir / "LOTG_Stats.xlsx")
        except Exception as e:
            _log_exc(debug, "excel_write", e)

        try:
            import zipfile
            with zipfile.ZipFile(out_dir / "LOTG_Exports.zip", "w", zipfile.ZIP_DEFLATED) as z:
                for f in out_dir.glob("*.csv"):
                    z.write(f, arcname=f.name)
                if (out_dir / "LOTG_Stats.xlsx").exists():
                    z.write(out_dir / "LOTG_Stats.xlsx", arcname="LOTG_Stats.xlsx")
                for f in (out_dir / "raw").glob("*"):
                    if f.is_file():
                        z.write(f, arcname=f"raw/{f.name}")
        except Exception as e:
            _log_exc(debug, "zip_exports", e)

        _log(debug, f"[{_now_iso()}] ===== Build end =====")

    if not leagues:
        fallback_dir = repo_root / "data"
        fallback_tables = []
        for fname, plan_key in [
            ("player_week.csv", "Player-Week"),
            ("player_year.csv", "Player-year"),
            ("player_all_time.csv", "Player-all-time"),
            ("team_week.csv", "team-week"),
            ("team_year.csv", "team-year"),
            ("team_all_time.csv", "team-all-time"),
            ("league_week.csv", "league-week"),
            ("league_year.csv", "league-year"),
            ("league_all_time.csv", "league-all-time"),
            ("transactions.csv", "transactions"),
            ("trades.csv", "trades"),
            ("pick_history.csv", "Pick History"),
        ]:
            src = fallback_dir / fname
            if src.exists():
                try:
                    fallback_tables.append((fname, pd.read_csv(src), plan_key))
                except Exception:
                    fallback_tables.append((fname, pd.DataFrame(), plan_key))
            else:
                fallback_tables.append((fname, pd.DataFrame(), plan_key))
        _log(debug, f"[{_now_iso()}] WARN no leagues found; using fallback data/ outputs")
        write_outputs(fallback_tables)
        return

    # ------------- Output rows -------------
    player_week_rows: List[Dict[str, Any]] = []
    team_week_rows: List[Dict[str, Any]] = []
    transactions_rows: List[Dict[str, Any]] = []
    trades_rows: List[Dict[str, Any]] = []
    pick_rows: List[Dict[str, Any]] = []

    # Internal ledger helpers
    # key: (season, week, roster_id) -> opponent roster_id + opponent points
    opp_rid_map: Dict[Tuple[int, int, int], Optional[int]] = {}
    opp_pf_map: Dict[Tuple[int, int, int], Optional[float]] = {}
    stage_label_map: Dict[Tuple[int, int, int], Optional[str]] = {}
    playoff_start_by_season: Dict[int, Optional[int]] = {}

    # Determine last completed week per league (robust, per Apps Script)
    def last_completed_week(league_id: str, season: int, max_weeks: int = 30) -> int:
        """Last week with any non-zero team points, excluding the fantasy-championship week.

        Sleeper leagues often expose an extra 'final' week that we intentionally exclude from all tables:
        - week 18 for seasons 2021+
        - week 17 for seasons <= 2020 (kept for future ESPN backfill)
        """
        excluded = 18 if int(season) >= 2021 else 17
        last = 0
        for wk in range(1, max_weeks + 1):
            if wk == excluded:
                continue
            try:
                mu = sc.matchups(wk, league_id)
            except Exception:
                mu = None
            if not mu:
                continue
            has_real = any((_to_float(m.get("points"), 0.0) or 0.0) > 0.0 for m in mu)
            if has_real:
                last = wk
        return last


    # ------------- Build each season -------------
    for lg in leagues:
        league_id = str(lg.get("league_id"))
        season = _to_int(lg.get("season"), 0) or 0

        # playoff start week (Sleeper setting)
        settings = lg.get("settings") or {}
        playoff_start = _to_int(settings.get("playoff_week_start"), None)
        playoff_start_by_season[season] = playoff_start

        # cache played_by_week
        if season not in played_by_week_by_season:
            played_by_week_by_season[season] = _played_teams_by_week(games, season) if not games.empty else {}
        played_by_week = played_by_week_by_season.get(season, {})

        # nflverse injuries (used for injury/suspension + historical NFL team fill)
        try:
            injuries = _safe_df(load_nflverse_injuries(ext, season))
        except Exception as e:
            injuries = pd.DataFrame()
            _log_exc(debug, f"load_nflverse_injuries_{season}", e)

        # Build season-level indexes keyed by (week, gsis_id) and (week, canon_name)
        injury_status_by_gsis: Dict[Tuple[int, str], str] = {}
        injury_status_by_name: Dict[Tuple[int, str], str] = {}
        nfl_team_by_gsis: Dict[Tuple[int, str], str] = {}
        nfl_team_by_name: Dict[Tuple[int, str], str] = {}

        if not injuries.empty:
            try:
                inj_df = injuries.copy()
                if "week" in inj_df.columns:
                    inj_df["week"] = pd.to_numeric(inj_df.get("week"), errors="coerce").astype("Int64")
                status_col = _first_col(inj_df, ["report_status", "status", "game_status", "injury_status", "practice_status"]) or ""
                name_col = _first_col(inj_df, ["full_name", "player_name", "name"]) or ""
                team_col = _first_col(inj_df, ["team", "club", "team_abbr"]) or ""
                if "gsis_id" in inj_df.columns:
                    inj_df["gsis_id"] = inj_df["gsis_id"].astype(str)
                for _, r in inj_df.iterrows():
                    wk = r.get("week")
                    if pd.isna(wk):
                        continue
                    wk_i = int(wk)
                    st = str(r.get(status_col) or "").strip()
                    gsis = str(r.get("gsis_id") or "").strip()
                    nm = str(r.get(name_col) or "").strip()
                    tm = str(r.get(team_col) or "").strip()
                    if gsis:
                        injury_status_by_gsis[(wk_i, gsis)] = st
                        if tm:
                            nfl_team_by_gsis[(wk_i, gsis)] = tm
                    if nm:
                        cn = _canon_name(nm)
                        if cn:
                            injury_status_by_name[(wk_i, cn)] = st
                            if tm:
                                nfl_team_by_name[(wk_i, cn)] = tm
            except Exception as e:
                _log_exc(debug, f"injury_index_build_{season}", e)

        # nflverse weekly player stats (best source for team-by-week)
        # If available, prefer this over Sleeper meta (which is not historical).
        try:
            pws = _safe_df(load_nflverse_player_week_stats(ext, season))
        except Exception as e:
            pws = pd.DataFrame()
            _log_exc(debug, f"load_nflverse_player_week_stats_{season}", e)

        if not pws.empty:
            try:
                pws2 = pws.copy()
                if "week" in pws2.columns:
                    pws2["week"] = pd.to_numeric(pws2.get("week"), errors="coerce").astype("Int64")
                # known columns in nflverse weekly stats
                team_col = _first_col(pws2, ["recent_team", "team", "posteam"]) or ""
                name_col = _first_col(pws2, ["player_name", "name", "full_name"]) or ""
                # gsis id column varies; try a few
                gsis_col = _first_col(pws2, ["gsis_id", "player_gsis_id", "gsis"])
                if gsis_col:
                    pws2[gsis_col] = pws2[gsis_col].astype(str)
                for _, r in pws2.iterrows():
                    wk = r.get("week")
                    if pd.isna(wk):
                        continue
                    wk_i = int(wk)
                    tm = str(r.get(team_col) or "").strip()
                    if not tm:
                        continue
                    gsis = str(r.get(gsis_col) or "").strip() if gsis_col else ""
                    nm = str(r.get(name_col) or "").strip()
                    if gsis:
                        nfl_team_by_gsis[(wk_i, gsis)] = _norm_team(tm)
                    if nm:
                        cn = _canon_name(nm)
                        if cn:
                            nfl_team_by_name[(wk_i, cn)] = _norm_team(tm)
            except Exception as e:
                _log_exc(debug, f"player_week_team_index_{season}", e)

        # users/rosters
        try:
            users = sc.users(league_id)
        except Exception as e:
            users = []
            _log_exc(debug, f"users_{season}", e)
        try:
            rosters = sc.rosters(league_id)
        except Exception as e:
            rosters = []
            _log_exc(debug, f"rosters_{season}", e)

        user_handle = _team_handle_map(users)

        roster_owner: Dict[int, str] = {}
        roster_to_team: Dict[int, str] = {}
        for r in rosters or []:
            rid = _to_int(r.get("roster_id"), None)
            if rid is None:
                continue
            roster_owner[rid] = str(r.get("owner_id") or "")
            raw_name = user_handle.get(roster_owner[rid], f"Roster {rid}")
            canon = _norm_team_name(raw_name)
            # Preserve a stable display name for this canonical team key.
            if 'team_display' not in locals():
                team_display = {}
            if canon not in team_display:
                team_display[canon] = str(raw_name)
            roster_to_team[rid] = team_display[canon]


        # traded picks snapshot (used for pick history reconstruction)
        try:
            traded_picks = sc.traded_picks(league_id) or []
            traded_picks_by_season = locals().get("traded_picks_by_season", {})
            traded_picks_by_season[season] = traded_picks
        except Exception as e:
            traded_picks_by_season = locals().get("traded_picks_by_season", {})
            traded_picks_by_season[season] = []
            _log_exc(debug, f"traded_picks_{season}", e)

        # raw snapshots
        try:
            (raw_dir / f"league_{season}.json").write_text(json.dumps(lg, indent=2))
            (raw_dir / f"users_{season}.json").write_text(json.dumps(users, indent=2))
            (raw_dir / f"rosters_{season}.json").write_text(json.dumps(rosters, indent=2))
        except Exception:
            pass

        
        # traded picks (for future draft capital / tanking)
        try:
            traded_picks = sc.traded_picks(league_id)
            try:
                (raw_dir / f"traded_picks_{season}.json").write_text(json.dumps(traded_picks, indent=2))
            except Exception:
                pass
        except Exception as e:
            traded_picks = []
            _log_exc(debug, f"traded_picks_{season}", e)

# draft picks history (rookie + startup as available in Sleeper; still partial)
        try:
            drafts = sc.drafts(league_id)
        except Exception as e:
            drafts = []
            _log_exc(debug, f"drafts_{season}", e)
        draft_picks_all: List[Dict[str, Any]] = []
        for d in drafts or []:
            did = str(d.get("draft_id") or "")
            if not did:
                continue
            try:
                picks = sc.draft_picks(did)
            except Exception as e:
                picks = []
                _log_exc(debug, f"draft_picks_{season}_{did}", e)
            for p in picks or []:
                p["draft_id"] = did
                p["draft_season"] = season
            draft_picks_all.extend(picks or [])

        for p in draft_picks_all:
            rnd = p.get("round")
            pick_no = p.get("pick_no")
            roster_id = p.get("roster_id")
            player = p.get("player_id")
            team = roster_to_team.get(_to_int(roster_id, -1), f"Roster {roster_id}") if roster_id is not None else None
            pick_rows.append({
                "Year": season,
                "Original Team": team,
                "Number": f"R{rnd}.{pick_no}",
                "Player Picked": pid_meta.get(str(player), {}).get("full_name") if player else None,
                "Trade 1": None, "Trade 2": None, "Trade 3": None, "Trade 4": None, "Trade 5": None,
                "Trade 6": None, "Trade 7": None, "Trade 8": None, "Trade 9": None, "Trade 10": None,
                "etc": None,
            })

        # robust last completed week
        try:
            last_week = last_completed_week(league_id, season)
        except Exception as e:
            last_week = 0
            _log_exc(debug, f"last_completed_week_{season}", e)

        if last_week <= 0:
            _log(debug, f"[{_now_iso()}] INFO season {season}: no completed weeks, skipping")
            continue

        # Exclude week 18 always; if season < 2021 exclude week 17 (kept for future ESPN import)
        def week_allowed(wk: int) -> bool:
            if wk == 18:
                return False
            if season < 2021 and wk == 17:
                return False
            return True

        # storage for week/team scoring for expected win, etc.
        team_pf_by_week: Dict[int, Dict[str, float]] = defaultdict(dict)

        # ------------- Pre-fetch all weekly matchups & transactions -------------
        matchups_by_week: Dict[int, List[Dict[str, Any]]] = {}
        tx_by_week: Dict[int, List[Dict[str, Any]]] = {}

        for wk in range(1, min(last_week, 30) + 1):
            if not week_allowed(wk):
                continue
            try:
                mu = sc.matchups(wk, league_id) or []
                matchups_by_week[wk] = mu
                for m in mu:
                    rid = _to_int(m.get("roster_id"), None)
                    if rid is None:
                        continue
                    tm = roster_to_team.get(rid, f"Roster {rid}")
                    team_pf_by_week[wk][tm] = float(_to_float(m.get("points"), 0.0) or 0.0)
            except Exception as e:
                _log_exc(debug, f"matchups_{season}_wk{wk}", e)

            try:
                tx_by_week[wk] = sc.transactions(wk, league_id) or []
            except Exception as e:
                tx_by_week[wk] = []
                _log_exc(debug, f"transactions_{season}_wk{wk}", e)

        # ------------- Build opponent roster mapping + playoff labels -------------
        for wk, mu in matchups_by_week.items():
            mdf = _safe_df(pd.DataFrame(mu))
            if mdf.empty:
                continue
            if "matchup_id" not in mdf.columns:
                continue
            # ensure numeric
            try:
                mdf["roster_id"] = pd.to_numeric(mdf["roster_id"], errors="coerce").astype("Int64")
                mdf["points"] = pd.to_numeric(mdf["points"], errors="coerce").fillna(0.0)
            except Exception:
                pass

            stage = _matchup_stage(wk, playoff_start)

            for mid, g in mdf.groupby("matchup_id"):
                rids = [int(x) for x in g["roster_id"].dropna().astype(int).tolist()]
                if len(rids) != 2:
                    continue
                a, b = rids
                pa = float(g.loc[g["roster_id"] == a, "points"].iloc[0])
                pb = float(g.loc[g["roster_id"] == b, "points"].iloc[0])

                opp_rid_map[(season, wk, a)] = b
                opp_rid_map[(season, wk, b)] = a
                opp_pf_map[(season, wk, a)] = pb
                opp_pf_map[(season, wk, b)] = pa

            # playoff/toilet naming per your rules:
            if stage:
                # Determine top4 by regular season W-L then PF tiebreaker, using weeks < playoff_start.
                try:
                    if playoff_start and wk == playoff_start:
                        reg = []
                        for rr in rosters or []:
                            rid = _to_int(rr.get("roster_id"), None)
                            if rid is None:
                                continue
                            tm = roster_to_team.get(rid, f"Roster {rid}")
                            # compute record up to reg season
                            wins = losses = ties = 0
                            pf_sum = 0.0
                            for w2 in range(1, playoff_start):
                                if not week_allowed(w2):
                                    continue
                                pf2 = team_pf_by_week.get(w2, {}).get(tm)
                                if pf2 is None:
                                    continue
                                pf_sum += float(pf2)
                                opp_pf2 = opp_pf_map.get((season, w2, rid))
                                if opp_pf2 is None:
                                    continue
                                if pf2 > opp_pf2:
                                    wins += 1
                                elif pf2 < opp_pf2:
                                    losses += 1
                                else:
                                    ties += 1
                            reg.append((tm, rid, wins, losses, ties, pf_sum))
                        reg.sort(key=lambda x: (x[2] + 0.5 * x[4], x[5]), reverse=True)
                        top4 = set([rid for _, rid, *_ in reg[:4]])
                        bottom4 = set([rid for _, rid, *_ in reg[4:]])
                        # annotate this week (semis) and next week (finals)
                        for rid in top4:
                            stage_label_map[(season, playoff_start, rid)] = "Semifinal"
                        for rid in bottom4:
                            stage_label_map[(season, playoff_start, rid)] = "Toilet Semis"
                        # next week labels depend on semis results
                        finals_week = playoff_start + 1
                        if week_allowed(finals_week) and finals_week in matchups_by_week:
                            # Determine winners/losers within those brackets
                            for rid in top4:
                                opp = opp_rid_map.get((season, playoff_start, rid))
                                if opp is None:
                                    continue
                                if rid < opp:  # handle each pair once
                                    pf_a = team_pf_by_week[playoff_start].get(roster_to_team[rid], 0.0)
                                    pf_b = team_pf_by_week[playoff_start].get(roster_to_team[opp], 0.0)
                                    win_a = (pf_a > pf_b)
                                    winner = rid if win_a else opp
                                    loser = opp if win_a else rid
                                    stage_label_map[(season, finals_week, winner)] = "Final"
                                    stage_label_map[(season, finals_week, loser)] = "3rd Place"
                            for rid in bottom4:
                                opp = opp_rid_map.get((season, playoff_start, rid))
                                if opp is None:
                                    continue
                                if rid < opp:
                                    pf_a = team_pf_by_week[playoff_start].get(roster_to_team[rid], 0.0)
                                    pf_b = team_pf_by_week[playoff_start].get(roster_to_team[opp], 0.0)
                                    win_a = (pf_a > pf_b)
                                    winner = rid if win_a else opp
                                    loser = opp if win_a else rid
                                    stage_label_map[(season, finals_week, winner)] = "Toilet Final"
                                    stage_label_map[(season, finals_week, loser)] = "Toilet Trash"
                except Exception as e:
                    _log_exc(debug, f"playoff_labeling_{season}_wk{wk}", e)

        # ------------- Weekly loop to build team_week & player_week -------------
        prev_starters_by_team: Dict[str, set] = {}
        prev_roster_by_team: Dict[str, set] = {}
        player_last5_healthy: Dict[str, deque] = defaultdict(lambda: deque(maxlen=5))  # for hardship later
        # for player awards
        awards_weekly = defaultdict(list)  # (season,wk) -> list of (pid, team, pts, started, pos)

        for wk, mu in matchups_by_week.items():
            stage = _matchup_stage(wk, playoff_start)
            # tx summaries
            faab_spent: Dict[str, float] = defaultdict(float)
            trade_count: Dict[str, int] = defaultdict(int)
            tx_count: Dict[str, int] = defaultdict(int)

            for t in tx_by_week.get(wk, []):
                try:
                    creator = str(t.get("creator") or "")
                    team = user_handle.get(creator) if creator else None
                    if team:
                        tx_count[team] += 1
                        if t.get("type") == "trade":
                            trade_count[team] += 1
                        meta = t.get("metadata") or {}
                        if isinstance(meta, dict):
                            bid = _to_float(meta.get("waiver_bid") or meta.get("faab") or 0.0, 0.0) or 0.0
                            faab_spent[team] += float(bid)
                except Exception as e:
                    _log_exc(debug, f"tx_summary_{season}_wk{wk}", e)

            # build team rows
            for m in mu:
                try:
                    rid = _to_int(m.get("roster_id"), None)
                    if rid is None:
                        continue
                    team = roster_to_team.get(rid, f"Roster {rid}")
                    pf = float(_to_float(m.get("points"), 0.0) or 0.0)

                    opp_rid = opp_rid_map.get((season, wk, rid))
                    opp_team = roster_to_team.get(opp_rid, f"Roster {opp_rid}") if opp_rid is not None else None
                    opp_points = opp_pf_map.get((season, wk, rid))
                    margin = (pf - opp_points) if opp_points is not None else None
                    win = None
                    if margin is not None:
                        win = 1 if margin > 0 else 0 if margin < 0 else 0.5

                    def _valid_pid(x: Any) -> Optional[str]:
                        if x is None:
                            return None
                        s = str(x).strip()
                        if not s or s.lower() in ("none", "nan"):
                            return None
                        # Sleeper sometimes uses 0 placeholders
                        if s == "0":
                            return None
                        return s

                    starters = [pid for pid in (_valid_pid(x) for x in (m.get("starters") or [])) if pid]
                    players = [pid for pid in (_valid_pid(x) for x in (m.get("players") or [])) if pid]
                    ppts_raw = m.get("players_points") or {}
                    ppts: Dict[str, float] = {}
                    if isinstance(ppts_raw, dict):
                        for k, v in ppts_raw.items():
                            try:
                                ppts[str(k)] = float(v)
                            except Exception:
                                pass

                    # Max PF: use proven algorithm (Apps Script parity)
                    max_pf = compute_optimal_lineup(ppts, pid_pos, season)
                    # Sanity check: if a team scored points, Max PF must be > 0 (otherwise per-player points were lost).
                    if (pf or 0.0) > 0.0 and (max_pf or 0.0) <= 0.0:
                        _log(
                            debug,
                            "ERROR: Max PF computed as 0 despite PF>0. league=%s season=%s week=%s roster_id=%s "
                            "pf=%s players_points_type=%s players_points_len=%s"
                            % (
                                league_id,
                                season,
                                wk,
                                rid,
                                pf,
                                type(ppts_raw).__name__,
                                (len(ppts) if isinstance(ppts, dict) else "NA"),
                            ),
                        )
                        LOG.warning(
                            "Max PF sanity: PF>0 but Max PF==0 for %s %s wk=%s roster=%s; leaving Max PF blank. Check raw exports.",
                            season,
                            league_id,
                            wk,
                            rid,
                        )
                    eff = safe_div(pf, max_pf, default=0.0)

                    # expected win percentile vs league that week
                    scores = list(team_pf_by_week.get(wk, {}).values())
                    expected = None
                    luck_raw = None
                    if scores and len(scores) > 1:
                        expected = sum(1 for s in scores if pf > s) / max(1, (len(scores) - 1))
                        if win is not None:
                            luck_raw = (win - expected)

                    prev = prev_starters_by_team.get(team, set())
                    cur = set(starters)
                    # turnover = number of *new* starters compared to prior week (bounded by starter slots)
                    turnover = len(cur - prev) if prev else None
                    prev_starters_by_team[team] = cur

                    prev_r = prev_roster_by_team.get(team, set())
                    cur_r = set(players)
                    # roster turnover = number of *new* rostered players compared to prior week
                    roster_turnover = len(cur_r - prev_r) if prev_r else None
                    prev_roster_by_team[team] = cur_r

                    starter_points = [ppts.get(pid, 0.0) for pid in starters]
                    donuts = sum(1 for x in starter_points if float(x) == 0.0)
                    under10 = sum(1 for x in starter_points if float(x) < 10.0)
                    over20 = sum(1 for x in starter_points if float(x) > 20.0)
                    over30 = sum(1 for x in starter_points if float(x) > 30.0)
                    over40 = sum(1 for x in starter_points if float(x) > 40.0)
                    over50 = sum(1 for x in starter_points if float(x) > 50.0)
                    diff_hi_lo = (max(starter_points) - min(starter_points)) if starter_points else None

                    def count_pos(pids, pos):
                        return sum(1 for pid in pids if (pid_pos.get(pid) or "") == pos)

                    qb_s, rb_s, wr_s, te_s = count_pos(starters, "QB"), count_pos(starters, "RB"), count_pos(starters, "WR"), count_pos(starters, "TE")
                    qb_r, rb_r, wr_r, te_r = count_pos(players, "QB"), count_pos(players, "RB"), count_pos(players, "WR"), count_pos(players, "TE")

                    rook_s = sum(1 for pid in starters if str(pid_meta.get(pid, {}).get("years_exp")) in ("0", "0.0"))
                    rook_r = sum(1 for pid in players if str(pid_meta.get(pid, {}).get("years_exp")) in ("0", "0.0"))

                    approx_date = date(season, 9, 1) + timedelta(days=7 * (wk - 1))
                    ages = [a for a in (_calc_age(pid_meta.get(pid, {}).get("birth_date"), approx_date) for pid in players) if a is not None]
                    avg_age = round(sum(ages) / len(ages), 2) if ages else None

                    roster_nfl_teams = [pid_meta.get(pid, {}).get("team") for pid in players if pid_meta.get(pid, {}).get("team")]
                    start_nfl_teams = [pid_meta.get(pid, {}).get("team") for pid in starters if pid_meta.get(pid, {}).get("team")]
                    most_start_same = max(Counter(start_nfl_teams).values()) if start_nfl_teams else None
                    most_roster_same = max(Counter(roster_nfl_teams).values()) if roster_nfl_teams else None
                    most_start_team = Counter(start_nfl_teams).most_common(1)[0][0] if start_nfl_teams else None
                    most_roster_team = Counter(roster_nfl_teams).most_common(1)[0][0] if roster_nfl_teams else None

                    def max_same_team_by_pos(pids, pos):
                        teams = [
                            pid_meta.get(pid, {}).get("team")
                            for pid in pids
                            if (pid_pos.get(pid) or "").upper() == pos and pid_meta.get(pid, {}).get("team")
                        ]
                        return max(Counter(teams).values()) if teams else None

                    # Opponent label per playoffs spec
                    opp_label = opp_team
                    label = stage_label_map.get((season, wk, rid))
                    week_name = label if label else f"Week {wk}"

                    team_week_rows.append({
                        "Team": team,
                        "Opponent Team (raw)": opp_team,
                        "Week": wk,
                        "Week Name": week_name,
                        "Year": season,
                        "PF": round(pf, 2),
                        "Win?": win,
                        "Opponent": opp_team,
                        "Week label": label,
                        "Round": _round_from_label(label),
                        "Points against": round(float(opp_points), 2) if opp_points is not None else None,
                        "Margin": round(float(margin), 2) if margin is not None else None,
                        "Max PF": round(max_pf, 2) if max_pf is not None else None,
                        "Efficiency": round(eff, 4) if eff is not None else None,
                        "Number of Injuries": None,         # computed later from player_week
                        "Number of suspensions": None,      # computed later from player_week
                        "Number of players on bye": None,   # computed later from player_week
                        "Largest deficit overcome (if win)": None,
                        "Starter turnover from previous week": turnover,
                        "Roster turnover from previous week": roster_turnover,
                        "Difference in pregame avg max PF from opponent": None,  # computed later
                        "UPST": None,                      # computed later
                        "Hardship": None,                  # computed later
                        "Tanking": None,                   # computed later (needs pick ledger; best effort later)
                        "Luck": round(luck_raw, 4) if luck_raw is not None else None,
                        "Brosenzweig": None,
                        "Sisenzweig": None,
                        "Number of donuts": donuts,
                        "Number of players under 10": under10,
                        "Number of players over 20": over20,
                        "Number of players over 30": over30,
                        "Number of players over 40": over40,
                        "Number of players over 50": over50,
                        "Number of QB started": qb_s,
                        "Number of WR started": wr_s,
                        "Number of RB started": rb_s,
                        "Number of TE started": te_s,
                        "Number of QB rostered": qb_r,
                        "Number of WR rostered": wr_r,
                        "Number of RB rostered": rb_r,
                        "Number of TE rostered": te_r,
                        "Number of transactions": int(tx_count.get(team, 0)),
                        "Number of trades": int(trade_count.get(team, 0)),
                        "Amount of FAAB spent": round(float(faab_spent.get(team, 0.0)), 2),
                        "Most number of players started from same NFL team": most_start_same,
                        "Most number of players started from same NFL team (team)": most_start_team,
                        "Most number of players rostered from same NFL team": most_roster_same,
                        "Most number of players rostered from same NFL team (team)": most_roster_team,
                        "Most number of QBs started from same NFL team": max_same_team_by_pos(starters, "QB"),
                        "Most number of QBs rostered from same NFL team": max_same_team_by_pos(players, "QB"),
                        "Most number of RBs started from same NFL team": max_same_team_by_pos(starters, "RB"),
                        "Most number of RBs rostered from same NFL team": max_same_team_by_pos(players, "RB"),
                        "Most number of WR started from same NFL team": max_same_team_by_pos(starters, "WR"),
                        "Most number of WR rostered from same NFL team": max_same_team_by_pos(players, "WR"),
                        "Most number of TE started from same NFL team": max_same_team_by_pos(starters, "TE"),
                        "Most number of TE rostered from same NFL team": max_same_team_by_pos(players, "TE"),
                        "Number of NFL teams among starting players": len(set(start_nfl_teams)) if start_nfl_teams else None,
                        "Number of NFL teams amoung rostered players": len(set(roster_nfl_teams)) if roster_nfl_teams else None,
                        "Number of rookies started": rook_s,
                        "Number of rookies rostered": rook_r,
                        "Player average age": avg_age,
                        "Difference between highest and lowest starters": round(diff_hi_lo, 2) if diff_hi_lo is not None else None,
                        "Combined matchup score": round(pf + opp_points, 2) if opp_points is not None else None,
                        "Win streak": None,
                        "Loss streak": None,
                        "Win streak counting previous season": None,
                        "Loss streak counting previous season": None,
                        "Top half of league?": None,
                        "Highest score?": None,
                        "Lowest score?": None,
                        "Narrowest victory?": None,
                        "Largest blowout?": None,
                        "Most efficient?": None,
                        "Least efficient?": None,
                        "Increase in points from previous week": None,
                        "Number of cuffs rostered": None,
                        "Number of cuffs started": None,
                        "Future draft capital": None,
                        "Startup draft players remaining": None,
                        # leave remaining plan columns to enforcement step
                    })

                    # starter slot labels are not provided reliably by Sleeper; we approximate by roster order
                    starter_slot = {}
                    # League has fixed slots; for this build we only need "Position started in" for analysis,
                    # so we store the player's NFL position as a stable proxy.
                    for pid in starters:
                        starter_slot[pid] = pid_pos.get(pid)

                    played_set = played_by_week.get(wk, set())

                    bench = [pid for pid in players if pid not in starters]
                    best_bench_pid = None
                    best_bench_pts = None
                    if bench:
                        best_bench_pid = max(bench, key=lambda pid: ppts.get(pid, 0.0))
                        best_bench_pts = float(ppts.get(best_bench_pid, 0.0))

                    worst_starter_pid = None
                    worst_starter_pts = None
                    if starters:
                        worst_starter_pid = min(starters, key=lambda pid: ppts.get(pid, 0.0))
                        worst_starter_pts = float(ppts.get(worst_starter_pid, 0.0))

                    for pid in players:
                        meta = pid_meta.get(pid, {})
                        full_name = meta.get("full_name") or pid
                        # NFL team should be the team the player was on *that week*.
                        # Prefer nflverse weekly stats (recent_team), then injury report team,
                        # and only then fall back to Sleeper meta (which is current-state).
                        nfl_team = None
                        pts = float(ppts.get(pid, 0.0))
                        started = pid in starters
                        pos = (pid_pos.get(pid) or meta.get("position") or "")
                        pos = str(pos).upper() if pos else None
                        lineup_slot = starter_slot.get(pid) if started else None

                        # gsis id lookup for nflverse
                        gsis = None
                        if not dp_ids.empty and "sleeper_id" in dp_ids.columns and "gsis_id" in dp_ids.columns:
                            try:
                                match = dp_ids.loc[dp_ids["sleeper_id"].astype(str) == pid]
                                if not match.empty:
                                    gsis = str(match["gsis_id"].iloc[0])
                            except Exception:
                                gsis = None
                        if not gsis:
                            gsis = meta.get("gsis_id")

                        # ---- Injury / Suspension (brand-new approach) ----
                        # Primary source: nflverse injuries keyed by (week, gsis_id) with name fallback.
                        # We mark injury/suspension when a player *missed* the NFL week (0 fantasy points and not a bye)
                        # and nflverse indicates OUT/IR/INACTIVE/PUP or SUSP.

                        if gsis:
                            nfl_team = nfl_team_by_gsis.get((wk, str(gsis)))
                        if not nfl_team:
                            nfl_team = nfl_team_by_name.get((wk, _canon_name(full_name)))
                        if not nfl_team:
                            nfl_team = meta.get("team")

                        bye = False
                        if nfl_team and played_set:
                            bye = (_norm_team(nfl_team) not in played_set)
                        if pts > 0:
                            bye = False

                        inj = False
                        susp = False
                        if (pts or 0.0) == 0.0 and (bye is False):
                            st = None
                            if gsis:
                                st = injury_status_by_gsis.get((wk, str(gsis)))
                            if not st:
                                st = injury_status_by_name.get((wk, _canon_name(full_name)))
                            inj, susp = _status_to_flags(st)

                        rookie = str(meta.get("years_exp")) in ("0", "0.0")
                        age = _calc_age(meta.get("birth_date"), approx_date)

                        diff_best_bench = (pts - best_bench_pts) if (started and best_bench_pts is not None) else None
                        diff_worst_starter = (pts - worst_starter_pts) if ((not started) and worst_starter_pts is not None) else None
                        ref_player = None
                        if started and best_bench_pid:
                            ref_player = pid_meta.get(best_bench_pid, {}).get("full_name") or best_bench_pid
                        elif (not started) and worst_starter_pid:
                            ref_player = pid_meta.get(worst_starter_pid, {}).get("full_name") or worst_starter_pid

                        rookie = str(meta.get("years_exp")) in ("0", "0.0")
                        age = _calc_age(meta.get("birth_date"), approx_date)

                        diff_best_bench = (pts - best_bench_pts) if (started and best_bench_pts is not None) else None
                        diff_worst_starter = (pts - worst_starter_pts) if ((not started) and worst_starter_pts is not None) else None
                        ref_player = None
                        if started and best_bench_pid:
                            ref_player = pid_meta.get(best_bench_pid, {}).get("full_name") or best_bench_pid
                        elif (not started) and worst_starter_pid:
                            ref_player = pid_meta.get(worst_starter_pid, {}).get("full_name") or worst_starter_pid

                        if inj is None:
                            inj = False
                        if susp is None:
                            susp = False
                        if bye is None:
                            bye = False

                        if inj is None:
                            inj = False
                        if susp is None:
                            susp = False
                        if bye is None:
                            bye = False

                        if inj is None:
                            inj = False
                        if susp is None:
                            susp = False
                        if bye is None:
                            bye = False

                        if inj is None:
                            inj = False
                        if susp is None:
                            susp = False
                        if bye is None:
                            bye = False

                        player_week_rows.append({
                            "Player": full_name,
                            "Team": team,
                            "Week": wk,
                            "Year": season,
                            "Points": round(pts, 2),
                            "Injury?": bool(inj),
                            "Suspension?": bool(susp),
                            "Bye?": bool(bye),
                            "Starter/Bench": "Starter" if started else "Bench",
                            "% of points (if starter)": round(pts / pf, 4) if started and pf else None,
                            # Requested: this column should be the player's actual position, not the lineup slot label.
                            "Position started in (if starter)": pos if started else None,
                            # Separate column for the player's position.
                            "Position": pos,
                            "Change from previous week": None,
                            "Change from previous 5 weeks avg": None,
                            "Change from career average to that point": None,
                            "Change from overall career average": None,
                            "Number of weeks on team": None,
                            "Number of consecutive weeks on bench before start (if starter)": None,
                            "Number of consecutive weeks on bench before start excluding injury/bye (if starter)": None,
                            "Total weeks as team starter to that point": None,
                            "Total weeks on bench to that point": None,
                            "Total weeks as team starter on that team this season": None,
                            "Total weeks on bench on that team this season": None,
                            "- Activated Cuff? (Was a player of the same nfl team/position & who averages >10 PPG more over last 5 played games injured? Only for players with avg <10 PPG)": 0,
                            "Difference from best startable bench (if starter)": round(diff_best_bench, 2) if diff_best_bench is not None else None,
                            "Difference from worst benchable starter (if bench)": round(diff_worst_starter, 2) if diff_worst_starter is not None else None,
                            "Reference player name": ref_player,
                            "Difference in averages of best/worst startables over previous 5 games": None,
                            "Cuff adjusted difference": None,
                            "Rookie?": 1 if rookie else 0,
                            "Age": age,
                            "NFL team": nfl_team,
                            # award flags (filled later)
                            "Player of the week?": None,
                            "QB of the week?": None,
                            "RB of the week?": None,
                            "WR of the week?": None,
                            "TE of the week?": None,
                            "Benchwarmer of the week?": None,
                            "Bench QB of the week?": None,
                            "Bench RB of the week?": None,
                            "Bench WR of the week?": None,
                            "Bench TE of the week?": None,
                            "Highest starter on team?": None,
                            "Lowest starter on team?": None,
                        })

                        # store for awards later
                        awards_weekly[(season, wk)].append((pid, team, pts, started, pid_pos.get(pid) or ""))

                except Exception as e:
                    _log_exc(debug, f"team_player_rows_{season}_wk{wk}", e)

            # Transactions rows (non-trade) + trades ledger
            for t in tx_by_week.get(wk, []):
                try:
                    ttype = t.get("type")
                    created_date = _epoch_ms_to_date(t.get("created"))
                    created_dt = _epoch_ms_to_dt(t.get("created"))
                    creator = str(t.get("creator") or "")
                    team = user_handle.get(creator) if creator else None

                    # Trades ledger (one row per involved team)
                    if ttype == "trade":
                        roster_ids = t.get("roster_ids") or []
                        if not isinstance(roster_ids, list):
                            roster_ids = []
                        roster_ids_int = [int(x) for x in roster_ids if _to_int(x, None) is not None]
                        adds = t.get("adds") or {}
                        if not isinstance(adds, dict):
                            adds = {}
                        draft_picks = t.get("draft_picks") or []
                        if not isinstance(draft_picks, list):
                            draft_picks = []

                        # received assets by team
                        recv_players: Dict[int, List[str]] = defaultdict(list)
                        for pid, rrid in adds.items():
                            rr = _to_int(rrid, None)
                            if rr is None:
                                continue
                            recv_players[rr].append(pid_meta.get(str(pid), {}).get("full_name") or str(pid))

                        recv_picks: Dict[int, List[str]] = defaultdict(list)
                        for dp in draft_picks:
                            if not isinstance(dp, dict):
                                continue
                            owner_id = _to_int(dp.get("owner_id"), None)
                            if owner_id is None:
                                continue
                            recv_picks[owner_id].append(f"{dp.get('season')} R{dp.get('round')}")

                        # Build row per roster in roster_ids_int
                        for rid in roster_ids_int:
                            tm = roster_to_team.get(rid, f"Roster {rid}")
                            others = [roster_to_team.get(o, f"Roster {o}") for o in roster_ids_int if o != rid]
                            received = []
                            received.extend(recv_players.get(rid, []))
                            received.extend(recv_picks.get(rid, []))
                            dropped = []
                            for o in roster_ids_int:
                                if o == rid:
                                    continue
                                dropped.extend(recv_players.get(o, []))
                                dropped.extend(recv_picks.get(o, []))
                            trades_rows.append({
                                "Team": tm,
                                "Team's traded with": "; ".join(sorted(set([x for x in others if x]))),
                                "Assets recieved": "; ".join(received) if received else None,
                                "Assets dropped": "; ".join(dropped) if dropped else None,
                                "Date": created_dt.isoformat() if created_dt else (str(created_date) if created_date else None),
                                # KTC/Oliver columns stay blank by design for now
                                "KTC value difference at deal time": None,
                                "KTC value difference at end of season": None,
                                "KTC value difference 1 year later": None,
                                "KTC value difference 2 years later": None,
                                "Oliver value difference at deal time": None,
                                "Oliver value difference at end of season": None,
                                "Oliver value difference 1 year later": None,
                                "Oliver value difference 2 years later": None,
                                "Pick value recieved": None,
                                "Change in pick value at draft time": None,
                                "Assets retained now": None,
                                "Assets traded away": None,
                                "Return from trades": None,
                                "Additional assets traded away in those deals": None,
                                "Return from trades of trades...of trades. Keep going until present day": None,
                                "Asset difference in average age": None,
                                "Tanking before": None,
                                "Tanking after": None,
                                "Link to next transaction": None,
                                "Link to previous transaction": None,
                            })

                        continue  # don't add to transactions.csv

                    # Non-trade transactions
                    if ttype not in ("waiver", "free_agent", "commissioner"):
                        # keep but label if unknown
                        pass

                    adds = t.get("adds") or {}
                    drops = t.get("drops") or {}
                    meta = t.get("metadata") or {}
                    faab = meta.get("waiver_bid") if isinstance(meta, dict) else None
                    num_bids = meta.get("num_bids") if isinstance(meta, dict) else None

                    if not isinstance(adds, dict):
                        adds = {}
                    if not isinstance(drops, dict):
                        drops = {}

                    for pid, rrid in adds.items():
                        pid = str(pid)
                        dropped = None
                        for dp, drid in drops.items():
                            if str(drid) == str(rrid):
                                dropped = str(dp)
                                break

                        transactions_rows.append({
                            "Team": team,
                            "Player Added": pid_meta.get(pid, {}).get("full_name") or pid,
                            "Player Dropped": pid_meta.get(dropped, {}).get("full_name") if dropped else None,
                            "type of transaction (waiver/free agency)": ttype,
                            "Faab": faab,
                            "Date": created_dt.isoformat() if created_dt else (str(created_date) if created_date else None),
                            "Number of bids": num_bids,
                            "Link to next transaction": None,
                            "Link to previous transaction": None,
                            "Average PPG on team": None,
                            "Average PPG of dropped player over same time": None,
                            "Difference of averages": None,
                            "Difference of averages adjusted by position": None,
                            "Age difference": None,
                            "Player addition value": None,
                            "Cuff at time of pickup?": None,
                            "Weeks between pickup and start": None,
                            "Number of starts before next drop": None,
                            "% of starts made while rostered": None,
                            "Injury adjusted % of starts made while rostered": None,
                            "Date dropped/traded": None,
                            "Tanking before": None,
                            "Tanking after": None,
                            "Number of times picked up by this team": None,
                        })
                except Exception as e:
                    _log_exc(debug, f"transactions_trades_rows_{season}_wk{wk}", e)

    # --------------------------
    # Convert to DataFrames
    # --------------------------
    pw = pd.DataFrame(player_week_rows)

    if not pw.empty and "Team" in pw.columns:
        pw["_team_canon"] = pw["Team"].apply(_norm_team_name)
        canon_to_disp = {}
        for t in pw["Team"].dropna().astype(str).tolist():
            c=_norm_team_name(t)
            if c and c not in canon_to_disp:
                canon_to_disp[c]=t
        pw["Team"] = pw["_team_canon"].map(canon_to_disp).fillna(pw["Team"])
        pw.drop(columns=["_team_canon"], inplace=True, errors="ignore")
    log_df(pw, 'player_week', sample_cols=['Points','Injury?','Suspension?','Bye?','Starter?'])
    tw = pd.DataFrame(team_week_rows)

    # Normalize team names (case-insensitive) across seasons so joins don't duplicate teams like 'Shmuel256' vs 'shmuel256'.
    if not tw.empty and "Team" in tw.columns:
        tw["_team_canon"] = tw["Team"].apply(_norm_team_name)
        canon_to_disp: Dict[str,str] = {}
        for t in tw["Team"].dropna().astype(str).tolist():
            c=_norm_team_name(t)
            if c and c not in canon_to_disp:
                canon_to_disp[c]=t
        tw["Team"] = tw["_team_canon"].map(canon_to_disp).fillna(tw["Team"])
        tw.drop(columns=["_team_canon"], inplace=True, errors="ignore")
    if not tw.empty:
        log_missing_cols(tw, "team_week", [
            "Year", "Week", "Team", "PF", "Points against", "Margin", "Max PF", "Efficiency"
        ])
        zero_max = int((pd.to_numeric(tw.get("Max PF"), errors="coerce").fillna(0) <= 0).sum())
        LOG.info("team_week: rows=%s zero_max_pf=%s", len(tw), zero_max)
    log_df(tw, 'team_week', sample_cols=['PF','Max PF','Efficiency'])

    # ---- Tanking (user formula)
    # Tanking(team, week) uses season-to-date averages through that week.
    # Tanking(team, year) is the final week value.
    def _safe_div(n, d):
        try:
            d = float(d)
            n = float(n)
            if d == 0 or (pd.isna(d) or pd.isna(n)):
                return 0.0
            return n / d
        except Exception:
            return 0.0

    def _tanking_score(avg_pf, avg_max_pf, avg_age, league_avg_pf, league_avg_max_pf, league_avg_age, pick_sum, future_cap):
        # 1/6 *(1 - (AvgPF-2/3 L)/(L-2/3 L))
        denom1 = (league_avg_pf - (2.0/3.0)*league_avg_pf)  # L/3
        term1 = 1.0 - _safe_div((avg_pf - (2.0/3.0)*league_avg_pf), denom1)

        # 1/6 *(1 - (AvgMaxPF-LPF)/(LMaxPF-LPF))
        denom2 = (league_avg_max_pf - league_avg_pf)
        term2 = 1.0 - _safe_div((avg_max_pf - league_avg_pf), denom2)

        # 1/6 *(1 - (AvgAge-21)/(LAvgAge-21))
        denom3 = (league_avg_age - 21.0)
        term3 = 1.0 - _safe_div((avg_age - 21.0), denom3)

        # 1/6 *(sum picks value)
        term4 = float(pick_sum or 0.0)

        # 1/9 *(future draft capital weights)
        term5 = float(future_cap or 0.0)

        return (1.0/6.0)*term1 + (1.0/6.0)*term2 + (1.0/6.0)*term3 + (1.0/6.0)*term4 + (1.0/9.0)*term5

    def _future_cap_from_traded(traded_picks, roster_id: int, season: int) -> float:
        # weights provided by user
        w = {1: 0.25, 2: 0.09, 3: 0.03, 4: 0.01}
        tot = 0.0
        for tp in traded_picks or []:
            try:
                tp_season = _to_int(tp.get("season"), None)
                if tp_season is None or tp_season <= season:
                    continue
                owner = _to_int(tp.get("owner_id") or tp.get("roster_id"), None)
                if owner is None or owner != roster_id:
                    continue
                rnd = _to_int(tp.get("round"), None)
                if rnd in w:
                    tot += w[rnd]
            except Exception:
                continue
        return tot

    # Build team->roster_id lookup for draft capital
    team_to_roster = {}
    for rid, tm in roster_to_team.items():
        try:
            team_to_roster[str(tm)] = int(rid)
        except Exception:
            continue

    # Per-season pick value (that year's draft picks)
    pick_value_by_team_season = {}
    for p in draft_picks_all or []:
        try:
            y = _to_int(p.get("draft_season"), None)
            if y is None:
                continue
            rid = _to_int(p.get("roster_id"), None)
            if rid is None:
                continue
            team = roster_to_team.get(rid)
            if not team:
                continue
            pick_no = p.get("pick_no")
            if pick_no is None:
                # fall back: approximate overall pick if not provided
                rnd = _to_int(p.get("round"), 0)
                slot = _to_int(p.get("draft_slot"), 0) or _to_int(p.get("pick_in_round"), 0)
                if rnd and slot:
                    pick_no = (rnd - 1) * max(1, len(roster_to_team)) + slot
            pick_no = _to_int(pick_no, None)
            if pick_no is None:
                continue
            # rookie drafts only (heuristic): ignore huge pick numbers
            if pick_no > 1000:
                continue
            val = 1.0 / (float(pick_no) + 1.0)
            pick_value_by_team_season[(str(team), int(y))] = pick_value_by_team_season.get((str(team), int(y)), 0.0) + val
        except Exception:
            continue

    # Team-week tanking: compute season-to-date
    if not tw.empty:
        tw["Tanking"] = pd.to_numeric(tw.get("Tanking"), errors="coerce")

        # weekly team age average (all rostered, starter+bench)
        # Guard against missing/empty age columns.
        if "Age" in pw.columns:
            age_week = pw.groupby(["Team", "Year", "Week"], dropna=False)["Age"].mean().reset_index()
            age_week.rename(columns={"Age": "TeamWeekAvgAge"}, inplace=True)
        else:
            age_week = pd.DataFrame(columns=["Team", "Year", "Week", "TeamWeekAvgAge"])

        tw2 = tw.merge(age_week, on=["Team", "Year", "Week"], how="left")
        tw2["TeamWeekAvgAge"] = pd.to_numeric(tw2["TeamWeekAvgAge"], errors="coerce")

        tanking_rows = []
        for season in sorted(tw2["Year"].dropna().unique()):
            g = tw2[tw2["Year"] == season].copy()
            if g.empty:
                continue

            # league averages season-to-date by week (equal-weight per week)
            g_sorted = g.sort_values("Week").copy()
            # ensure numeric to avoid object-mean failures
            g_sorted["PF"] = pd.to_numeric(g_sorted.get("PF"), errors="coerce")
            g_sorted["Max PF"] = pd.to_numeric(g_sorted.get("Max PF"), errors="coerce")
            g_sorted["TeamWeekAvgAge"] = pd.to_numeric(g_sorted.get("TeamWeekAvgAge"), errors="coerce")

            pf_week = g_sorted.groupby("Week")["PF"].mean().sort_index()
            maxpf_week = g_sorted.groupby("Week")["Max PF"].mean().sort_index()
            age_week_lg = g_sorted.groupby("Week")["TeamWeekAvgAge"].mean().sort_index()

            league_avg_pf_upto = pf_week.expanding().mean()
            league_avg_maxpf_upto = maxpf_week.expanding().mean()
            league_avg_age_upto = age_week_lg.expanding().mean()

            for team, tg in g.groupby("Team"):
                tg = tg.sort_values("Week").copy()
                # expanding means
                pf_exp = pd.to_numeric(tg.get("PF"), errors="coerce").expanding().mean()
                maxpf_exp = pd.to_numeric(tg.get("Max PF"), errors="coerce").expanding().mean()

                # age expanding: use weekly mean age
                age_exp = pd.to_numeric(tg.get("TeamWeekAvgAge"), errors="coerce").expanding().mean()

                rid = team_to_roster.get(str(team), None)
                pick_sum = pick_value_by_team_season.get((str(team), int(season)), 0.0)
                future_cap = _future_cap_from_traded(traded_picks, rid, int(season)) if rid is not None else 0.0

                for i, row in tg.reset_index(drop=True).iterrows():
                    # Week can be missing/NaN in some corrupt rows; guard to avoid crashes.
                    try:
                        wk_int = int(row["Week"]) if pd.notna(row["Week"]) else None
                    except Exception:
                        wk_int = None

                    lg_pf = league_avg_pf_upto.iloc[-1] if len(league_avg_pf_upto) else 0.0
                    lg_mx = league_avg_maxpf_upto.iloc[-1] if len(league_avg_maxpf_upto) else 0.0
                    lg_ag = league_avg_age_upto.iloc[-1] if len(league_avg_age_upto) else 0.0
                    if wk_int is not None:
                        lg_pf = league_avg_pf_upto.get(wk_int, lg_pf)
                        lg_mx = league_avg_maxpf_upto.get(wk_int, lg_mx)
                        lg_ag = league_avg_age_upto.get(wk_int, lg_ag)

                    score = _tanking_score(
                        avg_pf=pf_exp.iloc[i],
                        avg_max_pf=maxpf_exp.iloc[i],
                        avg_age=age_exp.iloc[i],
                        league_avg_pf=lg_pf,
                        league_avg_max_pf=lg_mx,
                        league_avg_age=lg_ag,
                        pick_sum=pick_sum,
                        future_cap=future_cap,
                    )
                    tanking_rows.append((row["Team"], row["Year"], row["Week"], float(score)))

        if tanking_rows:
            tank_df = pd.DataFrame(tanking_rows, columns=["Team", "Year", "Week", "Tanking"])
            tw = tw.drop(columns=["Tanking"], errors="ignore").merge(tank_df, on=["Team", "Year", "Week"], how="left")
        else:
            tw["Tanking"] = 0.0

    
    # ---- Week Name propagation (custom week naming)
    # Use team_week's Week Name where available.
    if (not tw.empty) and ("Week Name" in tw.columns):
        # player_week
        if "Week Name" not in pw.columns:
            pw = pw.merge(tw[["Team","Year","Week","Week Name","Round"]].drop_duplicates(), on=["Team","Year","Week"], how="left")

        # derive a league-wide Week Name per (Year,Week) for league_week rollups
        def _mode_nonnull(vals):
            vals = [v for v in vals if isinstance(v, str) and v and v != "N/A"]
            if not vals:
                return None
            # prefer non-generic labels
            nongeneric = [v for v in vals if not v.startswith("Week ")]
            base = nongeneric if nongeneric else vals
            return pd.Series(base).mode().iloc[0] if len(base) else None

        week_name_global = tw.groupby(["Year","Week"])["Week Name"].apply(_mode_nonnull).reset_index()
    else:
        week_name_global = pd.DataFrame(columns=["Year","Week","Week Name"])

    tx = pd.DataFrame(transactions_rows)
    tr = pd.DataFrame(trades_rows)
    ph = pd.DataFrame(pick_rows)

    # --------------------------
    # Reconstruct draft pick trade history (best-effort) from Sleeper traded_picks.
    # Sleeper traded_picks is per-round (no pick_no), so we apply the same chain to all picks in that round.
    # --------------------------
    try:
        if not ph.empty and 'traded_picks_by_season' in locals():
            # Build per (season, round, original_owner) -> chain of owners (including original)
            round_chain: Dict[tuple, List[int]] = {}
            for season, tps in traded_picks_by_season.items():
                for tp in (tps or []):
                    yr = _to_int(tp.get("season"), season)
                    rnd = _to_int(tp.get("round"), None)
                    prev = _to_int(tp.get("previous_owner_id") or tp.get("previous_owner") or tp.get("previous_owner_roster_id"), None)
                    owner = _to_int(tp.get("owner_id") or tp.get("roster_id") or tp.get("owner_roster_id"), None)
                    if rnd is None or prev is None or owner is None:
                        continue
                    key = (int(yr), int(rnd), int(prev))
                    chain = round_chain.get(key, [int(prev)])
                    if chain[-1] == int(prev) and int(owner) not in chain:
                        chain.append(int(owner))
                    round_chain[key] = chain
            # Apply to pick history rows
            for i, r in ph.iterrows():
                yr = _to_int(r.get("Year"), None)
                num = str(r.get("Number") or "")
                m = re.match(r"R(\d+)\.", num)
                if yr is None or not m:
                    continue
                rnd = int(m.group(1))
                # try to infer original roster_id by matching Original Team name to roster_to_team map for that season (if available)
                orig_team = str(r.get("Original Team") or "")
                orig_rid = None
                # roster_to_team is scoped inside season loop; we can't access it here. Use best-effort: parse "Roster X".
                m2 = re.search(r"(\d+)$", orig_team)
                if m2:
                    orig_rid = int(m2.group(1))
                if orig_rid is None:
                    continue
                key=(int(yr), rnd, orig_rid)
                chain=round_chain.get(key)
                if not chain or len(chain)<=1:
                    continue
                # Fill Trade 1.. with team names if we can, else roster ids
                for j in range(1, min(10, len(chain))):
                    ph.at[i, f"Trade {j}"] = f"Roster {chain[j]}"
    except Exception as e:
        _log_exc(debug, "pick_history_reconstruct", e)


    # --------------------------
    # Player-week derived columns (deltas, tenure, awards) + Hardship
    # --------------------------
    if not pw.empty:
        pw["Year"] = pd.to_numeric(pw["Year"], errors="coerce").astype("Int64")
        pw["Week"] = pd.to_numeric(pw["Week"], errors="coerce").astype("Int64")
        pw["Points"] = pd.to_numeric(pw["Points"], errors="coerce").fillna(0.0)

        pw = pw.sort_values(["Player", "Year", "Week"]).reset_index(drop=True)

        active = ~pw[["Injury?", "Suspension?", "Bye?"]].fillna(False).any(axis=1)

        # Change from previous active week (bench weeks count if active)
        last_active_pts: Dict[str, float] = {}
        out_prev = []
        for i, row in pw.iterrows():
            k = row["Player"]
            out_prev.append((float(row["Points"]) - last_active_pts[k]) if k in last_active_pts else None)
            if bool(active.iloc[i]):
                last_active_pts[k] = float(row["Points"])
        pw["Change from previous week"] = out_prev

        # Previous 5 active weeks avg (spans seasons)
        windows: Dict[str, deque] = defaultdict(lambda: deque(maxlen=5))
        out_prev5 = []
        for i, row in pw.iterrows():
            k = row["Player"]
            q = windows[k]
            out_prev5.append((float(row["Points"]) - (sum(q) / 5)) if len(q) == 5 else None)
            if bool(active.iloc[i]):
                q.append(float(row["Points"]))
        pw["Change from previous 5 weeks avg"] = out_prev5

        # Career avg to that point (active weeks only)
        sums: Dict[str, float] = defaultdict(float)
        counts: Dict[str, int] = defaultdict(int)
        out_cavg = []
        for i, row in pw.iterrows():
            k = row["Player"]
            out_cavg.append((float(row["Points"]) - (sums[k] / counts[k])) if counts[k] > 0 else None)
            if bool(active.iloc[i]):
                sums[k] += float(row["Points"])
                counts[k] += 1
        pw["Change from career average to that point"] = out_cavg

        # Overall career avg (active weeks only)
        try:
            full_avg = pw.loc[active].groupby("Player")["Points"].mean()
            pw["Change from overall career average"] = pw["Points"] - pw["Player"].map(full_avg)
        except Exception as e:
            _log_exc(repo_root / "exports/raw/build_debug.log", "overall_career_avg", e)
            pw["Change from overall career average"] = None

        # Team tenure + bench streaks (bench streak spans seasons)
        pw = pw.sort_values(["Team", "Player", "Year", "Week"]).reset_index(drop=True)
        stats: Dict[Tuple[str, str], Dict[str, Any]] = {}
        for i, row in pw.iterrows():
            key = (str(row["Team"]), str(row["Player"]))
            st = stats.get(key, {
                "weeks": 0,
                "start_all": 0,
                "bench_all": 0,
                "season": None,
                "start_season": 0,
                "bench_season": 0,
                "bench_streak": 0,
                "bench_streak_ex": 0
            })

            if st["season"] != row["Year"]:
                st["season"] = row["Year"]
                st["start_season"] = 0
                st["bench_season"] = 0

            st["weeks"] += 1
            is_starter = (row["Starter/Bench"] == "Starter")
            inactive = bool((row.get("Injury?") or False) or (row.get("Suspension?") or False) or (row.get("Bye?") or False))

            if is_starter:
                pw.at[i, "Number of consecutive weeks on bench before start (if starter)"] = st["bench_streak"]
                pw.at[i, "Number of consecutive weeks on bench before start excluding injury/bye (if starter)"] = st["bench_streak_ex"]
                st["bench_streak"] = 0
                st["bench_streak_ex"] = 0
                st["start_all"] += 1
                st["start_season"] += 1
            else:
                st["bench_all"] += 1
                st["bench_season"] += 1
                st["bench_streak"] += 1
                if not inactive:
                    st["bench_streak_ex"] += 1

            pw.at[i, "Number of weeks on team"] = st["weeks"]
            pw.at[i, "Total weeks as team starter to that point"] = st["start_all"]
            pw.at[i, "Total weeks on bench to that point"] = st["bench_all"]
            pw.at[i, "Total weeks as team starter on that team this season"] = st["start_season"]
            pw.at[i, "Total weeks on bench on that team this season"] = st["bench_season"]

            stats[key] = st

        # Awards (league + team). Ties -> all winners.
        # We compute off pw itself to avoid any mismatched ids.
        pw = pw.sort_values(["Year", "Week", "Team", "Player"]).reset_index(drop=True)

        award_cols = [
            "Player of the week?",
            "QB of the week?",
            "RB of the week?",
            "WR of the week?",
            "TE of the week?",
            "Benchwarmer of the week?",
            "Bench QB of the week?",
            "Bench RB of the week?",
            "Bench WR of the week?",
            "Bench TE of the week?",
            "Highest starter on team?",
            "Lowest starter on team?",
        ]

        def _set_flag(mask, col):
            pw.loc[mask, col] = 1
            pw.loc[~mask, col] = pw.loc[~mask, col].fillna(0)

        # league-level player of week among starters
        starters = pw["Starter/Bench"] == "Starter"
        for (yr, wk), g in pw.groupby(["Year", "Week"]):
            sg = g[starters.loc[g.index]]
            if sg.empty:
                continue
            mx = sg["Points"].max()
            mn = sg["Points"].min()
            _set_flag((pw["Year"] == yr) & (pw["Week"] == wk) & starters & (pw["Points"] == mx), "Player of the week?")
            _set_flag((pw["Year"] == yr) & (pw["Week"] == wk) & starters & (pw["Points"] == mn), "Benchwarmer of the week?")

            for pos, col in [("QB", "QB of the week?"), ("RB", "RB of the week?"), ("WR", "WR of the week?"), ("TE", "TE of the week?")]:
                pg = sg[sg["Position started in (if starter)"].astype(str).str.upper() == pos]
                if pg.empty:
                    continue
                mxp = pg["Points"].max()
                _set_flag(
                    (pw["Year"] == yr)
                    & (pw["Week"] == wk)
                    & starters
                    & (pw["Position started in (if starter)"].astype(str).str.upper() == pos)
                    & (pw["Points"] == mxp),
                    col,
                )

            # bench position awards (bench only)
            bg = g[~starters.loc[g.index]]
            for pos, col in [
                ("QB", "Bench QB of the week?"),
                ("RB", "Bench RB of the week?"),
                ("WR", "Bench WR of the week?"),
                ("TE", "Bench TE of the week?"),
            ]:
                pg = bg[bg["Position started in (if starter)"].astype(str).str.upper() == pos]
                if pg.empty:
                    continue
                mxp = pg["Points"].max()
                _set_flag(
                    (pw["Year"] == yr)
                    & (pw["Week"] == wk)
                    & (~starters)
                    & (pw["Position started in (if starter)"].astype(str).str.upper() == pos)
                    & (pw["Points"] == mxp),
                    col,
                )

            # team-level awards: highest/lowest starter per team per week
            for team, tg in sg.groupby("Team"):
                mx_t = tg["Points"].max()
                mn_t = tg["Points"].min()
                _set_flag(
                    (pw["Year"] == yr)
                    & (pw["Week"] == wk)
                    & (pw["Team"] == team)
                    & starters
                    & (pw["Points"] == mx_t),
                    "Highest starter on team?",
                )
                _set_flag(
                    (pw["Year"] == yr)
                    & (pw["Week"] == wk)
                    & (pw["Team"] == team)
                    & starters
                    & (pw["Points"] == mn_t),
                    "Lowest starter on team?",
                )

        pw[award_cols] = pw[award_cols].fillna(0)

        # Hardship engine per your definition
        # points lost when: Points==0 AND (Injury or Suspension) AND not Bye,
        # expected points = avg of last 5 HEALTHY games (points>0 and not Inj/Susp/Bye).
        pw = pw.sort_values(["Player", "Year", "Week"]).reset_index(drop=True)
        last5: Dict[str, deque] = defaultdict(lambda: deque(maxlen=5))
        exp_points: List[Optional[float]] = [None] * len(pw)
        points_lost: List[float] = [0.0] * len(pw)
        for i, row in pw.iterrows():
            player = row["Player"]
            pts = float(row["Points"])
            inj = bool(row.get("Injury?") or False)
            susp = bool(row.get("Suspension?") or False)
            bye = bool(row.get("Bye?") or False)
            hist = last5[player]
            expected = (sum(hist) / len(hist)) if len(hist) > 0 else None
            exp_points[i] = expected
            missed = (pts == 0.0) and (inj or susp) and (not bye)
            points_lost[i] = float(expected) if (missed and expected is not None) else 0.0
            if (pts > 0.0) and (not inj) and (not susp) and (not bye):
                hist.append(pts)
        pw["_expected_points_if_healthy"] = exp_points
        pw["_points_lost_inj_susp"] = points_lost

    # --------------------------
    # Recompute team-week injury/susp/bye counts and hardship from player-week
    # --------------------------
    if not tw.empty and not pw.empty:
        pw2 = pw.copy()
        pw2["Injury?"] = pw2["Injury?"].fillna(False).astype(bool)
        pw2["Suspension?"] = pw2["Suspension?"].fillna(False).astype(bool)
        pw2["Bye?"] = pw2["Bye?"].fillna(False).astype(bool)
        pw2["Points"] = pd.to_numeric(pw2["Points"], errors="coerce").fillna(0.0)

        pw2["_missed_injury"] = (pw2["Injury?"] & (~pw2["Bye?"]) & (pw2["Points"] == 0)).astype(int)
        pw2["_missed_susp"] = (pw2["Suspension?"] & (~pw2["Bye?"]) & (pw2["Points"] == 0)).astype(int)
        pw2["_on_bye"] = (pw2["Bye?"] & (pw2["Points"] == 0)).astype(int)

        pw2["Starter?"] = (pw2["Starter/Bench"] == "Starter").astype(int)
        pw2["Number_of_players_injured_or_suspended"] = pw2["_missed_injury"] + pw2["_missed_susp"]

        agg = pw2.groupby(["Team", "Year", "Week"], as_index=False).agg(
            Hardship_Points_Lost=("_points_lost_inj_susp", "sum"),
            Number_of_Injuries=("_missed_injury", "sum"),
            Number_of_suspensions=("_missed_susp", "sum"),
            Number_of_players_on_bye=("_on_bye", "sum"),
            Number_of_players_injured_or_suspended=("Number_of_players_injured_or_suspended", "sum"),
            Starter_Count=("Starter?", "sum"),
        )

        tw = tw.merge(agg, how="left", on=["Team", "Year", "Week"])
        # Harden numeric outputs + create friendly display columns (never crash on missing cols)
        for _c in [
            "Hardship_Points_Lost",
            "Number_of_Injuries",
            "Number_of_suspensions",
            "Number_of_players_injured_or_suspended",
            "Number_of_players_on_bye",
            "Starter_Count",
        ]:
            safe_to_numeric(tw, _c, default=0.0)

        tw["Hardship"] = pd.to_numeric(tw.get("Hardship_Points_Lost"), errors="coerce").fillna(0.0)
        tw["Number of Injuries"] = tw["Number_of_Injuries"].round(0).astype(int)
        tw["Number of suspensions"] = tw["Number_of_suspensions"].round(0).astype(int)
        tw["Number of players on bye"] = tw["Number_of_players_on_bye"].round(0).astype(int)

        tw.drop(columns=[
            "Hardship_Points_Lost",
            "Number_of_Injuries",
            "Number_of_suspensions",
            "Number_of_players_injured_or_suspended",
            "Number_of_players_on_bye",
            "Starter_Count",
        ], inplace=True, errors="ignore")

        # UPST: win with lower Max PF than opponent
        if "UPST" not in tw.columns:
            tw["UPST"] = None
        for (yr, wk), g in tw.groupby(["Year", "Week"]):
            g2 = g.copy()
            g2["PF"] = pd.to_numeric(g2["PF"], errors="coerce").fillna(0.0)
            g2["Points against"] = pd.to_numeric(g2["Points against"], errors="coerce").fillna(0.0)
            g2["Max PF"] = pd.to_numeric(g2["Max PF"], errors="coerce")
            for idx, row in g2.iterrows():
                opp = g2[(g2["PF"] == row["Points against"]) & (g2["Points against"] == row["PF"])]
                if len(opp) == 1:
                    opp_max = opp.iloc[0]["Max PF"]
                    if row["Win?"] == 1 and pd.notna(row["Max PF"]) and pd.notna(opp_max):
                        tw.loc[idx, "UPST"] = int(float(row["Max PF"]) < float(opp_max))
                    else:
                        tw.loc[idx, "UPST"] = 0
                else:
                    tw.loc[idx, "UPST"] = 0

        # Brosenzweig / Sisenzweig (correct definition)
        # Brosenzweig: LOSS while 2nd-highest scoring team of the week.
        # Sisenzweig: WIN while 2nd-lowest scoring team of the week.
        tw["Brosenzweig"] = 0
        tw["Sisenzweig"] = 0
        if "PF" in tw.columns and "Win?" in tw.columns:
            tw_pf = tw.copy()
            tw_pf["PF"] = pd.to_numeric(tw_pf["PF"], errors="coerce").fillna(0.0)
            tw_pf["Win?"] = pd.to_numeric(tw_pf["Win?"], errors="coerce")
            for (yr, wk), g in tw_pf.groupby(["Year", "Week"]):
                if g.empty:
                    continue
                # ranks: 1 = highest (desc), 1 = lowest (asc)
                r_desc = g["PF"].rank(method="min", ascending=False)
                r_asc = g["PF"].rank(method="min", ascending=True)
                mask_b = (g["Win?"] == 0) & (r_desc == 2)
                mask_s = (g["Win?"] == 1) & (r_asc == 2)
                tw.loc[g.index[mask_b], "Brosenzweig"] = 1
                tw.loc[g.index[mask_s], "Sisenzweig"] = 1


    # --------------------------
    
        # Fill remaining schema columns in team-week (best-effort)
        try:
            # Largest deficit overcome (no play-by-play available) -> 0 for wins, else None
            if "Largest deficit overcome (if win)" in tw.columns:
                tw.loc[tw["Win?"] == 1, "Largest deficit overcome (if win)"] = 0

            # Cuffs: use player-week activated cuff flag (rostered and started)
            cuff_col = "- Activated Cuff? (Was a player of ... 5 played games injured? Only for players with avg <10 PPG)"
            if (not pw.empty) and (cuff_col in pw.columns):
                pw_c = pw[["Team","Year","Week",cuff_col,"Starter/Bench"]].copy()
                pw_c[cuff_col] = pd.to_numeric(pw_c[cuff_col], errors="coerce").fillna(0.0)
                agg_c = pw_c.groupby(["Team","Year","Week"], as_index=False).agg(
                    **{
                        "Number of cuffs rostered": (cuff_col, "sum"),
                        "Number of cuffs started": (cuff_col, lambda s: float(s[pw_c.loc[s.index,"Starter/Bench"]=="Starter"].sum()) if len(s) else 0.0),
                    }
                )
                tw = tw.merge(agg_c, how="left", on=["Team","Year","Week"], suffixes=("","_c"))
                # fill if missing
                tw["Number of cuffs rostered"] = pd.to_numeric(tw.get("Number of cuffs rostered"), errors="coerce").fillna(0.0).round(0).astype(int)
                tw["Number of cuffs started"] = pd.to_numeric(tw.get("Number of cuffs started"), errors="coerce").fillna(0.0).round(0).astype(int)

            # Future draft capital / startup draft players remaining not yet modeled -> 0 for now
            for col in ["Future draft capital", "Startup draft players remaining"]:
                if col in tw.columns:
                    tw[col] = pd.to_numeric(tw[col], errors="coerce").fillna(0.0)
        except Exception as e:
            _log_exc(debug, "team_week_fill_schema_cols", e)
# Derived team-week columns: pregame avg maxPF diff
    # --------------------------
    if not tw.empty:
        tw["Difference in pregame avg max PF from opponent"] = None
        # build helper: map (Team,Year,Week) -> rid via raw opponent map is not kept, so approximate by using Opponent is label sometimes.
        # We'll compute using team-week history by team: avg MaxPF prior to this week in same season.
        tw = tw.sort_values(["Team", "Year", "Week"]).reset_index(drop=True)
        tw["Max PF"] = pd.to_numeric(tw["Max PF"], errors="coerce")
        # opponent avg maxPF requires opponent team identity; we can't recover from Opponent column in playoffs, so
        # compute against raw points map not in output. We'll approximate using the opponent points against mapping:
        # if Points against is present we can match to unique opponent that week by that score (in 8-team that's safe enough).
        for (yr, wk), g in tw.groupby(["Year", "Week"]):
            # build mapping by (Points against, PF) pair is ambiguous; fallback by symmetric join on Margin
            pass  # leave None if ambiguous (rare)
        # At minimum, compute own pregame avg maxPF, and set diff = own - league avg as a proxy when opponent unknown.
        # This prevents blanks and stays interpretable.
        tw["Pregame avg MaxPF (proxy)"] = None
        for (team, yr), g in tw.groupby(["Team", "Year"]):
            g = g.sort_values("Week")
            maxpfs = []
            for idx, row in g.iterrows():
                pre = (sum(maxpfs) / len(maxpfs)) if maxpfs else None
                tw.loc[idx, "Pregame avg MaxPF (proxy)"] = round(pre, 2) if pre is not None else None
                if not pd.isna(row["Max PF"]):
                    maxpfs.append(float(row["Max PF"]))
        # league avg pregame
        tw["Difference in pregame avg max PF from opponent"] = None
        for (yr, wk), g in tw.groupby(["Year", "Week"]):
            vals = pd.to_numeric(g["Pregame avg MaxPF (proxy)"], errors="coerce").dropna()
            lg_avg = float(vals.mean()) if len(vals) else None
            if lg_avg is None:
                continue
            for idx in g.index:
                v = tw.loc[idx, "Pregame avg MaxPF (proxy)"]
                if v is not None and not pd.isna(v):
                    tw.loc[idx, "Difference in pregame avg max PF from opponent"] = round(float(v) - lg_avg, 2)

    # --------------------------
    
    # --------------------------
    # Player-week: rolling 5-game diffs vs reference player + cuff adjusted diff
    # --------------------------
    if not pw.empty:
        try:
            pw["Year"] = pd.to_numeric(pw["Year"], errors="coerce").astype("Int64")
            pw["Week"] = pd.to_numeric(pw["Week"], errors="coerce").astype("Int64")
            pw["Points"] = pd.to_numeric(pw["Points"], errors="coerce").fillna(0.0)

            # "played games" exclude injury/susp/bye
            played_mask = ~pw[["Injury?", "Suspension?", "Bye?"]].fillna(False).any(axis=1)

            pw_sorted = pw.sort_values(["Player", "Year", "Week"]).reset_index()
            # map (player,year,week)-> rolling avg last5 played (including current if played)
            rolling_avg = {}
            # NOTE: do NOT import deque inside this function.
            # An inner import would make `deque` a local variable for the entire
            # enclosing scope, which breaks earlier lambdas that reference the
            # global `deque` (CI failure: cannot access free variable 'deque').
            hist = defaultdict(lambda: deque(maxlen=5))
            for _, r in pw_sorted.iterrows():
                p=str(r["Player"]); yr=int(r["Year"]) if pd.notna(r["Year"]) else None; wk=int(r["Week"]) if pd.notna(r["Week"]) else None
                if yr is None or wk is None:
                    continue
                key=(p,yr,wk)
                # compute avg of previous played games (last5) BEFORE adding current
                prev=list(hist[(p,yr)])
                avg_prev=float(np.mean(prev)) if prev else None
                # if played, include current for future
                if bool(played_mask.loc[r["index"]]):
                    hist[(p,yr)].append(float(r["Points"]))
                rolling_avg[key]=avg_prev

            def get_avg(p,yr,wk):
                return rolling_avg.get((str(p),int(yr),int(wk)))

            diffs=[]
            cuff_adj=[]
            for _, r in pw.iterrows():
                ref=r.get("Reference player name")
                if not isinstance(ref,str) or ref.strip()=="":
                    diffs.append(None); cuff_adj.append(None); continue
                yr=r.get("Year"); wk=r.get("Week"); player=r.get("Player")
                if pd.isna(yr) or pd.isna(wk):
                    diffs.append(None); cuff_adj.append(None); continue
                avg_p=get_avg(player,yr,wk)
                avg_r=get_avg(ref,yr,wk)
                if (avg_p is None) or (avg_r is None):
                    diffs.append(None); cuff_adj.append(None); continue
                started = (r.get("Starter/Bench") == "Starter")
                diff = (avg_r-avg_p) if started else (avg_p-avg_r)
                diffs.append(round(float(diff),2))
                cuff = float(r.get("- Activated Cuff? (Was a player of ... 5 played games injured? Only for players with avg <10 PPG)") or 0)
                cuff_adj.append(round(float(diff) * (0.5 if cuff else 1.0), 2))
            pw["Difference in averages of best/worst startables over previous 5 games"] = diffs
            pw["Cuff adjusted difference"] = cuff_adj
        except Exception as e:
            _log_exc(debug, "player_week_rolling_diffs", e)

    # --------------------------
    # Team-week: Tanking (best-effort heuristic)
    # --------------------------
    if not tw.empty:
        try:
            tw = tw.sort_values(["Year","Week","Team"]).reset_index(drop=True)
            tw["Max PF"] = pd.to_numeric(tw["Max PF"], errors="coerce")
            tw["Efficiency"] = pd.to_numeric(tw["Efficiency"], errors="coerce")
            tw["Number of transactions"] = pd.to_numeric(tw.get("Number of transactions"), errors="coerce").fillna(0.0)

            # pregame avg maxPF = season-to-date avg maxPF before current week
            tw = tw.sort_values(["Team","Year","Week"]).reset_index(drop=True)
            tw["Pregame Avg Max PF"] = tw.groupby(["Team","Year"])["Max PF"].apply(lambda s: s.shift(1).expanding().mean()).reset_index(level=[0,1], drop=True)

            # rank teams by pregame avg maxPF each week; bottom quartile eligible
            def _bottom_quartile(g):
                vals = pd.to_numeric(g["Pregame Avg Max PF"], errors="coerce")
                if vals.notna().sum() < 4:
                    return pd.Series([False]*len(g), index=g.index)
                thresh = vals.quantile(0.25)
                return vals <= thresh

            tw = tw.sort_values(["Year","Week","Team"]).reset_index(drop=True)
            bq = tw.groupby(["Year","Week"], group_keys=False).apply(_bottom_quartile)
            eff_bad = tw["Efficiency"].fillna(1.0) < 0.85
            tx_some = tw["Number of transactions"].fillna(0.0) >= 1
            tw["Tanking"] = ((bq) & eff_bad & tx_some).astype(int)
            tw.drop(columns=["Pregame Avg Max PF"], inplace=True, errors="ignore")
        except Exception as e:
            _log_exc(debug, "team_week_tanking", e)
# Team-week flags & streaks
    # --------------------------
    if not tw.empty:
        tw["Increase in points from previous week"] = None
        tw["Highest score?"] = 0
        tw["Lowest score?"] = 0
        tw["Narrowest victory?"] = 0
        tw["Largest blowout?"] = 0
        tw["Most efficient?"] = 0
        tw["Least efficient?"] = 0
        tw["Top half of league?"] = 0

        for (yr, wk), g in tw.groupby(["Year", "Week"]):
            g2 = g.copy()
            g2["PF"] = pd.to_numeric(g2["PF"], errors="coerce").fillna(0.0)
            g2["Margin"] = pd.to_numeric(g2["Margin"], errors="coerce")
            g2["Efficiency"] = pd.to_numeric(g2["Efficiency"], errors="coerce")
            if not g2.empty:
                max_pf = g2["PF"].max()
                min_pf = g2["PF"].min()
                tw.loc[(tw["Year"] == yr) & (tw["Week"] == wk) & (tw["PF"] == max_pf), "Highest score?"] = 1
                tw.loc[(tw["Year"] == yr) & (tw["Week"] == wk) & (tw["PF"] == min_pf), "Lowest score?"] = 1

                # narrowest victory (smallest positive margin)
                wins = g2[g2["Margin"] > 0]
                if not wins.empty:
                    min_margin = wins["Margin"].min()
                    tw.loc[(tw["Year"] == yr) & (tw["Week"] == wk) & (tw["Margin"] == min_margin), "Narrowest victory?"] = 1
                    max_margin = wins["Margin"].max()
                    tw.loc[(tw["Year"] == yr) & (tw["Week"] == wk) & (tw["Margin"] == max_margin), "Largest blowout?"] = 1

                # efficiency
                if g2["Efficiency"].notna().any():
                    max_eff = g2["Efficiency"].max()
                    min_eff = g2["Efficiency"].min()
                    tw.loc[(tw["Year"] == yr) & (tw["Week"] == wk) & (tw["Efficiency"] == max_eff), "Most efficient?"] = 1
                    tw.loc[(tw["Year"] == yr) & (tw["Week"] == wk) & (tw["Efficiency"] == min_eff), "Least efficient?"] = 1

                # top half of league by PF
                median_pf = g2["PF"].median()
                tw.loc[(tw["Year"] == yr) & (tw["Week"] == wk) & (tw["PF"] >= median_pf), "Top half of league?"] = 1

        # streaks + increase from previous week
        tw = tw.sort_values(["Team", "Year", "Week"]).reset_index(drop=True)
        tw["Win streak"] = 0
        tw["Loss streak"] = 0
        tw["Win streak counting previous season"] = 0
        tw["Loss streak counting previous season"] = 0
        for team, g in tw.groupby("Team"):
            win_streak = loss_streak = 0
            win_streak_season = loss_streak_season = 0
            current_year = None
            prev_pf_by_year = {}
            for idx, row in g.sort_values(["Year", "Week"]).iterrows():
                if current_year != row["Year"]:
                    current_year = row["Year"]
                    win_streak_season = 0
                    loss_streak_season = 0
                result = row.get("Win?")
                if result == 1:
                    win_streak_season += 1
                    loss_streak_season = 0
                    win_streak += 1
                    loss_streak = 0
                elif result == 0:
                    loss_streak_season += 1
                    win_streak_season = 0
                    loss_streak += 1
                    win_streak = 0
                else:
                    win_streak_season = 0
                    loss_streak_season = 0
                    win_streak = 0
                    loss_streak = 0

                tw.loc[idx, "Win streak"] = win_streak_season
                tw.loc[idx, "Loss streak"] = loss_streak_season
                tw.loc[idx, "Win streak counting previous season"] = win_streak
                tw.loc[idx, "Loss streak counting previous season"] = loss_streak

                prev_pf = prev_pf_by_year.get(row["Year"])
                if prev_pf is not None and pd.notna(row["PF"]):
                    tw.loc[idx, "Increase in points from previous week"] = round(float(row["PF"]) - float(prev_pf), 2)
                prev_pf_by_year[row["Year"]] = row["PF"]

    # --------------------------
    # Rollups: player_year/all_time, team_year/all_time, league_week/year/all_time
    # --------------------------
    if (
        not pw.empty
        and not tw.empty
        and {"Team", "Year", "Week"}.issubset(pw.columns)
        and {"Team", "Year", "Week", "Win?"}.issubset(tw.columns)
    ):
        tw_keys = tw[["Team", "Year", "Week", "Win?"]].copy()
        tw_keys["Team"] = tw_keys["Team"].astype(str)
        for col in ["Year", "Week"]:
            tw_keys[col] = pd.to_numeric(tw_keys[col], errors="coerce").astype("Int64").astype(object)
        win_map = tw_keys.set_index(["Team", "Year", "Week"])["Win?"].to_dict()

        pw_keys = pw[["Team", "Year", "Week"]].copy()
        pw_keys["Team"] = pw_keys["Team"].astype(str)
        for col in ["Year", "Week"]:
            pw_keys[col] = pd.to_numeric(pw_keys[col], errors="coerce").astype("Int64").astype(object)
        pw["Team win?"] = [
            win_map.get((team, year, week))
            for team, year, week in pw_keys.itertuples(index=False, name=None)
        ]

    player_year = pd.DataFrame()
    player_all = pd.DataFrame()
    if not pw.empty:
        pw_work = pw.copy()
        pw_work["Points"] = pd.to_numeric(pw_work["Points"], errors="coerce").fillna(0.0)
        pw_work["Missed_injury"] = (pw_work["Injury?"].fillna(False) & (pw_work["Points"] == 0)).astype(int)
        pw_work["Missed_suspension"] = (pw_work["Suspension?"].fillna(False) & (pw_work["Points"] == 0)).astype(int)
        pw_work["Starter?"] = (pw_work["Starter/Bench"] == "Starter").astype(int)

        award_cols = [
            "Player of the week?",
            "QB of the week?",
            "RB of the week?",
            "WR of the week?",
            "TE of the week?",
            "Benchwarmer of the week?",
            "Bench QB of the week?",
            "Bench RB of the week?",
            "Bench WR of the week?",
            "Bench TE of the week?",
            "Highest starter on team?",
            "Lowest starter on team?",
        ]

        pw_work[award_cols] = pw_work[award_cols].fillna(0)

        team_points = pw_work.groupby(["Player", "Year", "Team"], as_index=False)["Points"].sum()
        top_team = (
            team_points.sort_values(["Player", "Year", "Points"], ascending=[True, True, False])
            .drop_duplicates(["Player", "Year"])
            .rename(columns={"Team": "Top Team", "Points": "Top Team Points"})
        )
        last_team = (
            pw_work.sort_values(["Player", "Year", "Week"])
            .groupby(["Player", "Year"])
            .tail(1)[["Player", "Year", "Team"]]
            .rename(columns={"Team": "Last team"})
        )

        py_base = pw_work.groupby(["Player", "Year"], as_index=False).agg(
            Points=("Points", "sum"),
            Avg_points=("Points", "mean"),
            Weeks_missed_injury=("Missed_injury", "sum"),
            Weeks_missed_suspension=("Missed_suspension", "sum"),
            Weeks_as_starter=("Starter?", "sum"),
            Number_of_teams=("Team", "nunique"),
            Weeks=("Points", "count"),
            **{c: (c, "sum") for c in award_cols},
        )

        py = py_base.merge(top_team[["Player", "Year", "Top Team"]], on=["Player", "Year"], how="left")
        py = py.merge(last_team, on=["Player", "Year"], how="left")

        team_points_all = pw_work.groupby(["Player", "Year", "Team"])["Points"].sum()
        total_points = pw_work.groupby(["Player", "Year"])["Points"].sum()
        max_share = (team_points_all.groupby(["Player", "Year"]).max() / total_points).rename("% of points (highest team)")
        min_share = (team_points_all.groupby(["Player", "Year"]).min() / total_points).rename("% of points (lowest team)")
        py = py.merge(max_share.reset_index(), on=["Player", "Year"], how="left")
        py = py.merge(min_share.reset_index(), on=["Player", "Year"], how="left")

        py = py.sort_values(["Player", "Year"]).reset_index(drop=True)
        py["Change in points from previous season"] = py.groupby("Player")["Points"].diff()
        py["Change in avg points from previous season"] = py.groupby("Player")["Avg_points"].diff()

        py["Career_points_before"] = py.groupby("Player")["Points"].cumsum().shift(1)
        py["Career_years_before"] = py.groupby("Player").cumcount()
        py["Change in points from career"] = py.apply(
            lambda r: (r["Points"] - (r["Career_points_before"] / r["Career_years_before"]))
            if r["Career_years_before"] and r["Career_points_before"] is not None
            else None,
            axis=1,
        )

        py["Career_points_before_total"] = py.groupby("Player")["Points"].cumsum().shift(1)
        py["Career_weeks_before_total"] = py.groupby("Player")["Weeks"].cumsum().shift(1)
        py["Change in avg points from career"] = py.apply(
            lambda r: (r["Avg_points"] - (r["Career_points_before_total"] / r["Career_weeks_before_total"]))
            if r["Career_weeks_before_total"] and r["Career_points_before_total"] is not None
            else None,
            axis=1,
        )

        py = py.rename(
            columns={
                "Avg_points": "Avg points",
                "Weeks_missed_injury": "Weeks missed due to injury",
                "Weeks_missed_suspension": "Weeks missed due to suspension",
                "Weeks_as_starter": "Weeks as starter",
                "Number_of_teams": "Number of teams",
                "Top Team": "Top Team",
            }
        )

        py["Number of transactions"] = 0
        py["Number of trades"] = 0

        py = py.rename(
            columns={
                "Player of the week?": "Times as Player of the week?",
                "QB of the week?": "Times as QB of the week?",
                "RB of the week?": "Times as RB of the week?",
                "WR of the week?": "Times as WR of the week?",
                "TE of the week?": "Times as TE of the week?",
                "Benchwarmer of the week?": "Times as Benchwarmer of the week?",
                "Bench QB of the week?": "Times as Bench QB of the week?",
                "Bench RB of the week?": "Times as Bench RB of the week?",
                "Bench WR of the week?": "Times as Bench WR of the week?",
                "Bench TE of the week?": "Times as Bench TE of the week?",
                "Highest starter on team?": "Times as Highest starter on team?",
                "Lowest starter on team?": "Times as Lowest starter on team?",
            }
        )

        player_year = py

        pa = pw_work.groupby(["Player"], as_index=False).agg(
            Points=("Points", "sum"),
            Avg_points=("Points", "mean"),
            Weeks_missed_injury=("Missed_injury", "sum"),
            Weeks_missed_suspension=("Missed_suspension", "sum"),
            Weeks_as_starter=("Starter?", "sum"),
            Number_of_teams=("Team", "nunique"),
            **{c: (c, "sum") for c in award_cols},
        )

        top_team_all = (
            pw_work.groupby(["Player", "Team"], as_index=False)["Points"].sum()
            .sort_values(["Player", "Points"], ascending=[True, False])
            .drop_duplicates(["Player"])
            .rename(columns={"Team": "Top team", "Points": "Top team points"})
        )
        last_team_all = (
            pw_work.sort_values(["Year", "Week"])
            .groupby("Player")
            .tail(1)[["Player", "Team", "Rookie?", "Age"]]
            .rename(columns={"Team": "Last team"})
        )
        team_points_all_time = pw_work.groupby(["Player", "Team"])["Points"].sum()
        total_points_all_time = pw_work.groupby(["Player"])["Points"].sum()
        max_share_all = (team_points_all_time.groupby("Player").max() / total_points_all_time).rename("% of points (highest team)")
        min_share_all = (team_points_all_time.groupby("Player").min() / total_points_all_time).rename("% of points (lowest team)")

        pa = pa.merge(top_team_all[["Player", "Top team"]], on="Player", how="left")
        pa = pa.merge(last_team_all, on="Player", how="left")
        pa = pa.merge(max_share_all.reset_index(), on="Player", how="left")
        pa = pa.merge(min_share_all.reset_index(), on="Player", how="left")

        pa = pa.rename(
            columns={
                "Avg_points": "Avg points",
                "Weeks_missed_injury": "Weeks missed due to injury",
                "Weeks_missed_suspension": "Weeks missed due to suspension",
                "Weeks_as_starter": "Weeks as starter",
                "Number_of_teams": "Number of teams",
                "Player of the week?": "Times as Player of the week?",
                "QB of the week?": "Times as QB of the week?",
                "RB of the week?": "Times as RB of the week?",
                "WR of the week?": "Times as WR of the week?",
                "TE of the week?": "Times as TE of the week?",
                "Benchwarmer of the week?": "Times as Benchwarmer of the week?",
                "Bench QB of the week?": "Times as Bench QB of the week?",
                "Bench RB of the week?": "Times as Bench RB of the week?",
                "Bench WR of the week?": "Times as Bench WR of the week?",
                "Bench TE of the week?": "Times as Bench TE of the week?",
                "Highest starter on team?": "Times as Highest starter on team?",
                "Lowest starter on team?": "Times as Lowest starter on team?",
            }
        )

        pa["Number of transactions"] = 0
        pa["Number of trades"] = 0

        player_all = pa

    # Team-year: compute record and vs records using raw opp_rid_map (still available in closures above? not anymore)
    team_year = pd.DataFrame()
    team_all = pd.DataFrame()
    if not tw.empty:
        # reconstruct team list
        teams = sorted(tw["Team"].dropna().astype(str).unique().tolist())
        # Ensure head-to-head columns exist in the export schema.
        # The plan file doesn't include these dynamic fields, but the workbook expects them.
        try:
            for plan_key in ("team-year", "team-all-time"):
                cols = catalog.get(plan_key, [])
                if not cols:
                    continue
                # insert after the main Record column if present; otherwise append.
                insert_at = cols.index("Record") + 1 if "Record" in cols else len(cols)
                h2h_cols: List[str] = []
                for opp in teams:
                    h2h_cols.append(f"Record vs {opp}")
                    h2h_cols.append(f"Win % vs {opp}")
                # add only missing
                for c in h2h_cols:
                    if c not in cols:
                        cols.insert(insert_at, c)
                        insert_at += 1
                catalog[plan_key] = cols
        except Exception as e:
            _log_exc(debug, "catalog_h2h_extend", e)
        # compute per game outcomes using raw opponent team when available.
        game_rows = []
        for (yr, wk), g in tw.groupby(["Year", "Week"]):
            g2 = g.copy()
            g2["PF"] = pd.to_numeric(g2["PF"], errors="coerce").fillna(0.0)
            g2["Points against"] = pd.to_numeric(g2["Points against"], errors="coerce").fillna(0.0)
            for idx, row in g2.iterrows():
                opp = row.get("Opponent Team (raw)")
                if not opp or pd.isna(opp):
                    match = g2[g2["PF"] == row["Points against"]]
                    if len(match) == 1:
                        opp = str(match.iloc[0]["Team"])
                    elif len(match) > 1:
                        match2 = match[match["Points against"] == row["PF"]]
                        if len(match2) == 1:
                            opp = str(match2.iloc[0]["Team"])
                if opp:
                    game_rows.append({
                        "Year": int(yr),
                        "Week": int(wk),
                        "Team": str(row["Team"]),
                        "OppTeam": str(opp),
                        "Win?": row.get("Win?"),
                        "PF": float(row["PF"]),
                        "PA": float(row["Points against"]),
                    })
        games_df = pd.DataFrame(game_rows).drop_duplicates(subset=["Year","Week","Team"])

        def _record_str(w, l, t=0):
            return f"{int(w)}-{int(l)}" + (f"-{int(t)}" if t else "")

        def _record_from_games(df: pd.DataFrame) -> Tuple[int, int, int]:
            w = int((df["Win?"] == 1).sum())
            l = int((df["Win?"] == 0).sum())
            t = int((df["Win?"] == 0.5).sum())
            return w, l, t

        playoff_teams_by_season: Dict[int, set] = {}
        champion_by_season: Dict[int, Optional[str]] = {}
        last_place_by_season: Dict[int, Optional[str]] = {}
        for yr, g in tw.groupby("Year"):
            season = int(yr)
            playoff_start = playoff_start_by_season.get(season)
            reg = g.copy()
            if playoff_start:
                reg = reg[pd.to_numeric(reg["Week"], errors="coerce") < playoff_start]
            reg["PF"] = pd.to_numeric(reg["PF"], errors="coerce").fillna(0.0)
            reg["Win?"] = pd.to_numeric(reg["Win?"], errors="coerce")
            standings = []
            for team, tg in reg.groupby("Team"):
                wins = int((tg["Win?"] == 1).sum())
                losses = int((tg["Win?"] == 0).sum())
                ties = int((tg["Win?"] == 0.5).sum())
                pf = float(tg["PF"].sum())
                standings.append((team, wins, losses, ties, pf))
            standings.sort(key=lambda x: (x[1] + 0.5 * x[3], x[4]), reverse=True)
            playoff_teams_by_season[season] = set([t for t, *_ in standings[:4]])
            last_place_by_season[season] = standings[-1][0] if standings else None
            champ = None
            if "Week label" in g.columns:
                finals = g[g["Week label"] == "Final"]
                champ_row = finals[finals["Win?"] == 1]
                if not champ_row.empty:
                    champ = str(champ_row.iloc[0]["Team"])
            if not champ and standings:
                champ = standings[0][0]
            champion_by_season[season] = champ

        
        # Determine season finishing positions (Result) from playoff/toilet brackets when available.
        season_finish: Dict[int, Dict[str, str]] = {}
        try:
            for yr, g in tw.groupby("Year"):
                season = int(yr)
                playoff_start = playoff_start_by_season.get(season)
                if not playoff_start:
                    continue
                finals_week = playoff_start + 1
                fin_map: Dict[str, str] = {}
                # Finals
                gf = tw[(tw["Year"]==season) & (tw["Week"]==finals_week) & (tw["Week label"]=="Final")].copy()
                gf["PF"] = pd.to_numeric(gf["PF"], errors="coerce").fillna(0.0)
                if len(gf)==2:
                    gf = gf.sort_values("PF", ascending=False)
                    fin_map[str(gf.iloc[0]["Team"])] = "champion"
                    fin_map[str(gf.iloc[1]["Team"])] = "2nd"
                # 3rd place
                g3 = tw[(tw["Year"]==season) & (tw["Week"]==finals_week) & (tw["Week label"]=="3rd Place")].copy()
                g3["PF"] = pd.to_numeric(g3["PF"], errors="coerce").fillna(0.0)
                if len(g3)==2:
                    g3=g3.sort_values("PF", ascending=False)
                    fin_map[str(g3.iloc[0]["Team"])] = "3rd"
                    fin_map[str(g3.iloc[1]["Team"])] = "4th"

                # Non-playoff finishes (5th-8th) are based on regular-season record cutoff,
                # with PF as the tiebreaker. (Pre-2025: through 17 games; 2025+: through 15 games.)
                cutoff = 17 if season < 2025 else 15
                try:
                    all_teams = [str(t) for t in tw[tw["Year"] == season]["Team"].dropna().unique().tolist()]
                    playoff_teams = set([t for t, r in fin_map.items() if r in ("champion", "2nd", "3rd", "4th")])
                    non_playoff = [t for t in all_teams if t not in playoff_teams]
                    if non_playoff and (not games_df.empty):
                        reg = games_df[(games_df["Year"] == season) & (games_df["Week"] <= cutoff)].copy()
                        reg["PF"] = pd.to_numeric(reg.get("PF", 0.0), errors="coerce").fillna(0.0)
                        reg["Win?"] = pd.to_numeric(reg.get("Win?", 0.0), errors="coerce").fillna(0.0)
                        sub = reg[reg["Team"].astype(str).isin(non_playoff)]
                        rows_np = []
                        for team_np, gg in sub.groupby(sub["Team"].astype(str)):
                            w = int((gg["Win?"] == 1).sum())
                            l = int((gg["Win?"] == 0).sum())
                            t_ = int((gg["Win?"] == 0.5).sum())
                            pf_sum = float(gg["PF"].sum())
                            rows_np.append((team_np, w, l, t_, pf_sum))
                        # Sort: record (wins desc, losses asc), PF desc
                        rows_np.sort(key=lambda x: (-x[1], x[2], -x[4]))
                        place = 5
                        for team_np, *_ in rows_np:
                            if place == 5:
                                fin_map[team_np] = "5th"
                            elif place == 6:
                                fin_map[team_np] = "6th"
                            elif place == 7:
                                fin_map[team_np] = "7th"
                            elif place == 8:
                                fin_map[team_np] = "8th"
                            place += 1
                            if place > 8:
                                break
                except Exception:
                    pass
                season_finish[season] = fin_map
        except Exception as e:
            _log_exc(debug, "season_finish_map", e)

        # team-year rollup
        rows = []
        for (team, yr), g in tw.groupby(["Team", "Year"]):
            wins = int((g["Win?"] == 1).sum())
            losses = int((g["Win?"] == 0).sum())
            ties = int((g["Win?"] == 0.5).sum())
            gp = max(1, wins + losses + ties)
            pf = float(pd.to_numeric(g["PF"], errors="coerce").fillna(0.0).sum())
            pa = float(pd.to_numeric(g["Points against"], errors="coerce").fillna(0.0).sum())
            diff = pf - pa
            maxpf_sum = float(pd.to_numeric(g["Max PF"], errors="coerce").fillna(0.0).sum())
            maxpf_avg = float(pd.to_numeric(g["Max PF"], errors="coerce").fillna(0.0).mean())
            rec = _record_str(wins, losses, ties)
            winp = round((wins + 0.5 * ties) / gp, 4)
            record_vs = "N/A"
            if not games_df.empty:
                sub = games_df[(games_df["Team"] == str(team)) & (games_df["Year"] == int(yr))]
                pieces = []
                for opp in [t for t in teams if t != str(team)]:
                    sg = sub[sub["OppTeam"] == opp]
                    w = int((sg["Win?"] == 1).sum())
                    l = int((sg["Win?"] == 0).sum())
                    t_ = int((sg["Win?"] == 0.5).sum())
                    gp2 = max(0, w + l + t_)
                    winp2 = round((w + 0.5 * t_) / gp2, 4) if gp2 else 0.0
                    pieces.append(f"{opp}: {_record_str(w, l, t_)} ({winp2})")
                record_vs = "; ".join(pieces) if pieces else "N/A"

            row = {
                "Team": str(team),
                "Year": int(yr),
                "Result": "N/A",
                "Win %": winp,
                "Record": rec,
                "Record & win % vs each team": record_vs,
                "Record & win % vs playoff teams": "N/A",
                "Record & win % vs non-playoff teams": "N/A",
                "Record & win % vs champion": "N/A",
                "Record & win % vs last place": "N/A",
                "Change in win % from previous season": None,
                "Win Variance": float(pd.to_numeric(g["Win?"], errors="coerce").fillna(0.0).var()),
                "Week of playoff elimination": "N/A",
                "Draft Value": 0,
                "Number of first round picks made": 0,
                "Total number of picks made": 0,
                "Points": round(pf, 2),
                "Avg points": round(pf / gp, 2),
                "Points against": round(pa, 2),
                "Avg points against": round(pa / gp, 2),
                "Differential": round(diff, 2),
                "Avg differential": round(diff / gp, 2),
                "Max PF": round(maxpf_sum, 2),
                "Avg max PF": round(maxpf_avg, 2) if not math.isnan(maxpf_avg) else None,
                "Efficiency": round(pf / maxpf_sum, 4) if maxpf_sum else None,
                "Weeks of injuries": int(pd.to_numeric(g.get("Number of Injuries"), errors="coerce").fillna(0.0).sum()),
                "Weeks suspensions": int(pd.to_numeric(g.get("Number of suspensions"), errors="coerce").fillna(0.0).sum()),
                "Hardship": float(pd.to_numeric(g.get("Hardship"), errors="coerce").fillna(0.0).sum()),
                "Offseason starter turnover": 0,
                "Inseason starter turnover": 0,
            }
            rows.append(row)
        team_year = pd.DataFrame(rows)

        # Compute starter/roster turnover metrics using unique players
        try:
            if not pw.empty and "Starter/Bench" in pw.columns:
                pw_t = pw.copy()
                pw_t["Week"] = pd.to_numeric(pw_t["Week"], errors="coerce")
                # helper to get set of players for (team,year,week) for starters/roster
                def _set_for(team, year, week, starters_only):
                    df = pw_t[(pw_t["Team"]==team) & (pw_t["Year"]==year) & (pw_t["Week"]==week)]
                    if starters_only:
                        df = df[df["Starter/Bench"].astype(str).str.lower().eq("starter")]
                    return set(df["Player"].dropna().astype(str).tolist())
                # precompute weeks per (team,year)
                for (team, year), g in pw_t.groupby(["Team","Year"]):
                    weeks = sorted([int(w) for w in g["Week"].dropna().unique().tolist()])
                    if not weeks:
                        continue
                    first_w = weeks[0]
                    last_w = weeks[-1]
                    s_first = _set_for(team, year, first_w, True)
                    s_last = _set_for(team, year, last_w, True)
                    r_first = _set_for(team, year, first_w, False)
                    r_last = _set_for(team, year, last_w, False)
                    # Use removed-count so turnover is bounded by spot count
                    in_s = len(s_first - s_last)
                    in_r = len(r_first - r_last)
                    team_year.loc[(team_year["Team"]==team)&(team_year["Year"]==year), "Inseason starter turnover"] = in_s
                    team_year.loc[(team_year["Team"]==team)&(team_year["Year"]==year), "Inseason roster turnover"] = in_r
                    # offseason vs previous season
                    prev_year = int(year)-1
                    if ((pw_t["Team"]==team)&(pw_t["Year"]==prev_year)).any():
                        gprev = pw_t[(pw_t["Team"]==team)&(pw_t["Year"]==prev_year)]
                        prev_weeks = sorted([int(w) for w in gprev["Week"].dropna().unique().tolist()])
                        if prev_weeks:
                            prev_last = prev_weeks[-1]
                            s_prev = _set_for(team, prev_year, prev_last, True)
                            r_prev = _set_for(team, prev_year, prev_last, False)
                            off_s = len(s_prev - s_first)
                            off_r = len(r_prev - r_first)
                            team_year.loc[(team_year["Team"]==team)&(team_year["Year"]==year), "Offseason starter turnover"] = off_s
                            team_year.loc[(team_year["Team"]==team)&(team_year["Year"]==year), "Offseason roster turnover"] = off_r
        except Exception as e:
            _log_exc(debug, "turnover_metrics_team_year", e)

        # --------------------------
        # Fill missing Team-year columns from team-week (flags, tanking, luck, roster composition, etc.)
        # --------------------------
        try:
            agg_year = tw.groupby(["Team", "Year"], as_index=False).agg(
                **{
                    "Tanking": ("Tanking", "sum"),
                    "Luck": ("Luck", "sum"),
                    "Times Brosenzweig": ("Brosenzweig", "sum"),
                    "Times Sisenzweig": ("Sisenzweig", "sum"),
                    "Times Highest score?": ("Highest score?", "sum"),
                    "Times Lowest score?": ("Lowest score?", "sum"),
                    "Times Narrowest victory?": ("Narrowest victory?", "sum"),
                    "Times Largest blowout?": ("Largest blowout?", "sum"),
                    "Times Most efficient?": ("Most efficient?", "sum"),
                    "Times Least efficient?": ("Least efficient?", "sum"),
                    "Times Top half of league?": ("Top half of league?", "sum"),
                    "Number of QB started": ("Number of QB started", "sum"),
                    "Number of WR started": ("Number of WR started", "sum"),
                    "Number of RB started": ("Number of RB started", "sum"),
                    "Number of TE started": ("Number of TE started", "sum"),
                    "Number of QB rostered": ("Number of QB rostered", "sum"),
                    "Number of WR rostered": ("Number of WR rostered", "sum"),
                    "Number of RB rostered": ("Number of RB rostered", "sum"),
                    "Number of TE rostered": ("Number of TE rostered", "sum"),
                    "Most number of players started from same NFL team": ("Most number of players started from same NFL team", "max"),
                    "Most number of players rostered from same NFL team": ("Most number of players rostered from same NFL team", "max"),
                    "Most number of QBs started from same NFL team": ("Most number of QBs started from same NFL team", "max"),
                    "Most number of QBs rostered from same NFL team": ("Most number of QBs rostered from same NFL team", "max"),
                    "Most number of RBs started from same NFL team": ("Most number of RBs started from same NFL team", "max"),
                    "Most number of RBs rostered from same NFL team": ("Most number of RBs rostered from same NFL team", "max"),
                    "Most number of WR started from same NFL team": ("Most number of WR started from same NFL team", "max"),
                    "Most number of WR rostered from same NFL team": ("Most number of WR rostered from same NFL team", "max"),
                    "Most number of TE started from same NFL team": ("Most number of TE started from same NFL team", "max"),
                    "Most number of TE rostered from same NFL team": ("Most number of TE rostered from same NFL team", "max"),
                    "Number of NFL teams among starting players": ("Number of NFL teams among starting players", "max"),
                    "Number of NFL teams amoung rostered players": ("Number of NFL teams amoung rostered players", "max"),
                    "Number of rookies started": ("Number of rookies started", "sum"),
                    "Number of rookies rostered": ("Number of rookies rostered", "sum"),
                    "Player average age": ("Player average age", "mean"),
                    "Difference between highest and lowest starters": ("Difference between highest and lowest starters", "max"),
                    "Number of donuts": ("Number of donuts", "sum"),
                    "Number of players under 10": ("Number of players under 10", "sum"),
                    "Number of players over 20": ("Number of players over 20", "sum"),
                    "Number of players over 30": ("Number of players over 30", "sum"),
                    "Number of players over 40": ("Number of players over 40", "sum"),
                    "Number of players over 50": ("Number of players over 50", "sum"),
                    "Number of cuffs rostered": ("Number of cuffs rostered", "sum"),
                    "Number of cuffs started": ("Number of cuffs started", "sum"),
                }
            )
            team_year = team_year.merge(agg_year, how="left", on=["Team", "Year"])
        except Exception as e:
            _log_exc(debug, "team_year_aggregate_fill", e)


        # Unique-player rollups for positional counts (avoid counting the same player in multiple weeks).
        # IMPORTANT: do not sum weekly counts; count unique players over the season.
        try:
            if not pw.empty and "Position" in pw.columns and "Starter/Bench" in pw.columns:
                pw_u = pw.copy()
                pw_u["Position"] = pw_u["Position"].astype(str).str.upper()
                pw_u["Starter/Bench"] = pw_u["Starter/Bench"].astype(str)

                rostered = pw_u.groupby(["Team", "Year", "Position"], as_index=False)["Player"].nunique()
                started = pw_u[pw_u["Starter/Bench"].str.lower().eq("starter")].groupby(["Team", "Year", "Position"], as_index=False)["Player"].nunique()

                # Build lookup maps keyed by (team,year)
                r_map = {(str(t), int(y), str(p)): int(n) for t, y, p, n in rostered.itertuples(index=False, name=None)}
                s_map = {(str(t), int(y), str(p)): int(n) for t, y, p, n in started.itertuples(index=False, name=None)}

                for pos in ["QB", "RB", "WR", "TE"]:
                    team_year[f"Number of {pos} rostered"] = team_year.apply(lambda r: r_map.get((str(r["Team"]), int(r["Year"]), pos), 0), axis=1)
                    team_year[f"Number of {pos} started"] = team_year.apply(lambda r: s_map.get((str(r["Team"]), int(r["Year"]), pos), 0), axis=1)
        except Exception as e:
            _log_exc(debug, "team_year_unique_player_counts", e)


        # Result + vs-category records (best-effort using games_df + playoff/champion/last_place sets)
        try:

            # Fill Result from bracket-derived finish map when available.
            if "Result" in team_year.columns and season_finish:
                def _res(team, year):
                    m = season_finish.get(int(year), {})
                    return m.get(str(team))
                team_year["Result"] = team_year.apply(lambda r: _res(r["Team"], r["Year"]), axis=1)

            # Head-to-head records vs each team (per season)
            try:
                if not games_df.empty:
                    teams_in_year = team_year.groupby("Year")["Team"].apply(lambda s: sorted(s.dropna().astype(str).unique().tolist())).to_dict()
                    for year, teams_list in teams_in_year.items():
                        g = games_df[games_df["Year"]==int(year)]
                        if g.empty:
                            continue
                        for team in teams_list:
                            for opp in teams_list:
                                if opp == team:
                                    continue
                                gg = g[(g["Team"]==team) & (g["OppTeam"]==opp)]
                                if gg.empty:
                                    rec = "0-0-0"
                                    wp = None
                                else:
                                    w = int((pd.to_numeric(gg["Win?"], errors="coerce")==1).sum())
                                    l = int((pd.to_numeric(gg["Win?"], errors="coerce")==0).sum())
                                    t = int((pd.to_numeric(gg["Win?"], errors="coerce")==0.5).sum())
                                    gp = max(1, w+l+t)
                                    rec = f"{w}-{l}-{t}"
                                    wp = round((w + 0.5*t)/gp, 4)
                                team_year.loc[(team_year["Year"]==int(year)) & (team_year["Team"]==team), f"Record vs {opp}"] = rec
                                team_year.loc[(team_year["Year"]==int(year)) & (team_year["Team"]==team), f"Win % vs {opp}"] = wp
            except Exception as e:
                _log_exc(debug, "h2h_team_year", e)

            def _fmt_rec(wlt):
                w, l, t = wlt
                return f"{int(w)}-{int(l)}-{int(t)}"

            def _wlt_for(team: str, yr: int, opp_filter):
                if games_df.empty:
                    return (0, 0, 0)
                sub = games_df[(games_df["Year"] == yr) & (games_df["Team"] == team)]
                if opp_filter is not None:
                    sub = sub[sub["Opponent"].isin(opp_filter)]
                return (
                    int((sub["Win?"] == 1).sum()),
                    int((sub["Win?"] == 0).sum()),
                    int((sub["Win?"] == 0.5).sum()),
                )

            results = []
            for _, r in team_year.iterrows():
                team = str(r["Team"])
                yr = int(r["Year"])
                playoffs = playoff_teams_by_season.get(yr, set())
                champ = champion_by_season.get(yr)
                lastp = last_place_by_season.get(yr)

                if champ and team == champ:
                    res = "Champion"
                elif lastp and team == lastp:
                    res = "Last place"
                elif team in playoffs:
                    res = "Playoffs"
                else:
                    res = "Missed playoffs"

                wlt_play = _wlt_for(team, yr, playoffs if playoffs else None)
                wlt_non = _wlt_for(team, yr, set(team_year[team_year["Year"] == yr]["Team"]) - set(playoffs) if playoffs else None)
                wlt_champ = _wlt_for(team, yr, {champ} if champ else None)
                wlt_last = _wlt_for(team, yr, {lastp} if lastp else None)

                def winpct(wlt):
                    w, l, t = wlt
                    gp = max(1, w + l + t)
                    return round((w + 0.5 * t) / gp, 4)

                results.append({
                    "Team": team,
                    "Year": yr,
                    "Result": res,
                    "Record vs playoff teams": _fmt_rec(wlt_play) if playoffs else "N/A",
                    "Win % vs playoff teams": winpct(wlt_play) if playoffs else None,
                    "Record vs non-playoff teams": _fmt_rec(wlt_non) if playoffs else "N/A",
                    "Win % vs non-playoff teams": winpct(wlt_non) if playoffs else None,
                    "Record vs champion": _fmt_rec(wlt_champ) if champ else "N/A",
                    "Win % vs champion": winpct(wlt_champ) if champ else None,
                    "Record vs last place": _fmt_rec(wlt_last) if lastp else "N/A",
                    "Win % vs last place": winpct(wlt_last) if lastp else None,
                })
            extra = pd.DataFrame(results)
            team_year = team_year.drop(columns=[c for c in extra.columns if c in team_year.columns and c not in ["Team","Year"]], errors="ignore")
            team_year = team_year.merge(extra, how="left", on=["Team","Year"])
            team_year["Week of playoff elimination"] = team_year.get("Week of playoff elimination", "N/A")
            team_year["Offseason roster turnover"] = team_year.get("Offseason roster turnover", 0).fillna(0)
            team_year["Inseason roster turnover"] = team_year.get("Inseason roster turnover", 0).fillna(0)
        except Exception as e:
            _log_exc(debug, "team_year_results_vs_records", e)

        team_year = team_year.sort_values(["Team", "Year"]).reset_index(drop=True)
        team_year["Change in win % from previous season"] = team_year.groupby("Team")["Win %"].diff()

        # team-all-time rollup
        rows = []
        for team, g in tw.groupby("Team"):
            wins = int((g["Win?"] == 1).sum())
            losses = int((g["Win?"] == 0).sum())
            ties = int((g["Win?"] == 0.5).sum())
            gp = max(1, wins + losses + ties)
            pf = float(pd.to_numeric(g["PF"], errors="coerce").fillna(0.0).sum())
            pa = float(pd.to_numeric(g["Points against"], errors="coerce").fillna(0.0).sum())
            diff = pf - pa
            maxpf_sum = float(pd.to_numeric(g["Max PF"], errors="coerce").fillna(0.0).sum())
            maxpf_avg = float(pd.to_numeric(g["Max PF"], errors="coerce").fillna(0.0).mean())
            record_vs = "N/A"
            if not games_df.empty:
                sub = games_df[games_df["Team"] == str(team)]
                pieces = []
                for opp in [t for t in teams if t != str(team)]:
                    sg = sub[sub["OppTeam"] == opp]
                    w = int((sg["Win?"] == 1).sum())
                    l = int((sg["Win?"] == 0).sum())
                    t_ = int((sg["Win?"] == 0.5).sum())
                    gp2 = max(0, w + l + t_)
                    winp2 = round((w + 0.5 * t_) / gp2, 4) if gp2 else 0.0
                    pieces.append(f"{opp}: {_record_str(w, l, t_)} ({winp2})")
                record_vs = "; ".join(pieces) if pieces else "N/A"

            row = {
                "Team": str(team),
                "All time win %": round((wins + 0.5 * ties) / gp, 4),
                "All time record": _record_str(wins, losses, ties),
                "Record & win % vs each team": record_vs,
                "Record & win % vs playoff teams": "N/A",
                "Record & win % vs non-playoff teams": "N/A",
                "Record & win % vs champions": "N/A",
                "Record & win % vs last place": "N/A",
                "Win Variance": float(pd.to_numeric(g["Win?"], errors="coerce").fillna(0.0).var()),
                "Draft Value": 0,
                "Number of first round picks made": 0,
                "Total number of picks made": 0,
                "Points": round(pf, 2),
                "Avg points": round(pf / gp, 2),
                "Points against": round(pa, 2),
                "Avg points against": round(pa / gp, 2),
                "Differential": round(diff, 2),
                "Avg differential": round(diff / gp, 2),
                "Max PF": round(maxpf_sum, 2),
                "Avg max PF": round(maxpf_avg, 2) if not math.isnan(maxpf_avg) else None,
                "Efficiency": round(pf / maxpf_sum, 4) if maxpf_sum else None,
                "Weeks of injuries": int(pd.to_numeric(g.get("Number of Injuries"), errors="coerce").fillna(0.0).sum()),
                "Weeks suspensions": int(pd.to_numeric(g.get("Number of suspensions"), errors="coerce").fillna(0.0).sum()),
                "Hardship": float(pd.to_numeric(g.get("Hardship"), errors="coerce").fillna(0.0).sum()),
                "Offseason starter turnover": 0,
                "Inseason starter turnover": 0,
                "Offseason roster turnover": 0,
                "Inseason roster turnover": 0,
                "Tanking": float(pd.to_numeric(g.get("Tanking"), errors="coerce").fillna(0.0).sum()),
                "Luck": float(pd.to_numeric(g.get("Luck"), errors="coerce").fillna(0.0).sum()),
            }
            rows.append(row)
        team_all = pd.DataFrame(rows)

        # --------------------------
        # Fill missing Team-all-time columns from team-week (flags, roster composition, etc.)
        # --------------------------
        try:
            agg_all = tw.groupby("Team", as_index=False).agg(
                **{
                    "Times Brosenzweig": ("Brosenzweig", "sum"),
                    "Times Sisenzweig": ("Sisenzweig", "sum"),
                    "Times Highest score?": ("Highest score?", "sum"),
                    "Times Lowest score?": ("Lowest score?", "sum"),
                    "Times Narrowest victory?": ("Narrowest victory?", "sum"),
                    "Times Largest blowout?": ("Largest blowout?", "sum"),
                    "Times Most efficient?": ("Most efficient?", "sum"),
                    "Times Least efficient?": ("Least efficient?", "sum"),
                    "Times Top half of league?": ("Top half of league?", "sum"),
                    "Number of QB started": ("Number of QB started", "sum"),
                    "Number of WR started": ("Number of WR started", "sum"),
                    "Number of RB started": ("Number of RB started", "sum"),
                    "Number of TE started": ("Number of TE started", "sum"),
                    "Number of QB rostered": ("Number of QB rostered", "sum"),
                    "Number of WR rostered": ("Number of WR rostered", "sum"),
                    "Number of RB rostered": ("Number of RB rostered", "sum"),
                    "Number of TE rostered": ("Number of TE rostered", "sum"),
                    "Most number of players started from same NFL team": ("Most number of players started from same NFL team", "max"),
                    "Most number of players rostered from same NFL team": ("Most number of players rostered from same NFL team", "max"),
                    "Most number of QBs started from same NFL team": ("Most number of QBs started from same NFL team", "max"),
                    "Most number of QBs rostered from same NFL team": ("Most number of QBs rostered from same NFL team", "max"),
                    "Most number of RBs started from same NFL team": ("Most number of RBs started from same NFL team", "max"),
                    "Most number of RBs rostered from same NFL team": ("Most number of RBs rostered from same NFL team", "max"),
                    "Most number of WR started from same NFL team": ("Most number of WR started from same NFL team", "max"),
                    "Most number of WR rostered from same NFL team": ("Most number of WR rostered from same NFL team", "max"),
                    "Most number of TE started from same NFL team": ("Most number of TE started from same NFL team", "max"),
                    "Most number of TE rostered from same NFL team": ("Most number of TE rostered from same NFL team", "max"),
                    "Number of NFL teams among starting players": ("Number of NFL teams among starting players", "max"),
                    "Number of NFL teams amoung rostered players": ("Number of NFL teams amoung rostered players", "max"),
                    "Number of rookies started": ("Number of rookies started", "sum"),
                    "Number of rookies rostered": ("Number of rookies rostered", "sum"),
                    "Player average age": ("Player average age", "mean"),
                    "Difference between highest and lowest starters": ("Difference between highest and lowest starters", "max"),
                    "Combined matchup score": ("Combined matchup score", "max"),
                    "Number of donuts": ("Number of donuts", "sum"),
                    "Number of players under 10": ("Number of players under 10", "sum"),
                    "Number of players over 20": ("Number of players over 20", "sum"),
                    "Number of players over 30": ("Number of players over 30", "sum"),
                    "Number of players over 40": ("Number of players over 40", "sum"),
                    "Number of players over 50": ("Number of players over 50", "sum"),
                    "Number of cuffs rostered": ("Number of cuffs rostered", "sum"),
                    "Number of cuffs started": ("Number of cuffs started", "sum"),
                }
            )
            team_all = team_all.merge(agg_all, how="left", on="Team")
        except Exception as e:
            _log_exc(debug, "team_all_aggregate_fill", e)

        # vs-category records (all-time)
        try:
            def _fmt_rec(wlt):
                w, l, t = wlt
                return f"{int(w)}-{int(l)}-{int(t)}"

            def _wlt_all(team: str, opp_filter):
                if games_df.empty:
                    return (0, 0, 0)
                sub = games_df[games_df["Team"] == team]
                if opp_filter is not None:
                    sub = sub[sub["Opponent"].isin(opp_filter)]
                return (
                    int((sub["Win?"] == 1).sum()),
                    int((sub["Win?"] == 0).sum()),
                    int((sub["Win?"] == 0.5).sum()),
                )

            # compute sets across seasons
            all_teams = set(team_all["Team"].astype(str).tolist())
            champs = {c for c in champion_by_season.values() if c}
            lastps = {c for c in last_place_by_season.values() if c}
            playoffs = set().union(*[set(s) for s in playoff_teams_by_season.values()]) if playoff_teams_by_season else set()
            nonplayoffs = all_teams - playoffs if playoffs else None

            extra=[]
            for team in team_all["Team"].astype(str).tolist():
                wlt_play=_wlt_all(team, playoffs if playoffs else None)
                wlt_non=_wlt_all(team, nonplayoffs) if nonplayoffs is not None else (0,0,0)
                wlt_ch=_wlt_all(team, champs if champs else None)
                wlt_last=_wlt_all(team, lastps if lastps else None)

                def winpct(wlt):
                    w,l,t=wlt
                    gp=max(1,w+l+t)
                    return round((w+0.5*t)/gp,4)
                extra.append({
                    "Team": team,
                    "Record vs playoff teams": _fmt_rec(wlt_play) if playoffs else "N/A",
                    "Win % vs playoff teams": winpct(wlt_play) if playoffs else None,
                    "Record vs non-playoff teams": _fmt_rec(wlt_non) if nonplayoffs is not None else "N/A",
                    "Win % vs non-playoff teams": winpct(wlt_non) if nonplayoffs is not None else None,
                    "Record vs champions": _fmt_rec(wlt_ch) if champs else "N/A",
                    "Win % vs champions": winpct(wlt_ch) if champs else None,
                    "Record vs last place": _fmt_rec(wlt_last) if lastps else "N/A",
                    "Win % vs last place": winpct(wlt_last) if lastps else None,
                })
            extra=pd.DataFrame(extra)
            team_all = team_all.drop(columns=[c for c in extra.columns if c in team_all.columns and c!="Team"], errors="ignore")
            team_all = team_all.merge(extra, how="left", on="Team")
        except Exception as e:
            _log_exc(debug, "team_all_vs_records", e)

        # Head-to-head totals vs each opponent (all-time)
        try:
            if not games_df.empty and not team_all.empty:
                for team in team_all["Team"].dropna().astype(str).unique().tolist():
                    for opp in team_all["Team"].dropna().astype(str).unique().tolist():
                        if team == opp:
                            continue
                        gg = games_df[(games_df["Team"] == team) & (games_df["Opponent"] == opp)]
                        if gg.empty:
                            rec = "0-0-0"
                            wp = None
                        else:
                            w = int((pd.to_numeric(gg["Win?"], errors="coerce") == 1).sum())
                            l = int((pd.to_numeric(gg["Win?"], errors="coerce") == 0).sum())
                            t = int((pd.to_numeric(gg["Win?"], errors="coerce") == 0.5).sum())
                            gp = max(1, w + l + t)
                            rec = f"{w}-{l}-{t}"
                            wp = round((w + 0.5 * t) / gp, 4)
                        team_all.loc[team_all["Team"] == team, f"Record vs {opp}"] = rec
                        team_all.loc[team_all["Team"] == team, f"Win % vs {opp}"] = wp
        except Exception as e:
            _log_exc(debug, "h2h_team_all", e)

    # League rollups
    league_week = pd.DataFrame()
    league_year = pd.DataFrame()
    league_all = pd.DataFrame()
    if not tw.empty:
        g_week = tw.copy()
        g_week["PF"] = pd.to_numeric(g_week["PF"], errors="coerce").fillna(0.0)
        g_week["Margin"] = pd.to_numeric(g_week["Margin"], errors="coerce")
        g_week["Efficiency"] = pd.to_numeric(g_week["Efficiency"], errors="coerce")
        g_week["Max PF"] = pd.to_numeric(g_week["Max PF"], errors="coerce").fillna(0.0)

        rows = []
        for (yr, wk), g in g_week.groupby(["Year", "Week"]):
            margin_abs = g["Margin"].abs()
            rows.append({
                "Year": int(yr),
                "Week": int(wk),
                "PF": float(g["PF"].sum()),
                "PF Range": float(g["PF"].max() - g["PF"].min()) if not g.empty else 0.0,
                "Avg margin": (float(g.loc[pd.to_numeric(g.get("Win?"), errors="coerce") == 1, "Margin"].mean()) if (pd.to_numeric(g.get("Win?"), errors="coerce") == 1).any() else None),
                "Margin range": (float(g.loc[pd.to_numeric(g.get("Win?"), errors="coerce") == 1, "Margin"].max() - g.loc[pd.to_numeric(g.get("Win?"), errors="coerce") == 1, "Margin"].min()) if (pd.to_numeric(g.get("Win?"), errors="coerce") == 1).any() else None),
                "Number of games within 10": int((margin_abs <= 10).sum() / 2),
                "Number of games within 5": int((margin_abs <= 5).sum() / 2),
                "Max PF": float(g["Max PF"].sum()),
                "Efficiency": float(g["Efficiency"].mean()) if g["Efficiency"].notna().any() else None,
                "Number of Injuries": int(pd.to_numeric(g.get("Number of Injuries"), errors="coerce").fillna(0.0).sum()),
                "Number of suspensions": int(pd.to_numeric(g.get("Number of suspensions"), errors="coerce").fillna(0.0).sum()),
                "Number of players on bye": int(pd.to_numeric(g.get("Number of players on bye"), errors="coerce").fillna(0.0).sum()),
                "Starter turnover from previous week": float(pd.to_numeric(g.get("Starter turnover from previous week"), errors="coerce").fillna(0.0).mean()),
                "Number of wins with pregame avg max PF from opponent": int(pd.to_numeric(g.get("UPST"), errors="coerce").fillna(0.0).sum()),
                "UPST": int(pd.to_numeric(g.get("UPST"), errors="coerce").fillna(0.0).sum()),
                "Hardship": float(pd.to_numeric(g.get("Hardship"), errors="coerce").fillna(0.0).sum()),
                "Tanking": float(pd.to_numeric(g.get("Tanking"), errors="coerce").fillna(0.0).sum()),
                "Luck": float(pd.to_numeric(g.get("Luck"), errors="coerce").fillna(0.0).sum()),
                "Increase in points from previous week": float(pd.to_numeric(g.get("Increase in points from previous week"), errors="coerce").fillna(0.0).sum()),
                "Number of QB started": int(pd.to_numeric(g.get("Number of QB started"), errors="coerce").fillna(0.0).sum()),
                "Number of WR started": int(pd.to_numeric(g.get("Number of WR started"), errors="coerce").fillna(0.0).sum()),
                "Number of RB started": int(pd.to_numeric(g.get("Number of RB started"), errors="coerce").fillna(0.0).sum()),
                "Number of TE started": int(pd.to_numeric(g.get("Number of TE started"), errors="coerce").fillna(0.0).sum()),
                "Number of QB rostered": int(pd.to_numeric(g.get("Number of QB rostered"), errors="coerce").fillna(0.0).sum()),
                "Number of WR rostered": int(pd.to_numeric(g.get("Number of WR rostered"), errors="coerce").fillna(0.0).sum()),
                "Number of RB rostered": int(pd.to_numeric(g.get("Number of RB rostered"), errors="coerce").fillna(0.0).sum()),
                "Number of TE rostered": int(pd.to_numeric(g.get("Number of TE rostered"), errors="coerce").fillna(0.0).sum()),
                "Number of transactions": int(pd.to_numeric(g.get("Number of transactions"), errors="coerce").fillna(0.0).sum()),
                "Number of trades": int(pd.to_numeric(g.get("Number of trades"), errors="coerce").fillna(0.0).sum()),
            })
        league_week = pd.DataFrame(rows)

        # Attach Week Name (custom week naming) if available
        try:
            if 'week_name_global' in locals() and (not week_name_global.empty):
                league_week = league_week.merge(week_name_global, on=["Year","Week"], how="left")
        except Exception as e:
            _log_exc(debug, "league_week_week_name", e)


        # Fill additional league-week columns from team-week (schema completeness)
        try:
            agg_lw = g_week.groupby(["Year","Week"], as_index=False).agg(
                **{
                    "Most number of players started from same NFL team": ("Most number of players started from same NFL team", "max"),
                    "Most number of players rostered from same NFL team": ("Most number of players rostered from same NFL team", "max"),
                    "Most number of QBs started from same NFL team": ("Most number of QBs started from same NFL team", "max"),
                    "Most number of QBs rostered from same NFL team": ("Most number of QBs rostered from same NFL team", "max"),
                    "Most number of RBs started from same NFL team": ("Most number of RBs started from same NFL team", "max"),
                    "Most number of RBs rostered from same NFL team": ("Most number of RBs rostered from same NFL team", "max"),
                    "Most number of WR started from same NFL team": ("Most number of WR started from same NFL team", "max"),
                    "Most number of WR rostered from same NFL team": ("Most number of WR rostered from same NFL team", "max"),
                    "Most number of TE started from same NFL team": ("Most number of TE started from same NFL team", "max"),
                    "Most number of TE rostered from same NFL team": ("Most number of TE rostered from same NFL team", "max"),
                    "Number of NFL teams among starting players": ("Number of NFL teams among starting players", "max"),
                    "Number of NFL teams amoung rostered players": ("Number of NFL teams amoung rostered players", "max"),
                    "Number of rookies started": ("Number of rookies started", "sum"),
                    "Number of rookies rostered": ("Number of rookies rostered", "sum"),
                    "Player average age": ("Player average age", "mean"),
                    "Difference between highest and lowest starters": ("Difference between highest and lowest starters", "max"),
                    "Number of donuts": ("Number of donuts", "sum"),
                    "Number of players under 10": ("Number of players under 10", "sum"),
                    "Number of players over 20": ("Number of players over 20", "sum"),
                    "Number of players over 30": ("Number of players over 30", "sum"),
                    "Number of players over 40": ("Number of players over 40", "sum"),
                    "Number of players over 50": ("Number of players over 50", "sum"),
                    "Number of cuffs rostered": ("Number of cuffs rostered", "sum"),
                    "Number of cuffs started": ("Number of cuffs started", "sum"),
                    "Startup draft players remaining": ("Startup draft players remaining", "max"),
                }
            )
            league_week = league_week.merge(agg_lw, how="left", on=["Year","Week"])
            league_week["Amount of FAAB spent"] = 0
        except Exception as e:
            _log_exc(debug, "league_week_fill_extra", e)

        rows = []
        for yr, g in g_week.groupby("Year"):
            margin_abs = g["Margin"].abs()
            rows.append({
                "Year": int(yr),
                "(smallest) Playoff tiebreaker": "N/A",
                "PF": float(g["PF"].sum()),
                "Avg PF": float(g["PF"].mean()) if g["PF"].notna().any() else None,
                "PF Range": float(g["PF"].max() - g["PF"].min()) if g["PF"].notna().any() else None,
                "Avg margin": float(g["Margin"].mean()) if g["Margin"].notna().any() else None,
                "Margin range": float(g["Margin"].max() - g["Margin"].min()) if g["Margin"].notna().any() else None,
                "Number of games within 10": int((margin_abs <= 10).sum() / 2),
                "Number of games within 5": int((margin_abs <= 5).sum() / 2),
                "Max PF": float(g["Max PF"].sum()),
                "Avg max PF": float(g["Max PF"].mean()) if g["Max PF"].notna().any() else None,
                "Efficiency": float(g["Efficiency"].mean()) if g["Efficiency"].notna().any() else None,
                "Number of weeks missed due to injury": int(pd.to_numeric(g.get("Number of Injuries"), errors="coerce").fillna(0.0).sum()),
                "Number of weeks missed due to suspensions": int(pd.to_numeric(g.get("Number of suspensions"), errors="coerce").fillna(0.0).sum()),
                "Inseason starter turnover": float(pd.to_numeric(g.get("Starter turnover from previous week"), errors="coerce").fillna(0.0).mean()),
                "Offseason starter turnover": 0,
                "Inseason roster turnover": 0,
                "Offseason roster turnover": 0,
                "Number of wins with pregame avg max PF from opponent": int(pd.to_numeric(g.get("UPST"), errors="coerce").fillna(0.0).sum()),
                "Tanking": float(pd.to_numeric(g.get("Tanking"), errors="coerce").fillna(0.0).sum()),
                "Luck": float(pd.to_numeric(g.get("Luck"), errors="coerce").fillna(0.0).sum()),
                "Increase in points from previous week": float(pd.to_numeric(g.get("Increase in points from previous week"), errors="coerce").fillna(0.0).sum()),
                "Number of QB started": int(pd.to_numeric(g.get("Number of QB started"), errors="coerce").fillna(0.0).sum()),
                "Number of WR started": int(pd.to_numeric(g.get("Number of WR started"), errors="coerce").fillna(0.0).sum()),
                "Number of RB started": int(pd.to_numeric(g.get("Number of RB started"), errors="coerce").fillna(0.0).sum()),
                "Number of TE started": int(pd.to_numeric(g.get("Number of TE started"), errors="coerce").fillna(0.0).sum()),
                "Number of QB rostered": int(pd.to_numeric(g.get("Number of QB rostered"), errors="coerce").fillna(0.0).sum()),
                "Number of WR rostered": int(pd.to_numeric(g.get("Number of WR rostered"), errors="coerce").fillna(0.0).sum()),
                "Number of RB rostered": int(pd.to_numeric(g.get("Number of RB rostered"), errors="coerce").fillna(0.0).sum()),
                "Number of TE rostered": int(pd.to_numeric(g.get("Number of TE rostered"), errors="coerce").fillna(0.0).sum()),
            })
        league_year = pd.DataFrame(rows)

        # Fill additional league-year columns from team-week
        try:
            agg_ly = g_week.groupby(["Year"], as_index=False).agg(
                **{
                    "Most number of players started from same NFL team": ("Most number of players started from same NFL team", "max"),
                    "Most number of players rostered from same NFL team": ("Most number of players rostered from same NFL team", "max"),
                    "Most number of QBs started from same NFL team": ("Most number of QBs started from same NFL team", "max"),
                    "Most number of QBs rostered from same NFL team": ("Most number of QBs rostered from same NFL team", "max"),
                    "Most number of RBs started from same NFL team": ("Most number of RBs started from same NFL team", "max"),
                    "Most number of RBs rostered from same NFL team": ("Most number of RBs rostered from same NFL team", "max"),
                    "Most number of WR started from same NFL team": ("Most number of WR started from same NFL team", "max"),
                    "Most number of WR rostered from same NFL team": ("Most number of WR rostered from same NFL team", "max"),
                    "Most number of TE started from same NFL team": ("Most number of TE started from same NFL team", "max"),
                    "Most number of TE rostered from same NFL team": ("Most number of TE rostered from same NFL team", "max"),
                    "Number of NFL teams among starting players": ("Number of NFL teams among starting players", "max"),
                    "Number of NFL teams amoung rostered players": ("Number of NFL teams amoung rostered players", "max"),
                    "Number of rookies started": ("Number of rookies started", "sum"),
                    "Number of rookies rostered": ("Number of rookies rostered", "sum"),
                    "Player average age": ("Player average age", "mean"),
                    "Difference between highest and lowest starters": ("Difference between highest and lowest starters", "max"),
                    "Number of donuts": ("Number of donuts", "sum"),
                    "Number of players under 10": ("Number of players under 10", "sum"),
                    "Number of players over 20": ("Number of players over 20", "sum"),
                    "Number of players over 30": ("Number of players over 30", "sum"),
                    "Number of players over 40": ("Number of players over 40", "sum"),
                    "Number of players over 50": ("Number of players over 50", "sum"),
                    "Number of cuffs rostered": ("Number of cuffs rostered", "sum"),
                    "Number of cuffs started": ("Number of cuffs started", "sum"),
                    "Startup draft players remaining": ("Startup draft players remaining", "max"),
                }
            )
            league_year = league_year.merge(agg_ly, how="left", on=["Year"])
            league_year["Amount of FAAB spent"] = 0
        except Exception as e:
            _log_exc(debug, "league_year_fill_extra", e)

        league_all = pd.DataFrame([{
            "PF": float(pd.to_numeric(g_week["PF"], errors="coerce").fillna(0.0).sum()),
            "Avg PF": float(pd.to_numeric(g_week["PF"], errors="coerce").fillna(0.0).mean()),
            "PF Range": float(g_week["PF"].max() - g_week["PF"].min()) if g_week["PF"].notna().any() else None,
            "Avg margin": float(g_week["Margin"].mean()) if g_week["Margin"].notna().any() else None,
            "Margin range": float(g_week["Margin"].max() - g_week["Margin"].min()) if g_week["Margin"].notna().any() else None,
            "Number of games within 10": int((g_week["Margin"].abs() <= 10).sum() / 2),
            "Number of games within 5": int((g_week["Margin"].abs() <= 5).sum() / 2),
            "Max PF": float(pd.to_numeric(g_week["Max PF"], errors="coerce").fillna(0.0).sum()),
            "Avg max PF": float(pd.to_numeric(g_week["Max PF"], errors="coerce").fillna(0.0).mean()),
            "Efficiency": float(pd.to_numeric(g_week["Efficiency"], errors="coerce").dropna().mean()) if g_week["Efficiency"].notna().any() else None,
            "Number of weeks missed due to injury": int(pd.to_numeric(g_week.get("Number of Injuries"), errors="coerce").fillna(0.0).sum()),
            "Number of weeks missed due to suspensions": int(pd.to_numeric(g_week.get("Number of suspensions"), errors="coerce").fillna(0.0).sum()),
            "Number of wins with pregame avg max PF from opponent": int(pd.to_numeric(g_week.get("UPST"), errors="coerce").fillna(0.0).sum()),
            "Tanking": float(pd.to_numeric(g_week.get("Tanking"), errors="coerce").fillna(0.0).sum()),
            "Luck": float(pd.to_numeric(g_week.get("Luck"), errors="coerce").fillna(0.0).sum()),
            "Increase in points from previous week": float(pd.to_numeric(g_week.get("Increase in points from previous week"), errors="coerce").fillna(0.0).sum()),
            "Number of QB started": int(pd.to_numeric(g_week.get("Number of QB started"), errors="coerce").fillna(0.0).sum()),
            "Number of WR started": int(pd.to_numeric(g_week.get("Number of WR started"), errors="coerce").fillna(0.0).sum()),
            "Number of RB started": int(pd.to_numeric(g_week.get("Number of RB started"), errors="coerce").fillna(0.0).sum()),
            "Number of TE started": int(pd.to_numeric(g_week.get("Number of TE started"), errors="coerce").fillna(0.0).sum()),
            "Number of QB rostered": int(pd.to_numeric(g_week.get("Number of QB rostered"), errors="coerce").fillna(0.0).sum()),
            "Number of WR rostered": int(pd.to_numeric(g_week.get("Number of WR rostered"), errors="coerce").fillna(0.0).sum()),
            "Number of RB rostered": int(pd.to_numeric(g_week.get("Number of RB rostered"), errors="coerce").fillna(0.0).sum()),
            "Number of TE rostered": int(pd.to_numeric(g_week.get("Number of TE rostered"), errors="coerce").fillna(0.0).sum()),
            "Number of transactions": int(pd.to_numeric(g_week.get("Number of transactions"), errors="coerce").fillna(0.0).sum()),
            "Number of trades": int(pd.to_numeric(g_week.get("Number of trades"), errors="coerce").fillna(0.0).sum()),
            "Amount of FAAB spent": float(pd.to_numeric(g_week.get("Amount of FAAB spent"), errors="coerce").fillna(0.0).sum()),
            "Most number of players started from same NFL team": float(pd.to_numeric(g_week.get("Most number of players started from same NFL team"), errors="coerce").fillna(0.0).max()),
            "Most number of players rostered from same NFL team": float(pd.to_numeric(g_week.get("Most number of players rostered from same NFL team"), errors="coerce").fillna(0.0).max()),
            "Most number of QBs started from same NFL team": float(pd.to_numeric(g_week.get("Most number of QBs started from same NFL team"), errors="coerce").fillna(0.0).max()),
        }])

        # Fill additional league-all-time columns from team-week
        try:
            league_all["Most number of players started from same NFL team"] = float(pd.to_numeric(g_week.get("Most number of players started from same NFL team"), errors="coerce").max())
            league_all["Most number of players rostered from same NFL team"] = float(pd.to_numeric(g_week.get("Most number of players rostered from same NFL team"), errors="coerce").max())
            league_all["Most number of QBs started from same NFL team"] = float(pd.to_numeric(g_week.get("Most number of QBs started from same NFL team"), errors="coerce").max())
            league_all["Most number of QBs rostered from same NFL team"] = float(pd.to_numeric(g_week.get("Most number of QBs rostered from same NFL team"), errors="coerce").max())
            league_all["Most number of RBs started from same NFL team"] = float(pd.to_numeric(g_week.get("Most number of RBs started from same NFL team"), errors="coerce").max())
            league_all["Most number of RBs rostered from same NFL team"] = float(pd.to_numeric(g_week.get("Most number of RBs rostered from same NFL team"), errors="coerce").max())
            league_all["Most number of WR started from same NFL team"] = float(pd.to_numeric(g_week.get("Most number of WR started from same NFL team"), errors="coerce").max())
            league_all["Most number of WR rostered from same NFL team"] = float(pd.to_numeric(g_week.get("Most number of WR rostered from same NFL team"), errors="coerce").max())
            league_all["Most number of TE started from same NFL team"] = float(pd.to_numeric(g_week.get("Most number of TE started from same NFL team"), errors="coerce").max())
            league_all["Most number of TE rostered from same NFL team"] = float(pd.to_numeric(g_week.get("Most number of TE rostered from same NFL team"), errors="coerce").max())
            league_all["Number of NFL teams among starting players"] = float(pd.to_numeric(g_week.get("Number of NFL teams among starting players"), errors="coerce").max())
            league_all["Number of NFL teams amoung rostered players"] = float(pd.to_numeric(g_week.get("Number of NFL teams amoung rostered players"), errors="coerce").max())
            league_all["Number of rookies started"] = float(pd.to_numeric(g_week.get("Number of rookies started"), errors="coerce").sum())
            league_all["Number of rookies rostered"] = float(pd.to_numeric(g_week.get("Number of rookies rostered"), errors="coerce").sum())
            league_all["Player average age"] = float(pd.to_numeric(g_week.get("Player average age"), errors="coerce").mean())
            league_all["Difference between highest and lowest starters"] = float(pd.to_numeric(g_week.get("Difference between highest and lowest starters"), errors="coerce").max())
            league_all["Number of donuts"] = float(pd.to_numeric(g_week.get("Number of donuts"), errors="coerce").sum())
            league_all["Number of players under 10"] = float(pd.to_numeric(g_week.get("Number of players under 10"), errors="coerce").sum())
            league_all["Number of players over 20"] = float(pd.to_numeric(g_week.get("Number of players over 20"), errors="coerce").sum())
            league_all["Number of players over 30"] = float(pd.to_numeric(g_week.get("Number of players over 30"), errors="coerce").sum())
            league_all["Number of players over 40"] = float(pd.to_numeric(g_week.get("Number of players over 40"), errors="coerce").sum())
            league_all["Number of players over 50"] = float(pd.to_numeric(g_week.get("Number of players over 50"), errors="coerce").sum())
            league_all["Number of cuffs rostered"] = float(pd.to_numeric(g_week.get("Number of cuffs rostered"), errors="coerce").sum())
            league_all["Number of cuffs started"] = float(pd.to_numeric(g_week.get("Number of cuffs started"), errors="coerce").sum())
            league_all["Startup draft players remaining"] = float(pd.to_numeric(g_week.get("Startup draft players remaining"), errors="coerce").max())
            league_all["Amount of FAAB spent"] = 0
        except Exception as e:
            _log_exc(debug, "league_all_fill_extra", e)


    # --------------------------
    # Write outputs (schema contract)
    # --------------------------
    
    # --------------------------
    # Transactions / Trades: link columns + tanking (best-effort)
    # --------------------------
    try:
        # normalize date
        if not tx.empty and "Date" in tx.columns:
            tx["Date"] = pd.to_datetime(tx["Date"], errors="coerce", utc=True)
            tx = tx.sort_values(["Team","Date"]).reset_index(drop=True)
            # link columns as row numbers (1-indexed) for easy navigation
            tx["Link to previous transaction"] = tx.groupby("Team").cumcount().replace(0, np.nan)
            tx["Link to next transaction"] = tx.groupby("Team").cumcount().shift(-1) + 2
            tx.loc[tx.groupby("Team").tail(1).index, "Link to next transaction"] = np.nan

            # tanking before/after based on team-year tanking sum (0/1+) mapped from tx date year
            if not team_year.empty:
                ty_map = team_year.set_index(["Team","Year"])["Tanking"].to_dict()
                tx_year = tx["Date"].dt.year
                tx["Tanking before"] = [float(ty_map.get((str(t), int(y)), 0)) for t,y in zip(tx["Team"], tx_year)]
                tx["Tanking after"] = tx["Tanking before"]
        else:
            if "Tanking before" in tx.columns:
                tx["Tanking before"] = pd.to_numeric(tx["Tanking before"], errors="coerce").fillna(0.0)
            if "Tanking after" in tx.columns:
                tx["Tanking after"] = pd.to_numeric(tx["Tanking after"], errors="coerce").fillna(0.0)
    except Exception as e:
        _log_exc(debug, "transactions_links_tanking", e)

    try:
        if not tr.empty and "Date" in tr.columns:
            tr["Date"] = pd.to_datetime(tr["Date"], errors="coerce", utc=True)
            tr = tr.sort_values(["Team","Date"]).reset_index(drop=True)
            tr["Link to previous transaction"] = tr.groupby("Team").cumcount().replace(0, np.nan)
            tr["Link to next transaction"] = tr.groupby("Team").cumcount().shift(-1) + 2
            tr.loc[tr.groupby("Team").tail(1).index, "Link to next transaction"] = np.nan

            if not team_year.empty:
                ty_map = team_year.set_index(["Team","Year"])["Tanking"].to_dict()
                tr_year = tr["Date"].dt.year
                tr["Tanking before"] = [float(ty_map.get((str(t), int(y)), 0)) for t,y in zip(tr["Team"], tr_year)]
                tr["Tanking after"] = tr["Tanking before"]
    except Exception as e:
        _log_exc(debug, "trades_links_tanking", e)

    # --------------------------
    # Final schema tweaks (dynamic columns + ordering)
    # --------------------------
    try:
        # Ensure player_year/player_all_time carry a stable Position column (most common in that span)
        if "Position" in pw.columns:
            if "Position" not in player_year.columns and not player_year.empty:
                pos_map = (
                    pw.dropna(subset=["Player","Year"])
                      .groupby(["Player","Year"])["Position"]
                      .agg(lambda s: (s.dropna().mode().iloc[0] if not s.dropna().mode().empty else None))
                )
                player_year = player_year.merge(pos_map.rename("Position"), on=["Player","Year"], how="left")
            if "Position" not in player_all.columns and not player_all.empty:
                pos_map2 = (
                    pw.dropna(subset=["Player"])
                      .groupby(["Player"])["Position"]
                      .agg(lambda s: (s.dropna().mode().iloc[0] if not s.dropna().mode().empty else None))
                )
                player_all = player_all.merge(pos_map2.rename("Position"), on=["Player"], how="left")

        # Unique-player rollups for team_year/team_all_time positional counts
        if not pw.empty and "Position" in pw.columns and "Starter/Bench" in pw.columns:
            pw_u = pw.copy()
            pw_u["StarterFlag"] = pw_u["Starter/Bench"].astype(str).str.lower().eq("starter")
            for pos in ["QB","RB","WR","TE"]:
                started_col = f"Number of {pos} started"
                rostered_col = f"Number of {pos} rostered"
                started = (
                    pw_u[(pw_u["Position"]==pos) & (pw_u["StarterFlag"])]
                      .groupby(["Team","Year"])["Player"].nunique()
                      .rename(started_col)
                )
                rostered = (
                    pw_u[pw_u["Position"]==pos]
                      .groupby(["Team","Year"])["Player"].nunique()
                      .rename(rostered_col)
                )
                team_year = team_year.merge(started, on=["Team","Year"], how="left")
                team_year = team_year.merge(rostered, on=["Team","Year"], how="left")

            # team_all_time
            for pos in ["QB","RB","WR","TE"]:
                started_col = f"Number of {pos} started"
                rostered_col = f"Number of {pos} rostered"
                started = (
                    pw_u[(pw_u["Position"]==pos) & (pw_u["StarterFlag"])]
                      .groupby(["Team"])["Player"].nunique()
                      .rename(started_col)
                )
                rostered = (
                    pw_u[pw_u["Position"]==pos]
                      .groupby(["Team"])["Player"].nunique()
                      .rename(rostered_col)
                )
                team_all = team_all.merge(started, on=["Team"], how="left")
                team_all = team_all.merge(rostered, on=["Team"], how="left")

        # Dynamic opponent columns: Record/Win% vs each team
        team_names = []
        if not team_all.empty and "Team" in team_all.columns:
            team_names = sorted([str(x) for x in team_all["Team"].dropna().unique().tolist()])
        if team_names:
            extra_cols = []
            for t in team_names:
                extra_cols.append(f"Record vs {t}")
                extra_cols.append(f"Win % vs {t}")
            catalog["team-year"] = catalog.get("team-year", []) + [c for c in extra_cols if c not in catalog.get("team-year", [])]
            catalog["team-all-time"] = catalog.get("team-all-time", []) + [c for c in extra_cols if c not in catalog.get("team-all-time", [])]

        # Add Position column to all player sheets
        for k in ["Player-Week","Player-year","Player-all-time"]:
            cols = catalog.get(k, [])
            if "Position" not in cols:
                # place next to NFL team for readability
                if "NFL team" in cols:
                    i = cols.index("NFL team")
                    cols.insert(i, "Position")
                else:
                    cols.append("Position")
                catalog[k] = cols

        # Move Week Name next to Week (ordering only)
        for k in ["Player-Week","team-week","league-week"]:
            cols = catalog.get(k, [])
            if "Week" in cols and "Week Name" in cols:
                cols = [c for c in cols if c != "Week Name"]
                wi = cols.index("Week")
                cols.insert(wi+1, "Week Name")
                catalog[k] = cols
    except Exception as e:
        _log_exc(debug, "final_schema_tweaks", e)
    tables = [
        ("player_week.csv", pw, "Player-Week"),
        ("player_year.csv", player_year, "Player-year"),
        ("player_all_time.csv", player_all, "Player-all-time"),
        ("team_week.csv", tw, "team-week"),
        ("team_year.csv", team_year, "team-year"),
        ("team_all_time.csv", team_all, "team-all-time"),
        ("league_week.csv", league_week, "league-week"),
        ("league_year.csv", league_year, "league-year"),
        ("league_all_time.csv", league_all, "league-all-time"),
        ("transactions.csv", tx, "transactions"),
        ("trades.csv", tr, "trades"),
        ("pick_history.csv", ph, "Pick History"),
    ]
    write_outputs(tables)
