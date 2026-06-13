
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Any, Tuple, Optional, Set
from datetime import datetime, timezone, timedelta, date
from collections import Counter, deque, defaultdict
import json
import math
import re
import traceback
import logging
import warnings
import numpy as np
import os
import sys

warnings.filterwarnings("ignore", category=FutureWarning, message="Downcasting object dtype arrays")

LOG = logging.getLogger("lotg")
if not LOG.handlers:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")


import pandas as pd


def _round_from_label(label: Any) -> str:
    s = str(label or "").strip().lower()
    if s in ("semifinal", "semifinals"):
        return "semifinals"
    if s in ("final", "finals"):
        return "Finals"
    if s in ("toilet semis", "toilet semifinal", "toilet semifinals"):
        return "Toilet Semis"
    if s in ("toilet final", "toilet finals"):
        return "toilet finals"
    if s in ("3rd place", "third place"):
        return "3rd place game"
    if s in ("toilet trash", "trash"):
        return "toilet trash"
    return "regular season"
# ----------------------------
# DataFrame safety helpers
# ----------------------------
def ensure_cols(df: pd.DataFrame, cols, default=None):
    """Ensure columns exist; if missing, create with default."""
    for c in cols:
        if c not in df.columns:
            df[c] = default
    return df

def to_num_series(s, default=0.0):
    """Robust numeric coercion for series/array-like; returns float series."""
    if s is None:
        return pd.Series([], dtype='float64')
    out = pd.to_numeric(s, errors='coerce')
    if isinstance(out, pd.Series):
        return out.fillna(default)
    # scalar
    return pd.Series([out if pd.notna(out) else default], dtype='float64')

def safe_to_numeric(df: pd.DataFrame, col: str, default=0.0):
    """Convert df[col] to numeric if present; otherwise create with default."""
    if col not in df.columns:
        df[col] = default
    else:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(default)
    return df

def as_bool(df: pd.DataFrame, col: str, default=False):
    """Ensure a boolean column exists, with pandas BooleanDtype."""
    if col not in df.columns:
        df[col] = default
    df[col] = df[col].fillna(default).astype('boolean')
    return df

def log_df(df: pd.DataFrame, name: str, sample_cols=None, n=3):
    """Log basic df shape and missingness for debugging."""
    try:
        LOG.info('%s: shape=%s', name, df.shape)
        if sample_cols:
            miss = {c: int(df[c].isna().sum()) for c in sample_cols if c in df.columns}
            LOG.info('%s: missing=%s', name, miss)
    except Exception as e:
        LOG.warning('log_df failed for %s: %s', name, e)
    return df


def log_missing_cols(df: pd.DataFrame, name: str, required: list[str]) -> None:
    """Log missing required columns; helps catch silent schema drift."""
    missing = [c for c in required if c not in df.columns]
    if missing:
        LOG.warning("%s: missing expected columns: %s", name, missing)


import yaml
from dateutil import parser as dateparser

SRC_ROOT = Path(__file__).resolve().parent
if str(SRC_ROOT) not in sys.path:
    sys.path.insert(0, str(SRC_ROOT))

SUPPORT_ROOT = Path(__file__).resolve().parent.parent / "lib"
if str(SUPPORT_ROOT) not in sys.path:
    sys.path.insert(0, str(SUPPORT_ROOT))

from lotg_support.utils import HttpConfig, safe_div, clean_name, safe_bool
from lotg_support.sleeper import SleeperClient
from lotg_support.injury_tracker import (
    load_status_index as _load_injury_tracker,
    resolve_injury_flags as _resolve_injury_flags,
)
from lotg_support.external import (
    ExternalConfig,
    load_dynastyprocess_playerids,
    load_dynastyprocess_values_players,
    load_dynastyprocess_values_picks,
    load_nflverse_injuries,
    load_nflverse_player_ids,
    load_nflverse_stats_player_week,
    load_nflverse_weekly_rosters,
)
from lotg_support.lineup import compute_optimal_lineup
from lotg_support.plan import load_plan_catalog, require_columns

import league_all_time
import league_week
import league_year
import pick_history
import player_all_time
import player_week
import player_year
import team_all_time
import team_week
import team_year
import trades
import transactions
import formulas

DOCUMENT_MODULES = [
    formulas,
    player_week,
    player_year,
    player_all_time,
    team_week,
    team_year,
    team_all_time,
    league_week,
    league_year,
    league_all_time,
    transactions,
    trades,
    pick_history,
]


# ============================================================
# Build philosophy (two-step internally):
#  1) Pull & cache ALL Sleeper + supporting NFL data we need.
#  2) Compute every sheet deterministically from those caches.
# ============================================================


@dataclass
class RunConfig:
    league_id: str
    min_season: int | None
    max_season: int | None
    season_type: str = "regular"



def _norm_team_name(name: Any) -> str:
    """Normalize owner/team names for consistent joins (case-insensitive, space-insensitive)."""
    s = str(name or "").strip().lower()
    s = re.sub(r"\s+", "", s)
    return s



# --------------------------
# Logging helpers
# --------------------------

def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


# --------------------------
# In-season freshness gates (PR E follow-ups A & C)
# --------------------------
# These keep an actively-running (mid-season) build from finalizing games or
# standings that aren't actually done yet. They are PURELY TIME-BASED — we do
# NOT wait on nflverse (which lags ~Tue–Wed); a week/season is "done" once
# enough real time has passed that its NFL games are over. For any historical
# season these always return True, so completed seasons are unaffected.

def _nfl_kickoff_thursday(season: int) -> date:
    """NFL Week 1 Thursday ≈ the Thursday after Labor Day (1st Monday of Sept)."""
    sept1 = date(int(season), 9, 1)
    labor_day = sept1 + timedelta(days=(0 - sept1.weekday()) % 7)  # Monday == 0
    return labor_day + timedelta(days=3)


def _week_complete_cutoff(season: int, week: int) -> datetime:
    """UTC instant after which fantasy `week`'s NFL games are safely final:
    Tuesday 08:00 UTC (≈ 3am ET), a few hours after that week's Monday Night
    game so the in-progress week is never finalized early."""
    monday = _nfl_kickoff_thursday(season) + timedelta(days=4 + 7 * (int(week) - 1))
    tuesday = monday + timedelta(days=1)
    return datetime(tuesday.year, tuesday.month, tuesday.day, 8, 0, tzinfo=timezone.utc)


def _week_is_complete(season: int, week: int, now: Optional[datetime] = None) -> bool:
    """True once `week`'s games are over by the clock (independent of nflverse)."""
    now = now or datetime.now(timezone.utc)
    try:
        return now >= _week_complete_cutoff(int(season), int(week))
    except Exception:
        return True  # never block on a date we can't compute


def _season_is_complete(season: int, now: Optional[datetime] = None) -> bool:
    """True once the season's fantasy championship is over. Week 18's Tuesday
    cutoff is comfortably after the fantasy final (≤ NFL week 17), so we gate on
    that. Used to suppress provisional playoff/champion/finish for a live season."""
    return _week_is_complete(int(season), 18, now)

def _log(path: Path, msg: str) -> None:
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(path.read_text() + msg + "\n" if path.exists() else msg + "\n")
    except Exception:
        pass

def _log_exc(path: Path, where: str, e: Exception) -> None:
    _log(path, f"[{_now_iso()}] ERROR at {where}: {type(e).__name__}: {e}\n{traceback.format_exc()}")


def _fatal_log(repo_root: Path, where: str, e: Exception) -> None:
    """Best-effort fatal logging for CI visibility when process exits non-zero."""
    debug = repo_root / "exports" / "raw" / "build_debug.log"
    _log_exc(debug, where, e)
    try:
        print(f"FATAL [{where}] {type(e).__name__}: {e}", file=sys.stderr)
        traceback.print_exc()
    except Exception:
        pass

def _to_int(x: Any, default: Optional[int] = None) -> Optional[int]:
    try:
        return int(x)
    except Exception:
        return default

def _to_float(x: Any, default: Optional[float] = None) -> Optional[float]:
    try:
        return float(x)
    except Exception:
        return default


def _format_pick_number(round_no: Optional[int], pick_in_round: Optional[int]) -> Optional[str]:
    """Canonical pick notation: '1.05' for round 1 slot 5. '1' when slot unknown."""
    if round_no is None:
        return None
    if pick_in_round is None:
        return f"{int(round_no)}"
    return f"{int(round_no)}.{int(pick_in_round):02d}"


# Phase 12 #5: re-score nflverse weekly stats with the LEAGUE's own (Sleeper)
# scoring settings, so "Points (full season)" lives on the same scale as the
# rostered "Points" (which use Sleeper scoring). Offensive scoring only —
# nflverse stats_player_week is offense; the league rosters no K/DST. Each
# season uses its OWN scoring_settings, so a settings change is handled
# automatically (a build-time log flags when they differ year-to-year).
_LEAGUE_SCORE_MAP = {
    # Passing
    "pass_yd": ("passing_yards",),
    "pass_td": ("passing_tds",),
    "pass_int": ("passing_interceptions",),
    "pass_2pt": ("passing_2pt_conversions",),
    "pass_fd": ("passing_first_downs",),
    # Rushing
    "rush_yd": ("rushing_yards",),
    "rush_td": ("rushing_tds",),
    "rush_2pt": ("rushing_2pt_conversions",),
    "rush_fd": ("rushing_first_downs",),
    # Receiving
    "rec": ("receptions",),
    "rec_yd": ("receiving_yards",),
    "rec_td": ("receiving_tds",),
    "rec_2pt": ("receiving_2pt_conversions",),
    "rec_fd": ("receiving_first_downs",),
    # Fumbles. The league scores 'fum' (ANY fumble) on TOP of 'fum_lost', so a
    # lost fumble is fum + fum_lost. 'fum_rec' is opponent-fumble recovery
    # (recovering your OWN fumble is not awarded — verified vs Sleeper).
    "fum": ("sack_fumbles", "rushing_fumbles", "receiving_fumbles"),
    "fum_lost": ("sack_fumbles_lost", "rushing_fumbles_lost", "receiving_fumbles_lost"),
    "fum_rec": ("fumble_recovery_opp",),
    "fum_rec_td": ("fumble_recovery_tds",),
    "fum_ret_yd": ("fumble_recovery_yards_own", "fumble_recovery_yards_opp"),
    # Special teams (returners)
    "st_td": ("special_teams_tds",),
    # Kicking (future-proof — scored if a kicker is ever rostered)
    "fgm_0_19": ("fg_made_0_19",),
    "fgm_20_29": ("fg_made_20_29",),
    "fgm_30_39": ("fg_made_30_39",),
    "fgm_40_49": ("fg_made_40_49",),
    "fgm_50p": ("fg_made_50_59", "fg_made_60_"),
    "fgmiss": ("fg_missed",),
    "xpm": ("pat_made",),
    "xpmiss": ("pat_missed",),
    # Individual defense (IDP) — future-proof
    "def_td": ("def_tds",),
    "int": ("def_interceptions",),
    "sack": ("def_sacks",),
    "ff": ("def_fumbles_forced",),
    "safe": ("def_safeties",),
}
_LEAGUE_SCORE_BONUS = (  # (scoring_key, stat_col, threshold)
    ("bonus_pass_yd_300", "passing_yards", 300), ("bonus_pass_yd_400", "passing_yards", 400),
    ("bonus_rush_yd_100", "rushing_yards", 100), ("bonus_rush_yd_200", "rushing_yards", 200),
    ("bonus_rec_yd_100", "receiving_yards", 100), ("bonus_rec_yd_200", "receiving_yards", 200),
)


def _league_score(stats: Dict[str, Any], scoring: Dict[str, Any], position: Optional[str] = None) -> float:
    """Fantasy points for one nflverse stat row under `scoring` (Sleeper)."""
    pts = 0.0
    for key, cols in _LEAGUE_SCORE_MAP.items():
        mult = scoring.get(key)
        if not mult:
            continue
        val = 0.0
        for c in cols:
            v = stats.get(c)
            if v is not None:
                try:
                    val += float(v)
                except Exception:
                    pass
        pts += float(mult) * val
    if position == "TE" and scoring.get("bonus_rec_te"):
        try:
            pts += float(scoring["bonus_rec_te"]) * float(stats.get("receptions") or 0.0)
        except Exception:
            pass
    for bkey, col, thresh in _LEAGUE_SCORE_BONUS:
        if scoring.get(bkey):
            try:
                if float(stats.get(col) or 0.0) >= thresh:
                    pts += float(scoring[bkey])
            except Exception:
                pass
    return round(pts, 2)


def _valid_pid(pid: Any) -> bool:
    """Return True if pid is a real Sleeper player id (not placeholders like '0')."""
    if pid is None:
        return False
    # handle pandas/numpy NaN
    try:
        if isinstance(pid, float) and np.isnan(pid):
            return False
    except Exception:
        pass
    s = str(pid).strip()
    if s == "" or s.lower() == "nan":
        return False
    if s == "0" or s == "-1":
        return False
    return True
def _safe_df(obj: Any) -> pd.DataFrame:
    return obj if isinstance(obj, pd.DataFrame) else pd.DataFrame()

def _first_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    if not isinstance(df, pd.DataFrame) or df.empty:
        return None
    for c in candidates:
        if c in df.columns:
            return c
    return None

def _epoch_ms_to_dt(ms: Any) -> Optional[datetime]:
    try:
        ms_i = int(ms)
        if ms_i <= 0:
            return None
        return datetime.fromtimestamp(ms_i / 1000, tz=timezone.utc)
    except Exception:
        return None

def _epoch_ms_to_date(ms: Any) -> Optional[date]:
    dt = _epoch_ms_to_dt(ms)
    return dt.date() if dt else None

def _calc_age(birth_date_str: Optional[str], on_date: date) -> Optional[float]:
    if not birth_date_str:
        return None
    try:
        bd = dateparser.parse(str(birth_date_str)).date()
        return round((on_date - bd).days / 365.25, 2)
    except Exception:
        return None



def _pick_expected_age(year_of_pick: int, on_date: date) -> Optional[float]:
    """Synthetic age of a not-yet-known rookie at a given date.

    NFL rookies average ~22 at draft time and our league's rookie
    draft wraps up early September. Anchor each pick's expected
    birth date at Sept 1 of (Y - 22) so a typical rookie reads
    exactly 22 on draft day. Earlier trades of further-out picks
    naturally read younger.

    Used wherever picks need to participate in age calculations:
    trades 'Asset difference in average age', team_week 'Team age
    including picks', and the avg_age input to the tanking score.
    """
    try:
        born = date(int(year_of_pick) - 22, 9, 1)
        return round((on_date - born).days / 365.25, 2)
    except Exception:
        return None


def _rookie_season(meta: Dict[str, Any], current_season: Optional[int]) -> Optional[int]:
    draft_year = _to_int(meta.get("draft_year"), None)
    if draft_year:
        return draft_year
    years_exp = _to_float(meta.get("years_exp"), None)
    if years_exp is None or current_season is None or years_exp < 0:
        return None
    return int(current_season) - int(years_exp)


# --------------------------
# Team handle mapping (HANDLE, not franchise name)
# --------------------------

def _team_handle_map(users: List[Dict[str, Any]]) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for u in users or []:
        uid = str(u.get("user_id"))
        handle = u.get("display_name") or u.get("username")
        if not handle:
            meta = u.get("metadata") or {}
            handle = meta.get("team_name") or uid
        out[uid] = str(handle)
    return out


# --------------------------
# NFL team normalization + bye schedule support
# --------------------------

_TEAM_NORMALIZE = {
    "LA": "LAR", "STL": "LAR",
    "SD": "LAC",
    "WSH": "WAS",
    "ARZ": "ARI", "AZ": "ARI",
    "NWE": "NE",
    "KCC": "KC",
    "NOR": "NO",
    "SFO": "SF",
    "TAM": "TB",
    "GNB": "GB",
    "LVR": "LV",
    # already-normalized common codes:
    "ARI": "ARI", "ATL": "ATL", "BAL": "BAL", "BUF": "BUF", "CAR": "CAR",
    "CHI": "CHI", "CIN": "CIN", "CLE": "CLE", "DAL": "DAL", "DEN": "DEN",
    "DET": "DET", "GB": "GB", "HOU": "HOU", "IND": "IND", "JAX": "JAX",
    "KC": "KC", "LAC": "LAC", "LAR": "LAR", "LV": "LV", "MIA": "MIA",
    "MIN": "MIN", "NE": "NE", "NO": "NO", "NYG": "NYG", "NYJ": "NYJ",
    "PHI": "PHI", "PIT": "PIT", "SEA": "SEA", "SF": "SF", "TB": "TB",
    "TEN": "TEN", "WAS": "WAS",
}

def _norm_team(t: Any) -> Optional[str]:
    if not t:
        return None
    s = str(t).strip().upper()
    return _TEAM_NORMALIZE.get(s, s)

def _download_csv_best_effort(urls: List[str], path: Path, timeout: int = 120, debug: Optional[Path]=None) -> pd.DataFrame:
    import requests
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists() and path.stat().st_size > 0:
        try:
            return pd.read_csv(path)
        except Exception:
            pass

    last_err = None
    for url in urls:
        fetched = False
        for trust_env in (False, True):
            try:
                session = requests.Session()
                session.trust_env = trust_env
                kwargs = {"timeout": timeout}
                if not trust_env:
                    kwargs["proxies"] = {"http": None, "https": None}
                r = session.get(url, **kwargs)
                if r.status_code == 200 and r.content:
                    path.write_bytes(r.content)
                    fetched = True
                    break
                last_err = f"{r.status_code} {url} trust_env={trust_env}"
            except Exception as e:
                last_err = f"{type(e).__name__}: {e} {url} trust_env={trust_env}"

        if fetched:
            try:
                return pd.read_csv(path)
            except Exception:
                return pd.DataFrame()

    if debug:
        _log(debug, f"[{_now_iso()}] WARN csv download failed: {last_err}")
    return pd.DataFrame()


def _played_teams_by_week(games: pd.DataFrame, season: int) -> Dict[int, set]:
    games = _safe_df(games)
    out: Dict[int, set] = {}
    if games.empty:
        return out
    if not {"season", "week", "home_team", "away_team"}.issubset(set(games.columns)):
        return out
    sub = games.copy()
    try:
        sub["season"] = pd.to_numeric(sub["season"], errors="coerce").astype("Int64")
        sub["week"] = pd.to_numeric(sub["week"], errors="coerce").astype("Int64")
    except Exception:
        return out
    sub = sub[sub["season"] == season]
    if sub.empty:
        return out
    for wk, g in sub.groupby("week"):
        if pd.isna(wk):
            continue
        home = g["home_team"].dropna().astype(str).map(_norm_team).tolist()
        away = g["away_team"].dropna().astype(str).map(_norm_team).tolist()
        out[int(wk)] = set([t for t in (home + away) if t])
    return out


# --------------------------
# Injury/Suspension flags (platform designation at that week)
# --------------------------

def _infer_flags_from_sleeper_player_meta(meta: Dict[str, Any]) -> Tuple[Optional[bool], Optional[bool]]:
    """
    Uses Sleeper platform designations from /players/nfl (status/injury_status).
    Conservative: only True when it's clearly OUT/IR/PUP/NFI or SUSP.
    """
    if not isinstance(meta, dict):
        return (None, None)

    status = str(meta.get("status") or "").lower()
    injury_status = str(meta.get("injury_status") or "").lower()

    # suspension
    if "susp" in status or "susp" in injury_status:
        return (False, True)

    healthy_markers = {"active", "", "healthy", "none", "null"}
    if (status in healthy_markers) and (injury_status in healthy_markers):
        return (False, False)

    # Out/IR style
    injury_markers = ["ir", "out", "inactive", "pup", "nfi", "injured", "covid"]
    if any(k == status for k in injury_markers) or any(k in injury_status for k in injury_markers):
        return (True, False)

    # questionable/doubtful can still play -> do not mark True
    if ("question" in status) or ("question" in injury_status) or ("doubt" in injury_status):
        return (False, False)

    return (None, None)

def _infer_flags_from_nflverse(injuries: pd.DataFrame, gsis_id: Optional[str], season: int, week: int) -> Tuple[Optional[bool], Optional[bool]]:
    injuries = _safe_df(injuries)
    if injuries.empty or not gsis_id:
        return (None, None)
    if "gsis_id" not in injuries.columns:
        return (None, None)

    try:
        sub = injuries.copy()
        if "season" in sub.columns:
            sub["season"] = pd.to_numeric(sub["season"], errors="coerce").astype("Int64")
            sub = sub[sub["season"] == season]
        if "week" in sub.columns:
            sub["week"] = pd.to_numeric(sub["week"], errors="coerce").astype("Int64")
            sub = sub[sub["week"] == week]
        sub = sub[sub["gsis_id"].astype(str) == str(gsis_id)]
    except Exception:
        return (None, None)

    if sub.empty:
        return (None, None)

    status_col = _first_col(sub, ["report_status", "status", "game_status", "injury_status", "practice_status"])
    if not status_col:
        return (None, None)

    s = str(sub.iloc[0].get(status_col) or "").lower()
    if not s:
        return (None, None)

    suspension = ("susp" in s) or ("sspd" in s)
    injury = (("out" in s) or ("ir" in s) or ("inactive" in s) or ("pup" in s)) and not suspension
    return (injury, suspension)

def _merge_flags(primary: Tuple[Optional[bool], Optional[bool]], secondary: Tuple[Optional[bool], Optional[bool]]) -> Tuple[Optional[bool], Optional[bool]]:
    inj1, sus1 = primary
    inj2, sus2 = secondary
    if sus1 is True or sus2 is True:
        return (False, True)
    if inj1 is True or inj2 is True:
        return (True, False)
    if (inj1 is False and sus1 is False) or (inj2 is False and sus2 is False):
        return (False, False)
    return (None, None)


# --------------------------
# League chain
# --------------------------

def _walk_league_chain(sc: SleeperClient, start_league_id: str, min_season: int | None, max_season: int | None) -> List[Dict[str, Any]]:
    chain: List[Dict[str, Any]] = []
    lid = str(start_league_id)
    seen = set()
    while lid and lid not in seen:
        seen.add(lid)
        try:
            lg = sc.league(lid)
        except Exception:
            break
        if not isinstance(lg, dict):
            break
        season = _to_int(lg.get("season"), None)
        if season is not None and min_season is not None and season < min_season:
            break
        chain.append(lg)
        prev = lg.get("previous_league_id")
        lid = str(prev) if prev else ""
        if lid == "None":
            lid = ""
    chain = sorted(chain, key=lambda x: _to_int(x.get("season"), 0) or 0)
    if max_season is not None:
        chain = [x for x in chain if (_to_int(x.get("season"), 0) or 0) <= max_season]
    return chain


# --------------------------
# Plan column enforcement
# --------------------------

def _to_eastern_display(series: pd.Series) -> pd.Series:
    """Convert a UTC timestamp column to US Eastern wall-clock time for display.

    All raw Sleeper timestamps are UTC (e.g. '2021-09-07 22:42:20.159000+00:00').
    We render them in America/New_York — DST-aware, so EST (UTC-5) in winter and
    EDT (UTC-4) during daylight saving — formatted 'YYYY-MM-DD HH:MM:SS' with no
    offset and no microseconds. Values that don't parse (blank / 'N/A') pass
    through unchanged. This is display-only and runs after ALL date-based logic
    (window comparisons, links, sorting), so internal computations stay on UTC.
    """
    # format="mixed" parses each value independently so the 'T'-vs-space
    # separator and microsecond differences across rows all resolve (a single
    # inferred format would NaT the non-matching ones).
    try:
        parsed = pd.to_datetime(series, errors="coerce", utc=True, format="mixed")
    except (ValueError, TypeError):
        parsed = pd.to_datetime(series, errors="coerce", utc=True)
    out = parsed.dt.tz_convert("America/New_York").dt.strftime("%Y-%m-%d %H:%M:%S")
    return out.where(parsed.notna(), series)


def _bold_comment_verbs(xlsx_path: Path, debug: Optional[Path] = None) -> None:
    """Bold the action verbs inside every cell-comment. openpyxl writes a comment
    as a single plain `<text><t>…</t></text>`; we rewrite it with rich-text runs,
    bolding the event verbs. Each rewritten comment file is ET-validated; on ANY
    error the original workbook is restored so a bad rewrite can't corrupt it."""
    import zipfile
    import shutil
    import xml.etree.ElementTree as ET

    _vpat = re.compile(r"\b(drafted|traded|dropped|added|got|sent)\b", re.IGNORECASE)
    _tpat = re.compile(r"<text><t(?:[^>]*)>(.*?)</t></text>", re.DOTALL)
    _rpr = '<rPr><sz val="9"/><color indexed="81"/><rFont val="Tahoma"/><family val="2"/></rPr>'
    _rprb = '<rPr><b/><sz val="9"/><color indexed="81"/><rFont val="Tahoma"/><family val="2"/></rPr>'

    def _runs(_txt: str) -> str:
        parts: List[str] = []
        _last = 0
        for _m in _vpat.finditer(_txt):
            if _m.start() > _last:
                parts.append(f'<r>{_rpr}<t xml:space="preserve">{_txt[_last:_m.start()]}</t></r>')
            parts.append(f'<r>{_rprb}<t xml:space="preserve">{_m.group(0)}</t></r>')
            _last = _m.end()
        if _last < len(_txt):
            parts.append(f'<r>{_rpr}<t xml:space="preserve">{_txt[_last:]}</t></r>')
        return "<text>" + "".join(parts) + "</text>"

    _bak = Path(str(xlsx_path) + ".bak")
    _tmp = Path(str(xlsx_path) + ".tmp")
    shutil.copy(xlsx_path, _bak)
    try:
        with zipfile.ZipFile(xlsx_path, "r") as zin, zipfile.ZipFile(_tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for _it in zin.namelist():
                _data = zin.read(_it)
                if re.match(r"xl/comments/comment\d+\.xml$", _it):
                    _new = _tpat.sub(lambda m: _runs(m.group(1)), _data.decode("utf-8"))
                    ET.fromstring(_new)  # well-formedness guard
                    _data = _new.encode("utf-8")
                zout.writestr(_it, _data)
        _tmp.replace(xlsx_path)
        _bak.unlink()
        if debug:
            _log(debug, f"[{_now_iso()}] INFO bolded comment verbs in cell history popups")
    except Exception:
        shutil.copy(_bak, xlsx_path)  # restore the good file
        for _p in (_bak, _tmp):
            try:
                _p.unlink()
            except Exception:
                pass
        raise


def _ensure_plan_columns(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    df = _safe_df(df).copy()
    for c in cols:
        if c not in df.columns:
            df[c] = None
    return df[cols]


TEAM_VS_EXTREME_COLS = [
    "Highest Win % vs a team",
    "Team for highest Win %",
    "Lowest Win % vs a team",
    "Team for lowest Win %",
]


# Phase 11C-2: header color-banding by topic (mirrors the 11C-1 reorder groups),
# a consistent palette across every data sheet + family tab colors.
_TOPIC_FILL = {
    "Identity": "D9D9D9", "Outcome": "C6E0B4", "Scoring": "BDD7EE",
    "Value": "8EAADB", "Awards": "FFE699", "Streaks": "F4B183",
    "Roster": "D9C3E9", "Hardship & Luck": "FFC7CE", "Activity": "FFD966",
    "KTC": "9CD3DC", "Tenure": "C9E4CA", "Change": "F8CBAD",
}
_FAMILY_TAB = {
    "player": "5B9BD5", "team": "70AD47", "league": "FFC000",
    "transactions": "ED7D31", "trades": "7030A0", "picks": "808080",
    "formulas": "44546A",
}
_TOPIC_IDENTITY = {
    "player", "player id", "team", "year", "week", "week name", "position",
    "position started in (if starter)", "nfl team", "opponent", "opponent team (raw)",
    "starter/bench", "season", "date", "date dropped/traded", "top team", "last team",
    "top team points", "player picked", "player added", "player dropped",
    "reference player name", "team's traded with", "age", "rookie?", "taxi-eligible",
    "type of transaction (waiver/free agency)", "etc", "number", "original team",
}


def _col_topic(col: str) -> str:
    """Topic group for header color-banding — mirrors the 11C-1 reorder classifier."""
    n = re.sub(r"\s+", " ", str(col).strip().lower())
    if n in _TOPIC_IDENTITY or n.startswith("team's traded with"):
        return "Identity"
    if n.startswith("change from") or n.startswith("change in"):
        return "Change"
    if "streak" in n:
        return "Streaks"
    if n in ("injury?", "suspension?", "bye?", "loss from hardship?"):
        return "Hardship & Luck"
    if n == "win?":
        return "Outcome"
    if n in ("cuff when drafted?", "cuff at time of pickup?"):
        return "Roster"
    if n == "commissioner moved?":
        return "Activity"
    if n.endswith("?") or n.startswith("times ") or n.startswith("times as ") or n in ("brosenzweig", "sisenzweig"):
        return "Awards"
    if "ktc" in n or "value difference" in n or "net ktc" in n:
        return "KTC"
    if (n.startswith("link to") or "tenure" in n or "weeks on team" in n
            or "weeks as team starter" in n or "weeks on bench" in n
            or "consecutive weeks on bench" in n or "weeks before first start" in n
            or "weeks between pickup" in n):
        return "Tenure"
    if n == "tanking":
        return "Roster"
    if any(k in n for k in ("injur", "suspensi", "bye", "hardship", "luck", "weeks missed", "weeks of ")):
        return "Hardship & Luck"
    if any(k in n for k in ("startable bench", "benchable starter", "startables", "cuff adjusted",
                            "% of points", "ppg", "par", "volatility", "scoring floor", "scoring ceiling",
                            "boom", "bust", "addition value", "points added", "net points", "points lost",
                            "avg net", "o-score", "skill", "difference of averages", "% of starts",
                            "difference from best", "difference from worst", "weeks as starter")):
        return "Value"
    if any(k in n for k in ("age", "number of qb", "number of rb", "number of wr", "number of te",
                            "number of rookies", "number of cuff", "donut", "under 10", "over 20",
                            "over 30", "over 40", "over 50", "nfl team", "rostered from same",
                            "started from same", "startup draft", "difference between highest and lowest",
                            "number of players on bye", "number of teams")):
        return "Roster"
    if any(k in n for k in ("assets received", "assets sent", "assets retained", "assets traded",
                            "assets dropped", "additional assets", "return from trades", "pick value received",
                            "number of bids", "number of times picked up", "transaction", "trade", "drop",
                            "faab", "draft value", "picks made", "future draft", "turnover",
                            "first round picks", "number of picks")):
        return "Activity"
    if any(k in n for k in ("record vs", "win % vs", "win %", "record", "result", "week of playoff",
                            "championship", "margin", "differential", "all time", "upst", "win variance",
                            "playoff tiebreaker", "games within")):
        return "Outcome"
    if any(k in n for k in ("pf", "points against", "max pf", "efficiency", "points", "avg points",
                            "highest starter score", "lowest starter score", "increase in points")):
        return "Scoring"
    return "Identity"


def _col_number_format(col: str) -> Optional[str]:
    """Excel number format for a column (Phase 11E): uniform 2 decimals, NO
    thousands commas, on all value/stat/percent columns. Counts, streaks and
    Year/Week/Season stay whole numbers; text/date columns are left General."""
    n = re.sub(r"\s+", " ", str(col).strip().lower())
    # Identity / labels / dates / pick-number -> leave alone.
    if n in ("year", "week", "season", "number") or "date" in n:
        return None
    # Whole-number columns: counts, aggregates, streaks.
    if (n.startswith("number of ") or n.startswith("times ") or n.startswith("times as ")
            or n.startswith("most number of ") or n.startswith("weeks ")
            or n.endswith("streak") or n == "championships" or n == "upst"
            or "number of teams" in n):
        return "0"
    # Percent columns that are stored 0-100 (NOT fractions) -> show with a "%"
    # literal and no x100.
    if n in ("starter boom %", "starter bust %", "faab premium %"):
        return '0.00"%"'
    # Percent columns stored 0-1 -> Excel percent (x100).
    if (n.endswith("%") or n.startswith("win % vs ") or "win %" in n or n == "efficiency"
            or "% of points" in n or "% of starts" in n or "all-play win" in n
            or n in ("highest win % vs a team", "lowest win % vs a team")):
        return "0.00%"
    # Everything else numeric (PPG, points, PF, PAR, KTC, addition value, Luck,
    # skill, O-Score, …) -> 2 decimals, no commas. (Harmless on text cells.)
    return "0.00"


def _append_team_vs_columns(frame: pd.DataFrame, cols: List[str], plan_key: str = "team-year") -> List[str]:
    if frame.empty or "Team" not in frame.columns:
        return cols
    teams = sorted(frame["Team"].dropna().astype(str).unique().tolist())

    def _dedup(seq: List[str]) -> List[str]:
        seen: set = set()
        out: List[str] = []
        for x in seq:
            if x not in seen:
                seen.add(x)
                out.append(x)
        return out

    # Per-opponent vs columns (NOT in the plan — generated here). team-all-time
    # groups by stat type (all Win % then all Record); team-year interleaves.
    if plan_key == "team-all-time":
        per = _dedup([f"Win % vs {t}" for t in teams] + [f"Record vs {t}" for t in teams])
    else:
        per = _dedup([c for t in teams for c in (f"Record vs {t}", f"Win % vs {t}")])
    per = [c for c in per if c not in cols]
    if not per:
        return cols

    # Phase 11C-1: insert the per-opponent block right after the END OF THE
    # OUTCOME vs-cluster (the fixed Record/Win % vs buckets + the high/low
    # extremes, which the reorder already places at the end of Outcome) — not
    # at the end of the whole sheet.
    def _is_vs(c: str) -> bool:
        return (c.startswith("Record vs ") or c.startswith("Win % vs ")
                or c in TEAM_VS_EXTREME_COLS or c.startswith("Team for "))
    idxs = [i for i, c in enumerate(cols) if _is_vs(c)]
    anchor = max(idxs) if idxs else len(cols) - 1
    return cols[:anchor + 1] + per + cols[anchor + 1:]


def _column_kind(col: str) -> str:
    """Infer expected output type for a plan column: text | boolean | numeric."""
    col_l = str(col or "").strip().lower()

    # Explicit text columns first (some contain words like "week"/"year" but are labels).
    text_exact = {
        "week name",
        "team",
        "player",
        "opponent",
        "top team",
        "last team",
        "original team",
        "final team",
        "player picked",
        "reference player name",
        "nfl team",
        "position",
        "position started in (if starter)",
        "starter/bench",
        "type of transaction",
        "assets sent",
        "assets received",
        "assets dropped",
        "assets received",
        "assets sent",
        "assets retained now",
        "assets traded away",
        "assets dropped to fa",
        "return from trades",
        "additional assets traded away in those deals",
        "return from trades of trades...of trades. keep going until present day",
        "team's traded with",
        "reason",
        "date",
        "date dropped/traded",
        "record",
        "all time record",
        "(smallest) playoff tiebreaker",
        "round",
        "number",
        "result",
        "etc",
        # team-name holders for the highest/lowest Win% extremes (item 13)
        "team for highest win %",
        "team for lowest win %",
        # Formulas sheet — pure documentation, all four columns are text
        "stat",
        "sheet",
        "formula",
        "notes",
    }
    if col_l in text_exact:
        return "text"

    # Substring-based markers: a column counts as text when its label *contains*
    # any of these tokens. Important: keep these specific enough that they
    # don't grab numeric columns by accident.
    text_markers = [
        "record vs ",
        "link to ",
        "trade ",
        "pick",
        "type of transaction",  # 'type of transaction (waiver/free agency)' variants
        "player added",
        "player dropped",
        "assets received",
        "assets dropped",
        "assets received",
        "team's traded with",
        "teams traded with",
    ]
    if any(m in col_l for m in text_markers):
        return "text"

    # Boolean / flag-style columns.
    bool_exact = {
        "injury?",
        "suspension?",
        "bye?",
        "rookie?",
        "win?",
        "loss?",
        "player of the week?",
        "qb of the week?",
        "rb of the week?",
        "wr of the week?",
        "te of the week?",
        "benchwarmer of the week?",
        "bench qb of the week?",
        "bench rb of the week?",
        "bench wr of the week?",
        "bench te of the week?",
        "highest starter on team?",
        "lowest starter on team?",
        # No trailing '?' but still a flag — without this it falls through to
        # numeric and renders 1.0/0.0 while every other flag reads True/False
        # (audit run-2 F6).
        "taxi-eligible",
    }
    # "Times X of the week?" / "Times Top half of league?" etc. are aggregate
    # counts in player_year/team_year — they end in '?' but are integers.
    # Numeric kind first prevents boolean coercion of summed counts.
    if col_l.startswith("times ") or col_l.startswith("number of ") or col_l.startswith("weeks "):
        return "numeric"

    if col_l in bool_exact or col_l.endswith("?"):
        return "boolean"

    # Remaining metrics are numeric by contract.
    return "numeric"


def _is_text_column(col: str) -> bool:
    return _column_kind(col) == "text"


def _fill_empty_columns(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    for col in cols:
        if col not in df.columns:
            continue
        if df[col].isna().all():
            # Columns whose missing values carry semantic meaning
            # ('Change in X', '(if starter)' etc.) should render N/A,
            # not get filled with a default. Leave them NaN so
            # _fill_missing_values' preserve-na branch handles them.
            if _preserve_na(col):
                continue
            df[col] = _default_fill_for_column(col)
    return df


def _default_fill_for_column(col: str) -> Any:
    kind = _column_kind(col)
    if kind == "text":
        return "N/A"
    if kind == "boolean":
        return False
    return 0.0


def _terminalize_streaks(df: pd.DataFrame, group_cols: List[str],
                         order_cols: List[str], streak_cols: List[str]) -> pd.DataFrame:
    """Collapse running streak counters into a 'terminal' encoding so each
    maximal run is represented exactly once and the column stays sortable:

      * the FINAL row of a run carries the run's length (the most recent row if
        the streak is still active, else the peak row right before it reset),
      * every earlier row of the run reads the text 'In Progress',
      * rows not in any run (counter 0) stay 0.

    A descending sort then surfaces each streak once (numbers sort above the
    'In Progress' text), which is what feeds the top-N longest-streak lists."""
    cols_present = [c for c in streak_cols if c in df.columns]
    if not cols_present or df.empty:
        return df
    s = df.sort_values(group_cols + order_cols)
    idx = s.index.tolist()
    keys = list(s[group_cols].itertuples(index=False, name=None))
    n = len(idx)
    for col in cols_present:
        raw: List[int] = []
        for v in s[col].tolist():
            try:
                raw.append(int(v) if pd.notna(v) else 0)
            except Exception:
                raw.append(0)
        out: List[Any] = [0] * n
        for k in range(n):
            c = raw[k]
            same_next = (k + 1 < n) and (keys[k + 1] == keys[k])
            if c == 0:
                out[k] = 0
            elif same_next and raw[k + 1] == c + 1:
                out[k] = "In Progress"
            else:
                out[k] = c
        df[col] = df[col].astype(object)
        df.loc[idx, col] = pd.Series(out, index=idx, dtype=object)
    return df


def _encode_player_streaks(df: pd.DataFrame, group_col: str, order_cols: List[str],
                           played, specs: Dict[str, Any]) -> pd.DataFrame:
    """Compute terminal-encoded PLAYER streaks that SKIP non-played weeks.

    A bye / injury / suspension week (played[i] is False) does NOT count toward
    a streak and does NOT break it — the run simply bridges across it (those
    cells read 'N/A'). Among PLAYED weeks the encoding matches
    _terminalize_streaks: a run shows its length on its final played week,
    'In Progress' before, 0 on a played week that didn't qualify.

    `played` is a boolean array aligned to df (post-sort here); `specs` maps each
    output column -> a boolean array of 'did the player qualify that week'."""
    df = df.sort_values([group_col] + order_cols).reset_index(drop=True)
    n = len(df)
    import numpy as _np
    played = _np.asarray(played, dtype=bool)
    groups = df.groupby([group_col], sort=False).groups
    for out_col, qual in specs.items():
        qual = _np.asarray(qual, dtype=bool)
        vals: List[Any] = [0] * n
        for _gk, idx in groups.items():
            idx = list(idx)
            raw: Dict[int, int] = {}
            played_seq: List[int] = []
            c = 0
            for i in idx:
                if not played[i]:
                    continue
                c = c + 1 if qual[i] else 0
                raw[i] = c
                played_seq.append(i)
            for k, i in enumerate(played_seq):
                cc = raw[i]
                nxt = raw[played_seq[k + 1]] if k + 1 < len(played_seq) else None
                if cc == 0:
                    vals[i] = 0
                elif nxt is not None and nxt == cc + 1:
                    vals[i] = "In Progress"
                else:
                    vals[i] = cc
            for i in idx:
                if not played[i]:
                    vals[i] = "N/A"
        df[out_col] = pd.Series(vals, index=df.index, dtype=object)
    return df


def _preserve_na(col: str) -> bool:
    """Numeric columns where a missing value carries meaning (no prior data)
    and should render as 'N/A' instead of being filled with 0.0. Most of
    these describe a *change* relative to a prior period — for the first
    week/season the comparison literally doesn't exist."""
    col_l = str(col or "").strip().lower()
    if col_l.startswith("change from ") or col_l.startswith("change in "):
        return True
    # KTC value differences: 'at deal time' / 'at end of season' / '1 year
    # later' / '2 years later'. A blank here means the reference date is
    # in the future, or at least one side of the trade had no resolvable
    # KTC value (e.g. pre-2022 trades where DynastyProcess data is
    # sparse). Both cases are distinct from 'the diff is actually zero',
    # so don't collapse them to 0.0.
    if col_l.startswith("ktc value difference"):
        return True
    # KTC columns on transactions.csv. Blank means the player wasn't
    # tracked by dynasty-daddy (low-value / pre-2021), or the row was
    # a drop-only / no-drop row, or the reference date is in the
    # future. Distinct from 'KTC is actually zero'.
    if col_l.startswith("ktc value of player added") or col_l.startswith("ktc value of player dropped"):
        return True
    if col_l.startswith("net ktc value"):
        return True
    # Dropped avg/total points: N/A when the row dropped nobody; an explicit
    # 0 is real (the dropped player never played another NFL game).
    if col_l in {"dropped avg points", "dropped total points"}:
        return True
    # Trade impact score (trades): always computed, but keep the N/A-not-0
    # convention in case the composite is ever skipped for a row.
    if col_l == "trade impact score":
        return True
    # O-Score: N/A unless all four percentile components are present.
    if col_l == "o-score":
        return True
    # Manager skill (team_year / team_all_time): shrunk-mean O-Score of the
    # team's picks / trades / transactions. Blank = no events of that type
    # (didn't draft/trade/transact) — N/A, distinct from a real low score.
    if col_l in {"drafting skill", "trading skill", "transaction skill"}:
        return True
    # All-play win % / Losses from hardship: N/A for a (team, year) with no
    # games that season (e.g. the not-yet-played current/future season).
    if col_l in {"all-play win %", "all-play win % minus win %", "losses from hardship"}:
        return True
    # Week-over-week comparison columns: on the league's very first week
    # (2021 w1) the prior week doesn't exist, so the comparison is N/A, not
    # 0 (audit run-2 F5 — they rendered 0.0, reading as "no change").
    if col_l in {"increase in points from previous week",
                 "roster turnover from previous week",
                 "starter turnover from previous week",
                 "difference in pregame avg max pf from opponent"}:
        return True
    # Clutch index (team_all_time): playoff-vs-regular PF / win% delta. N/A when
    # the team never reached the winners'-bracket playoffs (no delta to take).
    if col_l in {"playoff pf minus regular-season pf",
                 "playoff win % minus regular-season win %"}:
        return True
    # Bracket-specific win % (regular season / playoff / toilet bowl): N/A for a
    # bracket with no games (e.g. a team that never made the playoffs, or the
    # not-yet-played current season). Distinct from a real 0% win rate.
    if col_l in {"regular season win %", "playoff win %", "toilet bowl win %"}:
        return True
    # 3-year roster retention rate: N/A when the +3-year week-1 roster doesn't
    # exist yet (recent classes) — distinct from a real 0% retention.
    if col_l == "3-year roster retention rate":
        return True
    # Player consistency + PAR: N/A for players who never started (volatility
    # also N/A with < 2 starts). Boom/Bust % keep a real 0 for players who did
    # start but never boomed/busted.
    if col_l in {"starter scoring volatility", "starter scoring floor", "starter scoring ceiling",
                 "starter boom %", "starter bust %", "starter par", "starter par per game",
                 "consistency percentile", "floor percentile", "ceiling percentile",
                 "rostered scoring volatility", "rostered scoring floor", "rostered scoring ceiling",
                 "rostered boom %", "rostered bust %",
                 "rostered consistency percentile", "rostered floor percentile", "rostered ceiling percentile"}:
        return True
    # Phase 12 fix #3: role-split scoring averages are N/A (not 0) when the
    # player never started / never benched / never played that period.
    if col_l in {"ppg starter", "ppg bench", "adjusted ppg starter", "adjusted ppg bench",
                 "adjusted avg points", "ppg starter vs bench diff"}:
        return True
    # Length of tenure on team: a blank means there is NO player whose tenure
    # to measure — a transactions pure drop (no added player) or an unmade pick
    # (no player drafted yet). Render those as N/A. A genuine 0-day tenure
    # (added/drafted then immediately moved) is computed as 0 and stays 0.
    if col_l == "length of tenure on team":
        return True
    # Faab columns: blank Faab means the transaction wasn't a waiver
    # (free-agent / commissioner adds aren't bid on) — surface as N/A
    # so it's not confused with 'won the claim with a $0 bid'.
    # Also covers 2021- transactions, where the league had no FAAB
    # system at all (Total FAAB bid / FAAB % difference also preserve
    # via the dedicated entries below).
    if col_l == "faab":
        return True
    if col_l == "total faab bid":
        return True
    # Faab-vs-second-place: blank means the row isn't a waiver, or the
    # waiver was uncontested (no runner-up). Either case is distinct
    # from 'won by zero' so don't collapse to 0.0.
    if col_l in {
        "faab difference over second place",
        "faab premium %",
    }:
        return True
    # 'Number of times dropped by this team' is blank on a transaction row that
    # didn't drop anyone (a pure pickup) — N/A, not 0.
    if col_l == "number of times dropped by this team":
        return True
    # Win % vs / Record vs <team>: blank means "no games played against
    # this opponent" — typically self-vs-self in team_year / team_all_time.
    # Surface as N/A so it's not confused with '0-0-0 record / 0% win'
    # against a team you actually faced.
    if col_l.startswith("win % vs ") or col_l.startswith("record vs "):
        return True
    # PPG / age / start-rate columns on transactions.csv. Blank = the
    # row had no Player Added/Dropped to compute against, or the
    # player has no pre-pickup game log. Distinct from 'value is
    # actually zero'.
    # Player-week conditional columns: by their column-name spec
    # ('(if starter)' / '(if bench)') they're only defined for one
    # half of the rows; the rest should render N/A, not 0.
    if "(if starter)" in col_l or "(if bench)" in col_l or "(if win)" in col_l:
        return True
    # Season-summary stats that should read N/A for in-progress
    # seasons (no games played yet) instead of 0.0 implying the team
    # played and scored zero.
    if col_l in {
        "points", "avg points",
        "points against", "avg points against",
        "differential", "avg differential",
        "max pf", "avg max pf",
        "efficiency",
        "hardship",
        "combined matchup score",
        "win %",
        "tanking",
        "luck",
        "win variance",
    }:
        return True
    if col_l in {
        "average ppg on team",
        "average ppg of dropped player over same time",
        "ppg of 5 games before pickup",
        "avg ppg of received players on team",
        "avg ppg of sent players over same time",
        "avg ppg of received players in 5 games before trade",
        "asset difference in average age",
        "trade addition value",
        "difference of averages",
        "difference of averages adjusted by position",
        "age difference",
    }:
        return True
    # NOTE: removed from preserve_na per Phase 1 spec — these should
    # never render as N/A:
    #   "player addition value"  -> defaults to 0.0 when not computable
    #   "number of starts before next drop" -> 0 when player never started
    #   "% of starts made while rostered" -> 0 when player never started
    #   "injury adjusted % of starts made while rostered" -> 0 when never started
    # Pick-value columns on trades.csv: blank means the trade had no
    # picks at all, or all picks failed to resolve (e.g., picks for a
    # draft too far in the future). Distinct from 'the pick value is
    # actually zero', so don't collapse to 0.0.
    if col_l in {
        "pick value received",
        "change in pick value at draft time",
    }:
        return True
    if col_l in {
        "win variance",
        "weeks between pickup and start",
    }:
        return True
    return False


def _fill_missing_values(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    truthy = {"true", "t", "1", "yes", "y"}
    falsy = {"false", "f", "0", "no", "n", ""}

    for col in cols:
        if col not in df.columns:
            continue

        kind = _column_kind(col)
        default = _default_fill_for_column(col)

        if kind == "text":
            df[col] = df[col].astype(object).replace("", default).fillna(default)
            continue

        if kind == "boolean":
            def _coerce_bool(v: Any) -> bool:
                if pd.isna(v):
                    return default
                if isinstance(v, bool):
                    return v
                if isinstance(v, (int, float)):
                    return bool(v)
                s = str(v).strip().lower()
                if s in truthy:
                    return True
                if s in falsy:
                    return False
                return default

            df[col] = df[col].map(_coerce_bool)
            continue

        # numeric path
        if _preserve_na(col):
            # First-occurrence values stay as NaN and render as 'N/A' in CSV.
            num = pd.to_numeric(df[col], errors="coerce")
            df[col] = num.astype(object).where(num.notna(), "N/A")
        else:
            # Guard: if any non-empty value parses as non-numeric (e.g.
            # pick_history's "2021 (vet)" Year tag), skip coercion to
            # preserve the intentional text tag.
            coerced = pd.to_numeric(df[col], errors="coerce")
            raw_str = df[col].astype(str).str.strip()
            non_numeric = df[col].notna() & coerced.isna() & raw_str.ne("") & raw_str.str.lower().ne("nan")
            if non_numeric.any():
                continue
            df[col] = coerced.fillna(default)

    return df


def _default_player_week_benchmark_cases() -> pd.DataFrame:
    """Fallback benchmark cases for player_week validation."""
    rows = [
        {"player": "Nick Chubb", "season": 2022, "week": 1, "expected_nfl_team": "CLE", "check_type": "HEALTHY_WEEK", "why_this_week": "Healthy baseline (Active/played; no flags)"},
        {"player": "Nick Chubb", "season": 2023, "week": 3, "expected_nfl_team": "CLE", "check_type": "INJURY_WEEK", "why_this_week": "Post-knee-injury non-start week (should show Injury/IR, not Played)"},
        {"player": "Nick Chubb", "season": 2024, "week": 16, "expected_nfl_team": "CLE", "check_type": "INJURY_WEEK", "why_this_week": "Post-broken-foot non-start week (clear Injury miss)"},
        {"player": "Nick Chubb", "season": 2025, "week": 1, "expected_nfl_team": "HOU", "check_type": "TEAM_WEEK", "why_this_week": "Team mapping check: should show Texans in 2025"},
        {"player": "Cooper Kupp", "season": 2021, "week": 1, "expected_nfl_team": "LAR", "check_type": "HEALTHY_WEEK", "why_this_week": "Healthy baseline"},
        {"player": "Cooper Kupp", "season": 2022, "week": 11, "expected_nfl_team": "LAR", "check_type": "INJURY_WEEK", "why_this_week": "On IR for high-ankle (non-start; Injury/IR)"},
        {"player": "Cooper Kupp", "season": 2023, "week": 1, "expected_nfl_team": "LAR", "check_type": "INJURY_WEEK", "why_this_week": "Start-of-season IR (non-start; Injury/IR)"},
        {"player": "Cooper Kupp", "season": 2024, "week": 3, "expected_nfl_team": "LAR", "check_type": "INJURY_WEEK", "why_this_week": "Post-ankle injury non-start week"},
        {"player": "Cooper Kupp", "season": 2025, "week": 1, "expected_nfl_team": "SEA", "check_type": "TEAM_WEEK", "why_this_week": "Team mapping check: Seahawks"},
        {"player": "Rashee Rice", "season": 2023, "week": 1, "expected_nfl_team": "KC", "check_type": "HEALTHY_WEEK", "why_this_week": "Healthy baseline"},
        {"player": "Rashee Rice", "season": 2024, "week": 5, "expected_nfl_team": "KC", "check_type": "INJURY_WEEK", "why_this_week": "After Week 4 knee injury (non-start; Injury/IR)"},
        {"player": "Rashee Rice", "season": 2025, "week": 1, "expected_nfl_team": "KC", "check_type": "SUSPENSION_WEEK", "why_this_week": "Start-of-season suspension (must show Suspension, not Injury)"},
        {"player": "Rashee Rice", "season": 2025, "week": 6, "expected_nfl_team": "KC", "check_type": "SUSPENSION_WEEK", "why_this_week": "Final week of 6-game suspension"},
        {"player": "Jordan Addison", "season": 2023, "week": 1, "expected_nfl_team": "MIN", "check_type": "HEALTHY_WEEK", "why_this_week": "Healthy rookie baseline"},
        {"player": "Jordan Addison", "season": 2024, "week": 2, "expected_nfl_team": "MIN", "check_type": "INJURY_WEEK", "why_this_week": "Ruled out (ankle); did not start"},
        # Audit run-2 F8: the old 2024 w9 "midseason suspension" case was
        # factually wrong — Addison PLAYED that week (16.1 pts); his real
        # suspension was the 3-game start of 2025 (covered below).
        {"player": "Jordan Addison", "season": 2024, "week": 12, "expected_nfl_team": "MIN", "check_type": "HEALTHY_WEEK", "why_this_week": "Healthy big game (30.2 pts); validates no stray flags"},
        {"player": "Jordan Addison", "season": 2025, "week": 1, "expected_nfl_team": "MIN", "check_type": "SUSPENSION_WEEK", "why_this_week": "Start-of-season 3-game suspension"},
        {"player": "Jordan Addison", "season": 2025, "week": 3, "expected_nfl_team": "MIN", "check_type": "SUSPENSION_WEEK", "why_this_week": "Final week of 3-game suspension"},
        {"player": "Allen Lazard", "season": 2022, "week": 4, "expected_nfl_team": "GB", "check_type": "HEALTHY_WEEK", "why_this_week": "Healthy Packers baseline"},
        {"player": "Allen Lazard", "season": 2022, "week": 1, "expected_nfl_team": "GB", "check_type": "INJURY_WEEK", "why_this_week": "Missed opener (ankle; non-start)"},
        {"player": "Allen Lazard", "season": 2023, "week": 1, "expected_nfl_team": "NYJ", "check_type": "TEAM_WEEK", "why_this_week": "Team mapping check: Jets"},
        {"player": "Allen Lazard", "season": 2024, "week": 8, "expected_nfl_team": "NYJ", "check_type": "INJURY_WEEK", "why_this_week": "Post-chest-injury IR week (non-start)"},
        # Audit run-2 F8: the old 2025 w1 case expected a row, but Lazard was
        # dropped 2024-11-10 and never re-rostered — no 2025 row is CORRECT.
        # Replaced with a week inside his real 2024 rostered window (Sep-Nov).
        {"player": "Allen Lazard", "season": 2024, "week": 6, "expected_nfl_team": "NYJ", "check_type": "TEAM_WEEK", "why_this_week": "Team mapping continuity check (rostered Sep-Nov 2024)"},
    ]
    return pd.DataFrame(rows)


def _load_player_week_benchmark_cases(repo_root: Path) -> pd.DataFrame:
    path = repo_root / "plan" / "player_week_benchmark_cases.csv"
    if path.exists():
        try:
            df = pd.read_csv(path)
            required = {"player", "season", "week", "expected_nfl_team", "check_type", "why_this_week"}
            if not df.empty and required.issubset(set(df.columns)):
                return df
        except Exception:
            pass
    return _default_player_week_benchmark_cases()


def _finalize_validation_output(out: List[Dict[str, Any]]) -> pd.DataFrame:
    df = pd.DataFrame(out)
    if df.empty:
        return df
    # keep detailed reason but force Error marker per request
    if "Error" in df.columns:
        df = df.rename(columns={"Error": "Reason"})
    df["Error"] = "error"
    wanted = [
        "Player", "Season", "Week", "Check Type", "Column", "Error", "Reason", "Expected", "Observed Example"
    ]
    cols = [c for c in wanted if c in df.columns] + [c for c in df.columns if c not in wanted]
    return df[cols]


def _known_player_column_errors(repo_root: Path, player_week_df: pd.DataFrame) -> pd.DataFrame:
    """Case-based benchmark validation focused on player_week rows."""
    cases = _load_player_week_benchmark_cases(repo_root)
    pw = _safe_df(player_week_df).copy()

    out: List[Dict[str, Any]] = []

    if pw.empty:
        for _, c in cases.iterrows():
            out.append({
                "Player": c.get("player"),
                "Season": _to_int(c.get("season"), None),
                "Week": _to_int(c.get("week"), None),
                "Check Type": c.get("check_type"),
                "Column": "ALL",
                "Error": "No player_week rows available for validation",
                "Expected": c.get("why_this_week"),
                "Observed Example": "No rows",
            })
        return _finalize_validation_output(out)

    pw["_player_norm"] = pw.get("Player", pd.Series(dtype=str)).astype(str).map(clean_name).str.lower()
    pw["_year"] = pd.to_numeric(pw.get("Year"), errors="coerce").astype("Int64")
    pw["_week"] = pd.to_numeric(pw.get("Week"), errors="coerce").astype("Int64")

    for _, c in cases.iterrows():
        player = clean_name(c.get("player")).lower()
        season = _to_int(c.get("season"), None)
        week = _to_int(c.get("week"), None)
        check_type = str(c.get("check_type") or "").upper().strip()
        expected_team = _norm_team(c.get("expected_nfl_team"))
        why = str(c.get("why_this_week") or "").strip()

        sub = pw[(pw["_player_norm"] == player) & (pw["_year"] == season) & (pw["_week"] == week)].copy()
        if sub.empty:
            out.append({
                "Player": clean_name(c.get("player")),
                "Season": season,
                "Week": week,
                "Check Type": check_type,
                "Column": "ALL",
                "Error": "Case row missing from output",
                "Expected": why,
                "Observed Example": "No matching row",
            })
            continue

        # team mapping check for all case types
        obs_team = sub.get("NFL team", pd.Series(dtype=object)).map(_norm_team)
        if expected_team and not bool((obs_team == expected_team).all()):
            out.append({
                "Player": clean_name(c.get("player")),
                "Season": season,
                "Week": week,
                "Check Type": check_type,
                "Column": "NFL team",
                "Error": "Team mismatch",
                "Expected": expected_team,
                "Observed Example": ", ".join(sorted(set(obs_team.astype(str).tolist()))),
            })

        inj = sub.get("Injury?", pd.Series(dtype=bool)).map(lambda v: safe_bool(v, default=False))
        sus = sub.get("Suspension?", pd.Series(dtype=bool)).map(lambda v: safe_bool(v, default=False))
        bye = sub.get("Bye?", pd.Series(dtype=bool)).map(lambda v: safe_bool(v, default=False))
        sb = sub.get("Starter/Bench", pd.Series(dtype=object)).astype(str).str.lower().str.strip()

        if check_type == "HEALTHY_WEEK":
            if bool(inj.any()):
                out.append({"Player": clean_name(c.get("player")), "Season": season, "Week": week, "Check Type": check_type, "Column": "Injury?", "Error": "Unexpected injury flag", "Expected": "False", "Observed Example": "True"})
            if bool(sus.any()):
                out.append({"Player": clean_name(c.get("player")), "Season": season, "Week": week, "Check Type": check_type, "Column": "Suspension?", "Error": "Unexpected suspension flag", "Expected": "False", "Observed Example": "True"})
            if bool(bye.any()):
                out.append({"Player": clean_name(c.get("player")), "Season": season, "Week": week, "Check Type": check_type, "Column": "Bye?", "Error": "Unexpected bye flag", "Expected": "False", "Observed Example": "True"})

        elif check_type == "INJURY_WEEK":
            if not bool(inj.any()):
                out.append({"Player": clean_name(c.get("player")), "Season": season, "Week": week, "Check Type": check_type, "Column": "Injury?", "Error": "Missing injury flag", "Expected": "True", "Observed Example": "False"})
            if bool(sus.any()):
                out.append({"Player": clean_name(c.get("player")), "Season": season, "Week": week, "Check Type": check_type, "Column": "Suspension?", "Error": "Suspension should not be set", "Expected": "False", "Observed Example": "True"})
            if (sb == "starter").any():
                out.append({"Player": clean_name(c.get("player")), "Season": season, "Week": week, "Check Type": check_type, "Column": "Starter/Bench", "Error": "Expected non-start week", "Expected": "Bench", "Observed Example": "Starter"})

        elif check_type == "SUSPENSION_WEEK":
            if not bool(sus.any()):
                out.append({"Player": clean_name(c.get("player")), "Season": season, "Week": week, "Check Type": check_type, "Column": "Suspension?", "Error": "Missing suspension flag", "Expected": "True", "Observed Example": "False"})
            if bool(inj.any()):
                out.append({"Player": clean_name(c.get("player")), "Season": season, "Week": week, "Check Type": check_type, "Column": "Injury?", "Error": "Injury should not be set", "Expected": "False", "Observed Example": "True"})
            if bool(bye.any()):
                out.append({"Player": clean_name(c.get("player")), "Season": season, "Week": week, "Check Type": check_type, "Column": "Bye?", "Error": "Bye should not be set during suspension", "Expected": "False", "Observed Example": "True"})
            if (sb == "starter").any():
                out.append({"Player": clean_name(c.get("player")), "Season": season, "Week": week, "Check Type": check_type, "Column": "Starter/Bench", "Error": "Expected non-start week", "Expected": "Bench", "Observed Example": "Starter"})

    return _finalize_validation_output(out)




# --------------------------
# Matchup naming for playoffs/toilet
# --------------------------

def _matchup_stage(week: int, playoff_start: Optional[int]) -> Optional[str]:
    if not playoff_start:
        return None
    if week < playoff_start:
        return None
    if week == playoff_start:
        return "SEMIS"
    if week == playoff_start + 1:
        return "FINALS"
    return None


def _trade_is_netzero_swap(t: Dict[str, Any]) -> bool:
    """A 'joke' trade where nothing actually changed hands: no players added/
    dropped, no draft picks moved, and FAAB nets to zero for every roster
    (e.g. a symmetric $5-for-$5 swap). These are deleted entirely from the
    dataset — not counted and not written to trades.csv (Phase 7A)."""
    if t.get("adds") or t.get("drops") or t.get("draft_picks"):
        return False
    net: Dict[int, int] = {}
    for wb in (t.get("waiver_budget") or []):
        if not isinstance(wb, dict):
            continue
        try:
            amt = int(wb.get("amount") or 0)
        except Exception:
            amt = 0
        r, s = wb.get("receiver"), wb.get("sender")
        if r is not None:
            net[int(r)] = net.get(int(r), 0) + amt
        if s is not None:
            net[int(s)] = net.get(int(s), 0) - amt
    # Every involved roster ends even (or no FAAB at all → empty no-op trade).
    return all(v == 0 for v in net.values())


# --------------------------
# Main build
# --------------------------

def build_all(repo_root: Path) -> None:
    debug = repo_root / "exports" / "raw" / "build_debug.log"
    _log(debug, f"\n[{_now_iso()}] ===== Build start =====")

    plan_csv = repo_root / "plan" / "LOTG Plan - Sheet1.csv"
    catalog = load_plan_catalog(plan_csv)

    # Phase 11A: flag any output column missing a Formulas-sheet entry so the
    # documentation can't silently drift as new stats ship. Non-fatal — logged.
    try:
        _undoc = formulas.undocumented_columns(catalog)
        if _undoc:
            _log(debug, f"WARNING formulas coverage: {len(_undoc)} undocumented column(s): {_undoc}")
        else:
            _log(debug, "formulas coverage: all non-obvious columns documented")
    except Exception as e:
        _log_exc(debug, "formulas_coverage_check", e)

    cfg = yaml.safe_load((repo_root / "config/league.yaml").read_text())
    run_cfg = RunConfig(
        league_id=str(cfg["league_id"]),
        min_season=cfg.get("min_season"),
        max_season=cfg.get("max_season"),
        season_type=str(cfg.get("season_type", "regular")).lower(),
    )

    http = HttpConfig(timeout_seconds=30, max_retries=10, backoff_base_seconds=0.7)
    cache_dir = repo_root / ".cache"
    cache_dir.mkdir(exist_ok=True)
    # Cache historical Sleeper responses to .cache/sleeper/ (URL-hashed JSON).
    # Current-league + /players/nfl URLs bypass the cache automatically; see
    # SleeperClient._should_cache. Restored across CI runs by the workflow's
    # .cache/ actions/cache step.
    sleeper_cache_dir = cache_dir / "sleeper"
    sleeper_cache_dir.mkdir(parents=True, exist_ok=True)
    sc = SleeperClient(run_cfg.league_id, http, cache_dir=sleeper_cache_dir)
    ext = ExternalConfig(cache_dir=cache_dir, timeout_seconds=120)

    # PR E fix B: in-house weekly Sleeper injury tracker (captured every Monday
    # night into data/injury_tracker.csv). Used as the PRIMARY injury/suspension
    # source per (player_id, season, week) in the player_week loop; nflverse is
    # the backup. Empty until 2026 wk1, so a no-op on all historical data.
    try:
        injury_tracker_idx = _load_injury_tracker(repo_root)
    except Exception as e:
        injury_tracker_idx = {}
        _log_exc(debug, "load_injury_tracker", e)
    _log(debug, f"injury tracker rows loaded: {len(injury_tracker_idx)}")

    # nflverse player id mapping (for injury + team-by-week)
    sleeper_to_gsis: Dict[str, str] = {}
    try:
        nfl_ids = _safe_df(load_nflverse_player_ids(ext))
        if (not nfl_ids.empty) and ("sleeper_id" in nfl_ids.columns) and ("gsis_id" in nfl_ids.columns):
            nfl_ids = nfl_ids.dropna(subset=["sleeper_id", "gsis_id"]).copy()
            nfl_ids["sleeper_id"] = nfl_ids["sleeper_id"].astype(str)
            nfl_ids["gsis_id"] = nfl_ids["gsis_id"].astype(str)
            sleeper_to_gsis = dict(zip(nfl_ids["sleeper_id"], nfl_ids["gsis_id"]))
    except Exception as e:
        _log_exc(debug, "load_nflverse_player_ids", e)

    # ------------- External data -------------

    try:
        dp_ids = _safe_df(load_dynastyprocess_playerids(ext))
    except Exception as e:
        dp_ids = pd.DataFrame()
        _log_exc(debug, "load_dynastyprocess_playerids", e)

    def _norm_id(x: Any) -> str:
        """
        Normalise an id read from a CSV to a bare integer-string.
        pandas reads columns containing NaN as float64, so a plain
        astype(str) on '7564' (int) yields '7564.0' (float). Strip
        the .0 suffix and any whitespace so dict lookups against
        bare Sleeper pids ('7564', '5859') succeed.
        """
        s = str(x).strip()
        if s.endswith(".0"):
            s = s[:-2]
        return s

    for c in ["sleeper_id", "gsis_id", "name"]:
        if c in dp_ids.columns:
            dp_ids[c] = dp_ids[c].astype(str)

    dp_sleeper_to_gsis: Dict[str, str] = {}
    if (not dp_ids.empty) and ("sleeper_id" in dp_ids.columns) and ("gsis_id" in dp_ids.columns):
        try:
            m = dp_ids[["sleeper_id", "gsis_id"]].dropna().copy()
            m["sleeper_id"] = m["sleeper_id"].astype(str).map(_norm_id)
            m["gsis_id"] = m["gsis_id"].astype(str).map(lambda v: str(v).strip())
            # Drop rows whose ids degenerated to empty/nan strings after coercion.
            m = m[(m["sleeper_id"] != "") & (m["sleeper_id"].str.lower() != "nan")]
            m = m[(m["gsis_id"] != "") & (m["gsis_id"].str.lower() != "nan")]
            dp_sleeper_to_gsis = dict(zip(m["sleeper_id"], m["gsis_id"]))
        except Exception:
            dp_sleeper_to_gsis = {}

    # nflverse games for byes
    games = _download_csv_best_effort(
        urls=[
            "https://raw.githubusercontent.com/nflverse/nfldata/master/data/games.csv",
            "https://github.com/nflverse/nfldata/raw/master/data/games.csv",
        ],
        path=cache_dir / "nfldata_games.csv",
        timeout=120,
        debug=debug,
    )
    played_by_week_by_season: Dict[int, Dict[int, set]] = {}
    if not games.empty:
        try:
            games["season"] = pd.to_numeric(games["season"], errors="coerce").astype("Int64")
            games["week"] = pd.to_numeric(games["week"], errors="coerce").astype("Int64")
        except Exception:
            pass

    # Sleeper NFL players (meta)
    try:
        players_nfl = sc.players_nfl()
    except Exception as e:
        players_nfl = {}
        _log_exc(debug, "players_nfl", e)

    pid_meta: Dict[str, Dict[str, Any]] = {}
    pid_pos: Dict[str, str] = {}
    for pid, meta in (players_nfl or {}).items():
        if not isinstance(meta, dict):
            continue
        pid = str(pid)
        full = meta.get("full_name") or (f"{meta.get('first_name','')} {meta.get('last_name','')}".strip())
        # Sleeper's /players/nfl returns gsis_id with leading/trailing whitespace for
        # roughly 22% of players (e.g. A.J. Brown -> ' 00-0035676'). Strip it here
        # so every downstream lookup (player_team_by_week, injuries_by_gsis_week,
        # nflverse enrichment) hits the dict keys correctly.
        raw_gsis = meta.get("gsis_id")
        clean_gsis = str(raw_gsis).strip() if raw_gsis is not None else None
        pid_meta[pid] = {
            "full_name": full or pid,
            "pos": (meta.get("position") or ""),
            "team": _norm_team(meta.get("team")),
            "birth_date": meta.get("birth_date") or meta.get("birthdate"),
            "years_exp": meta.get("years_exp"),
            "draft_year": meta.get("draft_year"),
            "status": meta.get("status"),
            "injury_status": meta.get("injury_status"),
            "gsis_id": clean_gsis if clean_gsis else None,
        }
        pid_pos[pid] = (pid_meta[pid]["pos"] or "").upper()

    # Enrich pid_meta with nflverse player metadata (authoritative rookie_season,
    # birth_date, position). Sleeper's draft_year/years_exp are often missing or
    # current-relative, which is why Rookie? produced False for every row even
    # for known rookies (Chase 2021, Williams 2024, etc.).
    try:
        nfl_players = _safe_df(load_nflverse_player_ids(ext))
    except Exception as e:
        nfl_players = pd.DataFrame()
        _log_exc(debug, "load_nflverse_player_ids_enrich", e)

    if not nfl_players.empty and "gsis_id" in nfl_players.columns:
        nfl_players = nfl_players.copy()
        nfl_players["gsis_id"] = nfl_players["gsis_id"].astype(str)
        nfl_by_gsis = {
            str(row.get("gsis_id")): row
            for _, row in nfl_players.iterrows()
            if isinstance(row.get("gsis_id"), str) and row.get("gsis_id")
        }

        # Last-name index for validating Sleeper's gsis_id against the
        # actual NFLverse player. Sleeper has a small number of swapped
        # gsis_ids — e.g. sid=5133 (Tyler Conklin TE) carries gsis
        # 00-0034439 which is actually Ryan Izzo; sid=5094 (Ryan Izzo)
        # carries 00-0034270 which is actually Conklin. Without a check
        # we silently look up the wrong player's NFLverse stats and
        # produce Points (full season) = 0 for one of them.
        def _last_name_key(s: Any) -> str:
            s2 = "".join(c if c.isalpha() or c == " " else "" for c in str(s).lower())
            parts = [p for p in s2.split() if p not in ("jr", "sr", "ii", "iii", "iv", "v")]
            return parts[-1] if parts else ""

        gsis_to_lastname: Dict[str, str] = {}
        name_col = "display_name" if "display_name" in nfl_players.columns else (
            "football_name" if "football_name" in nfl_players.columns else None
        )
        if name_col:
            for g, row in nfl_by_gsis.items():
                gsis_to_lastname[g] = _last_name_key(row.get(name_col))

        for pid, m in pid_meta.items():
            sleeper_gsis = m.get("gsis_id")
            # If Sleeper's gsis_id points to a player with a different
            # last name, treat it as bad and prefer DP/NFLverse mappings.
            # Compares against the canonical NFLverse player at that gsis.
            if sleeper_gsis and gsis_to_lastname:
                nf_last = gsis_to_lastname.get(str(sleeper_gsis))
                s_last = _last_name_key(m.get("full_name"))
                if nf_last and s_last and nf_last != s_last:
                    dp_gsis = dp_sleeper_to_gsis.get(str(pid))
                    nv_gsis = sleeper_to_gsis.get(str(pid))
                    # Pick the first replacement whose last name actually
                    # matches the Sleeper player.
                    for cand in (dp_gsis, nv_gsis):
                        if cand and gsis_to_lastname.get(str(cand)) == s_last:
                            _log(debug, f"[{_now_iso()}] INFO gsis_correction sid={pid} name={m.get('full_name')!r} sleeper_gsis={sleeper_gsis} -> {cand}")
                            m["gsis_id"] = cand
                            sleeper_gsis = cand
                            break
            # Look up gsis_id via the full chain — Sleeper's meta lacks gsis_id for
            # many players (e.g. Ja'Marr Chase, Caleb Williams). DP's db_playerids
            # provides the sleeper_id <-> gsis_id mapping that closes the gap.
            gsis = (
                sleeper_gsis
                or dp_sleeper_to_gsis.get(str(pid))
                or sleeper_to_gsis.get(str(pid))
            )
            if not gsis:
                continue
            # Backfill gsis_id onto pid_meta so the per-week injury / team
            # lookups also benefit (otherwise they re-resolve every row).
            if not m.get("gsis_id"):
                m["gsis_id"] = gsis
            nrow = nfl_by_gsis.get(str(gsis))
            if nrow is None:
                continue
            # rookie_season is authoritative; override Sleeper's draft_year when nflverse has it
            rs = nrow.get("rookie_season")
            try:
                rs_int = int(float(rs)) if rs is not None and str(rs) != "nan" else None
            except Exception:
                rs_int = None
            if rs_int is not None:
                m["draft_year"] = rs_int
            # birth_date: prefer nflverse if Sleeper is missing it
            if not m.get("birth_date"):
                bd = nrow.get("birth_date")
                if bd is not None and str(bd) != "nan":
                    m["birth_date"] = str(bd)
            # latest_team / draft_team fallback: when Sleeper shows team=None
            # (retired / free agent / out of league), use nflverse's last-known
            # team so per-week NFL team isn't NAN for players whose career ended
            # within the league window (Tarik Cohen 2021 - BAL, Josh Doctson 2021
            # - WAS, etc.). We only fill it when Sleeper has nothing.
            if not m.get("team"):
                lt = nrow.get("latest_team") or nrow.get("draft_team")
                if lt is not None and str(lt) != "nan" and str(lt).strip():
                    m["team"] = _norm_team(str(lt))

    # ------------- League chain -------------
    leagues = _walk_league_chain(sc, run_cfg.league_id, run_cfg.min_season, run_cfg.max_season)
    raw_dir = repo_root / "exports" / "raw"
    raw_dir.mkdir(parents=True, exist_ok=True)

    # Phase 8F: per picks row, the ordered trades-sheet refs ("T#N") for its
    # i5 (#15): full Sleeper-style asset history for hover-comments. Picks keyed
    # by picks.csv row index; players by display name. A player's history begins
    # with the history of the pick they were drafted at, then continues through
    # every later add / drop / trade. Populated in the per-asset link pass.
    pick_history_text: Dict[int, str] = {}
    player_history_text: Dict[str, str] = {}

    def write_outputs(tables: List[Tuple[str, pd.DataFrame, str]]) -> None:
        out_dir = repo_root / "exports"
        out_dir.mkdir(exist_ok=True)
        for fname, frame, plan_key in tables:
            cols = catalog.get(plan_key, [])
            if plan_key in {"team-year", "team-all-time"}:
                cols = _append_team_vs_columns(frame, cols, plan_key)
            frame = _safe_df(frame)
            out = _ensure_plan_columns(frame, cols)
            out = _fill_empty_columns(out, cols)
            out = _fill_missing_values(out, cols)
            # Round numeric columns at output to suppress float noise
            # (e.g., Luck = 188.49633333333333, change-in-win% = -0.05879999..).
            # 4 decimals is enough precision for fantasy football stats.
            # Previous version used errors='ignore' which pandas treats as
            # 'leave unchanged' for any column with a non-numeric value, so
            # rounding silently no-op'd on most columns.
            try:
                for c in out.columns:
                    if _column_kind(c) == "numeric":
                        try:
                            coerced = pd.to_numeric(out[c], errors="coerce")
                            # Skip coercion if any non-empty value failed to
                            # parse as numeric (e.g. pick_history's "2021
                            # (vet)" Year tag). Preserves intentional text
                            # tags in otherwise-numeric columns.
                            raw_str = out[c].astype(str).str.strip()
                            non_numeric = out[c].notna() & coerced.isna() & raw_str.ne("") & raw_str.str.lower().ne("nan")
                            if non_numeric.any():
                                continue
                            # Only overwrite when we actually got numeric data.
                            if coerced.notna().any():
                                rounded = coerced.round(4)
                                # COUNT columns render as integers ('1' not
                                # '1.0') so same-family columns read uniformly
                                # (audit run-2 F7: dtype luck made e.g. Most-QBs
                                # render 1 while Most-TE rendered 1.0).
                                _cl0 = str(c).strip().lower()
                                if (any(_cl0.startswith(p) for p in (
                                        "number of", "most number of", "times ",
                                        "weeks ", "total number"))
                                        and rounded.dropna().mod(1).eq(0).all()):
                                    rounded = rounded.astype("Int64")
                                if _preserve_na(c):
                                    # Keep 'N/A' as a string for first-occurrence
                                    # rows (no prior data to compare against).
                                    out[c] = rounded.astype(object).where(rounded.notna(), "N/A")
                                else:
                                    out[c] = rounded
                        except Exception:
                            continue
            except Exception as e:
                _log_exc(debug, f"round_numeric_{fname}", e)
            try:
                require_columns(out, cols, fname.replace(".csv", ""))
            except Exception as e:
                _log_exc(debug, f"require_columns_{fname}", e)
            out.to_csv(out_dir / fname, index=False)

        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment, PatternFill, Font
            from openpyxl.formatting.rule import ColorScaleRule

            wb = Workbook()
            wb.remove(wb.active)

            # Iterate the `tables` list (DOCUMENT_MODULES order) so xlsx tab
            # order matches the spec. Previously this used
            # sorted(out_dir.glob("*.csv")), which sorted tabs alphabetically
            # and ignored the configured order.
            _csv_order = [fname for (fname, _, _) in tables]

            # Phase 11D: player-name hyperlink anchors. Every single-name cell
            # links to that player's player_all_time row; per-week references
            # (Reference player name) link to the player_week row for that exact
            # (year, week). Built from the already-written CSVs (CSV order ==
            # xlsx row order; sheet row = csv index + 2 for the header).
            def _yw(v) -> str:
                s = str(v).strip()
                try:
                    return str(int(float(s)))
                except Exception:
                    return s
            name_to_patrow: Dict[str, int] = {}
            try:
                _pat = pd.read_csv(out_dir / "player_all_time.csv")
                if "Player" in _pat.columns:
                    for i, nm in enumerate(_pat["Player"].astype(str)):
                        nm = nm.strip()
                        if nm and nm not in name_to_patrow:
                            name_to_patrow[nm] = i + 2
            except Exception:
                pass
            # i7 (#27): team display name -> team_all_time row, for hyperlinking
            # single-team reference cells (Opponent, Top/Last team, Team-for-X).
            team_to_tatrow: Dict[str, int] = {}
            try:
                _tat = pd.read_csv(out_dir / "team_all_time.csv")
                if "Team" in _tat.columns:
                    for i, nm in enumerate(_tat["Team"].astype(str)):
                        nm = nm.strip()
                        if nm and nm not in team_to_tatrow:
                            team_to_tatrow[nm] = i + 2
            except Exception:
                pass
            # i7 (#32): normalized column name -> Formulas-sheet definition, for
            # the header hover-tooltips attached below.
            try:
                _col_defs = formulas.column_definitions()
            except Exception:
                _col_defs = {}
            nameyw_to_pwrow: Dict[Tuple[str, str, str], int] = {}
            try:
                _pwk = pd.read_csv(out_dir / "player_week.csv", low_memory=False)
                if {"Player", "Year", "Week"}.issubset(_pwk.columns):
                    for i, (nm, yr, wk) in enumerate(zip(_pwk["Player"].astype(str), _pwk["Year"], _pwk["Week"])):
                        key = (nm.strip(), _yw(yr), _yw(wk))
                        if key not in nameyw_to_pwrow:
                            nameyw_to_pwrow[key] = i + 2
            except Exception:
                pass

            _written = set()
            for fname in _csv_order:
                csvf = out_dir / fname
                if not csvf.exists():
                    continue
                _written.add(csvf.name)
                sheet_name = csvf.stem[:31]
                ws = wb.create_sheet(title=sheet_name)

                try:
                    d = pd.read_csv(csvf, low_memory=False)
                except Exception:
                    d = pd.DataFrame()

                # Terminal-encoded streak columns hold a mix of integers (a
                # run's length on its final week), the text "In Progress", and
                # 0 / "N/A". read_csv types the whole column as text, so the
                # xlsx would store the lengths as strings and they wouldn't sort
                # numerically (breaking the top-N longest-streak use case).
                # Re-type each cell: numbers -> int, keep "In Progress"/"N/A".
                for _scol in d.columns:
                    _cl = str(_scol)
                    if not (_cl.endswith(" streak") or _cl == "Win streak vs this opponent"):
                        continue
                    def _streak_cell(v: Any) -> Any:
                        s = str(v).strip()
                        if s in ("In Progress", "N/A"):
                            return s
                        # Blank/NaN only arises for season-grain streaks in a
                        # not-yet-played season (read_csv turns the CSV's "N/A"
                        # into NaN); surface it as "N/A" to match the sheet.
                        if s in ("", "nan", "None"):
                            return "N/A"
                        try:
                            f = float(s)
                            return int(f) if f == int(f) else f
                        except Exception:
                            return v
                    d[_scol] = d[_scol].map(_streak_cell).astype(object)

                # Phase 11B: the Formulas tab is a reference doc, not a data
                # table — render it grouped into color-coded sections by the
                # sheet each stat belongs to, with a styled header and wrapped
                # Formula/Notes. (CSV row order is unchanged; this is xlsx-only.)
                if sheet_name == "formulas" and {"Stat", "Sheet", "Formula", "Notes"}.issubset(set(d.columns)):
                    _fcols = ["Stat", "Sheet", "Formula", "Notes"]

                    def _fcat(sv):
                        s = str(sv).lower()
                        if "player" in s: return "Player sheets"
                        if "team" in s: return "Team sheets"
                        if "league" in s: return "League sheets"
                        if "transaction" in s: return "Transactions"
                        if "trade" in s: return "Trades"
                        if "pick" in s: return "Picks"
                        return "Other"
                    _forder = ["Player sheets", "Team sheets", "League sheets",
                               "Transactions", "Trades", "Picks", "Other"]
                    _fband = {  # (row fill, section-header fill)
                        "Player sheets": ("DDEBF7", "9DC3E6"),
                        "Team sheets":   ("E2EFDA", "A9D08E"),
                        "League sheets": ("FFF2CC", "FFD966"),
                        "Transactions":  ("FCE4D6", "F4B183"),
                        "Trades":        ("E4DFEC", "B4A7D6"),
                        "Picks":         ("F2F2F2", "BFBFBF"),
                        "Other":         ("F2F2F2", "BFBFBF"),
                    }
                    _wrap = Alignment(wrap_text=True, vertical="top")
                    # Header row
                    ws.append(_fcols)
                    for j in range(1, len(_fcols) + 1):
                        c = ws.cell(row=1, column=j)
                        c.fill = PatternFill("solid", fgColor="305496")
                        c.font = Font(bold=True, color="FFFFFF")
                        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
                    # Group rows by category (preserve _ROWS order within a group)
                    _by = {k: [] for k in _forder}
                    for rec in d.to_dict("records"):
                        _by[_fcat(rec.get("Sheet"))].append(rec)
                    _r = 1
                    for cat in _forder:
                        recs = _by.get(cat) or []
                        if not recs:
                            continue
                        _r += 1
                        ws.cell(row=_r, column=1, value=cat.upper())
                        ws.merge_cells(start_row=_r, start_column=1, end_row=_r, end_column=len(_fcols))
                        sh = ws.cell(row=_r, column=1)
                        sh.fill = PatternFill("solid", fgColor=_fband[cat][1])
                        sh.font = Font(bold=True)
                        sh.alignment = Alignment(vertical="center", horizontal="left")
                        _rowfill = PatternFill("solid", fgColor=_fband[cat][0])
                        for rec in recs:
                            _r += 1
                            for j, col in enumerate(_fcols, 1):
                                cell = ws.cell(row=_r, column=j, value=rec.get(col, ""))
                                cell.fill = _rowfill
                                cell.alignment = _wrap
                                if col == "Stat":
                                    cell.font = Font(bold=True)
                    for j, col in enumerate(_fcols, 1):
                        ws.column_dimensions[get_column_letter(j)].width = {
                            "Stat": 36, "Sheet": 30, "Formula": 90, "Notes": 70}.get(col, 20)
                    ws.freeze_panes = "A2"
                    ws.sheet_properties.tabColor = _FAMILY_TAB["formulas"]
                    continue

                # Cross-table row references in any "Link to ..." cell:
                # "#N" -> transactions row N, "T#N" -> trades row N, "PH#N" ->
                # pick_history row N (1-indexed; xlsx adds 1 for the header).
                _ref_re = re.compile(r"^(PH|T)?#(\d+)$")
                _ref_sheet = {"": "transactions", "T": "trades", "PH": "picks"}

                def _set_ref_link(cell, ref):
                    m = _ref_re.match(str(ref).strip())
                    if not m:
                        return
                    cell.hyperlink = f"#'{_ref_sheet[m.group(1) or '']}'!A{int(m.group(2)) + 1}"
                    cell.style = "Hyperlink"

                # Trades: explode each per-asset link column into one clickable
                # column PER received asset, under a merged group header. Each
                # cell shows the asset NAME and hyperlinks to that asset's
                # next/previous event. (The CSV keeps the ';'-joined ref list.)
                _pa_cols = [c for c in d.columns
                            if "link to " in str(c).lower() and "per asset" in str(c).lower()]
                _did_expand = False
                if sheet_name == "trades" and _pa_cols and "Assets received" in d.columns and not d.empty:
                    _did_expand = True
                    def _split(s):
                        return [t.strip() for t in str(s if pd.notna(s) else "").split(";")]
                    def _names(s):
                        return [t for t in _split(s) if t and t.upper() != "N/A"]
                    _recv = [_names(v) for v in d["Assets received"]]
                    K = max((len(n) for n in _recv), default=1) or 1
                    headers, spec = [], []
                    for c in d.columns:
                        if c in _pa_cols:
                            for s in range(K):
                                headers.append(c if s == 0 else "")
                                spec.append(("asset", c, s))
                        else:
                            headers.append(c)
                            spec.append(("plain", c))
                    ws.append(headers)
                    _records = d.to_dict("records")
                    _ref_lists = {c: [_split(v) for v in d[c]] for c in _pa_cols}
                    for ri, rec in enumerate(_records):
                        outrow = []
                        for sp in spec:
                            if sp[0] == "plain":
                                v = rec.get(sp[1])
                                outrow.append(None if (isinstance(v, float) and pd.isna(v)) else v)
                            else:
                                nm = _recv[ri]
                                outrow.append(nm[sp[2]] if sp[2] < len(nm) else None)
                        ws.append(outrow)
                    for ci, sp in enumerate(spec, start=1):
                        if sp[0] != "asset":
                            continue
                        for ri in range(len(d)):
                            cell = ws.cell(row=ri + 2, column=ci)
                            refs = _ref_lists[sp[1]][ri]
                            if sp[2] < len(refs):
                                _set_ref_link(cell, refs[sp[2]])
                            # Bug #6: an asset with no next/previous event ref
                            # (a chain endpoint) still shows its NAME — link that
                            # name to the asset's home so every non-FAAB cell is a
                            # working link. Players -> player_all_time row; picks
                            # already fall back to their PH# home ref above; FAAB
                            # ("$N FAAB") has no home page and stays unlinked.
                            if cell.hyperlink is None and cell.value is not None:
                                _home_row = name_to_patrow.get(str(cell.value).strip())
                                if _home_row:
                                    cell.hyperlink = f"#'player_all_time'!A{_home_row}"
                                    cell.style = "Hyperlink"
                    # NB: we intentionally do NOT merge the per-asset group
                    # headers. A merged cell in the header row makes Excel
                    # refuse to apply an auto-filter (the table stops being
                    # sortable / re-orderable). The group label already sits
                    # in the first sub-column header (blank on the rest), so
                    # the columns stay visually grouped while remaining a
                    # filterable, sortable table.
                else:
                    ws.append(list(d.columns))
                    for row in d.itertuples(index=False, name=None):
                        ws.append(list(row))
                    link_col_idx = [j for j, col in enumerate(d.columns, 1)
                                    if "link to " in str(col).lower()]
                    for r in range(2, ws.max_row + 1):
                        for j in link_col_idx:
                            cell = ws.cell(row=r, column=j)
                            if cell.value is None:
                                continue
                            # single ref, or first ref of a ';'-joined list
                            first = next((t.strip() for t in str(cell.value).split(";")
                                          if _ref_re.match(t.strip())), None)
                            if first:
                                _set_ref_link(cell, first)
                # Phase 11C-2: freeze through the pinned columns (team_week also
                # pins Opponent -> 5 cols).
                _pin_n = 5 if sheet_name == "team_week" else 4
                ws.freeze_panes = f"{get_column_letter(_pin_n + 1)}2"

                # Family tab color.
                _fam = next((k for k in _FAMILY_TAB if k in sheet_name.lower()), None)
                if _fam:
                    ws.sheet_properties.tabColor = _FAMILY_TAB[_fam]

                # Style the header row: bold + wrap + a per-topic color band so
                # adjacent same-topic columns read as a group. Also (i7) band the
                # DATA columns in a subtle two-tone WITHIN each topic run (#34) and
                # tint "In Progress" cells amber so active streaks stand out (#33).
                try:
                    from openpyxl.comments import Comment as _HdrComment  # i7 #32
                    _data_wrap = Alignment(wrap_text=True, vertical="top")  # Phase 12 #7: wrap ALL cells
                    _band_fill = PatternFill("solid", fgColor="F2F2F2")  # 2nd tone (light gray)
                    _inprog_fill = PatternFill("solid", fgColor="C6EFCE")  # opaque light green (#33)
                    _topics = [(_col_topic(ws.cell(row=1, column=j).value)
                                if ws.cell(row=1, column=j).value else "Identity")
                               for j in range(1, ws.max_column + 1)]
                    # Parity of each column within its contiguous same-topic run.
                    _parity = []
                    _runi = 0
                    for _ci in range(len(_topics)):
                        _runi = _runi + 1 if (_ci and _topics[_ci] == _topics[_ci - 1]) else 0
                        _parity.append(_runi % 2)
                    for j in range(1, ws.max_column + 1):
                        hc = ws.cell(row=1, column=j)
                        topic = _topics[j - 1]
                        hc.fill = PatternFill("solid", fgColor=_TOPIC_FILL.get(topic, "D9D9D9"))
                        hc.font = Font(bold=True)
                        hc.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                        # i7 (#32): hover-tooltip on the header pulling this column's
                        # Formulas-sheet definition, so cryptic headers explain
                        # themselves. Identity/label columns (Team, Player, Year,
                        # Points, …) are self-evident -> skipped even if some entry
                        # happens to document them.
                        _hn = re.sub(r"\s+", " ", str(hc.value).strip().lower()) if hc.value else ""
                        # Prefer the definition documented for THIS sheet (so e.g.
                        # picks 'Number of trades' doesn't get the team note), then
                        # fall back to the global one.
                        _def = ((_col_defs.get((sheet_name, _hn)) or _col_defs.get((None, _hn)))
                                if _hn and _hn not in formulas.IDENTITY_ALLOWLIST else None)
                        if _def and hc.comment is None:
                            _hc_cm = _HdrComment(_def, "LOTG Formulas")
                            _hc_cm.width = 460
                            _hc_nl = _def.count("\n") + 1 + len(_def) // 70
                            _hc_cm.height = min(520, max(80, 16 * _hc_nl))
                            hc.comment = _hc_cm
                        # Wrap every data cell (#7) + conservative number format.
                        nf = _col_number_format(hc.value) if hc.value else None
                        _band = _band_fill if _parity[j - 1] else None
                        for r in range(2, ws.max_row + 1):
                            dc = ws.cell(row=r, column=j)
                            dc.alignment = _data_wrap
                            if nf:
                                dc.number_format = nf
                            if str(dc.value).strip() == "In Progress":
                                dc.fill = _inprog_fill
                            elif _band is not None:
                                dc.fill = _band
                except Exception:
                    pass

                # Phase 11D: hyperlink single player-name cells. Most link to the
                # player's player_all_time row; 'Reference player name' (a
                # specific player in a specific week) links to the player_week
                # row for that (year, week). Multi-name list cells (trades
                # Assets …) are left to their existing per-asset event links.
                try:
                    _link_pat = []   # columns -> player_all_time
                    _link_pw = []    # columns -> player_week (per-week ref)
                    if sheet_name in ("player_week", "player_year", "player_all_time"):
                        _link_pat.append("Player")
                    if sheet_name == "picks":
                        _link_pat.append("Player Picked")
                    if sheet_name == "transactions":
                        _link_pat += ["Player Added", "Player Dropped"]
                    if sheet_name == "player_week":
                        _link_pw.append("Reference player name")
                    if _link_pat or _link_pw:
                        _hdr = [ws.cell(row=1, column=j).value for j in range(1, ws.max_column + 1)]
                        _yr_j = _hdr.index("Year") + 1 if "Year" in _hdr else None
                        _wk_j = _hdr.index("Week") + 1 if "Week" in _hdr else None
                        _bluefont = Font(color="0563C1")

                        def _linkcell(cell, target):
                            cell.hyperlink = target
                            cell.font = _bluefont

                        for colname in _link_pat:
                            if colname not in _hdr:
                                continue
                            j = _hdr.index(colname) + 1
                            for r in range(2, ws.max_row + 1):
                                nm = str(ws.cell(row=r, column=j).value or "").strip()
                                if not nm or nm.lower() in ("nan", "n/a"):
                                    continue
                                row = name_to_patrow.get(nm)
                                if row:
                                    _linkcell(ws.cell(row=r, column=j), f"#'player_all_time'!A{row}")
                        for colname in _link_pw:
                            if colname not in _hdr or _yr_j is None or _wk_j is None:
                                continue
                            j = _hdr.index(colname) + 1
                            for r in range(2, ws.max_row + 1):
                                nm = str(ws.cell(row=r, column=j).value or "").strip()
                                if not nm or nm.lower() in ("nan", "n/a"):
                                    continue
                                key = (nm, _yw(ws.cell(row=r, column=_yr_j).value), _yw(ws.cell(row=r, column=_wk_j).value))
                                row = nameyw_to_pwrow.get(key)
                                if row:
                                    _linkcell(ws.cell(row=r, column=j), f"#'player_week'!A{row}")
                except Exception as e:
                    _log_exc(debug, "player_name_hyperlinks", e)

                # i7 (#27): team-name cells link by CONTEXT, not always all-time.
                #   - team_week "Opponent" -> the opponent's row of that SAME
                #     (year, week) matchup.
                #   - trades "Team's traded with N" -> the counterparty's row of
                #     that SAME trade (matched on the shared Date timestamp).
                #   - every OTHER single-team reference cell (Top/Last team, Team
                #     for highest/lowest %/Win%) -> that team's team_all_time row.
                _bluefont2 = Font(color="0563C1")
                _hdr2 = [str(ws.cell(row=1, column=j).value or "")
                         for j in range(1, ws.max_column + 1)]

                def _hcol(name):
                    return _hdr2.index(name) + 1 if name in _hdr2 else None

                try:
                    # Opponent -> same-week matchup row.
                    if sheet_name == "team_week" and {"Opponent", "Year", "Week"} <= set(_hdr2):
                        _tj, _yj, _wj, _oj = _hcol("Team"), _hcol("Year"), _hcol("Week"), _hcol("Opponent")
                        _twrow: Dict[Tuple[str, str, str], int] = {}
                        for r in range(2, ws.max_row + 1):
                            _twrow[(str(ws.cell(row=r, column=_tj).value or "").strip(),
                                    _yw(ws.cell(row=r, column=_yj).value),
                                    _yw(ws.cell(row=r, column=_wj).value))] = r
                        for r in range(2, ws.max_row + 1):
                            _opp = str(ws.cell(row=r, column=_oj).value or "").strip()
                            _row = _twrow.get((_opp, _yw(ws.cell(row=r, column=_yj).value),
                                               _yw(ws.cell(row=r, column=_wj).value)))
                            if _row:
                                _c = ws.cell(row=r, column=_oj)
                                _c.hyperlink = f"#'team_week'!A{_row}"
                                _c.font = _bluefont2
                    # Trade counterparties -> the other team's row of the same trade.
                    if sheet_name == "trades" and "Date" in _hdr2:
                        _dj, _tj = _hcol("Date"), _hcol("Team")
                        _trrow: Dict[Tuple[str, str], int] = {}
                        for r in range(2, ws.max_row + 1):
                            _trrow[(str(ws.cell(row=r, column=_dj).value or "").strip(),
                                    str(ws.cell(row=r, column=_tj).value or "").strip())] = r
                        for _cn in [c for c in _hdr2 if c.startswith("Team's traded with")]:
                            _cj = _hcol(_cn)
                            for r in range(2, ws.max_row + 1):
                                _ct = str(ws.cell(row=r, column=_cj).value or "").strip()
                                if not _ct:
                                    continue
                                _row = _trrow.get((str(ws.cell(row=r, column=_dj).value or "").strip(), _ct))
                                if _row:
                                    _c = ws.cell(row=r, column=_cj)
                                    _c.hyperlink = f"#'trades'!A{_row}"
                                    _c.font = _bluefont2
                except Exception as e:
                    _log_exc(debug, "matchup_trade_hyperlinks", e)

                # Pure team-reference cells -> team_all_time. A column qualifies
                # when EVERY non-empty cell is a known team; the identity "Team"
                # column and the context columns above (Opponent / Team's traded
                # with N) are excluded so they keep their contextual links.
                try:
                    if team_to_tatrow:
                        _teams = set(team_to_tatrow)
                        _blank = {"", "nan", "n/a", "in progress"}
                        for _cn in d.columns:
                            _cns = str(_cn).strip()
                            if (_cns in ("Team", "Opponent") or _cns.startswith("Team's traded with")
                                    or _cn not in _hdr2):
                                continue
                            _nb = [str(v).strip() for v in d[_cn].tolist()
                                   if str(v).strip().lower() not in _blank]
                            if not _nb or any(v not in _teams for v in _nb):
                                continue  # not an all-team reference column
                            j = _hdr2.index(_cn) + 1
                            for r in range(2, ws.max_row + 1):
                                if ws.cell(row=r, column=j).hyperlink is not None:
                                    continue
                                _row = team_to_tatrow.get(str(ws.cell(row=r, column=j).value or "").strip())
                                if _row:
                                    _c = ws.cell(row=r, column=j)
                                    _c.hyperlink = f"#'team_all_time'!A{_row}"
                                    _c.font = _bluefont2
                except Exception as e:
                    _log_exc(debug, "team_name_hyperlinks", e)

                # i5 (#15): full asset-history hover-comment on COLUMN 1 — picks
                # (per pick) and player_all_time (per player). A small red marker
                # on the cell; hover/click reveals the whole history, else hidden.
                try:
                    from openpyxl.comments import Comment

                    def _attach_hist(_cell, _txt):
                        _cm = Comment(_txt, "LOTG")
                        # Size the box for the WRAPPED line count, not the logical
                        # one: trade lines run ~150 chars and wrap to 2-3 visual
                        # rows at this width, so a flat 15px/line box clipped long
                        # histories (only ~6 of 11 trades showed). Estimate visual
                        # rows from each line's length at ~88 chars/row.
                        _w = 560
                        _cpr = 88
                        _vis = sum(max(1, -(-len(_l) // _cpr)) for _l in _txt.split("\n"))
                        _cm.width = _w
                        _cm.height = min(1100, max(90, 15 * _vis + 12))
                        _cell.comment = _cm

                    if sheet_name == "picks" and pick_history_text and not d.empty:
                        for _ri in range(len(d)):
                            _txt = pick_history_text.get(_ri)
                            if _txt:
                                _attach_hist(ws.cell(row=_ri + 2, column=1), _txt)
                    elif sheet_name == "player_all_time" and player_history_text and "Player" in d.columns:
                        _pj = list(d.columns).index("Player") + 1
                        for _r in range(2, ws.max_row + 1):
                            _nm = str(ws.cell(row=_r, column=_pj).value or "").strip()
                            _txt = player_history_text.get(_nm)
                            if _txt:
                                _attach_hist(ws.cell(row=_r, column=1), _txt)
                except Exception as e:
                    _log_exc(debug, "asset_history_comments", e)

                # Phase 11E: light red->yellow->green 3-color scale on each
                # sheet's headline value column (O-Score; team Win %/PF; player
                # PAR/Points) so good/bad pops at a glance.
                try:
                    _scale_cols = {
                        "picks": "O-Score", "transactions": "O-Score", "trades": "O-Score",
                        "team_year": "Win %", "team_all_time": "All time win %", "team_week": "PF",
                        "player_year": "Starter PAR", "player_all_time": "Starter PAR",
                        "player_week": "Points",
                    }
                    _scol = _scale_cols.get(sheet_name)
                    if _scol:
                        _hdr2 = [ws.cell(row=1, column=j).value for j in range(1, ws.max_column + 1)]
                        if _scol in _hdr2 and ws.max_row >= 2:
                            _cl = get_column_letter(_hdr2.index(_scol) + 1)
                            ws.conditional_formatting.add(
                                f"{_cl}2:{_cl}{ws.max_row}",
                                ColorScaleRule(
                                    start_type="percentile", start_value=5, start_color="F8696B",
                                    mid_type="percentile", mid_value=50, mid_color="FFEB84",
                                    end_type="percentile", end_value=95, end_color="63BE7B"),
                            )
                except Exception as e:
                    _log_exc(debug, "color_scale", e)

                # i7 (#30): highlight RECORDS — the highest (gold) and lowest
                # (blue) value of EVERY comparable numeric column on EVERY sheet,
                # so each stat's record / anti-record holder pops. Gold/blue mean
                # simply highest/lowest value, not good/bad. All ties are colored,
                # EXCEPT a degenerate extreme tied by a large share of the column
                # (e.g. the thousands of 0-point weeks on player_week) — that's the
                # common value, not a record, so it's skipped to avoid flooding.
                try:
                    if ws.max_row >= 3:
                        _hi_fill = PatternFill("solid", fgColor="FFD966")  # record high
                        _lo_fill = PatternFill("solid", fgColor="BDD7EE")  # record low
                        _rec_skip = {"player id", "season", "year", "week", "number"}
                        _hdr2 = [str(ws.cell(row=1, column=k).value or "")
                                 for k in range(1, ws.max_column + 1)]
                        _hpos = {c: i + 1 for i, c in enumerate(_hdr2)}
                        for _cn in d.columns:
                            if str(_cn).strip().lower() in _rec_skip or _cn not in _hpos:
                                continue
                            _num = pd.to_numeric(
                                d[_cn].replace({"N/A": None, "In Progress": None, "": None}),
                                errors="coerce")
                            _n_valid = int(_num.notna().sum())
                            if _num.nunique(dropna=True) < 2:
                                continue  # all-equal / single value isn't a record
                            # A record can be tied by a few; an extreme shared by a
                            # large fraction is just the common value -> not a record.
                            _cap = max(8, int(_n_valid * 0.02))
                            _arr = _num.to_numpy()
                            j = _hpos[_cn]
                            for _ext, _fill in ((np.nanmax(_arr), _hi_fill),
                                                (np.nanmin(_arr), _lo_fill)):
                                _pos = np.where(_arr == _ext)[0]
                                if len(_pos) > _cap:
                                    continue
                                for _p in _pos:
                                    ws.cell(row=int(_p) + 2, column=j).fill = _fill
                except Exception as e:
                    _log_exc(debug, "record_highlight", e)

                # Auto-filter every sheet (incl. trades) so the tables stay
                # sortable/re-orderable. The per-asset expansion no longer
                # merges header cells, so the filter applies cleanly.
                if ws.max_column >= 1:
                    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{max(1, ws.max_row)}"

                try:
                    for j in range(1, ws.max_column + 1):
                        vals = [str(ws.cell(row=1, column=j).value or "")]
                        for r in range(2, min(ws.max_row, 201) + 1):
                            vals.append(str(ws.cell(row=r, column=j).value or ""))
                        ws.column_dimensions[get_column_letter(j)].width = min(40, max(10, max(len(v) for v in vals) + 2))
                except Exception:
                    pass

                # Phase 11C-2: tame the trades per-asset link explosion — narrow
                # the blank-header slot columns and hide any that are fully empty.
                if sheet_name == "trades":
                    try:
                        for j in range(1, ws.max_column + 1):
                            if (ws.cell(row=1, column=j).value or "") != "":
                                continue  # only the exploded (blank-header) slots
                            letter = get_column_letter(j)
                            ws.column_dimensions[letter].width = 16
                            if all((ws.cell(row=r, column=j).value in (None, "")) for r in range(2, ws.max_row + 1)):
                                ws.column_dimensions[letter].hidden = True
                    except Exception:
                        pass

            wb.save(out_dir / "LOTG_Stats.xlsx")
            # i5: bold the action verbs in the history hover-comments (openpyxl
            # flattens comments to plain text). Self-restoring on any failure.
            try:
                _bold_comment_verbs(out_dir / "LOTG_Stats.xlsx", debug)
            except Exception as e:
                _log_exc(debug, "bold_comment_verbs", e)
        except Exception as e:
            _log_exc(debug, "excel_write", e)

        try:
            import zipfile
            with zipfile.ZipFile(out_dir / "LOTG_Exports.zip", "w", zipfile.ZIP_DEFLATED) as z:
                for f in out_dir.glob("*.csv"):
                    z.write(f, arcname=f.name)
                if (out_dir / "LOTG_Stats.xlsx").exists():
                    z.write(out_dir / "LOTG_Stats.xlsx", arcname="LOTG_Stats.xlsx")
                for f in (out_dir / "raw").glob("*"):
                    if f.is_file():
                        z.write(f, arcname=f"raw/{f.name}")
        except Exception as e:
            _log_exc(debug, "zip_exports", e)

        _log(debug, f"[{_now_iso()}] ===== Build end =====")

    if not leagues:
        fallback_dir = repo_root / "data"
        fallback_tables = []
        for doc in DOCUMENT_MODULES:
            src = fallback_dir / doc.FILE_NAME
            if src.exists():
                try:
                    fallback_tables.append((doc.FILE_NAME, pd.read_csv(src), doc.PLAN_KEY))
                except Exception:
                    fallback_tables.append((doc.FILE_NAME, pd.DataFrame(), doc.PLAN_KEY))
            else:
                fallback_tables.append((doc.FILE_NAME, pd.DataFrame(), doc.PLAN_KEY))
        _log(debug, f"[{_now_iso()}] WARN no leagues found; using fallback data/ outputs")
        write_outputs(fallback_tables)
        return

    league_seasons = [_to_int(lg.get("season"), None) for lg in leagues]
    league_seasons = [s for s in league_seasons if s is not None]
    current_season_for_rookies = max(league_seasons) if league_seasons else (_to_int(run_cfg.max_season, None) or date.today().year)
    rookie_season_by_pid = {
        pid: _rookie_season(meta, current_season_for_rookies)
        for pid, meta in pid_meta.items()
    }

    def is_rookie_pid(pid: Any, season: int) -> bool:
        rookie_season = rookie_season_by_pid.get(str(pid))
        return rookie_season is not None and int(season) == int(rookie_season)

    # ------------- Output rows -------------
    player_week_rows: List[Dict[str, Any]] = []
    team_week_rows: List[Dict[str, Any]] = []
    transactions_rows: List[Dict[str, Any]] = []
    # Cross-season per-player NFL game log (sourced from nflverse, so
    # weeks count as "played" whether or not the player was on a
    # fantasy roster at the time). Used by the transactions polish
    # pass to compute pre-pickup PPG without being limited to pw —
    # rookies and UFAs picked up after their NFL debut now resolve.
    # Keyed by Sleeper sleeper_id; each entry: {year, week, points,
    # _wk_date}. Points use fantasy_points_ppr as the approximation
    # (most leagues are 0.5-PPR or full PPR; rankings are stable
    # between the two for trend purposes).
    nfl_log_by_sid: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    trades_rows: List[Dict[str, Any]] = []
    # Orphan drops: a player dropped without a corresponding add in the same
    # transaction (pure waiver-to-FA). These don't get a transactions.csv row
    # but they still matter for the per-player event log when computing
    # 'Weeks between pickup and start' / 'Date dropped/traded'.
    orphan_drop_events: List[Dict[str, Any]] = []
    pick_rows: List[Dict[str, Any]] = []
    player_tx_week: Dict[Tuple[str, int, int], int] = defaultdict(int)
    player_drop_week: Dict[Tuple[str, int, int], int] = defaultdict(int)
    player_tx_year: Dict[Tuple[str, int], int] = defaultdict(int)
    player_drop_year: Dict[Tuple[str, int], int] = defaultdict(int)
    player_tx_all: Dict[str, int] = defaultdict(int)
    player_drop_all: Dict[str, int] = defaultdict(int)

    def _player_display_name(pid: Any) -> str:
        pid_str = str(pid)
        return pid_meta.get(pid_str, {}).get("full_name") or pid_str

    # Internal ledger helpers
    # key: (season, week, roster_id) -> opponent roster_id + opponent points
    opp_rid_map: Dict[Tuple[int, int, int], Optional[int]] = {}
    opp_pf_map: Dict[Tuple[int, int, int], Optional[float]] = {}
    stage_label_map: Dict[Tuple[int, int, int], Optional[str]] = {}
    playoff_start_by_season: Dict[int, Optional[int]] = {}
    roster_ids_by_season: Dict[int, List[int]] = {}
    roster_to_team_by_season: Dict[int, Dict[int, str]] = {}
    draft_rounds_by_season: Dict[int, int] = {}
    included_draft_rounds_by_season: Dict[int, int] = {}

    # Phase 10 — synthetic draft-day picks from commissioner-forced adds.
    # draft_dates_by_season: actual draft day(s) per season (for matching adds).
    # draft_day_commish_adds: per season, the commissioner-type adds landing on a
    #   draft day, sorted by time -> [(created_ms, roster_id, player_id, tx_id)].
    #   From 2024 the FIRST add is the 2.09 (toilet-bracket reward); from 2025 the
    #   rest are 5.0X picks (20-FAAB draft-day buys). toilet_winner_by_season holds
    #   each season's losers-bracket champion (p=1 winner) = next year's 2.09 owner.
    draft_dates_by_season: Dict[int, Set] = {}
    draft_day_commish_adds: Dict[int, List[Tuple[int, int, str, str]]] = {}
    toilet_winner_by_season: Dict[int, Optional[int]] = {}

    # Draft pick ownership ledger
    # key: (season, round, original_owner_id) -> current_owner_id
    pick_current_owner: Dict[Tuple[int, int, int], int] = {}
    pick_trade_events: Dict[Tuple[int, int, int], List[Tuple[Optional[datetime], int, int, Optional[int]]]] = {}
    pick_holdings: Dict[Tuple[int, int, int], List[int]] = defaultdict(list)

    # Commissioner-moved picks: pick keyed (year, round, original_owner)
    # whose latest traded_picks owner differs from original but no
    # trade transaction explains the move. We synthesize 'always
    # owned by current owner from the earliest data point we have',
    # which is the right behavior for ownership queries (commissioner
    # picks were typically moved during pre-Sleeper offseason).
    commissioner_pick_moves: Dict[Tuple[int, int, int], int] = {}

    def _detect_commissioner_moves(season_now: int) -> None:
        """Find picks whose current ownership in traded_picks_by_season
        isn't fully explained by recorded trade events. These are
        commissioner-driven moves (Sleeper doesn't track them as
        trades when they happen more than ~3 years before the
        pick's draft year).

        For each detected pick, set commissioner_pick_moves[key] =
        current owner. The helpers above treat the pick as 'always
        belonged to current owner' for ownership queries — matches
        the user's guidance to assume a single move from original
        to current.
        """
        snapshot = traded_picks_by_season.get(int(season_now), []) or []
        snapshot_owner: Dict[Tuple[int, int, int], int] = {}
        for ev in snapshot:
            if not isinstance(ev, dict):
                continue
            ps = _to_int(ev.get("season"), None)
            rnd_e = _to_int(ev.get("round"), None)
            orig = _to_int(ev.get("roster_id"), None)
            nw = _to_int(ev.get("owner_id"), None)
            if ps is None or rnd_e is None or orig is None or nw is None:
                continue
            if ps < int(season_now):
                continue
            # Sleeper returns events in chronological order; last wins.
            snapshot_owner[(int(ps), int(rnd_e), int(orig))] = int(nw)
        for key, snap_owner in snapshot_owner.items():
            events = pick_trade_events.get(key) or []
            # A pick is commissioner-moved only if its snapshot owner is NOT
            # reachable through any RECORDED trade — i.e. snap_owner is neither
            # the original owner nor any team a trade event handed it to. This
            # is membership (not chain-END equality) on purpose: a pick traded
            # again in a LATER season has a chain that continues past this
            # season's snapshot, yet the snapshot owner is still a real hop in
            # that chain, so it must NOT be flagged. Only a snapshot owner the
            # ledger never explains (off-platform / commissioner reassignment)
            # counts. NB: requires pick_trade_events to be fully populated, so
            # the authoritative pass runs AFTER the season loop.
            chain_owners = {int(key[2])}
            for ev in events:
                try:
                    chain_owners.add(int(ev[2]))
                except Exception:
                    pass
            if int(snap_owner) not in chain_owners:
                commissioner_pick_moves[key] = int(snap_owner)

    def _future_picks_owned(team_name: str, season_now: int, at_date: date) -> List[Tuple[int, int]]:
        """(pick_year, round) for every FUTURE pick (next 3 seasons) owned
        by this team at `at_date`. Single ownership source shared by
        'Team age including picks' and 'Future draft capital'.

        Walks pick_trade_events per (pick_year, round, original_owner)
        to find the owner at the query date. A pick with no trade events
        defaults to its original owner — so a team's OWN retained picks
        are counted (this is exactly what _future_cap_from_traded missed,
        since the Sleeper traded_picks snapshot only lists picks that have
        actually moved). Commissioner-moved picks layer on top: if a pick
        has no trade events but its latest traded_picks snapshot shows a
        different owner, treat it as 'always with the current owner'.
        """
        # season_team_to_roster is keyed by _norm_team_name(...) (lowercased,
        # space-stripped), but callers pass the display handle (e.g.
        # "AceMatthew", "BROsenzweig"). Without normalizing here, every team
        # whose display name isn't already lowercase resolved to rid=None and
        # silently contributed ZERO picks to "Team age including picks" — only
        # the all-lowercase handles (plehv79/shmuel256/stevenb123) worked.
        rid = season_team_to_roster.get(int(season_now), {}).get(_norm_team_name(team_name))
        if rid is None:
            return []
        rosters_in_season = season_roster_to_team.get(int(season_now), {}) or {}
        rounds_in_draft = draft_rounds_by_season.get(int(season_now), 4) or 4
        owned: List[Tuple[int, int]] = []
        for offset in range(1, 4):  # 3-year horizon
            ps = int(season_now) + offset
            for rnd_e in range(1, int(rounds_in_draft) + 1):
                for orig_rid in rosters_in_season.keys():
                    key = (int(ps), int(rnd_e), int(orig_rid))
                    cur_owner = int(orig_rid)
                    events = pick_trade_events.get(key) or []
                    # Walk chronologically, applying events <= at_date.
                    for ev in events:
                        ev_dt = ev[0]
                        if ev_dt is None:
                            continue
                        try:
                            ev_date = ev_dt.date() if hasattr(ev_dt, "date") else ev_dt
                        except Exception:
                            continue
                        if ev_date <= at_date:
                            cur_owner = int(ev[2])  # new_owner
                    # Commissioner override
                    if key in commissioner_pick_moves and not events:
                        cur_owner = commissioner_pick_moves[key]
                    if cur_owner == int(rid):
                        owned.append((int(ps), int(rnd_e)))
        return owned

    def _picks_held_by_team_at(team_name: str, season_now: int, at_date: date) -> List[float]:
        """Pick ages (via _pick_expected_age) for all FUTURE picks owned by
        this team at `at_date`. Thin wrapper over _future_picks_owned."""
        ages: List[float] = []
        for (ps, _rnd) in _future_picks_owned(team_name, season_now, at_date):
            pa = _pick_expected_age(ps, at_date)
            if pa is not None:
                ages.append(pa)
        return ages

    # Round weights for future draft capital (provided by user). Picks in
    # rounds beyond this map (e.g. a 5th rookie round) contribute 0.
    _FUTURE_PICK_WEIGHTS = {1: 0.25, 2: 0.09, 3: 0.03, 4: 0.01}

    def _future_cap_held(team_name: str, season_now: int, at_date: date) -> float:
        """Weighted future draft capital actually held by the team at
        `at_date` — own retained picks + acquired − traded away. Returns
        0.0 only when the team holds no picks across the next 3 seasons,
        per spec. Replaces _future_cap_from_traded, which undercounted by
        omitting un-traded own picks."""
        return round(
            sum(_FUTURE_PICK_WEIGHTS.get(rnd, 0.0)
                for (_ps, rnd) in _future_picks_owned(team_name, season_now, at_date)),
            4,
        )

    def _ensure_pick_bases(target_season: int, source_season: int) -> None:
        if target_season < 2021:
            return
        rounds = draft_rounds_by_season.get(target_season) or draft_rounds_by_season.get(source_season)
        roster_ids = roster_ids_by_season.get(target_season) or roster_ids_by_season.get(source_season) or []
        if not rounds or not roster_ids:
            return
        draft_rounds_by_season.setdefault(target_season, int(rounds))
        roster_ids_by_season.setdefault(target_season, list(roster_ids))
        for rnd in range(1, int(rounds) + 1):
            for rid in roster_ids:
                key = (int(target_season), int(rnd), int(rid))
                if key in pick_current_owner:
                    continue
                pick_current_owner[key] = int(rid)
                pick_trade_events[key] = []
                pick_holdings[(int(target_season), int(rnd), int(rid))].append(int(rid))

    def _format_pick_number_for_season(season: int, round_num: Optional[int], pick_no: Optional[int]) -> Optional[str]:
        if round_num is None:
            return None
        team_count = len(roster_ids_by_season.get(season, [])) or None
        if pick_no is None or team_count is None:
            return f"{int(round_num)}.??"
        slot = ((int(pick_no) - 1) % int(team_count)) + 1
        return f"{int(round_num)}.{slot:02d}"

    def _format_pick_label(season: int, round_num: Optional[int], pick_no: Optional[int]) -> Optional[str]:
        num = _format_pick_number_for_season(season, round_num, pick_no)
        if num is None:
            return None
        return f"{int(season)} {num}"

    def _slot_for_roster(season: int, roster_id: int) -> Optional[int]:
        roster_ids = sorted(roster_ids_by_season.get(season, []))
        if not roster_ids:
            return None
        try:
            return roster_ids.index(int(roster_id)) + 1
        except ValueError:
            return None

    def _select_pick_key(
        season: int,
        round_num: int,
        prev_owner: int,
        original_owner: Optional[int],
    ) -> Optional[Tuple[int, int, int]]:
        if original_owner is not None:
            key = (int(season), int(round_num), int(original_owner))
            if key in pick_current_owner:
                return key
        candidates = pick_holdings.get((int(season), int(round_num), int(prev_owner)), [])
        if len(candidates) == 1:
            return (int(season), int(round_num), int(candidates[0]))
        if candidates:
            return (int(season), int(round_num), int(sorted(candidates)[0]))
        return None

    # Determine last completed week per league (robust, per Apps Script)
    def last_completed_week(league_id: str, season: int, max_weeks: int = 30) -> int:
        """Last week with any non-zero team points, excluding the fantasy-championship week.

        Sleeper leagues often expose an extra 'final' week that we intentionally exclude from all tables:
        - week 18 for seasons 2021+
        - week 17 for seasons <= 2020 (kept for future ESPN backfill)
        """
        excluded = 18 if int(season) >= 2021 else 17
        last = 0
        for wk in range(1, max_weeks + 1):
            if wk == excluded:
                continue
            try:
                mu = sc.matchups(wk, league_id)
            except Exception:
                mu = None
            if not mu:
                continue
            has_real = any((_to_float(m.get("points"), 0.0) or 0.0) > 0.0 for m in mu)
            # PR E fix A: don't finalize a week whose games aren't over yet.
            # Sleeper points go live DURING games, so a week can have points
            # while still in progress. Gate on a purely time-based cutoff
            # (Tuesday after the week's MNF) so a mid-season build drops the
            # in-progress week instead of treating partial scores as final.
            # Historical seasons are always past the cutoff -> no change.
            if has_real and _week_is_complete(season, wk):
                last = wk
        return last


    # Phase 2 audit followup: NFLverse pre-LOTG backfill for hardship.
    # Players already in the NFL before our LOTG history begins (e.g.
    # Deshaun Watson, suspended for the entirety of 2021-2022, so pw
    # has zero active weeks for him) need their pre-LOTG NFL game log
    # to seed the hardship baseline. Without this, Hardship reports 0
    # contribution for such players even when they really did "lose"
    # an expected score's worth of points to suspension/injury.
    #
    # Dynamic window: backfill the 2 NFL seasons immediately before
    # the earliest LOTG league season. Currently leagues start at 2021
    # so we pull 2019 + 2020. When 2020 ESPN data is added later
    # (Phase 13), the league window shifts to 2020 and the backfill
    # auto-adjusts to 2018 + 2019. Two seasons is enough to cover the
    # 5-active-game baseline window for any player.
    _lotg_seasons_for_backfill = [
        s for s in (
            _to_int(_lg.get("season"), None) for _lg in leagues
        ) if s is not None
    ]
    _earliest_lotg = min(_lotg_seasons_for_backfill) if _lotg_seasons_for_backfill else None
    # The current (most-recent) LOTG season is the one whose NFLverse stats
    # are still updating week-to-week — force a re-download for it so the
    # CI cache restore (which would otherwise return last week's file)
    # doesn't pin in-progress data. Historical seasons are immutable, so
    # they read straight from cache.
    _current_lotg_season = max(_lotg_seasons_for_backfill) if _lotg_seasons_for_backfill else None
    _nflverse_backfill_yrs = (
        list(range(_earliest_lotg - 2, _earliest_lotg))
        if _earliest_lotg is not None else []
    )
    for _bk_yr in _nflverse_backfill_yrs:
        try:
            _bk_spw = _safe_df(load_nflverse_stats_player_week(ext, _bk_yr))
            if _bk_spw.empty or "player_id" not in _bk_spw.columns or "week" not in _bk_spw.columns:
                continue
            _bk_spw["week"] = pd.to_numeric(_bk_spw["week"], errors="coerce").astype("Int64")
            _bk_spw["player_id"] = _bk_spw["player_id"].astype(str)
            _gsis_to_sid = {
                str((meta or {}).get("gsis_id") or "").strip(): str(sid)
                for sid, meta in pid_meta.items()
                if (meta or {}).get("gsis_id")
            }
            _gsis_to_sid.pop("", None)
            _pts_col = "fantasy_points_ppr" if "fantasy_points_ppr" in _bk_spw.columns else (
                "fantasy_points" if "fantasy_points" in _bk_spw.columns else None
            )
            if not _gsis_to_sid or not _pts_col:
                continue
            for r in _bk_spw[["player_id", "week", _pts_col]].dropna(subset=["player_id", "week"]).itertuples(index=False):
                _gsis = str(r.player_id)
                _sid_bk = _gsis_to_sid.get(_gsis)
                if not _sid_bk:
                    continue
                try:
                    _wk_bk = int(r.week)
                    _pts_bk = float(getattr(r, _pts_col))
                except Exception:
                    continue
                try:
                    _wk_d = date(int(_bk_yr), 9, 7) + timedelta(days=7 * (_wk_bk - 1))
                    _wk_iso = _wk_d.isoformat()
                except Exception:
                    _wk_iso = ""
                nfl_log_by_sid[_sid_bk].append({
                    "year": int(_bk_yr),
                    "week": _wk_bk,
                    "points": _pts_bk,
                    "_wk_date": _wk_iso,
                })
        except Exception as e:
            _log_exc(debug, f"nflverse_backfill_{_bk_yr}", e)

    # ------------- Build each season -------------
    traded_picks_by_season: Dict[int, List[Dict[str, Any]]] = {}
    season_roster_to_team: Dict[int, Dict[int, str]] = {}
    season_team_to_roster: Dict[int, Dict[str, int]] = {}
    season_draft_picks_all: Dict[int, List[Dict[str, Any]]] = {}
    draft_picks_records: List[Dict[str, Any]] = []

    # Cross-season state for "from previous week" stats. Hoisted out of the
    # per-season loop so Week 1 turnover & PF-increase reference the last
    # played week of the prior season (≈ championship week) instead of None.
    prev_starters_by_team_xseason: Dict[str, set] = {}
    prev_roster_by_team_xseason: Dict[str, set] = {}

    for lg in leagues:
        league_id = str(lg.get("league_id"))
        season = _to_int(lg.get("season"), 0) or 0

        # playoff start week (Sleeper setting)
        settings = lg.get("settings") or {}
        scoring_settings = lg.get("scoring_settings") or {}  # Phase 12 #5: league scoring
        # Auto-detect a scoring-settings change vs the prior season (the
        # re-score already uses each year's own table; this just flags it).
        try:
            _relevant = set(_LEAGUE_SCORE_MAP) | {"bonus_rec_te"} | {b[0] for b in _LEAGUE_SCORE_BONUS}
            _cur_sc = {k: round(float(scoring_settings[k]), 4) for k in _relevant if k in scoring_settings}
            _prev_sc = locals().get("_prev_scoring_sig")
            if _prev_sc is not None and _prev_sc != _cur_sc:
                _diff = {k: (_prev_sc.get(k), _cur_sc.get(k)) for k in set(_prev_sc) | set(_cur_sc) if _prev_sc.get(k) != _cur_sc.get(k)}
                _log(debug, f"INFO league scoring settings changed in {season}: {_diff}")
            _prev_scoring_sig = _cur_sc
        except Exception:
            pass
        playoff_start = _to_int(settings.get("playoff_week_start"), None)
        playoff_start_by_season[season] = playoff_start

        # cache played_by_week
        if season not in played_by_week_by_season:
            played_by_week_by_season[season] = _played_teams_by_week(games, season) if not games.empty else {}
        played_by_week = played_by_week_by_season.get(season, {})

        # nflverse weekly player stats (team-by-week + played detection)
        player_team_by_week: Dict[Tuple[str, int], str] = {}
        player_pos_by_week: Dict[Tuple[str, int], str] = {}
        # Season-level fallback team for each player. Used when a specific week
        # has no stats row (player injured / inactive / on bye / suspended) —
        # the right answer is "the team they were on that season," not "the
        # team they're on today." Fixes A.J. Brown 2021 wks 4/12-15 (was PHI,
        # should be TEN), Cooper Kupp 2022 wk11+ (was SEA, should be LAR), etc.
        player_season_team: Dict[str, str] = {}
        # nflverse WEEKLY rosters: (gsis, season, week) -> team, plus a per-
        # (gsis, season) fallback. These list players who were ON a roster but
        # never accumulated stats (IR / suspended / PUP) — so a player like
        # Calvin Ridley 2022 (suspended all year, on JAX) gets his real team
        # instead of the 'NFL' free-agent sentinel.
        roster_team_by_week: Dict[Tuple[str, int, int], str] = {}
        roster_team_by_season: Dict[Tuple[str, int], str] = {}
        played_players_by_week: Dict[int, set] = {}
        try:
            spw = _safe_df(load_nflverse_stats_player_week(
                ext, season,
                force_refresh=(season == _current_lotg_season),
            ))
            if (not spw.empty) and ("player_id" in spw.columns) and ("week" in spw.columns):
                spw["week"] = pd.to_numeric(spw["week"], errors="coerce").astype("Int64")
                spw["player_id"] = spw["player_id"].astype(str)
                # nflverse has used 'recent_team' historically and 'team' in current releases.
                # Accept either; older code only checked 'recent_team' and silently produced
                # an empty mapping for every season, which is why NFL team / Position columns
                # were stuck on the live Sleeper snapshot for all historical rows.
                team_col = "recent_team" if "recent_team" in spw.columns else ("team" if "team" in spw.columns else None)
                if team_col:
                    spw[team_col] = spw[team_col].astype(str)
                    for r in spw[["player_id", "week", team_col]].dropna().itertuples(index=False):
                        player_team_by_week[(str(r.player_id), int(r.week))] = str(getattr(r, team_col))
                    try:
                        modes = (
                            spw[["player_id", team_col]].dropna()
                            .groupby("player_id")[team_col]
                            .agg(lambda s: s.mode().iat[0] if not s.mode().empty else None)
                        )
                        for pid_x, tm_x in modes.items():
                            if tm_x and str(tm_x) != "nan":
                                player_season_team[str(pid_x)] = str(tm_x)
                    except Exception as e:
                        _log_exc(debug, f"player_season_team_{season}", e)
                if "position" in spw.columns:
                    spw["position"] = spw["position"].astype(str)
                    for r in spw[["player_id", "week", "position"]].dropna().itertuples(index=False):
                        player_pos_by_week[(str(r.player_id), int(r.week))] = str(r.position).upper()
                for wk_i, g in spw.groupby("week"):
                    if pd.isna(wk_i):
                        continue
                    played_players_by_week[int(wk_i)] = set(g["player_id"].dropna().astype(str).tolist())

                # Build cross-season per-player game log so the
                # transactions polish pass can compute pre-pickup PPG
                # for players who weren't yet rostered (rookies, UFAs).
                # nflverse uses GSIS player_ids; bridge to Sleeper IDs
                # via pid_meta.
                try:
                    gsis_to_sid = {
                        str((meta or {}).get("gsis_id") or "").strip(): str(sid)
                        for sid, meta in pid_meta.items()
                        if (meta or {}).get("gsis_id")
                    }
                    gsis_to_sid.pop("", None)
                    if gsis_to_sid:
                        # Phase 12 #5: score each week with the LEAGUE's scoring
                        # settings from the raw nflverse stats (same scale as the
                        # rostered "Points"); fall back to nflverse's PPR total
                        # only if scoring_settings is unavailable.
                        _score_cols = [c for cols in _LEAGUE_SCORE_MAP.values()
                                       for c in cols if c in spw.columns]
                        _pos_col = "position" if "position" in spw.columns else None
                        _ppr_col = "fantasy_points_ppr" if "fantasy_points_ppr" in spw.columns else (
                            "fantasy_points" if "fantasy_points" in spw.columns else None
                        )
                        _read_cols = list(dict.fromkeys(
                            [c for c in (["player_id", "week", _pos_col, _ppr_col] + _score_cols) if c]
                        ))
                        _use_league = bool(scoring_settings) and bool(_score_cols)
                        if _use_league or _ppr_col:
                            for r in spw[_read_cols].dropna(subset=["player_id", "week"]).itertuples(index=False):
                                gsis = str(getattr(r, "player_id"))
                                sid = gsis_to_sid.get(gsis)
                                if not sid:
                                    continue
                                try:
                                    wk = int(getattr(r, "week"))
                                except Exception:
                                    continue
                                if _use_league:
                                    _stats = {c: getattr(r, c, None) for c in _score_cols}
                                    _pos = getattr(r, _pos_col, None) if _pos_col else (pid_meta.get(sid, {}) or {}).get("position")
                                    pts = _league_score(_stats, scoring_settings, _pos)
                                else:
                                    try:
                                        pts = float(getattr(r, _ppr_col))
                                    except Exception:
                                        continue
                                # Approx Thursday-of-week date for sorting.
                                try:
                                    wk_d = date(int(season), 9, 7) + timedelta(days=7 * (wk - 1))
                                    wk_iso = wk_d.isoformat()
                                except Exception:
                                    wk_iso = ""
                                nfl_log_by_sid[sid].append({
                                    "year": int(season),
                                    "week": wk,
                                    "points": pts,
                                    "_wk_date": wk_iso,
                                })
                except Exception as e:
                    _log_exc(debug, f"nfl_log_by_sid_{season}", e)
        except Exception as e:
            _log_exc(debug, f"load_nflverse_stats_player_week_{season}", e)

        # nflverse weekly rosters — players on a team that week even with no
        # stats (IR / suspended / PUP). Lets unrostered-but-suspended players
        # keep their real team instead of the 'NFL' sentinel.
        try:
            wr = _safe_df(load_nflverse_weekly_rosters(
                ext, season, force_refresh=(season == _current_lotg_season),
            ))
            if not wr.empty and "gsis_id" in wr.columns and "team" in wr.columns and "week" in wr.columns:
                wr = wr[["gsis_id", "week", "team"]].dropna()
                wr["week"] = pd.to_numeric(wr["week"], errors="coerce").astype("Int64")
                for _r in wr.dropna(subset=["week"]).itertuples(index=False):
                    _gs = str(_r.gsis_id); _tm = str(_r.team)
                    if _gs and _gs != "nan" and _tm and _tm != "nan":
                        roster_team_by_week[(_gs, int(season), int(_r.week))] = _tm
                try:
                    _modes = wr[["gsis_id", "team"]].dropna().groupby("gsis_id")["team"].agg(
                        lambda s: s.mode().iat[0] if not s.mode().empty else None)
                    for _gs, _tm in _modes.items():
                        if _tm and str(_tm) != "nan":
                            roster_team_by_season[(str(_gs), int(season))] = str(_tm)
                except Exception:
                    pass
        except Exception as e:
            _log_exc(debug, f"load_nflverse_weekly_rosters_{season}", e)

        # nflverse injuries (optional; used as secondary signal)
        try:
            injuries = _safe_df(load_nflverse_injuries(
                ext, season,
                force_refresh=(season == _current_lotg_season),
            ))
        except Exception as e:
            injuries = pd.DataFrame()
            _log_exc(debug, f"load_nflverse_injuries_{season}", e)

            # placeholder to anchor following logic (keep in scope)
        injuries_by_gsis_week: Dict[Tuple[str, int], Tuple[Optional[bool], Optional[bool]]] = {}
        if not injuries.empty and "gsis_id" in injuries.columns:
            try:
                inj_df = injuries.copy()
                inj_df["season"] = pd.to_numeric(inj_df.get("season"), errors="coerce").astype("Int64")
                inj_df["week"] = pd.to_numeric(inj_df.get("week"), errors="coerce").astype("Int64")
                inj_df["gsis_id"] = inj_df["gsis_id"].astype(str)
                status_col = _first_col(
                    inj_df,
                    ["report_status", "status", "game_status", "injury_status", "practice_status"],
                )
                if status_col:
                    for (gsis, yr, wk), grp in inj_df.groupby(["gsis_id", "season", "week"]):
                        if pd.isna(wk) or pd.isna(yr):
                            continue
                        s = str(grp.iloc[0].get(status_col) or "").lower()
                        suspension = ("susp" in s) or ("sspd" in s)
                        injury = (("out" in s) or ("ir" in s) or ("inactive" in s) or ("pup" in s)) and not suspension
                        injuries_by_gsis_week[(gsis, int(yr), int(wk))] = (injury, suspension)
            except Exception as e:
                _log_exc(debug, f"injury_index_{season}", e)

        # Overlay curated suspensions (data/suspensions.csv). nflverse's injury
        # feed is the NFL's official game-status report and does NOT list
        # suspended players (they're simply absent from the roster). Without
        # this overlay, Suspension? would always be False for every player and
        # year. Each row expands to (gsis_id, season, week) keys for the
        # inclusive range [week_start, week_end].
        try:
            susp_path = repo_root / "data" / "suspensions.csv"
            if susp_path.exists():
                susp_df = pd.read_csv(susp_path)
                for _, srow in susp_df.iterrows():
                    if int(srow.get("season", 0)) != int(season):
                        continue
                    g = str(srow.get("gsis_id") or "").strip()
                    if not g:
                        continue
                    try:
                        wks = int(srow.get("week_start", 0))
                        wke = int(srow.get("week_end", 0))
                    except Exception:
                        continue
                    if wks <= 0 or wke <= 0 or wke < wks:
                        continue
                    for wk_n in range(wks, wke + 1):
                        injuries_by_gsis_week[(g, int(season), int(wk_n))] = (False, True)
        except Exception as e:
            _log_exc(debug, f"suspensions_overlay_{season}", e)

        # Overlay curated injuries (data/injuries.csv). The nflverse injury
        # report only covers weeks a player is on the active roster with a
        # game-status question — players on season-ending IR usually drop off
        # the report entirely, so the file doesn't say "Out" for those weeks.
        # This overlay marks season-ending and multi-week IR stints that the
        # nflverse feed misses. Only sets Injury? if the key doesn't already
        # carry a suspension (suspension always wins).
        try:
            inj_path = repo_root / "data" / "injuries.csv"
            if inj_path.exists():
                inj_overlay_df = pd.read_csv(inj_path)
                for _, irow in inj_overlay_df.iterrows():
                    if int(irow.get("season", 0)) != int(season):
                        continue
                    g = str(irow.get("gsis_id") or "").strip()
                    if not g:
                        continue
                    try:
                        wks = int(irow.get("week_start", 0))
                        wke = int(irow.get("week_end", 0))
                    except Exception:
                        continue
                    if wks <= 0 or wke <= 0 or wke < wks:
                        continue
                    for wk_n in range(wks, wke + 1):
                        existing = injuries_by_gsis_week.get((g, int(season), int(wk_n)))
                        # Don't overwrite a confirmed suspension
                        if existing is not None and existing[1] is True:
                            continue
                        injuries_by_gsis_week[(g, int(season), int(wk_n))] = (True, False)
        except Exception as e:
            _log_exc(debug, f"injuries_overlay_{season}", e)

        # Gap-fill heuristic: for every player who has at least one nflverse
        # weekly stats row this season AND every week between the season's first
        # and last played week, mark Injury?=True for the missing weeks.
        # Catches mid-season IR stints (Aaron Jones 2025 wks 3-5, Adam Thielen
        # 2024 wks 4-9) AND end-of-season IRs (Travis Hunter 2025 wks 10-17,
        # Michael Penix Jr 2025 wks 12-17) in one pass.
        #
        # Conservative: only fires for players who played at least one nflverse
        # game in the season (so we never invent injuries for never-active
        # backups), and never overwrites an existing key (so the curated
        # suspensions / injuries overlays and nflverse Out reports still win).
        played_by_gsis_season: Dict[str, set] = defaultdict(set)
        try:
            if played_players_by_week:
                season_max_week = 0
                for wk_i, played_set in played_players_by_week.items():
                    try:
                        wk_int = int(wk_i)
                    except Exception:
                        continue
                    if wk_int > season_max_week:
                        season_max_week = wk_int
                    for pid_s in played_set:
                        pid_clean = str(pid_s).strip()
                        if pid_clean:
                            played_by_gsis_season[pid_clean].add(wk_int)
                for gsis_s, played_weeks in played_by_gsis_season.items():
                    if not played_weeks:
                        continue
                    for wk in range(1, int(season_max_week) + 1):
                        if wk in played_weeks:
                            continue
                        key = (gsis_s, int(season), int(wk))
                        if key in injuries_by_gsis_week:
                            continue
                        injuries_by_gsis_week[key] = (True, False)
        except Exception as e:
            _log_exc(debug, f"injury_gap_fill_{season}", e)

        # users/rosters
        try:
            users = sc.users(league_id)
        except Exception as e:
            users = []
            _log_exc(debug, f"users_{season}", e)
        try:
            rosters = sc.rosters(league_id)
        except Exception as e:
            rosters = []
            _log_exc(debug, f"rosters_{season}", e)

        user_handle = _team_handle_map(users)

        roster_owner: Dict[int, str] = {}
        roster_to_team: Dict[int, str] = {}
        for r in rosters or []:
            rid = _to_int(r.get("roster_id"), None)
            if rid is None:
                continue
            roster_owner[rid] = str(r.get("owner_id") or "")
            raw_name = user_handle.get(roster_owner[rid], f"Roster {rid}")
            canon = _norm_team_name(raw_name)
            # Preserve a stable display name for this canonical team key.
            if 'team_display' not in locals():
                team_display = {}
            if canon not in team_display:
                team_display[canon] = str(raw_name)
            roster_to_team[rid] = team_display[canon]

        season_roster_to_team[season] = dict(roster_to_team)
        season_team_to_roster[season] = {
            _norm_team_name(v): k for k, v in roster_to_team.items() if v is not None
        }
        # Persist the season's roster ids for downstream pick-history /
        # draft-frame logic. (Prior code only populated this via
        # _ensure_pick_bases.setdefault, which had a chicken-and-egg gate
        # — it requires roster_ids to already exist for some season to
        # propagate them. Without a direct seed here, the dict could stay
        # empty for the entire run, which broke future-year pick
        # synthesis.)
        roster_ids_by_season.setdefault(season, sorted(roster_to_team.keys()))

        # Gap-fill (part 2): full-season absences. A player on a fantasy roster
        # who NEVER appeared in nflverse weekly stats this season is almost
        # always a season-ending injury sustained before he had a chance to
        # play (Gus Edwards 2021 ACL, J.J. McCarthy 2024 meniscus, Cam Akers
        # 2021 Achilles, Jordan Travis 2024 broken leg, etc.). Mark every week
        # as Injury?=True so Hardship / Weeks of injuries reflect reality.
        try:
            if season_max_week_for_fill := (max(played_players_by_week.keys()) if played_players_by_week else 0):
                rostered_gsis: set = set()
                for r in rosters or []:
                    for pid_r in (r.get("players") or []):
                        gsis_r = (pid_meta.get(str(pid_r), {}) or {}).get("gsis_id")
                        if not gsis_r:
                            # Try DP and legacy mappings for completeness.
                            gsis_r = (
                                dp_sleeper_to_gsis.get(str(pid_r))
                                or sleeper_to_gsis.get(str(pid_r))
                            )
                        if gsis_r:
                            rostered_gsis.add(str(gsis_r).strip())
                for gsis_r in rostered_gsis:
                    if not gsis_r:
                        continue
                    # If this player has any nflverse appearance, the earlier
                    # gap-fill (part 1) already filled their gaps. Skip.
                    if played_by_gsis_season.get(gsis_r):
                        continue
                    for wk in range(1, int(season_max_week_for_fill) + 1):
                        key = (gsis_r, int(season), int(wk))
                        if key in injuries_by_gsis_week:
                            continue
                        injuries_by_gsis_week[key] = (True, False)
        except Exception as e:
            _log_exc(debug, f"injury_gap_fill_never_played_{season}", e)

        # traded picks snapshot (used for pick history reconstruction)
        try:
            traded_picks = sc.traded_picks(league_id) or []
            traded_picks_by_season[season] = traded_picks
        except Exception as e:
            traded_picks = []
            traded_picks_by_season[season] = []
            _log_exc(debug, f"traded_picks_{season}", e)

        # Detect commissioner-moved picks for this season's snapshot.
        # Has to run AFTER the snapshot is captured but can use prior
        # seasons' pick_trade_events (mid-season trades in this
        # season are captured later inside the per-week sub-loop and
        # will be merged into the same map). The team_week emission
        # downstream will see the up-to-date commissioner moves.
        try:
            _detect_commissioner_moves(int(season))
        except Exception as e:
            _log_exc(debug, f"detect_commissioner_moves_{season}", e)

        # Per-pick movement tracking happens via pick_trade_events
        # (populated as each trade transaction is processed later in
        # this season loop). Per-week pick ownership for the
        # 'Team age including picks' column gets computed by querying
        # those event lists at week-level snapshots — see the
        # _picks_held_by_team_at helper defined at function scope
        # below the per-season loop.
        #
        # Commissioner-moved picks (whose ownership change isn't in
        # any trade transaction but does appear in traded_picks) are
        # synthesized as 'always belonged to current owner' below
        # and added to pick_history.

        # raw snapshots
        try:
            (raw_dir / f"league_{season}.json").write_text(json.dumps(lg, indent=2))
            (raw_dir / f"users_{season}.json").write_text(json.dumps(users, indent=2))
            (raw_dir / f"rosters_{season}.json").write_text(json.dumps(rosters, indent=2))
        except Exception:
            pass

        
        # traded picks raw snapshot (for future draft capital / tanking + debugging)
        try:
            (raw_dir / f"traded_picks_{season}.json").write_text(json.dumps(traded_picks, indent=2))
        except Exception:
            pass

        # draft picks history (rookie + startup as available in Sleeper; still partial)
        try:
            drafts = sc.drafts(league_id)
        except Exception as e:
            drafts = []
            _log_exc(debug, f"drafts_{season}", e)
        # Capture actual draft day(s) — when picks were really made — for
        # matching commissioner-forced adds (the 2.09 / 5.0X synthetic picks).
        try:
            _dd: Set = set()
            for _dr in drafts or []:
                for _k in ("last_picked", "start_time"):
                    _x = _epoch_ms_to_dt(_dr.get(_k))
                    if _x is not None:
                        _dd.add(_x.date())
            draft_dates_by_season[int(season)] = _dd
        except Exception as e:
            _log_exc(debug, f"draft_dates_{season}", e)
        # Losers-bracket champion (p=1 winner) = the toilet-bracket winner, who
        # earns NEXT season's 2.09 pick.
        try:
            _lb = sc.losers_bracket(league_id) or []
            _tw = None
            for _m in _lb:
                if _m.get("p") == 1:
                    _tw = _to_int(_m.get("w"), None)
                    break
            toilet_winner_by_season[int(season)] = _tw
        except Exception as e:
            _log_exc(debug, f"losers_bracket_{season}", e)
        # Hydrate each draft with its FULL object via the /draft/{id}
        # endpoint. The /league/{id}/drafts list view returns drafts
        # without `slot_to_roster_id` (it's null there) — the full
        # draft object has it. Without this, every season fell back
        # to the canonical sorted-roster slot ordering and Original
        # Team didn't reflect the real standings-based draft slot.
        for _d in drafts or []:
            _did = str(_d.get("draft_id") or "")
            if not _did:
                continue
            if _d.get("slot_to_roster_id"):
                continue
            try:
                _full = sc.draft(_did)
                if isinstance(_full, dict) and _full.get("slot_to_roster_id"):
                    _d["slot_to_roster_id"] = _full.get("slot_to_roster_id")
            except Exception as e:
                _log_exc(debug, f"draft_detail_{season}_{_did}", e)
        # Dump raw drafts for audit (slot_to_roster_id verification, etc).
        try:
            (raw_dir / f"drafts_{season}.json").write_text(json.dumps(drafts or [], indent=2))
        except Exception:
            pass
        draft_picks_all: List[Dict[str, Any]] = []
        draft_slot_to_roster_by_did: Dict[str, Dict[int, int]] = {}
        for d in drafts or []:
            did = str(d.get("draft_id") or "")
            if not did:
                continue
            slot_map_raw = d.get("slot_to_roster_id") if isinstance(d, dict) else {}
            slot_map: Dict[int, int] = {}
            if isinstance(slot_map_raw, dict):
                for k, v in slot_map_raw.items():
                    kk = _to_int(k, None)
                    vv = _to_int(v, None)
                    if kk is not None and vv is not None:
                        slot_map[int(kk)] = int(vv)
            if slot_map:
                draft_slot_to_roster_by_did[did] = slot_map
            try:
                picks = sc.draft_picks(did)
            except Exception as e:
                picks = []
                _log_exc(debug, f"draft_picks_{season}_{did}", e)
            max_round = 0
            picks_with_players = 0
            picks_with_names = 0
            rookie_picks = 0
            for p in picks or []:
                rnd = _to_int(p.get("round"), None)
                if rnd is not None:
                    max_round = max(max_round, int(rnd))
                pid = p.get("player_id")
                if _valid_pid(pid):
                    picks_with_players += 1
                    if is_rookie_pid(pid, season):
                        rookie_picks += 1
                md = p.get("metadata") if isinstance(p.get("metadata"), dict) else {}
                fname = str(md.get("first_name") or "").strip()
                lname = str(md.get("last_name") or "").strip()
                if fname or lname:
                    picks_with_names += 1

            # Prefer explicit rookie-type drafts when available.
            draft_type = str(d.get("type") or d.get("draft_type") or "").strip().lower()
            is_rookie_draft = draft_type in {"rookie", "dynasty"}

            # Cap at 5-round rookie drafts (per league rules). Startup drafts
            # are excluded — those have many more rounds.
            if max_round > 0 and max_round > 5:
                continue
            if (picks_with_players + picks_with_names) == 0 and not is_rookie_draft:
                continue
            # Note: we deliberately do NOT exclude vet/supplemental drafts.
            # The 2021 league had both a rookie draft AND a 4-round
            # supplemental veteran draft — both belong in pick_history.
            # Dedupe below uses Player Picked to keep them distinct.
            if max_round:
                included_draft_rounds_by_season[season] = max(
                    included_draft_rounds_by_season.get(season, 0),
                    int(max_round),
                )
                # Mirror into draft_rounds_by_season — the ledger helpers
                # (_ensure_pick_bases, _picks_held_by_team_at) read from
                # that dict to know how many rounds to seed per season.
                # Without this mirror it stayed empty, which broke
                # _select_pick_key for every future-pick trade processed
                # before any season ran _ensure_pick_bases successfully.
                draft_rounds_by_season[season] = max(
                    draft_rounds_by_season.get(season, 0),
                    int(max_round),
                )
            # Tag supplemental veteran drafts: any draft NOT marked rookie
            # type whose selections are <50% NFL rookies for the season
            # (the 2021 league held one). Vet picks get a "(vet)" suffix
            # in the Year column so they're visually distinct from rookie
            # picks at the same slot.
            is_vet_draft = (
                not is_rookie_draft
                and picks_with_players > 0
                and (rookie_picks / float(picks_with_players)) <= 0.5
            )
            # Snake drafts reverse pick ORDER on even rounds (the team that
            # picked last in round 1 picks first in round 2). Capture the
            # type + reversal_round so the pick-number rebuild can label
            # picks by true draft order, not by raw draft_slot. The 2021
            # rookie drafts are snake; all later rookie drafts are linear.
            _reversal_round = _to_int((d.get("settings") or {}).get("reversal_round"), 0) or 0
            for p in picks or []:
                p["draft_id"] = did
                p["draft_season"] = season
                p["_is_vet_draft"] = bool(is_vet_draft)
                p["_draft_type"] = draft_type
                p["_reversal_round"] = int(_reversal_round)
                if did in draft_slot_to_roster_by_did:
                    p["slot_to_roster_id"] = draft_slot_to_roster_by_did.get(did)
            draft_picks_all.extend(picks or [])

        season_draft_picks_all[int(season)] = list(draft_picks_all)
        draft_picks_records.extend(draft_picks_all)

        for p in draft_picks_all:
            rnd = _to_int(p.get("round"), None)
            pick_no = _to_int(p.get("pick_no"), None)
            if pick_no is None:
                pick_no = _to_int(p.get("pick_in_round"), None) or _to_int(p.get("draft_slot"), None)
            roster_id = _to_int(p.get("roster_id"), None)
            # Note: Sleeper's `picked_by` is a USER_ID (long string), not a
            # roster_id. The team that ACTUALLY drafted the player is in
            # `roster_id` (= Final Team for the pick). For "Original Team"
            # (who owned the pick before any trades), prefer the draft's
            # static slot_to_roster_id mapping — it survives ESPN-era picks
            # that have no event in traded_picks.
            player = p.get("player_id")

            slot_map = p.get("slot_to_roster_id") if isinstance(p.get("slot_to_roster_id"), dict) else {}
            team_count = len(slot_map) or len(roster_ids_by_season.get(season, [])) or 8
            slot_no = _to_int(p.get("draft_slot"), None) or _to_int(p.get("pick_in_round"), None)
            if slot_no is None and pick_no is not None and team_count:
                try:
                    slot_no = ((int(pick_no) - 1) % int(team_count)) + 1
                except Exception:
                    slot_no = None

            # Original team: slot ownership map first; fallback to picker roster.
            origin_rid: Optional[int] = None
            if slot_no is not None and slot_map:
                origin_rid = _to_int(slot_map.get(int(slot_no)), None)
            if origin_rid is None:
                origin_rid = roster_id
            origin_team = roster_to_team.get(origin_rid, f"Roster {origin_rid}") if origin_rid is not None else None

            # Final team: the roster that actually made the selection.
            final_rid = roster_id if roster_id is not None else origin_rid
            final_team = roster_to_team.get(final_rid, f"Roster {final_rid}") if final_rid is not None else None

            # Display number — canonical '{round}.{slot:02d}', '{round}' when slot unknown.
            number = _format_pick_number(rnd, slot_no)

            # Resolve player name from Sleeper player map first, then pick metadata.
            player_name = pid_meta.get(str(player), {}).get("full_name") if player else None
            if not player_name:
                md = p.get("metadata") if isinstance(p.get("metadata"), dict) else {}
                player_name = md.get("first_name") and f"{md.get('first_name')} {md.get('last_name','').strip()}".strip()
            if not player_name:
                player_name = p.get("player") or p.get("player_name")
            if not player_name:
                player_name = "Unknown"

            _year_display: Any = f"{int(season)} (vet)" if p.get("_is_vet_draft") else int(season)
            pick_rows.append({
                "Year": _year_display,
                "Number": number,
                "Original Team": origin_team,
                "Final Team": final_team,
                "Player Picked": player_name,
                "Trade 1": None, "Trade 2": None, "Trade 3": None, "Trade 4": None, "Trade 5": None,
                "Trade 6": None, "Trade 7": None, "Trade 8": None, "Trade 9": None, "Trade 10": None,                "Commissioner moved?": False,  # filled later if applicable
            })

        # NOTE: the prior traded_picks-fallback row generation was removed
        # here. It emitted one row per trade EVENT (with Original Team set
        # to the previous owner, which is often an intermediate owner in a
        # multi-trade chain — not the chain origin). That produced
        # incomplete future-year pick frames (only the traded picks, with
        # wrong origins).
        # Future-year pick frames are now built by the dedicated
        # synthesis block below, which starts from the canonical rule
        # "every roster owns one pick in each round of each upcoming
        # draft" and lets the chain-reconstruction pass fill in trades.

        # robust last completed week
        try:
            last_week = last_completed_week(league_id, season)
        except Exception as e:
            last_week = 0
            _log_exc(debug, f"last_completed_week_{season}", e)

        # Preseason / in-progress league with no completed games: still scan a
        # few weeks so offseason transactions (rookie-draft pick swaps, trades,
        # free-agent adds) flow into trades.csv / transactions.csv. The
        # team_week / player_week construction inside the loop will produce
        # empty rows for unplayed weeks; those self-filter at output time.
        preseason_only = last_week <= 0
        if preseason_only:
            _log(debug, f"[{_now_iso()}] INFO season {season}: preseason — scanning offseason transactions only")
            last_week = 1

        # Exclude week 18 always; if season < 2021 exclude week 17 (kept for future ESPN import)
        def week_allowed(wk: int) -> bool:
            if wk == 18:
                return False
            if season < 2021 and wk == 17:
                return False
            return True

        # storage for week/team scoring for expected win, etc.
        team_pf_by_week: Dict[int, Dict[str, float]] = defaultdict(dict)
        semis_bonus_by_week: Dict[int, Dict[int, float]] = defaultdict(dict)

        # ------------- Pre-fetch all weekly matchups & transactions -------------
        matchups_by_week: Dict[int, List[Dict[str, Any]]] = {}
        tx_by_week: Dict[int, List[Dict[str, Any]]] = {}
        seen_tx_ids: Set[str] = set()
        # bids_per_player_week[(season, wk, player_id)] = total waiver
        # attempts (complete + failed) targeting that player in that week.
        # Sleeper doesn't ship a 'num_bids' field; we derive it by counting
        # waiver transactions BEFORE filtering out failed claims, so the
        # winning row can carry the contested-bid count for downstream
        # display.
        bids_per_player_week: Dict[Tuple[int, int, str], int] = defaultdict(int)
        # total_bids_amount_per_player_week — sum of waiver_bid amounts
        # across all attempts (winner + losers) on the same player that
        # week. Surfaces on transactions.csv as 'Total FAAB bid' so the
        # context for a winning bid is visible: e.g. a 5-FAAB win that
        # beat a single 1-FAAB bid is much less impressive than a
        # 5-FAAB win that beat 4 other bids summing to 50.
        total_bids_amount_per_player_week: Dict[Tuple[int, int, str], float] = defaultdict(float)
        # bid_amounts_per_player_week — full sorted list of competing bid
        # amounts. Powers the 'FAAB difference over second place' and
        # 'FAAB % difference over second place' columns: how decisively
        # did the winner outbid the runner-up.
        bid_amounts_per_player_week: Dict[Tuple[int, int, str], List[float]] = defaultdict(list)

        for wk in range(1, min(last_week, 30) + 1):
            if not week_allowed(wk):
                continue
            try:
                mu = sc.matchups(wk, league_id) or []
                matchups_by_week[wk] = mu
                for m in mu:
                    rid = _to_int(m.get("roster_id"), None)
                    if rid is None:
                        continue
                    tm = roster_to_team.get(rid, f"Roster {rid}")
                    team_pf_by_week[wk][tm] = float(_to_float(m.get("points"), 0.0) or 0.0)
            except Exception as e:
                _log_exc(debug, f"matchups_{season}_wk{wk}", e)

            try:
                raw_tx = sc.transactions(wk, league_id) or []
                deduped_tx: List[Dict[str, Any]] = []
                for t in raw_tx:
                    tx_id = t.get("transaction_id")
                    if tx_id is None:
                        deduped_tx.append(t)
                        continue
                    tx_id_str = str(tx_id)
                    if tx_id_str in seen_tx_ids:
                        continue
                    seen_tx_ids.add(tx_id_str)
                    deduped_tx.append(t)

                # Sleeper-data quirk dedup. The transaction_id check above
                # catches verbatim duplicates only — Sleeper sometimes emits
                # two distinct transaction_ids for what's logically the same
                # event:
                #   (a) Duplicate waiver claim — same creator, same adds set,
                #       within 60 seconds; one has Player Dropped=None and
                #       the other carries the real drop. Keep the row with
                #       the drop, discard the bare one.
                #   (b) Inverted commissioner swap — same timestamp, two
                #       rows with adds/drops mirrored. Keep the one whose
                #       smallest-added-pid is alphabetically first.
                # Doing this here (before the per-week tx_count / faab
                # counters consume the list) keeps team_week / team_year
                # FAAB consistent with transactions.csv.
                def _tx_ts(t: Dict[str, Any]) -> int:
                    try:
                        return int(t.get("created") or 0)
                    except Exception:
                        return 0
                def _tx_adds(t: Dict[str, Any]) -> Tuple[str, ...]:
                    a = t.get("adds") or {}
                    if not isinstance(a, dict):
                        return tuple()
                    return tuple(sorted(str(k) for k in a.keys()))
                def _tx_drops(t: Dict[str, Any]) -> Tuple[str, ...]:
                    d = t.get("drops") or {}
                    if not isinstance(d, dict):
                        return tuple()
                    return tuple(sorted(str(k) for k in d.keys()))

                deduped_tx.sort(key=lambda t: (_tx_ts(t), t.get("transaction_id") or ""))
                pruned: List[Dict[str, Any]] = []
                for t in deduped_tx:
                    t_ts = _tx_ts(t)
                    t_adds = _tx_adds(t)
                    t_drops = _tx_drops(t)
                    t_creator = str(t.get("creator") or "")
                    replaced = False
                    drop_self = False
                    for idx in range(len(pruned) - 1, -1, -1):
                        p = pruned[idx]
                        p_ts = _tx_ts(p)
                        if t_ts and p_ts and abs(t_ts - p_ts) > 60_000:
                            break  # outside 60-second window
                        p_adds = _tx_adds(p)
                        p_drops = _tx_drops(p)
                        p_creator = str(p.get("creator") or "")
                        # Case (a): same creator, same adds, drops differ by
                        # one side being empty.
                        if (
                            t_creator and p_creator and t_creator == p_creator
                            and t_adds == p_adds
                        ):
                            if not p_drops and t_drops:
                                pruned[idx] = t
                                replaced = True
                                break
                            if not t_drops and p_drops:
                                drop_self = True
                                break
                        # Case (b): identical timestamp, inverted swap.
                        if t_ts and p_ts and t_ts == p_ts:
                            if t_adds == p_drops and t_drops == p_adds and t_adds and t_drops:
                                # Inverted view of one event. Keep the
                                # alphabetically-earlier adds tuple.
                                if t_adds < p_adds:
                                    pruned[idx] = t
                                drop_self = True
                                break
                    if replaced or drop_self:
                        continue
                    pruned.append(t)

                # Count waiver attempts per player BEFORE filtering out the
                # failed claims. Sleeper returns every team's waiver bid as
                # its own transaction (one status=complete winner, plus
                # status=failed for each losing bid). The losing bids never
                # actually moved the player and shouldn't pollute tx_count,
                # FAAB-spent, or transactions.csv — but we want the count
                # of contested bids on the winning row.
                for t in pruned:
                    if t.get("type") != "waiver":
                        continue
                    adds_for_count = t.get("adds") or {}
                    if not isinstance(adds_for_count, dict):
                        continue
                    bid_settings = t.get("settings") or {}
                    bid_amt = 0.0
                    if isinstance(bid_settings, dict):
                        try:
                            bid_amt = float(bid_settings.get("waiver_bid") or 0)
                        except Exception:
                            bid_amt = 0.0
                    for pid_key in adds_for_count.keys():
                        key = (int(season), int(wk), str(pid_key))
                        bids_per_player_week[key] += 1
                        total_bids_amount_per_player_week[key] += bid_amt
                        bid_amounts_per_player_week[key].append(float(bid_amt))

                # Drop failed transactions. Sleeper's status taxonomy:
                #   complete -> the move actually happened
                #   failed   -> rejected (lost waiver, roster full, etc.)
                # Anything other than 'complete' didn't move players or
                # money and should not appear in the dataset.
                pruned = [t for t in pruned if (t.get("status") or "complete") == "complete"]

                tx_by_week[wk] = pruned
            except Exception as e:
                tx_by_week[wk] = []
                _log_exc(debug, f"transactions_{season}_wk{wk}", e)

        # ------------- Manual 2021 botched-trade merge -------------
        # Sleeper split ONE draft-day pick trade (shmuel256's 2021 2.08 for
        # LWebs53's 3.06 + a 2022 4th + a 2023 4.08) into a pick swap PLUS a
        # phantom PLAYER swap (Michael Carter <-> Rhamondre Stevenson). The
        # players never changed hands — each was simply drafted with the pick the
        # other team sent (Carter by LWebs53 off the 2.08, Stevenson by shmuel256
        # off the 3.06; see _DRAFTER_FIX_2021). Merge the phantom trade's draft
        # pick(s) into the real pick trade and drop the phantom player legs so the
        # deal reads as picks-only and both rookies stay traceable from their true
        # drafters. The 2023 4.08 is kept (dropping it would orphan its later
        # hops). Matched precisely by the two players' Sleeper ids.
        if int(season) == 2021:
            try:
                _MC, _RS = "7607", "7611"  # Michael Carter, Rhamondre Stevenson
                _all_trades = [(wk, t) for wk, txs in tx_by_week.items()
                               for t in txs if t.get("type") == "trade"]
                _phantom = None
                for _wk, _t in _all_trades:
                    _pids = (set(str(k) for k in (_t.get("adds") or {}))
                             | set(str(k) for k in (_t.get("drops") or {})))
                    if {_MC, _RS} <= _pids:
                        _phantom = (_wk, _t)
                        break
                if _phantom is not None:
                    _pw, _pt = _phantom
                    _rosters = set(int(r) for r in (_pt.get("roster_ids") or []))
                    _pickswap = None
                    for _wk, _t in _all_trades:
                        if _t is _pt or not _t.get("draft_picks"):
                            continue
                        if set(int(r) for r in (_t.get("roster_ids") or [])) == _rosters:
                            _pickswap = _t
                            break
                    if _pickswap is not None:
                        _pickswap.setdefault("draft_picks", [])
                        _pickswap["draft_picks"].extend(_pt.get("draft_picks") or [])
                        for _wb in (_pt.get("waiver_budget") or []):
                            _pickswap.setdefault("waiver_budget", []).append(_wb)
                        tx_by_week[_pw] = [t for t in tx_by_week[_pw] if t is not _pt]
                        _log(debug, f"[{_now_iso()}] INFO merged the 2021 phantom "
                                    f"Carter/Stevenson player swap into the pick trade")
            except Exception as e:
                _log_exc(debug, "merge_2021_phantom_trade", e)

        # ------------- Commissioner "wash" detection (Phase 6B) -------------
        # A single-day commissioner action that ends with the roster exactly
        # as it started is a no-op correction, not a real transaction. We flag
        # any transaction whose every player movement nets to zero ON ITS OWN
        # roster that day AND where a commissioner action was involved, then
        # exclude those transactions from the counts (and from trades_rows /
        # the trade split). Covers: commish add+drop of a player, a player a
        # team dropped and the commish re-added same day, an add the commish
        # immediately undid, and a trade the commish reversed.
        wash_tx_ids: Set[str] = set()
        try:
            _pd_net: Dict[Tuple[str, str], Dict[int, int]] = defaultdict(lambda: defaultdict(int))
            _pd_commish: Dict[Tuple[str, str], bool] = defaultdict(bool)
            _tx_pdays: Dict[str, set] = {}
            for _wk_txs in tx_by_week.values():
                for _t in _wk_txs:
                    _dt = _epoch_ms_to_dt(_t.get("created"))
                    if _dt is None:
                        continue
                    _day = _dt.date().isoformat()
                    _is_comm = (str(_t.get("type") or "") == "commissioner")
                    _txid = str(_t.get("transaction_id") or id(_t))
                    _adds = _t.get("adds") if isinstance(_t.get("adds"), dict) else {}
                    _drops = _t.get("drops") if isinstance(_t.get("drops"), dict) else {}
                    _pset = set()
                    for _pid, _rid in (_adds or {}).items():
                        _ri = _to_int(_rid, None)
                        if _ri is None:
                            continue
                        _k = (str(_pid), _day); _pd_net[_k][_ri] += 1
                        if _is_comm:
                            _pd_commish[_k] = True
                        _pset.add(_k)
                    for _pid, _rid in (_drops or {}).items():
                        _ri = _to_int(_rid, None)
                        if _ri is None:
                            continue
                        _k = (str(_pid), _day); _pd_net[_k][_ri] -= 1
                        if _is_comm:
                            _pd_commish[_k] = True
                        _pset.add(_k)
                    _tx_pdays[_txid] = _pset
            _wash_pdays = {
                _k for _k, _nets in _pd_net.items()
                if _pd_commish.get(_k) and all(_v == 0 for _v in _nets.values())
            }
            for _txid, _pset in _tx_pdays.items():
                if _pset and all(_k in _wash_pdays for _k in _pset):
                    wash_tx_ids.add(_txid)
            if wash_tx_ids:
                _log(debug, f"[{_now_iso()}] INFO commissioner-wash {season}: excluded {len(wash_tx_ids)} no-op txns")
        except Exception as e:
            _log_exc(debug, f"commish_wash_{season}", e)

        # Phase 10 — capture commissioner-forced adds that land on a draft day.
        # These become synthetic picks (the 2.09 toilet reward from 2024+, and
        # 5.0X FAAB buys from 2025+) and are removed from the transactions sheet.
        # We capture for every season but only CONVERT/REMOVE the qualifying ones,
        # so the rule fires automatically in future years.
        try:
            _ddays = draft_dates_by_season.get(int(season)) or set()
            _dda: List[Tuple[int, int, str, str]] = []
            if _ddays:
                for _wk_txs in tx_by_week.values():
                    for _t in _wk_txs:
                        if str(_t.get("type") or "") != "commissioner":
                            continue
                        _a = _t.get("adds") if isinstance(_t.get("adds"), dict) else {}
                        if not _a:
                            continue
                        _dt = _epoch_ms_to_dt(_t.get("created"))
                        if _dt is None or not any(abs((_dt.date() - _d).days) <= 1 for _d in _ddays):
                            continue
                        _txid = str(_t.get("transaction_id") or id(_t))
                        for _pid, _rid in _a.items():
                            _ri = _to_int(_rid, None)
                            if _ri is not None:
                                _dda.append((int(_t.get("created") or 0), _ri, str(_pid), _txid))
                _dda.sort(key=lambda x: x[0])
            draft_day_commish_adds[int(season)] = _dda
            # Convert: 2.09 (idx 0) from 2024+, 5.0X (idx 1+) from 2025+. Remove
            # only the converted adds from the transactions output.
            _conv: List[Tuple[int, int, str, str]] = []
            if int(season) >= 2024 and _dda:
                _conv.append(_dda[0])
                if int(season) >= 2025:
                    _conv.extend(_dda[1:])
            for _ca in _conv:
                wash_tx_ids.add(_ca[3])
            if _conv:
                _log(debug, f"[{_now_iso()}] INFO draft-day synthetic picks {season}: {len(_conv)} commish-add(s) -> picks, removed from transactions")
        except Exception as e:
            _log_exc(debug, f"draft_day_commish_adds_{season}", e)

        # ------------- Build opponent roster mapping + playoff labels -------------
        for wk, mu in matchups_by_week.items():
            mdf = _safe_df(pd.DataFrame(mu))
            if mdf.empty:
                continue
            if "matchup_id" not in mdf.columns:
                continue
            # ensure numeric
            try:
                mdf["roster_id"] = pd.to_numeric(mdf["roster_id"], errors="coerce").astype("Int64")
                mdf["points"] = pd.to_numeric(mdf["points"], errors="coerce").fillna(0.0)
            except Exception:
                pass

            stage = _matchup_stage(wk, playoff_start)

            for mid, g in mdf.groupby("matchup_id"):
                rids = [int(x) for x in g["roster_id"].dropna().astype(int).tolist()]
                if len(rids) != 2:
                    continue
                a, b = rids
                pa = float(g.loc[g["roster_id"] == a, "points"].iloc[0])
                pb = float(g.loc[g["roster_id"] == b, "points"].iloc[0])

                opp_rid_map[(season, wk, a)] = b
                opp_rid_map[(season, wk, b)] = a
                opp_pf_map[(season, wk, a)] = pb
                opp_pf_map[(season, wk, b)] = pa

            # playoff/toilet naming per your rules:
            if stage:
                # Determine top4 by regular season W-L then PF tiebreaker, using weeks < playoff_start.
                try:
                    if playoff_start and wk == playoff_start:
                        reg = []
                        for rr in rosters or []:
                            rid = _to_int(rr.get("roster_id"), None)
                            if rid is None:
                                continue
                            tm = roster_to_team.get(rid, f"Roster {rid}")
                            # compute record up to reg season
                            wins = losses = ties = 0
                            pf_sum = 0.0
                            for w2 in range(1, playoff_start):
                                if not week_allowed(w2):
                                    continue
                                pf2 = team_pf_by_week.get(w2, {}).get(tm)
                                if pf2 is None:
                                    continue
                                pf_sum += float(pf2)
                                opp_pf2 = opp_pf_map.get((season, w2, rid))
                                if opp_pf2 is None:
                                    continue
                                if pf2 > opp_pf2:
                                    wins += 1
                                elif pf2 < opp_pf2:
                                    losses += 1
                                else:
                                    ties += 1
                            reg.append((tm, rid, wins, losses, ties, pf_sum))
                        reg.sort(key=lambda x: (x[2] + 0.5 * x[4], x[5]), reverse=True)
                        seed_by_rid = {rid: idx + 1 for idx, (_, rid, *_ ) in enumerate(reg)}
                        top4 = set([rid for _, rid, *_ in reg[:4]])
                        bottom4 = set([rid for _, rid, *_ in reg[4:]])
                        # apply semifinal bonus to higher seeded teams (playoff_start week only)
                        for rid in list(top4):
                            opp = opp_rid_map.get((season, playoff_start, rid))
                            if opp is None:
                                continue
                            if rid < opp and opp in top4:
                                seed_a = seed_by_rid.get(rid)
                                seed_b = seed_by_rid.get(opp)
                                if seed_a is None or seed_b is None:
                                    continue
                                higher = rid if seed_a < seed_b else opp
                                semis_bonus_by_week[playoff_start][higher] = 5.0
                                higher_team = roster_to_team.get(higher, f"Roster {higher}")
                                lower = opp if higher == rid else rid
                                lower_team = roster_to_team.get(lower, f"Roster {lower}")
                                team_pf_by_week[playoff_start][higher_team] = (
                                    team_pf_by_week[playoff_start].get(higher_team, 0.0) + 5.0
                                )
                                opp_pf_map[(season, playoff_start, lower)] = team_pf_by_week[playoff_start].get(higher_team, 0.0)
                                opp_pf_map[(season, playoff_start, higher)] = team_pf_by_week[playoff_start].get(lower_team, 0.0)
                        # annotate this week (semis) and next week (finals)
                        for rid in top4:
                            stage_label_map[(season, playoff_start, rid)] = "Semifinal"
                        for rid in bottom4:
                            stage_label_map[(season, playoff_start, rid)] = "Toilet Semis"
                        # next week labels depend on semis results
                        finals_week = playoff_start + 1
                        if week_allowed(finals_week) and finals_week in matchups_by_week:
                            # Determine winners/losers within those brackets
                            for rid in top4:
                                opp = opp_rid_map.get((season, playoff_start, rid))
                                if opp is None:
                                    continue
                                if rid < opp:  # handle each pair once
                                    pf_a = team_pf_by_week[playoff_start].get(roster_to_team[rid], 0.0)
                                    pf_b = team_pf_by_week[playoff_start].get(roster_to_team[opp], 0.0)
                                    win_a = (pf_a > pf_b)
                                    winner = rid if win_a else opp
                                    loser = opp if win_a else rid
                                    stage_label_map[(season, finals_week, winner)] = "Final"
                                    stage_label_map[(season, finals_week, loser)] = "3rd Place"
                            for rid in bottom4:
                                opp = opp_rid_map.get((season, playoff_start, rid))
                                if opp is None:
                                    continue
                                if rid < opp:
                                    pf_a = team_pf_by_week[playoff_start].get(roster_to_team[rid], 0.0)
                                    pf_b = team_pf_by_week[playoff_start].get(roster_to_team[opp], 0.0)
                                    win_a = (pf_a > pf_b)
                                    winner = rid if win_a else opp
                                    loser = opp if win_a else rid
                                    stage_label_map[(season, finals_week, winner)] = "Toilet Final"
                                    stage_label_map[(season, finals_week, loser)] = "Toilet Trash"
                except Exception as e:
                    _log_exc(debug, f"playoff_labeling_{season}_wk{wk}", e)

        # ------------- Weekly loop to build team_week & player_week -------------
        # Inherit cross-season state: Week 1 of THIS season compares against
        # the last played week of the prior season (≈ championship week).
        # Each week's loop updates the cross-season dicts in place via the
        # `prev_starters_by_team[team] = cur_s` / `prev_roster_by_team[team]
        # = cur_r` assignments below, so by the time the season ends these
        # carry the championship-week state forward to next season's wk 1.
        prev_starters_by_team: Dict[str, set] = prev_starters_by_team_xseason
        prev_roster_by_team: Dict[str, set] = prev_roster_by_team_xseason
        player_last5_healthy: Dict[str, deque] = defaultdict(lambda: deque(maxlen=5))  # for hardship later
        # for player awards
        awards_weekly = defaultdict(list)  # (season,wk) -> list of (pid, team, pts, started, pos)

        for wk, mu in matchups_by_week.items():
            stage = _matchup_stage(wk, playoff_start)
            # tx summaries
            faab_spent: Dict[str, float] = defaultdict(float)
            trade_count: Dict[str, int] = defaultdict(int)
            tx_count: Dict[str, int] = defaultdict(int)

            for t in tx_by_week.get(wk, []):
                try:
                    ttype = t.get("type")

                    # Commissioner wash (Phase 6B): a no-op same-day correction —
                    # don't count it (tx / trade / FAAB).
                    if str(t.get("transaction_id") or id(t)) in wash_tx_ids:
                        continue

                    # Net-zero FAAB swap (Phase 7A): a joke trade where nothing
                    # changed hands — delete it entirely (no count, no row).
                    if ttype == "trade" and _trade_is_netzero_swap(t):
                        continue

                    # Date-validity gate. If Sleeper's 'created' timestamp
                    # is missing or unparseable, this transaction can't be
                    # anchored in time — it would emit a row with Date='N/A'
                    # which breaks Link prev/next ordering, Tanking joins,
                    # and (Team, Season) reconciliation. Skip the whole
                    # transaction so tx_count / trade_count / faab_spent
                    # stay row-for-row consistent with the detail CSVs.
                    if _epoch_ms_to_dt(t.get("created")) is None:
                        continue

                    # Resolve every team participating in this transaction via
                    # roster_ids -> roster_to_team. We deliberately do NOT use
                    # user_handle.get(creator) here: user_handle is rebuilt per
                    # season from that season's users feed, so it returns this
                    # season's display name (e.g. "Shmuel256"), which can drift
                    # year-to-year. roster_to_team uses the canonical stable
                    # display name that team_week / team_year rows are keyed by.
                    # Without this, the same manager's tx fragment under two
                    # different team labels and the rollup is wrong.
                    roster_ids_in_tx = [_to_int(r, None) for r in (t.get("roster_ids") or [])]
                    roster_ids_in_tx = [r for r in roster_ids_in_tx if r is not None]
                    teams_in_tx: list[str] = []
                    for rid_ in roster_ids_in_tx:
                        nm = roster_to_team.get(int(rid_))
                        if nm and nm not in teams_in_tx:
                            teams_in_tx.append(nm)

                    # creator_team is only used as a tiebreaker when roster_ids
                    # is empty (rare; mostly placeholder transactions). Resolve
                    # creator -> owner_id -> roster_id -> stable team name when
                    # possible so we still don't go through user_handle directly.
                    creator = str(t.get("creator") or "")
                    creator_team = None
                    if creator and not teams_in_tx:
                        for rid_, owner_ in roster_owner.items():
                            if owner_ == creator:
                                creator_team = roster_to_team.get(rid_)
                                break
                        if creator_team is None:
                            creator_team = user_handle.get(creator)

                    if ttype == "trade":
                        # Weekly trade rollup (Phase 5C item 9): a deep-offseason
                        # trade (created > 7 days before kickoff) does NOT count
                        # toward this week's WEEKLY trade tally; it still appears
                        # in the season/all-time totals (which are counted from
                        # the distinct trade ledger, not the weekly sum).
                        _tr_dt = _epoch_ms_to_dt(t.get("created"))
                        _kick = date(int(season), 9, 7)
                        _deep_offseason = bool(
                            _tr_dt is not None
                            and _tr_dt.date() < _kick
                            and (_kick - _tr_dt.date()).days > 7
                        )
                        # Credit both sides.
                        for tm in teams_in_tx:
                            if not _deep_offseason:
                                trade_count[tm] += 1
                            tx_count[tm] += 1
                    else:
                        # waiver / free_agent / commissioner — credit the
                        # destination roster of EACH add AND the dropping
                        # roster of each drop that doesn't pair with an add
                        # in this same transaction. Net effect:
                        #   - swap (1 add + 1 drop on same roster): +1
                        #   - multi-roster commish swap (2 adds + 2 drops
                        #     across 2 rosters): +1 per roster
                        #   - pure drop (0 adds + N drops on same roster):
                        #     +N (each visible in transactions.csv too)
                        # team_week / team_year 'Number of transactions'
                        # reconciles row-for-row with transactions.csv.
                        adds_dict = t.get("adds") if isinstance(t.get("adds"), dict) else {}
                        drops_dict = t.get("drops") if isinstance(t.get("drops"), dict) else {}
                        # rosters that received an add — each gets +1 here.
                        add_rosters: Set[int] = set()
                        for _pid, _rrid in (adds_dict or {}).items():
                            _rid_i = _to_int(_rrid, None)
                            tm_name = roster_to_team.get(int(_rid_i)) if _rid_i is not None else None
                            if not tm_name:
                                tm_name = (teams_in_tx[0] if teams_in_tx else None) or creator_team
                            if tm_name:
                                tx_count[tm_name] += 1
                                if _rid_i is not None:
                                    add_rosters.add(int(_rid_i))
                        # remaining drops (per roster) become "orphan" drops
                        # for tx_count purposes. Group by roster and credit
                        # +1 per leftover drop after pairing one drop with
                        # each add on the same roster.
                        drops_by_rid_count: Dict[int, int] = defaultdict(int)
                        for _dpid, _drid in (drops_dict or {}).items():
                            _drid_i = _to_int(_drid, None)
                            if _drid_i is None:
                                continue
                            drops_by_rid_count[int(_drid_i)] += 1
                        adds_by_rid_count: Dict[int, int] = defaultdict(int)
                        for _pid, _rrid in (adds_dict or {}).items():
                            _rid_i = _to_int(_rrid, None)
                            if _rid_i is None:
                                continue
                            adds_by_rid_count[int(_rid_i)] += 1
                        for _rid_i, n_drops in drops_by_rid_count.items():
                            n_orphan = max(0, n_drops - adds_by_rid_count.get(_rid_i, 0))
                            if n_orphan <= 0:
                                continue
                            tm_name = roster_to_team.get(int(_rid_i))
                            if not tm_name:
                                tm_name = (teams_in_tx[0] if teams_in_tx else None) or creator_team
                            if tm_name:
                                tx_count[tm_name] += n_orphan

                    # FAAB lives under settings.waiver_bid on Sleeper transactions
                    # (legacy code looked under metadata, which was always empty).
                    settings_obj = t.get("settings") or {}
                    if isinstance(settings_obj, dict):
                        bid = _to_float(settings_obj.get("waiver_bid"), 0.0) or 0.0
                    else:
                        bid = 0.0
                    if not bid:
                        meta = t.get("metadata") or {}
                        if isinstance(meta, dict):
                            bid = _to_float(
                                meta.get("waiver_bid") or meta.get("faab") or 0.0, 0.0
                            ) or 0.0
                    if bid:
                        primary = (teams_in_tx[0] if teams_in_tx else None) or creator_team
                        if primary:
                            faab_spent[primary] += float(bid)
                except Exception as e:
                    _log_exc(debug, f"tx_summary_{season}_wk{wk}", e)

            # build team rows
            for m in mu:
                try:
                    rid = _to_int(m.get("roster_id"), None)
                    if rid is None:
                        continue
                    team = roster_to_team.get(rid, f"Roster {rid}")
                    pf = float(_to_float(m.get("points"), 0.0) or 0.0)
                    bonus = semis_bonus_by_week.get(wk, {}).get(rid)
                    if bonus:
                        pf += bonus

                    opp_rid = opp_rid_map.get((season, wk, rid))
                    opp_team = roster_to_team.get(opp_rid, f"Roster {opp_rid}") if opp_rid is not None else None
                    opp_points = opp_pf_map.get((season, wk, rid))
                    margin = (pf - opp_points) if opp_points is not None else None
                    win = None
                    if margin is not None:
                        win = 1 if margin > 0 else 0 if margin < 0 else 0.5

                    starters = [str(x) for x in (m.get("starters") or []) if x]
                    players = [str(x) for x in (m.get("players") or []) if x]
                    ppts_raw = m.get("players_points") or {}
                    ppts: Dict[str, float] = {}
                    if isinstance(ppts_raw, dict):
                        for k, v in ppts_raw.items():
                            try:
                                ppts[str(k)] = float(v)
                            except Exception:
                                pass

                    # Max PF: use proven algorithm (Apps Script parity)
                    max_pf = compute_optimal_lineup(ppts, pid_pos, season)
                    # Sanity check: if a team scored points, Max PF must be > 0 (otherwise per-player points were lost).
                    if (pf or 0.0) > 0.0 and (max_pf or 0.0) <= 0.0:
                        _log(
                            debug,
                            "ERROR: Max PF computed as 0 despite PF>0. league=%s season=%s week=%s roster_id=%s "
                            "pf=%s players_points_type=%s players_points_len=%s"
                            % (
                                league_id,
                                season,
                                wk,
                                rid,
                                pf,
                                type(ppts_raw).__name__,
                                (len(ppts) if isinstance(ppts, dict) else "NA"),
                            ),
                        )
                        LOG.warning(
                            "Max PF sanity: PF>0 but Max PF==0 for %s %s wk=%s roster=%s; leaving Max PF blank. Check raw exports.",
                            season,
                            league_id,
                            wk,
                            rid,
                        )
                    eff = safe_div(pf, max_pf, default=0.0)

                    # expected win percentile vs league that week
                    scores = list(team_pf_by_week.get(wk, {}).values())
                    expected = None
                    luck_raw = None
                    if scores and len(scores) > 1:
                        expected = sum(1 for s in scores if pf > s) / max(1, (len(scores) - 1))
                        if win is not None:
                            luck_raw = (win - expected)

                    prev = prev_starters_by_team.get(team)
                    cur_s = set([pid for pid in starters if _valid_pid(pid)])
                    if prev is None:
                        turnover = None
                    else:
                        inter = cur_s.intersection(prev)
                        turnover = max(len(cur_s), len(prev)) - len(inter)
                    prev_starters_by_team[team] = cur_s

                    prev_r = prev_roster_by_team.get(team)
                    cur_r = set([pid for pid in players if _valid_pid(pid)])
                    if prev_r is None:
                        roster_turnover = None
                    else:
                        inter_r = cur_r.intersection(prev_r)
                        roster_turnover = max(len(cur_r), len(prev_r)) - len(inter_r)
                    prev_roster_by_team[team] = cur_r

                    # Exclude empty/placeholder lineup slots (Sleeper fills an
                    # unset starter slot with "0"); otherwise an empty starter
                    # slot counts as a starter "donut" but not a roster donut,
                    # which made "Number of starter donuts" exceed "Number of
                    # donuts" in championship weeks with unfilled lineups
                    # (e.g. plehv79 2022 wk16).
                    starter_points = [ppts.get(pid, 0.0) for pid in starters if _valid_pid(pid)]
                    # "Number of players ..." count ALL rostered players (item 6);
                    # "Number of starters ..." are the starter-only companions.
                    roster_points = [ppts.get(pid, 0.0) for pid in players if _valid_pid(pid)]
                    donuts = sum(1 for x in roster_points if float(x) == 0.0)
                    under10 = sum(1 for x in roster_points if float(x) < 10.0)
                    over20 = sum(1 for x in roster_points if float(x) > 20.0)
                    over30 = sum(1 for x in roster_points if float(x) > 30.0)
                    over40 = sum(1 for x in roster_points if float(x) > 40.0)
                    over50 = sum(1 for x in roster_points if float(x) > 50.0)
                    s_donuts = sum(1 for x in starter_points if float(x) == 0.0)
                    s_under10 = sum(1 for x in starter_points if float(x) < 10.0)
                    s_over20 = sum(1 for x in starter_points if float(x) > 20.0)
                    s_over30 = sum(1 for x in starter_points if float(x) > 30.0)
                    s_over40 = sum(1 for x in starter_points if float(x) > 40.0)
                    s_over50 = sum(1 for x in starter_points if float(x) > 50.0)
                    diff_hi_lo = (max(starter_points) - min(starter_points)) if starter_points else None

                    def count_pos(pids, pos):
                        return sum(1 for pid in pids if (pid_pos.get(pid) or "") == pos)

                    qb_s, rb_s, wr_s, te_s = count_pos(starters, "QB"), count_pos(starters, "RB"), count_pos(starters, "WR"), count_pos(starters, "TE")
                    qb_r, rb_r, wr_r, te_r = count_pos(players, "QB"), count_pos(players, "RB"), count_pos(players, "WR"), count_pos(players, "TE")

                    rook_s = len({pid for pid in starters if is_rookie_pid(pid, season)})
                    rook_r = len({pid for pid in players if is_rookie_pid(pid, season)})

                    approx_date = date(season, 9, 1) + timedelta(days=7 * (wk - 1))
                    ages = [a for a in (_calc_age(pid_meta.get(pid, {}).get("birth_date"), approx_date) for pid in players) if a is not None]
                    avg_age = round(sum(ages) / len(ages), 2) if ages else None
                    # 'Team age including picks' uses the per-week
                    # pick ownership tracking. The lookup helper
                    # _picks_held_by_team_at walks pick_trade_events
                    # (which has real dates from the source trade
                    # transactions) and synthesizes commissioner-moved
                    # picks as 'always with current owner'. Each pick
                    # contributes _pick_expected_age(pick_year,
                    # approx_date) — a future-rookie age that grows
                    # as the draft date approaches.
                    pick_ages_this_week = _picks_held_by_team_at(
                        team, int(season), approx_date,
                    )
                    combined_ages = ages + pick_ages_this_week
                    avg_age_inc_picks = (
                        round(sum(combined_ages) / len(combined_ages), 2) if combined_ages else None
                    )

                    def _nfl_team_for_pid_week(pid_val: Any) -> Optional[str]:
                        m = pid_meta.get(str(pid_val), {}) if isinstance(pid_meta, dict) else {}
                        gs = dp_sleeper_to_gsis.get(str(pid_val)) or m.get("gsis_id") or sleeper_to_gsis.get(str(pid_val))
                        tm = (
                            (player_team_by_week.get((str(gs), int(wk))) if gs else None)
                            or (player_season_team.get(str(gs)) if gs else None)
                            # On a roster that week/season but with no stats
                            # (IR / suspended / PUP) -> keep their real team.
                            or (roster_team_by_week.get((str(gs), int(season), int(wk))) if gs else None)
                            or (roster_team_by_season.get((str(gs), int(season))) if gs else None)
                            # A real NFL identity (gsis) but on NO roster that
                            # season — a free agent or retired. Assign the 33rd
                            # "NFL" sentinel rather than the live Sleeper snapshot
                            # (current-team, churns between builds and is wrong
                            # for a past season). Players with no gsis (team DSTs
                            # / unmapped) keep the Sleeper team.
                            or ("NFL" if gs else None)
                            or m.get("team")
                        )
                        return _norm_team(tm)

                    roster_nfl_teams = [t for t in (_nfl_team_for_pid_week(pid) for pid in players) if t]
                    start_nfl_teams = [t for t in (_nfl_team_for_pid_week(pid) for pid in starters) if t]
                    most_start_same = max(Counter(start_nfl_teams).values()) if start_nfl_teams else None
                    most_roster_same = max(Counter(roster_nfl_teams).values()) if roster_nfl_teams else None
                    most_start_team = Counter(start_nfl_teams).most_common(1)[0][0] if start_nfl_teams else None
                    most_roster_team = Counter(roster_nfl_teams).most_common(1)[0][0] if roster_nfl_teams else None

                    def max_same_team_by_pos(pids, pos):
                        # Resolve via the same deterministic helper as the other
                        # NFL-team columns (week stats -> season stats -> "NFL"
                        # sentinel), not the raw Sleeper snapshot, so unrostered
                        # players don't churn this count between builds.
                        teams = []
                        for pid in pids:
                            if (pid_pos.get(pid) or "").upper() == pos:
                                _t = _nfl_team_for_pid_week(pid)
                                if _t:
                                    teams.append(_t)
                        return max(Counter(teams).values()) if teams else None

                    # Opponent label per playoffs spec
                    opp_label = opp_team
                    label = stage_label_map.get((season, wk, rid))
                    week_name = label if label else f"Week {wk}"

                    # Skip team_week / player_week rows for preseason weeks
                    # where no games have been played. Transactions / trades
                    # for the same week were already harvested earlier in this
                    # iteration, so 2026 offseason trades still reach
                    # trades.csv without polluting the score tables with
                    # phantom 0-PF rows.
                    if preseason_only:
                        continue
                    team_week_rows.append({
                        "Team": team,
                        "Opponent Team (raw)": opp_team,
                        "Week": wk,
                        "Week Name": week_name,
                        "Year": season,
                        "PF": round(pf, 2),
                        "Win?": win,
                        "Opponent": opp_team,
                        "Week label": label,
                        "Round": _round_from_label(label),
                        "Points against": round(float(opp_points), 2) if opp_points is not None else None,
                        "Margin": round(float(margin), 2) if margin is not None else None,
                        "Max PF": round(max_pf, 2) if max_pf is not None else None,
                        "Efficiency": round(eff, 4) if eff is not None else None,
                        "Number of Injuries": None,         # computed later from player_week
                        "Number of suspensions": None,      # computed later from player_week
                        "Number of players on bye": None,   # computed later from player_week
                        "Largest deficit overcome (if win)": None,
                        "Starter turnover from previous week": turnover,
                        "Roster turnover from previous week": roster_turnover,
                        "Difference in pregame avg max PF from opponent": None,  # computed later
                        "UPST": None,                      # computed later
                        "Hardship": None,                  # computed later
                        "Tanking": None,                   # computed later (needs pick ledger; best effort later)
                        # Internal-only (filtered from CSV by the plan catalog):
                        # entity count behind "Team age including picks" =
                        # rostered players + future picks held. Phase 6E uses
                        # this as the denominator N when computing the marginal
                        # tanking delta of a transaction/trade.
                        "_RosterNInclPicks": len(combined_ages),
                        "Luck": round(luck_raw, 4) if luck_raw is not None else None,
                        "Win Variance": round(luck_raw, 4) if luck_raw is not None else None,
                        "Brosenzweig": None,
                        "Sisenzweig": None,
                        "Number of donuts": donuts,
                        "Number of starter donuts": s_donuts,
                        "Number of players under 10": under10,
                        "Number of starters under 10": s_under10,
                        "Number of players over 20": over20,
                        "Number of starters over 20": s_over20,
                        "Number of players over 30": over30,
                        "Number of starters over 30": s_over30,
                        "Number of players over 40": over40,
                        "Number of starters over 40": s_over40,
                        "Number of players over 50": over50,
                        "Number of starters over 50": s_over50,
                        "Number of QB started": qb_s,
                        "Number of WR started": wr_s,
                        "Number of RB started": rb_s,
                        "Number of TE started": te_s,
                        "Number of QB rostered": qb_r,
                        "Number of WR rostered": wr_r,
                        "Number of RB rostered": rb_r,
                        "Number of TE rostered": te_r,
                        "Number of transactions": int(tx_count.get(team, 0)),
                        "Number of trades": int(trade_count.get(team, 0)),
                        "Amount of FAAB spent": round(float(faab_spent.get(team, 0.0)), 2),
                        "Most number of players started from same NFL team": most_start_same,
                        "Most number of players started from same NFL team (team)": most_start_team,
                        "Most number of players rostered from same NFL team": most_roster_same,
                        "Most number of players rostered from same NFL team (team)": most_roster_team,
                        "Most number of QBs started from same NFL team": max_same_team_by_pos(starters, "QB"),
                        "Most number of QBs rostered from same NFL team": max_same_team_by_pos(players, "QB"),
                        "Most number of RBs started from same NFL team": max_same_team_by_pos(starters, "RB"),
                        "Most number of RBs rostered from same NFL team": max_same_team_by_pos(players, "RB"),
                        "Most number of WR started from same NFL team": max_same_team_by_pos(starters, "WR"),
                        "Most number of WR rostered from same NFL team": max_same_team_by_pos(players, "WR"),
                        "Most number of TE started from same NFL team": max_same_team_by_pos(starters, "TE"),
                        "Most number of TE rostered from same NFL team": max_same_team_by_pos(players, "TE"),
                        "Number of NFL teams among starting players": len(set(start_nfl_teams)) if start_nfl_teams else None,
                        "Number of NFL teams among rostered players": len(set(roster_nfl_teams)) if roster_nfl_teams else None,
                        "Number of rookies started": rook_s,
                        "Number of rookies rostered": rook_r,
                        "Player average age": avg_age,
                        "Team age including picks": avg_age_inc_picks,
                        "Difference between highest and lowest starters": round(diff_hi_lo, 2) if diff_hi_lo is not None else None,
                        "Combined matchup score": round(pf + opp_points, 2) if opp_points is not None else None,
                        "Win streak": None,
                        "Loss streak": None,
                        "Win streak counting previous season": None,
                        "Loss streak counting previous season": None,
                        "Top half of league?": None,
                        "Highest score?": None,
                        "Lowest score?": None,
                        "Narrowest victory?": None,
                        "Largest blowout?": None,
                        "Most efficient?": None,
                        "Least efficient?": None,
                        "Increase in points from previous week": None,
                        "Number of cuffs rostered": None,
                        "Number of cuffs started": None,
                        "Future draft capital": None,
                        "Startup draft players remaining": None,
                        # leave remaining plan columns to enforcement step
                    })

                    # starter slot labels: map roster positions to labeled starter slots
                    starter_slot = {}
                    roster_positions = [str(x) for x in (lg.get("roster_positions") or []) if x]
                    non_start_slots = {"BN", "BE", "BENCH", "IR", "TAXI"}
                    starter_slots = [pos for pos in roster_positions if str(pos).upper() not in non_start_slots]

                    def _base_slot(pos: str) -> str:
                        upper = str(pos or "").upper()
                        if upper in ("SUPER_FLEX", "SUPERFLEX", "SFLEX", "SFLX"):
                            return "SFLX"
                        if upper in ("FLEX", "FLX"):
                            return "FLX"
                        return upper

                    if starter_slots and starters:
                        counts: Dict[str, int] = {}
                        for pid, slot in zip(starters, starter_slots):
                            base = _base_slot(slot)
                            if not base:
                                continue
                            counts[base] = counts.get(base, 0) + 1
                            idx = counts[base]

                            label = base
                            if base == "RB":
                                label = f"RB{idx}"
                            elif base == "WR":
                                label = f"WR{idx}"
                            elif base == "FLX":
                                label = f"FLX{idx}"
                            elif base == "SFLX":
                                label = "SFLX"
                            elif base == "QB":
                                label = "QB"
                            starter_slot[pid] = label

                    # fallback to player position if starter slot data is missing
                    if not starter_slot:
                        for pid in starters:
                            starter_slot[pid] = pid_pos.get(pid)

                    played_set = played_by_week.get(wk, set())

                    bench = [pid for pid in players if pid not in starters]
                    best_bench_pid = None
                    best_bench_pts = None
                    if bench:
                        best_bench_pid = max(bench, key=lambda pid: ppts.get(pid, 0.0))
                        best_bench_pts = float(ppts.get(best_bench_pid, 0.0))

                    worst_starter_pid = None
                    worst_starter_pts = None
                    if starters:
                        worst_starter_pid = min(starters, key=lambda pid: ppts.get(pid, 0.0))
                        worst_starter_pts = float(ppts.get(worst_starter_pid, 0.0))

                    for pid in players:
                        meta = pid_meta.get(pid, {})
                        full_name = meta.get("full_name") or pid
                        position = pid_pos.get(pid)

                        # gsis id lookup for nflverse (pre-indexed for speed/reliability)
                        gsis = (
                            dp_sleeper_to_gsis.get(str(pid))
                            or meta.get("gsis_id")
                            or sleeper_to_gsis.get(str(pid))
                        )

                        # Prefer week-specific nflverse team when available;
                        # then a "NFL" sentinel for a player with a gsis but no
                        # nflverse team that season (free agent / retired —
                        # unrostered in the real NFL), so they show up as such
                        # deterministically instead of via the churny Sleeper
                        # snapshot. No-gsis players (DSTs / unmapped) keep it.
                        nfl_team = (
                            (player_team_by_week.get((str(gsis), int(wk))) if gsis else None)
                            or (player_season_team.get(str(gsis)) if gsis else None)
                            # On a roster but no stats (IR / suspended / PUP).
                            or (roster_team_by_week.get((str(gsis), int(season), int(wk))) if gsis else None)
                            or (roster_team_by_season.get((str(gsis), int(season))) if gsis else None)
                            or ("NFL" if gsis else None)
                            or meta.get("team")
                        )
                        nfl_team = _norm_team(nfl_team)

                        pts = float(ppts.get(pid, 0.0))
                        started = pid in starters
                        slot = starter_slot.get(pid) if started else "N/A"
                        # Position-as-of-week from nflverse weekly stats when present; otherwise
                        # fall back to Sleeper's live position. Handles in-career switchers
                        # (e.g. Taysom Hill QB/TE) and rookies whose Sleeper meta lags.
                        player_position = (
                            (player_pos_by_week.get((str(gsis), int(wk))) if gsis else None)
                            or pid_pos.get(pid)
                            or None
                        )

                        # Bye is schedule-based. If player scored >0 -> not a bye.
                        # The "NFL" sentinel (no active NFL team: retired /
                        # unsigned / out of the league) has no game at all, so
                        # the week is a BYE, not an injury — audit run-2 F4:
                        # retired meme-pickups (Brady '24/'25, Brees '24, OBJ
                        # '22, …) were counted as injuries, inflating team
                        # injury tallies and deciding a Most injured? award.
                        # bye=True also keeps these weeks out of the played-week
                        # denominators (Adjusted Avg, rostered floors), same as
                        # before when they were flagged injured.
                        bye = None
                        if nfl_team and _norm_team(nfl_team) == "NFL":
                            bye = True
                        elif nfl_team and played_set:
                            bye = (_norm_team(nfl_team) not in played_set)
                        if pts > 0:
                            # Player scored points; not a bye week. Do NOT override injury/suspension flags:
                            # a player can play while injured or returning from suspension.
                            bye = False

                        # Flags (platform primary, nflverse secondary)
                        # Injury/suspension (new approach): authoritative nflverse injuries, keyed by gsis_id (via player_ids mapping).
                        # Only mark if the player did NOT play that week (no stats row) and it is not a bye.
                        inj = False
                        susp = False
                        try:
                            played_players = played_players_by_week.get(int(wk), set())
                            played = bool(gsis) and (str(gsis) in played_players)
                        except Exception:
                            played = False
                        # bye may be None when we couldn't compute it (no NFL team
                        # resolvable + player never had stats this season — e.g.
                        # career-ending injury cases like Gus Edwards 2021, Tarik
                        # Cohen 2021). Treat None like False so we still consider
                        # the player injured when they have pts=0 and no
                        # contradicting suspension entry.
                        if ((pts or 0.0) == 0.0) and (bye is not True) and (not played) and gsis:
                            existing = injuries_by_gsis_week.get((str(gsis), season, int(wk)))
                            if existing is not None and existing[1] is True:
                                # Confirmed suspension wins.
                                susp = True
                                inj = False
                            else:
                                # Player was rostered, scored 0, not on bye, did
                                # not appear in nflverse stats — they were
                                # missing for some reason. Default to injury.
                                # Captures full-season IR cases the per-season
                                # gap-fill couldn't see because the player was
                                # dropped from the roster before season end.
                                inj = True
                                susp = False

                        # Fallback injury/suspension inference (Sleeper metadata is not historical but fixes obvious cases):
                        # If the player scored 0, was not on bye, and Sleeper marks them OUT/IR/SUSP, treat as missed.
                        if (pts or 0.0) == 0.0 and bye is False:
                            meta_p = pid_meta.get(str(pid), {}) if isinstance(pid_meta, dict) else {}
                            st = str(meta_p.get("status") or "").lower()
                            inj_st = str(meta_p.get("injury_status") or meta_p.get("injuryStatus") or "").lower()
                            # suspension signals
                            if ("susp" in st) or ("susp" in inj_st):
                                susp = True
                                inj = False if inj is None else inj
                            # injury signals
                            elif any(x in (st + " " + inj_st) for x in ["out", "ir", "inactive", "pup", "doubtful"]):
                                inj = True

                        # PR E fix B: the in-house weekly Sleeper injury tracker
                        # is the PRIMARY source — it's the historical, per-week
                        # snapshot of Sleeper's own diagnoses, so it overrides the
                        # nflverse/meta inference above when it has this exact
                        # (player, season, week). Covers injury, suspension AND
                        # bye (bye via the captured NFL team vs the fixed
                        # schedule, so traded players get the right bye). Empty
                        # until 2026 wk1 -> a no-op on historical data.
                        _trk = injury_tracker_idx.get((str(pid), int(season), int(wk)))
                        if _trk:
                            _ov = _resolve_injury_flags(_trk.get("status"), _trk.get("bye"), pts)
                            if _ov is not None:
                                _oi, _os, _ob = _ov
                                if _ob is True:
                                    bye = True
                                    inj = False
                                    susp = False
                                elif bye is not True:
                                    # Don't override a real bye with injury/suspension.
                                    inj = _oi
                                    susp = _os

                        if inj is None:
                            inj = False
                        if susp is None:
                            susp = False
                        if susp:
                            inj = False
                        if bye is None:
                            bye = False

                        rookie = is_rookie_pid(pid, season)
                        age = _calc_age(meta.get("birth_date"), approx_date)

                        diff_best_bench = (pts - best_bench_pts) if (started and best_bench_pts is not None) else None
                        diff_worst_starter = (pts - worst_starter_pts) if ((not started) and worst_starter_pts is not None) else None
                        ref_player = None
                        if started and best_bench_pid:
                            ref_player = pid_meta.get(best_bench_pid, {}).get("full_name") or best_bench_pid
                        elif (not started) and worst_starter_pid:
                            ref_player = pid_meta.get(worst_starter_pid, {}).get("full_name") or worst_starter_pid

                        if preseason_only:
                            continue
                        player_week_rows.append({
                            "Player ID": str(pid) if pid is not None else None,
                            "Player": full_name,
                            "Team": team,
                            "Week": wk,
                            "Year": season,
                            "Points": round(pts, 2),
                            "Injury?": bool(inj),
                            "Suspension?": bool(susp),
                            "Bye?": bool(bye),
                            "Starter/Bench": "Starter" if started else "Bench",
                            "% of points (if starter)": round(pts / pf, 4) if started and pf else None,
                            "Position started in (if starter)": slot,
                            "Position": player_position,
                            "Change from previous week": None,
                            "Change from previous 5 weeks avg": None,
                            "Change from career average to that point": None,
                            "Change from overall career average": None,
                            "Number of weeks on team": None,
                            "Number of consecutive weeks on bench before start (if starter)": None,
                            "Number of consecutive weeks on bench before start excluding injury/bye (if starter)": None,
                            "Total weeks as team starter to that point": None,
                            "Total weeks on bench to that point": None,
                            "Total weeks as team starter on that team this season": None,
                            "Total weeks on bench on that team this season": None,
                            "- Activated Cuff? (Was a player of the same nfl team/position & who averages >10 PPG more over last 5 played games injured? Only for players with avg <10 PPG)": 0,
                            "Difference from best startable bench (if starter)": round(diff_best_bench, 2) if diff_best_bench is not None else None,
                            "Difference from worst benchable starter (if bench)": round(diff_worst_starter, 2) if diff_worst_starter is not None else None,
                            "Reference player name": ref_player,
                            "Difference in averages of best/worst startables over previous 5 games": None,
                            "Cuff adjusted difference": None,
                            "Rookie?": 1 if rookie else 0,
                            "Age": age,
                            "NFL team": nfl_team,
                            # award flags (filled later)
                            "Player of the week?": None,
                            "QB of the week?": None,
                            "RB of the week?": None,
                            "WR of the week?": None,
                            "TE of the week?": None,
                            "Benchwarmer of the week?": None,
                            "Bench QB of the week?": None,
                            "Bench RB of the week?": None,
                            "Bench WR of the week?": None,
                            "Bench TE of the week?": None,
                            "Highest starter on team?": None,
                            "Lowest starter on team?": None,
                        })

                        # store for awards later
                        awards_weekly[(season, wk)].append((pid, team, pts, started, pid_pos.get(pid) or ""))

                except Exception as e:
                    _log_exc(debug, f"team_player_rows_{season}_wk{wk}", e)

            # Transactions rows (non-trade) + trades ledger
            for t in tx_by_week.get(wk, []):
                try:
                    ttype = t.get("type")
                    # Commissioner wash (Phase 6B): a no-op same-day correction
                    # is not a real transaction — omit it from the detail rows
                    # and the trades ledger as well as the counts.
                    if str(t.get("transaction_id") or id(t)) in wash_tx_ids:
                        continue
                    # Net-zero FAAB swap (Phase 7A): joke trade, delete entirely.
                    if ttype == "trade" and _trade_is_netzero_swap(t):
                        continue
                    # For waivers, 'created' is when the bid was
                    # submitted but 'status_updated' is when the waiver
                    # actually ran and the player moved. A single
                    # submission date can be misleading when waivers
                    # span multiple processing days — we've seen pairs
                    # of claims submitted within minutes that actually
                    # resolved on different days. Prefer status_updated
                    # for waiver-type transactions; for free_agent,
                    # commissioner, and trades the events resolve at
                    # creation, so 'created' is correct.
                    _t_type = t.get("type")
                    _resolve_ms = t.get("status_updated") if _t_type == "waiver" else None
                    if _resolve_ms is None:
                        _resolve_ms = t.get("created")
                    created_date = _epoch_ms_to_date(_resolve_ms)
                    created_dt = _epoch_ms_to_dt(_resolve_ms)
                    # Mirror the date-validity gate from Loop 1 (per-week
                    # counters). If we can't anchor the transaction in
                    # time, don't emit a row — the matching tx_count
                    # entry was already skipped upstream.
                    if created_dt is None:
                        continue
                    creator = str(t.get("creator") or "")
                    # Resolve via the canonical roster_to_team mapping so the
                    # 'Team' column in the output stays consistent across
                    # seasons even if the manager renames themselves on
                    # Sleeper (e.g. shmuel256 / Shmuel256). roster_ids in the
                    # transaction usually has exactly one entry for non-trade
                    # actions; fall back to user_handle when there's no
                    # resolvable roster_id (rare).
                    team = None
                    try:
                        tx_roster_ids = [_to_int(r, None) for r in (t.get("roster_ids") or [])]
                        for rid_ in tx_roster_ids:
                            if rid_ is None:
                                continue
                            nm = roster_to_team.get(int(rid_))
                            if nm:
                                team = nm
                                break
                    except Exception:
                        team = None
                    if not team and creator:
                        # creator -> owner_id -> roster_id -> canonical team
                        for rid_, owner_ in roster_owner.items():
                            if owner_ == creator:
                                team = roster_to_team.get(rid_)
                                if team:
                                    break
                    if not team and creator:
                        team = user_handle.get(creator)

                    # Trades ledger (one row per involved team)
                    if ttype == "trade":
                        roster_ids = t.get("roster_ids") or []
                        if not isinstance(roster_ids, list):
                            roster_ids = []
                        roster_ids_int = [int(x) for x in roster_ids if _to_int(x, None) is not None]
                        adds = t.get("adds") or {}
                        if not isinstance(adds, dict):
                            adds = {}
                        draft_picks = t.get("draft_picks") or []
                        if not isinstance(draft_picks, list):
                            draft_picks = []

                        # update pick ledger with trade order
                        if draft_picks:
                            for dp in [x for x in draft_picks if isinstance(x, dict)]:
                                dp_season = _to_int(dp.get("season"), season)
                                dp_round = _to_int(dp.get("round"), None)
                                prev_owner = _to_int(
                                    dp.get("previous_owner_id") or dp.get("previous_owner") or dp.get("previous_owner_roster_id"),
                                    None,
                                )
                                new_owner = _to_int(
                                    dp.get("owner_id") or dp.get("owner_roster_id"),
                                    None,
                                )
                                # Sleeper transaction draft_picks carry `roster_id` =
                                # the pick's ORIGINAL owner (same semantics as
                                # traded_picks). Using it as a fallback for owner_id
                                # was a long-standing misread — that swapped origin
                                # and current owner whenever owner_id was absent
                                # and dropped the actual origin signal.
                                original_owner = _to_int(
                                    dp.get("original_owner_id")
                                    or dp.get("original_owner")
                                    or dp.get("roster_id"),
                                    None,
                                )
                                if prev_owner is None and new_owner is not None and len(roster_ids_int) == 2:
                                    prev_owner = [rid for rid in roster_ids_int if rid != new_owner][0]
                                if dp_season is None or dp_round is None or prev_owner is None or new_owner is None:
                                    continue
                                _ensure_pick_bases(int(dp_season), season)
                                key = _select_pick_key(int(dp_season), int(dp_round), int(prev_owner), original_owner)
                                if not key:
                                    _log(
                                        debug,
                                        f"[{_now_iso()}] WARN pick ledger unresolved: season={dp_season} round={dp_round} prev={prev_owner} new={new_owner} orig={original_owner}",
                                    )
                                    continue
                                pick_holdings[(key[0], key[1], int(prev_owner))] = [
                                    oid for oid in pick_holdings.get((key[0], key[1], int(prev_owner)), []) if oid != key[2]
                                ]
                                pick_holdings[(key[0], key[1], int(new_owner))].append(int(key[2]))
                                pick_holdings[(key[0], key[1], int(new_owner))] = sorted(
                                    set(pick_holdings[(key[0], key[1], int(new_owner))])
                                )
                                pick_current_owner[key] = int(new_owner)
                                pick_trade_events.setdefault(key, []).append(
                                    (created_dt, int(prev_owner), int(new_owner), int(wk)),
                                )

                        # received assets by team (display names) plus a
                        # parallel mapping of sleeper player ids for KTC
                        # value lookups downstream.
                        recv_players: Dict[int, List[str]] = defaultdict(list)
                        recv_player_ids: Dict[int, List[str]] = defaultdict(list)
                        for pid, rrid in adds.items():
                            rr = _to_int(rrid, None)
                            if rr is None:
                                continue
                            recv_players[rr].append(pid_meta.get(str(pid), {}).get("full_name") or str(pid))
                            recv_player_ids[rr].append(str(pid))

                        recv_picks: Dict[int, List[str]] = defaultdict(list)
                        # Parallel metadata so a post-pass can substitute
                        # pick labels with specific-slot + drafted player
                        # ('2024 1.??' -> '2024 1.05(B. Robinson)') once
                        # pick_history is built. dp.get('roster_id') is
                        # the pick's ORIGINAL owner per Sleeper's API.
                        recv_pick_meta: Dict[int, List[Tuple[int, int, str]]] = defaultdict(list)
                        for dp in draft_picks:
                            if not isinstance(dp, dict):
                                continue
                            owner_id = _to_int(dp.get("owner_id"), None)
                            if owner_id is None:
                                continue
                            dp_season = _to_int(dp.get("season"), season)
                            dp_round = _to_int(dp.get("round"), None)
                            if dp_round is None:
                                continue
                            label = _format_pick_label(int(dp_season), dp_round, None)
                            if label:
                                recv_picks[owner_id].append(label)
                                # Original team name as of the pick's
                                # season (roster IDs are stable within
                                # a league iteration in Sleeper).
                                orig_roster = _to_int(dp.get("roster_id"), None)
                                orig_team = (
                                    season_roster_to_team.get(int(dp_season), {}).get(int(orig_roster))
                                    if orig_roster is not None else None
                                ) or roster_to_team.get(int(orig_roster) if orig_roster is not None else -1, "")
                                recv_pick_meta[owner_id].append(
                                    (int(dp_season), int(dp_round), str(orig_team or ""))
                                )

                        # FAAB as a tradeable asset (Phase 7A). Sleeper records
                        # moved waiver budget in `waiver_budget` as
                        # {amount, sender(roster), receiver(roster)} entries.
                        # Render received FAAB PER SENDER (one "$N FAAB" asset
                        # per source roster) rather than lumping a receiver's
                        # total into a single "$N FAAB". This mirrors the sent
                        # side (which sums per sender) so a multi-sender 3-team
                        # deal shows e.g. "$4 FAAB" + "$15 FAAB" on the receiver
                        # instead of a lumped "$19 FAAB". Dollars still conserve.
                        # The per-roster dropped pass below picks each up as
                        # 'sent' for the counterparty.
                        _faab_rcv_by_snd: Dict[int, Dict[int, int]] = defaultdict(lambda: defaultdict(int))
                        for wb in (t.get("waiver_budget") or []):
                            if not isinstance(wb, dict):
                                continue
                            _amt = _to_int(wb.get("amount"), None)
                            _rcv = _to_int(wb.get("receiver"), None)
                            _snd = _to_int(wb.get("sender"), None)
                            if _amt is None or _rcv is None or int(_amt) <= 0:
                                continue
                            # Group by sender so multiple budget entries from the
                            # same sender collapse into one asset (sender id may
                            # be None — bucket those together under -1).
                            _faab_rcv_by_snd[int(_rcv)][int(_snd) if _snd is not None else -1] += int(_amt)
                        recv_faab: Dict[int, List[str]] = {
                            _r: [f"${_a} FAAB" for _s, _a in sorted(_by_snd.items())]
                            for _r, _by_snd in _faab_rcv_by_snd.items()
                        }

                        # SENT side, keyed by the roster that actually GAVE UP
                        # each asset (Phase 7 fix). Building "sent" from "what
                        # the OTHER teams received" double-counts in 3+ team
                        # trades — each team then lists every other team's haul.
                        # Instead attribute each asset to its real source: a
                        # player to the roster that DROPPED it, a pick to its
                        # previous owner, FAAB to its sender.
                        _drops_dict = t.get("drops") if isinstance(t.get("drops"), dict) else {}
                        drop_players: Dict[int, List[str]] = defaultdict(list)
                        drop_player_ids: Dict[int, List[str]] = defaultdict(list)
                        for _pid, _rrid in _drops_dict.items():
                            _rr = _to_int(_rrid, None)
                            if _rr is None:
                                continue
                            drop_players[_rr].append(pid_meta.get(str(_pid), {}).get("full_name") or str(_pid))
                            drop_player_ids[_rr].append(str(_pid))
                        drop_picks: Dict[int, List[str]] = defaultdict(list)
                        drop_pick_meta: Dict[int, List[Tuple[int, int, str]]] = defaultdict(list)
                        for dp in draft_picks:
                            if not isinstance(dp, dict):
                                continue
                            _prev = _to_int(
                                dp.get("previous_owner_id") or dp.get("previous_owner") or dp.get("previous_owner_roster_id"),
                                None,
                            )
                            if _prev is None and len(roster_ids_int) == 2:
                                _no = _to_int(dp.get("owner_id"), None)
                                if _no is not None:
                                    _prev = next((rid for rid in roster_ids_int if rid != _no), None)
                            if _prev is None:
                                continue
                            _ds = _to_int(dp.get("season"), season)
                            _dr = _to_int(dp.get("round"), None)
                            if _dr is None:
                                continue
                            _lbl = _format_pick_label(int(_ds), _dr, None)
                            if not _lbl:
                                continue
                            _or = _to_int(dp.get("roster_id"), None)
                            _ot = (
                                season_roster_to_team.get(int(_ds), {}).get(int(_or))
                                if _or is not None else None
                            ) or roster_to_team.get(int(_or) if _or is not None else -1, "")
                            drop_picks[_prev].append(_lbl)
                            drop_pick_meta[_prev].append((int(_ds), int(_dr), str(_ot or "")))
                        _faab_by_snd: Dict[int, int] = defaultdict(int)
                        for wb in (t.get("waiver_budget") or []):
                            if not isinstance(wb, dict):
                                continue
                            _amt = _to_int(wb.get("amount"), None)
                            _snd = _to_int(wb.get("sender"), None)
                            if _amt is None or _snd is None or int(_amt) <= 0:
                                continue
                            _faab_by_snd[int(_snd)] += int(_amt)
                        drop_faab: Dict[int, List[str]] = {
                            _r: [f"${_a} FAAB"] for _r, _a in _faab_by_snd.items()
                        }

                        # Build row per roster in roster_ids_int
                        for rid in roster_ids_int:
                            tm = roster_to_team.get(rid, f"Roster {rid}")
                            others = [roster_to_team.get(o, f"Roster {o}") for o in roster_ids_int if o != rid]
                            received = []
                            received.extend(recv_players.get(rid, []))
                            received.extend(recv_picks.get(rid, []))
                            received.extend(recv_faab.get(rid, []))
                            received_ids = list(recv_player_ids.get(rid, []))
                            received_picks = list(recv_picks.get(rid, []))
                            received_pick_meta = list(recv_pick_meta.get(rid, []))
                            dropped = []
                            dropped.extend(drop_players.get(rid, []))
                            dropped.extend(drop_picks.get(rid, []))
                            dropped.extend(drop_faab.get(rid, []))
                            dropped_ids: List[str] = list(drop_player_ids.get(rid, []))
                            dropped_picks: List[str] = list(drop_picks.get(rid, []))
                            dropped_pick_meta: List[Tuple[int, int, str]] = list(drop_pick_meta.get(rid, []))
                            trades_rows.append({
                                "Team": tm,
                                # i7 (#27): counterparties go in their OWN numbered
                                # columns ("Team's traded with 1/2/…") — one per
                                # other team in the deal — so each can hyperlink to
                                # that team's row of the SAME trade. Injected just
                                # below the append (dynamic count: 1 for a swap, 2+
                                # for a multi-team trade).
                                # Distinct teams in the deal (this team + every
                                # counterparty) — 2 for a normal swap, 3+ for a
                                # multi-team trade (Phase 7B).
                                "Number of teams involved": len({x for x in others if x}) + 1,
                                "Assets received": "; ".join(received) if received else None,
                                "Assets sent": "; ".join(dropped) if dropped else None,
                                "Date": created_dt.isoformat() if created_dt else (str(created_date) if created_date else None),
                                "Season": int(season),
                                # Internal-only sleeper ID lists used by the
                                # KTC value lookup pass. Not in the schema
                                # catalog, so they get filtered out before
                                # write_outputs serializes the CSV.
                                "_recv_player_ids": received_ids,
                                "_drop_player_ids": dropped_ids,
                                "_recv_picks": received_picks,
                                "_drop_picks": dropped_picks,
                                "_recv_pick_meta": received_pick_meta,
                                "_drop_pick_meta": dropped_pick_meta,
                                # FAAB $ moved (numeric) — valued in KTC via the
                                # avg KTC-per-$ implied by FAAB-for-asset trades.
                                "_recv_faab": float(sum(_faab_rcv_by_snd.get(rid, {}).values())),
                                "_drop_faab": float(_faab_by_snd.get(rid, 0)),
                                # Shared id for all rows of the SAME trade event
                                # (a 3-team trade emits N rows). Lets the link
                                # chains skip a trade's own mirror rows when
                                # finding the next/previous DISTINCT event.
                                "_tx_id": str(t.get("transaction_id") or id(t)),
                                # KTC columns — 'at deal time' is computed
                                # by a post-processing pass; the other three
                                # time points stay None until a follow-up
                                # PR populates them.
                                "KTC value difference at deal time": None,
                                "KTC value difference at end of season": None,
                                "KTC value difference 1 year later": None,
                                "KTC value difference 2 years later": None,
                                "Pick value received": None,
                                "Change in pick value at draft time": None,
                                "Assets retained now": None,
                                "Assets traded away": None,
                                "Return from trades": None,
                                "Additional assets traded away in those deals": None,
                                "Return from trades of trades...of trades. Keep going until present day": None,
                                "Asset difference in average age": None,
                                "Tanking": None,
                                "Link to next transaction": None,
                                "Link to previous transaction": None,
                            })
                            # Split counterparties into numbered columns (one team
                            # each) so the writer can hyperlink each to the other
                            # side of this same trade.
                            for _ti, _to in enumerate(sorted({x for x in others if x}), 1):
                                trades_rows[-1][f"Team's traded with {_ti}"] = _to

                        continue  # don't add to transactions.csv

                    # Non-trade transactions
                    if ttype not in ("waiver", "free_agent", "commissioner"):
                        # keep but label if unknown
                        pass

                    adds = t.get("adds") or {}
                    drops = t.get("drops") or {}
                    meta = t.get("metadata") or {}
                    settings_obj = t.get("settings") or {}
                    # Sleeper exposes the waiver bid on transaction.settings.waiver_bid
                    # (not on metadata, which is almost always null). The previous
                    # code looked at metadata only, so Faab was None for every
                    # transaction, and downstream coercion stamped it as 0.0.
                    faab = None
                    if isinstance(settings_obj, dict):
                        faab = settings_obj.get("waiver_bid")
                    if faab is None and isinstance(meta, dict):
                        faab = meta.get("waiver_bid") or meta.get("faab")
                    # Number of bids: derived in the fetch loop from all
                    # waiver attempts (complete + failed) against the same
                    # player in the same week. Sleeper's API doesn't carry
                    # this on the transaction settings — every field they
                    # do expose (priority, seq, waiver_bid) describes only
                    # THIS team's claim. Resolved per-add below where we
                    # have the player id in hand.
                    num_bids = None  # placeholder; set per-add

                    if not isinstance(adds, dict):
                        adds = {}
                    if not isinstance(drops, dict):
                        drops = {}

                    drops_by_roster: Dict[str, List[str]] = defaultdict(list)
                    for dp, drid in drops.items():
                        drops_by_roster[str(drid)].append(str(dp))
                    for pid, rrid in adds.items():
                        pid = str(pid)
                        rrid_str = str(rrid)
                        player_tx_week[(pid, season, wk)] += 1
                        player_tx_year[(pid, season)] += 1
                        player_tx_all[pid] += 1
                        dropped = None
                        drop_list = drops_by_roster.get(rrid_str)
                        if drop_list:
                            dropped = drop_list.pop(0)
                            if not drop_list:
                                drops_by_roster.pop(rrid_str, None)

                        if dropped:
                            dropped_id = str(dropped)
                            player_tx_week[(dropped_id, season, wk)] += 1
                            player_tx_year[(dropped_id, season)] += 1
                            player_tx_all[dropped_id] += 1
                            player_drop_week[(dropped_id, season, wk)] += 1
                            player_drop_year[(dropped_id, season)] += 1
                            player_drop_all[dropped_id] += 1

                        # Team for THIS specific add = the roster the player
                        # is going TO (rrid). For multi-roster transactions
                        # like commissioner-driven swaps, the outer 'team'
                        # variable points at only the first roster, so a
                        # second add to a different roster needs its own
                        # team resolution. Fall back to the outer 'team' if
                        # rrid doesn't resolve.
                        row_rrid_int = _to_int(rrid_str, None)
                        row_team = (
                            roster_to_team.get(int(row_rrid_int)) if row_rrid_int is not None else None
                        ) or team

                        # Resolve Number of bids, Total FAAB bid, and the
                        # winner-vs-second-place columns from the per-week
                        # tallies. Only meaningful for waiver claims;
                        # free-agent and commissioner adds aren't bid on.
                        row_num_bids = None
                        row_total_faab = None
                        row_faab_diff_2nd = None
                        row_faab_pct_2nd = None
                        if ttype == "waiver":
                            key = (int(season), int(wk), str(pid))
                            row_num_bids = bids_per_player_week.get(key, 0) or None
                            row_total_faab = total_bids_amount_per_player_week.get(key, 0.0) or None
                            # Compute the runner-up from the pre-filter
                            # tally. Two important details:
                            # 1) The winner is the team owning THIS row
                            #    (status=complete). Its bid amount is the
                            #    'faab' value we just read off settings.
                            #    Don't take amounts[0] — Sleeper records
                            #    failed bids alongside the winner, and a
                            #    higher-amount failed bid (roster full,
                            #    not enough FAAB, etc.) would otherwise
                            #    look like 'the winner'.
                            # 2) Any bid > winner_bid_val that ended up
                            #    losing must have been INVALIDATED — a
                            #    valid higher bid would have won the
                            #    auction. Exclude those from runner-up
                            #    consideration; they aren't real competing
                            #    bids. This is what was producing diffs
                            #    larger than the winning bid (e.g. Dont'e
                            #    Thornton showed diff=40 on a 16-FAAB win).
                            winner_bid_val = float(faab) if faab is not None else 0.0
                            competing = list(bid_amounts_per_player_week.get(key, []))
                            try:
                                # Remove exactly one occurrence of the
                                # winning bid (the row we're emitting).
                                # If multiple bids tied at winner_bid_val,
                                # the others are legitimate runners-up.
                                competing.remove(winner_bid_val)
                            except ValueError:
                                pass
                            competing = [b for b in competing if b <= winner_bid_val]
                            if competing:
                                second = max(competing)
                                row_faab_diff_2nd = round(winner_bid_val - second, 2)
                                # FAAB premium % (Phase 6A): the winning bid's
                                # margin over the runner-up as a share of the
                                # WINNING bid — normalized by bid size so it's
                                # comparable across big and small auctions
                                # ($50 over $40 reads the same 20% as $5 over
                                # $4). Bounded 0–100; defined whenever the
                                # winning bid > 0 (premium = 100% vs a $0
                                # runner-up).
                                if winner_bid_val > 0:
                                    row_faab_pct_2nd = round((winner_bid_val - second) / winner_bid_val * 100.0, 2)

                        # FAAB column semantics (user spec / audit fix):
                        # - 2021 and earlier: league had no FAAB; ALL faab
                        #   columns -> N/A regardless of transaction type.
                        # - 2022+:
                        #     waiver       -> numeric (0 if Sleeper didn't
                        #                     record a bid amount — uncontested
                        #                     $0 claim, not 'missing data').
                        #     free_agent   -> N/A (no bidding)
                        #     commissioner -> N/A (no bidding)
                        _faab_emit = faab
                        _total_faab_emit = row_total_faab
                        _faab_diff_emit = row_faab_diff_2nd
                        _faab_pct_emit = row_faab_pct_2nd
                        if int(season) < 2022:
                            _faab_emit = None
                            _total_faab_emit = None
                            _faab_diff_emit = None
                            _faab_pct_emit = None
                        elif ttype == "waiver" and _faab_emit is None:
                            _faab_emit = 0
                        transactions_rows.append({
                            "Team": row_team,
                            "Player Added": pid_meta.get(pid, {}).get("full_name") or pid,
                            "Player Dropped": pid_meta.get(dropped, {}).get("full_name") if dropped else None,
                            "type of transaction (waiver/free agency)": ttype,
                            "Faab": _faab_emit,
                            "Total FAAB bid": _total_faab_emit,
                            "FAAB difference over second place": _faab_diff_emit,
                            "FAAB premium %": _faab_pct_emit,
                            "Date": created_dt.isoformat() if created_dt else (str(created_date) if created_date else None),
                            "Season": int(season),
                            # Internal-only sleeper IDs for the KTC pass.
                            # Filtered out before write_outputs (not in the
                            # plan catalog).
                            "_added_pid": str(pid) if pid else None,
                            "_dropped_pid": str(dropped) if dropped else None,
                            "Number of bids": row_num_bids,
                            "Link to next transaction": None,
                            "Link to previous transaction": None,
                            "Average PPG on team": None,
                            "Average PPG of dropped player over same time": None,
                            "Difference of averages": None,
                            "Difference of averages adjusted by position": None,
                            "Age difference": None,
                            "Player addition value": None,
                            "Cuff at time of pickup?": None,
                            "Weeks between pickup and start": None,
                            "Number of starts before next drop": None,
                            "% of starts made while rostered": None,
                            "Injury adjusted % of starts made while rostered": None,
                            "Date dropped/traded": None,
                            "Tanking": None,
                            "Number of times picked up by this team": None,
                        })

                    for rrid_orphan_str, drop_list in drops_by_roster.items():
                        for dp_str in drop_list:
                            dropped_id = str(dp_str)
                            player_tx_week[(dropped_id, season, wk)] += 1
                            player_tx_year[(dropped_id, season)] += 1
                            player_tx_all[dropped_id] += 1
                            player_drop_week[(dropped_id, season, wk)] += 1
                            player_drop_year[(dropped_id, season)] += 1
                            player_drop_all[dropped_id] += 1
                            # tx_count for orphan drops is credited in Loop 1
                            # (the per-week tx summary pass earlier) so that
                            # team_week's read of tx_count sees the right
                            # totals.
                            # Record an orphan-drop event so the transactions
                            # polish pass can bound 'Date dropped/traded' /
                            # 'Weeks between pickup and start' correctly.
                            try:
                                rid_int = _to_int(rrid_orphan_str, None)
                                drop_team = roster_to_team.get(int(rid_int)) if rid_int is not None else None
                                drop_player_name = (pid_meta.get(dropped_id, {}) or {}).get("full_name") or dropped_id
                                drop_dt = created_dt.isoformat() if created_dt else (str(created_date) if created_date else None)
                                # Skip orphan-drop emission when we have no
                                # timestamp — the row would be unanchored in
                                # time and pollutes downstream reconciliation.
                                if drop_team and drop_player_name and drop_dt:
                                    orphan_drop_events.append({
                                        "Team": drop_team,
                                        "Player Dropped": drop_player_name,
                                        "Date": drop_dt,
                                    })
                                    # Also emit a transactions.csv row so the
                                    # detail file reconciles with team_year's
                                    # 'Number of transactions' count. Pure drops
                                    # are real transactions in Sleeper and were
                                    # previously invisible — tx_count would
                                    # increment but no row would be written.
                                    transactions_rows.append({
                                        "Team": drop_team,
                                        "Player Added": None,
                                        "Player Dropped": drop_player_name,
                                        "type of transaction (waiver/free agency)": ttype,
                                        "Faab": None,
                                        "Total FAAB bid": None,
                                        "FAAB difference over second place": None,
                                        "FAAB premium %": None,
                                        "Date": drop_dt,
                                        "Season": int(season),
                                        "_added_pid": None,
                                        "_dropped_pid": str(dropped_id) if dropped_id else None,
                                        "Number of bids": None,
                                        "Link to next transaction": None,
                                        "Link to previous transaction": None,
                                        "Average PPG on team": None,
                                        "Average PPG of dropped player over same time": None,
                                        "Difference of averages": None,
                                        "Difference of averages adjusted by position": None,
                                        "Age difference": None,
                                        "Player addition value": None,
                                        "Cuff at time of pickup?": None,
                                        "Weeks between pickup and start": None,
                                        "Number of starts before next drop": None,
                                        "% of starts made while rostered": None,
                                        "Injury adjusted % of starts made while rostered": None,
                                        "Date dropped/traded": None,
                                        "Tanking": None,
                                        "Number of times picked up by this team": None,
                                    })
                            except Exception:
                                pass
                except Exception as e:
                    _log_exc(debug, f"transactions_trades_rows_{season}_wk{wk}", e)

    # Authoritative commissioner-move detection — RERUN now that every
    # trade transaction across all seasons has been folded into
    # pick_trade_events. The per-season calls above ran before each
    # season's own trades were recorded, so they over-flagged ordinary
    # traded picks (snapshot owner present, but the chain not yet built).
    # Clear those provisional entries and rebuild from the complete ledger.
    commissioner_pick_moves.clear()
    for _cm_season in sorted(traded_picks_by_season.keys()):
        try:
            _detect_commissioner_moves(int(_cm_season))
        except Exception as e:
            _log_exc(debug, f"detect_commissioner_moves_final_{_cm_season}", e)

    # Ensure pick ledger includes seasons from 2021 through three years after latest draft.
    latest_draft_season = max(
        [r.get("draft_season") for r in draft_picks_records if r.get("draft_season") is not None],
        default=None,
    )
    latest_league_season = max(roster_ids_by_season.keys(), default=None)
    base_season = latest_draft_season or latest_league_season
    if base_season is not None:
        max_future_season = int(base_season) + 3
        seed_source = max(draft_rounds_by_season.keys(), default=int(base_season))
        for yr in range(2021, max_future_season + 1):
            _ensure_pick_bases(int(yr), int(seed_source))

    # =====================================================================
    # Pick History — full rebuild from primary data sources.
    #
    # Approach (three steps):
    #   1) Build the canonical 288-pick frame: for every Sleeper draft we
    #      know about (rookie + 2021 supplemental vet) and every future
    #      rookie draft through max_future_season, enumerate all
    #      (round, slot) pairs and assign the slot's original owner via
    #      slot_to_roster_id (or, for future drafts, "every roster owns
    #      its own pick in each round").
    #   2) Walk each pick's trade history by chaining traded_picks events
    #      (combined across all season snapshots, since Sleeper removes
    #      used picks from each year's snapshot once that draft completes).
    #   3) Emit one row per pick into pick_rows. The earlier per-season
    #      pick_rows accumulation is discarded — primary data is the
    #      single source of truth here.
    # =====================================================================
    pick_rows = []

    # Earliest-week rosterer per (year, player_id) — the team that first held
    # a player that season. For a drafted rookie this is the team that drafted
    # them (they join via the draft, not a transaction). Used to repair the
    # 2021 rookie draft, whose EVEN-round (snake) picks have corrupted
    # roster_id / picked_by in Sleeper's data: Sleeper failed to track those
    # picks' trades and defaulted the picker to the slot's original owner, so
    # e.g. Trey Sermon (actually drafted by BROsenzweig) was attributed to
    # plehv79. The empirical roster ledger is the source of truth.
    _first_team_by_pid_year: Dict[Tuple[int, str], str] = {}
    _first_week_by_pid_year: Dict[Tuple[int, str], int] = {}
    for _pwr in player_week_rows:
        _pid = _pwr.get("Player ID")
        _yr = _to_int(_pwr.get("Year"), None)
        _wk = _to_int(_pwr.get("Week"), None)
        _tm = _pwr.get("Team")
        if _pid is None or _yr is None or _wk is None or not _tm:
            continue
        _k = (int(_yr), str(_pid))
        if _k not in _first_week_by_pid_year or _wk < _first_week_by_pid_year[_k]:
            _first_week_by_pid_year[_k] = int(_wk)
            _first_team_by_pid_year[_k] = str(_tm)

    try:
        # ---- Step 1: collect draft frames -----------------------------
        # Real Sleeper drafts: group picks by draft_id, capture slot_map,
        # max_round, vet flag.
        drafts_by_id: Dict[str, Dict[str, Any]] = {}
        for _season_key, _picks_for_season in season_draft_picks_all.items():
            for _p in _picks_for_season or []:
                _did = str(_p.get("draft_id") or "")
                if not _did:
                    continue
                if _did not in drafts_by_id:
                    drafts_by_id[_did] = {
                        "year": int(_season_key),
                        "is_vet": bool(_p.get("_is_vet_draft")),
                        "is_snake": (str(_p.get("_draft_type") or "").strip().lower() == "snake"),
                        "reversal_round": int(_to_int(_p.get("_reversal_round"), 0) or 0),
                        "slot_map": {},
                        "max_round": 0,
                        "picks": [],
                    }
                _info = drafts_by_id[_did]
                _sm = _p.get("slot_to_roster_id") if isinstance(_p.get("slot_to_roster_id"), dict) else None
                if _sm and not _info["slot_map"]:
                    _safe_sm: Dict[int, int] = {}
                    for _k, _v in _sm.items():
                        _kk = _to_int(_k, None)
                        _vv = _to_int(_v, None)
                        if _kk is not None and _vv is not None:
                            _safe_sm[_kk] = _vv
                    if _safe_sm:
                        _info["slot_map"] = _safe_sm
                _rnd = _to_int(_p.get("round"), None)
                if _rnd is not None:
                    _info["max_round"] = max(_info["max_round"], int(_rnd))
                _info["picks"].append(_p)

        draft_frames: List[Dict[str, Any]] = []
        for _did, _info in drafts_by_id.items():
            if _info["max_round"] <= 0:
                continue
            # If slot_map is missing/partial, synthesize one from the
            # league's roster ids for that year (stable canonical ordering).
            _smap = dict(_info["slot_map"]) if _info["slot_map"] else {}
            if not _smap:
                _rids_for_year = roster_ids_by_season.get(_info["year"]) or []
                if _rids_for_year:
                    _smap = {i + 1: int(rid) for i, rid in enumerate(sorted(_rids_for_year))}
            if not _smap:
                continue
            draft_frames.append({
                "year": int(_info["year"]),
                "is_vet": bool(_info["is_vet"]),
                "is_snake": bool(_info.get("is_snake")),
                "reversal_round": int(_info.get("reversal_round") or 0),
                "slot_map": _smap,
                "real_picks": _info["picks"],
                "max_round": min(int(_info["max_round"]), 4),  # cap at 4 for rookie/supplemental drafts
            })

        # Future-year frames: every roster owns one pick per round per
        # upcoming year, default slot assignment 1..N = roster_ids in
        # canonical sort order. Stops at base_season + 3.
        # NB: scan season_roster_to_team (only processed seasons) for
        # the fallback year — roster_ids_by_season also contains
        # synthetic future-season entries created by _ensure_pick_bases,
        # which would pick a future year that has no team-name map.
        default_year: Optional[int] = None
        default_rosters: List[int] = []
        for _y in sorted(season_roster_to_team.keys(), reverse=True):
            _rids = roster_ids_by_season.get(_y) or list(season_roster_to_team.get(_y, {}).keys())
            if _rids:
                default_rosters = list(_rids)
                default_year = int(_y)
                break
        if default_rosters and base_season is not None:
            default_slot_map = {i + 1: int(rid) for i, rid in enumerate(sorted(default_rosters))}
            _latest_real_year = max(
                (_f["year"] for _f in draft_frames if not _f["is_vet"]),
                default=int(latest_draft_season) if latest_draft_season is not None else int(base_season),
            )
            _max_fy = int(base_season) + 3
            for _yr in range(int(_latest_real_year) + 1, int(_max_fy) + 1):
                draft_frames.append({
                    "year": int(_yr),
                    "is_vet": False,
                    "slot_map": dict(default_slot_map),
                    "real_picks": [],
                    "max_round": 4,
                })

        # ---- Step 2: chain-by-origin from trade transactions ----------
        # Walk pick_trade_events, which is built from the actual Sleeper
        # transactions ledger (each entry recorded when a trade
        # involving this pick is processed). Goal: capture every time
        # a pick changed hands.
        #
        # Each event is (created_date, prev_owner, new_owner, week).
        # Sort chronologically by date, then build the chain by
        # appending each new_owner that differs from the running tail.
        #
        # commissioner_pick_moves layers on top: picks whose final
        # ownership in Sleeper's snapshot doesn't match the transaction
        # ledger (typical for off-platform / commissioner-executed
        # moves). For those we model a single hop from origin to the
        # snapshot owner, matching the rule "treat as one trade from
        # original to current".
        _chain_by_origin: Dict[Tuple[int, int, int], List[int]] = {}

        for _key, _events in (pick_trade_events or {}).items():
            if not _events:
                continue
            try:
                _sorted = sorted(
                    _events,
                    key=lambda _e: (_e[0] is None, _e[0] or datetime.min.replace(tzinfo=timezone.utc)),
                )
            except Exception:
                _sorted = list(_events)
            _chain = [int(_key[2])]
            for _e in _sorted:
                try:
                    _new = int(_e[2])
                except Exception:
                    continue
                if _chain[-1] != _new:
                    _chain.append(_new)
            if len(_chain) > 1:
                _chain_by_origin[_key] = _chain

        # Layer commissioner moves (off-platform single-hops).
        for _cm_key, _cm_new in (commissioner_pick_moves or {}).items():
            if _cm_key in _chain_by_origin:
                continue
            try:
                if int(_cm_new) != int(_cm_key[2]):
                    _chain_by_origin[_cm_key] = [int(_cm_key[2]), int(_cm_new)]
            except Exception:
                continue

        # ---- Step 3: emit one row per (frame, round, slot) ------------
        # Player-name resolution helper.
        def _resolve_player_name(_p: Dict[str, Any]) -> str:
            _player_id = _p.get("player_id")
            _name = pid_meta.get(str(_player_id), {}).get("full_name") if _player_id else None
            if not _name:
                _md = _p.get("metadata") if isinstance(_p.get("metadata"), dict) else {}
                _fn = (_md.get("first_name") or "").strip()
                _ln = (_md.get("last_name") or "").strip()
                if _fn or _ln:
                    _name = (f"{_fn} {_ln}").strip()
            if not _name:
                _name = _p.get("player") or _p.get("player_name")
            return str(_name) if _name else "Unknown"

        # Manual drafter corrections for the 2021 rookie draft. Sleeper's draft
        # data is doubly wrong here: it is snake-encoded AND it never reflected
        # the draft-DAY pick trades, so the recorded drafter (roster_id) for a
        # few picks can't be recovered automatically. Each correction was
        # established by tracing the rookie's first real transaction back to its
        # origin team (start-to-finish career check). Keyed by (round, position)
        # -> correct drafting team. NONE is a commissioner move: three are self-
        # drafts (the owner drafted, then traded the PLAYER away) and one is a
        # recorded draft-day PICK trade:
        #   2.05 Elijah Moore      -> plehv79    (plehv drafted; player traded)
        #   4.05 Nico Collins      -> plehv79    (plehv drafted; player traded)
        #   4.03 Dyami Brown       -> AceMatthew (AceMatthew drafted; player traded)
        #   3.06 Rhamondre Stevenson -> shmuel256 (shmuel got the pick in the
        #        2.08-for-3.06 + 2022-4th draft-day swap; players never moved)
        _DRAFTER_FIX_2021: Dict[Tuple[int, int], str] = {
            (2, 5): "plehv79",
            (4, 5): "plehv79",
            (4, 3): "AceMatthew",
            (3, 6): "shmuel256",
        }

        for _frame in draft_frames:
            _year = int(_frame["year"])
            _is_vet = bool(_frame["is_vet"])
            _slot_map = _frame["slot_map"]
            _real_picks = _frame["real_picks"]
            _max_rnd = min(int(_frame.get("max_round") or 4), 4)
            _team_count = len(_slot_map) or 8
            _is_snake = bool(_frame.get("is_snake"))
            # This league runs LINEAR rookie drafts: the team picking Pth in a
            # round OWNS pick N.0P, so original ownership follows the pick NUMBER
            # (position), and each team appears exactly once per round. Sleeper
            # mislabeled the 2021 ROOKIE draft as snake (the 2021 startup/vet
            # draft is genuinely snake), which left the original owner reading off
            # the reversed even-round draft_slot. Treat any non-vet snake draft
            # as linear for ORIGINAL-TEAM-by-position only — the displayed pick
            # number + drafted player (keyed by Sleeper's draft_slot) are already
            # correct and must NOT change.
            _force_linear = _is_snake and not _is_vet

            # In a snake draft the pick ORDER reverses on even rounds (the team
            # at draft_slot 1 picks last in round 2, first in round 3). draft_slot
            # stays constant per team, so the player keyed by draft_slot and the
            # displayed pick NUMBER (by draft order) are correct as-is.
            def _pick_position(_round: int, _draft_slot: int) -> int:
                if _is_snake and (int(_round) % 2 == 0):
                    return int(_team_count) + 1 - int(_draft_slot)
                return int(_draft_slot)

            # Index real selections by (round, slot) → (player_name, picker_rid).
            # picker_rid is Sleeper's `roster_id` on a draft pick — the team
            # that actually drafted the player (i.e. the true final owner
            # of the pick at draft time). For completed drafts this is
            # the source of truth and overrides the chain end below.
            _real_by_slot: Dict[Tuple[int, int], Tuple[str, Optional[int]]] = {}
            for _p in _real_picks:
                _rnd = _to_int(_p.get("round"), None)
                _slot = _to_int(_p.get("draft_slot"), None) or _to_int(_p.get("pick_in_round"), None)
                if _slot is None:
                    _pn = _to_int(_p.get("pick_no"), None)
                    if _pn is not None and _team_count:
                        _slot = ((int(_pn) - 1) % int(_team_count)) + 1
                if _rnd is None or _slot is None:
                    continue
                _picker_rid = _to_int(_p.get("roster_id"), None)
                _real_by_slot[(int(_rnd), int(_slot))] = (
                    _resolve_player_name(_p), _picker_rid, _p.get("player_id"),
                )

            # Year roster→team map. Future years fall back to the most-recent
            # known season's map so display names work for 2026-2028 rows.
            _rid_to_team = season_roster_to_team.get(_year, {}) or season_roster_to_team.get(default_year or _year, {})

            for _rnd in range(1, _max_rnd + 1):
                # Iterate in pick-ORDER position (1 = first pick of the
                # round) so rows emit in true draft order. _pick_position
                # is an involution, so it maps order position -> draft_slot
                # too (snake reverses even rounds; linear is identity).
                for _pos in range(1, _team_count + 1):
                    _si = _pick_position(int(_rnd), int(_pos))
                    # Original owner = the team at this pick's NUMBER position in
                    # the round (linear). For a genuine linear draft _pos == _si,
                    # so this is unchanged; for the snake-encoded 2021 rookie
                    # draft it corrects the reversed even-round owner (e.g. the
                    # first pick of round 2 belongs to slot 1, not slot N).
                    _ori = _to_int(_slot_map.get(_pos if _force_linear else _si), None)
                    if _si is None or _ori is None:
                        continue

                    _info = _real_by_slot.get((_rnd, _si))
                    if _info is None:
                        _player, _picker_rid, _player_id = "Unknown", None, None
                    else:
                        _player, _picker_rid, _player_id = _info

                    # Walk the chain (vet picks aren't traded by Sleeper —
                    # treat as a single-step chain).
                    _chain = [_ori]
                    if not _is_vet:
                        _chain = list(_chain_by_origin.get((_year, _rnd, _ori), [_ori]))
                        if not _chain:
                            _chain = [_ori]
                    # Owners reachable through RECORDED trades (origin + every
                    # ledger hop). Used below to tell a real (tracked) trade from
                    # an untracked move on the linearized rookie draft.
                    _ledger_owners = set(int(_o) for _o in _chain)

                    # If we know the actual picker (completed draft) and it
                    # doesn't match the chain end, append it. Sleeper's
                    # transactions ledger occasionally misses the final hop
                    # of a multi-trade pick (off-platform reassignments,
                    # late offseason trades not yet ingested), but the
                    # draft pick's roster_id is authoritative — that's who
                    # actually made the selection.
                    if _picker_rid is not None and _chain[-1] != int(_picker_rid):
                        _chain.append(int(_picker_rid))

                    # Repair corrupted even-round picks of the 2021 rookie
                    # snake draft. Sleeper mis-attributed those picks to the
                    # slot's original owner (it didn't track the pick trade),
                    # so the drafter is wrong. The team that actually first
                    # rostered the drafted player IS the drafter — override
                    # to it (and collapse the chain, since the slot-based
                    # origin is unreliable for these picks). Only fires when
                    # the roster ledger disagrees with Sleeper, so untouched
                    # / correctly-traded picks (e.g. DeVonta Smith) are left
                    # alone with their real origin->drafter chain intact.
                    if _is_snake and not _is_vet and (int(_rnd) % 2 == 0) and _player_id:
                        _rost_team = _first_team_by_pid_year.get((int(_year), str(_player_id)))
                        if _rost_team:
                            _rost_rid = season_team_to_roster.get(int(_year), {}).get(_norm_team_name(_rost_team))
                            if _rost_rid is not None and int(_rost_rid) != int(_chain[-1]):
                                # The slot's ORIGINAL owner (_ori) is correct;
                                # only the DRAFTER was corrupted by Sleeper. Set
                                # the final owner to the real (empirical) drafter
                                # and keep original -> drafter as a chain hop, so
                                # Original Team stays the slot's linear owner.
                                _chain = ([int(_ori)] if int(_rost_rid) == int(_ori)
                                          else [int(_ori), int(_rost_rid)])

                    # Manual 2021-rookie drafter corrections (see _DRAFTER_FIX_2021):
                    # override the chain's final owner with the true drafter and
                    # mark it non-commish (self-draft or recorded draft-day pick
                    # trade). Keeps the player's career traceable from its drafter.
                    _manual_fix = False
                    if _force_linear and int(_year) == 2021:
                        _fix_team = _DRAFTER_FIX_2021.get((int(_rnd), int(_pos)))
                        if _fix_team is not None:
                            _fix_rid = season_team_to_roster.get(int(_year), {}).get(_norm_team_name(_fix_team))
                            if _fix_rid is not None:
                                _chain = ([int(_ori)] if int(_fix_rid) == int(_ori)
                                          else [int(_ori), int(_fix_rid)])
                                _manual_fix = True

                    _final_rid = int(_chain[-1])
                    _orig_team = _rid_to_team.get(_ori, f"Roster {_ori}")
                    _final_team = _rid_to_team.get(_final_rid, f"Roster {_final_rid}")

                    # Commissioner-move determination. The standard detector
                    # (_detect_commissioner_moves) only sees UN-drafted future
                    # picks in Sleeper's traded_picks snapshot, so it can't catch
                    # the linearized 2021 rookie picks (already drafted). For
                    # those, a pick whose final owner (drafter) differs from its
                    # original slot owner and was NOT reached through a recorded
                    # trade (picker_rid append / empirical-roster repair) is an
                    # off-platform startup move = commissioner move. A genuinely
                    # traded pick (e.g. 2.08 J. Fields, shmuel256->LWebs53, in the
                    # trade ledger) lands in _ledger_owners and stays un-flagged.
                    # If the position owner MADE the selection themselves
                    # (Sleeper's raw picker == owner), the pick never moved — any
                    # later change of hands is a PLAYER trade, not a pick move
                    # (e.g. 4.03 Dyami Brown: AceMatthew drafted their own pick,
                    # then traded the player to plehv79). Don't flag those.
                    _owner_drafted = (_picker_rid is not None and int(_picker_rid) == int(_ori))
                    _untracked_move = bool(
                        _force_linear
                        and not _owner_drafted
                        and not _manual_fix
                        and _final_rid != int(_ori)
                        and _final_rid not in _ledger_owners
                    )
                    _commish = (
                        False if _is_vet
                        else (((_year, _rnd, _ori) in commissioner_pick_moves) or _untracked_move)
                    )

                    _row: Dict[str, Any] = {
                        "Year": (f"{_year} (vet)" if _is_vet else _year),
                        "Number": _format_pick_number(int(_rnd), int(_pos)),
                        "Original Team": _orig_team,
                        "Final Team": _final_team,
                        "Player Picked": _player,
                        # Internal (dropped from output by _ensure_plan_columns):
                        # the Sleeper player_id of the drafted player, so the
                        # picks PPG/KTC passes resolve by ID (robust to name
                        # suffixes like "III" and duplicate names) instead of by
                        # display name.
                        "_player_id": (str(_player_id) if _player_id else None),                        "Commissioner moved?": _commish,
                    }
                    # Trade 1..N from intermediate + final owners. Emit
                    # as many columns as the chain has hops — output
                    # writer extends the schema to match the longest
                    # chain across all picks.
                    for _j, _ownr in enumerate(_chain[1:], start=1):
                        _row[f"Trade {_j}"] = _rid_to_team.get(int(_ownr), f"Roster {_ownr}")
                    pick_rows.append(_row)
    except Exception as e:
        _log_exc(debug, "pick_history_rebuild", e)

    # ----- Phase 10: inject synthetic draft-day picks (2.09 + 5.0X) -----
    # Built from the commissioner-forced adds captured per season. They carry a
    # _player_id, so they flow through every downstream pick stat pass (PPG /
    # KTC / addition value / pick-adjusted / O-Score) exactly like real picks.
    # 2.09: original team = prior-season toilet-bracket winner, final team = the
    #   roster the player was force-added to, one synthetic trade hop if they
    #   differ. 5.0X: original = final = the buyer, no trade.
    try:
        def _pid_name(_pid: Any) -> str:
            return pid_meta.get(str(_pid), {}).get("full_name") or str(_pid)
        for _sea, _adds in sorted(draft_day_commish_adds.items()):
            if int(_sea) < 2024 or not _adds:
                continue
            _rid2tm = season_roster_to_team.get(int(_sea), {})
            # 2.09 (toilet-bracket reward)
            _ts0, _rid0, _pid0, _txid0 = _adds[0]
            _final0 = _rid2tm.get(_rid0, f"Roster {_rid0}")
            _tw_rid = toilet_winner_by_season.get(int(_sea) - 1)
            _orig0 = _rid2tm.get(_tw_rid, _final0) if _tw_rid is not None else _final0
            _row209: Dict[str, Any] = {
                "Year": int(_sea),
                "Number": "2.09",
                "Original Team": _orig0,
                "Final Team": _final0,
                "Player Picked": _pid_name(_pid0),
                "_player_id": str(_pid0),                "Commissioner moved?": False,
            }
            if _orig0 != _final0:
                _row209["Trade 1"] = _final0
            pick_rows.append(_row209)
            # 5.0X (FAAB draft-day buys) — only from 2025; chronological order.
            if int(_sea) >= 2025:
                for _j, (_tsj, _ridj, _pidj, _txidj) in enumerate(_adds[1:], start=1):
                    _tmj = _rid2tm.get(_ridj, f"Roster {_ridj}")
                    pick_rows.append({
                        "Year": int(_sea),
                        "Number": f"5.{_j:02d}",
                        "Original Team": _tmj,
                        "Final Team": _tmj,
                        "Player Picked": _pid_name(_pidj),
                        "_player_id": str(_pidj),                        "Commissioner moved?": False,
                    })
    except Exception as e:
        _log_exc(debug, "synthetic_draft_day_picks", e)

    # --------------------------
    # Convert to DataFrames
    # --------------------------
    pw = pd.DataFrame(player_week_rows)

    if not pw.empty and "Team" in pw.columns:
        pw["_team_canon"] = pw["Team"].apply(_norm_team_name)
        canon_to_disp = {}
        for t in pw["Team"].dropna().astype(str).tolist():
            c=_norm_team_name(t)
            if c and c not in canon_to_disp:
                canon_to_disp[c]=t
        pw["Team"] = pw["_team_canon"].map(canon_to_disp).fillna(pw["Team"])

    # --------------------------------------------------------------
    # Unique-player position counts (Phase 1B, item 5).
    # Build lookup dicts so team_year / team_all_time / league_year /
    # league_all_time can report DISTINCT QBs/WRs/RBs/TEs started or
    # rostered over the period — not the sum of weekly counts (which
    # double-counts a QB who started 5 weeks as "5 QBs started").
    # Runs AFTER pw Team canonicalization above so the dict keys use
    # the same display-name form that team_year groupby will lookup
    # against. Earlier placement saw pre-canonical names and the
    # (Team, Year) lookups all missed.
    # --------------------------------------------------------------
    def _build_unique_position_counts(pw_df: pd.DataFrame, group_cols: List[str]) -> Dict[Tuple, Dict[str, int]]:
        out: Dict[Tuple, Dict[str, int]] = {}
        if pw_df.empty or "Player ID" not in pw_df.columns or "Position" not in pw_df.columns:
            return out
        df = pw_df[pw_df["Player ID"].notna() & pw_df["Position"].notna()].copy()
        df["_pos"] = df["Position"].astype(str).str.upper().str.strip()
        df["_starter"] = df.get("Starter/Bench", "").astype(str).str.lower() == "starter"
        for col in group_cols:
            if col == "Year":
                df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64").astype(object)
                df[col] = df[col].map(lambda x: int(x) if pd.notna(x) else None)
            elif col == "Team":
                df[col] = df[col].astype(str)
            else:
                df[col] = df[col].astype(object)
        df = df.dropna(subset=group_cols)
        for pos in ["QB", "WR", "RB", "TE"]:
            pos_df = df[df["_pos"] == pos]
            if pos_df.empty:
                continue
            rostered = pos_df.groupby(group_cols)["Player ID"].nunique()
            for key, val in rostered.items():
                k = key if isinstance(key, tuple) else (key,)
                out.setdefault(k, {})[f"Number of {pos} rostered"] = int(val)
            started = pos_df[pos_df["_starter"]].groupby(group_cols)["Player ID"].nunique()
            for key, val in started.items():
                k = key if isinstance(key, tuple) else (key,)
                out.setdefault(k, {})[f"Number of {pos} started"] = int(val)
        return out

    unique_pos_by_team_year: Dict[Tuple, Dict[str, int]] = {}
    unique_pos_by_team_all: Dict[Tuple, Dict[str, int]] = {}
    unique_pos_by_year: Dict[Tuple, Dict[str, int]] = {}
    unique_pos_league_all: Dict[str, int] = {}
    if not pw.empty:
        try:
            unique_pos_by_team_year = _build_unique_position_counts(pw, ["Team", "Year"])
            unique_pos_by_team_all  = _build_unique_position_counts(pw, ["Team"])
            unique_pos_by_year      = _build_unique_position_counts(pw, ["Year"])
            _all = _build_unique_position_counts(pw.assign(_all="all"), ["_all"])
            unique_pos_league_all = _all.get(("all",), {})
        except Exception as e:
            _log_exc(debug, "unique_position_counts", e)

    if not pw.empty and "Team" in pw.columns:
        pw.drop(columns=["_team_canon"], inplace=True, errors="ignore")

    # Phase 3A item 5: Number of trades per player-week.
    # Each trades_row has a Date and an "Assets received" string listing
    # the players (and pick labels — skip those) the row's Team gained
    # in the trade. Counting once per received player per trade matches
    # the player_trade_year aggregation downstream.
    player_trade_week: Dict[Tuple[str, int, int], int] = defaultdict(int)
    try:
        # Build a name -> sid resolver. Prefer the longest match (some
        # pid_meta entries share the same full_name; we just need any
        # one valid sid since trade counts aren't per-sleeper-id).
        _name_to_sid_for_trades: Dict[str, str] = {}
        for _sid, _meta in pid_meta.items():
            _fn = (_meta or {}).get("full_name")
            if _fn:
                _name_to_sid_for_trades.setdefault(str(_fn), str(_sid))

        def _trade_week_for_date(_dt_str: Optional[str], _season: int) -> int:
            # Returns the in-season week a trade falls in, or 0 for a deep-
            # offseason trade that should NOT roll into any weekly bucket
            # (Phase 5C item 9: offseason trades count in Week 1's WEEKLY rollup
            # only if within 7 days before kickoff). Year/all-time totals are
            # counted separately and still include these trades.
            if not _dt_str:
                return 1
            try:
                _d = datetime.fromisoformat(str(_dt_str).replace("Z", "+00:00")).date()
                _season_start = date(int(_season), 9, 7)
                if _d < _season_start:
                    return 1 if (_season_start - _d).days <= 7 else 0
                _wk = (_d - _season_start).days // 7 + 1
                return max(1, min(17, int(_wk)))
            except Exception:
                return 1

        for _tr_row in trades_rows:
            _yr_val = _tr_row.get("Season")
            _date_s = _tr_row.get("Date")
            if _yr_val is None and _date_s:
                try:
                    _yr_val = int(str(_date_s)[:4])
                except Exception:
                    continue
            try:
                _yr_i = int(_yr_val)
            except Exception:
                continue
            _wk_i = _trade_week_for_date(_date_s, _yr_i)
            _recv = _tr_row.get("Assets received")
            if not _recv or str(_recv) == "0.0":
                continue
            for _asset in str(_recv).split(";"):
                _asset = _asset.strip()
                if not _asset:
                    continue
                # Skip pick labels ('2025 1.05(B. Robinson)' etc).
                if re.match(r"^\d{4}\b", _asset):
                    continue
                _sid = _name_to_sid_for_trades.get(_asset)
                if not _sid:
                    continue
                player_trade_week[(str(_sid), _yr_i, _wk_i)] += 1
    except Exception as e:
        _log_exc(debug, "player_trade_week_build", e)

    if not pw.empty and {"Player ID", "Year", "Week"}.issubset(pw.columns):
        pw_keys = pw[["Player ID", "Year", "Week"]].copy()
        pw_keys["Player ID"] = pw_keys["Player ID"].astype(str)
        pw_keys["Year"] = pd.to_numeric(pw_keys["Year"], errors="coerce").astype("Int64")
        pw_keys["Week"] = pd.to_numeric(pw_keys["Week"], errors="coerce").astype("Int64")
        pw["Number of transactions"] = [
            int(player_tx_week.get(
                (
                    str(player_id),
                    int(year) if pd.notna(year) else None,
                    int(week) if pd.notna(week) else None,
                ),
                0,
            ))
            for player_id, year, week in pw_keys.itertuples(index=False, name=None)
        ]
        pw["Number of drops"] = [
            int(player_drop_week.get(
                (
                    str(player_id),
                    int(year) if pd.notna(year) else None,
                    int(week) if pd.notna(week) else None,
                ),
                0,
            ))
            for player_id, year, week in pw_keys.itertuples(index=False, name=None)
        ]
        pw["Number of trades"] = [
            int(player_trade_week.get(
                (
                    str(player_id),
                    int(year) if pd.notna(year) else None,
                    int(week) if pd.notna(week) else None,
                ),
                0,
            ))
            for player_id, year, week in pw_keys.itertuples(index=False, name=None)
        ]
    log_df(pw, 'player_week', sample_cols=['Points','Injury?','Suspension?','Bye?','Starter?'])
    tw = pd.DataFrame(team_week_rows)

    # Normalize team names (case-insensitive) across seasons so joins don't duplicate teams like 'Shmuel256' vs 'shmuel256'.
    if not tw.empty and "Team" in tw.columns:
        tw["_team_canon"] = tw["Team"].apply(_norm_team_name)
        canon_to_disp: Dict[str,str] = {}
        for t in tw["Team"].dropna().astype(str).tolist():
            c=_norm_team_name(t)
            if c and c not in canon_to_disp:
                canon_to_disp[c]=t
        tw["Team"] = tw["_team_canon"].map(canon_to_disp).fillna(tw["Team"])
        tw.drop(columns=["_team_canon"], inplace=True, errors="ignore")
    if not tw.empty:
        log_missing_cols(tw, "team_week", [
            "Year", "Week", "Team", "PF", "Points against", "Margin", "Max PF", "Efficiency"
        ])
        zero_max = int((pd.to_numeric(tw.get("Max PF"), errors="coerce").fillna(0) <= 0).sum())
        LOG.info("team_week: rows=%s zero_max_pf=%s", len(tw), zero_max)
    log_df(tw, 'team_week', sample_cols=['PF','Max PF','Efficiency'])

    # Override rookie counts to ensure unique rookie IDs per team-week (v3).
    try:
        if not pw.empty and {"Team", "Year", "Week", "Player ID", "Starter/Bench", "Rookie?"}.issubset(pw.columns):
            rookies = pw[["Team", "Year", "Week", "Player ID", "Starter/Bench", "Rookie?"]].copy()
            rookies = rookies.dropna(subset=["Team", "Year", "Week", "Player ID"])
            rookies["Player ID"] = rookies["Player ID"].astype(str)
            rookies["Starter/Bench"] = rookies["Starter/Bench"].astype(str)
            rookies["Rookie?"] = pd.to_numeric(rookies["Rookie?"], errors="coerce").fillna(0).astype(int)
            rookies = rookies[rookies["Rookie?"] == 1]

            rookies_rostered = (
                rookies.groupby(["Team", "Year", "Week"])["Player ID"]
                .nunique()
                .reset_index()
                .rename(columns={"Player ID": "Number of rookies rostered"})
            )
            rookies_started = (
                rookies[rookies["Starter/Bench"].str.lower().eq("starter")]
                .groupby(["Team", "Year", "Week"])["Player ID"]
                .nunique()
                .reset_index()
                .rename(columns={"Player ID": "Number of rookies started"})
            )

            if not tw.empty:
                tw = tw.drop(columns=["Number of rookies started", "Number of rookies rostered"], errors="ignore")
                tw = tw.merge(rookies_rostered, on=["Team", "Year", "Week"], how="left")
                tw = tw.merge(rookies_started, on=["Team", "Year", "Week"], how="left")
                tw["Number of rookies rostered"] = (
                    pd.to_numeric(tw.get("Number of rookies rostered"), errors="coerce").fillna(0).astype(int)
                )
                tw["Number of rookies started"] = (
                    pd.to_numeric(tw.get("Number of rookies started"), errors="coerce").fillna(0).astype(int)
                )
    except Exception as e:
        _log_exc(debug, "team_week_rookie_unique_counts", e)

    # ---- Tanking (user formula)
    # Tanking(team, week) uses season-to-date averages through that week.
    # Tanking(team, year) is the final week value.
    def _safe_div(n, d):
        try:
            d = float(d)
            n = float(n)
            if d == 0 or (pd.isna(d) or pd.isna(n)):
                return 0.0
            return n / d
        except Exception:
            return 0.0

    def _tanking_score(avg_pf, avg_max_pf, avg_age, league_avg_pf, league_avg_max_pf, league_avg_age, pick_sum, future_cap):
        # 1/6 *(1 - (AvgPF-2/3 L)/(L-2/3 L))
        denom1 = (league_avg_pf - (2.0/3.0)*league_avg_pf)  # L/3
        term1 = 1.0 - _safe_div((avg_pf - (2.0/3.0)*league_avg_pf), denom1)

        # 1/6 *(1 - (AvgMaxPF-LPF)/(LMaxPF-LPF))
        denom2 = (league_avg_max_pf - league_avg_pf)
        term2 = 1.0 - _safe_div((avg_max_pf - league_avg_pf), denom2)

        # 1/6 *(1 - (AvgAge-21)/(LAvgAge-21))
        denom3 = (league_avg_age - 21.0)
        term3 = 1.0 - _safe_div((avg_age - 21.0), denom3)

        # 1/6 *(sum picks value)
        term4 = float(pick_sum or 0.0)

        # 1/9 *(future draft capital weights)
        term5 = float(future_cap or 0.0)

        return (1.0/6.0)*term1 + (1.0/6.0)*term2 + (1.0/6.0)*term3 + (1.0/6.0)*term4 + (1.0/9.0)*term5

    # (Future draft capital is now computed by _future_cap_held, defined
    # near _future_picks_owned above. The old _future_cap_from_traded
    # helper was removed — it undercounted by ignoring un-traded own picks.)

    # Per-season pick value (that year's draft picks)
    pick_value_by_team_season = {}
    for season_key, picks_for_season in season_draft_picks_all.items():
        season_map = season_roster_to_team.get(int(season_key), {})
        for p in picks_for_season or []:
            try:
                y = _to_int(p.get("draft_season"), season_key)
                if y is None:
                    continue
                rid = _to_int(p.get("roster_id"), None)
                if rid is None:
                    continue
                team = season_map.get(rid)
                if not team:
                    continue
                pick_no = p.get("pick_no")
                if pick_no is None:
                    # fall back: approximate overall pick if not provided
                    rnd = _to_int(p.get("round"), 0)
                    slot = _to_int(p.get("draft_slot"), 0) or _to_int(p.get("pick_in_round"), 0)
                    if rnd and slot:
                        pick_no = (rnd - 1) * max(1, len(season_map)) + slot
                pick_no = _to_int(pick_no, None)
                if pick_no is None:
                    continue
                # rookie drafts only (heuristic): ignore huge pick numbers
                if pick_no > 1000:
                    continue
                val = 1.0 / (float(pick_no) + 1.0)
                pick_value_by_team_season[(str(team), int(y))] = pick_value_by_team_season.get((str(team), int(y)), 0.0) + val
            except Exception:
                continue

    # Team-week tanking: compute season-to-date
    if not tw.empty:
        tw["Tanking"] = pd.to_numeric(tw.get("Tanking"), errors="coerce")

        # Weekly team age average for tanking. Prefer the
        # 'Team age including picks' column we now emit per
        # (team, year, week) — that includes held draft capital,
        # which is the right denominator for the tank-detection
        # heuristic (a team accumulating picks IS tanking, not just
        # one running an old roster). Fall back to the old
        # rostered-player-only average when the new column isn't
        # available for some reason.
        if "Team age including picks" in tw.columns:
            tw2 = tw.copy()
            tw2["TeamWeekAvgAge"] = pd.to_numeric(tw2["Team age including picks"], errors="coerce")
            # Fill missing pick-inclusive ages with the player-only
            # average so the per-week tank score still has signal.
            if "Age" in pw.columns:
                fallback = pw.groupby(["Team", "Year", "Week"], dropna=False)["Age"].mean().reset_index()
                fallback.rename(columns={"Age": "_TeamWeekAvgAgeFallback"}, inplace=True)
                tw2 = tw2.merge(fallback, on=["Team", "Year", "Week"], how="left")
                tw2["TeamWeekAvgAge"] = tw2["TeamWeekAvgAge"].fillna(
                    pd.to_numeric(tw2["_TeamWeekAvgAgeFallback"], errors="coerce")
                )
        elif "Age" in pw.columns:
            age_week = pw.groupby(["Team", "Year", "Week"], dropna=False)["Age"].mean().reset_index()
            age_week.rename(columns={"Age": "TeamWeekAvgAge"}, inplace=True)
            tw2 = tw.merge(age_week, on=["Team", "Year", "Week"], how="left")
            tw2["TeamWeekAvgAge"] = pd.to_numeric(tw2["TeamWeekAvgAge"], errors="coerce")
        else:
            tw2 = tw.copy()
            tw2["TeamWeekAvgAge"] = pd.NA

        tanking_rows = []
        for season in sorted(tw2["Year"].dropna().unique()):
            g = tw2[tw2["Year"] == season].copy()
            if g.empty:
                continue

            # league averages season-to-date by week (equal-weight per week)
            g_sorted = g.sort_values("Week").copy()
            # ensure numeric to avoid object-mean failures
            g_sorted["PF"] = pd.to_numeric(g_sorted.get("PF"), errors="coerce")
            g_sorted["Max PF"] = pd.to_numeric(g_sorted.get("Max PF"), errors="coerce")
            g_sorted["TeamWeekAvgAge"] = pd.to_numeric(g_sorted.get("TeamWeekAvgAge"), errors="coerce")

            pf_week = g_sorted.groupby("Week")["PF"].mean().sort_index()
            maxpf_week = g_sorted.groupby("Week")["Max PF"].mean().sort_index()
            age_week_lg = g_sorted.groupby("Week")["TeamWeekAvgAge"].mean().sort_index()

            league_avg_pf_upto = pf_week.expanding().mean()
            league_avg_maxpf_upto = maxpf_week.expanding().mean()
            league_avg_age_upto = age_week_lg.expanding().mean()

            for team, tg in g.groupby("Team"):
                tg = tg.sort_values("Week").copy()
                # expanding means
                pf_exp = pd.to_numeric(tg.get("PF"), errors="coerce").expanding().mean()
                maxpf_exp = pd.to_numeric(tg.get("Max PF"), errors="coerce").expanding().mean()

                # age expanding: use weekly mean age
                age_exp = pd.to_numeric(tg.get("TeamWeekAvgAge"), errors="coerce").expanding().mean()

                rid = season_team_to_roster.get(int(season), {}).get(_norm_team_name(team))
                pick_sum = pick_value_by_team_season.get((str(team), int(season)), 0.0)
                # End-of-season holdings (own retained + acquired − traded away);
                # corrected helper replaces _future_cap_from_traded which omitted
                # un-traded own picks. Feb 1 anchor = after Week 18, before the
                # next rookie draft, so all in-season pick trades are reflected.
                future_cap = _future_cap_held(str(team), int(season), date(int(season) + 1, 2, 1))

                for i, row in tg.reset_index(drop=True).iterrows():
                    # Week can be missing/NaN in some corrupt rows; guard to avoid crashes.
                    try:
                        wk_int = int(row["Week"]) if pd.notna(row["Week"]) else None
                    except Exception:
                        wk_int = None

                    lg_pf = league_avg_pf_upto.iloc[-1] if len(league_avg_pf_upto) else 0.0
                    lg_mx = league_avg_maxpf_upto.iloc[-1] if len(league_avg_maxpf_upto) else 0.0
                    lg_ag = league_avg_age_upto.iloc[-1] if len(league_avg_age_upto) else 0.0
                    if wk_int is not None:
                        lg_pf = league_avg_pf_upto.get(wk_int, lg_pf)
                        lg_mx = league_avg_maxpf_upto.get(wk_int, lg_mx)
                        lg_ag = league_avg_age_upto.get(wk_int, lg_ag)

                    score = _tanking_score(
                        avg_pf=pf_exp.iloc[i],
                        avg_max_pf=maxpf_exp.iloc[i],
                        avg_age=age_exp.iloc[i],
                        league_avg_pf=lg_pf,
                        league_avg_max_pf=lg_mx,
                        league_avg_age=lg_ag,
                        pick_sum=pick_sum,
                        future_cap=future_cap,
                    )
                    tanking_rows.append((row["Team"], row["Year"], row["Week"], float(score)))

        if tanking_rows:
            tank_df = pd.DataFrame(tanking_rows, columns=["Team", "Year", "Week", "Tanking"])
            tw = tw.drop(columns=["Tanking"], errors="ignore").merge(tank_df, on=["Team", "Year", "Week"], how="left")
        else:
            tw["Tanking"] = 0.0

    
    # ---- Week Name propagation (custom week naming)
    # Use team_week's Week Name where available.
    if (not tw.empty) and ("Week Name" in tw.columns):
        # player_week
        if "Week Name" not in pw.columns:
            pw = pw.merge(tw[["Team","Year","Week","Week Name","Round"]].drop_duplicates(), on=["Team","Year","Week"], how="left")

        # derive a league-wide Week Name per (Year,Week) for league_week rollups
        def _mode_nonnull(vals):
            vals = [v for v in vals if isinstance(v, str) and v and v != "N/A"]
            if not vals:
                return None
            # prefer non-generic labels
            nongeneric = [v for v in vals if not v.startswith("Week ")]
            base = nongeneric if nongeneric else vals
            return pd.Series(base).mode().iloc[0] if len(base) else None

        week_name_global = tw.groupby(["Year","Week"])["Week Name"].apply(_mode_nonnull).reset_index()
    else:
        week_name_global = pd.DataFrame(columns=["Year","Week","Week Name"])

    # --------------------------
    # Merge manual transactions overrides. Sleeper's transactions API
    # occasionally omits real pickups (we've seen one case: Puka Nacua's
    # 2023-09-04 add to Shmuel256, confirmed via Week 1 matchup rosters
    # but missing from the transactions endpoint across every leg).
    # data/manual_transactions.csv lets us drop in those rows by hand.
    # We append them BEFORE the polish + KTC passes so they participate
    # in pickup-counter, drop-date, link-prev/next, and KTC enrichment.
    # --------------------------
    try:
        manual_path = repo_root / "data" / "manual_transactions.csv"
        if manual_path.exists():
            mdf = pd.read_csv(manual_path)
            # Build a quick name -> sleeper_id lookup so KTC enrichment
            # can resolve added/dropped players downstream.
            name_to_sid: Dict[str, str] = {}
            for sid, meta in pid_meta.items():
                fn = (meta or {}).get("full_name")
                if fn:
                    name_to_sid.setdefault(str(fn), str(sid))
            # Build the canonical set of team names so we can refuse to
            # emit manual rows whose Team doesn't match. A typo silently
            # showed up as a phantom team in an earlier iteration.
            canonical_teams: Set[str] = set()
            if not tw.empty and "Team" in tw.columns:
                canonical_teams = {str(t) for t in tw["Team"].dropna().unique()}

            n_added = 0
            for _, mrow in mdf.iterrows():
                added_name = (str(mrow.get("Player Added") or "").strip() or None)
                dropped_name = (str(mrow.get("Player Dropped") or "").strip() or None)
                if added_name in ("", "nan"): added_name = None
                if dropped_name in ("", "nan"): dropped_name = None
                if not added_name and not dropped_name:
                    continue
                # Sanity-check the Date — a corrupted CSV (e.g.
                # unquoted commas in Notes) can shift every field one
                # column left, which surfaced once as Puka Nacua's row
                # claiming Date=2023 and Faab="waiver". Fail loud.
                try:
                    datetime.fromisoformat(str(mrow.get("Date")).replace("Z","+00:00"))
                except Exception:
                    _log(debug, f"[{_now_iso()}] WARN manual transactions: skipping row with bad Date {mrow.get('Date')!r} (CSV likely malformed)")
                    continue
                team_val = str(mrow.get("Team") or "").strip()
                if canonical_teams and team_val not in canonical_teams:
                    _log(debug, f"[{_now_iso()}] WARN manual transactions: Team {team_val!r} doesn't match any canonical team — skipping")
                    continue
                added_pid = name_to_sid.get(added_name) if added_name else None
                dropped_pid = name_to_sid.get(dropped_name) if dropped_name else None
                row: Dict[str, Any] = {
                    "Team": str(mrow.get("Team")),
                    "Player Added": added_name,
                    "Player Dropped": dropped_name,
                    "type of transaction (waiver/free agency)": str(mrow.get("Type") or "free_agent"),
                    "Faab": _to_float(mrow.get("Faab"), None),
                    "Total FAAB bid": _to_float(mrow.get("Total FAAB bid"), None),
                    "FAAB difference over second place": None,
                    "FAAB premium %": None,
                    "Date": str(mrow.get("Date")),
                    "Season": int(mrow.get("Season")) if pd.notna(mrow.get("Season")) else None,
                    "_added_pid": added_pid,
                    "_dropped_pid": dropped_pid,
                    "Number of bids": _to_float(mrow.get("Number of bids"), None),
                    "Link to next transaction": None,
                    "Link to previous transaction": None,
                    "Average PPG on team": None,
                    "Average PPG of dropped player over same time": None,
                    "Difference of averages": None,
                    "Difference of averages adjusted by position": None,
                    "Age difference": None,
                    "Player addition value": None,
                    "Cuff at time of pickup?": None,
                    "Weeks between pickup and start": None,
                    "Number of starts before next drop": None,
                    "% of starts made while rostered": None,
                    "Injury adjusted % of starts made while rostered": None,
                    "Date dropped/traded": None,
                    "Tanking": None,
                    "Number of times picked up by this team": None,
                }
                transactions_rows.append(row)
                # Also credit per-week/year/all counters so player
                # rollups stay consistent with the manual entries.
                try:
                    season_i = int(mrow.get("Season"))
                except Exception:
                    season_i = None
                if added_pid and season_i is not None:
                    player_tx_year[(str(added_pid), season_i)] += 1
                    player_tx_all[str(added_pid)] += 1
                if dropped_pid and season_i is not None:
                    player_tx_year[(str(dropped_pid), season_i)] += 1
                    player_tx_all[str(dropped_pid)] += 1
                    player_drop_year[(str(dropped_pid), season_i)] += 1
                    player_drop_all[str(dropped_pid)] += 1
                n_added += 1
            if n_added:
                _log(debug, f"[{_now_iso()}] INFO merged {n_added} manual transaction(s) from data/manual_transactions.csv")
            # Bump team_week / team_year counters so the rollup tables
            # reflect the manual rows we just added. Find the Year+Week
            # rows that match each manual row's Date and increment.
            if n_added and not tw.empty and "Year" in tw.columns and "Week" in tw.columns:
                from datetime import date as _dt_date
                # NFL week 1 starts the first Thursday of September. For
                # the dates we care about, a simple approximation is
                # enough: week N starts roughly Sep 7 + 7*(N-1) of that
                # year. Pre-season / week 1 prep dates map to week 1.
                def _week_for_date(dstr: str, season: int):
                    try:
                        d = datetime.fromisoformat(str(dstr).replace("Z","+00:00")).date()
                    except Exception:
                        return None
                    season_start = _dt_date(int(season), 9, 5)
                    if d < season_start:
                        return 1
                    diff = (d - season_start).days // 7 + 1
                    return min(max(1, diff), 17)

                for _, mrow in mdf.iterrows():
                    season = mrow.get("Season")
                    if pd.isna(season):
                        continue
                    season = int(season)
                    wk = _week_for_date(str(mrow.get("Date")), season)
                    team = str(mrow.get("Team"))
                    if not wk:
                        continue
                    mask = (tw["Team"]==team) & (tw["Year"]==season) & (tw["Week"]==wk)
                    matches = tw[mask]
                    if matches.empty:
                        continue
                    idx_ = matches.index[0]
                    cur = pd.to_numeric(tw.at[idx_, "Number of transactions"], errors="coerce")
                    tw.at[idx_, "Number of transactions"] = int((0 if pd.isna(cur) else cur) + 1)
                    faab = _to_float(mrow.get("Faab"), 0.0) or 0.0
                    if faab:
                        cur_f = pd.to_numeric(tw.at[idx_, "Amount of FAAB spent"], errors="coerce")
                        tw.at[idx_, "Amount of FAAB spent"] = round((0.0 if pd.isna(cur_f) else cur_f) + faab, 2)
    except Exception as e:
        _log_exc(debug, "manual_transactions_merge", e)

    # --------------------------
    # transactions_rows post-processing: fill polish columns now that the
    # full history is available.
    #
    #  - Number of times picked up by this team: running per-(team, player)
    #    pickup count, chronological.
    #  - Date dropped/traded: for each pickup, the next date this team
    #    dropped or traded away this player (None if still rostered).
    #  - Weeks between pickup and start: count of player_week rows for
    #    (Team, Player) that occur after the pickup date but before the
    #    player's first start on that team. None if never started.
    # --------------------------
    try:
        if transactions_rows:
            # Sort by date so the running counters scan chronologically.
            def _date_key(r):
                d = r.get("Date") or ""
                return str(d)
            transactions_rows.sort(key=_date_key)

            # (Source-level dedup of Sleeper transactions now runs in the
            # per-week fetch loop, so transactions_rows are already deduped
            # by the time we get here.)

            # 1) Number of times picked up / dropped by this team (Phase 6C).
            #    Both now INCLUDE trades: a player received in a trade counts as
            #    a pickup, a player traded away counts as a drop. We interleave
            #    transaction adds/drops with trade in/out events chronologically
            #    so the running count emitted on a transaction row reflects every
            #    prior acquisition/departure (incl. trades) of that player.
            def _split_assets(_s):
                for _a in str(_s or "").split(";"):
                    _a = _a.strip()
                    if (_a and _a not in ("0.0", "None", "N/A")
                            and not re.match(r"^\d{4}\b", _a) and not _a.endswith("FAAB")):
                        yield _a

            acq_events = []  # (date, team, player, row_or_None)
            drop_events = []
            for r in transactions_rows:
                team = r.get("Team")
                if not team:
                    continue
                d = str(r.get("Date") or "")
                add = r.get("Player Added")
                if add and str(add) != "N/A":
                    acq_events.append((d, str(team), str(add), r))
                dropped = r.get("Player Dropped")
                if dropped and str(dropped) != "N/A":
                    drop_events.append((d, str(team), str(dropped), r))
            for tr_row in trades_rows:
                team = tr_row.get("Team")
                if not team:
                    continue
                d = str(tr_row.get("Date") or "")
                for asset in _split_assets(tr_row.get("Assets received")):
                    acq_events.append((d, str(team), asset, None))
                for asset in _split_assets(tr_row.get("Assets sent")):
                    drop_events.append((d, str(team), asset, None))

            pickup_count: Dict[Tuple[str, str], int] = defaultdict(int)
            for d, team, player, row in sorted(acq_events, key=lambda e: e[0]):
                pickup_count[(team, player)] += 1
                if row is not None:
                    row["Number of times picked up by this team"] = pickup_count[(team, player)]
            drop_count: Dict[Tuple[str, str], int] = defaultdict(int)
            for d, team, player, row in sorted(drop_events, key=lambda e: e[0]):
                drop_count[(team, player)] += 1
                if row is not None:
                    row["Number of times dropped by this team"] = drop_count[(team, player)]

            # 2) Build per-(team, player) event log to find next drop/trade-out.
            # NOTE: don't name a local 'date' here — that would shadow the
            # imported datetime.date class for the entire build_all function
            # (Python decides locals at compile time), which broke player_week
            # construction earlier with an UnboundLocalError. Use dt_str.
            event_log: Dict[Tuple[str, str], List[Tuple[str, str]]] = defaultdict(list)
            for r in transactions_rows:
                team = r.get("Team")
                add = r.get("Player Added")
                dropped = r.get("Player Dropped")
                dt_str = r.get("Date") or ""
                if team and add:
                    event_log[(str(team), str(add))].append((str(dt_str), "add"))
                if team and dropped:
                    event_log[(str(team), str(dropped))].append((str(dt_str), "drop"))
            # Players traded away count as "left this team" too.
            for tr_row in trades_rows:
                team = tr_row.get("Team")
                dt_str = tr_row.get("Date") or ""
                dropped_assets = str(tr_row.get("Assets sent") or "")
                if dropped_assets in ("0.0", "None", ""):
                    continue
                for asset in dropped_assets.split(";"):
                    asset = asset.strip()
                    if not asset or re.match(r"^\d{4}\b", asset) or asset.endswith("FAAB"):
                        continue
                    if team:
                        event_log[(str(team), asset)].append((str(dt_str), "trade_out"))
            # Orphan drops (player dropped without a same-transaction add) —
            # these never made it into transactions_rows but they are real
            # departures and must close the event window for the prior pickup.
            for od in orphan_drop_events:
                t_od = od.get("Team")
                p_od = od.get("Player Dropped")
                d_od = od.get("Date") or ""
                if t_od and p_od and d_od:
                    event_log[(str(t_od), str(p_od))].append((str(d_od), "drop"))
            for k in event_log:
                event_log[k].sort()

            for r in transactions_rows:
                team = r.get("Team")
                add = r.get("Player Added")
                add_date = r.get("Date") or ""
                if not (team and add and add_date):
                    continue
                # Next departure event strictly after add_date.
                next_evt = None
                for ev_date, ev_type in event_log.get((str(team), str(add)), []):
                    if ev_type == "add":
                        continue
                    if ev_date and ev_date > add_date:
                        next_evt = ev_date
                        break
                if next_evt:
                    r["Date dropped/traded"] = next_evt

            # 3) Weeks between pickup and start. Needs pw which exists at this
            # point in build_all. Map (Team, Player Name) -> sorted list of
            # (Year, Week, Starter?) rows. For each pickup, count player_week
            # rows that fall between pickup date and first start on that team.
            if not pw.empty and {"Team", "Player", "Year", "Week", "Starter/Bench"}.issubset(set(pw.columns)):
                # Approximate fantasy-week date: NFL week 1 starts ~Sept 7,
                # subsequent weeks each Thursday after. Good enough for an
                # "is this player_week row after the pickup date" gate.
                def _approx_week_date(year: int, week: int) -> str:
                    try:
                        d = date(int(year), 9, 7) + timedelta(days=7 * (int(week) - 1))
                        return d.isoformat()
                    except Exception:
                        return ""

                pw_min = pw[["Team", "Player", "Year", "Week", "Starter/Bench"]].copy()
                pw_min = pw_min.sort_values(["Team", "Player", "Year", "Week"]).reset_index(drop=True)
                # Bucket by (team, player) for fast lookup
                pw_by_tp: Dict[Tuple[str, str], List[Tuple[int, int, bool, str]]] = defaultdict(list)
                for _, prow in pw_min.iterrows():
                    try:
                        yr = int(prow["Year"]); wk = int(prow["Week"])
                    except Exception:
                        continue
                    started = str(prow["Starter/Bench"]) == "Starter"
                    pw_by_tp[(str(prow["Team"]), str(prow["Player"]))].append(
                        (yr, wk, started, _approx_week_date(yr, wk))
                    )

                for r in transactions_rows:
                    team = r.get("Team")
                    add = r.get("Player Added")
                    add_date = r.get("Date") or ""
                    if not (team and add and add_date):
                        continue
                    rows = pw_by_tp.get((str(team), str(add)))
                    if not rows:
                        continue
                    # Counting restarts on each pickup. If the player was
                    # dropped/traded away after this pickup, the bench window
                    # ends at that drop. So bound the search at the NEXT
                    # departure for the same (team, player) after add_date.
                    drop_after = r.get("Date dropped/traded") or ""
                    weeks_before_start = 0
                    found_start = False
                    for yr, wk, started, wk_date in rows:
                        if wk_date and wk_date < add_date:
                            continue
                        # Stop counting once the player was let go again.
                        if drop_after and wk_date and wk_date >= drop_after:
                            break
                        if started:
                            found_start = True
                            break
                        weeks_before_start += 1
                    if found_start:
                        r["Weeks between pickup and start"] = weeks_before_start
    except Exception as e:
        _log_exc(debug, "transactions_polish", e)

    # --------------------------
    # transactions_rows polish pass 2: PPG-derived columns, age diff,
    # cuff detection, post-pickup start-rate metrics, and composite
    # 'Player addition value'.
    #
    # Definitions (per league owner):
    #   Avg PPG (added/dropped): mean of points scored in the last 5
    #     played games BEFORE the pickup date, regardless of fantasy
    #     team. Bye and injury weeks are not "played games" — skip them.
    #   Difference of averages: added_avg - dropped_avg
    #   Difference adjusted by position:
    #     added_adj  = added_avg * all_starter_avg / pos_avg[added_pos]
    #     dropped_adj = dropped_avg * all_starter_avg / pos_avg[dropped_pos]
    #     adjusted_diff = added_adj - dropped_adj
    #     (normalises positions to a common scale)
    #   Cuff at time of pickup?: another STARTER on the picking team
    #     at the pickup week shares NFL team + position with the added
    #     player AND averaged 10+ PPG more in last 5 played games.
    #   Number of starts before next drop: count of pw rows for
    #     (Team, Player Added) with Starter/Bench=='Starter' that fall
    #     between Date and Date dropped/traded.
    #   % of starts made while rostered: starts / weeks_rostered.
    #   Injury adjusted version: same, but exclude Bye? and Injury?
    #     rows from BOTH numerator and denominator.
    #   Player addition value:
    #     adjusted_diff * (1 + pct_starts) * (1 + pct_starts_inj_adj)
    #       + CUFF_BONUS  (only added when 'Cuff at time of pickup?'=True)
    # --------------------------
    try:
        if transactions_rows and not pw.empty:
            from datetime import date as _date_cls
            # --- precompute per-player game logs across the whole pw ---
            pw_min_p = pw[["Player", "Team", "Year", "Week", "Points",
                           "Position", "NFL team", "Starter/Bench",
                           "Injury?", "Bye?"]].copy()

            def _approx_week_date2(year, week):
                try:
                    d = _date_cls(int(year), 9, 7) + timedelta(days=7 * (int(week) - 1))
                    return d.isoformat()
                except Exception:
                    return ""

            pw_min_p["_wk_date"] = [
                _approx_week_date2(y, w) for y, w in zip(pw_min_p["Year"], pw_min_p["Week"])
            ]

            # Index by player (regardless of fantasy team) for the
            # "last 5 played games before date" lookup.
            pw_by_player: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
            # Index by (team, week, year) for the cuff lookup — gives
            # us the picking team's roster snapshot at the pickup week.
            pw_by_team_week: Dict[Tuple[str, int, int], List[Dict[str, Any]]] = defaultdict(list)
            # Index by (team, player) for the start-rate walk.
            pw_by_team_player: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)

            for _, prow in pw_min_p.iterrows():
                d = {
                    "Player": str(prow.get("Player") or ""),
                    "Team": str(prow.get("Team") or ""),
                    "Year": int(prow["Year"]) if pd.notna(prow["Year"]) else None,
                    "Week": int(prow["Week"]) if pd.notna(prow["Week"]) else None,
                    "Points": float(prow["Points"]) if pd.notna(prow["Points"]) else 0.0,
                    "Position": str(prow.get("Position") or "").upper(),
                    "NFL team": str(prow.get("NFL team") or ""),
                    "Starter/Bench": str(prow.get("Starter/Bench") or ""),
                    "Injury?": bool(prow.get("Injury?")),
                    "Bye?": bool(prow.get("Bye?")),
                    "_wk_date": str(prow.get("_wk_date") or ""),
                }
                pw_by_player[d["Player"]].append(d)
                if d["Year"] is not None and d["Week"] is not None:
                    pw_by_team_week[(d["Team"], d["Year"], d["Week"])].append(d)
                pw_by_team_player[(d["Team"], d["Player"])].append(d)

            # All-time starter averages for the position adjustment.
            starters = pw[pw["Starter/Bench"] == "Starter"]
            all_starter_avg = float(starters["Points"].mean()) if not starters.empty else 0.0
            pos_avg: Dict[str, float] = {}
            if not starters.empty and "Position" in starters.columns:
                pos_avg = {
                    str(k).upper(): float(v)
                    for k, v in starters.groupby("Position")["Points"].mean().to_dict().items()
                }

            def _pos_adjust(ppg: Optional[float], pos: Optional[str]) -> Optional[float]:
                if ppg is None:
                    return None
                p = pos_avg.get((pos or "").upper())
                if not p or p <= 0 or not all_starter_avg:
                    return ppg
                return round(ppg * all_starter_avg / p, 4)

            # Build name -> sleeper_id once so we can hit nfl_log_by_sid
            # (which is keyed by Sleeper id) from a player display name.
            name_to_sid_local: Dict[str, str] = {}
            for sid, meta in pid_meta.items():
                fn = (meta or {}).get("full_name")
                if fn:
                    name_to_sid_local.setdefault(str(fn), str(sid))

            def _player_games(player_name: Optional[str]) -> List[Dict[str, Any]]:
                """Return the player's NFL game log: list of dicts with
                _wk_date and Points. Prefer nflverse (covers all NFL
                weeks regardless of fantasy roster status); fall back
                to pw (filtering out bye / injury / suspension rows
                since those aren't 'played' for our purposes)."""
                if not player_name:
                    return []
                sid = name_to_sid_local.get(player_name)
                games: List[Dict[str, Any]] = []
                if sid:
                    for entry in nfl_log_by_sid.get(sid, []):
                        if entry["_wk_date"]:
                            games.append({"_wk_date": entry["_wk_date"], "Points": entry["points"]})
                if not games:
                    for r in pw_by_player.get(player_name, []):
                        if r["_wk_date"] and not r["Bye?"] and not r["Injury?"] and not r.get("Suspension?"):
                            games.append({"_wk_date": r["_wk_date"], "Points": r["Points"]})
                return games

            def _avg_ppg_last5_before(player_name: Optional[str], pickup_date_iso: str) -> Optional[float]:
                """Mean fantasy points over the 5 most-recent played
                games BEFORE the pickup date. If fewer than 5 games
                exist on record, averages whatever's available (e.g.,
                2 games -> mean of 2). Used for cuff detection and the
                'PPG of 5 games before pickup' column."""
                games = _player_games(player_name)
                played = [g for g in games if g["_wk_date"] < pickup_date_iso]
                if not played:
                    return None
                played.sort(key=lambda g: g["_wk_date"], reverse=True)
                window = played[:5]
                return round(sum(g["Points"] for g in window) / len(window), 4)

            def _avg_ppg_in_window(
                player_name: Optional[str],
                start_iso: str,
                end_iso: Optional[str],
            ) -> Optional[float]:
                """Mean fantasy points across NFL games in [start, end).
                end_iso=None means 'no upper bound' — counts through
                end of dataset. Used for the forward-looking 'Average
                PPG on team' and 'over same time' columns: the window
                is the time the added player was on the picking team."""
                games = _player_games(player_name)
                in_window: List[float] = []
                for g in games:
                    if not g["_wk_date"]:
                        continue
                    if g["_wk_date"] < start_iso:
                        continue
                    if end_iso and g["_wk_date"] >= end_iso:
                        continue
                    in_window.append(g["Points"])
                if not in_window:
                    return None
                return round(sum(in_window) / len(in_window), 4)

            def _player_pos_nfl_team_at(player_name: Optional[str], pickup_date_iso: str) -> Tuple[Optional[str], Optional[str]]:
                """Best-effort position+NFL team for the player AS OF the
                pickup date (latest pw row before that)."""
                if not player_name:
                    return None, None
                rows = pw_by_player.get(player_name, [])
                before = [r for r in rows if r["_wk_date"] and r["_wk_date"] <= pickup_date_iso]
                if not before:
                    # fall back to first pw row
                    if rows:
                        r = rows[0]
                        return (r["Position"] or None), (r["NFL team"] or None)
                    return None, None
                before.sort(key=lambda r: r["_wk_date"], reverse=True)
                r = before[0]
                return (r["Position"] or None), (r["NFL team"] or None)

            CUFF_BONUS = 5.0  # PPG-equivalent bonus when player is a cuff at pickup

            for r in transactions_rows:
                team = r.get("Team")
                added = r.get("Player Added")
                dropped = r.get("Player Dropped")
                pickup_date = r.get("Date") or ""
                # Date is an ISO string at this stage; convert for comparison.
                pickup_iso = str(pickup_date)
                # Pull a YYYY-MM-DD prefix for week-date string comparison.
                pickup_iso_prefix = pickup_iso[:10] if len(pickup_iso) >= 10 else pickup_iso

                # Compute the player's tenure window on this team.
                # 'Average PPG on team' and 'over same time' are
                # forward-looking: from pickup date to the next
                # drop/trade of the added player (or open-ended if
                # the player is still rostered).
                drop_after_iso = str(r.get("Date dropped/traded") or "")
                drop_after_prefix = drop_after_iso[:10] if len(drop_after_iso) >= 10 else (
                    drop_after_iso or None
                )

                # --- Avg PPG family ---
                # Forward-looking: how did the added player do on this
                # team, and what was the dropped player doing in NFL
                # over the same window?
                added_on_team = (
                    _avg_ppg_in_window(added, pickup_iso_prefix, drop_after_prefix)
                    if added else None
                )
                dropped_same_window = (
                    _avg_ppg_in_window(dropped, pickup_iso_prefix, drop_after_prefix)
                    if dropped else None
                )
                # Pre-pickup snapshot: trailing 5-game average. Useful
                # for evaluating the pickup decision (what did the
                # market know about this guy at the time).
                added_pre5 = _avg_ppg_last5_before(added, pickup_iso_prefix) if added else None
                # Pre-pickup average for the dropped player too (used
                # in the cuff comparison; still computed once).
                dropped_pre5 = _avg_ppg_last5_before(dropped, pickup_iso_prefix) if dropped else None

                if added_on_team is not None:
                    r["Average PPG on team"] = added_on_team
                if dropped_same_window is not None:
                    r["Average PPG of dropped player over same time"] = dropped_same_window
                if added_pre5 is not None:
                    r["PPG of 5 games before pickup"] = added_pre5
                if added_on_team is not None or dropped_same_window is not None:
                    r["Difference of averages"] = round(
                        (added_on_team or 0.0) - (dropped_same_window or 0.0), 4
                    )

                # --- Position adjustment ---
                # Position adjustment now uses the forward-looking
                # tenure averages (matches user's clarification:
                # 'adjust adjusted averages to match this').
                added_pos, added_nfl = _player_pos_nfl_team_at(added, pickup_iso_prefix)
                dropped_pos, dropped_nfl = _player_pos_nfl_team_at(dropped, pickup_iso_prefix)
                added_adj = _pos_adjust(added_on_team, added_pos)
                dropped_adj = _pos_adjust(dropped_same_window, dropped_pos)
                adj_diff = None
                if added_adj is not None or dropped_adj is not None:
                    adj_diff = round((added_adj or 0.0) - (dropped_adj or 0.0), 4)
                    r["Difference of averages adjusted by position"] = adj_diff

                # --- Points Added / Lost / Net (+ per-week averages) ---
                # Points Added: the added player's fantasy points in the weeks
                # they STARTED for this team, from pickup until their next exit.
                # Points Lost: the dropped player's REAL NFL points (game log,
                # 0 for any DNP/bye) over exactly those same started weeks — the
                # opportunity cost of starting the add instead of the drop. The
                # averages divide by the number of started weeks so swaps of
                # different lengths are comparable. Pure drop (no add) -> all 0.
                def _is_name(_v):
                    _s = str(_v).strip()
                    return bool(_s) and _s.upper() != "N/A" and _s.lower() not in ("nan", "none")
                pts_added = 0.0
                started_yw: List[Tuple[int, int]] = []
                if _is_name(added):
                    for _d in pw_by_team_player.get((str(team), str(added)), []):
                        if _d.get("Starter/Bench") != "Starter":
                            continue
                        _wd = _d.get("_wk_date") or ""
                        if _wd < pickup_iso_prefix:
                            continue
                        if drop_after_prefix and _wd >= drop_after_prefix:
                            continue
                        if _d.get("Year") is None or _d.get("Week") is None:
                            continue
                        started_yw.append((int(_d["Year"]), int(_d["Week"])))
                        pts_added += float(_d.get("Points") or 0.0)
                pts_lost = 0.0
                if _is_name(dropped) and started_yw:
                    _dlog = {
                        (int(_e["year"]), int(_e["week"])): float(_e.get("points") or 0.0)
                        for _e in nfl_log_by_sid.get(str(r.get("_dropped_pid")), [])
                        if _e.get("year") is not None and _e.get("week") is not None
                    }
                    for _yw in started_yw:
                        pts_lost += _dlog.get(_yw, 0.0)
                _nwk = len(started_yw)
                _net = pts_added - pts_lost
                r["Points Added"] = round(pts_added, 2)
                r["Points Lost"] = round(pts_lost, 2)
                r["Net points"] = round(_net, 2)
                r["Avg points added"] = round(pts_added / _nwk, 2) if _nwk else 0.0
                r["Avg points lost"] = round(pts_lost / _nwk, 2) if _nwk else 0.0
                r["Avg net points"] = round(_net / _nwk, 2) if _nwk else 0.0
                # Position-adjusted variants (Item 4): scale each side by the
                # mover's position (added player for added, dropped for lost)
                # via the same league_starter_avg / pos_avg normalizer used by
                # 'Difference of averages adjusted by position'.
                _adj_add = _pos_adjust(pts_added, added_pos) if added_pos else pts_added
                _adj_lost = _pos_adjust(pts_lost, dropped_pos) if dropped_pos else pts_lost
                _adj_add = pts_added if _adj_add is None else _adj_add
                _adj_lost = pts_lost if _adj_lost is None else _adj_lost
                r["Avg points added adjusted by position"] = round(_adj_add / _nwk, 2) if _nwk else 0.0
                r["Avg points lost adjusted by position"] = round(_adj_lost / _nwk, 2) if _nwk else 0.0
                r["Avg net points adjusted by position"] = round((_adj_add - _adj_lost) / _nwk, 2) if _nwk else 0.0

                # --- Dropped avg / total points (next 17 PLAYED games) ---
                # The dropped player's realized NFL output after leaving, NEGATED
                # (points that walked out the door): his next 17 games actually
                # played per the nflverse log, from the drop date on. Dynamic —
                # fewer than 17 games so far averages/sums what exists; a player
                # who never played again scores a real 0 (the perfect drop).
                if _is_name(dropped):
                    _post = sorted(
                        (g for g in _player_games(dropped)
                         if g["_wk_date"] and g["_wk_date"] >= pickup_iso_prefix),
                        key=lambda g: g["_wk_date"],
                    )[:17]
                    _ptot = sum(float(g["Points"] or 0.0) for g in _post)
                    r["Dropped avg points"] = round(-_ptot / len(_post), 4) if _post else 0.0
                    r["Dropped total points"] = round(-_ptot, 2) if _post else 0.0

                # Length of tenure on team (added player): days from pickup to
                # the next exit (or to today if still rostered). Blank if no add.
                if _is_name(added):
                    try:
                        _pk_d = datetime.fromisoformat(pickup_iso_prefix).date()
                        _end_d = (datetime.fromisoformat(drop_after_prefix).date()
                                  if drop_after_prefix else datetime.utcnow().date())
                        r["Length of tenure on team"] = max(0, (_end_d - _pk_d).days)
                    except Exception:
                        pass

                # The cuff comparison still uses the pre-pickup 5-game
                # snapshot — that's the form info available at pickup.
                # Bind it under the original name the cuff code reads.
                added_avg = added_pre5

                # --- Age difference (added - dropped, years) ---
                def _age_for(name):
                    if not name:
                        return None
                    # Find this player's sleeper_id via pid_meta full_name
                    for sid, meta in pid_meta.items():
                        if (meta or {}).get("full_name") == name:
                            bd = meta.get("birth_date")
                            if bd:
                                try:
                                    pickup_d = datetime.fromisoformat(pickup_iso.replace("Z","+00:00")).date()
                                    born = dateparser.parse(str(bd)).date()
                                    return round((pickup_d - born).days / 365.25, 2)
                                except Exception:
                                    return None
                            return None
                    return None
                added_age = _age_for(added)
                dropped_age = _age_for(dropped)
                if added_age is not None or dropped_age is not None:
                    r["Age difference"] = round((added_age or 0.0) - (dropped_age or 0.0), 2)

                # --- Tanking-delta inputs (Phase 6E) ---
                # Marginal change this transaction makes to the team's roster
                # age (entities = players + picks). Only count sides whose age
                # is known, so the swap math stays consistent with how
                # "Team age including picks" averages known ages. Waiver/FA
                # transactions never move draft picks, so the future-capital
                # delta is 0. The late tanking-delta pass combines these with
                # the team-week roster age/size to produce the final score.
                r["_tank_recv_age_sum"] = float(added_age) if added_age is not None else 0.0
                r["_tank_recv_n"] = 1 if added_age is not None else 0
                r["_tank_sent_age_sum"] = float(dropped_age) if dropped_age is not None else 0.0
                r["_tank_sent_n"] = 1 if dropped_age is not None else 0
                r["_tank_fcap_delta"] = 0.0

                # --- Cuff at time of pickup? ---
                # Identify the pickup's NFL Year+Week (best effort) so
                # we can look at the picking team's roster that week.
                # The pw _wk_date approximation gives us a 7-day bucket
                # we can match against the pickup_iso.
                cuff = False
                if added and team and added_nfl and added_pos:
                    # Find the team's pw rows for the same week
                    candidate_team_rows: List[Dict[str, Any]] = []
                    for r2 in pw_by_team_week.values():
                        for entry in r2:
                            if entry["Team"] != team:
                                continue
                            # Pickup must fall within ~7 days of this week
                            if entry["_wk_date"] and entry["_wk_date"] <= pickup_iso_prefix:
                                if not candidate_team_rows or entry["_wk_date"] > candidate_team_rows[0]["_wk_date"]:
                                    candidate_team_rows = [entry]
                                elif entry["_wk_date"] == candidate_team_rows[0]["_wk_date"]:
                                    candidate_team_rows.append(entry)
                    # Item 11 (relaxed): the qualifying teammate need only have
                    # been a STARTER at some point in the previous 3 weeks — the
                    # pickup week and the two before it — not necessarily the
                    # exact pickup week. Catches handcuffs added right after the
                    # starter they back up goes down.
                    if candidate_team_rows:
                        yr = candidate_team_rows[0]["Year"]
                        wk = candidate_team_rows[0]["Week"]
                        # The reference (handcuff) player must STILL be rostered
                        # by the team at the pickup week — a teammate who was a
                        # starter two weeks ago but has since been dropped is no
                        # longer insurance you hold (Item 8).
                        pickup_roster = {
                            m["Player"] for m in pw_by_team_week.get((team, yr, wk), [])
                        }
                        for _w in (wk, wk - 1, wk - 2):
                            if _w is None or _w < 1:
                                continue
                            for mate in pw_by_team_week.get((team, yr, _w), []):
                                if mate["Player"] == added:
                                    continue
                                if mate["Player"] not in pickup_roster:
                                    continue
                                if mate["Starter/Bench"] != "Starter":
                                    continue
                                if mate["NFL team"] != added_nfl:
                                    continue
                                if mate["Position"] != added_pos:
                                    continue
                                # Check mate's last-5 PPG > added's last-5 PPG + 10
                                mate_avg = _avg_ppg_last5_before(mate["Player"], pickup_iso_prefix)
                                if mate_avg is not None and (added_avg or 0.0) + 10 <= mate_avg:
                                    cuff = True
                                    break
                            if cuff:
                                break
                r["Cuff at time of pickup?"] = bool(cuff)

                # --- Start-rate metrics after the pickup ---
                # Walk pw rows for (Team, Player Added) between Date and
                # Date dropped/traded. Count starts, weeks rostered,
                # injury-adjusted starts/weeks.
                if added and team:
                    drop_after = str(r.get("Date dropped/traded") or "")
                    drop_after_prefix = drop_after[:10] if len(drop_after) >= 10 else drop_after
                    weeks_played = 0
                    starts = 0
                    inj_weeks_played = 0
                    inj_starts = 0
                    for entry in pw_by_team_player.get((team, added), []):
                        wk_date = entry["_wk_date"]
                        if not wk_date or wk_date < pickup_iso_prefix:
                            continue
                        if drop_after_prefix and wk_date >= drop_after_prefix:
                            break
                        weeks_played += 1
                        if entry["Starter/Bench"] == "Starter":
                            starts += 1
                        if not entry["Bye?"] and not entry["Injury?"]:
                            inj_weeks_played += 1
                            if entry["Starter/Bench"] == "Starter":
                                inj_starts += 1
                    r["Number of starts before next drop"] = int(starts)
                    if weeks_played > 0:
                        r["% of starts made while rostered"] = round(starts / weeks_played, 4)
                    if inj_weeks_played > 0:
                        r["Injury adjusted % of starts made while rostered"] = round(inj_starts / inj_weeks_played, 4)

                # --- Player addition value composite ---
                # Only meaningful when we have the adjusted diff.
                if adj_diff is not None:
                    pct_starts = float(r.get("% of starts made while rostered") or 0.0)
                    pct_inj = float(r.get("Injury adjusted % of starts made while rostered") or 0.0)
                    cuff_bonus = CUFF_BONUS if cuff else 0.0
                    addition_val = adj_diff * (1.0 + pct_starts) * (1.0 + pct_inj) + cuff_bonus
                    r["Player addition value"] = round(addition_val, 4)
                elif added and team:
                    # An added player with no measurable PPG impact (no
                    # position-adjusted average and no dropped counterpart —
                    # e.g. never rostered a full week, or rostered but logged
                    # no fantasy points the whole tenure) added nothing → 0
                    # (not N/A), so the O-Score still scores them off net
                    # points / KTC. Pure drops (no added player) stay N/A.
                    r["Player addition value"] = 0.0
    except Exception as e:
        _log_exc(debug, "transactions_polish_v2", e)

    # --------------------------
    # KTC value pass — single pass that powers all KTC columns on both
    # trades.csv and transactions.csv. Data source: dynasty-daddy.com's
    # public API, which scrapes KeepTradeCut daily and exposes per-player
    # daily history back to April 2021. We use trade_value (1QB format).
    #
    # Reference points per trade (4):
    #   - deal time:       the trade date itself
    #   - end of season:   the Monday after (trade.Season)'s championship game
    #                      (the NEXT championship after the deal)
    #   - 1 year later:    exactly 1 calendar year after the deal date
    #   - 2 years later:   exactly 2 calendar years after the deal date
    # 'End of season' is anchored to the Monday after the season's fantasy
    # championship game (Phase 6F) rather than a fixed Jan-5, so it tracks the
    # actual season end. The 1- and 2-year columns are simply the deal date
    # plus 1/2 calendar years — a fixed horizon from the move itself.
    #
    # Per transaction row: KTC of added, dropped, and net at deal time.
    # --------------------------
    try:
        from lotg_support.ktc import build_index, asset_value_at
        today = datetime.utcnow().date()

        # Monday after a season's championship game. Sleeper playoffs end at
        # NFL week 17 (2021+); NFL week 1's Sunday is 6 days after Labor Day
        # (the first Monday of September), the championship Sunday is 16 weeks
        # later, and the snapshot Monday is the day after. e.g. 2021 -> Jan 3
        # 2022, 2023 -> Jan 1 2024, 2024 -> Dec 30 2024.
        def _championship_monday(season_year: int) -> date:
            sept1 = date(int(season_year), 9, 1)
            first_monday = sept1 + timedelta(days=(7 - sept1.weekday()) % 7)
            week1_sunday = first_monday + timedelta(days=6)
            return week1_sunday + timedelta(weeks=16) + timedelta(days=1)

        # Exactly N calendar years after a date (Feb 29 -> Feb 28 in non-leap
        # years). Used for the '1 year later' / '2 years later' KTC references,
        # which are a fixed horizon from the deal date itself.
        def _plus_years(d: date, n: int) -> date:
            try:
                return d.replace(year=d.year + n)
            except ValueError:
                return d.replace(year=d.year + n, day=28)

        # League-format detection so the KTC values we pull match the
        # user's setup. dynasty-daddy publishes two value series per
        # asset: trade_value (1QB) and sf_trade_value (superflex). A
        # superflex league should use sf_trade_value or QBs read way
        # too low and trade rollups read backwards.
        ktc_value_col = "trade_value"
        try:
            for lg in leagues or []:
                rp = [str(x).upper() for x in (lg.get("roster_positions") or [])]
                if any(p in {"SUPER_FLEX", "SUPERFLEX", "SFLEX", "SFLX"} for p in rp):
                    ktc_value_col = "sf_trade_value"
                    break
                if rp.count("QB") >= 2:
                    ktc_value_col = "sf_trade_value"
                    break
        except Exception:
            pass
        _log(debug, f"[{_now_iso()}] INFO ktc value column: {ktc_value_col}")

        # Collect every sleeper_id and pick label we'll need across both
        # detail tables. dynasty-daddy serves per-player histories one
        # API call at a time, so pre-fetching only what we use keeps the
        # build fast (the cache makes subsequent runs cheaper still).
        needed_sids: Set[str] = set()
        needed_picks: Set[str] = set()
        for row in trades_rows:
            for sid in (row.get("_recv_player_ids") or []):
                if sid:
                    needed_sids.add(str(sid))
            for sid in (row.get("_drop_player_ids") or []):
                if sid:
                    needed_sids.add(str(sid))
            for plabel in (row.get("_recv_picks") or []):
                if plabel:
                    needed_picks.add(str(plabel))
            for plabel in (row.get("_drop_picks") or []):
                if plabel:
                    needed_picks.add(str(plabel))
        for tx_row in transactions_rows:
            sid = tx_row.get("_added_pid")
            if sid:
                needed_sids.add(str(sid))
            sid = tx_row.get("_dropped_pid")
            if sid:
                needed_sids.add(str(sid))
        # Phase 8D: include every DRAFTED player so the KTC index also covers
        # the picks sheet's KTC-over-time columns. Resolve by the pick's Sleeper
        # player_id (threaded through pick_rows) — robust to name suffixes and
        # duplicate names — falling back to a name lookup only when the id is
        # absent.
        _ktc_name_to_sid: Dict[str, str] = {}
        for _ks, _km in pid_meta.items():
            _kfn = (_km or {}).get("full_name")
            if _kfn:
                _ktc_name_to_sid.setdefault(str(_kfn), str(_ks))
        for _prow in pick_rows:
            _psid = _prow.get("_player_id")
            if not _psid:
                _pname = str(_prow.get("Player Picked") or "").strip()
                if _pname and _pname.lower() not in ("unknown", "nan", "n/a", ""):
                    _psid = _ktc_name_to_sid.get(_pname)
            if _psid:
                needed_sids.add(str(_psid))

        # Pass full_name+pos for every sleeper_id we need so the KTC
        # module can derive name_id slugs for retired/aged-out players
        # not in dynasty-daddy's active directory.
        sid_to_meta = {
            str(sid): {
                "full_name": (pid_meta.get(str(sid)) or {}).get("full_name") or "",
                "pos": (pid_meta.get(str(sid)) or {}).get("pos") or "",
            }
            for sid in needed_sids
        }
        idx = build_index(
            repo_root,
            needed_sids,
            needed_picks,
            value_col=ktc_value_col,
            sid_to_meta=sid_to_meta,
        )
        # Capture under a stable name: `idx` is later reused as a loop variable
        # (the trades enumerate loop), so the picks-KTC pass uses _ktc_idx.
        _ktc_idx = idx

        def _side_total(
            target: date,
            player_ids: List[str],
            pick_labels: List[str],
        ) -> Tuple[float, int]:
            total = 0.0
            hits = 0
            for sid in player_ids:
                v = asset_value_at(None, str(sid), target, idx)
                if v is not None:
                    total += v
                    hits += 1
            for plabel in pick_labels:
                v = asset_value_at(str(plabel), None, target, idx)
                if v is not None:
                    total += v
                    hits += 1
            return total, hits

        def _side_values(
            target: date,
            player_ids: List[str],
            pick_labels: List[str],
            faab: float = 0.0,
        ) -> List[float]:
            """Per-asset KTC values on one side (for the package-tax diff). FAAB
            dollars are valued at the league-wide avg KTC-per-$ (Fix 3)."""
            out: List[float] = []
            for sid in player_ids:
                v = asset_value_at(None, str(sid), target, idx)
                if v is not None and v > 0:
                    out.append(float(v))
            for plabel in pick_labels:
                v = asset_value_at(str(plabel), None, target, idx)
                if v is not None and v > 0:
                    out.append(float(v))
            if faab and faab > 0 and _ktc_per_faab > 0:
                out.append(float(faab) * _ktc_per_faab)
            return out

        # --- Fix 3: average KTC value of $1 FAAB ---
        # From "similar trades": a side that is PURE FAAB exchanged for a side of
        # KTC-valued assets implies KTC-per-$ = (asset-side KTC) / FAAB$. Median
        # across all such clean FAAB-for-asset trades.
        _ktc_per_faab = 0.0
        try:
            _faab_ratios: List[float] = []
            for _r in trades_rows:
                _ds = _r.get("Date")
                try:
                    _dt = datetime.fromisoformat(str(_ds).replace("Z", "+00:00")).date()
                except Exception:
                    continue
                if _dt > today:
                    continue
                _rf = float(_r.get("_recv_faab") or 0.0)
                _df = float(_r.get("_drop_faab") or 0.0)
                _rv = _side_values(_dt, _r.get("_recv_player_ids") or [], _r.get("_recv_picks") or [])
                _sv = _side_values(_dt, _r.get("_drop_player_ids") or [], _r.get("_drop_picks") or [])
                if _rf > 0 and _df == 0 and _sv and not _rv:        # got FAAB for assets
                    _faab_ratios.append(sum(_sv) / _rf)
                elif _df > 0 and _rf == 0 and _rv and not _sv:      # gave FAAB for assets
                    _faab_ratios.append(sum(_rv) / _df)
            _est = 0.0
            if _faab_ratios:
                _faab_ratios.sort()
                _est = _faab_ratios[len(_faab_ratios) // 2]
            # The data-derived median (~329) over-values FAAB: it's skewed by a
            # few small-$ overpays. Use a flat 100 KTC per $1 FAAB — arbitrary
            # but a more accurate, conservative valuation. (Estimate still
            # logged for reference.)
            _ktc_per_faab = 100.0
            _log(debug, f"[{_now_iso()}] INFO KTC per $1 FAAB = {_ktc_per_faab:.0f} (fixed; data-derived median {_est:.1f} from {len(_faab_ratios)} FAAB-for-asset trades)")
        except Exception as e:
            _log_exc(debug, "ktc_per_faab", e)

        # --- Package-tax adjustment for uneven multi-asset trades (Item 2) ---
        # A naive Σreceived − Σsent over-values the side with MORE pieces: in
        # dynasty you can only start so many, so three scrubs are worth less than
        # their summed KTC. KTC's calculator reverse-engineers to a per-asset
        # "raw adjustment" (~10-42% of KTC) but applying that literally also
        # rewards merely holding the single best asset on an otherwise balanced
        # even-count trade, which we don't want.
        # Depth-tax (per user): on each side keep the BEST asset at full KTC and
        # discount every subsequent asset geometrically (2nd × f, 3rd × f², …),
        # so three scrubs aren't worth their summed KTC. Even-count balanced
        # trades stay ≈ the naive diff; only the deeper side is taxed. f tunable.
        _KTC_DEPTH_FACTOR = 0.6

        def _depth_adjusted_value(vals: List[float]) -> float:
            _tot = 0.0
            for _i, _v in enumerate(sorted(vals, reverse=True)):
                _tot += _v * (_KTC_DEPTH_FACTOR ** _i)
            return _tot

        def _ktc_adjusted_diff(recv_vals: List[float], sent_vals: List[float]) -> Optional[float]:
            """Depth-tax-adjusted KTC margin (received − sent), in KTC value
            units. Positive ⇒ received side got more value. Best asset per side
            counts in full; each lesser asset is discounted, so a scrubs-for-stud
            package is penalised while a balanced even-count swap stays ≈ naive."""
            if not (recv_vals and sent_vals):
                return None
            return round(_depth_adjusted_value(recv_vals) - _depth_adjusted_value(sent_vals), 1)

        def _diff_at(
            target: date,
            recv_ids: List[str],
            drop_ids: List[str],
            recv_picks: List[str],
            drop_picks: List[str],
            recv_faab: float = 0.0,
            drop_faab: float = 0.0,
        ) -> Optional[float]:
            if target > today:
                return None
            recv_vals = _side_values(target, recv_ids, recv_picks, recv_faab)
            sent_vals = _side_values(target, drop_ids, drop_picks, drop_faab)
            if not (recv_vals and sent_vals):
                return None
            # Package-tax-adjusted difference (Item 2), replacing the old naive
            # Σreceived − Σsent so uneven multi-asset trades value correctly.
            return _ktc_adjusted_diff(recv_vals, sent_vals)

        # --- Trades pass ---
        for row in trades_rows:
            ds = row.get("Date")
            if not ds:
                continue
            try:
                trade_dt = datetime.fromisoformat(str(ds).replace("Z", "+00:00"))
                trade_date = trade_dt.date()
            except Exception:
                continue
            season_i = row.get("Season")
            try:
                season_i = int(season_i) if season_i is not None else None
            except Exception:
                season_i = None

            recv_ids = list(row.get("_recv_player_ids") or [])
            drop_ids = list(row.get("_drop_player_ids") or [])
            recv_picks = list(row.get("_recv_picks") or [])
            drop_picks = list(row.get("_drop_picks") or [])
            recv_faab = float(row.get("_recv_faab") or 0.0)
            drop_faab = float(row.get("_drop_faab") or 0.0)

            # Deal time
            if trade_date <= today:
                diff = _diff_at(trade_date, recv_ids, drop_ids, recv_picks, drop_picks, recv_faab, drop_faab)
                if diff is not None:
                    row["KTC value difference at deal time"] = diff

                # 'Pick value received' = sum of received-side pick
                # values at deal time. Player side intentionally excluded.
                pick_recv_total = 0.0
                pick_recv_hits = 0
                for plabel in recv_picks:
                    v = asset_value_at(str(plabel), None, trade_date, idx)
                    if v is not None:
                        pick_recv_total += v
                        pick_recv_hits += 1
                if pick_recv_hits:
                    row["Pick value received"] = round(pick_recv_total, 1)

            # 'Change in pick value at draft time' — for each received
            # pick, (value at the draft snapshot for that pick's year)
            # minus (value at trade date). The draft snapshot is Sept 1 of
            # the pick's year (Phase 6F): rookie drafts have completed and
            # the pick has resolved into a settled rookie value by then.
            # Sums across picks. Picks whose drafts are in the future don't
            # contribute.
            pick_change_total = 0.0
            pick_change_hits = 0
            for plabel in recv_picks:
                parts = str(plabel).strip().split()
                if len(parts) != 2:
                    continue
                try:
                    pick_year = int(parts[0])
                except Exception:
                    continue
                post_draft = date(pick_year, 9, 1)
                if post_draft > today:
                    continue
                v_before = asset_value_at(str(plabel), None, trade_date, idx)
                v_after = asset_value_at(str(plabel), None, post_draft, idx)
                if v_before is None or v_after is None:
                    continue
                pick_change_total += (v_after - v_before)
                pick_change_hits += 1
            if pick_change_hits:
                row["Change in pick value at draft time"] = round(pick_change_total, 1)

            if season_i is None:
                continue
            ref_points = [
                ("KTC value difference at end of season", _championship_monday(season_i)),
                ("KTC value difference 1 year later",     _plus_years(trade_date, 1)),
                ("KTC value difference 2 years later",     _plus_years(trade_date, 2)),
            ]
            for col_name, ref_date in ref_points:
                # Floor at the trade date — an end-of-season ref earlier
                # than the trade itself doesn't make sense.
                if ref_date < trade_date:
                    ref_date = trade_date
                diff = _diff_at(ref_date, recv_ids, drop_ids, recv_picks, drop_picks, recv_faab, drop_faab)
                if diff is not None:
                    row[col_name] = diff

        # --- Transactions pass: 4 reference points × 3 columns each ---
        # For each transaction we look up the added and dropped players'
        # KTC values at four moments: the transaction date itself, then
        # the same championship-Monday ladder we use for trades. Net
        # is added − dropped (missing side treated as zero). Future-
        # dated references stay N/A via _preserve_na.
        def _tx_value_at(
            target: date,
            added_pid: Optional[str],
            dropped_pid: Optional[str],
        ) -> Tuple[Optional[float], Optional[float], Optional[float]]:
            v_a = asset_value_at(None, str(added_pid), target, idx) if added_pid else None
            v_d = asset_value_at(None, str(dropped_pid), target, idx) if dropped_pid else None
            v_net = None
            if v_a is not None or v_d is not None:
                v_net = round((v_a or 0.0) - (v_d or 0.0), 1)
            return v_a, v_d, v_net

        tx_ref_points = [
            ("at deal time",     "deal"),
            ("at end of season", "end"),
            ("1 year later",     "y1"),
            ("2 years later",    "y2"),
        ]
        for tx_row in transactions_rows:
            ds = tx_row.get("Date")
            if not ds:
                continue
            try:
                tx_dt = datetime.fromisoformat(str(ds).replace("Z", "+00:00"))
                tx_date = tx_dt.date()
            except Exception:
                continue
            season_i = tx_row.get("Season")
            try:
                season_i = int(season_i) if season_i is not None else None
            except Exception:
                season_i = None
            added_pid = tx_row.get("_added_pid")
            dropped_pid = tx_row.get("_dropped_pid")

            for label, tag in tx_ref_points:
                if tag == "deal":
                    ref = tx_date
                elif tag == "end":
                    if season_i is None:
                        continue
                    # End of season = Monday after this season's championship
                    # (the next championship after the move).
                    ref = _championship_monday(season_i)
                    # Never refer to a date earlier than the transaction
                    # itself — matches the floor used in the trades pass.
                    if ref < tx_date:
                        ref = tx_date
                else:
                    # 1/2 years later = exactly 1/2 calendar years after the
                    # transaction date itself.
                    ref = _plus_years(tx_date, 1 if tag == "y1" else 2)
                if ref > today:
                    continue
                v_a, v_d, v_net = _tx_value_at(ref, added_pid, dropped_pid)
                if v_a is not None:
                    tx_row[f"KTC value of player added {label}"] = round(v_a, 1)
                if v_d is not None:
                    tx_row[f"KTC value of player dropped {label}"] = round(v_d, 1)
                if v_net is not None:
                    tx_row[f"Net KTC value {label}"] = v_net
    except Exception as e:
        _log_exc(debug, "ktc_value_diff", e)
    # Surface any dynasty-daddy fetch failures so a silent partial
    # population is visible in the build log.
    try:
        from lotg_support.ktc import get_http_errors
        for err in get_http_errors()[:25]:
            _log(debug, f"[{_now_iso()}] WARN ktc fetch: {err}")
    except Exception:
        pass

    tx = pd.DataFrame(transactions_rows)
    tr = pd.DataFrame(trades_rows)
    ph = pd.DataFrame(pick_rows)
    # No dedupe / chain reconstruction needed — the pick_history rebuild
    # block (above) emits exactly one row per (frame, round, slot) with
    # chains already resolved.

    # Pick-history chain reconstruction lives in the rebuild block above
    # — this legacy in-place mutator is preserved (gated off) for
    # reference and quick rollback. The new rebuild handles chain +
    # commissioner flag in one pass.
    try:
        if False:
            # Use the most-recent season's snapshot; it accumulates history.
            # Sleeper REMOVES used picks from traded_picks once the
            # draft happens. So the latest_season snapshot is missing
            # all historical pick movements (2021-2025 in our chain).
            # Combine events from every season's snapshot — picks that
            # were tracked at any point during their lifecycle show up.
            # Dedup since the same event appears in multiple snapshots
            # (a 2025 1.05 trade in 2023 shows in the 2023, 2024, and
            # 2025 snapshots until the draft used the pick).
            all_events_raw: List[Dict[str, Any]] = []
            for _snap_yr, _snap in traded_picks_by_season.items():
                if _snap:
                    all_events_raw.extend(_snap)
            seen_ev: Set[Tuple[Any, Any, Any, Any, Any]] = set()
            all_events: List[Dict[str, Any]] = []
            for ev in all_events_raw:
                if not isinstance(ev, dict):
                    continue
                key = (
                    ev.get("season"), ev.get("round"),
                    ev.get("roster_id"),
                    ev.get("previous_owner_id"), ev.get("owner_id"),
                )
                if key in seen_ev:
                    continue
                seen_ev.add(key)
                all_events.append(ev)
            # Group events by (season, round). Sleeper returns events in
            # roughly chronological order within a snapshot; we'll preserve
            # insertion order.
            # Sleeper's traded_picks entries carry `roster_id` = the
            # pick's ORIGINAL owner. We key events by (season, round,
            # original_owner) so each pick is its own chain. The old
            # (season, round)-only keying collapsed all picks of the
            # same round into one graph, which dropped chains whenever
            # two picks passed through the same intermediate team.
            events_by_pick: Dict[Tuple[int, int, int], List[Tuple[int, int]]] = defaultdict(list)
            for tp in all_events:
                try:
                    s = int(tp.get("season"))
                    rd = int(tp.get("round"))
                    orig = _to_int(tp.get("roster_id"), None)
                    prev = _to_int(tp.get("previous_owner_id"), None)
                    new = _to_int(tp.get("owner_id"), None)
                    if orig is None or prev is None or new is None:
                        continue
                    events_by_pick[(s, rd, int(orig))].append((int(prev), int(new)))
                except Exception:
                    continue

            # For each pick, walk its chain by following prev → new
            # starting from the original owner.
            # chain_by_final[(season, round, final_rid)] = [origin, mid1, ..., final]
            # chain_by_origin[(season, round, orig_rid)] = same chain
            chain_by_final: Dict[Tuple[int, int, int], List[int]] = {}
            chain_by_origin: Dict[Tuple[int, int, int], List[int]] = {}
            for (s, rd, orig), events in events_by_pick.items():
                chain = [int(orig)]
                cur = int(orig)
                remaining = list(events)
                # Greedy walk: at each step find the next event whose
                # prev_owner matches our current cursor. Avoids needing
                # events to be in chronological order.
                guard = 0
                while remaining and guard < 50:
                    guard += 1
                    next_idx = None
                    for idx, (p, n) in enumerate(remaining):
                        if p == cur:
                            next_idx = idx
                            break
                    if next_idx is None:
                        break
                    _, n = remaining.pop(next_idx)
                    if n in chain:  # cycle guard
                        break
                    chain.append(int(n))
                    cur = int(n)
                final = chain[-1]
                chain_by_final[(s, rd, int(final))] = chain
                chain_by_origin[(s, rd, int(orig))] = chain

            # Apply to ph
            for i, r in ph.iterrows():
                yr = _to_int(r.get("Year"), None)
                num = str(r.get("Number") or "")
                # Accept both legacy 'R1.5' format and current '1.05'.
                m = re.match(r"^R?(\d+)(?:\.|$)", num)
                if yr is None or not m:
                    continue
                rnd = int(m.group(1))

                # Pick-row construction now seeds "Original Team" with the
                # slot owner (true origin) when slot_to_roster_id is
                # available — convert back to roster_id and look up by
                # origin. Fall back to Final Team if Original Team is
                # missing.
                rid_to_team = season_roster_to_team.get(int(yr), {})
                t2r = season_team_to_roster.get(int(yr), {})

                orig_team_disp = str(r.get("Original Team") or "")
                orig_rid = t2r.get(_norm_team_name(orig_team_disp))
                final_team_disp = str(r.get("Final Team") or "")
                final_rid_row = t2r.get(_norm_team_name(final_team_disp))

                chain = None
                if orig_rid is not None:
                    chain = chain_by_origin.get((int(yr), rnd, int(orig_rid)))
                if chain is None and final_rid_row is not None:
                    chain = chain_by_final.get((int(yr), rnd, int(final_rid_row)))
                    if chain and orig_rid is None:
                        orig_rid = int(chain[0])

                if not chain or len(chain) < 2:
                    # No chain found OR chain has just the origin (no trades).
                    # Still make sure Final Team reflects chain end if we
                    # learned one.
                    if chain and orig_rid is not None:
                        ph.at[i, "Original Team"] = rid_to_team.get(int(chain[0]), f"Roster {chain[0]}")
                    # Commissioner flag still applies even without a chain.
                    if orig_rid is not None and (int(yr), rnd, int(orig_rid)) in commissioner_pick_moves:
                        ph.at[i, "Commissioner moved?"] = True
                    continue

                # Rewrite Original Team to the chain origin (covers the
                # rare case where slot map disagreed with the event log).
                ph.at[i, "Original Team"] = rid_to_team.get(int(chain[0]), f"Roster {chain[0]}")
                # Final Team = last owner in the chain (post-trades).
                ph.at[i, "Final Team"] = rid_to_team.get(int(chain[-1]), f"Roster {chain[-1]}")
                # Trade 1..N = intermediate owners (exclude the origin at index 0)
                for j, owner_rid in enumerate(chain[1:11], start=1):
                    try:
                        ph.at[i, f"Trade {j}"] = rid_to_team.get(int(owner_rid), f"Roster {owner_rid}")
                    except Exception:
                        continue

                # Commissioner-moved flag: any (year, round, original_owner)
                # in commissioner_pick_moves means this pick's ownership
                # shift wasn't a normal transaction-recorded trade.
                # Surface this on pick_history so it's visible.
                if (int(yr), rnd, int(chain[0])) in commissioner_pick_moves:
                    ph.at[i, "Commissioner moved?"] = True

        # Apply Commissioner-moved? flag to any remaining picks not
        # covered by the chain reconstruction loop above (e.g.,
        # picks whose final owner happens to be the original owner
        # but a commissioner override still applies).
        for key, new_owner in commissioner_pick_moves.items():
            pick_yr, pick_rnd, pick_orig = key
            rid_to_team = season_roster_to_team.get(int(pick_yr), {})
            orig_team_disp = rid_to_team.get(int(pick_orig))
            if not orig_team_disp:
                continue
            for i, r in ph.iterrows():
                try:
                    if int(r.get("Year")) != int(pick_yr):
                        continue
                except Exception:
                    continue
                num = str(r.get("Number") or "")
                rm = re.match(r"^R?(\d+)", num)
                if not rm or int(rm.group(1)) != int(pick_rnd):
                    continue
                if str(r.get("Original Team") or "").strip() != orig_team_disp:
                    continue
                ph.at[i, "Commissioner moved?"] = True
                break
    except Exception as e:
        _log_exc(debug, "pick_history_reconstruct", e)


    # --------------------------
    # Trades polish: pick-label substitution + new PPG / age / value
    # columns + return-from-trades chain. Runs after pick_history is
    # reconstructed (we need its (Year, Round, Original Team) -> drafted
    # player mapping to substitute already-made picks with the specific
    # slot and player they became, e.g. '2024 1.??' -> '2024 1.05(B. Robinson)').
    # --------------------------
    try:
        # 1) Build pick -> 'R.NN(F. Last)' lookup from pick_history.
        pick_lookup: Dict[Tuple[int, int, str], str] = {}
        if not ph.empty:
            for _, prow in ph.iterrows():
                try:
                    yr_i = int(prow.get("Year"))
                except Exception:
                    continue
                num = str(prow.get("Number") or "")
                m = re.match(r"^R?(\d+)(?:\.(\d+))?", num)
                if not m:
                    continue
                rnd_i = int(m.group(1))
                slot = m.group(2) or "??"
                # Synthetic draft-day picks (2.09 toilet reward, 5.0X FAAB buys)
                # are off-platform and never referenced in Sleeper trades; keep
                # them out of the trade pick-label map so they can't shadow a
                # real same-(year,round,owner) pick.
                if num.strip() == "2.09" or rnd_i >= 5:
                    continue
                orig_team = str(prow.get("Original Team") or "").strip()
                if not orig_team:
                    continue
                player = str(prow.get("Player Picked") or "").strip()
                # Initial + last name: 'B. Robinson'. Fall back to full
                # name when we can't tokenize. Skip 'Unknown' / blank
                # since those aren't drafted yet.
                if player and player.lower() not in ("unknown", "nan"):
                    parts = player.split()
                    if len(parts) >= 2:
                        short = f"{parts[0][0]}. {' '.join(parts[1:])}"
                    else:
                        short = player
                    label = f"{yr_i} {rnd_i}.{slot}({short})"
                else:
                    # Not yet drafted (future pick) — the draft ORDER isn't
                    # finalized, so a slot like '.07' is just a guess off the
                    # owner's roster position. Reference it by ORIGINAL TEAM
                    # instead: '2027 2(Oliverwkw)'. Dynamic — once the season is
                    # drafted, the branch above emits the real slot + player.
                    label = f"{yr_i} {int(rnd_i)}({orig_team})"
                pick_lookup[(yr_i, rnd_i, orig_team)] = label

        def _substitute_picks(asset_str: Optional[str], meta_list: List[Tuple[int, int, str]]) -> Optional[str]:
            """Walk a '; '-joined assets string. For each token that
            looks like a generic pick label ('YYYY R.??'), pop the
            next entry from meta_list and substitute the lookup result
            if available. Player names pass through unchanged."""
            if not asset_str:
                return asset_str
            meta_idx = 0
            tokens = [t.strip() for t in str(asset_str).split(";")]
            out_tokens: List[str] = []
            for tok in tokens:
                if not tok:
                    out_tokens.append(tok)
                    continue
                # Pick label heuristic: starts with 4-digit year + space.
                if len(tok) >= 6 and tok[:4].isdigit() and tok[4] == " " and meta_idx < len(meta_list):
                    yr_i, rnd_i, orig_team = meta_list[meta_idx]
                    meta_idx += 1
                    sub = pick_lookup.get((int(yr_i), int(rnd_i), str(orig_team)))
                    out_tokens.append(sub or tok)
                else:
                    out_tokens.append(tok)
            return "; ".join(t for t in out_tokens if t)

        # 2) Build a per-team event log for the return-from-trades
        # chain. Each entry: (date, role, trade_idx).
        team_player_events: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
        team_pick_events: Dict[Tuple[str, Tuple[int, int, str]], List[Dict[str, Any]]] = defaultdict(list)
        for idx, row in enumerate(trades_rows):
            team = str(row.get("Team") or "")
            d_iso = str(row.get("Date") or "")
            if not team or not d_iso:
                continue
            for pid in (row.get("_recv_player_ids") or []):
                team_player_events[(team, str(pid))].append({"date": d_iso, "role": "in", "tx_idx": idx})
            for pid in (row.get("_drop_player_ids") or []):
                team_player_events[(team, str(pid))].append({"date": d_iso, "role": "out", "tx_idx": idx})
            for pmeta in (row.get("_recv_pick_meta") or []):
                team_pick_events[(team, tuple(pmeta))].append({"date": d_iso, "role": "in", "tx_idx": idx})
            for pmeta in (row.get("_drop_pick_meta") or []):
                team_pick_events[(team, tuple(pmeta))].append({"date": d_iso, "role": "out", "tx_idx": idx})
        for k in team_player_events:
            team_player_events[k].sort(key=lambda e: e["date"])
        for k in team_pick_events:
            team_pick_events[k].sort(key=lambda e: e["date"])

        # V2 chain: also gather FA-drop events from transactions_rows.
        # When a team RECEIVES a player in a trade and later DROPS that
        # player to free agency (not traded), the V1 chain wrongly
        # called them 'retained'. Now they go into a new 'Assets
        # dropped to FA' bucket and the chain terminates there.
        team_fa_drops: Dict[Tuple[str, str], List[str]] = defaultdict(list)
        for r in transactions_rows:
            team_t = str(r.get("Team") or "")
            sid_d = r.get("_dropped_pid")
            if not team_t or not sid_d:
                continue
            d_iso = str(r.get("Date") or "")
            if not d_iso:
                continue
            team_fa_drops[(team_t, str(sid_d))].append(d_iso)
        for k in team_fa_drops:
            team_fa_drops[k].sort()

        def _next_out_player(team: str, pid: str, after: str) -> Optional[Dict[str, Any]]:
            for e in team_player_events.get((team, str(pid)), []):
                if e["role"] == "out" and e["date"] > after:
                    return e
            return None

        def _next_fa_drop_player(team: str, pid: str, after: str) -> Optional[str]:
            for d in team_fa_drops.get((team, str(pid)), []):
                if d > after:
                    return d
            return None

        def _next_out_pick(team: str, pmeta: Tuple[int, int, str], after: str) -> Optional[Dict[str, Any]]:
            for e in team_pick_events.get((team, tuple(pmeta)), []):
                if e["role"] == "out" and e["date"] > after:
                    return e
            return None

        def _player_display(pid: str) -> str:
            return str((pid_meta.get(str(pid)) or {}).get("full_name") or pid)

        def _pick_display(pmeta: Tuple[int, int, str]) -> str:
            sub = pick_lookup.get((int(pmeta[0]), int(pmeta[1]), str(pmeta[2])))
            return sub or f"{pmeta[0]} {pmeta[1]}.??"

        def _trade_assets_received(tr_row: Dict[str, Any]) -> Tuple[List[Tuple[str, Any]], List[str]]:
            """Return (asset_keys, display_labels) for the received
            side of a trade row. Asset key is ('player', sid) or
            ('pick', meta_tuple)."""
            keys: List[Tuple[str, Any]] = []
            disp: List[str] = []
            for pid in (tr_row.get("_recv_player_ids") or []):
                keys.append(("player", str(pid)))
                disp.append(_player_display(pid))
            for pmeta in (tr_row.get("_recv_pick_meta") or []):
                keys.append(("pick", tuple(pmeta)))
                disp.append(_pick_display(tuple(pmeta)))
            return keys, disp

        # 3) Per-row trades polish.
        # Reuse the nflverse log + name lookup from the transactions
        # polish block. They were function-locals there; rebuild here.
        name_to_sid_local2: Dict[str, str] = {}
        for sid, meta in pid_meta.items():
            fn = (meta or {}).get("full_name")
            if fn:
                name_to_sid_local2.setdefault(str(fn), str(sid))

        # --- Phase 8B: "Length of tenure on team" on picks ---
        # How long the DRAFTED player stayed on the team that drafted it (Final
        # Team): days from the draft (≈ late August of the pick year, mirroring
        # the 7D draft anchor) to that player's next exit from the team (or
        # today if still rostered).
        #
        # An UNMADE pick (no player drafted yet — "Unknown"/blank) is left blank
        # so the output coercion renders it "N/A" (via _preserve_na): there is
        # no player whose tenure to measure — the same as a transactions pure
        # drop (no added player). Every MADE pick gets a number ≥ 0: a genuine
        # 0-day tenure (drafted then immediately moved) is 0, and a made pick we
        # can't otherwise resolve also falls back to 0 (there IS a player).
        try:
            if not ph.empty and {"Year", "Final Team", "Player Picked"}.issubset(set(ph.columns)):
                _today_iso = datetime.utcnow().date().isoformat()
                for _i, _pr in ph.iterrows():
                    _ply = str(_pr.get("Player Picked") or "").strip()
                    if not _ply or _ply.lower() in ("unknown", "nan", "n/a", ""):
                        # Pick not made yet → leave blank → N/A.
                        continue
                    # Made pick → always a number ≥ 0 (default 0 if unresolved).
                    _tenure = 0
                    _ft = str(_pr.get("Final Team") or "").strip()
                    _ym = re.match(r"\s*(\d{4})", str(_pr.get("Year") or ""))
                    _sid = name_to_sid_local2.get(_ply)
                    if _ft and _ym and _sid:
                        _draft_iso = f"{int(_ym.group(1))}-08-28"
                        _nx = _next_out_player(_ft, _sid, _draft_iso)
                        _end_iso = (_nx["date"][:10] if _nx else _today_iso)
                        try:
                            _d0 = datetime.fromisoformat(_draft_iso).date()
                            _d1 = datetime.fromisoformat(_end_iso).date()
                            _tenure = max(0, (_d1 - _d0).days)
                        except Exception:
                            _tenure = 0
                    ph.at[_i, "Length of tenure on team"] = _tenure
        except Exception as e:
            _log_exc(debug, "picks_tenure_8b", e)

        # League-wide all-starter avg + per-position avg for position adjustment.
        try:
            starters_pw = pw[pw.get("Starter/Bench").astype(str).str.lower() == "starter"] if not pw.empty and "Starter/Bench" in pw.columns else pd.DataFrame()
            league_starter_avg = float(pd.to_numeric(starters_pw.get("Points"), errors="coerce").mean()) if not starters_pw.empty else 0.0
            pos_avg_map: Dict[str, float] = {}
            if not starters_pw.empty and "Position" in starters_pw.columns:
                for pos_g, gpos in starters_pw.groupby(starters_pw["Position"].astype(str).str.upper()):
                    pos_avg_map[str(pos_g)] = float(pd.to_numeric(gpos.get("Points"), errors="coerce").mean() or 0.0)
        except Exception:
            league_starter_avg = 0.0
            pos_avg_map = {}

        def _player_games(name: Optional[str]) -> List[Tuple[str, float]]:
            if not name:
                return []
            sid = name_to_sid_local2.get(name)
            out: List[Tuple[str, float]] = []
            if sid:
                for entry in nfl_log_by_sid.get(sid, []):
                    if entry.get("_wk_date"):
                        out.append((entry["_wk_date"], float(entry["points"])))
            return out

        def _player_age_at(name: Optional[str], at_iso: str) -> Optional[float]:
            if not name:
                return None
            sid = name_to_sid_local2.get(name)
            if not sid:
                return None
            meta = pid_meta.get(sid) or {}
            bd = meta.get("birth_date")
            if not bd:
                return None
            try:
                d = datetime.fromisoformat(at_iso.replace("Z", "+00:00")).date()
                born = dateparser.parse(str(bd)).date()
                return round((d - born).days / 365.25, 2)
            except Exception:
                return None

        def _player_pos(name: Optional[str]) -> Optional[str]:
            if not name:
                return None
            sid = name_to_sid_local2.get(name)
            if not sid:
                return None
            return ((pid_meta.get(sid) or {}).get("pos") or "").upper() or None

        # Drafted-pick lookup (Phase 7D): canonical pick identity
        # (year, round, original owner) -> (drafted player, team that actually
        # made the pick, draft year). Used to fold the PPG of the player drafted
        # with a received pick into "Avg PPG of received players on team" —
        # but only for the team that ended up making the selection.
        _pick_to_drafted: Dict[Tuple[int, int, str], Tuple[str, str, int]] = {}
        if not ph.empty and {"Year", "Number", "Original Team", "Final Team", "Player Picked"}.issubset(set(ph.columns)):
            for _, _phr in ph.iterrows():
                _ym = re.match(r"\s*(\d{4})", str(_phr.get("Year", "")))
                _nm = re.match(r"\s*(\d+)\.(\d+)", str(_phr.get("Number", "")))
                _plp = str(_phr.get("Player Picked", "")).strip()
                if not _ym or not _nm or not _plp or _plp.upper() == "N/A":
                    continue
                # Skip synthetic off-platform picks (2.09 / 5.0X) — they aren't
                # Sleeper-traded and must not shadow a real (year,round,owner) pick.
                if str(_phr.get("Number", "")).strip() == "2.09" or int(_nm.group(1)) >= 5:
                    continue
                _key = (int(_ym.group(1)), int(_nm.group(1)), _norm_team_name(_phr.get("Original Team", "")))
                _pick_to_drafted[_key] = (_plp, _norm_team_name(_phr.get("Final Team", "")), int(_ym.group(1)))

        # (fantasy team, player name) -> [(year, week, points, wk_date)] for
        # weeks that player STARTED for that team — powers the trade Points
        # Added (received assets' starter output, per week).
        _started_idx: Dict[Tuple[str, str], List[Tuple[int, int, float, str]]] = defaultdict(list)
        _scols = ["Team", "Player", "Year", "Week", "Points", "Starter/Bench"]
        if not pw.empty and set(_scols).issubset(pw.columns):
            for _t, _p, _y, _w, _pt, _sb in zip(*[pw[c] for c in _scols]):
                if str(_sb) != "Starter" or pd.isna(_y) or pd.isna(_w):
                    continue
                try:
                    _yi, _wi = int(_y), int(_w)
                except Exception:
                    continue
                _wkd = (date(_yi, 9, 7) + timedelta(days=7 * (_wi - 1))).isoformat()
                _started_idx[(str(_t), str(_p))].append((_yi, _wi, float(_pt or 0.0), _wkd))

        # (fantasy team, player name) -> sorted wk_dates the player was ROSTERED
        # for that team in ANY NFL week (starter OR bench). Lets the picks pass
        # tell "cut after the draft before week 1" (never rostered → on-team PPG
        # N/A) apart from "rostered but no game production" (→ 0).
        _pw_rostered_idx: Dict[Tuple[str, str], List[str]] = defaultdict(list)
        _rcols = ["Team", "Player", "Year", "Week"]
        if not pw.empty and set(_rcols).issubset(pw.columns):
            for _t, _p, _y, _w in zip(*[pw[c] for c in _rcols]):
                if pd.isna(_y) or pd.isna(_w):
                    continue
                try:
                    _yi, _wi = int(_y), int(_w)
                except Exception:
                    continue
                _wkd = (date(_yi, 9, 7) + timedelta(days=7 * (_wi - 1))).isoformat()
                _pw_rostered_idx[(str(_t), str(_p))].append(_wkd)

        # ---- Item 7E indexes (V2 Trade addition value: leverage + cuff) ----
        # (fantasy team, player) -> per-week roster rows with starter + injury
        # flags, so we can compute a received player's % of starts made while
        # rostered (and the injury-adjusted variant) exactly like the
        # transaction "Player addition value" leverage multipliers.
        _pwfull_idx: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
        # (fantasy team, year, week) -> teammates that week (for cuff teammate
        # scan) and player -> (wk_date, NFL team, position) (for a received
        # player's own NFL team/position near the trade).
        _twk_idx: Dict[Tuple[str, int, int], List[Dict[str, Any]]] = defaultdict(list)
        _player_nflpos_idx: Dict[str, List[Tuple[str, str, str]]] = defaultdict(list)
        _fcols = ["Team", "Player", "Year", "Week", "Starter/Bench",
                  "NFL team", "Position", "Injury?", "Bye?"]
        if not pw.empty and set(_fcols).issubset(pw.columns):
            for _t, _p, _y, _w, _sb, _nt, _ps, _inj, _bye in zip(*[pw[c] for c in _fcols]):
                if pd.isna(_y) or pd.isna(_w):
                    continue
                try:
                    _yi, _wi = int(_y), int(_w)
                except Exception:
                    continue
                _wkd = (date(_yi, 9, 7) + timedelta(days=7 * (_wi - 1))).isoformat()
                _is_starter = str(_sb) == "Starter"
                _inj_free = (not bool(_inj)) and (not bool(_bye))
                _nflt = str(_nt or "")
                _posu = str(_ps or "").upper()
                _pname = str(_p)
                _pwfull_idx[(str(_t), _pname)].append(
                    {"wkd": _wkd, "starter": _is_starter, "inj_free": _inj_free})
                _twk_idx[(str(_t), _yi, _wi)].append(
                    {"Player": _pname, "starter": _is_starter, "nfl": _nflt,
                     "pos": _posu, "wkd": _wkd})
                _player_nflpos_idx[_pname].append((_wkd, _nflt, _posu))

        _TRADE_CUFF_BONUS = 5.0  # mirror the transaction CUFF_BONUS
        # Item 7E (pick value): future draft picks received/sent are valued with
        # the SAME round weights tanking uses (_FUTURE_PICK_WEIGHTS: 1st=0.25,
        # 2nd=0.09, 3rd=0.03, 4th=0.01, next-3-seasons only) and converted to a
        # PPG-equivalent by this coefficient, then added to Trade addition value
        # so pick-heavy hauls register. 20.0 => a future 1st ≈ +5 (≈ one cuff
        # bonus). TUNABLE — adjust to taste.
        _TRADE_PICK_COEFF = 20.0

        def _recv_is_cuff(name: str, team: str, at_prefix: str) -> bool:
            """A received player is a cuff (mirrors 'Cuff at time of pickup?'):
            at the trade the team already rosters a STARTER on the SAME NFL
            team + position whose last-5 PPG is 10+ above the received player's
            last-5, and that teammate is still rostered at the trade week."""
            if not name or not at_prefix:
                return False
            # received player's own NFL team + position near the trade
            _own = [e for e in _player_nflpos_idx.get(name, []) if e[0] <= at_prefix]
            if _own:
                _own.sort(key=lambda e: e[0])
                _nflt, _pos = _own[-1][1], _own[-1][2]
            else:
                _nflt, _pos = "", (_player_pos(name) or "")
            if not _nflt or not _pos:
                return False

            def _pre5(_nm: str) -> Optional[float]:
                _g = [(d, p) for d, p in _player_games(_nm) if d < at_prefix]
                if not _g:
                    return None
                _g.sort(key=lambda kv: kv[0], reverse=True)
                _win = _g[:5]
                return sum(p for _, p in _win) / len(_win)

            # latest roster week for THIS team at/just before the trade
            _cand = None
            for (_t, _yi, _wi), _ents in _twk_idx.items():
                if _t != team:
                    continue
                _wkd = _ents[0]["wkd"]
                if _wkd <= at_prefix and (_cand is None or _wkd > _cand[0]):
                    _cand = (_wkd, _yi, _wi)
            if _cand is None:
                return False
            _, _yr, _wk = _cand
            _rostered = {m["Player"] for m in _twk_idx.get((team, _yr, _wk), [])}
            _added_avg = _pre5(name) or 0.0
            for _w in (_wk, _wk - 1, _wk - 2):
                if _w < 1:
                    continue
                for mate in _twk_idx.get((team, _yr, _w), []):
                    if mate["Player"] == name or mate["Player"] not in _rostered:
                        continue
                    if (not mate["starter"]) or mate["nfl"] != _nflt or mate["pos"] != _pos:
                        continue
                    _mavg = _pre5(mate["Player"])
                    if _mavg is not None and _added_avg + 10 <= _mavg:
                        return True
            return False

        def _nfl_wk_pts(sid=None, name=None) -> Dict[Tuple[int, int], float]:
            """A player's real NFL fantasy points keyed by (year, week)."""
            if not sid and name:
                sid = name_to_sid_local2.get(name)
            out: Dict[Tuple[int, int], float] = {}
            for _e in nfl_log_by_sid.get(str(sid), []):
                if _e.get("year") is not None and _e.get("week") is not None:
                    out[(int(_e["year"]), int(_e["week"]))] = float(_e.get("points") or 0.0)
            return out

        # --- Phase 8C: PPG / points columns on picks ---
        # For each MADE pick, evaluating the player it became. Two PPG flavours
        # (each with a position-adjusted variant), then the on-team points added:
        #   Avg PPG on team                          mean nflverse PPG over the
        #                                            games played while on the
        #                                            DRAFTING team (draft ≈ Aug 28
        #                                            of the pick year → next exit)
        #   Avg PPG on team adjusted by position     × league_starter_avg/pos_avg
        #   Avg career PPG                           mean nflverse PPG over EVERY
        #                                            game the player has played
        #                                            (whole career on record —
        #                                            injury-adjusted by construct-
        #                                            ion, the log has only games
        #                                            played; NOT team-scoped)
        #   Avg career PPG adjusted by position      × league_starter_avg/pos_avg
        #   Points added                             Σ points in weeks STARTED for
        #                                            the drafting team in the
        #                                            post-draft tenure window
        #   Avg points added                         Points added / started weeks
        #   Avg points added adjusted by position    same, position-adjusted
        # N/A only for an unmade pick (no player) or a PPG with no games to
        # average (on-team PPG can be N/A while career PPG is a number — e.g. a
        # pick traded before ever playing here). Every made pick's point columns
        # are a number ≥ 0. Scoped to ph via explicit "N/A" so it never touches
        # the identically-named trades columns.
        _pick_ppg_cols = [
            "Avg PPG on team", "Avg PPG on team adjusted by position",
            "Avg career PPG", "Avg career PPG adjusted by position",
            "Points added", "Avg points added",
            "Avg points added adjusted by position",
        ]
        try:
            if not ph.empty and {"Year", "Final Team", "Player Picked"}.issubset(set(ph.columns)):
                for _c in _pick_ppg_cols:
                    ph[_c] = "N/A"
                    ph[_c] = ph[_c].astype(object)
                _today_iso2 = datetime.utcnow().date().isoformat()
                for _i, _pr in ph.iterrows():
                    _ply = str(_pr.get("Player Picked") or "").strip()
                    if not _ply or _ply.lower() in ("unknown", "nan", "n/a", ""):
                        continue  # unmade pick → all stay "N/A"
                    # Made pick → these are numbers ≥ 0 (default 0). Career PPG
                    # is NEVER N/A for a made pick: a drafted player who logged
                    # no post-draft games (e.g. a vet drafted near end of career
                    # who never played again) is 0.0, not N/A.
                    ph.at[_i, "Points added"] = 0.0
                    ph.at[_i, "Avg points added"] = 0.0
                    ph.at[_i, "Avg points added adjusted by position"] = 0.0
                    ph.at[_i, "Avg career PPG"] = 0.0
                    ph.at[_i, "Avg career PPG adjusted by position"] = 0.0
                    _ft = str(_pr.get("Final Team") or "").strip()
                    _ym = re.match(r"\s*(\d{4})", str(_pr.get("Year") or ""))
                    # Resolve by Sleeper player_id (threaded from the draft data)
                    # so suffixes ("III") and duplicate names don't silently
                    # drop the game log; fall back to a name lookup if absent.
                    _sid = _pr.get("_player_id") or name_to_sid_local2.get(_ply)
                    if not (_ft and _ym and _sid):
                        continue  # PPG stays N/A (no data), points stay 0
                    _sid = str(_sid)
                    _draft_iso = f"{int(_ym.group(1))}-08-28"
                    _nx = _next_out_player(_ft, _sid, _draft_iso)
                    _end_iso = (_nx["date"][:10] if _nx else _today_iso2)
                    _pos = ((pid_meta.get(_sid) or {}).get("pos") or "").upper() or None
                    _posa = pos_avg_map.get(_pos or "", 0.0)
                    _fac = (league_starter_avg / _posa) if (_posa and league_starter_avg) else 1.0
                    _all_games = [
                        (_e["_wk_date"], float(_e["points"]))
                        for _e in nfl_log_by_sid.get(_sid, [])
                        if _e.get("_wk_date")
                    ]
                    # Avg PPG on team: PPG over the games the player played while
                    # on the drafting team (draft → next exit). N/A ONLY if the
                    # player was never on the team's roster for an NFL week (cut
                    # after the draft before week 1); if they were rostered for
                    # ≥1 NFL week but logged no games (injured all year, etc.)
                    # it's 0, not N/A.
                    _on_team = [
                        _p for _d, _p in _all_games
                        if _d >= _draft_iso and (not _end_iso or _d < _end_iso)
                    ]
                    _rostered_wk = [
                        _wkd for _wkd in _pw_rostered_idx.get((_ft, _ply), [])
                        if _wkd >= _draft_iso and (not _end_iso or _wkd < _end_iso)
                    ]
                    # ROSTER PRESENCE GATES FIRST. If the player was never on the
                    # drafting team's roster for an NFL week, on-team PPG is N/A —
                    # even if they have a game log (a pick drafted then dropped/
                    # traded before week 1 has no on-team exit, so the window
                    # would otherwise degenerate to their whole career and mis-
                    # report that career PPG as on-team). Only when they WERE
                    # rostered do we average their games here (0 if rostered but
                    # no game production).
                    if not _rostered_wk:
                        pass  # never rostered an NFL week here → stays N/A
                    elif _on_team:
                        _avg = sum(_on_team) / len(_on_team)
                        ph.at[_i, "Avg PPG on team"] = round(_avg, 2)
                        ph.at[_i, "Avg PPG on team adjusted by position"] = round(_avg * _fac, 2)
                    else:
                        # rostered for ≥1 NFL week but no game production → 0
                        ph.at[_i, "Avg PPG on team"] = 0.0
                        ph.at[_i, "Avg PPG on team adjusted by position"] = 0.0
                    # Avg career PPG: every nflverse game the player played FROM
                    # THE DRAFT ONWARD (across all teams, not just the drafting
                    # one; injury-adjusted — DNP weeks aren't in the log). Vets
                    # are treated as rookies here: only post-draft games count,
                    # so their pre-draft history is excluded. Stays 0.0 (set
                    # above) when there are no post-draft games.
                    _career = [_p for _d, _p in _all_games if _d >= _draft_iso]
                    if _career:
                        _cavg = sum(_career) / len(_career)
                        ph.at[_i, "Avg career PPG"] = round(_cavg, 2)
                        ph.at[_i, "Avg career PPG adjusted by position"] = round(_cavg * _fac, 2)
                    # Points added: points in weeks the player STARTED here.
                    _pa = 0.0
                    _nwk = 0
                    for (_yi, _wi, _pt, _wkd) in _started_idx.get((_ft, _ply), []):
                        if _wkd >= _draft_iso and (not _end_iso or _wkd < _end_iso):
                            _pa += _pt
                            _nwk += 1
                    ph.at[_i, "Points added"] = round(_pa, 2)
                    if _nwk:
                        ph.at[_i, "Avg points added"] = round(_pa / _nwk, 2)
                        ph.at[_i, "Avg points added adjusted by position"] = round(_pa * _fac / _nwk, 2)
        except Exception as e:
            _log_exc(debug, "picks_ppg_8c", e)

        # --- Phase 8D: KTC-over-time columns on picks ---
        # The drafted player's KeepTradeCut value (1QB trade_value) at five
        # checkpoints relative to the draft (anchor ≈ Aug 28 of the pick year):
        #   KTC on draft day            the draft anchor itself
        #   KTC at end of rookie year   ≈ Feb 1 of the year AFTER the draft
        #   KTC 1 / 2 / 3 / 4 years after draft day
        # Uses the same dynasty-daddy KTC index (_ktc_idx) that powers the
        # transactions/trades KTC columns. N/A for an unmade pick, an untracked
        # player, or a checkpoint date in the future / before KTC history began
        # (≈ April 2021) — explicit "N/A" scoped to ph.
        _pick_ktc_cols = [
            "KTC on draft day", "KTC at end of rookie year",
            "KTC 1 year after draft day", "KTC 2 years after draft day",
            "KTC 3 years after draft day", "KTC 4 years after draft day",
        ]
        try:
            from lotg_support.ktc import asset_value_at as _ktc_value_at
            if not ph.empty and {"Year", "Player Picked"}.issubset(set(ph.columns)) and _ktc_idx is not None:
                for _c in _pick_ktc_cols:
                    ph[_c] = "N/A"
                    ph[_c] = ph[_c].astype(object)
                _ktc_today = datetime.utcnow().date()
                for _i, _pr in ph.iterrows():
                    _ply = str(_pr.get("Player Picked") or "").strip()
                    if not _ply or _ply.lower() in ("unknown", "nan", "n/a", ""):
                        continue  # unmade pick → all stay "N/A"
                    _ym = re.match(r"\s*(\d{4})", str(_pr.get("Year") or ""))
                    _sid = _pr.get("_player_id") or name_to_sid_local2.get(_ply)
                    if not (_ym and _sid):
                        continue
                    _sid = str(_sid)
                    _yr = int(_ym.group(1))
                    _checkpoints = [
                        ("KTC on draft day", date(_yr, 8, 28)),
                        ("KTC at end of rookie year", date(_yr + 1, 2, 1)),
                        ("KTC 1 year after draft day", date(_yr + 1, 8, 28)),
                        ("KTC 2 years after draft day", date(_yr + 2, 8, 28)),
                        ("KTC 3 years after draft day", date(_yr + 3, 8, 28)),
                        ("KTC 4 years after draft day", date(_yr + 4, 8, 28)),
                    ]
                    for _col, _tgt in _checkpoints:
                        # A checkpoint still in the FUTURE has no KTC yet — leave
                        # it N/A. Without this guard the index lookup (latest
                        # value on-or-before target) clamps to today's value and
                        # mislabels it as the future checkpoint. Mirrors the
                        # transactions/trades KTC pass (`if ref > today: skip`).
                        if _tgt > _ktc_today:
                            continue
                        try:
                            _v = _ktc_value_at(None, str(_sid), _tgt, _ktc_idx)
                        except Exception:
                            _v = None
                        if _v is not None:
                            ph.at[_i, _col] = round(_v, 1)
        except Exception as e:
            _log_exc(debug, "picks_ktc_8d", e)

        # --- Phase 8E: draft/usage cluster on picks ---
        # Characterise the drafted player's usage on the drafting team over the
        # post-draft tenure (draft ≈ Aug 28 → next exit), mirroring the
        # transaction usage columns:
        #   Age when drafted
        #   Player addition value   on-team adj PPG × (1+%starts)
        #                           × (1+injury-adj %starts) + CUFF_BONUS (5 when
        #                           cuff-when-drafted). On-team baseline (per
        #                           user): a pick gives up no player, so there is
        #                           no "dropped" side.
        #   Cuff when drafted?      handcuff test at the draft (best effort; the
        #                           rookie's NFL team/pos may be unknown until
        #                           they play, so this skews toward False)
        #   Weeks before first start
        #   Number of starts before next transaction
        #   % of starts made while rostered by drafting team
        #   Injury adjusted % of starts made while rostered by drafting team
        # N/A for unmade picks and for ratios with no rostered/uninjured weeks
        # (never rostered an NFL week here). "Number of starts" is a count (≥0).
        # Scoped to ph via explicit "N/A".
        _PICK_CUFF_BONUS = 5.0
        _pick_usage_na = [
            "Age when drafted", "Weeks before first start",
            "Number of starts before next transaction",
            "% of starts made while rostered by drafting team",
            "Injury adjusted % of starts made while rostered by drafting team",
            "Player addition value",
        ]
        try:
            if not ph.empty and {"Year", "Final Team", "Player Picked"}.issubset(set(ph.columns)):
                for _c in _pick_usage_na:
                    ph[_c] = "N/A"
                    ph[_c] = ph[_c].astype(object)
                ph["Cuff when drafted?"] = False
                ph["Cuff when drafted?"] = ph["Cuff when drafted?"].astype(object)
                _today_iso3 = datetime.utcnow().date().isoformat()
                for _i, _pr in ph.iterrows():
                    _ply = str(_pr.get("Player Picked") or "").strip()
                    if not _ply or _ply.lower() in ("unknown", "nan", "n/a", ""):
                        continue  # unmade pick → all stay N/A / False
                    _ft = str(_pr.get("Final Team") or "").strip()
                    _ym = re.match(r"\s*(\d{4})", str(_pr.get("Year") or ""))
                    _sid = _pr.get("_player_id") or name_to_sid_local2.get(_ply)
                    if not (_ft and _ym and _sid):
                        continue
                    _sid = str(_sid)
                    _yr = int(_ym.group(1))
                    _draft_iso = f"{_yr}-08-28"
                    # Number of starts is a count → made picks get 0+ (not N/A).
                    ph.at[_i, "Number of starts before next transaction"] = 0
                    # Age when drafted (years at the draft anchor).
                    _bd = (pid_meta.get(_sid) or {}).get("birth_date")
                    if _bd:
                        try:
                            _born = dateparser.parse(str(_bd)).date()
                            ph.at[_i, "Age when drafted"] = round(
                                (date(_yr, 8, 28) - _born).days / 365.25, 2)
                        except Exception:
                            pass
                    # Post-draft tenure window + per-week roster/start stats.
                    _nx = _next_out_player(_ft, _sid, _draft_iso)
                    _end_iso = (_nx["date"][:10] if _nx else _today_iso3)
                    _ents = [
                        _e for _e in _pwfull_idx.get((_ft, _ply), [])
                        if _e["wkd"] >= _draft_iso and (not _end_iso or _e["wkd"] < _end_iso)
                    ]
                    # Cuff when drafted? — the handcuff test evaluated at the
                    # player's FIRST week on the drafting team's roster (per
                    # user), not the draft date: at the draft a rookie's NFL
                    # team/position isn't known yet, but by their first rostered
                    # week it is. False if never rostered here.
                    _cuff = False
                    if _ents:
                        _first_roster_wkd = min(_e["wkd"] for _e in _ents)
                        try:
                            _cuff = bool(_recv_is_cuff(_ply, _ft, _first_roster_wkd))
                        except Exception:
                            _cuff = False
                    if _cuff:
                        ph.at[_i, "Cuff when drafted?"] = True
                    _wk = len(_ents)
                    _st = sum(1 for _e in _ents if _e["starter"])
                    _iwk = sum(1 for _e in _ents if _e["inj_free"])
                    _ist = sum(1 for _e in _ents if _e["inj_free"] and _e["starter"])
                    ph.at[_i, "Number of starts before next transaction"] = int(_st)
                    _pct = (_st / _wk) if _wk > 0 else 0.0
                    if _wk > 0:
                        ph.at[_i, "% of starts made while rostered by drafting team"] = round(_pct, 4)
                    _ipct = (_ist / _iwk) if _iwk > 0 else 0.0
                    if _iwk > 0:
                        ph.at[_i, "Injury adjusted % of starts made while rostered by drafting team"] = round(_ipct, 4)
                    # Weeks before first start = rostered weeks before the first start.
                    _start_wkds = sorted(_e["wkd"] for _e in _ents if _e["starter"])
                    if _start_wkds:
                        _first = _start_wkds[0]
                        ph.at[_i, "Weeks before first start"] = int(
                            sum(1 for _e in _ents if _e["wkd"] < _first))
                    # Player addition value (on-team baseline). N/A when there's
                    # no on-team production to value (Avg PPG on team adj is N/A).
                    _otadj = ph.at[_i, "Avg PPG on team adjusted by position"]
                    try:
                        _otadj_f = float(_otadj)
                    except Exception:
                        _otadj_f = None
                    if _otadj_f is not None:
                        _cuffb = _PICK_CUFF_BONUS if _cuff else 0.0
                        ph.at[_i, "Player addition value"] = round(
                            _otadj_f * (1.0 + _pct) * (1.0 + _ipct) + _cuffb, 4)
        except Exception as e:
            _log_exc(debug, "picks_usage_8e", e)

        # --- Item 1: "Pick-adjusted Difference in [...]" columns on picks ---
        # For each position-adjusted average and each KTC column, a companion
        # column = (this pick's stat) − (mean of that stat over ALL non-vet
        # picks, all years, made in the 3-SLOT window around this pick). The
        # window is by OVERALL draft position (crossing round boundaries), with
        # the user's edge rule: the very first pick uses {this, next} (2 slots),
        # the very last uses {prev-2, prev, this} (3), everything else uses
        # {prev, this, next}. The 2021 vet/startup draft is excluded entirely —
        # its rows are N/A and it is left out of every reference average.
        _padj_diff_stats = [
            "Avg PPG on team adjusted by position",
            "Avg career PPG adjusted by position",
            "Avg points added adjusted by position",
            "Player addition value",
            "KTC on draft day", "KTC at end of rookie year",
            "KTC 1 year after draft day", "KTC 2 years after draft day",
            "KTC 3 years after draft day", "KTC 4 years after draft day",
        ]
        try:
            if not ph.empty and {"Year", "Number"}.issubset(set(ph.columns)):
                _vals_by_slot: Dict[str, Dict[Tuple[int, int], List[float]]] = {
                    _st: defaultdict(list) for _st in _padj_diff_stats
                }
                _rs_of: Dict[Any, Tuple[int, int]] = {}
                _max_slot = 0
                _max_round = 0
                for _pi in ph.index:
                    if "(vet)" in str(ph.at[_pi, "Year"]).lower():
                        continue
                    _m = re.match(r"\s*(\d+)\.(\d+)", str(ph.at[_pi, "Number"]))
                    if not _m:
                        continue
                    _R, _S = int(_m.group(1)), int(_m.group(2))
                    # Synthetic picks compare as their equivalent slot: the 2.09
                    # toilet pick counts as a 2.08, and every 5.0X counts as a
                    # 4.08 — both for its OWN comparison window and for the pools
                    # the real 2.08 / 4.08 picks compare against. Mapping here
                    # (before _max_slot/_max_round) also keeps slot 9 / round 5
                    # out of the window-position math.
                    if (_R, _S) == (2, 9):
                        _R, _S = 2, 8
                    elif _R == 5:
                        _R, _S = 4, 8
                    _rs_of[_pi] = (_R, _S)
                    _max_slot = max(_max_slot, _S)
                    _max_round = max(_max_round, _R)
                    for _st in _padj_diff_stats:
                        try:
                            _vals_by_slot[_st][(_R, _S)].append(float(ph.at[_pi, _st]))
                        except Exception:
                            pass
                _rsize = _max_slot or 1
                _maxpos = _rsize * (_max_round or 1)

                def _window_slots(_R: int, _S: int) -> List[Tuple[int, int]]:
                    _pos = (_R - 1) * _rsize + _S
                    if _pos <= 1:
                        _ps = [1, 2]
                    elif _pos >= _maxpos:
                        _ps = [_maxpos - 2, _maxpos - 1, _maxpos]
                    else:
                        _ps = [_pos - 1, _pos, _pos + 1]
                    _out = []
                    for _p in _ps:
                        if 1 <= _p <= _maxpos:
                            _out.append(((_p - 1) // _rsize + 1, (_p - 1) % _rsize + 1))
                    return _out

                for _st in _padj_diff_stats:
                    _col = f"Pick-adjusted Difference in {_st}"
                    ph[_col] = "N/A"
                    ph[_col] = ph[_col].astype(object)
                for _pi in ph.index:
                    _rs = _rs_of.get(_pi)
                    if not _rs:
                        continue  # vet / unparseable → leave N/A
                    _ws = _window_slots(_rs[0], _rs[1])
                    _pos0 = (_rs[0] - 1) * _rsize + _rs[1]
                    for _st in _padj_diff_stats:
                        _col = f"Pick-adjusted Difference in {_st}"
                        try:
                            _my = float(ph.at[_pi, _st])
                        except Exception:
                            continue  # this pick's stat is N/A → diff stays N/A
                        # From overall pick 1.05 onward (_pos0 >= 5), the two
                        # OUTER slots of the 3-slot window are averaged into a
                        # synthetic 4th comparison pick, and the baseline is the
                        # mean of the four per-slot means (3 window slots + the
                        # outer-average), up-weighting the window's edges. Picks
                        # 1.01-1.04 keep the original rule: a flat pooled mean of
                        # every value in the edge-clamped window.
                        if _pos0 >= 5 and len(_ws) == 3:
                            _means = []
                            for _sl in _ws:
                                _vv = _vals_by_slot[_st].get(_sl, [])
                                _means.append(sum(_vv) / len(_vv) if _vv else None)
                            if _means[0] is not None and _means[2] is not None:
                                _fourth = (_means[0] + _means[2]) / 2.0
                                _terms = [_m for _m in _means if _m is not None] + [_fourth]
                                ph.at[_pi, _col] = round(_my - sum(_terms) / len(_terms), 4)
                            else:
                                _present = [_m for _m in _means if _m is not None]
                                if _present:
                                    ph.at[_pi, _col] = round(_my - sum(_present) / len(_present), 4)
                        else:
                            _ref = [v for _sl in _ws for v in _vals_by_slot[_st].get(_sl, [])]
                            if _ref:
                                ph.at[_pi, _col] = round(_my - sum(_ref) / len(_ref), 4)
        except Exception as e:
            _log_exc(debug, "picks_pickadj_diff_item1", e)

        # --- Item 3: "Trade impact score" = wins the trade actually
        # flipped (player stats → win impact). For each week AFTER the trade in
        # which a received asset started for this team, we already know the
        # week's net points (received starters − the top-k players traded away,
        # the existing maximize rule). Counterfactual: had the team NOT made the
        # trade, that week's margin would be (actual margin − net points). If the
        # win/loss outcome flips between the two, the trade is credited (+1) or
        # debited (−1) that game. Picks count via the players drafted with them
        # (future weeks), so a rebuild is credited when its rookies win games —
        # it is not punished for shedding veterans. Σ over weeks × coefficient.
        _team_margin: Dict[Tuple[str, int, int], float] = {}
        _mcols = ["Team", "Year", "Week", "Margin"]
        if not tw.empty and set(_mcols).issubset(tw.columns):
            for _t, _y, _w, _mg in zip(*[tw[c] for c in _mcols]):
                if pd.isna(_y) or pd.isna(_w):
                    continue
                try:
                    _team_margin[(str(_t), int(_y), int(_w))] = float(_mg)
                except Exception:
                    continue
        _TPI_COEFF = 1000.0  # KTC-points per net game flipped (tunable)

        for idx, row in enumerate(trades_rows):
            team = str(row.get("Team") or "")
            trade_iso = str(row.get("Date") or "")
            trade_prefix = trade_iso[:10] if len(trade_iso) >= 10 else trade_iso
            if not team:
                continue

            # ----- (a) Substitute pick labels in Assets received/dropped -----
            row["Assets received"] = _substitute_picks(
                row.get("Assets received"), row.get("_recv_pick_meta") or []
            )
            row["Assets sent"] = _substitute_picks(
                row.get("Assets sent"), row.get("_drop_pick_meta") or []
            )

            # ----- (b) Asset difference in average age -----
            # Players use birth_date as usual. Picks count too —
            # treat each pick as an unknown future rookie. NFL
            # rookies are ~22 years old at draft time, so a pick for
            # year Y represents a player born ~Sept 1 of (Y - 22).
            # Age at the trade date = (trade_date - that birth date)
            # in years. This makes a pick get YOUNGER (in expected
            # age) the further out from its draft year it's traded,
            # which lines up with intuition: a 2027 pick today is
            # worth a younger eventual player than a 2024 pick.
            recv_player_names = [_player_display(pid) for pid in (row.get("_recv_player_ids") or [])]
            drop_player_names = [_player_display(pid) for pid in (row.get("_drop_player_ids") or [])]
            recv_ages = [a for a in (_player_age_at(n, trade_iso) for n in recv_player_names) if a is not None]
            drop_ages = [a for a in (_player_age_at(n, trade_iso) for n in drop_player_names) if a is not None]

            # Uses the module-level _pick_expected_age helper. Takes a
            # date (not an ISO string), so parse trade_iso once.
            try:
                _trade_dt_for_age = datetime.fromisoformat(trade_iso.replace("Z", "+00:00")).date()
            except Exception:
                _trade_dt_for_age = None

            for pmeta in (row.get("_recv_pick_meta") or []):
                yr_i = int(pmeta[0]) if pmeta and pmeta[0] is not None else None
                if yr_i is None or _trade_dt_for_age is None:
                    continue
                a = _pick_expected_age(yr_i, _trade_dt_for_age)
                if a is not None:
                    recv_ages.append(a)
            for pmeta in (row.get("_drop_pick_meta") or []):
                yr_i = int(pmeta[0]) if pmeta and pmeta[0] is not None else None
                if yr_i is None or _trade_dt_for_age is None:
                    continue
                a = _pick_expected_age(yr_i, _trade_dt_for_age)
                if a is not None:
                    drop_ages.append(a)

            if recv_ages and drop_ages:
                row["Asset difference in average age"] = round(
                    (sum(recv_ages) / len(recv_ages)) - (sum(drop_ages) / len(drop_ages)), 2
                )
            else:
                # Never blank (Phase 7C): one side has no aged asset (it was
                # FAAB-only or empty — a give-away), so there is no measurable
                # average-age differential. Report 0 rather than leaving it
                # blank. Players and picks (as future rookies) both carry ages;
                # only FAAB / nothing sides land here.
                row["Asset difference in average age"] = 0.0

            # --- Tanking-delta inputs (Phase 6E) ---
            # recv_ages / drop_ages already include BOTH players (birth-date
            # age) and picks (expected future-rookie age), so they are exactly
            # the entities that move in/out of "Team age including picks".
            row["_tank_recv_age_sum"] = float(sum(recv_ages))
            row["_tank_recv_n"] = len(recv_ages)
            row["_tank_sent_age_sum"] = float(sum(drop_ages))
            row["_tank_sent_n"] = len(drop_ages)
            # Future-capital delta = round-weighted future picks received minus
            # sent. Only picks for the next 3 seasons count toward future
            # capital (matches _future_cap_held / _future_picks_owned); a
            # current-season rookie pick (year == season) is not "future".
            _tr_season = _to_int(row.get("Season"), None)
            def _fcap_of(meta_list, _s=_tr_season) -> float:
                if _s is None:
                    return 0.0
                tot = 0.0
                for pmeta in (meta_list or []):
                    try:
                        yr_i = int(pmeta[0]); rnd_i = int(pmeta[1])
                    except Exception:
                        continue
                    if _s < yr_i <= _s + 3:
                        tot += _FUTURE_PICK_WEIGHTS.get(rnd_i, 0.0)
                return tot
            row["_tank_fcap_delta"] = round(
                _fcap_of(row.get("_recv_pick_meta")) - _fcap_of(row.get("_drop_pick_meta")), 4
            )

            # ----- (c) Forward-looking tenure window + PPG averages -----
            # For each received player, find their next drop/trade-out
            # by THIS team. Build per-player [trade_date, drop_date)
            # windows. The collective window for the dropped side is
            # [trade_date, latest_drop_among_received].
            recv_windows: Dict[str, Tuple[str, Optional[str]]] = {}
            latest_end: Optional[str] = None
            for pid in (row.get("_recv_player_ids") or []):
                nx = _next_out_player(team, pid, trade_iso)
                end_iso = nx["date"][:10] if nx else None
                recv_windows[str(pid)] = (trade_prefix, end_iso)
                if end_iso is not None:
                    latest_end = end_iso if (latest_end is None or end_iso > latest_end) else latest_end

            def _avg_ppg_window(name: str, start: str, end: Optional[str]) -> Optional[float]:
                games = _player_games(name)
                pts = [p for d, p in games if d >= start and (end is None or d < end)]
                if not pts:
                    return None
                return sum(pts) / len(pts)

            def _avg_ppg_pre5(name: str, before: str) -> Optional[float]:
                games = [(d, p) for d, p in _player_games(name) if d < before]
                if not games:
                    return None
                games.sort(key=lambda kv: kv[0], reverse=True)
                window = games[:5]
                return sum(p for _, p in window) / len(window)

            recv_on_team_avgs: List[float] = []
            recv_adj_on_team_avgs: List[float] = []
            recv_pre5_avgs: List[float] = []
            for pid in (row.get("_recv_player_ids") or []):
                name = _player_display(pid)
                start_i, end_i = recv_windows.get(str(pid), (trade_prefix, None))
                avg_on = _avg_ppg_window(name, start_i, end_i)
                if avg_on is not None:
                    recv_on_team_avgs.append(avg_on)
                    pos = _player_pos(name)
                    pos_a = pos_avg_map.get(pos or "", 0.0)
                    if pos_a and league_starter_avg:
                        recv_adj_on_team_avgs.append(avg_on * league_starter_avg / pos_a)
                    else:
                        recv_adj_on_team_avgs.append(avg_on)
                pre5 = _avg_ppg_pre5(name, trade_prefix)
                if pre5 is not None:
                    recv_pre5_avgs.append(pre5)

            # Phase 7D: fold in the PPG of players DRAFTED with received picks,
            # over their post-draft tenure on THIS team. Only when this team
            # actually made the selection (Final Team == team) — a pick flipped
            # again before the draft never became a player here. Undrafted
            # future picks contribute nothing. The drafted player's window
            # starts at the draft (late August of the pick year) and ends at
            # their next exit from this team; injured/bye/suspended weeks are
            # already absent from the nflverse game log the avg is built on.
            for _pm in (row.get("_recv_pick_meta") or []):
                try:
                    _pk = (int(_pm[0]), int(_pm[1]), _norm_team_name(_pm[2]))
                except Exception:
                    continue
                _drafted = _pick_to_drafted.get(_pk)
                if not _drafted:
                    continue
                _dpl, _dfinal, _dyear = _drafted
                if _dfinal != _norm_team_name(team):
                    continue  # pick was flipped before the draft
                _dstart = f"{_dyear}-08-28"
                _dsid = name_to_sid_local2.get(_dpl)
                _dend = None
                if _dsid:
                    _nxo = _next_out_player(team, _dsid, _dstart)
                    _dend = _nxo["date"][:10] if _nxo else None
                _davg = _avg_ppg_window(_dpl, _dstart, _dend)
                if _davg is not None:
                    recv_on_team_avgs.append(_davg)
                    _dpos = _player_pos(_dpl)
                    _dposa = pos_avg_map.get(_dpos or "", 0.0)
                    if _dposa and league_starter_avg:
                        recv_adj_on_team_avgs.append(_davg * league_starter_avg / _dposa)
                    else:
                        recv_adj_on_team_avgs.append(_davg)

            drop_over_avgs: List[float] = []
            drop_adj_avgs: List[float] = []
            # If no received player has been dropped yet, the window
            # is open-ended — use today as the upper bound so the
            # sent-side PPG still has a meaningful window.
            effective_end = latest_end or datetime.utcnow().date().isoformat()
            if effective_end:
                for pid in (row.get("_drop_player_ids") or []):
                    name = _player_display(pid)
                    avg_over = _avg_ppg_window(name, trade_prefix, effective_end)
                    if avg_over is not None:
                        drop_over_avgs.append(avg_over)
                        pos = _player_pos(name)
                        pos_a = pos_avg_map.get(pos or "", 0.0)
                        if pos_a and league_starter_avg:
                            drop_adj_avgs.append(avg_over * league_starter_avg / pos_a)
                        else:
                            drop_adj_avgs.append(avg_over)

            if recv_on_team_avgs:
                row["Avg PPG of received players on team"] = round(
                    sum(recv_on_team_avgs) / len(recv_on_team_avgs), 4
                )
            if drop_over_avgs:
                row["Avg PPG of sent players over same time"] = round(
                    sum(drop_over_avgs) / len(drop_over_avgs), 4
                )
            if recv_pre5_avgs:
                row["Avg PPG of received players in 5 games before trade"] = round(
                    sum(recv_pre5_avgs) / len(recv_pre5_avgs), 4
                )

            diff_avg = None
            if recv_on_team_avgs or drop_over_avgs:
                a = (sum(recv_on_team_avgs) / len(recv_on_team_avgs)) if recv_on_team_avgs else 0.0
                b = (sum(drop_over_avgs) / len(drop_over_avgs)) if drop_over_avgs else 0.0
                diff_avg = round(a - b, 4)
                row["Difference of averages"] = diff_avg
            adj_diff = None
            if recv_adj_on_team_avgs or drop_adj_avgs:
                a_adj = (sum(recv_adj_on_team_avgs) / len(recv_adj_on_team_avgs)) if recv_adj_on_team_avgs else 0.0
                b_adj = (sum(drop_adj_avgs) / len(drop_adj_avgs)) if drop_adj_avgs else 0.0
                adj_diff = round(a_adj - b_adj, 4)
                row["Difference of averages adjusted by position"] = adj_diff

            # ----- Trade addition value (V2, Item 7E) -----
            # Mirror the transaction "Player addition value" composite:
            #   adj_diff * (1 + pct_starts) * (1 + pct_starts_inj_adj)
            #     + CUFF_BONUS   (added once if ANY received player was a cuff
            #                     at the trade — same handcuff test as the
            #                     transaction "Cuff at time of pickup?").
            # The leverage multipliers use the received PLAYERS' % of starts
            # made while rostered on THIS team over their post-trade tenure
            # (trade -> next exit); the injury-adjusted variant divides by
            # injury/bye-free weeks. Players drafted from received picks feed
            # adj_diff (above) but not the leverage term.
            #
            # Pick value: future picks (next 3 seasons) are valued with the
            # tanking round weights and added in via _TRADE_PICK_COEFF so
            # pick-heavy hauls register. Current-season picks that get drafted
            # are already captured by adj_diff (7D), so only future capital is
            # added here -> no double count. This term applies even when
            # adj_diff is None (a pick-only haul is no longer flat 0).
            _pick_val = _TRADE_PICK_COEFF * float(row.get("_tank_fcap_delta") or 0.0)
            if adj_diff is not None:
                _pct_list: List[float] = []
                _pinj_list: List[float] = []
                for _rpid in (row.get("_recv_player_ids") or []):
                    _rnm = _player_display(_rpid)
                    if not _rnm:
                        continue
                    _wstart, _wend = recv_windows.get(str(_rpid), (trade_prefix, None))
                    _wk_n = _st_n = _iwk_n = _ist_n = 0
                    for _e in _pwfull_idx.get((team, _rnm), []):
                        _wkd = _e["wkd"]
                        if _wkd < _wstart:
                            continue
                        if _wend and _wkd >= _wend:
                            continue
                        _wk_n += 1
                        if _e["starter"]:
                            _st_n += 1
                        if _e["inj_free"]:
                            _iwk_n += 1
                            if _e["starter"]:
                                _ist_n += 1
                    if _wk_n > 0:
                        _pct_list.append(_st_n / _wk_n)
                    if _iwk_n > 0:
                        _pinj_list.append(_ist_n / _iwk_n)
                _pct_starts = (sum(_pct_list) / len(_pct_list)) if _pct_list else 0.0
                _pct_inj = (sum(_pinj_list) / len(_pinj_list)) if _pinj_list else 0.0
                _cuff_hit = any(
                    _recv_is_cuff(_player_display(_rpid), team, trade_prefix)
                    for _rpid in (row.get("_recv_player_ids") or [])
                    if _player_display(_rpid)
                )
                _cuff_bonus = _TRADE_CUFF_BONUS if _cuff_hit else 0.0
                row["Trade addition value"] = round(
                    adj_diff * (1.0 + _pct_starts) * (1.0 + _pct_inj)
                    + _cuff_bonus + _pick_val, 4)
            else:
                row["Trade addition value"] = round(_pick_val, 4)

            # --- Points Added / Lost / Net (+ per-week averages) ---
            # Received assets = received players + players THIS team drafted
            # with received picks (window starts at the draft). Per week,
            # Points Added = points of received assets that STARTED that week;
            # k = how many started. Points Lost = the top-k players-traded-away
            # by real NFL points that week (each sent asset once, capped at the
            # number sent) — the best plays forgone. Sent picks contribute the
            # player drafted with them. Averages divide by the matched weeks.
            _trade_dt10 = trade_iso[:10] if len(trade_iso) >= 10 else trade_iso
            _recv_assets: List[Tuple[str, str]] = []
            for _pid in (row.get("_recv_player_ids") or []):
                _nm = _player_display(_pid)
                if _nm:
                    _recv_assets.append((_nm, _trade_dt10))
            for _m in (row.get("_recv_pick_meta") or []):
                try:
                    _pk = (int(_m[0]), int(_m[1]), _norm_team_name(_m[2]))
                except Exception:
                    continue
                _dr = _pick_to_drafted.get(_pk)
                if _dr and _dr[1] == _norm_team_name(team):
                    _recv_assets.append((_dr[0], f"{_dr[2]}-08-28"))
            # Position-adjustment factor: scale an asset's points by the same
            # league_starter_avg / pos_avg normalizer used elsewhere (Item 4).
            def _afac(_n):
                _pa = pos_avg_map.get(_player_pos(_n) or "", 0.0)
                return (league_starter_avg / _pa) if (_pa and league_starter_avg) else 1.0
            # received: per week, list of (raw, position-adjusted) points
            _recv_week: Dict[Tuple[int, int], List[Tuple[float, float]]] = defaultdict(list)
            for _nm, _wstart in _recv_assets:
                _f = _afac(_nm)
                for (_yi, _wi, _pt, _wkd) in _started_idx.get((str(team), _nm), []):
                    if _wkd >= _wstart:
                        _recv_week[(_yi, _wi)].append((_pt, _pt * _f))
            # sent: per asset, {(year, week): (raw, position-adjusted)}
            _sent_logs: List[Dict[Tuple[int, int], Tuple[float, float]]] = []
            for _pid in (row.get("_drop_player_ids") or []):
                _f = _afac(_player_display(_pid))
                _sent_logs.append({_k2: (_v, _v * _f) for _k2, _v in _nfl_wk_pts(sid=_pid).items()})
            for _m in (row.get("_drop_pick_meta") or []):
                try:
                    _pk = (int(_m[0]), int(_m[1]), _norm_team_name(_m[2]))
                except Exception:
                    continue
                _dr = _pick_to_drafted.get(_pk)
                if _dr:
                    _f = _afac(_dr[0])
                    _sent_logs.append({_k2: (_v, _v * _f) for _k2, _v in _nfl_wk_pts(name=_dr[0]).items()})
            _tp_added = _tp_lost = _tadj_added = _tadj_lost = 0.0
            _tnwk = 0
            _wins_flipped = 0  # Item 3: net games the trade actually flipped
            _has_margin = False
            for _w, _rpts in _recv_week.items():
                _k = len(_rpts)
                if _k == 0:
                    continue
                _tnwk += 1
                _wk_added = sum(p[0] for p in _rpts)
                _tp_added += _wk_added
                _tadj_added += sum(p[1] for p in _rpts)
                # top-k players-traded-away by RAW points that week; sum their
                # raw AND position-adjusted points for the two Lost variants.
                _cand = sorted((_lg.get(_w, (0.0, 0.0)) for _lg in _sent_logs),
                               key=lambda x: x[0], reverse=True)[:_k]
                _wk_lost = sum(c[0] for c in _cand)
                _tp_lost += _wk_lost
                _tadj_lost += sum(c[1] for c in _cand)
                # Item 3: did this week's net swing flip the team's result? The
                # counterfactual margin (no trade) = actual margin − net points.
                _mg = _team_margin.get((str(team), int(_w[0]), int(_w[1])))
                if _mg is not None:
                    _has_margin = True
                    _net_wk = _wk_added - _wk_lost
                    _wins_flipped += int(_mg > 0) - int((_mg - _net_wk) > 0)
            # Raw win-impact (games the trade flipped). Folded into the
            # "Trade impact score" composite in a post-loop pass below.
            row["_tpi_wins"] = float(_wins_flipped) if _has_margin else None
            _tnet = _tp_added - _tp_lost
            _tadj_net = _tadj_added - _tadj_lost
            row["Points added"] = round(_tp_added, 2)
            row["Points lost"] = round(_tp_lost, 2)
            row["Net points"] = round(_tnet, 2)
            row["Avg points added"] = round(_tp_added / _tnwk, 2) if _tnwk else 0.0
            row["Avg points lost"] = round(_tp_lost / _tnwk, 2) if _tnwk else 0.0
            row["Avg net points"] = round(_tnet / _tnwk, 2) if _tnwk else 0.0
            row["Avg points added adjusted by position"] = round(_tadj_added / _tnwk, 2) if _tnwk else 0.0
            row["Avg points lost adjusted by position"] = round(_tadj_lost / _tnwk, 2) if _tnwk else 0.0
            row["Avg net points adjusted by position"] = round(_tadj_net / _tnwk, 2) if _tnwk else 0.0

            # ----- (d) Return-from-trades chain (V2: three buckets) -----
            # Each received asset terminates in one of:
            #   - Assets retained now    (still on roster, no exit)
            #   - Assets traded away     (next exit was a trade)
            #   - Assets dropped to FA   (next exit was a drop, players only)
            # Picks never drop to FA — they're either used (drafted)
            # or traded, so the FA classification is player-only.
            recv_keys, recv_disp = _trade_assets_received(row)
            # Item 3: for each received asset that this team later RE-TRADED,
            # record (downstream trade idx, the asset, the re-trade date) so the
            # win-impact can be credited downstream, weighted by the asset's KTC
            # value SHARE of that downstream trade (no separate hop discount —
            # the share itself decays the credit each link).
            _tpi_down_list: List[Tuple[int, Tuple[str, Any], str]] = []
            retained: List[str] = []
            traded_away: List[str] = []
            dropped_to_fa: List[str] = []
            return_immediate: List[str] = []
            additional_immediate: List[str] = []
            return_full: List[str] = []
            visited: Set[Tuple] = set()

            def _next_event(key_tuple: Tuple[str, Any], after: str) -> Optional[Dict[str, Any]]:
                if key_tuple[0] == "player":
                    return _next_out_player(team, key_tuple[1], after)
                return _next_out_pick(team, key_tuple[1], after)

            def _next_trade_received(nx_idx: int) -> Tuple[List[Tuple[str, Any]], List[str]]:
                return _trade_assets_received(trades_rows[nx_idx])

            def _next_trade_dropped(nx_idx: int) -> Tuple[List[Tuple[str, Any]], List[str]]:
                tr_row = trades_rows[nx_idx]
                keys: List[Tuple[str, Any]] = []
                disp: List[str] = []
                for pid in (tr_row.get("_drop_player_ids") or []):
                    keys.append(("player", str(pid)))
                    disp.append(_player_display(pid))
                for pmeta in (tr_row.get("_drop_pick_meta") or []):
                    keys.append(("pick", tuple(pmeta)))
                    disp.append(_pick_display(tuple(pmeta)))
                return keys, disp

            for asset_key, asset_disp in zip(recv_keys, recv_disp):
                nx = _next_event(asset_key, trade_iso)
                # For player assets, also check FA drop events.
                fa_drop_date: Optional[str] = None
                if asset_key[0] == "player":
                    fa_drop_date = _next_fa_drop_player(team, asset_key[1], trade_iso)

                # Choose the EARLIER of trade-out vs FA-drop as the exit.
                trade_date = nx["date"] if nx else None
                if trade_date and fa_drop_date:
                    if fa_drop_date < trade_date:
                        # FA drop came first → chain ends here.
                        dropped_to_fa.append(asset_disp)
                        continue
                    # else trade came first → fall through to trade handling
                elif fa_drop_date and not trade_date:
                    dropped_to_fa.append(asset_disp)
                    continue

                if nx is None:
                    retained.append(asset_disp)
                    continue

                traded_away.append(asset_disp)
                try:
                    _tpi_down_list.append((int(nx["tx_idx"]), asset_key, str(nx["date"])))
                except Exception:
                    pass
                nx_keys, nx_disp = _next_trade_received(nx["tx_idx"])
                return_immediate.extend(nx_disp)
                # Additional assets traded away in the next trade
                # alongside our asset (anything else this team gave up).
                dr_keys, dr_disp = _next_trade_dropped(nx["tx_idx"])
                for dk, dd in zip(dr_keys, dr_disp):
                    if dk != asset_key:
                        additional_immediate.append(dd)
                # Recurse to build full chain.
                queue: List[Tuple[Tuple[str, Any], str, str]] = []
                for k, d in zip(nx_keys, nx_disp):
                    queue.append((k, d, nx["date"]))
                while queue:
                    cur_k, cur_d, cur_after = queue.pop(0)
                    if cur_k in visited:
                        continue
                    visited.add(cur_k)
                    return_full.append(cur_d)
                    nxt = _next_event(cur_k, cur_after)
                    if nxt is None:
                        continue
                    more_k, more_d = _next_trade_received(nxt["tx_idx"])
                    for k2, d2 in zip(more_k, more_d):
                        queue.append((k2, d2, nxt["date"]))

            row["_tpi_down"] = _tpi_down_list  # Item 3 downstream re-trade links
            if retained:
                row["Assets retained now"] = "; ".join(retained)
            if traded_away:
                row["Assets traded away"] = "; ".join(traded_away)
            if dropped_to_fa:
                row["Assets dropped to FA"] = "; ".join(dropped_to_fa)
            if return_immediate:
                row["Return from trades"] = "; ".join(dict.fromkeys(return_immediate))
            if additional_immediate:
                row["Additional assets traded away in those deals"] = "; ".join(dict.fromkeys(additional_immediate))
            if return_full:
                row["Return from trades of trades...of trades. Keep going until present day"] = "; ".join(dict.fromkeys(return_full))

        # --- Item 3: "Trade impact score" composite. ---
        # Win impact = the games the trade flipped (counterfactual weekly margins
        # from received-asset production) PLUS a share of the games flipped by
        # LATER trades that re-used the received assets. Each downstream trade is
        # credited by the FRACTION of its sent-side KTC value (on the later
        # trade's day) that came from THIS trade's assets — so a minor add-on
        # later bundled for a stud earns only its small share, recursively. That
        # downstream-aware win impact is HEAVILY weighted (per user) and blended
        # (z-scored across all trades) with realized production (Avg net points),
        # overall trade value incl. picks (Trade addition value), future pick
        # capital (Pick value received), and youth (−Asset diff in avg age) into
        # one continuous, percentile-rankable score. The output uses no KTC
        # directly (KTC only proportions the downstream credit); all signals
        # credit a rebuild, so tank trades aren't punished.
        try:
            def _tpi_f(_v):
                try:
                    return float(_v)
                except Exception:
                    return None
            try:
                from lotg_support.ktc import asset_value_at as _tpi_kv
            except Exception:
                _tpi_kv = None
            try:
                _tpi_idx = _ktc_idx  # may be undefined if the KTC pass was skipped
            except NameError:
                _tpi_idx = None

            def _tpi_pdate(_s):
                try:
                    return datetime.fromisoformat(str(_s).replace("Z", "+00:00")).date()
                except Exception:
                    return None

            # Phase 12 #45a: count win-impact KTC lookups that come back EMPTY
            # (no history at that date) vs total. A high empty rate means the
            # build fetched KTC incompletely (rate-limit / cold cache), which
            # silently 0-fills the downstream-credit proportioning and makes
            # "Trade impact score" non-deterministic across builds. Logged after
            # the loop so an incomplete build is obvious instead of silent.
            _tpi_ktc_stats = {"total": 0, "empty": 0}

            def _tpi_kv_counted(_pl, _pid, _d):
                _v = _tpi_kv(_pl, _pid, _d, _tpi_idx)
                _tpi_ktc_stats["total"] += 1
                if _v is None:
                    _tpi_ktc_stats["empty"] += 1
                return float(_v or 0.0)

            def _tpi_asset_ktc(_ak, _jidx, _dstr):
                if _tpi_kv is None or _tpi_idx is None:
                    return 0.0
                _d = _tpi_pdate(_dstr)
                if _d is None:
                    return 0.0
                try:
                    if _ak[0] == "player":
                        return _tpi_kv_counted(None, str(_ak[1]), _d)
                    _rj = trades_rows[_jidx]
                    for _m, _l in zip(_rj.get("_drop_pick_meta") or [], _rj.get("_drop_picks") or []):
                        if tuple(_m) == tuple(_ak[1]):
                            return _tpi_kv_counted(str(_l), None, _d)
                except Exception:
                    return 0.0
                return 0.0

            _sent_tot_cache: Dict[int, float] = {}

            def _tpi_sent_total(_jidx):
                if _jidx in _sent_tot_cache:
                    return _sent_tot_cache[_jidx]
                _tot = 0.0
                if _tpi_kv is not None and _tpi_idx is not None:
                    _rj = trades_rows[_jidx]
                    _d = _tpi_pdate(_rj.get("Date"))
                    if _d is not None:
                        for _pid in (_rj.get("_drop_player_ids") or []):
                            _tot += _tpi_kv_counted(None, str(_pid), _d)
                        for _l in (_rj.get("_drop_picks") or []):
                            _tot += _tpi_kv_counted(str(_l), None, _d)
                _sent_tot_cache[_jidx] = _tot
                return _tot

            _winimpact: Dict[int, float] = {}
            _wi_active: Set[int] = set()

            def _tpi_total(_i):
                if _i in _winimpact:
                    return _winimpact[_i]
                if _i in _wi_active:
                    return 0.0  # cycle guard (trades are time-ordered; shouldn't fire)
                _wi_active.add(_i)
                _val = _tpi_f(trades_rows[_i].get("_tpi_wins")) or 0.0
                _by_j: Dict[int, List[Tuple[Tuple[str, Any], str]]] = defaultdict(list)
                for (_j, _ak, _ds) in (trades_rows[_i].get("_tpi_down") or []):
                    if 0 <= _j < len(trades_rows):
                        _by_j[_j].append((_ak, _ds))
                for _j, _assets in _by_j.items():
                    _st = _tpi_sent_total(_j)
                    if _st > 0:
                        _num = sum(_tpi_asset_ktc(_ak, _j, _ds) for (_ak, _ds) in _assets)
                        _share = max(0.0, min(1.0, _num / _st))
                        if _share > 0:
                            _val += _share * _tpi_total(_j)
                _wi_active.discard(_i)
                _winimpact[_i] = _val
                return _val

            for _i in range(len(trades_rows)):
                trades_rows[_i]["_tpi_winimpact"] = _tpi_total(_i)

            _tt = _tpi_ktc_stats["total"]
            _te = _tpi_ktc_stats["empty"]
            if _tt:
                _rate = 100.0 * _te / _tt
                _log(debug, f"[{_now_iso()}] INFO KTC win-impact lookups: {_te}/{_tt} empty ({_rate:.1f}%)")
                if _rate > 25.0:
                    _log(debug, f"[{_now_iso()}] WARN KTC win-impact incomplete ({_te}/{_tt} empty) "
                                f"— Trade impact score may be non-deterministic this build; check the KTC cache/fetch")

            _tpi_specs = [
                ("_tpi_winimpact", 2.0, True),                        # downstream-aware win impact (HEAVY)
                ("Avg net points", 0.8, False),                       # realized production
                ("Trade addition value", 0.5, False),                 # value incl. picks
                ("Pick value received", 0.5, True),                   # future capital
                ("Asset difference in average age", -0.3, False),     # youth (neg weight)
            ]
            _tpi_stats = {}
            for _col, _w, _fill0 in _tpi_specs:
                _vals = []
                for _r in trades_rows:
                    _fv = _tpi_f(_r.get(_col))
                    if _fv is None and _fill0:
                        _fv = 0.0
                    if _fv is not None:
                        _vals.append(_fv)
                if _vals:
                    _mu = sum(_vals) / len(_vals)
                    _sd = (sum((x - _mu) ** 2 for x in _vals) / len(_vals)) ** 0.5 or 1.0
                else:
                    _mu, _sd = 0.0, 1.0
                _tpi_stats[_col] = (_mu, _sd)
            for _r in trades_rows:
                _score = 0.0
                for _col, _w, _fill0 in _tpi_specs:
                    _fv = _tpi_f(_r.get(_col))
                    if _fv is None:
                        if _fill0:
                            _fv = 0.0
                        else:
                            continue
                    _mu, _sd = _tpi_stats[_col]
                    _score += _w * ((_fv - _mu) / _sd)
                _r["Trade impact score"] = round(1500.0 * _score, 1)
                for _k in ("_tpi_wins", "_tpi_down", "_tpi_winimpact"):
                    _r.pop(_k, None)
        except Exception as e:
            _log_exc(debug, "team_perf_improvement_item3", e)

        # 4) Rebuild tr DataFrame with the polished trades_rows.
        tr = pd.DataFrame(trades_rows)
    except Exception as e:
        _log_exc(debug, "trades_polish_v2", e)


    # --------------------------
    # Player-week derived columns (deltas, tenure, awards) + Hardship
    # --------------------------
    if not pw.empty:
        pw["Year"] = pd.to_numeric(pw["Year"], errors="coerce").astype("Int64")
        pw["Week"] = pd.to_numeric(pw["Week"], errors="coerce").astype("Int64")
        pw["Points"] = pd.to_numeric(pw["Points"], errors="coerce").fillna(0.0)

        pw = pw.sort_values(["Player", "Year", "Week"]).reset_index(drop=True)

        # Fill missing NFL team with nearest known value for same player-season, then player-level mode.
        if "NFL team" in pw.columns:
            pw["NFL team"] = pw["NFL team"].replace("", np.nan)
            pw["NFL team"] = pw.groupby(["Player", "Year"])["NFL team"].transform(lambda s: s.ffill().bfill())
            team_mode = pw.groupby("Player")["NFL team"].agg(lambda s: s.dropna().mode().iloc[0] if len(s.dropna()) else np.nan)
            pw["NFL team"] = pw["NFL team"].fillna(pw["Player"].map(team_mode))
            pw["NFL team"] = pw["NFL team"].map(_norm_team)

        active = ~pw[["Injury?", "Suspension?", "Bye?"]].fillna(False).any(axis=1)

        # Change from previous active week (bench weeks count if active)
        last_active_pts: Dict[str, float] = {}
        out_prev = []
        for i, row in pw.iterrows():
            k = row["Player"]
            out_prev.append((float(row["Points"]) - last_active_pts[k]) if k in last_active_pts else None)
            if bool(active.iloc[i]):
                last_active_pts[k] = float(row["Points"])
        pw["Change from previous week"] = out_prev

        # Previous 5 active weeks avg (spans seasons)
        windows: Dict[str, deque] = defaultdict(lambda: deque(maxlen=5))
        out_prev5 = []
        for i, row in pw.iterrows():
            k = row["Player"]
            q = windows[k]
            out_prev5.append((float(row["Points"]) - (sum(q) / 5)) if len(q) == 5 else None)
            if bool(active.iloc[i]):
                q.append(float(row["Points"]))
        pw["Change from previous 5 weeks avg"] = out_prev5

        # Career avg to that point (active weeks only)
        sums: Dict[str, float] = defaultdict(float)
        counts: Dict[str, int] = defaultdict(int)
        out_cavg = []
        for i, row in pw.iterrows():
            k = row["Player"]
            out_cavg.append((float(row["Points"]) - (sums[k] / counts[k])) if counts[k] > 0 else None)
            if bool(active.iloc[i]):
                sums[k] += float(row["Points"])
                counts[k] += 1
        pw["Change from career average to that point"] = out_cavg

        # Overall career avg (active weeks only)
        try:
            full_avg = pw.loc[active].groupby("Player")["Points"].mean()
            pw["Change from overall career average"] = pw["Points"] - pw["Player"].map(full_avg)
        except Exception as e:
            _log_exc(repo_root / "exports/raw/build_debug.log", "overall_career_avg", e)
            pw["Change from overall career average"] = None

        # Team tenure + bench streaks (bench streak spans seasons)
        pw = pw.sort_values(["Team", "Player", "Year", "Week"]).reset_index(drop=True)
        stats: Dict[Tuple[str, str], Dict[str, Any]] = {}
        for i, row in pw.iterrows():
            key = (str(row["Team"]), str(row["Player"]))
            st = stats.get(key, {
                "weeks": 0,
                "start_all": 0,
                "bench_all": 0,
                "season": None,
                "start_season": 0,
                "bench_season": 0,
                "bench_streak": 0,
                "bench_streak_ex": 0
            })

            if st["season"] != row["Year"]:
                st["season"] = row["Year"]
                st["start_season"] = 0
                st["bench_season"] = 0

            st["weeks"] += 1
            is_starter = (row["Starter/Bench"] == "Starter")
            inactive = bool((row.get("Injury?") or False) or (row.get("Suspension?") or False) or (row.get("Bye?") or False))

            if is_starter:
                pw.at[i, "Number of consecutive weeks on bench before start (if starter)"] = st["bench_streak"]
                pw.at[i, "Number of consecutive weeks on bench before start excluding injury/bye (if starter)"] = st["bench_streak_ex"]
                st["bench_streak"] = 0
                st["bench_streak_ex"] = 0
                st["start_all"] += 1
                st["start_season"] += 1
            else:
                st["bench_all"] += 1
                st["bench_season"] += 1
                st["bench_streak"] += 1
                if not inactive:
                    st["bench_streak_ex"] += 1

            pw.at[i, "Number of weeks on team"] = st["weeks"]
            pw.at[i, "Total weeks as team starter to that point"] = st["start_all"]
            pw.at[i, "Total weeks on bench to that point"] = st["bench_all"]
            pw.at[i, "Total weeks as team starter on that team this season"] = st["start_season"]
            pw.at[i, "Total weeks on bench on that team this season"] = st["bench_season"]

            stats[key] = st

        # Awards (league + team). Ties -> all winners.
        # We compute off pw itself to avoid any mismatched ids.
        pw = pw.sort_values(["Year", "Week", "Team", "Player"]).reset_index(drop=True)

        award_cols = [
            "Player of the week?",
            "QB of the week?",
            "RB of the week?",
            "WR of the week?",
            "TE of the week?",
            "Benchwarmer of the week?",
            "Bench QB of the week?",
            "Bench RB of the week?",
            "Bench WR of the week?",
            "Bench TE of the week?",
            "Highest starter on team?",
            "Lowest starter on team?",
            "Captain?",
        ]

        def _set_flag(mask, col):
            pw.loc[mask, col] = 1
            pw.loc[~mask, col] = pw.loc[~mask, col].fillna(0)

        # league-level player of week among starters.
        # Pick exactly ONE winner per (Year, Week) per award. Tie-breaker rules:
        #   - Player of the week (max score): on tie, alphabetical Player name first
        #   - Benchwarmer (min score): if 2+ starters tie at 0 points, no winner
        #     (it's not meaningful to call out one of many who all sat out). If 2+
        #     tie at a non-zero low, alphabetical first wins.
        starters = pw["Starter/Bench"] == "Starter"
        pos_col = "Position" if "Position" in pw.columns else "Position started in (if starter)"
        pos_series = pw[pos_col].astype(str).str.upper()

        def _pick_one(sub_df: pd.DataFrame, by_max: bool) -> Optional[int]:
            """Return the row index of the single award winner, or None."""
            if sub_df.empty:
                return None
            extreme = sub_df["Points"].max() if by_max else sub_df["Points"].min()
            tied = sub_df[sub_df["Points"] == extreme]
            # Multi-way tie at 0 for min: don't award.
            if (not by_max) and float(extreme) == 0.0 and len(tied) >= 2:
                return None
            tied_sorted = tied.assign(_p=tied["Player"].astype(str)).sort_values("_p", kind="stable")
            return int(tied_sorted.index[0])

        for (yr, wk), g in pw.groupby(["Year", "Week"]):
            sg = g[starters.loc[g.index]]
            if sg.empty:
                continue
            # Player of the week
            idx = _pick_one(sg, by_max=True)
            if idx is not None:
                _set_flag(pw.index == idx, "Player of the week?")
            # Captain of the week: the starter, league-wide, who supplied the
            # highest share of his team's PF (most indispensable single player).
            _sh = pd.to_numeric(sg.get("% of points (if starter)"), errors="coerce")
            if _sh is not None and _sh.notna().any():
                _mx = _sh.max()
                _cand = sg.loc[_sh[_sh == _mx].index]
                _cand = _cand.assign(_p=_cand["Player"].astype(str)).sort_values("_p", kind="stable")
                _set_flag(pw.index == int(_cand.index[0]), "Captain?")
            # Benchwarmer of the week
            idx = _pick_one(sg, by_max=False)
            if idx is not None:
                _set_flag(pw.index == idx, "Benchwarmer of the week?")

            for pos, col in [("QB", "QB of the week?"), ("RB", "RB of the week?"), ("WR", "WR of the week?"), ("TE", "TE of the week?")]:
                pg = sg[pos_series.loc[sg.index] == pos]
                idx = _pick_one(pg, by_max=True)
                if idx is not None:
                    _set_flag(pw.index == idx, col)

            # bench position awards (bench only)
            bg = g[~starters.loc[g.index]]
            for pos, col in [
                ("QB", "Bench QB of the week?"),
                ("RB", "Bench RB of the week?"),
                ("WR", "Bench WR of the week?"),
                ("TE", "Bench TE of the week?"),
            ]:
                pg = bg[pos_series.loc[bg.index] == pos]
                idx = _pick_one(pg, by_max=True)
                if idx is not None:
                    _set_flag(pw.index == idx, col)

            # team-level awards: highest/lowest starter per team per week
            for team, tg in sg.groupby("Team"):
                mx_t = tg["Points"].max()
                mn_t = tg["Points"].min()
                _set_flag(
                    (pw["Year"] == yr)
                    & (pw["Week"] == wk)
                    & (pw["Team"] == team)
                    & starters
                    & (pw["Points"] == mx_t),
                    "Highest starter on team?",
                )
                _set_flag(
                    (pw["Year"] == yr)
                    & (pw["Week"] == wk)
                    & (pw["Team"] == team)
                    & starters
                    & (pw["Points"] == mn_t),
                    "Lowest starter on team?",
                )

        pw[award_cols] = pw[award_cols].fillna(0)

        # Player streaks (award + point-threshold). ALL-TIME (continuous across
        # seasons) and terminal-encoded so each run lists once. Per user, these
        # SKIP bye / injury / suspension weeks: a week the player didn't play
        # doesn't count toward a streak and doesn't break it (the run bridges
        # across it; those cells read 'N/A').
        try:
            _grp_key = "Player ID" if "Player ID" in pw.columns else "Player"
            pw = pw.sort_values([_grp_key, "Year", "Week"]).reset_index(drop=True)
            _played = ~(
                pw.get("Injury?", pd.Series(False, index=pw.index)).fillna(False).astype(bool)
                | pw.get("Suspension?", pd.Series(False, index=pw.index)).fillna(False).astype(bool)
                | pw.get("Bye?", pd.Series(False, index=pw.index)).fillna(False).astype(bool)
            ).to_numpy()
            # Award streaks: "<award>?" -> "<award> streak"
            _specs = {}
            for _fc in award_cols:
                _specs[_fc[:-1] + " streak"] = (pd.to_numeric(pw[_fc], errors="coerce").fillna(0) == 1).to_numpy()
            # Point-threshold streaks: >= 10/20/30/40/50 (starter OR bench)
            _pts = pd.to_numeric(pw["Points"], errors="coerce").fillna(0.0).to_numpy()
            for _t in (10, 20, 30, 40, 50):
                _specs[f"{_t}+ point streak"] = (_pts >= _t)
            pw = _encode_player_streaks(pw, _grp_key, ["Year", "Week"], _played, _specs)
        except Exception as e:
            _log_exc(debug, "player_streaks", e)

        # Hardship engine (Phase 2 rebuild + Phase 2A refinement +
        # NFLverse-seeded baseline fix).
        #
        # Per-player expected points = mean of last 5 ACTIVE NFL games
        # (excluding the most recent active game to avoid using a
        # potentially-injury-shortened game as part of its own baseline).
        # "Active" = the player actually played in the NFL that week
        # (positive games appear in nflverse). Bye / Injury / Suspension
        # weeks are excluded from baseline.
        #
        # Critically: baseline uses BOTH LOTG-rostered weeks (from pw)
        # AND NFLverse-only games (player played in NFL but wasn't on
        # any LOTG roster that week). Without the NFLverse seed,
        # newly-picked-up players and recently-suspended players (e.g.
        # Deshaun Watson 2022 wk3: 2 yrs in pw all-suspended, but he
        # had 5 active games in 2020 pre-suspension) get an empty
        # baseline and contribute 0 to Hardship — which understates
        # the real injury hit.
        #
        # starter_hist tracks last 5 active weeks' starter status;
        # NFLverse-only games count as "not starter" (player wasn't on
        # any LOTG roster).
        pw = pw.sort_values(["Player", "Year", "Week"]).reset_index(drop=True)

        # Build a per-player chronological NFL game log from nfl_log_by_sid
        # (already populated season-by-season above). Each entry has
        # year/week/points/_wk_date. Index by player display name so we
        # can look up alongside pw which keys by name.
        name_to_sid_h: Dict[str, str] = {}
        for _sid_h, _meta_h in pid_meta.items():
            _fn = (_meta_h or {}).get("full_name")
            if _fn:
                name_to_sid_h.setdefault(str(_fn), str(_sid_h))

        nfl_games_by_name: Dict[str, List[Tuple[int, int, float]]] = defaultdict(list)
        for _name, _sid in name_to_sid_h.items():
            for _e in nfl_log_by_sid.get(_sid, []):
                try:
                    _yr_h = int(_e.get("year"))
                    _wk_h = int(_e.get("week"))
                    _pts_h = float(_e.get("points") or 0.0)
                except Exception:
                    continue
                nfl_games_by_name[_name].append((_yr_h, _wk_h, _pts_h))
        for _name in nfl_games_by_name:
            nfl_games_by_name[_name].sort()

        # Pre-seed each player's history deque with NFLverse games BEFORE
        # their first pw row. Then as we iterate pw chronologically, also
        # inject any nflverse-only games (year, week) that fall between
        # the prior pw row and the current one. Dedup logic: a pw row
        # always represents the (year, week) authoritatively (it captures
        # the LOTG-rostered status + starter flag), so we skip nfl entries
        # that match an existing pw row's (year, week) for the same player.
        first_pw_yrwk: Dict[str, Tuple[int, int]] = {}
        for _, row in pw.iterrows():
            _p = row["Player"]
            _y = int(row["Year"]) if pd.notna(row.get("Year")) else None
            _w = int(row["Week"]) if pd.notna(row.get("Week")) else None
            if _y is None or _w is None:
                continue
            if _p not in first_pw_yrwk:
                first_pw_yrwk[_p] = (_y, _w)

        # Map (player, year, week) -> True if pw has a row (to skip dup nfl entries)
        pw_yrwk_set: Set[Tuple[str, int, int]] = set()
        for _, row in pw.iterrows():
            _p = row["Player"]
            _y = int(row["Year"]) if pd.notna(row.get("Year")) else None
            _w = int(row["Week"]) if pd.notna(row.get("Week")) else None
            if _y is None or _w is None:
                continue
            pw_yrwk_set.add((_p, _y, _w))

        last6: Dict[str, deque] = defaultdict(lambda: deque(maxlen=6))
        starter_hist: Dict[str, deque] = defaultdict(lambda: deque(maxlen=6))
        # Track an iterator position into nfl_games_by_name per player.
        nfl_cursor: Dict[str, int] = defaultdict(int)
        # Seed deques with NFLverse games BEFORE the player's first pw row.
        for _p, (_fy, _fw) in first_pw_yrwk.items():
            games = nfl_games_by_name.get(_p, [])
            i_cur = 0
            for (gy, gw, gp) in games:
                if (gy, gw) < (_fy, _fw):
                    # NFLverse-only pre-pw game: append to history.
                    # Skip if pw will have a matching row for this (year, week)
                    # — pw is authoritative.
                    if (_p, gy, gw) in pw_yrwk_set:
                        i_cur += 1
                        continue
                    last6[_p].append(gp)
                    starter_hist[_p].append(0)  # not on LOTG roster -> not starter
                    i_cur += 1
                else:
                    break
            nfl_cursor[_p] = i_cur

        exp_points: List[Optional[float]] = [None] * len(pw)
        points_lost: List[float] = [0.0] * len(pw)
        starter_adj_lost: List[float] = [0.0] * len(pw)
        # 1 when, at this row, the player counts as a "starter" under the
        # SAME heuristic Starter-adjusted Hardship uses (starter_pct > 0 over
        # the SA baseline window). Used for "Weeks of starter injuries/
        # suspensions".
        was_recent_starter: List[int] = [0] * len(pw)

        # Early-season forward-looking starter capture (user spec). In the very
        # first weeks of the first season a player has no backward LOTG active
        # history, so starter_pct is 0 and an injured starter contributes ~0 to
        # Starter-adjusted Hardship. To fix that we pre-build each player's
        # ordered (year, active, is_starter) sequence; when the backward window
        # is empty we look FORWARD to their next few active weeks (same season)
        # to recover their starter share. pw is sorted by Player, Year, Week so
        # rows for a player arrive in order and align with this sequence.
        player_seq: Dict[str, List[Tuple[int, bool, int]]] = defaultdict(list)
        for _, _r in pw.iterrows():
            _i2 = bool(_r.get("Injury?") or False)
            _s2 = bool(_r.get("Suspension?") or False)
            _b2 = bool(_r.get("Bye?") or False)
            _act = not (_i2 or _s2 or _b2)
            _st = 1 if (str(_r.get("Starter/Bench") or "").strip().lower() == "starter") else 0
            _yr2 = _to_int(_r.get("Year"), None)
            player_seq[_r["Player"]].append((_yr2 if _yr2 is not None else -1, _act, _st))
        player_pos: Dict[str, int] = defaultdict(int)
        lotg_active_seen: Dict[str, int] = defaultdict(int)

        def _forward_starter_frac(pl: str, pos: int, year: Optional[int]) -> Optional[float]:
            """Starter share over the player's next up-to-5 active weeks in the
            same season (rows strictly after `pos`)."""
            seq = player_seq.get(pl, [])
            flags: List[int] = []
            for j in range(pos + 1, len(seq)):
                yj, aj, sj = seq[j]
                if year is not None and yj != year:
                    break
                if aj:
                    flags.append(sj)
                    if len(flags) >= 5:
                        break
            if not flags:
                return None
            return sum(flags) / len(flags)

        prev_yrwk_by_player: Dict[str, Tuple[int, int]] = {}

        for i, row in pw.iterrows():
            player = row["Player"]
            pts = float(row["Points"])
            inj = bool(row.get("Injury?") or False)
            susp = bool(row.get("Suspension?") or False)
            bye = bool(row.get("Bye?") or False)
            is_starter = (str(row.get("Starter/Bench") or "").strip().lower() == "starter")
            yr = int(row["Year"]) if pd.notna(row.get("Year")) else None
            wk = int(row["Week"]) if pd.notna(row.get("Week")) else None
            hist = last6[player]
            s_hist = starter_hist[player]

            # Inject any NFLverse-only games between this player's previous
            # pw row and the current pw row (catches mid-season games where
            # the player was on no LOTG roster — they played in NFL, scored
            # points, but didn't appear in pw).
            if yr is not None and wk is not None:
                games = nfl_games_by_name.get(player, [])
                cur_idx = nfl_cursor[player]
                while cur_idx < len(games):
                    gy, gw, gp = games[cur_idx]
                    if (gy, gw) >= (yr, wk):
                        break
                    cur_idx += 1
                    if (player, gy, gw) in pw_yrwk_set:
                        continue  # pw will/did supply this week
                    # NFLverse-only active week: append.
                    hist.append(gp)
                    s_hist.append(0)
                nfl_cursor[player] = cur_idx

            # Baseline: drop most recent active week when ≥ 2; single
            # fallback for N=1; None for empty history.
            hist_list = list(hist)
            s_hist_list = list(s_hist)
            if len(hist_list) >= 2:
                baseline_pts = hist_list[:-1]
                baseline_starter = s_hist_list[:-1]
                expected = sum(baseline_pts) / len(baseline_pts)
                starter_pct = (sum(baseline_starter) / len(baseline_starter)) if baseline_starter else 0.0
            elif len(hist_list) == 1:
                expected = hist_list[0]
                starter_pct = float(s_hist_list[0]) if s_hist_list else 0.0
            else:
                expected = None
                starter_pct = 0.0
            exp_points[i] = expected
            # Effective starter share for this week. Normally starter_pct (the
            # SA baseline started-share). Early-season fallback: if the backward
            # window gives no starter signal (starter_pct == 0) AND the player
            # has < 5 prior LOTG active weeks (i.e. not enough history to judge),
            # look FORWARD at their next active weeks this season. Week 1 is left
            # at 0 by design (no signal we trust for the opening week).
            eff_starter_pct = float(starter_pct)
            if (
                eff_starter_pct == 0.0
                and wk is not None and wk >= 2
                and lotg_active_seen[player] < 5
            ):
                fwd = _forward_starter_frac(player, player_pos[player], yr)
                if fwd is not None:
                    eff_starter_pct = float(fwd)
            # Same starter heuristic as Starter-adjusted Hardship (now including
            # the early-season forward fallback): the player counts as a starter
            # for this week when eff_starter_pct > 0. Drives "Weeks of starter
            # injuries/suspensions" too, keeping the two in lockstep.
            was_recent_starter[i] = 1 if eff_starter_pct > 0.0 else 0
            missed = (pts == 0.0) and (inj or susp) and (not bye)
            if missed and expected is not None:
                # Clamp expected ≥ 0 (negative-baseline edge case).
                exp_clamped = max(0.0, float(expected))
                points_lost[i] = exp_clamped
                starter_adj_lost[i] = exp_clamped * eff_starter_pct
            # Advance active-week history when LOTG status is active. We
            # skip the nflverse dedup here because the current pw row is
            # the authoritative entry for this (year, week).
            if (not inj) and (not susp) and (not bye):
                hist.append(pts)
                s_hist.append(1 if is_starter else 0)
                lotg_active_seen[player] += 1
            player_pos[player] += 1
            prev_yrwk_by_player[player] = (yr, wk) if (yr is not None and wk is not None) else prev_yrwk_by_player.get(player)
        pw["_expected_points_if_healthy"] = exp_points
        pw["_points_lost_inj_susp"] = points_lost
        pw["_starter_adj_points_lost"] = starter_adj_lost
        pw["_was_recent_starter_injsusp"] = was_recent_starter

        # Healthy-lineup score per team-week for the "Loss from hardship?" flag.
        # Pool = the team's ACTUAL STARTERS (at their real points) + the hurt
        # would-be-starters who MISSED (injury/suspension, 0 pts), each subbed in
        # at their STARTER-ADJUSTED hardship value. compute_optimal_lineup then
        # picks the best valid lineup from that pool — so a hurt player only
        # helps by displacing a weaker actual starter (bounded to the lineup
        # slots, nets the replacement). Healthy bench players are deliberately
        # EXCLUDED: this asks "what if their hurt guys were available?", NOT
        # "what if they had also start/sat optimally".
        _healthy_opt_by_tw: Dict[Tuple[str, str, str], float] = {}
        try:
            _need = {"Team", "Year", "Week", "Position", "Points", "Injury?",
                     "Suspension?", "Starter/Bench", "Player ID", "_starter_adj_points_lost"}
            if _need.issubset(pw.columns):
                _pwm = pw[list(_need)].copy()
                _pwm["_pt"] = pd.to_numeric(_pwm["Points"], errors="coerce").fillna(0.0)
                _pwm["_sa"] = pd.to_numeric(_pwm["_starter_adj_points_lost"], errors="coerce")
                _pwm["_st"] = _pwm["Starter/Bench"].astype(str).str.strip().str.lower().eq("starter")
                _pwm["_hurt"] = (
                    (_pwm["Injury?"].astype(str).str.strip().str.lower().isin(["true", "1"])
                     | _pwm["Suspension?"].astype(str).str.strip().str.lower().isin(["true", "1"]))
                    & (_pwm["_pt"] == 0.0) & _pwm["_sa"].notna() & (_pwm["_sa"] > 0)
                )
                _pwm = _pwm[_pwm["_st"] | _pwm["_hurt"]]
                for (_tm, _yr, _wk), _g in _pwm.groupby(["Team", "Year", "Week"]):
                    _hp, _pp = {}, {}
                    for _pid, _pos, _st, _pt, _sa in zip(
                        _g["Player ID"].astype(str), _g["Position"].astype(str),
                        _g["_st"], _g["_pt"], _g["_sa"]):
                        if not _pid or _pid in ("None", "nan"):
                            continue
                        _hp[_pid] = float(_pt) if _st else float(_sa)
                        _pp[_pid] = _pos
                    if not _hp:
                        continue
                    try:
                        _ho = compute_optimal_lineup(_hp, _pp, int(_yr) if str(_yr).strip().isdigit() else _yr)
                    except Exception:
                        _ho = None
                    if _ho is not None:
                        _healthy_opt_by_tw[(str(_tm), str(_yr), str(_wk))] = float(_ho)
        except Exception as e:
            _log_exc(debug, "healthy_optimal_lineup", e)

        # --------------------------
        # Activated Cuff detection
        # --------------------------
        # A player has an "activated cuff" in week W if all of:
        #   - their own last-5-played PPG average is < 10 (i.e. low-scorer)
        #   - another NFL teammate (same NFL team AND same position) is
        #     injured or suspended in W
        #   - that teammate's last-5-played avg exceeds this player's by >10 PPG
        # _expected_points_if_healthy is the rolling 5-played-game mean built by
        # the hardship engine just above, so reuse it as last-5-avg.
        try:
            cuff_col = (
                "- Activated Cuff? (Was a player of the same nfl team/position "
                "& who averages >10 PPG more over last 5 played games injured? "
                "Only for players with avg <10 PPG)"
            )
            pw_c = pw.copy()
            pw_c["_avg"] = pd.to_numeric(pw_c.get("_expected_points_if_healthy"), errors="coerce")
            pw_c["_inj"] = pw_c.get("Injury?", False).fillna(False).astype(bool)
            pw_c["_sus"] = pw_c.get("Suspension?", False).fillna(False).astype(bool)
            pw_c["_inj_or_sus"] = pw_c["_inj"] | pw_c["_sus"]
            pw_c["_nfl_team"] = pw_c.get("NFL team").astype(str)
            pw_c["_pos"] = pw_c.get("Position").astype(str)

            # Build the injured-teammate index: highest last-5 avg of any
            # injured/suspended player per (Year, Week, NFL team, Position).
            # Restricting to max() per group keeps the per-row comparison O(1).
            inj_rows = pw_c[
                pw_c["_inj_or_sus"]
                & pw_c["_avg"].notna()
                & (pw_c["_nfl_team"] != "")
                & (pw_c["_nfl_team"].str.lower() != "nan")
                & (pw_c["_pos"] != "")
            ]
            inj_max_by_group: Dict[Tuple[int, int, str, str], float] = {}
            if not inj_rows.empty:
                inj_grp = inj_rows.groupby(["Year", "Week", "_nfl_team", "_pos"])["_avg"].max()
                for (yr, wk, nt, ps), v in inj_grp.items():
                    try:
                        inj_max_by_group[(int(yr), int(wk), str(nt), str(ps))] = float(v)
                    except Exception:
                        continue

            # Two signals (item 10):
            #  _cuff_rostered = the player is a handcuff this week — low scorer
            #     (<10 avg) with a same-NFL-team/position teammate who is
            #     injured/suspended and averages >10 PPG more. The injured
            #     teammate does NOT need to have been a starter.
            #  activated ("Activated Cuff?") = a rostered cuff who BECOMES A
            #     STARTER this week (Starter/Bench == "Starter").
            pw_c["_is_starter"] = pw_c.get("Starter/Bench", "").astype(str).str.lower() == "starter"
            cuff_rostered: List[int] = [0] * len(pw_c)
            activated: List[int] = [0] * len(pw_c)
            if inj_max_by_group:
                _starter_vals = pw_c["_is_starter"].tolist()
                for pos_i, (i, row) in enumerate(pw_c.iterrows()):
                    my_avg = row["_avg"]
                    if pd.isna(my_avg) or my_avg >= 10.0:
                        continue
                    nt = row["_nfl_team"]
                    ps = row["_pos"]
                    if not nt or nt.lower() == "nan" or not ps:
                        continue
                    try:
                        yr = int(row["Year"])
                        wk = int(row["Week"])
                    except Exception:
                        continue
                    best = inj_max_by_group.get((yr, wk, nt, ps))
                    if best is not None and best > float(my_avg) + 10.0:
                        cuff_rostered[pos_i] = 1
                        if bool(_starter_vals[pos_i]):
                            activated[pos_i] = 1
            pw["_cuff_rostered_flag"] = cuff_rostered
            pw[cuff_col] = activated  # "Activated Cuff?" now requires starting
        except Exception as e:
            _log_exc(debug, "cuff_detection", e)

    # Unique cuff player counts (item 9): "Number of cuffs rostered/started"
    # at the year / all-time levels count DISTINCT players, not player-weeks.
    # Built here (after cuff detection) and applied as overrides when each
    # sheet is assembled below.
    def _build_unique_cuff_counts(pw_df: pd.DataFrame, group_cols: List[str]) -> Dict[Tuple, Dict[str, int]]:
        out: Dict[Tuple, Dict[str, int]] = {}
        _cuff_col = (
            "- Activated Cuff? (Was a player of the same nfl team/position "
            "& who averages >10 PPG more over last 5 played games injured? "
            "Only for players with avg <10 PPG)"
        )
        if pw_df.empty or "Player ID" not in pw_df.columns:
            return out
        df = pw_df[pw_df["Player ID"].notna()].copy()
        df["_ros"] = pd.to_numeric(df.get("_cuff_rostered_flag"), errors="coerce").fillna(0.0) > 0
        df["_act"] = pd.to_numeric(df.get(_cuff_col), errors="coerce").fillna(0.0) > 0
        for col in group_cols:
            if col == "Year":
                df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64").map(
                    lambda x: int(x) if pd.notna(x) else None
                )
            elif col == "Team":
                df[col] = df[col].astype(str)
            else:
                df[col] = df[col].astype(object)
        df = df.dropna(subset=group_cols)
        ros = df[df["_ros"]].groupby(group_cols)["Player ID"].nunique()
        for key, val in ros.items():
            k = key if isinstance(key, tuple) else (key,)
            out.setdefault(k, {})["Number of cuffs rostered"] = int(val)
        act = df[df["_act"]].groupby(group_cols)["Player ID"].nunique()
        for key, val in act.items():
            k = key if isinstance(key, tuple) else (key,)
            out.setdefault(k, {})["Number of cuffs started"] = int(val)
        return out

    unique_cuffs_by_team_year: Dict[Tuple, Dict[str, int]] = {}
    unique_cuffs_by_team_all: Dict[Tuple, Dict[str, int]] = {}
    unique_cuffs_by_year: Dict[Tuple, Dict[str, int]] = {}
    unique_cuffs_league_all: Dict[str, int] = {}
    if not pw.empty:
        try:
            unique_cuffs_by_team_year = _build_unique_cuff_counts(pw, ["Team", "Year"])
            unique_cuffs_by_team_all = _build_unique_cuff_counts(pw, ["Team"])
            unique_cuffs_by_year = _build_unique_cuff_counts(pw, ["Year"])
            _allc = _build_unique_cuff_counts(pw.assign(_all="all"), ["_all"])
            unique_cuffs_league_all = _allc.get(("all",), {})
        except Exception as e:
            _log_exc(debug, "unique_cuff_counts", e)

    # --------------------------
    # Recompute team-week injury/susp/bye counts and hardship from player-week
    # --------------------------
    if not tw.empty and not pw.empty:
        pw2 = pw.copy()
        pw2["Injury?"] = pw2["Injury?"].fillna(False).astype(bool)
        pw2["Suspension?"] = pw2["Suspension?"].fillna(False).astype(bool)
        pw2["Bye?"] = pw2["Bye?"].fillna(False).astype(bool)
        pw2["Points"] = pd.to_numeric(pw2["Points"], errors="coerce").fillna(0.0)

        pw2["_missed_injury"] = (pw2["Injury?"] & (~pw2["Bye?"]) & (pw2["Points"] == 0)).astype(int)
        pw2["_missed_susp"] = (pw2["Suspension?"] & (~pw2["Bye?"]) & (pw2["Points"] == 0)).astype(int)
        pw2["_on_bye"] = (pw2["Bye?"] & (pw2["Points"] == 0)).astype(int)

        # Starter injury/suspension weeks (item 5): a missed week counts only
        # if the player was a starter under the SA-Hardship heuristic
        # (_was_recent_starter_injsusp == 1).
        pw2["_was_recent_starter"] = pd.to_numeric(
            pw2.get("_was_recent_starter_injsusp"), errors="coerce"
        ).fillna(0).astype(int)
        pw2["_missed_injury_starter"] = (pw2["_missed_injury"] & pw2["_was_recent_starter"]).astype(int)
        pw2["_missed_susp_starter"] = (pw2["_missed_susp"] & pw2["_was_recent_starter"]).astype(int)

        pw2["Starter?"] = (pw2["Starter/Bench"] == "Starter").astype(int)
        pw2["Number_of_players_injured_or_suspended"] = pw2["_missed_injury"] + pw2["_missed_susp"]

        agg = pw2.groupby(["Team", "Year", "Week"], as_index=False).agg(
            Hardship_Points_Lost=("_points_lost_inj_susp", "sum"),
            Starter_Adj_Hardship=("_starter_adj_points_lost", "sum"),
            Number_of_Injuries=("_missed_injury", "sum"),
            Number_of_suspensions=("_missed_susp", "sum"),
            Number_of_starter_injuries=("_missed_injury_starter", "sum"),
            Number_of_starter_suspensions=("_missed_susp_starter", "sum"),
            Number_of_players_on_bye=("_on_bye", "sum"),
            Number_of_players_injured_or_suspended=("Number_of_players_injured_or_suspended", "sum"),
            Starter_Count=("Starter?", "sum"),
        )

        tw = tw.merge(agg, how="left", on=["Team", "Year", "Week"])
        # Harden numeric outputs + create friendly display columns (never crash on missing cols)
        for _c in [
            "Hardship_Points_Lost",
            "Starter_Adj_Hardship",
            "Number_of_Injuries",
            "Number_of_suspensions",
            "Number_of_players_injured_or_suspended",
            "Number_of_players_on_bye",
            "Starter_Count",
        ]:
            safe_to_numeric(tw, _c, default=0.0)

        tw["Hardship"] = pd.to_numeric(tw.get("Hardship_Points_Lost"), errors="coerce").fillna(0.0)
        tw["Starter-adjusted Hardship"] = pd.to_numeric(tw.get("Starter_Adj_Hardship"), errors="coerce").fillna(0.0).round(4)
        # Loss from hardship? — a LOSS the team would have WON if its hurt
        # would-be-starters had been available: the healthy-lineup score
        # (_healthy_opt_by_tw — actual starters + hurt players at SA-hardship,
        # best valid lineup, bounded to the slots, NO healthy-bench re-optimizing)
        # beats the opponent's actual PF.
        try:
            _won = tw.get("Win?").astype(str).str.lower().isin(["true", "1", "yes"]) if "Win?" in tw.columns else pd.Series(False, index=tw.index)
            _pa = pd.to_numeric(tw.get("Points against"), errors="coerce")
            _ho = pd.Series(
                [_healthy_opt_by_tw.get((str(t), str(y), str(w)))
                 for t, y, w in zip(tw.get("Team"), tw.get("Year"), tw.get("Week"))],
                index=tw.index, dtype="float64",
            )
            tw["Loss from hardship?"] = (~_won) & _pa.notna() & _ho.notna() & (_ho > _pa)
        except Exception as e:
            _log_exc(debug, "loss_from_hardship_flag", e)
        # (Previously had a defensive SA ≤ H clamp here for a
        # plehv79 2022 wk3 anomaly. Audit traced it to negative
        # `expected` from a single-game-history rookie baseline;
        # fixed at the player level via expected = max(0, expected)
        # above. Clamp removed.)
        tw["Number of Injuries"] = tw["Number_of_Injuries"].round(0).astype(int)
        tw["Number of suspensions"] = tw["Number_of_suspensions"].round(0).astype(int)
        safe_to_numeric(tw, "Number_of_starter_injuries", default=0.0)
        safe_to_numeric(tw, "Number_of_starter_suspensions", default=0.0)
        tw["Number of starter injuries"] = tw["Number_of_starter_injuries"].round(0).astype(int)
        tw["Number of starter suspensions"] = tw["Number_of_starter_suspensions"].round(0).astype(int)
        tw["Number of players on bye"] = tw["Number_of_players_on_bye"].round(0).astype(int)

        tw.drop(columns=[
            "Hardship_Points_Lost",
            "Starter_Adj_Hardship",
            "Number_of_Injuries",
            "Number_of_suspensions",
            "Number_of_players_injured_or_suspended",
            "Number_of_players_on_bye",
            "Starter_Count",
        ], inplace=True, errors="ignore")

        # UPST: win with lower Max PF than opponent
        if "UPST" not in tw.columns:
            tw["UPST"] = None
        for (yr, wk), g in tw.groupby(["Year", "Week"]):
            g2 = g.copy()
            g2["PF"] = pd.to_numeric(g2["PF"], errors="coerce").fillna(0.0)
            g2["Points against"] = pd.to_numeric(g2["Points against"], errors="coerce").fillna(0.0)
            g2["Max PF"] = pd.to_numeric(g2["Max PF"], errors="coerce")
            for idx, row in g2.iterrows():
                opp = g2[(g2["PF"] == row["Points against"]) & (g2["Points against"] == row["PF"])]
                if len(opp) == 1:
                    opp_max = opp.iloc[0]["Max PF"]
                    if row["Win?"] == 1 and pd.notna(row["Max PF"]) and pd.notna(opp_max):
                        tw.loc[idx, "UPST"] = int(float(row["Max PF"]) < float(opp_max))
                    else:
                        tw.loc[idx, "UPST"] = 0
                else:
                    tw.loc[idx, "UPST"] = 0

        # Brosenzweig / Sisenzweig (correct definition)
        # Brosenzweig: LOSS while 2nd-highest scoring team of the week.
        # Sisenzweig: WIN while 2nd-lowest scoring team of the week.
        tw["Brosenzweig"] = 0
        tw["Sisenzweig"] = 0
        if "PF" in tw.columns and "Win?" in tw.columns:
            tw_pf = tw.copy()
            tw_pf["PF"] = pd.to_numeric(tw_pf["PF"], errors="coerce").fillna(0.0)
            tw_pf["Win?"] = pd.to_numeric(tw_pf["Win?"], errors="coerce")
            for (yr, wk), g in tw_pf.groupby(["Year", "Week"]):
                if g.empty:
                    continue
                # ranks: 1 = highest (desc), 1 = lowest (asc)
                r_desc = g["PF"].rank(method="min", ascending=False)
                r_asc = g["PF"].rank(method="min", ascending=True)
                mask_b = (g["Win?"] == 0) & (r_desc == 2)
                mask_s = (g["Win?"] == 1) & (r_asc == 2)
                tw.loc[g.index[mask_b], "Brosenzweig"] = 1
                tw.loc[g.index[mask_s], "Sisenzweig"] = 1

        # ── Luck (the "G2" model; full derivation + scorecard in
        #    plan/LUCK_REWORK.md). All inputs are existing team_week columns.
        #
        #  WeeklyLuck = (W_OUT·OUT + W_SIS·SIS − W_BROS·BROS)·POSTBOOST   # result surprise
        #             + (W_OPP·OPP + W_OWN·OWN)·GATE                       # scoring variance, closeness-gated
        #             − W_ADV·ADV + W_EFF·EFF + W_CLOSE·CLOSE
        #  where
        #    OUT   = Win − pregame_p   (pregame_p = logistic(1.5·blend of full-season
        #            MaxPF/PF/win% diffs) — calibrated so winning-to-talent nets ~0,
        #            which folds Win-Variance behaviour into Σweekly)
        #    OPP   = z_wk(−(opp PF − opp season-avg PF))   own/opp scoring vs norm
        #    OWN   = z_wk(PF − own season-avg PF)
        #    ADV   = z_wk(Hardship + Starter-adj Hardship + 3·byes)  (always subtracted, heavy)
        #    EFF   = z_wk(Efficiency)
        #    CLOSE = sign(Margin)·max(0, 1−|Margin|/8)   nail-biter term
        #    GATE  = 1/(1+|Margin|/15)   variance only "counts" if it swung a close game
        #    POST  = championship-bracket week (Final/Semifinal/3rd Place)
        #  Season / all-time Luck = Σ WeeklyLuck (no win% multiplier).
        try:
            L = tw.copy().reset_index(drop=True)
            L["_PF"] = pd.to_numeric(L.get("PF"), errors="coerce")
            L["_MAXPF"] = pd.to_numeric(L.get("Max PF"), errors="coerce")
            L["_OPPPF"] = pd.to_numeric(L.get("Points against"), errors="coerce")
            L["_MARGIN"] = pd.to_numeric(L.get("Margin"), errors="coerce").fillna(0.0)
            L["_EFF"] = pd.to_numeric(L.get("Efficiency"), errors="coerce")
            L["_HARD"] = pd.to_numeric(L.get("Hardship"), errors="coerce").fillna(0.0)
            L["_SA"] = pd.to_numeric(L.get("Starter-adjusted Hardship"), errors="coerce").fillna(0.0)
            L["_BYE"] = pd.to_numeric(L.get("Number of players on bye"), errors="coerce").fillna(0.0)
            L["_BROS"] = pd.to_numeric(L.get("Brosenzweig"), errors="coerce").fillna(0.0)
            L["_SIS"] = pd.to_numeric(L.get("Sisenzweig"), errors="coerce").fillna(0.0)
            # Win? is 1/0/0.5 here; coerce, with a boolean/string fallback.
            L["_WIN"] = pd.to_numeric(L.get("Win?"), errors="coerce")
            if L["_WIN"].isna().all():
                L["_WIN"] = L.get("Win?").astype(str).str.lower().map(
                    {"true": 1.0, "false": 0.0, "1": 1.0, "0": 0.0, "0.5": 0.5}
                )
            _wn = L["Week Name"].astype(str) if "Week Name" in L.columns else pd.Series("", index=L.index)
            L["_POST"] = _wn.isin(["Final", "Semifinal", "3rd Place"]).astype(float)

            def _z_all(s):
                s = pd.to_numeric(s, errors="coerce")
                sd = s.std()
                return (s - s.mean()) / (sd if sd and sd > 0 else 1.0)

            def _z_wk(s):
                s = pd.to_numeric(pd.Series(s, index=L.index), errors="coerce")
                m = s.groupby([L["Year"], L["Week"]]).transform("mean")
                sd = s.groupby([L["Year"], L["Week"]]).transform("std").replace(0, np.nan)
                return (((s - m) / sd).clip(-2.5, 2.5) / 2.5).fillna(0.0)

            _g = L.groupby(["Team", "Year"])
            L["FS_pf"] = _g["_PF"].transform("mean")
            L["FS_maxpf"] = _g["_MAXPF"].transform("mean")
            L["FS_win"] = _g["_WIN"].transform("mean")

            # opponent full-season values via (Year, Week, Team) lookup
            _idx = {(L.at[i, "Year"], L.at[i, "Week"], str(L.at[i, "Team"])): i for i in L.index}
            def _opp(col):
                vals = []
                for i in L.index:
                    j = _idx.get((L.at[i, "Year"], L.at[i, "Week"], str(L.at[i, "Opponent"])))
                    vals.append(L.at[j, col] if j is not None else np.nan)
                return pd.Series(vals, index=L.index)
            L["opp_FS_pf"] = _opp("FS_pf")
            L["opp_FS_maxpf"] = _opp("FS_maxpf")
            L["opp_FS_win"] = _opp("FS_win")

            dblend = (
                _z_all(L["FS_maxpf"] - L["opp_FS_maxpf"]).fillna(0.0)
                + _z_all(L["FS_pf"] - L["opp_FS_pf"]).fillna(0.0)
                + _z_all(L["FS_win"] - L["opp_FS_win"]).fillna(0.0)
            ) / 3.0
            pregame_p = 1.0 / (1.0 + np.exp(-1.5 * dblend))

            OUT = L["_WIN"] - pregame_p
            OPP = _z_wk(-(L["_OPPPF"] - L["opp_FS_pf"]))
            OWN = _z_wk(L["_PF"] - L["FS_pf"])
            ADV = _z_wk(L["_HARD"] + L["_SA"] + 3.0 * L["_BYE"])
            EFF = _z_wk(L["_EFF"])
            _mar = L["_MARGIN"]
            CLOSE = np.sign(_mar) * np.maximum(0.0, 1.0 - _mar.abs() / 8.0)
            GATE = 1.0 / (1.0 + _mar.abs() / 15.0)
            postboost = np.where(L["_POST"] > 0, 1.8, 1.0)
            # Flat penalty for a "Loss from hardship?" week — a winnable game lost
            # to injured starters is bad luck. 0.25 ≈ a typical weekly luck swing
            # (per-week |luck| ≈ 0.20), so one such loss roughly doubles that
            # week's misfortune without dominating the season total.
            LFH = (L.get("Loss from hardship?").astype(str).str.lower().isin(["true", "1"]).astype(float)
                   if "Loss from hardship?" in L.columns else 0.0)

            luck = (
                (0.27 * OUT + 0.14 * L["_SIS"] - 0.14 * L["_BROS"]) * postboost
                + (0.36 * OPP + 0.10 * OWN) * GATE
                - 0.36 * ADV
                + 0.12 * EFF
                + 0.16 * CLOSE
                - 0.25 * LFH
            )
            # No opponent / no game → 0 luck.
            _has_game = L["Opponent"].astype(str).str.strip().ne("") & L["opp_FS_pf"].notna()
            luck = luck.where(_has_game, 0.0)
            # Round weekly Luck so the season/all-time SUMS are deterministic
            # (floating-point summation order otherwise leaves ~1e-16 noise that
            # pollutes every build diff). 6 dp keeps full display precision.
            tw["Luck"] = pd.to_numeric(pd.Series(luck.values), errors="coerce").fillna(0.0).round(6).values
        except Exception as e:
            _log_exc(debug, "team_week_luck_formula", e)

    # --------------------------
    
        # Fill remaining schema columns in team-week (best-effort)
        try:
            # Cuffs: use player-week activated cuff flag (rostered and started)
            cuff_col = (
                "- Activated Cuff? (Was a player of the same nfl team/position "
                "& who averages >10 PPG more over last 5 played games injured? "
                "Only for players with avg <10 PPG)"
            )
            if (not pw.empty) and (cuff_col in pw.columns):
                # "Number of cuffs rostered" counts handcuffs on the roster
                # (_cuff_rostered_flag); "Number of cuffs started" counts the
                # activated cuffs (cuff_col, which already requires starting).
                _ros_col = "_cuff_rostered_flag" if "_cuff_rostered_flag" in pw.columns else cuff_col
                pw_c = pw[["Team","Year","Week",cuff_col,_ros_col]].copy()
                pw_c[cuff_col] = pd.to_numeric(pw_c[cuff_col], errors="coerce").fillna(0.0)
                pw_c[_ros_col] = pd.to_numeric(pw_c[_ros_col], errors="coerce").fillna(0.0)
                agg_c = pw_c.groupby(["Team","Year","Week"], as_index=False).agg(
                    **{
                        "Number of cuffs rostered": (_ros_col, "sum"),
                        "Number of cuffs started": (cuff_col, "sum"),
                    }
                )
                # Drop any pre-existing cuff cols before merge so the aggregated
                # values land in the natural column names (not '_c' suffixes).
                # The original merge used suffixes=("","_c") which kept the all-zero
                # placeholder column and put the real data under "Number of cuffs
                # rostered_c" — meaning team_week / team_year / league rollups all
                # read zeros despite the per-player flag being correct.
                tw = tw.drop(columns=["Number of cuffs rostered", "Number of cuffs started"], errors="ignore")
                tw = tw.merge(agg_c, how="left", on=["Team","Year","Week"])
                tw["Number of cuffs rostered"] = pd.to_numeric(tw.get("Number of cuffs rostered"), errors="coerce").fillna(0.0).round(0).astype(int)
                tw["Number of cuffs started"] = pd.to_numeric(tw.get("Number of cuffs started"), errors="coerce").fillna(0.0).round(0).astype(int)

            # Future draft capital (weighted rounds) + startup placeholder left blank for now
            if "Future draft capital" in tw.columns:
                # Per-week holdings: capital reflects picks owned as of that
                # week, so a mid-season pick trade updates the column from
                # that week forward ("updates on trade"). 0.0 only when the
                # team holds no picks in the next 3 seasons.
                fdc_vals = []
                for _, r in tw.iterrows():
                    yr = _to_int(r.get("Year"), None)
                    wk = _to_int(r.get("Week"), None)
                    if yr is None:
                        fdc_vals.append(0.0)
                        continue
                    wk_date = date(int(yr), 9, 1) + timedelta(days=7 * ((wk or 1) - 1))
                    fdc_vals.append(float(_future_cap_held(str(r.get("Team")), int(yr), wk_date)))
                tw["Future draft capital"] = pd.to_numeric(pd.Series(fdc_vals, index=tw.index), errors="coerce").fillna(0.0)

            if "Startup draft players remaining" in tw.columns:
                tw["Startup draft players remaining"] = None
        except Exception as e:
            _log_exc(debug, "team_week_fill_schema_cols", e)
# Derived team-week columns: pregame avg maxPF diff
    # --------------------------
    if not tw.empty:
        tw = tw.sort_values(["Team", "Year", "Week"]).reset_index(drop=True)
        tw["Max PF"] = pd.to_numeric(tw["Max PF"], errors="coerce")

        # Own pregame average Max PF (season-to-date, excluding current week)
        tw["Pregame avg MaxPF"] = tw.groupby(["Team", "Year"])["Max PF"].apply(
            lambda s: s.shift(1).expanding().mean()
        ).reset_index(level=[0, 1], drop=True)

        # Opponent-aware difference where matchup mapping is available.
        tw["Difference in pregame avg max PF from opponent"] = None

        tw_key = tw.copy()
        tw_key["_TeamNorm"] = tw_key["Team"].astype(str).map(_norm_team_name)
        tw_key["_OppNorm"] = tw_key.get("Opponent", pd.Series(dtype=str)).astype(str).map(_norm_team_name)

        by_key = {
            (int(r["Year"]), int(r["Week"]), str(r["_TeamNorm"])): int(i)
            for i, r in tw_key[["Year", "Week", "_TeamNorm"]].iterrows()
            if pd.notna(r["Year"]) and pd.notna(r["Week"])
        }

        # Primary: direct opponent lookup. Fallback: league-average pregame baseline.
        for i, r in tw_key.iterrows():
            yr = _to_int(r.get("Year"), None)
            wk = _to_int(r.get("Week"), None)
            opp_n = str(r.get("_OppNorm") or "")
            own_pre = _to_float(r.get("Pregame avg MaxPF"), None)
            if yr is None or wk is None or own_pre is None:
                continue

            opp_pre = None
            if opp_n:
                j = by_key.get((int(yr), int(wk), opp_n))
                if j is not None:
                    opp_pre = _to_float(tw_key.loc[j, "Pregame avg MaxPF"], None)

            if opp_pre is None:
                same_week = tw_key[(tw_key["Year"] == yr) & (tw_key["Week"] == wk)]
                vals = pd.to_numeric(same_week["Pregame avg MaxPF"], errors="coerce").dropna()
                if len(vals):
                    opp_pre = float(vals.mean())

            if opp_pre is not None:
                tw.loc[i, "Difference in pregame avg max PF from opponent"] = round(float(own_pre) - float(opp_pre), 2)

        # Align UPST with intended meaning: win despite lower pregame maxPF profile.
        tw["UPST"] = (
            (pd.to_numeric(tw.get("Win?"), errors="coerce") == 1)
            & (pd.to_numeric(tw.get("Difference in pregame avg max PF from opponent"), errors="coerce") < 0)
        ).astype(int)

        tw.drop(columns=["Pregame avg MaxPF"], inplace=True, errors="ignore")

    # --------------------------
    
    # --------------------------
    # Player-week: rolling 5-game diffs vs reference player + cuff adjusted diff
    # --------------------------
    if not pw.empty:
        try:
            pw["Year"] = pd.to_numeric(pw["Year"], errors="coerce").astype("Int64")
            pw["Week"] = pd.to_numeric(pw["Week"], errors="coerce").astype("Int64")
            pw["Points"] = pd.to_numeric(pw["Points"], errors="coerce").fillna(0.0)

            # "played games" exclude injury/susp/bye
            played_mask = ~pw[["Injury?", "Suspension?", "Bye?"]].fillna(False).any(axis=1)

            pw_sorted = pw.sort_values(["Player", "Year", "Week"]).reset_index()
            # map (player,year,week)-> rolling avg last5 played (including current if played)
            rolling_avg = {}
            # NOTE: do NOT import deque inside this function.
            # An inner import would make `deque` a local variable for the entire
            # enclosing scope, which breaks earlier lambdas that reference the
            # global `deque` (CI failure: cannot access free variable 'deque').
            hist = defaultdict(lambda: deque(maxlen=5))
            for _, r in pw_sorted.iterrows():
                p=str(r["Player"]); yr=int(r["Year"]) if pd.notna(r["Year"]) else None; wk=int(r["Week"]) if pd.notna(r["Week"]) else None
                if yr is None or wk is None:
                    continue
                key=(p,yr,wk)
                # compute avg of previous played games (last5) BEFORE adding current
                prev=list(hist[(p,yr)])
                avg_prev=float(np.mean(prev)) if prev else None
                # if played, include current for future
                if bool(played_mask.loc[r["index"]]):
                    hist[(p,yr)].append(float(r["Points"]))
                rolling_avg[key]=avg_prev

            def get_avg(p,yr,wk):
                return rolling_avg.get((str(p),int(yr),int(wk)))

            diffs=[]
            cuff_adj=[]
            for _, r in pw.iterrows():
                ref=r.get("Reference player name")
                if not isinstance(ref,str) or ref.strip()=="":
                    diffs.append(None); cuff_adj.append(None); continue
                yr=r.get("Year"); wk=r.get("Week"); player=r.get("Player")
                if pd.isna(yr) or pd.isna(wk):
                    diffs.append(None); cuff_adj.append(None); continue
                avg_p=get_avg(player,yr,wk)
                avg_r=get_avg(ref,yr,wk)
                if (avg_p is None) or (avg_r is None):
                    diffs.append(None); cuff_adj.append(None); continue
                started = (r.get("Starter/Bench") == "Starter")
                diff = (avg_r-avg_p) if started else (avg_p-avg_r)
                diffs.append(round(float(diff),2))
                cuff = float(r.get(
                    "- Activated Cuff? (Was a player of the same nfl team/position "
                    "& who averages >10 PPG more over last 5 played games injured? "
                    "Only for players with avg <10 PPG)"
                ) or 0)
                cuff_adj.append(round(float(diff) * (0.5 if cuff else 1.0), 2))
            pw["Difference in averages of best/worst startables over previous 5 games"] = diffs
            pw["Cuff adjusted difference"] = cuff_adj
        except Exception as e:
            _log_exc(debug, "player_week_rolling_diffs", e)

    # --------------------------
    # Team-week: Tanking (math formula computed earlier in build_all by
    # _tanking_score and merged into tw via tank_df). Keep that score as-is
    # here — no more binary override. Round to 4 decimals for readability.
    # --------------------------
    if not tw.empty:
        try:
            tw["Tanking"] = pd.to_numeric(tw.get("Tanking"), errors="coerce").fillna(0.0).round(4)
        except Exception as e:
            _log_exc(debug, "team_week_tanking_round", e)
# Team-week flags & streaks
    # --------------------------
    if not tw.empty:
        tw["Increase in points from previous week"] = None
        tw["Highest score?"] = 0
        tw["Lowest score?"] = 0
        tw["Narrowest victory?"] = 0
        tw["Largest blowout?"] = 0
        tw["Most efficient?"] = 0
        tw["Least efficient?"] = 0
        tw["Top half of league?"] = 0

        for (yr, wk), g in tw.groupby(["Year", "Week"]):
            g2 = g.copy()
            g2["PF"] = pd.to_numeric(g2["PF"], errors="coerce").fillna(0.0)
            g2["Margin"] = pd.to_numeric(g2["Margin"], errors="coerce")
            g2["Efficiency"] = pd.to_numeric(g2["Efficiency"], errors="coerce")
            if not g2.empty:
                max_pf = g2["PF"].max()
                min_pf = g2["PF"].min()
                tw.loc[(tw["Year"] == yr) & (tw["Week"] == wk) & (tw["PF"] == max_pf), "Highest score?"] = 1
                tw.loc[(tw["Year"] == yr) & (tw["Week"] == wk) & (tw["PF"] == min_pf), "Lowest score?"] = 1

                # narrowest victory (smallest positive margin)
                wins = g2[g2["Margin"] > 0]
                if not wins.empty:
                    min_margin = wins["Margin"].min()
                    tw.loc[(tw["Year"] == yr) & (tw["Week"] == wk) & (tw["Margin"] == min_margin), "Narrowest victory?"] = 1
                    max_margin = wins["Margin"].max()
                    tw.loc[(tw["Year"] == yr) & (tw["Week"] == wk) & (tw["Margin"] == max_margin), "Largest blowout?"] = 1

                # efficiency
                if g2["Efficiency"].notna().any():
                    max_eff = g2["Efficiency"].max()
                    min_eff = g2["Efficiency"].min()
                    tw.loc[(tw["Year"] == yr) & (tw["Week"] == wk) & (tw["Efficiency"] == max_eff), "Most efficient?"] = 1
                    tw.loc[(tw["Year"] == yr) & (tw["Week"] == wk) & (tw["Efficiency"] == min_eff), "Least efficient?"] = 1

                # top half of league by PF
                median_pf = g2["PF"].median()
                tw.loc[(tw["Year"] == yr) & (tw["Week"] == wk) & (tw["PF"] >= median_pf), "Top half of league?"] = 1

        # Three new weekly team awards derived from player_week:
        #   One-man army?  = the team whose single top starter supplied the
        #                    greatest share of its PF that week (most top-heavy).
        #   Most bench points? = the team with the highest total bench points.
        #   Most injured?  = the team with the most injured players on its
        #                    roster (starters + bench) that week.
        tw["One-man army?"] = 0
        tw["Most bench points?"] = 0
        tw["Most injured?"] = 0
        try:
            _p = pw.copy()
            _p["Year"] = pd.to_numeric(_p["Year"], errors="coerce")
            _p["Week"] = pd.to_numeric(_p["Week"], errors="coerce")
            _p["Points"] = pd.to_numeric(_p["Points"], errors="coerce").fillna(0.0)
            _p["_inj"] = _p["Injury?"].fillna(False).astype(bool).astype(int)
            _is_st = _p["Starter/Bench"] == "Starter"
            _st = _p[_is_st].copy()
            _st["_share"] = pd.to_numeric(_st["% of points (if starter)"], errors="coerce")
            _tops = _st.groupby(["Year", "Week", "Team"])["_share"].max().reset_index()
            _tops = _tops.dropna(subset=["_share"])
            _tops["_rk"] = _tops.groupby(["Year", "Week"])["_share"].transform("max")
            _oma = _tops[_tops["_share"] == _tops["_rk"]]
            _bn = _p[~_is_st].groupby(["Year", "Week", "Team"])["Points"].sum().reset_index()
            _bn["_rk"] = _bn.groupby(["Year", "Week"])["Points"].transform("max")
            _mbp = _bn[(_bn["Points"] == _bn["_rk"]) & (_bn["_rk"] > 0)]
            _ij = _p.groupby(["Year", "Week", "Team"])["_inj"].sum().reset_index()
            _ij["_rk"] = _ij.groupby(["Year", "Week"])["_inj"].transform("max")
            _mij = _ij[(_ij["_inj"] == _ij["_rk"]) & (_ij["_rk"] > 0)]
            _oma_keys = set(zip(_oma["Year"], _oma["Week"], _oma["Team"].astype(str)))
            _mbp_keys = set(zip(_mbp["Year"], _mbp["Week"], _mbp["Team"].astype(str)))
            _mij_keys = set(zip(_mij["Year"], _mij["Week"], _mij["Team"].astype(str)))
            _twy = pd.to_numeric(tw["Year"], errors="coerce")
            _tww = pd.to_numeric(tw["Week"], errors="coerce")
            _twt = tw["Team"].astype(str)
            tw["One-man army?"] = [1 if k in _oma_keys else 0 for k in zip(_twy, _tww, _twt)]
            tw["Most bench points?"] = [1 if k in _mbp_keys else 0 for k in zip(_twy, _tww, _twt)]
            tw["Most injured?"] = [1 if k in _mij_keys else 0 for k in zip(_twy, _tww, _twt)]
        except Exception as e:
            _log_exc(debug, "team_week_new_awards", e)

        # Per-week regular-season standings leader (cumulative through that
        # week): used for the "Standings leader streak" below.
        _leader_set = set()
        try:
            _t = tw.copy()
            _t["_Y"] = pd.to_numeric(_t["Year"], errors="coerce")
            _t["_W"] = pd.to_numeric(_t["Week"], errors="coerce")
            _t["_PF"] = pd.to_numeric(_t["PF"], errors="coerce").fillna(0.0)
            _t["_win"] = pd.to_numeric(_t["Win?"], errors="coerce")
            for _yr, _g in _t.groupby("_Y"):
                if pd.isna(_yr):
                    continue
                _ps = playoff_start_by_season.get(int(_yr))
                _gg = _g[_g["_W"] < _ps] if _ps else _g
                for _wk in sorted(w for w in _gg["_W"].dropna().unique()):
                    _cur = _gg[_gg["_W"] <= _wk]
                    _rank = []
                    for _tm, _d in _cur.groupby("Team"):
                        _w = float((_d["_win"] == 1).sum()) + 0.5 * float((_d["_win"] == 0.5).sum())
                        _rank.append((str(_tm), _w, float(_d["_PF"].sum())))
                    if _rank:
                        _rank.sort(key=lambda r: (r[1], r[2]), reverse=True)
                        _leader_set.add((int(_yr), int(_wk), _rank[0][0]))
        except Exception as e:
            _log_exc(debug, "standings_leader_calc", e)

        # streaks + increase from previous week
        tw = tw.sort_values(["Team", "Year", "Week"]).reset_index(drop=True)
        tw["Win streak"] = 0
        tw["Loss streak"] = 0
        tw["Win streak counting previous season"] = 0
        tw["Loss streak counting previous season"] = 0
        # Award streaks (one running column per weekly team award) + dedicated
        # streaks. Per the design these are ALL-TIME: they run continuously and
        # do NOT reset at the season boundary (only Win/Loss keep a separate
        # within-season variant). The running counters are computed here and
        # then terminal-encoded below so each run is listed once.
        _team_award_streaks = [
            ("Highest score?", "Highest score streak"),
            ("Lowest score?", "Lowest score streak"),
            ("Narrowest victory?", "Narrowest victory streak"),
            ("Largest blowout?", "Largest blowout streak"),
            ("Most efficient?", "Most efficient streak"),
            ("Least efficient?", "Least efficient streak"),
            ("Top half of league?", "Top half streak"),
            ("One-man army?", "One-man army streak"),
            ("Most bench points?", "Most bench points streak"),
            ("Most injured?", "Most injured streak"),
        ]
        _ts_dedicated = ["Bottom half streak", "150+ PF streak",
                         "Standings leader streak", "Quiet streak",
                         "Win streak vs this opponent"]
        for _fc, _sc in _team_award_streaks:
            tw[_sc] = 0
        for _sc in _ts_dedicated:
            tw[_sc] = 0
        for team, g in tw.groupby("Team"):
            win_streak = loss_streak = 0
            win_streak_season = loss_streak_season = 0
            current_year = None
            # prev_pf carries across seasons: Week 1's "Increase in points
            # from previous week" compares to the team's last played week
            # of the prior season (≈ championship week).
            prev_pf = None
            _aw = {_sc: 0 for _fc, _sc in _team_award_streaks}  # all-time award counters
            _bot = _pf150 = _lead = _quiet = 0  # all-time dedicated counters
            _rival: Dict[str, int] = {}  # opponent -> consecutive H2H wins
            for idx, row in g.sort_values(["Year", "Week"]).iterrows():
                if current_year != row["Year"]:
                    current_year = row["Year"]
                    win_streak_season = 0
                    loss_streak_season = 0
                result = row.get("Win?")
                if result == 1:
                    win_streak_season += 1
                    loss_streak_season = 0
                    win_streak += 1
                    loss_streak = 0
                elif result == 0:
                    loss_streak_season += 1
                    win_streak_season = 0
                    loss_streak += 1
                    win_streak = 0
                else:
                    win_streak_season = 0
                    loss_streak_season = 0
                    win_streak = 0
                    loss_streak = 0

                tw.loc[idx, "Win streak"] = win_streak_season
                tw.loc[idx, "Loss streak"] = loss_streak_season
                tw.loc[idx, "Win streak counting previous season"] = win_streak
                tw.loc[idx, "Loss streak counting previous season"] = loss_streak

                # Award streaks (all-time, no season reset)
                for _fc, _sc in _team_award_streaks:
                    _v = row.get(_fc)
                    _aw[_sc] = (_aw[_sc] + 1) if (pd.notna(_v) and float(_v) == 1) else 0
                    tw.loc[idx, _sc] = _aw[_sc]

                # Dedicated streaks (all-time, no season reset)
                _played = result in (0, 1, 0.5)
                _top = pd.notna(row.get("Top half of league?")) and float(row.get("Top half of league?")) == 1
                _bot = (_bot + 1) if (_played and not _top) else 0
                tw.loc[idx, "Bottom half streak"] = _bot

                _pf = pd.to_numeric(pd.Series([row.get("PF")]), errors="coerce").iloc[0]
                _pf150 = (_pf150 + 1) if (pd.notna(_pf) and float(_pf) >= 150) else 0
                tw.loc[idx, "150+ PF streak"] = _pf150

                # Standings leader: defined only on regular-season weeks; on
                # playoff weeks hold the value steady rather than break it.
                _yk = pd.to_numeric(pd.Series([row.get("Year")]), errors="coerce").iloc[0]
                _wk = pd.to_numeric(pd.Series([row.get("Week")]), errors="coerce").iloc[0]
                _ps = playoff_start_by_season.get(int(_yk)) if pd.notna(_yk) else None
                _is_reg = (_ps is None) or (pd.notna(_wk) and _wk < _ps)
                if _is_reg:
                    _is_lead = pd.notna(_yk) and pd.notna(_wk) and (int(_yk), int(_wk), str(team)) in _leader_set
                    _lead = (_lead + 1) if _is_lead else 0
                tw.loc[idx, "Standings leader streak"] = _lead

                _ntx = pd.to_numeric(pd.Series([row.get("Number of transactions")]), errors="coerce").iloc[0]
                _ntr = pd.to_numeric(pd.Series([row.get("Number of trades")]), errors="coerce").iloc[0]
                _moves = (0 if pd.isna(_ntx) else float(_ntx)) + (0 if pd.isna(_ntr) else float(_ntr))
                _quiet = (_quiet + 1) if (_played and _moves == 0) else 0
                tw.loc[idx, "Quiet streak"] = _quiet

                # Rivalry: consecutive H2H wins vs this week's opponent.
                _opp = row.get("Opponent")
                _opp = str(_opp).strip() if (_opp is not None and pd.notna(_opp)) else ""
                if _opp:
                    if result == 1:
                        _rival[_opp] = _rival.get(_opp, 0) + 1
                    elif result in (0, 0.5):
                        _rival[_opp] = 0
                    tw.loc[idx, "Win streak vs this opponent"] = _rival.get(_opp, 0)

                if prev_pf is not None and pd.notna(row["PF"]):
                    tw.loc[idx, "Increase in points from previous week"] = round(float(row["PF"]) - float(prev_pf), 2)
                if pd.notna(row["PF"]):
                    prev_pf = row["PF"]

        # Terminal-encode the all-time streaks: list each run once (length on
        # its final week, 'In Progress' before that, 0 when not streaking).
        # Win/Loss streaks are intentionally left as running counts.
        try:
            _non_rival = [_sc for _fc, _sc in _team_award_streaks] + \
                         ["Bottom half streak", "150+ PF streak", "Standings leader streak", "Quiet streak"]
            tw = _terminalize_streaks(tw, ["Team"], ["Year", "Week"], _non_rival)
            tw = _terminalize_streaks(tw, ["Team", "Opponent"], ["Year", "Week"], ["Win streak vs this opponent"])
        except Exception as e:
            _log_exc(debug, "team_week_terminalize", e)

    # --------------------------
    # Distinct trade events split into Offseason / Inseason / Total (user
    # request) for the team_year/all_time and league_year/all_time sheets.
    # Offseason = trade dated before that season's kickoff (Sept 7).
    def _trade_is_offseason(dt_str, season) -> Optional[bool]:
        try:
            d = datetime.fromisoformat(str(dt_str).replace("Z", "+00:00")).date()
            return d < date(int(season), 9, 7)
        except Exception:
            return None
    _team_trade_dates_split: Dict[Tuple[str, int], Dict[str, set]] = defaultdict(
        lambda: {"off": set(), "in": set(), "tot": set()}
    )
    _league_trade_dates_split: Dict[int, Dict[str, set]] = defaultdict(
        lambda: {"off": set(), "in": set(), "tot": set()}
    )
    for _tr in trades_rows:
        _tm = _tr.get("Team"); _d = _tr.get("Date"); _sy = _to_int(_tr.get("Season"), None)
        if not (_tm and _d and _sy is not None):
            continue
        _off = _trade_is_offseason(_d, _sy)
        _bucket = "off" if _off else "in"
        _team_trade_dates_split[(str(_tm), int(_sy))][_bucket].add(str(_d))
        _team_trade_dates_split[(str(_tm), int(_sy))]["tot"].add(str(_d))
        _league_trade_dates_split[int(_sy)][_bucket].add(str(_d))
        _league_trade_dates_split[int(_sy)]["tot"].add(str(_d))

    # Rollups: player_year/all_time, team_year/all_time, league_week/year/all_time
    # --------------------------
    if (
        not pw.empty
        and not tw.empty
        and {"Team", "Year", "Week"}.issubset(pw.columns)
        and {"Team", "Year", "Week", "Win?"}.issubset(tw.columns)
    ):
        tw_keys = tw[["Team", "Year", "Week", "Win?"]].copy()
        tw_keys["Team"] = tw_keys["Team"].astype(str)
        for col in ["Year", "Week"]:
            tw_keys[col] = pd.to_numeric(tw_keys[col], errors="coerce").astype("Int64").astype(object)
        win_map = tw_keys.set_index(["Team", "Year", "Week"])["Win?"].to_dict()

        pw_keys = pw[["Team", "Year", "Week"]].copy()
        pw_keys["Team"] = pw_keys["Team"].astype(str)
        for col in ["Year", "Week"]:
            pw_keys[col] = pd.to_numeric(pw_keys[col], errors="coerce").astype("Int64").astype(object)
        pw["Team win?"] = [
            win_map.get((team, year, week))
            for team, year, week in pw_keys.itertuples(index=False, name=None)
        ]

    # -------------------------------------------------------------------
    # Phase 3A.2: Player tenures from transactions + trades.
    # A tenure = (team, start_iso, end_iso?) representing a continuous
    # span where the player was on that team's roster. Built from the
    # transactions/trades event ledger so it catches partial-week
    # sessions invisible to player_week (e.g. Hunter Renfrow's 5th
    # team — added and dropped between roster snapshots).
    #
    # Used downstream to:
    #   - augment "Number of teams" (union with pw-derived teams)
    #   - populate top_team / last_team for pad rows of tx-only players
    #     in player_year / player_all_time
    #   - identify "truly never rostered" players (no tenure + no pw row)
    #     whose pad rows should be dropped
    # -------------------------------------------------------------------
    player_tenures: Dict[str, List[Dict[str, Any]]] = {}
    try:
        events: List[Tuple[str, str, str, str]] = []  # (pid, date_iso, team, kind)
        for _r in transactions_rows:
            _date_s = _r.get("Date")
            if not _date_s:
                continue
            _team = _r.get("Team")
            _added = _r.get("_added_pid")
            _dropped = _r.get("_dropped_pid")
            if _added and _team:
                events.append((str(_added), str(_date_s), str(_team), "add"))
            # The dropped player on the same row was on `team` before the
            # tx — its tenure on this team ends at the tx date.
            if _dropped and _team:
                events.append((str(_dropped), str(_date_s), str(_team), "drop"))
        for _r in trades_rows:
            _date_s = _r.get("Date")
            if not _date_s:
                continue
            _team = _r.get("Team")
            if not _team:
                continue
            for _pid in (_r.get("_recv_player_ids") or []):
                if _pid:
                    events.append((str(_pid), str(_date_s), str(_team), "add"))
            for _pid in (_r.get("_drop_player_ids") or []):
                if _pid:
                    events.append((str(_pid), str(_date_s), str(_team), "drop"))
        # EVERY made pick (real rookie picks + the synthetic 2.09 / 5.0X) is a
        # roster acquisition: the drafted/awarded player JOINED the final team at
        # the draft. Without this the draft has no 'add' event, so a player later
        # traded/dropped off the drafting team loses that whole tenure span (the
        # 'drop from drafter' has no matching add and is skipped) — e.g. Najee
        # Harris's 56 weeks on stevenb123 vanished and his Top team read LWebs53.
        # Anchor the add at the draft (≈ Aug 28 of the pick year), matching the
        # pick PPG/tenure window. Pure draft-and-hold players (no later events)
        # are unchanged — they previously fell back to player_week and now have
        # an equivalent open tenure on the drafter.
        if not ph.empty and {"Year", "Final Team", "Player Picked"}.issubset(set(ph.columns)):
            for _, _phr in ph.iterrows():
                _ply = str(_phr.get("Player Picked") or "").strip()
                _pid = _phr.get("_player_id")
                _tm = str(_phr.get("Final Team") or "").strip()
                if not _pid or not _tm or _ply.lower() in ("", "unknown", "n/a", "nan"):
                    continue
                _ym = re.match(r"\s*(\d{4})", str(_phr.get("Year") or ""))
                if not _ym:
                    continue
                # tz-AWARE anchor (other tenure events are offset-aware; a bare
                # date string parses naive and breaks the start/end comparison).
                _anchor = datetime(int(_ym.group(1)), 8, 28, tzinfo=timezone.utc).isoformat()
                events.append((str(_pid), _anchor, _tm, "add"))

        by_pid: Dict[str, List[Tuple[str, str, str]]] = defaultdict(list)
        for _pid, _d, _team, _kind in events:
            by_pid[_pid].append((_d, _team, _kind))

        for _pid, _evs in by_pid.items():
            _evs.sort(key=lambda e: e[0])
            tenures: List[Dict[str, Any]] = []
            open_team: Optional[str] = None
            open_start: Optional[str] = None
            for _d, _team, _kind in _evs:
                if _kind == "add":
                    # If we already have an open tenure on a different
                    # team, close it (a missing drop tx); ignore re-adds
                    # to the same team (idempotent).
                    if open_team and open_team != _team:
                        tenures.append({"team": open_team, "start": open_start, "end": _d})
                        open_team = None
                        open_start = None
                    if not open_team:
                        open_team = _team
                        open_start = _d
                elif _kind == "drop":
                    if open_team == _team:
                        tenures.append({"team": _team, "start": open_start, "end": _d})
                        open_team = None
                        open_start = None
                    # else: drop without matching add (player joined
                    # before our tx window) — skipped silently
            if open_team:
                tenures.append({"team": open_team, "start": open_start, "end": None})
            player_tenures[_pid] = tenures
    except Exception as e:
        _log_exc(debug, "player_tenures_build", e)

    # Derived aggregates: number-of-teams, time-per-team-year,
    # time-per-team-all-time, last team per year / all-time.
    def _iso_to_dt(s: Optional[str]) -> Optional[datetime]:
        if not s:
            return None
        try:
            return datetime.fromisoformat(str(s).replace("Z", "+00:00"))
        except Exception:
            return None

    _now_dt = datetime.now(timezone.utc)

    # Phase 3A.3 rewrite — fantasy year throughout. Calendar year is
    # never used here. Per user mandate: "we should never being using
    # calendar year for anything - always fantasy year".
    #
    # Fantasy year FY for a date d:
    #   d.month >= 9  ->  FY = d.year
    #   else          ->  FY = d.year - 1
    #
    # Two windows per FY:
    #   In-season   = [Sept 1 FY, Feb 1 FY+1]
    #   Full FY     = [Sept 1 FY, Sept 1 FY+1]   (in-season + offseason)
    #
    # Three FY-keyed aggregates per tenure:
    #   tenure_time_team_fy[(pid, FY)]           Full-FY seconds per team
    #                                            (used for Number of teams
    #                                            per FY — partial-week
    #                                            offseason stints count)
    #   tenure_inseason_time_team_fy[(pid, FY)]  In-season seconds per team
    #                                            (used for Top team per FY)
    #   tenure_last_event_fy[(pid, FY)]          Latest tenure end that
    #                                            falls inside that FY's
    #                                            full window (used for
    #                                            Last team per FY — Jan
    #                                            championship rolls back
    #                                            to the season's FY)
    # All-time variants collapse the FY axis.
    def _fy_for_date(_d: datetime) -> int:
        return _d.year if _d.month >= 9 else _d.year - 1

    def _fy_window(_fy: int, _tz_for_window) -> Tuple[datetime, datetime]:
        return (
            datetime(_fy, 9, 1, tzinfo=_tz_for_window),
            datetime(_fy + 1, 9, 1, tzinfo=_tz_for_window),
        )

    def _fy_inseason_window(_fy: int, _tz_for_window) -> Tuple[datetime, datetime]:
        return (
            datetime(_fy, 9, 1, tzinfo=_tz_for_window),
            datetime(_fy + 1, 2, 1, tzinfo=_tz_for_window),
        )

    tenure_teams_all: Dict[str, Set[str]] = defaultdict(set)
    tenure_time_team_all: Dict[str, Dict[str, float]] = defaultdict(lambda: defaultdict(float))
    tenure_time_team_fy: Dict[Tuple[str, int], Dict[str, float]] = defaultdict(lambda: defaultdict(float))
    tenure_inseason_time_team_fy: Dict[Tuple[str, int], Dict[str, float]] = defaultdict(lambda: defaultdict(float))
    tenure_inseason_time_team_all: Dict[str, Dict[str, float]] = defaultdict(lambda: defaultdict(float))
    tenure_last_event_fy: Dict[Tuple[str, int], Tuple[datetime, str]] = {}
    tenure_last_event_all: Dict[str, Tuple[datetime, str]] = {}

    for _pid, tenures in player_tenures.items():
        for t in tenures:
            tm = t.get("team")
            if not tm:
                continue
            tenure_teams_all[_pid].add(str(tm))
            s_dt = _iso_to_dt(t.get("start")) or _now_dt
            e_dt = _iso_to_dt(t.get("end")) or _now_dt
            if e_dt < s_dt:
                e_dt = s_dt
            _tz = s_dt.tzinfo or timezone.utc

            # All-time tenure time + last event.
            tenure_time_team_all[_pid][str(tm)] += (e_dt - s_dt).total_seconds()
            cur_last = tenure_last_event_all.get(_pid)
            if cur_last is None or e_dt > cur_last[0]:
                tenure_last_event_all[_pid] = (e_dt, str(tm))

            # Per-FY accumulation. A tenure can straddle multiple FYs.
            for _fy in range(_fy_for_date(s_dt), _fy_for_date(e_dt) + 1):
                _fy_start, _fy_end = _fy_window(_fy, _tz)
                _ovl_s = max(s_dt, _fy_start)
                _ovl_e = min(e_dt, _fy_end)
                if _ovl_e > _ovl_s:
                    _secs = (_ovl_e - _ovl_s).total_seconds()
                    # Full-FY tenure time — used for Number of teams
                    # per FY (partial-week offseason stints should
                    # count for uniqueness).
                    tenure_time_team_fy[(_pid, _fy)][str(tm)] += _secs

                # In-season overlap within this FY.
                _is_start, _is_end = _fy_inseason_window(_fy, _tz)
                _is_ovl_s = max(s_dt, _is_start)
                _is_ovl_e = min(e_dt, _is_end)
                if _is_ovl_e > _is_ovl_s:
                    _is_secs = (_is_ovl_e - _is_ovl_s).total_seconds()
                    tenure_inseason_time_team_fy[(_pid, _fy)][str(tm)] += _is_secs
                    tenure_inseason_time_team_all[_pid][str(tm)] += _is_secs
                    # Last team per FY — gated to IN-SEASON only.
                    # An offseason trade in March/April would otherwise
                    # outrank a championship-week event for "Last team
                    # in FY Y" (Aaron Jones 2023 audit case). Limiting
                    # last-event to the in-season window correctly
                    # attributes the season-ending team.
                    cur_fy_last = tenure_last_event_fy.get((_pid, _fy))
                    if cur_fy_last is None or _is_ovl_e > cur_fy_last[0]:
                        tenure_last_event_fy[(_pid, _fy)] = (_is_ovl_e, str(tm))

    # Phase 3B: NFLverse full-season points aggregation, used for
    # "Points (full season)" + change-in columns + career stats on
    # player_year and player_all_time. nfl_log_by_sid entries have
    # year/week/points/_wk_date keys (built per-season during the
    # NFLverse fetch). Summing gives the player's full NFL season —
    # not just the weeks they were rostered in our league.
    nfl_full_season_points: Dict[Tuple[str, int], float] = defaultdict(float)
    nfl_full_season_games: Dict[Tuple[str, int], int] = defaultdict(int)
    nfl_career_total_points: Dict[str, float] = defaultdict(float)
    nfl_career_total_games: Dict[str, int] = defaultdict(int)
    # Phase 12 #5: full-season points blend. For weeks the player was on a LOTG
    # roster we have the league's OWN (Sleeper) score — authoritative, exact —
    # so use it; fall back to the nflverse re-score only for weeks we don't have
    # Sleeper data (non-rostered NFL games). This is exact on rostered weeks
    # (matches the rostered "Points" to the cent) and guarantees full ≥ rostered.
    _sleeper_by_sid: Dict[str, Dict[Tuple[int, int], float]] = defaultdict(dict)
    try:
        if not pw.empty and {"Player ID", "Year", "Week", "Points"}.issubset(pw.columns):
            _pwp = pd.to_numeric(pw["Points"], errors="coerce")
            _pwy = pd.to_numeric(pw["Year"], errors="coerce")
            _pww = pd.to_numeric(pw["Week"], errors="coerce")
            for _i, _sid_v in enumerate(pw["Player ID"].astype(str)):
                _yv, _wv, _pv = _pwy.iloc[_i], _pww.iloc[_i], _pwp.iloc[_i]
                if pd.notna(_yv) and pd.notna(_wv) and pd.notna(_pv):
                    _sleeper_by_sid[_sid_v][(int(_yv), int(_wv))] = float(_pv)
    except Exception as e:
        _log_exc(debug, "sleeper_pts_index", e)
    for _sid, _entries in nfl_log_by_sid.items():
        # Per (sid): nflverse re-scored weeks are the "games" (weeks the player
        # actually played); OVERRIDE each with the exact Sleeper score where the
        # player was rostered. Add Sleeper weeks nflverse MISSED only if they
        # carry a non-zero score (a real game nflverse lacks) — never bye/0
        # weeks, which would wrongly inflate the games denominator.
        _by_yw: Dict[Tuple[int, int], float] = {}
        for _e in _entries:
            try:
                _by_yw[(int(_e.get("year")), int(_e.get("week")))] = float(_e.get("points") or 0.0)
            except Exception:
                continue
        _sl = _sleeper_by_sid.get(str(_sid)) or {}
        for _yw, _spts in _sl.items():
            if _yw in _by_yw:
                _by_yw[_yw] = _spts          # rostered + played -> exact Sleeper score
            elif _spts != 0.0:
                _by_yw[_yw] = _spts          # game nflverse missed (skip 0/bye weeks)
        for (_yr, _wk), _pts in _by_yw.items():
            nfl_full_season_points[(str(_sid), _yr)] += _pts
            nfl_full_season_games[(str(_sid), _yr)] += 1
            nfl_career_total_points[str(_sid)] += _pts
            nfl_career_total_games[str(_sid)] += 1

    # Phase 3A.3: top/last team lookups, all FY-keyed.
    # Top team = max in-season tenure time per FY.
    # Last team = team at the most recent tenure event whose end falls
    # inside that FY's full window (Sept Y - Sept Y+1).
    tenure_top_team_by_year: Dict[Tuple[str, int], str] = {}
    tenure_last_team_by_year: Dict[Tuple[str, int], str] = {}
    tenure_top_team_all: Dict[str, str] = {}
    tenure_last_team_all: Dict[str, str] = {}
    for (_pid, _fy), team_secs in tenure_inseason_time_team_fy.items():
        if team_secs:
            tenure_top_team_by_year[(_pid, _fy)] = max(team_secs.items(), key=lambda kv: kv[1])[0]
    for (_pid, _fy), (_dt, _team) in tenure_last_event_fy.items():
        tenure_last_team_by_year[(_pid, _fy)] = _team
    for _pid, team_secs in tenure_inseason_time_team_all.items():
        if team_secs:
            tenure_top_team_all[_pid] = max(team_secs.items(), key=lambda kv: kv[1])[0]
    for _pid, (_dt, _team) in tenure_last_event_all.items():
        tenure_last_team_all[_pid] = _team

    # Phase 3B build fix: precompute career cumulative full-season
    # points / games BEFORE each (sid, year). The previous code did
    # a nested py.groupby("Player ID").apply(g: g.apply(lambda r:
    # nfl_full_season_games.get((str(r["Player ID"]), ...))))  which
    # broke because the inner row Series didn't always carry the
    # Player ID column under the runner's pandas version. Plain dict
    # lookups in an .apply callback are stable and faster.
    nfl_career_points_before: Dict[Tuple[str, int], float] = {}
    nfl_career_games_before: Dict[Tuple[str, int], int] = {}
    nfl_career_years_before: Dict[Tuple[str, int], int] = {}
    _by_sid_yr: Dict[str, List[Tuple[int, int, float]]] = defaultdict(list)
    for (_sid_k, _yr_k), _pts in nfl_full_season_points.items():
        _by_sid_yr[_sid_k].append((_yr_k, nfl_full_season_games[(_sid_k, _yr_k)], _pts))
    for _sid_k, _entries in _by_sid_yr.items():
        _entries.sort()
        _cum_g = 0
        _cum_p = 0.0
        _cum_y = 0
        for _yr_k, _g, _p in _entries:
            nfl_career_points_before[(_sid_k, _yr_k)] = _cum_p
            nfl_career_games_before[(_sid_k, _yr_k)] = _cum_g
            nfl_career_years_before[(_sid_k, _yr_k)] = _cum_y
            _cum_g += _g
            _cum_p += _p
            _cum_y += 1

    player_year = pd.DataFrame()
    player_all = pd.DataFrame()
    if not pw.empty:
        pw_work = pw.copy()
        if "Player ID" in pw_work.columns:
            pw_work["Player ID"] = pw_work["Player ID"].astype(str)
        pw_work["Points"] = pd.to_numeric(pw_work["Points"], errors="coerce").fillna(0.0)
        pw_work["Missed_injury"] = (pw_work["Injury?"].fillna(False) & (pw_work["Points"] == 0)).astype(int)
        pw_work["Missed_suspension"] = (pw_work["Suspension?"].fillna(False) & (pw_work["Points"] == 0)).astype(int)
        pw_work["Starter?"] = (pw_work["Starter/Bench"] == "Starter").astype(int)

        award_cols = [
            "Player of the week?",
            "QB of the week?",
            "RB of the week?",
            "WR of the week?",
            "TE of the week?",
            "Benchwarmer of the week?",
            "Bench QB of the week?",
            "Bench RB of the week?",
            "Bench WR of the week?",
            "Bench TE of the week?",
            "Highest starter on team?",
            "Lowest starter on team?",
            "Captain?",
        ]

        pw_work[award_cols] = pw_work[award_cols].fillna(0)

        team_points = pw_work.groupby(["Player ID", "Year", "Team"], as_index=False)["Points"].sum()
        top_team = (
            team_points.sort_values(["Player ID", "Year", "Points"], ascending=[True, True, False])
            .drop_duplicates(["Player ID", "Year"])
            .rename(columns={"Team": "Top Team", "Points": "Top Team Points"})
        )
        last_team = (
            pw_work.sort_values(["Player ID", "Year", "Week"])
            .groupby(["Player ID", "Year"])
            .tail(1)[["Player ID", "Year", "Team"]]
            .rename(columns={"Team": "Last team"})
        )

        # Per-season rookie flag is taken as max across the season's weeks (a player
        # either was a rookie in YYYY or wasn't — every week of that season agrees).
        # Per-season age is the mean of weekly age (effectively mid-season age).
        pw_work["_rookie_int"] = pd.to_numeric(pw_work.get("Rookie?"), errors="coerce").fillna(0).astype(int)
        pw_work["_age_num"] = pd.to_numeric(pw_work.get("Age"), errors="coerce")

        # PPG split by starter/bench needs separate sums + counts.
        pw_work["_starter_points"] = pw_work["Points"].where(pw_work["Starter?"] == 1, other=0.0)
        pw_work["_bench_points"] = pw_work["Points"].where(pw_work["Starter?"] == 0, other=0.0)
        pw_work["_bench_weeks"] = (pw_work["Starter?"] == 0).astype(int)
        # Phase 1C (items 1+2): ALSO carry an injury/suspension/bye-adjusted
        # variant for every per-player average. Non-adjusted columns stay
        # as-is (unchanged for player_year display). Adjusted columns get
        # surfaced next to them AND are what downstream consumers
        # (transactions / trades / change-from-prev) read.
        pw_work["_played"] = (~pw_work[["Injury?", "Suspension?", "Bye?"]].fillna(False).any(axis=1)).astype(int)
        pw_work["_played_points"] = pw_work["Points"] * pw_work["_played"]
        pw_work["_played_starter_points"] = pw_work["_starter_points"] * pw_work["_played"]
        pw_work["_played_starter_weeks"] = ((pw_work["Starter?"] == 1) & (pw_work["_played"] == 1)).astype(int)
        pw_work["_played_bench_points"] = pw_work["_bench_points"] * pw_work["_played"]
        pw_work["_played_bench_weeks"] = ((pw_work["Starter?"] == 0) & (pw_work["_played"] == 1)).astype(int)

        py_base = pw_work.groupby(["Player ID", "Year"], as_index=False).agg(
            Player=("Player", "first"),
            Points=("Points", "sum"),
            Avg_points=("Points", "mean"),
            Weeks_missed_injury=("Missed_injury", "sum"),
            Weeks_missed_suspension=("Missed_suspension", "sum"),
            Weeks_as_starter=("Starter?", "sum"),
            Weeks_as_bench=("_bench_weeks", "sum"),
            Number_of_teams=("Team", "nunique"),
            Weeks=("Points", "count"),
            Rookie_flag=("_rookie_int", "max"),
            Age_avg=("_age_num", "mean"),
            Starter_points_sum=("_starter_points", "sum"),
            Bench_points_sum=("_bench_points", "sum"),
            Played_points=("_played_points", "sum"),
            Played_weeks=("_played", "sum"),
            Played_starter_points=("_played_starter_points", "sum"),
            Played_starter_weeks=("_played_starter_weeks", "sum"),
            Played_bench_points=("_played_bench_points", "sum"),
            Played_bench_weeks=("_played_bench_weeks", "sum"),
            **{c: (c, "sum") for c in award_cols},
        )

        # Convert sums + counts -> averages, then drop the helper sum columns.
        py_base["PPG starter"] = py_base.apply(
            lambda r: round(r["Starter_points_sum"] / r["Weeks_as_starter"], 4)
            if r["Weeks_as_starter"] else None,
            axis=1,
        )
        py_base["PPG bench"] = py_base.apply(
            lambda r: round(r["Bench_points_sum"] / r["Weeks_as_bench"], 4)
            if r["Weeks_as_bench"] else None,
            axis=1,
        )
        # Adjusted variants (Phase 1C): bye/injury/suspension excluded.
        py_base["Adjusted Avg points"] = py_base.apply(
            lambda r: round(r["Played_points"] / r["Played_weeks"], 4)
            if r["Played_weeks"] else None,
            axis=1,
        )
        py_base["Adjusted PPG starter"] = py_base.apply(
            lambda r: round(r["Played_starter_points"] / r["Played_starter_weeks"], 4)
            if r["Played_starter_weeks"] else None,
            axis=1,
        )
        py_base["Adjusted PPG bench"] = py_base.apply(
            lambda r: round(r["Played_bench_points"] / r["Played_bench_weeks"], 4)
            if r["Played_bench_weeks"] else None,
            axis=1,
        )
        # PPG starter vs bench diff: per Phase 1C clarification, derived
        # consumers of player averages use the bye/injury/suspension-
        # adjusted variants. Difference is therefore between
        # Adjusted PPG starter and Adjusted PPG bench.
        # Treat missing side as 0 (a player with no played bench weeks
        # has effective PPG bench = 0; ditto starter). Return None only
        # when BOTH are missing (player has no played weeks at all).
        # NB: pandas NaN is NOT `is None` — must use pd.isna() to catch
        # both. Last audit (Rashee Rice 2024 with no played bench
        # weeks) caught this: bn came through as NaN, the prior
        # `is None` check passed False, and `NaN or 0` returned NaN
        # which propagated to a NaN diff and ultimately a 0 in the
        # filled output.
        def _ppg_diff(r):
            st = r.get("Adjusted PPG starter")
            bn = r.get("Adjusted PPG bench")
            st_na = pd.isna(st)
            bn_na = pd.isna(bn)
            if st_na and bn_na:
                return None
            return round(float(0 if st_na else st) - float(0 if bn_na else bn), 4)
        py_base["PPG starter vs bench diff"] = py_base.apply(_ppg_diff, axis=1)
        py_base["Rookie?"] = py_base["Rookie_flag"].astype(bool)
        py_base["Age"] = py_base["Age_avg"].round(2)
        # Drop intermediate helpers but keep Played_* — change-in-career
        # below needs them; output catalog filter drops them at write.
        py_base = py_base.drop(columns=[
            "Rookie_flag", "Age_avg", "Starter_points_sum", "Bench_points_sum",
            "Played_starter_points", "Played_starter_weeks",
            "Played_bench_points", "Played_bench_weeks",
        ])

        py = py_base.merge(top_team[["Player ID", "Year", "Top Team"]], on=["Player ID", "Year"], how="left")
        py = py.merge(last_team, on=["Player ID", "Year"], how="left")

        # Phase 3A.3: replace pw-derived top_team / last_team (weeks-on-
        # roster + last-week heuristic) with tenure-time / tenure-last-
        # event values for every player_year row. Falls back to the
        # pw-based value when no tenure exists for the (pid, year)
        # — players on startup roster with no transactions yet remain
        # unchanged because they only have ONE team either way.
        try:
            def _tenure_top(pid: Any, yr: Any, fallback: Any) -> Any:
                try:
                    return tenure_top_team_by_year.get((str(pid), int(yr)), fallback)
                except Exception:
                    return fallback
            def _tenure_last(pid: Any, yr: Any, fallback: Any) -> Any:
                try:
                    return tenure_last_team_by_year.get((str(pid), int(yr)), fallback)
                except Exception:
                    return fallback
            py["Top Team"] = py.apply(
                lambda r: _tenure_top(r.get("Player ID"), r.get("Year"), r.get("Top Team")),
                axis=1,
            )
            py["Last team"] = py.apply(
                lambda r: _tenure_last(r.get("Player ID"), r.get("Year"), r.get("Last team")),
                axis=1,
            )
        except Exception as e:
            _log_exc(debug, "player_year_top_last_team_tenure_override", e)

        # Phase 3A.2: augment Number of teams with tenure-based teams.
        # Catches partial-week sessions that pw misses (Renfrow's 5th
        # team was added + dropped between weekly snapshots).
        try:
            def _expanded_team_count(_pid: Any, _yr: Any, _existing: Any) -> int:
                _pid_s = str(_pid)
                try:
                    _yr_i = int(_yr)
                except Exception:
                    _yr_i = None
                pw_teams: Set[str] = set()
                # The existing count is from pw groupby. We also want the
                # tenure teams overlapping this year.
                tenure_teams_this_year = set(tenure_time_team_fy.get((_pid_s, _yr_i), {}).keys()) if _yr_i else set()
                # Combine: take the larger of existing count vs the tenure
                # set size — if tenures show MORE teams (i.e. include
                # partial-week stints not in pw), use that.
                try:
                    existing_n = int(_existing) if pd.notna(_existing) else 0
                except Exception:
                    existing_n = 0
                return max(existing_n, len(tenure_teams_this_year))
            py["Number_of_teams"] = py.apply(
                lambda r: _expanded_team_count(r.get("Player ID"), r.get("Year"), r.get("Number_of_teams")),
                axis=1,
            )
        except Exception as e:
            _log_exc(debug, "player_year_teams_with_tenures", e)

        # Phase 3A item 4: % of points redefined.
        # OLD: max/min share of the PLAYER'S own points across their teams.
        # NEW: % contribution as a starter to that TEAM'S total starter
        #      points over the year. Adds the team name alongside each
        #      percentage so it's clear which team each number describes.
        try:
            starter_pw = pw_work[pw_work.get("Starter?") == 1].copy()
            if not starter_pw.empty:
                player_starter_per_team = starter_pw.groupby(
                    ["Player ID", "Year", "Team"]
                )["Points"].sum().rename("_player_pts")
                team_starter_totals = starter_pw.groupby(["Team", "Year"])["Points"].sum().rename("_team_pts")
                # Join the team total onto each (Player, Year, Team) row.
                shares = (
                    player_starter_per_team.reset_index()
                    .merge(team_starter_totals.reset_index(), on=["Team", "Year"], how="left")
                )
                shares["_share"] = shares["_player_pts"] / shares["_team_pts"].replace(0, np.nan)
                # Per (Player, Year): find the team with highest and lowest share.
                hi = (
                    shares.sort_values(["Player ID", "Year", "_share"], ascending=[True, True, False])
                    .drop_duplicates(["Player ID", "Year"])
                    [["Player ID", "Year", "Team", "_share"]]
                    .rename(columns={"Team": "Team for highest % of points", "_share": "% of points (highest team)"})
                )
                lo = (
                    shares.sort_values(["Player ID", "Year", "_share"], ascending=[True, True, True])
                    .drop_duplicates(["Player ID", "Year"])
                    [["Player ID", "Year", "Team", "_share"]]
                    .rename(columns={"Team": "Team for lowest % of points", "_share": "% of points (lowest team)"})
                )
                py = py.merge(hi, on=["Player ID", "Year"], how="left")
                py = py.merge(lo, on=["Player ID", "Year"], how="left")
                py["% of points (highest team)"] = py["% of points (highest team)"].round(4)
                py["% of points (lowest team)"] = py["% of points (lowest team)"].round(4)
            else:
                py["Team for highest % of points"] = None
                py["% of points (highest team)"] = None
                py["Team for lowest % of points"] = None
                py["% of points (lowest team)"] = None
        except Exception as e:
            _log_exc(debug, "player_year_pct_of_points", e)

        py = py.sort_values(["Player ID", "Year"]).reset_index(drop=True)
        # Phase 3B: NFLverse full-season columns + change-in rewrites.
        # "Points" stays as rostered-only (LOTG league totals); the new
        # "Points (full season)" sums the player's entire NFL season
        # from nflverse, filling the gap for weeks they weren't on a
        # fantasy roster.
        py["Points (full season)"] = py.apply(
            lambda r: round(float(nfl_full_season_points.get((str(r.get("Player ID")), int(r.get("Year"))), 0.0)), 2),
            axis=1,
        )
        py["Avg points (full season)"] = py.apply(
            lambda r: round(
                nfl_full_season_points.get((str(r.get("Player ID")), int(r.get("Year"))), 0.0)
                / nfl_full_season_games.get((str(r.get("Player ID")), int(r.get("Year"))), 1),
                4,
            ) if nfl_full_season_games.get((str(r.get("Player ID")), int(r.get("Year"))), 0) else None,
            axis=1,
        )

        # Change-in (per user spec): use FULL-SEASON values, not rostered.
        # Rookie years stay N/A because the diff against an absent prior
        # season is undefined.
        py["Change in points from previous season"] = (
            py.groupby("Player ID")["Points (full season)"].diff()
        )
        py["Change in avg points from previous season"] = (
            py.groupby("Player ID")["Avg points (full season)"].diff()
        )

        # "Change in points from career" — full-season points this year
        # minus the average season-total of prior NFL years.
        # "Change in avg points from career" — full-season per-game avg
        # this year minus prior career per-game avg.
        # Both lookups are O(1) against the precomputed nfl_career_*_before
        # dicts (built right after the NFLverse aggregation) — replaces
        # an unstable nested groupby.apply that broke the build under the
        # CI pandas version.
        def _change_pts_career(r):
            pid_s = str(r.get("Player ID"))
            try:
                yr_i = int(r.get("Year"))
            except Exception:
                return None
            pre_pts = nfl_career_points_before.get((pid_s, yr_i))
            pre_yrs = nfl_career_years_before.get((pid_s, yr_i), 0)
            if not pre_yrs or pre_pts is None:
                return None
            try:
                cur = r.get("Points (full season)")
                if cur is None:
                    return None
                return round(float(cur) - (pre_pts / pre_yrs), 4)
            except Exception:
                return None
        py["Change in points from career"] = py.apply(_change_pts_career, axis=1)

        def _change_avg_career(r):
            pid_s = str(r.get("Player ID"))
            try:
                yr_i = int(r.get("Year"))
            except Exception:
                return None
            pre_pts = nfl_career_points_before.get((pid_s, yr_i))
            pre_g = nfl_career_games_before.get((pid_s, yr_i), 0)
            if not pre_g or pre_pts is None:
                return None
            cur = r.get("Avg points (full season)")
            if cur is None:
                return None
            try:
                return round(float(cur) - (pre_pts / pre_g), 4)
            except Exception:
                return None
        py["Change in avg points from career"] = py.apply(_change_avg_career, axis=1)

        py = py.rename(
            columns={
                "Avg_points": "Avg points",
                "Weeks_missed_injury": "Weeks missed due to injury",
                "Weeks_missed_suspension": "Weeks missed due to suspension",
                "Weeks_as_starter": "Weeks as starter",
                "Number_of_teams": "Number of teams",
                "Top Team": "Top Team",
            }
        )

        py["Number of transactions"] = [
            int(player_tx_year.get((str(player_id), int(year) if pd.notna(year) else None), 0))
            for player_id, year in py[["Player ID", "Year"]].itertuples(index=False, name=None)
        ]
        py["Number of drops"] = [
            int(player_drop_year.get((str(player_id), int(year) if pd.notna(year) else None), 0))
            for player_id, year in py[["Player ID", "Year"]].itertuples(index=False, name=None)
        ]

        # Per-player trade counts from trades_rows. Each trade in trades.csv
        # has one row per team involved; the players sent to that team are
        # listed in that row's 'Assets received' (sic — column name is
        # misspelled in the schema and we preserve it). To avoid double
        # counting we only read the 'received' side: every traded player
        # appears in exactly one team's received cell per trade.
        player_trade_year: Dict[Tuple[str, int], int] = defaultdict(int)
        player_trade_all: Dict[str, int] = defaultdict(int)
        try:
            for tr_row in trades_rows:
                # Prefer the fantasy-year Season (which keeps offseason
                # trades attributed to the league iteration they belong to);
                # fall back to the calendar year of the Date for safety.
                yr_val = tr_row.get("Season")
                if yr_val is None:
                    date_s = tr_row.get("Date")
                    if not date_s:
                        continue
                    try:
                        yr_val = int(str(date_s)[:4])
                    except Exception:
                        continue
                try:
                    yr = int(yr_val)
                except Exception:
                    continue
                recv = tr_row.get("Assets received")
                if not recv or str(recv) == "0.0":
                    continue
                for asset in str(recv).split(";"):
                    asset = asset.strip()
                    if not asset:
                        continue
                    # Skip pick labels (start with a 4-digit year, e.g.
                    # '2025 1.??' or '2026 R2'). Player names never start
                    # with a 4-digit number.
                    if re.match(r"^\d{4}\b", asset):
                        continue
                    player_trade_year[(asset, yr)] += 1
                    player_trade_all[asset] += 1
        except Exception as e:
            _log_exc(debug, "player_trade_count", e)

        py["Number of trades"] = [
            int(player_trade_year.get((str(player_name), int(year) if pd.notna(year) else 0), 0))
            for player_name, year in py[["Player", "Year"]].itertuples(index=False, name=None)
        ]

        py = py.rename(
            columns={
                "Player of the week?": "Times as Player of the week?",
                "QB of the week?": "Times as QB of the week?",
                "RB of the week?": "Times as RB of the week?",
                "WR of the week?": "Times as WR of the week?",
                "TE of the week?": "Times as TE of the week?",
                "Benchwarmer of the week?": "Times as Benchwarmer of the week?",
                "Bench QB of the week?": "Times as Bench QB of the week?",
                "Bench RB of the week?": "Times as Bench RB of the week?",
                "Bench WR of the week?": "Times as Bench WR of the week?",
                "Bench TE of the week?": "Times as Bench TE of the week?",
                "Highest starter on team?": "Times as Highest starter on team?",
                "Lowest starter on team?": "Times as Lowest starter on team?",
                "Captain?": "Times as Captain?",
            }
        )

        player_year = py

        # Ensure player_year has a row for every player×year combo
        # that appears in transactions/trades, not just in player_week.
        # Concrete case: Tom Brady was dropped to FA in 2023 (a real
        # transactions.csv row) but had no roster appearance, so the
        # pw-derived player_year was missing 2023 Brady — and Brady's
        # 'Number of transactions' summed across years was 8 vs the
        # 9 in player_all_time. Pad here with skeleton rows.
        try:
            existing = set()
            if not player_year.empty and "Player ID" in player_year.columns and "Year" in player_year.columns:
                for pid_val, yr_val in player_year[["Player ID", "Year"]].itertuples(index=False, name=None):
                    try:
                        existing.add((str(pid_val), int(yr_val)))
                    except Exception:
                        continue

            # Walk transactions / trades for every (player_id, year)
            # that contributed to the per-year counters.
            tx_pairs: Set[Tuple[str, int]] = set()
            # The team(s) each (sid, season) was actually ON per its real moves —
            # the row's `Team` is the team the player joined (add/received) or was
            # removed from (drop/sent). Used as a realized-roster signal + a
            # Top/Last-team source for players with no player_tenures stint (a
            # vet dropped off an INITIAL roster has a real drop but no recorded
            # 'add', so player_tenures can't see them). Fix #1.
            tx_team_events_by_pair: Dict[Tuple[str, int], List[Tuple[str, str]]] = defaultdict(list)

            def _add_tx_team_event(sid_v, season_v, date_v, team_v):
                if sid_v and team_v:
                    tx_team_events_by_pair[(str(sid_v), int(season_v))].append(
                        (str(date_v or ""), str(team_v))
                    )

            for r in transactions_rows:
                try:
                    season_i = int(r.get("Season")) if r.get("Season") is not None else None
                except Exception:
                    continue
                if season_i is None:
                    continue
                for fld in ("_added_pid", "_dropped_pid"):
                    sid = r.get(fld)
                    if sid:
                        tx_pairs.add((str(sid), int(season_i)))
                        _add_tx_team_event(sid, season_i, r.get("Date"), r.get("Team"))
            for r in trades_rows:
                try:
                    season_i = int(r.get("Season")) if r.get("Season") is not None else None
                except Exception:
                    continue
                if season_i is None:
                    continue
                for sid in (r.get("_recv_player_ids") or []):
                    if sid:
                        tx_pairs.add((str(sid), int(season_i)))
                        _add_tx_team_event(sid, season_i, r.get("Date"), r.get("Team"))
                for sid in (r.get("_drop_player_ids") or []):
                    if sid:
                        tx_pairs.add((str(sid), int(season_i)))
                        _add_tx_team_event(sid, season_i, r.get("Date"), r.get("Team"))

            missing_pairs = [p for p in tx_pairs if p not in existing]
            # Phase 3A.2: drop pads for players truly never rostered.
            # If sid has no tenure for the given year AND no pw row, the
            # transaction was purely phantom (e.g. an add+drop with no
            # roster realization). Skip those rows per spec — "If the
            # player was never rostered, they shouldn't get a row for
            # that year."
            def _season_leadin_tenure(sid: str, yr: int):
                """Realized stints in season `yr`'s LEAD-IN offseason window
                [Jan 1 yr, min(now, Sept 1 yr)).

                Sleeper tags a pre-season move (Jan–Aug yr) as Season=yr, but
                `_fy_for_date` buckets those dates into FY (yr-1) — an offseason
                date's fantasy year is the prior year. So a player whose ONLY
                Season-yr activity is a pre-season offseason move has no
                (sid, yr) FY tenure entry and would be dropped by the realized-
                roster filter below. Recover their membership + Top/Last team
                straight from player_tenures. Bounded at Sept 1 yr so it captures
                only the lead-in (in-season players already have an FY-yr entry +
                a pw-derived row, so they never reach this path). For the live
                season Sept 1 yr is in the future, so the bound is `now`.
                Returns (secs_by_team, last_event). Fix #1."""
                _lead_start = datetime(int(yr), 1, 1, tzinfo=timezone.utc)
                _lead_end = min(_now_dt, datetime(int(yr), 9, 1, tzinfo=timezone.utc))
                _secs: Dict[str, float] = defaultdict(float)
                _last: Optional[Tuple[datetime, str]] = None
                for t in player_tenures.get(str(sid), []):
                    tm = t.get("team")
                    if not tm:
                        continue
                    s_dt = _iso_to_dt(t.get("start"))
                    e_dt = _iso_to_dt(t.get("end")) or _now_dt
                    if s_dt is None:
                        continue
                    ov_s = max(s_dt, _lead_start)
                    ov_e = min(e_dt, _lead_end)
                    if ov_e > ov_s:
                        _secs[str(tm)] += (ov_e - ov_s).total_seconds()
                        if _last is None or ov_e > _last[0]:
                            _last = (ov_e, str(tm))
                return dict(_secs), _last

            def _tx_team_summary(sid: str, yr: int):
                """Top/Last team + distinct-team count from the (sid, yr) real
                transaction `Team`s — the fallback realized source when the
                player has no player_tenures stint (initial-roster vet dropped).
                Top = most-frequent team, Last = team of the latest-dated move."""
                evs = tx_team_events_by_pair.get((str(sid), int(yr)))
                if not evs:
                    return None, None, 0
                counts: Dict[str, int] = {}
                for _d, tm in evs:
                    counts[tm] = counts.get(tm, 0) + 1
                top = max(counts.items(), key=lambda kv: kv[1])[0]
                last = max(evs, key=lambda e: e[0])[1]
                return top, last, len(counts)

            def _has_tenure_in_year(sid: str, yr: int) -> bool:
                tenure_yr_map = tenure_time_team_fy.get((str(sid), int(yr)))
                if tenure_yr_map:
                    return True
                # Phase 12 fix #1: a Season-yr offseason-only realization is
                # bucketed under FY (yr-1), so the FY-keyed lookup above misses
                # it (any completed season, not just the live one). Check the
                # lead-in window directly.
                _secs, _ = _season_leadin_tenure(str(sid), int(yr))
                if _secs:
                    return True
                # A real move with a Team (e.g. a vet dropped off an initial
                # roster — a genuine roster removal with no recorded 'add', so no
                # tenure stint) is still a realized rostering.
                return bool(tx_team_events_by_pair.get((str(sid), int(yr))))
            missing_pairs = [(s, y) for s, y in missing_pairs if _has_tenure_in_year(s, y)]
            if missing_pairs:
                pad_rows = []
                for sid, yr in missing_pairs:
                    meta = pid_meta.get(str(sid)) or {}
                    name = meta.get("full_name") or str(sid)
                    # Derive top/last team from tenure aggregates for this
                    # (sid, yr) — top = team with the most tenure time in
                    # year, last = team in tenure_last_event_fy.
                    # Top team prefers in-season tenure time; falls back
                    # to full-FY tenure time if the player had no in-
                    # season tenure at all (rare).
                    tenure_team_secs_full = tenure_time_team_fy.get((str(sid), int(yr)), {})
                    _tx_top = _tx_last = None
                    _tx_nteams = 0
                    if tenure_team_secs_full:
                        tenure_team_secs_is = tenure_inseason_time_team_fy.get((str(sid), int(yr)), {})
                        tenure_team_secs = tenure_team_secs_is or tenure_team_secs_full
                        last_event = tenure_last_event_fy.get((str(sid), int(yr)))
                    else:
                        # Offseason-only Season-yr player: the FY dicts bucket the
                        # lead-in under yr-1, so recover teams from the lead-in
                        # window. Fix #1 (covers the live season AND historical
                        # pre-season-only stints, e.g. Dee Eskridge 2022).
                        _lead_secs, _lead_last = _season_leadin_tenure(str(sid), int(yr))
                        tenure_team_secs = _lead_secs
                        tenure_team_secs_full = _lead_secs
                        last_event = _lead_last
                        if not _lead_secs:
                            # No tenure stint at all (initial-roster vet with only
                            # a drop). Fall back to the real transaction Team(s).
                            _tx_top, _tx_last, _tx_nteams = _tx_team_summary(str(sid), int(yr))
                    top_team_pad = (
                        max(tenure_team_secs.items(), key=lambda kv: kv[1])[0]
                        if tenure_team_secs else _tx_top
                    )
                    last_team_pad = (
                        last_event[1] if last_event else (_tx_last or top_team_pad)
                    )
                    # Phase 12 fix #2: padded (tx-only) rows had no Age (defaulted
                    # to 0). Compute the player's age as of mid-season of that
                    # year from birth_date (same source the weekly path uses).
                    _pad_age = _calc_age(meta.get("birth_date"), date(int(yr), 11, 1))
                    pad_rows.append({
                        "Player": name,
                        "Player ID": str(sid),
                        "Year": int(yr),
                        "Age": round(_pad_age, 2) if _pad_age is not None else None,
                        "Top Team": top_team_pad,
                        "Last team": last_team_pad,
                        # Use full-FY teams for the uniqueness count so
                        # partial-week offseason stints still register; fall back
                        # to the real-transaction team count for initial-roster
                        # vets with no tenure stint (Fix #1).
                        "Number of teams": len(tenure_team_secs_full) or _tx_nteams,
                        # Counters mirror what the main pw-derived
                        # rows would have gotten from the same dicts.
                        "Number of transactions": int(player_tx_year.get((str(sid), int(yr)), 0)),
                        "Number of drops": int(player_drop_year.get((str(sid), int(yr)), 0)),
                        # player_trade_year is name-keyed (see comment
                        # on the main-path Number of trades fill).
                        "Number of trades": int(player_trade_year.get((str(name), int(yr)), 0)),
                    })
                if pad_rows:
                    pad_df = pd.DataFrame(pad_rows)
                    player_year = pd.concat([player_year, pad_df], ignore_index=True)
                    _log(debug, f"[{_now_iso()}] INFO seeded {len(pad_rows)} player_year rows for tx-only player×year combos")
        except Exception as e:
            _log_exc(debug, "player_year_tx_only_pad", e)

        pa = pw_work.groupby(["Player ID"], as_index=False).agg(
            Player=("Player", "first"),
            Points=("Points", "sum"),
            Avg_points=("Points", "mean"),
            Weeks_missed_injury=("Missed_injury", "sum"),
            Weeks_missed_suspension=("Missed_suspension", "sum"),
            Weeks_as_starter=("Starter?", "sum"),
            Weeks_as_bench=("_bench_weeks", "sum"),
            Number_of_teams=("Team", "nunique"),
            Starter_points_sum=("_starter_points", "sum"),
            Bench_points_sum=("_bench_points", "sum"),
            Played_points=("_played_points", "sum"),
            Played_weeks=("_played", "sum"),
            Played_starter_points=("_played_starter_points", "sum"),
            Played_starter_weeks=("_played_starter_weeks", "sum"),
            Played_bench_points=("_played_bench_points", "sum"),
            Played_bench_weeks=("_played_bench_weeks", "sum"),
            **{c: (c, "sum") for c in award_cols},
        )
        # Same formula as player_year: derive PPG starter / bench / diff
        # from the per-week sums, drop the helper columns at the end.
        pa["PPG starter"] = pa.apply(
            lambda r: round(r["Starter_points_sum"] / r["Weeks_as_starter"], 4)
            if r["Weeks_as_starter"] else None,
            axis=1,
        )
        pa["PPG bench"] = pa.apply(
            lambda r: round(r["Bench_points_sum"] / r["Weeks_as_bench"], 4)
            if r["Weeks_as_bench"] else None,
            axis=1,
        )
        # Phase 1C — bye/injury/suspension-adjusted variants must be
        # computed BEFORE PPG starter vs bench diff since the diff
        # consumes them.
        pa["Adjusted Avg points"] = pa.apply(
            lambda r: round(r["Played_points"] / r["Played_weeks"], 4)
            if r["Played_weeks"] else None,
            axis=1,
        )
        pa["Adjusted PPG starter"] = pa.apply(
            lambda r: round(r["Played_starter_points"] / r["Played_starter_weeks"], 4)
            if r["Played_starter_weeks"] else None,
            axis=1,
        )
        pa["Adjusted PPG bench"] = pa.apply(
            lambda r: round(r["Played_bench_points"] / r["Played_bench_weeks"], 4)
            if r["Played_bench_weeks"] else None,
            axis=1,
        )
        # PPG starter vs bench diff uses the adjusted variants. Same
        # NaN-as-zero handling as player_year (pd.isna catches both
        # None and NaN; raw `is None` missed pandas NaN in the last
        # audit).
        def _pa_ppg_diff(r):
            st = r.get("Adjusted PPG starter")
            bn = r.get("Adjusted PPG bench")
            st_na = pd.isna(st)
            bn_na = pd.isna(bn)
            if st_na and bn_na:
                return None
            return round(float(0 if st_na else st) - float(0 if bn_na else bn), 4)
        pa["PPG starter vs bench diff"] = pa.apply(_pa_ppg_diff, axis=1)
        pa = pa.drop(columns=[
            "Starter_points_sum", "Bench_points_sum", "Weeks_as_bench",
            "Played_points", "Played_weeks",
            "Played_starter_points", "Played_starter_weeks",
            "Played_bench_points", "Played_bench_weeks",
        ], errors="ignore")

        top_team_all = (
            pw_work.groupby(["Player ID", "Team"], as_index=False)["Points"].sum()
            .sort_values(["Player ID", "Points"], ascending=[True, False])
            .drop_duplicates(["Player ID"])
            .rename(columns={"Team": "Top team", "Points": "Top team points"})
        )
        last_team_all = (
            pw_work.sort_values(["Year", "Week"])
            .groupby("Player ID")
            .tail(1)[["Player ID", "Team", "Rookie?", "Age"]]
            .rename(columns={"Team": "Last team"})
        )
        pa = pa.merge(top_team_all[["Player ID", "Top team"]], on="Player ID", how="left")
        pa = pa.merge(last_team_all, on="Player ID", how="left")

        # Phase 3A.3: replace pw-derived Top team / Last team with
        # tenure-based values (time rostered + last tenure end) for
        # every row. pw-based remains as the fallback when no tenure
        # exists for the player.
        try:
            pa["Top team"] = pa.apply(
                lambda r: tenure_top_team_all.get(str(r.get("Player ID")), r.get("Top team")),
                axis=1,
            )
            pa["Last team"] = pa.apply(
                lambda r: tenure_last_team_all.get(str(r.get("Player ID")), r.get("Last team")),
                axis=1,
            )
        except Exception as e:
            _log_exc(debug, "player_all_top_last_team_tenure_override", e)

        # Phase 3A.2: augment Number of teams with tenure-based teams
        # (partial-week sessions invisible to pw-derived nunique).
        try:
            def _expanded_team_count_all(_pid: Any, _existing: Any) -> int:
                _pid_s = str(_pid)
                tenure_teams = tenure_teams_all.get(_pid_s, set())
                try:
                    existing_n = int(_existing) if pd.notna(_existing) else 0
                except Exception:
                    existing_n = 0
                return max(existing_n, len(tenure_teams))
            pa["Number_of_teams"] = pa.apply(
                lambda r: _expanded_team_count_all(r.get("Player ID"), r.get("Number_of_teams")),
                axis=1,
            )
        except Exception as e:
            _log_exc(debug, "player_all_teams_with_tenures", e)

        # Phase 3A item 4 (all-time variant): % of points = player's
        # starter contribution to the TEAM's all-time starter total.
        # Adds team-name columns for highest and lowest %.
        try:
            starter_pw_at = pw_work[pw_work.get("Starter?") == 1].copy()
            if not starter_pw_at.empty:
                player_starter_at = starter_pw_at.groupby(
                    ["Player ID", "Team"]
                )["Points"].sum().rename("_player_pts")
                team_starter_at = starter_pw_at.groupby("Team")["Points"].sum().rename("_team_pts")
                shares_at = (
                    player_starter_at.reset_index()
                    .merge(team_starter_at.reset_index(), on="Team", how="left")
                )
                shares_at["_share"] = shares_at["_player_pts"] / shares_at["_team_pts"].replace(0, np.nan)
                hi_at = (
                    shares_at.sort_values(["Player ID", "_share"], ascending=[True, False])
                    .drop_duplicates(["Player ID"])
                    [["Player ID", "Team", "_share"]]
                    .rename(columns={"Team": "Team for highest % of points", "_share": "% of points (highest team)"})
                )
                lo_at = (
                    shares_at.sort_values(["Player ID", "_share"], ascending=[True, True])
                    .drop_duplicates(["Player ID"])
                    [["Player ID", "Team", "_share"]]
                    .rename(columns={"Team": "Team for lowest % of points", "_share": "% of points (lowest team)"})
                )
                pa = pa.merge(hi_at, on="Player ID", how="left")
                pa = pa.merge(lo_at, on="Player ID", how="left")
                pa["% of points (highest team)"] = pa["% of points (highest team)"].round(4)
                pa["% of points (lowest team)"] = pa["% of points (lowest team)"].round(4)
            else:
                pa["Team for highest % of points"] = None
                pa["% of points (highest team)"] = None
                pa["Team for lowest % of points"] = None
                pa["% of points (lowest team)"] = None
        except Exception as e:
            _log_exc(debug, "player_all_pct_of_points", e)

        # Phase 3B: NFLverse career columns on player_all_time. Named "(full
        # career)" (not "(full season)" as on player_year) because all-time pools
        # every season: "Points (full career)" = career NFL points; "Avg points
        # (full career)" = career per-game average.
        pa["Points (full career)"] = pa.apply(
            lambda r: round(float(nfl_career_total_points.get(str(r.get("Player ID")), 0.0)), 2),
            axis=1,
        )
        pa["Avg points (full career)"] = pa.apply(
            lambda r: round(
                nfl_career_total_points.get(str(r.get("Player ID")), 0.0)
                / nfl_career_total_games.get(str(r.get("Player ID")), 1),
                4,
            ) if nfl_career_total_games.get(str(r.get("Player ID")), 0) else None,
            axis=1,
        )

        # Phase 3B: Taxi-eligible boolean.
        # TRUE if: player is currently in their first year in the
        # league AND has never started for any team. Resets at week 1
        # of the following season (so rookies finishing year 1 stay
        # eligible until first snap of year 2).
        try:
            # Latest season we have data for (proxy for "current").
            current_season = int(pw["Year"].max()) if not pw.empty and "Year" in pw.columns else None
            # First SEASON per player — from player_year only. (Earlier this also
            # folded in tenure_time_team_fy, but an OFFSEASON acquisition is filed
            # under the PRIOR fantasy year, so a rookie added in the offseason read
            # as a year early — wrongly failing the first-year test below. A player
            # in player_all_time always has player_year rows, so min(player_year)
            # is the correct, pollution-free first season.)
            first_year_by_pid: Dict[str, int] = {}
            if not player_year.empty and "Player ID" in player_year.columns and "Year" in player_year.columns:
                _g = player_year.dropna(subset=["Player ID", "Year"]).groupby("Player ID")["Year"].min()
                for _p, _y in _g.items():
                    try:
                        first_year_by_pid[str(_p)] = int(_y)
                    except Exception:
                        continue

            def _is_taxi_eligible(row) -> bool:
                if current_season is None:
                    return False
                pid = str(row.get("Player ID"))
                fy = first_year_by_pid.get(pid)
                if fy is None or fy != current_season:
                    return False
                # NB: pa is still pre-rename here — the column is 'Weeks_as_starter'
                # (it becomes 'Weeks as starter' a few lines below). Read both so
                # the no-start test actually fires (it silently read 0 before).
                _ws = row.get("Weeks_as_starter")
                if _ws is None:
                    _ws = row.get("Weeks as starter")
                try:
                    weeks_started = int(pd.to_numeric(_ws, errors="coerce") or 0)
                except Exception:
                    weeks_started = 0
                return weeks_started == 0
            pa["Taxi-eligible"] = pa.apply(_is_taxi_eligible, axis=1)
        except Exception as e:
            _log_exc(debug, "player_all_taxi_eligible", e)

        pa = pa.rename(
            columns={
                "Avg_points": "Avg points",
                "Weeks_missed_injury": "Weeks missed due to injury",
                "Weeks_missed_suspension": "Weeks missed due to suspension",
                "Weeks_as_starter": "Weeks as starter",
                "Number_of_teams": "Number of teams",
                "Player of the week?": "Times as Player of the week?",
                "QB of the week?": "Times as QB of the week?",
                "RB of the week?": "Times as RB of the week?",
                "WR of the week?": "Times as WR of the week?",
                "TE of the week?": "Times as TE of the week?",
                "Benchwarmer of the week?": "Times as Benchwarmer of the week?",
                "Bench QB of the week?": "Times as Bench QB of the week?",
                "Bench RB of the week?": "Times as Bench RB of the week?",
                "Bench WR of the week?": "Times as Bench WR of the week?",
                "Bench TE of the week?": "Times as Bench TE of the week?",
                "Highest starter on team?": "Times as Highest starter on team?",
                "Lowest starter on team?": "Times as Lowest starter on team?",
                "Captain?": "Times as Captain?",
            }
        )

        pa["Number of transactions"] = [
            int(player_tx_all.get(str(player_id), 0))
            for player_id in pa["Player ID"].tolist()
        ]
        pa["Number of drops"] = [
            int(player_drop_all.get(str(player_id), 0))
            for player_id in pa["Player ID"].tolist()
        ]
        # Career trade count, keyed by Player name (matches trades.csv assets).
        pa["Number of trades"] = [
            int(player_trade_all.get(str(name), 0)) for name in pa["Player"].tolist()
        ]

        player_all = pa

        # Mirror the player_year skeleton-row fix on player_all_time:
        # any player who had a transaction or trade but never appeared
        # on a roster is missing from pa entirely. Add them with
        # career counters from player_tx_all / player_drop_all /
        # player_trade_all. Other columns stay NaN (no roster data
        # to derive from).
        try:
            existing_pa: Set[str] = set()
            if not player_all.empty and "Player ID" in player_all.columns:
                existing_pa = {str(p) for p in player_all["Player ID"].astype(str).tolist()}

            tx_sids: Set[str] = set()
            for r in transactions_rows:
                for fld in ("_added_pid", "_dropped_pid"):
                    sid = r.get(fld)
                    if sid:
                        tx_sids.add(str(sid))
            for r in trades_rows:
                for sid in (r.get("_recv_player_ids") or []):
                    if sid:
                        tx_sids.add(str(sid))
                for sid in (r.get("_drop_player_ids") or []):
                    if sid:
                        tx_sids.add(str(sid))

            missing_pa = [s for s in tx_sids if s not in existing_pa]
            # Phase 3A.2: drop pads for players truly never rostered.
            missing_pa = [s for s in missing_pa if tenure_teams_all.get(str(s))]
            if missing_pa:
                pad_rows = []
                for sid in missing_pa:
                    meta = pid_meta.get(str(sid)) or {}
                    name = meta.get("full_name") or str(sid)
                    tenure_team_secs = tenure_time_team_all.get(str(sid), {})
                    top_team_pad = (
                        max(tenure_team_secs.items(), key=lambda kv: kv[1])[0]
                        if tenure_team_secs else None
                    )
                    last_event = tenure_last_event_all.get(str(sid))
                    last_team_pad = last_event[1] if last_event else top_team_pad
                    pad_rows.append({
                        "Player": name,
                        "Player ID": str(sid),
                        "Top team": top_team_pad,
                        "Last team": last_team_pad,
                        "Number of teams": len(tenure_team_secs),
                        "Number of transactions": int(player_tx_all.get(str(sid), 0)),
                        "Number of drops": int(player_drop_all.get(str(sid), 0)),
                        "Number of trades": int(player_trade_all.get(str(name), 0)),
                    })
                if pad_rows:
                    player_all = pd.concat([player_all, pd.DataFrame(pad_rows)], ignore_index=True)
                    _log(debug, f"[{_now_iso()}] INFO seeded {len(pad_rows)} player_all_time rows for tx-only players")
        except Exception as e:
            _log_exc(debug, "player_all_tx_only_pad", e)

        # ----- PR C: player consistency + PAR (player_year + player_all_time) -----
        # Over STARTED weeks only. Volatility = std of started-week points; floor /
        # ceiling = min / max started-week points ever; Boom % / Bust % = share of
        # starts >= 20 / <= 5. PAR = points above positional replacement, where the
        # replacement for (year, week, position) = mean of the BOTTOM THIRD of that
        # week's started scores at the position (the "last startable" tier); PAR is
        # the season/all-time total, PAR per game its mean. All N/A for players who
        # never started (volatility also N/A with < 2 starts).
        try:
            _newc = ["Starter scoring volatility", "Starter scoring floor", "Starter scoring ceiling",
                     "Starter boom %", "Starter bust %", "Starter PAR", "Starter PAR per game"]
            _st = pw_work[pw_work.get("Starter?") == 1].copy()
            _st["Points"] = pd.to_numeric(_st["Points"], errors="coerce")
            _st = _st.dropna(subset=["Points"])
            if {"Player ID", "Year", "Week", "Position"}.issubset(_st.columns) and not _st.empty:
                def _repl(_s):
                    _s = _s.sort_values()
                    _k = max(1, int(np.ceil(len(_s) / 3.0)))
                    return float(_s.iloc[:_k].mean())
                _rep = (_st.groupby(["Year", "Week", "Position"])["Points"]
                        .apply(_repl).rename("_repl").reset_index())
                _stp = _st.merge(_rep, on=["Year", "Week", "Position"], how="left")
                _stp["_par"] = _stp["Points"] - _stp["_repl"]

                def _agg(_keys, _src_par):
                    _g = _st.groupby(_keys)["Points"]
                    _cnt = _g.count()
                    _d = pd.DataFrame({
                        "Starter scoring volatility": _g.std().round(2),
                        "Starter scoring floor": _g.min().round(2),
                        "Starter scoring ceiling": _g.max().round(2),
                        "Starter boom %": (_st[_st["Points"] >= 20].groupby(_keys)["Points"].count().reindex(_cnt.index).fillna(0) / _cnt * 100).round(1),
                        "Starter bust %": (_st[_st["Points"] <= 5].groupby(_keys)["Points"].count().reindex(_cnt.index).fillna(0) / _cnt * 100).round(1),
                        "Starter PAR": _src_par.groupby(_keys)["_par"].sum().round(2),
                        "Starter PAR per game": _src_par.groupby(_keys)["_par"].mean().round(2),
                    })
                    return _d
                _dy = _agg(["Player ID", "Year"], _stp)
                _da = _agg(["Player ID"], _stp)
                _yd = {(str(k[0]), int(k[1])): r for k, r in _dy.iterrows()}
                _ad = {str(k): r for k, r in _da.iterrows()}
                for _c in _newc:
                    if not player_year.empty and {"Player ID", "Year"}.issubset(player_year.columns):
                        _yy = pd.to_numeric(player_year["Year"], errors="coerce")
                        player_year[_c] = [
                            (_yd[(str(p), int(y))][_c] if (pd.notna(y) and (str(p), int(y)) in _yd) else None)
                            for p, y in zip(player_year["Player ID"], _yy)
                        ]
                    if not player_all.empty and "Player ID" in player_all.columns:
                        player_all[_c] = [(_ad[str(p)][_c] if str(p) in _ad else None) for p in player_all["Player ID"]]

                # i3 (#10): position-adjusted, league-wide consistency percentiles.
                # Higher percentile = better WITHIN the player's position:
                # Consistency = LOW volatility; Floor / Ceiling = HIGH. player_year
                # ranks within (Year, position); player_all_time within position.
                # NaN source (never started) -> NaN percentile (renders N/A).
                _pct_specs = [
                    ("Starter scoring volatility", "Consistency percentile", False),  # low vol -> high
                    ("Starter scoring floor", "Floor percentile", True),
                    ("Starter scoring ceiling", "Ceiling percentile", True),
                ]
                for _frame, _grp in ((player_year, ["Year", "_pos"]), (player_all, ["_pos"])):
                    if _frame.empty or "Player ID" not in _frame.columns:
                        continue
                    _frame["_pos"] = [pid_pos.get(str(p)) for p in _frame["Player ID"]]
                    for _src, _out, _asc in _pct_specs:
                        if _src in _frame.columns:
                            _v = pd.to_numeric(_frame[_src], errors="coerce")
                            _frame[_out] = (
                                _frame.assign(_v=_v)
                                .groupby(_grp)["_v"]
                                .rank(pct=True, ascending=_asc) * 100
                            ).round(1)
                    _frame.drop(columns=["_pos"], inplace=True)

                # While-ROSTERED consistency: the same metrics over every week the
                # player actually PLAYED while rostered — started OR benched —
                # excluding bye / injury / suspension weeks (no NFL game = a 0 that
                # isn't a real scoring week and would tank the floor / inflate
                # volatility). Captures the player's true scoring range, including
                # the low weeks a manager benched them for, vs the "Starter scoring"
                # set which only sees weeks they were started.
                def _flag(_df, _c):
                    if _c in _df.columns:
                        return _df[_c].map(lambda v: safe_bool(v, default=False))
                    return pd.Series(False, index=_df.index)
                _ro = pw_work.copy()
                _ro["Points"] = pd.to_numeric(_ro["Points"], errors="coerce")
                _played = ~(_flag(_ro, "Bye?") | _flag(_ro, "Injury?") | _flag(_ro, "Suspension?"))
                _ro = _ro[_played].dropna(subset=["Points"])
                if {"Player ID", "Year"}.issubset(_ro.columns) and not _ro.empty:
                    def _ragg(_keys):
                        _g = _ro.groupby(_keys)["Points"]
                        _cnt = _g.count()
                        return pd.DataFrame({
                            "Rostered scoring volatility": _g.std().round(2),
                            "Rostered scoring floor": _g.min().round(2),
                            "Rostered scoring ceiling": _g.max().round(2),
                            "Rostered boom %": (_ro[_ro["Points"] >= 20].groupby(_keys)["Points"].count().reindex(_cnt.index).fillna(0) / _cnt * 100).round(1),
                            "Rostered bust %": (_ro[_ro["Points"] <= 5].groupby(_keys)["Points"].count().reindex(_cnt.index).fillna(0) / _cnt * 100).round(1),
                        })
                    _rdy = _ragg(["Player ID", "Year"])
                    _rda = _ragg(["Player ID"])
                    _rydict = {(str(k[0]), int(k[1])): r for k, r in _rdy.iterrows()}
                    _radict = {str(k): r for k, r in _rda.iterrows()}
                    _rcols = ["Rostered scoring volatility", "Rostered scoring floor",
                              "Rostered scoring ceiling", "Rostered boom %", "Rostered bust %"]
                    for _c in _rcols:
                        if not player_year.empty and {"Player ID", "Year"}.issubset(player_year.columns):
                            _yy = pd.to_numeric(player_year["Year"], errors="coerce")
                            player_year[_c] = [
                                (_rydict[(str(p), int(y))][_c] if (pd.notna(y) and (str(p), int(y)) in _rydict) else None)
                                for p, y in zip(player_year["Player ID"], _yy)
                            ]
                        if not player_all.empty and "Player ID" in player_all.columns:
                            player_all[_c] = [(_radict[str(p)][_c] if str(p) in _radict else None) for p in player_all["Player ID"]]
                    # Position-adjusted percentiles, mirroring i3.
                    _rpct_specs = [
                        ("Rostered scoring volatility", "Rostered consistency percentile", False),
                        ("Rostered scoring floor", "Rostered floor percentile", True),
                        ("Rostered scoring ceiling", "Rostered ceiling percentile", True),
                    ]
                    for _frame, _grp in ((player_year, ["Year", "_pos"]), (player_all, ["_pos"])):
                        if _frame.empty or "Player ID" not in _frame.columns:
                            continue
                        _frame["_pos"] = [pid_pos.get(str(p)) for p in _frame["Player ID"]]
                        for _src, _out, _asc in _rpct_specs:
                            if _src in _frame.columns:
                                _v = pd.to_numeric(_frame[_src], errors="coerce")
                                _frame[_out] = (
                                    _frame.assign(_v=_v).groupby(_grp)["_v"]
                                    .rank(pct=True, ascending=_asc) * 100
                                ).round(1)
                        _frame.drop(columns=["_pos"], inplace=True)
        except Exception as e:
            _log_exc(debug, "player_consistency_par", e)

    # 3-year roster retention rate (improvement #16): of a team's WEEK-1 roster in
    # year Y, the fraction still on that team's WEEK-1 roster in year Y+3. Rate
    # (not count) because roster sizes have grown over the years. N/A when the
    # Y+3 week-1 roster doesn't exist yet (so currently only 2021->2024 and
    # 2022->2025 are measurable). Keyed by (team, Y) for team_year; team_all_time
    # averages a team's measurable rates.
    retention_3yr_by_ty: Dict[Tuple[str, int], float] = {}
    try:
        if not pw.empty and {"Player ID", "Year", "Week", "Team"}.issubset(pw.columns):
            _w1 = pw[pd.to_numeric(pw["Week"], errors="coerce") == 1]
            _wk1_roster: Dict[Tuple[str, int], set] = {}
            for _tm, _yr, _pid in _w1[["Team", "Year", "Player ID"]].itertuples(index=False, name=None):
                try:
                    _k = (str(_tm), int(_yr))
                except Exception:
                    continue
                if _pid is None or str(_pid) == "nan":
                    continue
                _wk1_roster.setdefault(_k, set()).add(str(_pid))
            for (_tm, _yr), _ros in _wk1_roster.items():
                _future = _wk1_roster.get((_tm, _yr + 3))
                if _future and _ros:
                    retention_3yr_by_ty[(_tm, _yr)] = round(len(_ros & _future) / len(_ros), 4)
    except Exception as e:
        _log_exc(debug, "retention_3yr", e)

    # Team-year: compute record and vs records using raw opp_rid_map (still available in closures above? not anymore)
    team_year = pd.DataFrame()
    team_all = pd.DataFrame()
    if not tw.empty:
        # reconstruct team list
        teams = sorted(tw["Team"].dropna().astype(str).unique().tolist())
        # compute per game outcomes using raw opponent team when available.
        game_rows = []
        for (yr, wk), g in tw.groupby(["Year", "Week"]):
            g2 = g.copy()
            g2["PF"] = pd.to_numeric(g2["PF"], errors="coerce").fillna(0.0)
            g2["Points against"] = pd.to_numeric(g2["Points against"], errors="coerce").fillna(0.0)
            for idx, row in g2.iterrows():
                opp = row.get("Opponent Team (raw)")
                if not opp or pd.isna(opp):
                    match = g2[g2["PF"] == row["Points against"]]
                    if len(match) == 1:
                        opp = str(match.iloc[0]["Team"])
                    elif len(match) > 1:
                        match2 = match[match["Points against"] == row["PF"]]
                        if len(match2) == 1:
                            opp = str(match2.iloc[0]["Team"])
                if opp:
                    game_rows.append({
                        "Year": int(yr),
                        "Week": int(wk),
                        "Team": str(row["Team"]),
                        "OppTeam": str(opp),
                        "Win?": row.get("Win?"),
                        "Week Name": row.get("Week Name"),
                        "PF": float(row["PF"]),
                        "PA": float(row["Points against"]),
                    })
        games_df = pd.DataFrame(game_rows).drop_duplicates(subset=["Year","Week","Team"])

        def _record_str(w, l, t=0):
            return f"{int(w)}-{int(l)}" + (f"-{int(t)}" if t else "")

        def _place_map(rows: List[Tuple[Any, ...]]) -> Dict[str, int]:
            if not rows:
                return {}
            return {str(team): idx + 1 for idx, (team, *_rest) in enumerate(rows)}

        def _normalize_games(df: pd.DataFrame) -> pd.DataFrame:
            if df.empty:
                return df.copy()
            gdf = df.copy()
            gdf["Team"] = gdf["Team"].astype(str)
            gdf["OppTeam"] = gdf["OppTeam"].astype(str)
            gdf["Win?"] = pd.to_numeric(gdf["Win?"], errors="coerce")
            gdf["Year"] = pd.to_numeric(gdf["Year"], errors="coerce").fillna(0).astype(int)
            return gdf

        def _wlt_for_team(df: pd.DataFrame, team: str, year: Optional[int] = None, opps: Optional[set] = None) -> Tuple[int, int, int]:
            if df.empty:
                return (0, 0, 0)
            sub = df[df["Team"] == str(team)]
            if year is not None:
                sub = sub[sub["Year"] == int(year)]
            if opps is not None:
                sub = sub[sub["OppTeam"].isin({str(o) for o in opps})]
            w = int((sub["Win?"] == 1).sum())
            l = int((sub["Win?"] == 0).sum())
            t = int((sub["Win?"] == 0.5).sum())
            return (w, l, t)

        def _win_pct(wlt: Tuple[int, int, int]) -> float:
            w, l, t = wlt
            gp = w + l + t
            return round((w + 0.5 * t) / gp, 4) if gp else 0.0

        # Bracket-specific W-L records. The CONSOLATION games are split out:
        #   regular  = the 'Week N' games
        #   playoff  = winners' bracket CHAMPIONSHIP path (Semifinal + Final);
        #              the 3rd-place game is NOT counted here.
        #   toilet   = losers' bracket (Toilet Semis + Toilet Final); the
        #              toilet-losers game (Toilet Trash) is NOT counted here.
        #   third_place / toilet_losers = the two consolation games on their own.
        _BRACKET = {
            "regular": lambda wn: wn.startswith("Week"),
            "playoff": lambda wn: wn in ("Semifinal", "Final"),
            "toilet": lambda wn: wn in ("Toilet Semis", "Toilet Final"),
            "third_place": lambda wn: wn == "3rd Place",
            "toilet_losers": lambda wn: wn == "Toilet Trash",
        }

        def _wlt_bracket(df: pd.DataFrame, team: str, kind: str, year: Optional[int] = None) -> Tuple[int, int, int]:
            if df.empty or "Win?" not in df.columns or "Week Name" not in df.columns:
                return (0, 0, 0)
            d = df.loc[:, ~df.columns.duplicated()]  # robust to a duplicated col
            sub = d[d["Team"].astype(str) == str(team)]
            if year is not None:
                sub = sub[sub["Year"] == int(year)]
            if sub.empty:
                return (0, 0, 0)
            keep = sub["Week Name"].astype(str).map(_BRACKET[kind])
            win = pd.to_numeric(sub["Win?"], errors="coerce")
            w = int((keep & (win == 1)).sum())
            l = int((keep & (win == 0)).sum())
            t = int((keep & (win == 0.5)).sum())
            return (w, l, t)

        def _win_pct_or_na(wlt: Tuple[int, int, int]):
            return _win_pct(wlt) if sum(wlt) else None

        def _wlt_for_team_pairs(df: pd.DataFrame, team: str, year_team_pairs: set) -> Tuple[int, int, int]:
            if df.empty or not year_team_pairs:
                return (0, 0, 0)
            sub = df[df["Team"] == str(team)]
            sub = sub[sub["YearOpp"].isin(year_team_pairs)]
            w = int((sub["Win?"] == 1).sum())
            l = int((sub["Win?"] == 0).sum())
            t = int((sub["Win?"] == 0.5).sum())
            return (w, l, t)

        def _playoff_elimination_weeks(
            df: pd.DataFrame,
            teams_by_year: Dict[int, set],
            playoff_starts: Dict[int, Optional[int]],
            playoff_teams: Dict[int, set],
        ) -> Dict[int, Dict[str, Optional[int]]]:
            elim_by_year: Dict[int, Dict[str, Optional[int]]] = {}
            if df.empty:
                return elim_by_year
            gdf = df.copy()
            gdf["Week"] = pd.to_numeric(gdf["Week"], errors="coerce")
            gdf["Win?"] = pd.to_numeric(gdf["Win?"], errors="coerce")
            for season, teams in teams_by_year.items():
                season_games = gdf[gdf["Year"] == int(season)].copy()
                playoff_start = playoff_starts.get(int(season))
                if playoff_start:
                    season_games = season_games[season_games["Week"] < playoff_start]
                if season_games.empty:
                    continue
                season_games = season_games.dropna(subset=["Week"])
                weeks = sorted({int(w) for w in season_games["Week"].tolist()})
                if not weeks:
                    continue

                season_games["wins"] = (season_games["Win?"] == 1).astype(int)
                season_games["losses"] = (season_games["Win?"] == 0).astype(int)
                season_games["ties"] = (season_games["Win?"] == 0.5).astype(int)
                week_results = season_games.groupby(["Team", "Week"], as_index=False)[["wins", "losses", "ties"]].sum()

                all_rows = pd.MultiIndex.from_product(
                    [sorted({str(t) for t in teams}), weeks], names=["Team", "Week"]
                ).to_frame(index=False)
                week_results = all_rows.merge(week_results, on=["Team", "Week"], how="left").fillna(0)
                week_results[["wins", "losses", "ties"]] = week_results[["wins", "losses", "ties"]].astype(int)

                week_results["cum_wins"] = week_results.groupby("Team")["wins"].cumsum()
                week_results["cum_losses"] = week_results.groupby("Team")["losses"].cumsum()
                week_results["cum_ties"] = week_results.groupby("Team")["ties"].cumsum()
                total_games = season_games.groupby("Team")["Week"].count().to_dict()
                week_results["total_games"] = week_results["Team"].map(total_games).fillna(0).astype(int)
                week_results["games_played"] = (
                    week_results["cum_wins"] + week_results["cum_losses"] + week_results["cum_ties"]
                )
                week_results["remaining"] = (week_results["total_games"] - week_results["games_played"]).clip(lower=0)
                week_results["max_win_pct"] = (
                    week_results["cum_wins"]
                    + week_results["remaining"]
                    + 0.5 * week_results["cum_ties"]
                ) / week_results["total_games"].replace(0, np.nan)
                week_results["min_win_pct"] = (
                    week_results["cum_wins"] + 0.5 * week_results["cum_ties"]
                ) / week_results["total_games"].replace(0, np.nan)

                elim_map: Dict[str, Optional[int]] = {}
                season_playoffs = {str(t) for t in playoff_teams.get(int(season), set())}
                for team in sorted({str(t) for t in teams}):
                    if team in season_playoffs:
                        elim_map[team] = None
                        continue
                    elim_week = None
                    for wk in weeks:
                        t_row = week_results[(week_results["Team"] == team) & (week_results["Week"] == wk)]
                        if t_row.empty:
                            continue
                        t_max = float(t_row["max_win_pct"].iloc[0])
                        others = week_results[(week_results["Week"] == wk) & (week_results["Team"] != team)]
                        if int((others["min_win_pct"] > t_max).sum()) >= 4:
                            elim_week = int(wk)
                            break
                    elim_map[team] = elim_week
                elim_by_year[int(season)] = elim_map
            return elim_by_year

        games_df = _normalize_games(games_df)
        if not games_df.empty:
            games_df["YearOpp"] = list(zip(games_df["Year"], games_df["OppTeam"]))

        playoff_teams_by_season: Dict[int, set] = {}
        champion_by_season: Dict[int, Optional[str]] = {}
        last_place_by_season: Dict[int, Optional[str]] = {}
        teams_by_season: Dict[int, set] = {}
        place_record_by_year: Dict[int, Dict[str, int]] = {}
        place_pf_by_year: Dict[int, Dict[str, int]] = {}
        place_maxpf_by_year: Dict[int, Dict[str, int]] = {}
        standings_place_by_season: Dict[int, Dict[str, int]] = {}
        pf_place_by_season: Dict[int, Dict[str, int]] = {}
        maxpf_place_by_season: Dict[int, Dict[str, int]] = {}
        for yr, g in tw.groupby("Year"):
            season = int(yr)
            teams_by_season[season] = set(g["Team"].dropna().astype(str).tolist())
            playoff_start = playoff_start_by_season.get(season)
            reg = g.copy()
            if playoff_start:
                reg = reg[pd.to_numeric(reg["Week"], errors="coerce") < playoff_start]
            reg["PF"] = pd.to_numeric(reg["PF"], errors="coerce").fillna(0.0)
            reg["Max PF"] = pd.to_numeric(reg["Max PF"], errors="coerce").fillna(0.0)
            reg["Win?"] = pd.to_numeric(reg["Win?"], errors="coerce")
            standings = []
            for team, tg in reg.groupby("Team"):
                wins = int((tg["Win?"] == 1).sum())
                losses = int((tg["Win?"] == 0).sum())
                ties = int((tg["Win?"] == 0.5).sum())
                pf = float(tg["PF"].sum())
                maxpf = float(tg["Max PF"].sum())
                standings.append((team, wins, losses, ties, pf, maxpf))
            standings.sort(key=lambda x: (x[1] + 0.5 * x[3], x[4]), reverse=True)
            standings_place_by_season[season] = {
                str(team): idx + 1 for idx, (team, *_rest) in enumerate(standings)
            }
            pf_sorted = sorted(standings, key=lambda x: x[4], reverse=True)
            pf_place_by_season[season] = {
                str(team): idx + 1 for idx, (team, *_rest) in enumerate(pf_sorted)
            }
            reg["Max PF"] = pd.to_numeric(reg.get("Max PF"), errors="coerce").fillna(0.0)
            maxpf_totals = []
            for team, tg in reg.groupby("Team"):
                maxpf_totals.append((team, float(tg["Max PF"].sum())))
            maxpf_totals.sort(key=lambda x: x[1], reverse=True)
            maxpf_place_by_season[season] = {
                str(team): idx + 1 for idx, (team, _maxpf) in enumerate(maxpf_totals)
            }
            # PR E fix C: playoff seeding / champion / last place / finish are
            # only meaningful once the season is OVER. Mid-season the standings
            # are provisional and `champ` would fall back to the current leader
            # (mislabeling the leader "Champion"). Gate these to completed
            # seasons; a live season is left out of the dicts so every
            # "vs playoff / vs champion / vs last place / Result" downstream
            # renders N/A for it. Standings/place maps below stay (they are
            # legitimately "current standings").
            _season_done = _season_is_complete(season)
            place_record_by_year[season] = _place_map(standings)
            standings_pf = sorted(standings, key=lambda x: x[4], reverse=True)
            standings_maxpf = sorted(standings, key=lambda x: x[5], reverse=True)
            place_pf_by_year[season] = _place_map(standings_pf)
            place_maxpf_by_year[season] = _place_map(standings_maxpf)
            if _season_done:
                playoff_teams_by_season[season] = set([t for t, *_ in standings[:4]])
                last_place_by_season[season] = standings[-1][0] if standings else None
                champ = None
                if "Week label" in g.columns:
                    finals = g[g["Week label"] == "Final"]
                    champ_row = finals[finals["Win?"] == 1]
                    if not champ_row.empty:
                        champ = str(champ_row.iloc[0]["Team"])
                if not champ and standings:
                    champ = standings[0][0]
                champion_by_season[season] = champ

        playoff_elimination_by_season: Dict[int, Dict[str, Optional[int]]] = {}
        try:
            playoff_elimination_by_season = _playoff_elimination_weeks(
                games_df, teams_by_season, playoff_start_by_season, playoff_teams_by_season
            )
        except Exception as e:
            _log_exc(debug, "playoff_elimination_calc", e)

        
        # Determine season finishing positions (Result) from playoff/toilet brackets when available.
        season_finish: Dict[int, Dict[str, str]] = {}
        try:
            for yr, g in tw.groupby("Year"):
                season = int(yr)
                # PR E fix C: a finish/Result is only real once the season is
                # over. Skip a live season so its Result renders N/A.
                if not _season_is_complete(season):
                    continue
                playoff_start = playoff_start_by_season.get(season)
                if not playoff_start:
                    continue
                finals_week = playoff_start + 1
                fin_map: Dict[str, str] = {}
                # Finals
                gf = tw[(tw["Year"]==season) & (tw["Week"]==finals_week) & (tw["Week label"]=="Final")].copy()
                gf["PF"] = pd.to_numeric(gf["PF"], errors="coerce").fillna(0.0)
                gf["Win?"] = pd.to_numeric(gf["Win?"], errors="coerce")
                if len(gf) == 2:
                    if gf["Win?"].notna().any():
                        gf = gf.sort_values("Win?", ascending=False)
                    else:
                        gf = gf.sort_values("PF", ascending=False)
                    fin_map[str(gf.iloc[0]["Team"])] = "Champion"
                    fin_map[str(gf.iloc[1]["Team"])] = "2nd"
                # 3rd place
                g3 = tw[(tw["Year"]==season) & (tw["Week"]==finals_week) & (tw["Week label"]=="3rd Place")].copy()
                g3["PF"] = pd.to_numeric(g3["PF"], errors="coerce").fillna(0.0)
                g3["Win?"] = pd.to_numeric(g3["Win?"], errors="coerce")
                if len(g3) == 2:
                    if g3["Win?"].notna().any():
                        g3 = g3.sort_values("Win?", ascending=False)
                    else:
                        g3 = g3.sort_values("PF", ascending=False)
                    fin_map[str(g3.iloc[0]["Team"])] = "3rd"
                    fin_map[str(g3.iloc[1]["Team"])] = "4th"

                # Non-playoff finishes (5th-8th) are based on regular-season record cutoff,
                # with PF as the tiebreaker. (Pre-2025: through 17 games; 2025+: through 15 games.)
                cutoff = 17 if season < 2025 else 15
                try:
                    all_teams = [str(t) for t in tw[tw["Year"] == season]["Team"].dropna().unique().tolist()]
                    playoff_teams = set([t for t, r in fin_map.items() if r in ("Champion", "2nd", "3rd", "4th")])
                    non_playoff = [t for t in all_teams if t not in playoff_teams]
                    if non_playoff and (not games_df.empty):
                        reg = games_df[(games_df["Year"] == season) & (games_df["Week"] <= cutoff)].copy()
                        reg["PF"] = pd.to_numeric(reg.get("PF", 0.0), errors="coerce").fillna(0.0)
                        reg["Win?"] = pd.to_numeric(reg.get("Win?", 0.0), errors="coerce").fillna(0.0)
                        sub = reg[reg["Team"].astype(str).isin(non_playoff)]
                        rows_np = []
                        for team_np, gg in sub.groupby(sub["Team"].astype(str)):
                            w = int((gg["Win?"] == 1).sum())
                            l = int((gg["Win?"] == 0).sum())
                            t_ = int((gg["Win?"] == 0.5).sum())
                            pf_sum = float(gg["PF"].sum())
                            win_pct = (w + 0.5 * t_) / max(1, w + l + t_)
                            rows_np.append((team_np, win_pct, pf_sum))
                        # Sort: record (win %) then PF
                        rows_np.sort(key=lambda x: (x[1], x[2]), reverse=True)
                        place = 5
                        for team_np, *_ in rows_np:
                            if place == 5:
                                fin_map[team_np] = "5th"
                            elif place == 6:
                                fin_map[team_np] = "6th"
                            elif place == 7:
                                fin_map[team_np] = "7th"
                            elif place == 8:
                                fin_map[team_np] = "8th"
                            place += 1
                            if place > 8:
                                break
                except Exception:
                    pass
                season_finish[season] = fin_map
        except Exception as e:
            _log_exc(debug, "season_finish_map", e)

        # team-year rollup
        all_time_rows = []
        for team, g in tw.groupby("Team"):
            wins = int((g["Win?"] == 1).sum())
            losses = int((g["Win?"] == 0).sum())
            ties = int((g["Win?"] == 0.5).sum())
            pf = float(pd.to_numeric(g["PF"], errors="coerce").fillna(0.0).sum())
            maxpf_sum = float(pd.to_numeric(g["Max PF"], errors="coerce").fillna(0.0).sum())
            all_time_rows.append((team, wins, losses, ties, pf, maxpf_sum))
        all_time_rows.sort(key=lambda x: (x[1] + 0.5 * x[3], x[4]), reverse=True)
        all_time_place_record = _place_map(all_time_rows)
        all_time_place_pf = _place_map(sorted(all_time_rows, key=lambda x: x[4], reverse=True))
        all_time_place_maxpf = _place_map(sorted(all_time_rows, key=lambda x: x[5], reverse=True))

        rows = []
        for (team, yr), g in tw.groupby(["Team", "Year"]):
            wins = int((g["Win?"] == 1).sum())
            losses = int((g["Win?"] == 0).sum())
            ties = int((g["Win?"] == 0.5).sum())
            gp = max(1, wins + losses + ties)
            pf = float(pd.to_numeric(g["PF"], errors="coerce").fillna(0.0).sum())
            pa = float(pd.to_numeric(g["Points against"], errors="coerce").fillna(0.0).sum())
            diff = pf - pa
            maxpf_sum = float(pd.to_numeric(g["Max PF"], errors="coerce").fillna(0.0).sum())
            maxpf_avg = float(pd.to_numeric(g["Max PF"], errors="coerce").fillna(0.0).mean())
            rec = _record_str(wins, losses, ties)
            winp = round((wins + 0.5 * ties) / gp, 4)
            place_map = standings_place_by_season.get(int(yr), {})
            pf_place_map = pf_place_by_season.get(int(yr), {})
            maxpf_place_map = maxpf_place_by_season.get(int(yr), {})
            place = place_map.get(str(team))
            pf_place = pf_place_map.get(str(team))
            maxpf_place = maxpf_place_map.get(str(team))
            win_variance = None
            if place is not None and pf_place is not None and maxpf_place is not None:
                win_variance = -1 * float(place - ((pf_place + maxpf_place) / 2))
            row = {
                "Team": str(team),
                "Year": int(yr),
                "Result": "N/A",
                "Win %": winp,
                "Record": rec,
                "Record & win % vs each team": "N/A",
                "Record & win % vs playoff teams": "N/A",
                "Record & win % vs non-playoff teams": "N/A",
                "Record & win % vs champion": "N/A",
                "Record & win % vs last place": "N/A",
                "Change in win % from previous season": None,
                "Win Variance": win_variance,
                "Week of playoff elimination": playoff_elimination_by_season.get(int(yr), {}).get(str(team)),
                "Draft Value": 0,
                "Number of first round picks made": 0,
                "Total number of picks made": 0,
                "Points": round(pf, 2),
                "Avg points": round(pf / gp, 2),
                "Points against": round(pa, 2),
                "Avg points against": round(pa / gp, 2),
                "Differential": round(diff, 2),
                "Avg differential": round(diff / gp, 2),
                "Max PF": round(maxpf_sum, 2),
                "Avg max PF": round(maxpf_avg, 2) if not math.isnan(maxpf_avg) else None,
                "Efficiency": round(pf / maxpf_sum, 4) if maxpf_sum else None,
                "Weeks of injuries": int(pd.to_numeric(g.get("Number of Injuries"), errors="coerce").fillna(0.0).sum()),
                "Weeks suspensions": int(pd.to_numeric(g.get("Number of suspensions"), errors="coerce").fillna(0.0).sum()),
                "Hardship": float(pd.to_numeric(g.get("Hardship"), errors="coerce").fillna(0.0).sum()),
                "Starter-adjusted Hardship": round(float(pd.to_numeric(g.get("Starter-adjusted Hardship"), errors="coerce").fillna(0.0).sum()), 4),
                "Offseason starter turnover": 0,
                "Inseason starter turnover": 0,
                # Roll up per-week activity (already computed in team_week) into
                # the season total. Previously these stayed at N/A because no
                # aggregation step carried them across.
                "Number of transactions": int(pd.to_numeric(g.get("Number of transactions"), errors="coerce").fillna(0.0).sum()),
                "Number of trades": int(pd.to_numeric(g.get("Number of trades"), errors="coerce").fillna(0.0).sum()),
                "Amount of FAAB spent": round(float(pd.to_numeric(g.get("Amount of FAAB spent"), errors="coerce").fillna(0.0).sum()), 2),
                # Highest combined matchup score (own PF + opponent PF) the
                # team participated in during the season. team_week already
                # computes Combined matchup score per game.
                "Combined matchup score": float(pd.to_numeric(g.get("Combined matchup score"), errors="coerce").fillna(0.0).max()),
            }
            # NOTE: Unique-player Number of {QB,WR,RB,TE} {started,rostered}
            # columns are added later (line ~7977 "team_unique_player_counts"
            # block) via a merge from pw groupby. Adding them here too
            # creates _x/_y suffix collisions on that merge.
            rows.append(row)
        team_year = pd.DataFrame(rows)

        # In-progress seasons (e.g., 2026 in May 2026 — drafts done,
        # no games yet) have no team_week rows, so the groupby above
        # produces no team_year row either. That leaves transactions
        # and trades for the in-progress season homeless when the
        # detail tables look for their per-season rollup, and breaks
        # the "team_year ↔ detail" reconciliation we've been keeping
        # clean. Emit minimal placeholder rows so the in-progress
        # season is represented.
        try:
            seasons_in_tw = set()
            if not tw.empty and "Year" in tw.columns:
                seasons_in_tw = {int(y) for y in tw["Year"].dropna().unique()}
            # All seasons present in the loaded league chain
            chain_seasons = {int(lg.get("season")) for lg in (leagues or []) if lg.get("season")}
            in_progress = chain_seasons - seasons_in_tw
            if in_progress:
                # Canonical team list from the most recent completed
                # season's rosters (or from any team_year row we have).
                canonical_teams: List[str] = []
                if not team_year.empty and "Team" in team_year.columns:
                    canonical_teams = sorted(team_year["Team"].dropna().astype(str).unique().tolist())
                # Fallback to roster_to_team if team_year was empty
                if not canonical_teams:
                    canonical_teams = sorted({str(v) for v in roster_to_team.values() if v})

                placeholder_rows = []
                for season_i in sorted(in_progress):
                    for team_name in canonical_teams:
                        # Play-derived stats stay None (render as N/A)
                        # for in-progress seasons — they're undefined
                        # rather than 'team played and scored 0'. Count
                        # / FAAB columns get real 0 because real-world
                        # offseason transactions actually count.
                        placeholder_rows.append({
                            "Team": str(team_name),
                            "Year": int(season_i),
                            "Result": "N/A",
                            "Win %": None,
                            "Record": "0-0-0",
                            "Record & win % vs each team": "N/A",
                            "Record & win % vs playoff teams": "N/A",
                            "Record & win % vs non-playoff teams": "N/A",
                            "Record & win % vs champion": "N/A",
                            "Record & win % vs last place": "N/A",
                            "Change in win % from previous season": None,
                            "Win Variance": None,
                            "Week of playoff elimination": None,
                            "Draft Value": 0,
                            "Number of first round picks made": 0,
                            "Total number of picks made": 0,
                            "Points": None,
                            "Avg points": None,
                            "Points against": None,
                            "Avg points against": None,
                            "Differential": None,
                            "Avg differential": None,
                            "Max PF": None,
                            "Avg max PF": None,
                            "Efficiency": None,
                            "Weeks of injuries": 0,
                            "Weeks suspensions": 0,
                            "Hardship": None,
                            "Offseason starter turnover": 0,
                            "Inseason starter turnover": 0,
                            # Number of transactions / trades / FAAB
                            # spent will be backfilled from the detail
                            # tables by the post-build reconciliation
                            # step we already run for completed seasons.
                            "Number of transactions": 0,
                            "Number of trades": 0,
                            "Amount of FAAB spent": 0.0,
                            "Combined matchup score": None,
                        })
                if placeholder_rows:
                    team_year = pd.concat([team_year, pd.DataFrame(placeholder_rows)], ignore_index=True)
                    _log(debug, f"[{_now_iso()}] INFO seeded {len(placeholder_rows)} team_year placeholder rows for in-progress season(s) {sorted(in_progress)}")

                    # Backfill transaction / trade / FAAB counts for
                    # the in-progress rows from the detail tables.
                    # Without this they'd stay at 0 even though we have
                    # real offseason activity (drafts, FA pickups,
                    # trades) for the in-progress season.
                    tx_count_ip: Dict[Tuple[str, int], int] = defaultdict(int)
                    tx_faab_ip: Dict[Tuple[str, int], float] = defaultdict(float)
                    for r in transactions_rows:
                        try:
                            t = str(r.get("Team") or "")
                            s = int(r.get("Season")) if r.get("Season") is not None else None
                        except Exception:
                            continue
                        if not t or s is None or s not in in_progress:
                            continue
                        tx_count_ip[(t, s)] += 1
                        try:
                            f = float(r.get("Faab") or 0.0)
                        except Exception:
                            f = 0.0
                        tx_faab_ip[(t, s)] += f
                    tr_count_ip: Dict[Tuple[str, int], int] = defaultdict(int)
                    for r in trades_rows:
                        try:
                            t = str(r.get("Team") or "")
                            s = int(r.get("Season")) if r.get("Season") is not None else None
                        except Exception:
                            continue
                        if not t or s is None or s not in in_progress:
                            continue
                        tr_count_ip[(t, s)] += 1
                    # Trades are counted by both Number of trades and
                    # Number of transactions in our schema.
                    for idx, row in team_year.iterrows():
                        try:
                            s = int(row.get("Year"))
                            t = str(row.get("Team"))
                        except Exception:
                            continue
                        if s not in in_progress:
                            continue
                        n_tx = tx_count_ip.get((t, s), 0)
                        n_tr = tr_count_ip.get((t, s), 0)
                        team_year.at[idx, "Number of transactions"] = int(n_tx + n_tr)
                        team_year.at[idx, "Number of trades"] = int(n_tr)
                        team_year.at[idx, "Amount of FAAB spent"] = round(tx_faab_ip.get((t, s), 0.0), 2)
        except Exception as e:
            _log_exc(debug, "team_year_in_progress_seed", e)

        # Pick-history-based rollups (Draft Value, # picks made by round).
        # Uses the final-owner team (who actually drafted) for each pick row,
        # which after the trade-chain rewrite above lives in pick_rows' Trade
        # columns (last non-empty Trade N) or in 'Original Team' when no trade
        # chain exists.
        try:
            if not ph.empty:
                phx = ph.copy()
                # Exclude the 2021 supplemental veteran draft from TEAM draft
                # stats — Draft Value / # first round picks made / total picks
                # made should count rookie-draft selections only. The vet picks
                # remain in pick_history itself (Year tagged "2021 (vet)"); we
                # just drop them from these rollups.
                phx = phx[
                    ~phx["Year"].astype(str).str.contains("vet", case=False, na=False)
                ].copy()
                # Resolve picker = Final Team (chain end), falling back to the
                # last non-empty Trade column, then Original Team.
                def _picker(row):
                    ft = row.get("Final Team")
                    if ft and str(ft) != "nan" and str(ft).strip() not in ("", "N/A"):
                        return str(ft)
                    for j in range(10, 0, -1):
                        v = row.get(f"Trade {j}")
                        if v and str(v) != "nan" and str(v).strip() not in ("", "N/A"):
                            return str(v)
                    return str(row.get("Original Team") or "")
                phx["_Picker"] = phx.apply(_picker, axis=1)
                # Strip "(vet)" suffix so pick rollups merge against
                # integer-Year team_year rows. The vet draft picks still
                # count toward the picker's 2021 totals.
                phx["_YearKey"] = phx["Year"].astype(str).str.extract(r"^(\d+)").astype(float).astype("Int64")
                # Parse round from Number — accept legacy 'R1.2' and current '1.05'.
                phx["_Round"] = phx["Number"].astype(str).str.extract(r"^R?(\d+)").astype(float)
                # Parse slot for draft value (1/(slot+1)).
                phx["_PickNo"] = phx["Number"].astype(str).str.extract(r"^R?\d+\.(\d+)").astype(float)
                # Synthetic picks count as their equivalent slot for draft
                # value / round tallies: 2.09 -> round 2 slot 8, 5.0X -> round 4
                # slot 8 (so a fifth never inflates round counts as a "round 5").
                _num = phx["Number"].astype(str).str.strip()
                phx.loc[_num == "2.09", "_PickNo"] = 8.0
                _is5 = phx["_Round"] == 5.0
                phx.loc[_is5, "_Round"] = 4.0
                phx.loc[_is5, "_PickNo"] = 8.0
                phx["_DraftVal"] = phx["_PickNo"].apply(lambda x: 1.0 / (x + 1.0) if pd.notna(x) else 0.0)
                pick_agg = phx.groupby(["_Picker", "_YearKey"], dropna=False).agg(
                    _draft_value=("_DraftVal", "sum"),
                    _total_picks=("Number", "count"),
                    _r1=("_Round", lambda s: int((s == 1.0).sum())),
                ).reset_index().rename(columns={"_Picker": "Team", "_YearKey": "Year"})
                # team_year.Year is Int64; align dtype.
                pick_agg["Year"] = pick_agg["Year"].astype("Int64")
                team_year = team_year.merge(pick_agg, on=["Team", "Year"], how="left")
                team_year["Draft Value"] = team_year["_draft_value"].fillna(0.0).round(4)
                team_year["Number of first round picks made"] = team_year["_r1"].fillna(0).astype(int)
                team_year["Total number of picks made"] = team_year["_total_picks"].fillna(0).astype(int)
                team_year.drop(columns=["_draft_value", "_total_picks", "_r1"], inplace=True, errors="ignore")
        except Exception as e:
            _log_exc(debug, "team_year_pick_rollups", e)

        # 3-year roster retention rate (improvement #16) — per (team, year).
        try:
            if not team_year.empty and {"Team", "Year"}.issubset(team_year.columns):
                team_year["3-year roster retention rate"] = [
                    retention_3yr_by_ty.get((str(t), int(y)))
                    if pd.notna(y) and str(y).strip() not in ("", "nan") else None
                    for t, y in zip(team_year["Team"], pd.to_numeric(team_year["Year"], errors="coerce"))
                ]
        except Exception as e:
            _log_exc(debug, "team_year_retention", e)

        # Future draft capital: weighted future picks held by the team at
        # END of each season (own retained + acquired − traded away). The
        # corrected _future_cap_held walks the pick-ownership ledger; the
        # old _future_cap_from_traded only saw picks present in Sleeper's
        # traded_picks snapshot, so every team's own un-traded picks were
        # silently dropped. Feb 1 anchor = post-Week-18, pre-next-draft.
        try:
            future_cap_vals = []
            for idx, row in team_year.iterrows():
                team = str(row["Team"])
                year = int(row["Year"])
                future_cap_vals.append(_future_cap_held(team, int(year), date(int(year) + 1, 2, 1)))
            team_year["Future draft capital"] = future_cap_vals
        except Exception as e:
            _log_exc(debug, "team_year_future_cap", e)

        # Rebuild all win % and record columns for team-year (single source of truth)
        try:
            teams_by_year = team_year.groupby("Year")["Team"].apply(lambda s: sorted(s.dropna().astype(str).unique().tolist())).to_dict()
            for idx, row in team_year.iterrows():
                team = str(row["Team"])
                year = int(row["Year"])
                wlt = _wlt_for_team(games_df, team, year=year)
                team_year.at[idx, "Record"] = _record_str(*wlt)
                # No games played → leave Win % as None so it renders
                # as N/A. _win_pct returns 0.0 for an empty record,
                # which would mistakenly read 'team lost everything'.
                if sum(wlt) == 0:
                    team_year.at[idx, "Win %"] = None
                else:
                    team_year.at[idx, "Win %"] = _win_pct(wlt)
                # Regular-season-only record + win % (the standings record).
                _reg = _wlt_bracket(games_df, team, "regular", year=year)
                team_year.at[idx, "Regular season record"] = _record_str(*_reg)
                team_year.at[idx, "Regular season win %"] = _win_pct_or_na(_reg)

                opp_list = [t for t in teams_by_year.get(year, []) if t != team]
                pieces = []
                for opp in opp_list:
                    wlt_opp = _wlt_for_team(games_df, team, year=year, opps={opp})
                    pieces.append(f"{opp}: {_record_str(*wlt_opp)} ({_win_pct(wlt_opp)})")
                team_year.at[idx, "Record & win % vs each team"] = "; ".join(pieces) if pieces else "N/A"

                playoffs = playoff_teams_by_season.get(year, set())
                champ = champion_by_season.get(year)
                lastp = last_place_by_season.get(year)
                non_playoffs = set(teams_by_season.get(year, set())) - set(playoffs) if playoffs else set()

                wlt_play = _wlt_for_team(games_df, team, year=year, opps=playoffs if playoffs else None)
                wlt_non = _wlt_for_team(games_df, team, year=year, opps=non_playoffs if non_playoffs else None)
                wlt_champ = _wlt_for_team(games_df, team, year=year, opps={champ} if champ else None)
                wlt_last = _wlt_for_team(games_df, team, year=year, opps={lastp} if lastp else None)

                team_year.at[idx, "Record & win % vs playoff teams"] = f"{_record_str(*wlt_play)} ({_win_pct(wlt_play)})" if playoffs else "N/A"
                team_year.at[idx, "Record & win % vs non-playoff teams"] = f"{_record_str(*wlt_non)} ({_win_pct(wlt_non)})" if non_playoffs else "N/A"
                team_year.at[idx, "Record & win % vs champion"] = f"{_record_str(*wlt_champ)} ({_win_pct(wlt_champ)})" if champ else "N/A"
                team_year.at[idx, "Record & win % vs last place"] = f"{_record_str(*wlt_last)} ({_win_pct(wlt_last)})" if lastp else "N/A"

            if not games_df.empty:
                for year, teams_list in teams_by_year.items():
                    for team in teams_list:
                        for opp in teams_list:
                            if opp == team:
                                continue
                            wlt_opp = _wlt_for_team(games_df, team, year=int(year), opps={opp})
                            team_year.loc[
                                (team_year["Year"] == int(year)) & (team_year["Team"] == team),
                                f"Record vs {opp}",
                            ] = _record_str(*wlt_opp)
                            team_year.loc[
                                (team_year["Year"] == int(year)) & (team_year["Team"] == team),
                                f"Win % vs {opp}",
                            ] = _win_pct(wlt_opp)
        except Exception as e:
            _log_exc(debug, "team_year_rebuild_records", e)

        # Compute starter/roster turnover metrics (item 4 refactor).
        # Definitions (all "unique players changed between two endpoints" =
        # symmetric difference of the two rosters/lineups):
        #  - Inseason  = roster/lineup at Week 1 vs at the championship (final
        #                played week) of the SAME season.
        #  - Offseason = roster/lineup at the prior season's championship week
        #                vs this season's Week 1. The Week-1 boundary is the
        #                shared seam between the two metrics.
        #  - Average weekly starter/roster turnover = mean across the season of
        #    team_week "… turnover from previous week".
        try:
            if not pw.empty and "Starter/Bench" in pw.columns:
                pw_t = pw.copy()
                pw_t["Week"] = pd.to_numeric(pw_t["Week"], errors="coerce")

                def _set_for(team, year, week, starters_only):
                    df = pw_t[(pw_t["Team"] == team) & (pw_t["Year"] == year) & (pw_t["Week"] == week)]
                    if starters_only:
                        df = df[df["Starter/Bench"].astype(str).str.lower().eq("starter")]
                    return set(df["Player"].dropna().astype(str).tolist())

                # Per (team, year) mean of the team_week weekly turnover columns.
                twk_s_avg: Dict[Tuple[str, int], float] = {}
                twk_r_avg: Dict[Tuple[str, int], float] = {}
                try:
                    if not tw.empty:
                        _g = tw.groupby(["Team", "Year"])
                        _sa = _g["Starter turnover from previous week"].mean()
                        _ra = _g["Roster turnover from previous week"].mean()
                        for (t_, y_), v_ in _sa.items():
                            twk_s_avg[(str(t_), int(y_))] = float(v_) if pd.notna(v_) else 0.0
                        for (t_, y_), v_ in _ra.items():
                            twk_r_avg[(str(t_), int(y_))] = float(v_) if pd.notna(v_) else 0.0
                except Exception as e:
                    _log_exc(debug, "turnover_weekly_avg", e)

                for (team, year), g in pw_t.groupby(["Team", "Year"]):
                    weeks = sorted([int(w) for w in g["Week"].dropna().unique().tolist()])
                    if not weeks:
                        continue
                    first_w, champ_w = weeks[0], weeks[-1]
                    mask = (team_year["Team"] == team) & (team_year["Year"] == year)

                    # In-season: Wk1 vs championship (final) week, unique changed.
                    s_first = _set_for(team, year, first_w, True)
                    r_first = _set_for(team, year, first_w, False)
                    s_champ = _set_for(team, year, champ_w, True)
                    r_champ = _set_for(team, year, champ_w, False)
                    team_year.loc[mask, "Inseason starter turnover"] = len(s_first.symmetric_difference(s_champ))
                    team_year.loc[mask, "Inseason roster turnover"] = len(r_first.symmetric_difference(r_champ))

                    # Offseason: prior season's championship week vs this Wk1.
                    prev_year = int(year) - 1
                    if ((pw_t["Team"] == team) & (pw_t["Year"] == prev_year)).any():
                        gprev = pw_t[(pw_t["Team"] == team) & (pw_t["Year"] == prev_year)]
                        prev_weeks = sorted([int(w) for w in gprev["Week"].dropna().unique().tolist()])
                        if prev_weeks:
                            prev_champ = prev_weeks[-1]
                            s_prev = _set_for(team, prev_year, prev_champ, True)
                            r_prev = _set_for(team, prev_year, prev_champ, False)
                            team_year.loc[mask, "Offseason starter turnover"] = len(s_prev.symmetric_difference(s_first))
                            team_year.loc[mask, "Offseason roster turnover"] = len(r_prev.symmetric_difference(r_first))

                    # Weekly averages.
                    team_year.loc[mask, "Average weekly starter turnover"] = round(twk_s_avg.get((str(team), int(year)), 0.0), 2)
                    team_year.loc[mask, "Average weekly roster turnover"] = round(twk_r_avg.get((str(team), int(year)), 0.0), 2)
        except Exception as e:
            _log_exc(debug, "turnover_metrics_team_year", e)

        # --------------------------
        # Fill missing Team-year columns from team-week (flags, tanking, luck, roster composition, etc.)
        # --------------------------
        try:
            # Tanking helper: pick the final week's expanding-mean
            # value rather than summing every week. The per-week
            # Tanking score is an expanding mean from season start
            # through that week (see _tanking_score), so the last
            # value already represents the team's season-final tank
            # signal. Summing weeks was wildly inflating the number
            # and made cross-team comparisons meaningless.
            def _tank_last_week(s: pd.Series) -> Optional[float]:
                vals = pd.to_numeric(s, errors="coerce").dropna()
                if vals.empty:
                    # No team_week rows for this team-year (e.g. an
                    # in-progress season placeholder). Return None so
                    # the team_year cell renders N/A rather than 0.0.
                    return None
                return float(vals.iloc[-1])

            # Sort by Week first so 'last' actually means latest week.
            tw_sorted = tw.sort_values(["Team", "Year", "Week"])
            agg_year = tw_sorted.groupby(["Team", "Year"], as_index=False).agg(
                **{
                    "Tanking": ("Tanking", _tank_last_week),
                    "Luck": ("Luck", "sum"),
                    "Times Brosenzweig": ("Brosenzweig", "sum"),
                    "Times Sisenzweig": ("Sisenzweig", "sum"),
                    "Times Highest score?": ("Highest score?", "sum"),
                    "Times Lowest score?": ("Lowest score?", "sum"),
                    "Times Narrowest victory?": ("Narrowest victory?", "sum"),
                    "Times Largest blowout?": ("Largest blowout?", "sum"),
                    "Times Most efficient?": ("Most efficient?", "sum"),
                    "Times Least efficient?": ("Least efficient?", "sum"),
                    "Times Top half of league?": ("Top half of league?", "sum"),
                    "Times One-man army?": ("One-man army?", "sum"),
                    "Times Most bench points?": ("Most bench points?", "sum"),
                    "Times Most injured?": ("Most injured?", "sum"),
                    "Most number of players started from same NFL team": ("Most number of players started from same NFL team", "max"),
                    "Most number of players rostered from same NFL team": ("Most number of players rostered from same NFL team", "max"),
                    "Most number of QBs started from same NFL team": ("Most number of QBs started from same NFL team", "max"),
                    "Most number of QBs rostered from same NFL team": ("Most number of QBs rostered from same NFL team", "max"),
                    "Most number of RBs started from same NFL team": ("Most number of RBs started from same NFL team", "max"),
                    "Most number of RBs rostered from same NFL team": ("Most number of RBs rostered from same NFL team", "max"),
                    "Most number of WR started from same NFL team": ("Most number of WR started from same NFL team", "max"),
                    "Most number of WR rostered from same NFL team": ("Most number of WR rostered from same NFL team", "max"),
                    "Most number of TE started from same NFL team": ("Most number of TE started from same NFL team", "max"),
                    "Most number of TE rostered from same NFL team": ("Most number of TE rostered from same NFL team", "max"),
                    "Number of NFL teams among starting players": ("Number of NFL teams among starting players", "max"),
                    "Number of NFL teams among rostered players": ("Number of NFL teams among rostered players", "max"),
                    "Number of rookies started": ("Number of rookies started", "sum"),
                    "Number of rookies rostered": ("Number of rookies rostered", "sum"),
                    "Player average age": ("Player average age", "mean"),
                    "Team age including picks": ("Team age including picks", "mean"),
                    "Difference between highest and lowest starters": ("Difference between highest and lowest starters", "max"),
                    "Number of donuts": ("Number of donuts", "sum"),
                    "Number of starter donuts": ("Number of starter donuts", "sum"),
                    "Number of players under 10": ("Number of players under 10", "sum"),
                    "Number of starters under 10": ("Number of starters under 10", "sum"),
                    "Number of players over 20": ("Number of players over 20", "sum"),
                    "Number of starters over 20": ("Number of starters over 20", "sum"),
                    "Number of players over 30": ("Number of players over 30", "sum"),
                    "Number of starters over 30": ("Number of starters over 30", "sum"),
                    "Number of players over 40": ("Number of players over 40", "sum"),
                    "Number of starters over 40": ("Number of starters over 40", "sum"),
                    "Number of players over 50": ("Number of players over 50", "sum"),
                    "Number of starters over 50": ("Number of starters over 50", "sum"),
                    "Number of cuffs rostered": ("Number of cuffs rostered", "sum"),
                    "Number of cuffs started": ("Number of cuffs started", "sum"),
                    "Weeks of starter injuries": ("Number of starter injuries", "sum"),
                    "Weeks of starter suspensions": ("Number of starter suspensions", "sum"),
                }
            )
            team_year = team_year.merge(agg_year, how="left", on=["Team", "Year"])

        except Exception as e:
            _log_exc(debug, "team_year_aggregate_fill", e)

        # Yearly Luck = Σ of weekly Luck (the agg above). No win% multiplier:
        # the weekly model's calibrated pregame-prob already nets out winning,
        # so summing tracks Win-Variance-style over/under-achievement on its own
        # (see plan/LUCK_REWORK.md).

        # Override cuff counts with UNIQUE player counts (item 9).
        try:
            if not team_year.empty and unique_cuffs_by_team_year:
                for _col in ["Number of cuffs rostered", "Number of cuffs started"]:
                    if _col in team_year.columns:
                        team_year[_col] = [
                            int(unique_cuffs_by_team_year.get((str(t), int(y)), {}).get(_col, 0))
                            if pd.notna(y) else 0
                            for t, y in zip(team_year["Team"], pd.to_numeric(team_year["Year"], errors="coerce"))
                        ]
        except Exception as e:
            _log_exc(debug, "team_year_unique_cuffs", e)

        # Trades split (user request): Offseason / Inseason / Total trades, as
        # DISTINCT trade events the team was in that season. Offseason = trade
        # dated before that season's kickoff (Sept 7); Inseason = on/after.
        # team_all_time sums these per-season counts (a trade lives in one
        # season). Replaces the single "Number of trades" on the year/all-time
        # sheets; the per-week sheets keep "Number of trades".
        try:
            for _t, _y in [("Offseason trades", "off"), ("Inseason trades", "in"), ("Total trades", "tot")]:
                if not team_year.empty:
                    team_year[_t] = [
                        int(len(_team_trade_dates_split.get((str(t), int(y)), {}).get(_y, set()))) if pd.notna(y) else 0
                        for t, y in zip(team_year["Team"], pd.to_numeric(team_year["Year"], errors="coerce"))
                    ]
        except Exception as e:
            _log_exc(debug, "team_year_trades_split", e)


        # Result + vs-category records (rebuilt using normalized games)
        try:

            # Fill Result from bracket-derived finish map when available.
            if "Result" in team_year.columns and season_finish:
                def _res(team, year):
                    m = season_finish.get(int(year), {})
                    return m.get(str(team))
                team_year["Result"] = team_year.apply(lambda r: _res(r["Team"], r["Year"]), axis=1)

            results = []
            for _, r in team_year.iterrows():
                team = str(r["Team"])
                yr = int(r["Year"])
                playoffs = playoff_teams_by_season.get(yr, set())
                champ = champion_by_season.get(yr)
                lastp = last_place_by_season.get(yr)
                teams_in_year = set(team_year[team_year["Year"] == yr]["Team"].astype(str).tolist())
                non_playoffs = teams_in_year - set(playoffs) if playoffs else set()

                res = season_finish.get(yr, {}).get(team, r.get("Result"))
                # PR E fix C: a live (incomplete) season has no finish yet —
                # force N/A rather than letting the fallback below brand every
                # team "Missed playoffs".
                if not _season_is_complete(yr):
                    res = "N/A"
                elif not res:
                    if champ and team == champ:
                        res = "Champion"
                    elif lastp and team == lastp:
                        res = "Last place"
                    elif team in playoffs:
                        res = "Playoffs"
                    else:
                        res = "Missed playoffs"

                wlt_play = _wlt_for_team(games_df, team, year=yr, opps=playoffs if playoffs else None)
                wlt_non = _wlt_for_team(games_df, team, year=yr, opps=non_playoffs if non_playoffs else None)
                wlt_champ = _wlt_for_team(games_df, team, year=yr, opps={champ} if champ else None)
                wlt_last = _wlt_for_team(games_df, team, year=yr, opps={lastp} if lastp else None)

                results.append({
                    "Team": team,
                    "Year": yr,
                    "Result": res,
                    "Record vs playoff teams": _record_str(*wlt_play) if playoffs else "N/A",
                    "Win % vs playoff teams": _win_pct(wlt_play) if playoffs else None,
                    "Record vs non-playoff teams": _record_str(*wlt_non) if non_playoffs else "N/A",
                    "Win % vs non-playoff teams": _win_pct(wlt_non) if non_playoffs else None,
                    "Record vs champion": _record_str(*wlt_champ) if champ else "N/A",
                    "Win % vs champion": _win_pct(wlt_champ) if champ else None,
                    "Record vs last place": _record_str(*wlt_last) if lastp else "N/A",
                    "Win % vs last place": _win_pct(wlt_last) if lastp else None,
                })
            extra = pd.DataFrame(results)
            team_year = team_year.drop(columns=[c for c in extra.columns if c in team_year.columns and c not in ["Team","Year"]], errors="ignore")
            team_year = team_year.merge(extra, how="left", on=["Team","Year"])
            team_year["Week of playoff elimination"] = team_year.get("Week of playoff elimination", "N/A")
            team_year["Offseason roster turnover"] = team_year.get("Offseason roster turnover", 0).fillna(0)
            team_year["Inseason roster turnover"] = team_year.get("Inseason roster turnover", 0).fillna(0)
            team_year["Average weekly starter turnover"] = team_year.get("Average weekly starter turnover", 0.0).fillna(0.0)
            team_year["Average weekly roster turnover"] = team_year.get("Average weekly roster turnover", 0.0).fillna(0.0)
        except Exception as e:
            _log_exc(debug, "team_year_results_vs_records", e)

        team_year = team_year.sort_values(["Team", "Year"]).reset_index(drop=True)
        team_year["Change in win % from previous season"] = team_year.groupby("Team")["Win %"].diff()
        team_year_win_variance = {}
        for team, val in team_year.groupby("Team")["Win Variance"].mean().items():
            team_year_win_variance[str(team)] = float(val) if pd.notna(val) else None

        # team-all-time rollup
        championship_counts = {}
        for champ in champion_by_season.values():
            if champ:
                champ_name = str(champ)
                championship_counts[champ_name] = championship_counts.get(champ_name, 0) + 1
        # New team_all_time counts (per user): playoff appearances (made the
        # top-4 bracket), championship-game appearances (reached the Final =
        # Champion or runner-up), and last-place finishes. Gated to completed
        # seasons via the same dicts that drive Result, so 2026 isn't counted.
        # All three counts derive from `season_finish` — the SAME dict that
        # drives the displayed team_year `Result` (set at line ~12058) — so the
        # counts always agree with the finishes a user can read off the sheet.
        # (last_place_by_season = standings[-1] uses a different, full-standings
        # ordering than Result's regular-season "8th"; using season_finish keeps
        # the three columns mutually + Result-consistent.)
        _finish_rank = {"Champion": 1, "2nd": 2, "3rd": 3, "4th": 4,
                        "5th": 5, "6th": 6, "7th": 7, "8th": 8}
        playoff_appearance_counts: Dict[str, int] = {}
        champ_appearance_counts: Dict[str, int] = {}
        last_place_counts: Dict[str, int] = {}
        for _fs, _fin in season_finish.items():
            if not _fin:
                continue
            # Last place = the worst ordinal actually present that season
            # (robust to leagues with !=8 teams, e.g. the ESPN 2020 backfill).
            _worst = max((_finish_rank.get(v, 0) for v in _fin.values()), default=0)
            for _ft, _fres in _fin.items():
                if _fres in ("Champion", "2nd", "3rd", "4th"):
                    playoff_appearance_counts[str(_ft)] = playoff_appearance_counts.get(str(_ft), 0) + 1
                if _fres in ("Champion", "2nd"):
                    champ_appearance_counts[str(_ft)] = champ_appearance_counts.get(str(_ft), 0) + 1
                if _worst and _finish_rank.get(_fres, 0) == _worst:
                    last_place_counts[str(_ft)] = last_place_counts.get(str(_ft), 0) + 1

        rows = []
        for team, g in tw.groupby("Team"):
            wins = int((g["Win?"] == 1).sum())
            losses = int((g["Win?"] == 0).sum())
            ties = int((g["Win?"] == 0.5).sum())
            gp = max(1, wins + losses + ties)
            pf = float(pd.to_numeric(g["PF"], errors="coerce").fillna(0.0).sum())
            pa = float(pd.to_numeric(g["Points against"], errors="coerce").fillna(0.0).sum())
            diff = pf - pa
            maxpf_sum = float(pd.to_numeric(g["Max PF"], errors="coerce").fillna(0.0).sum())
            maxpf_avg = float(pd.to_numeric(g["Max PF"], errors="coerce").fillna(0.0).mean())
            all_time_place_record_value = all_time_place_record.get(str(team))
            all_time_place_pf_value = all_time_place_pf.get(str(team))
            all_time_place_maxpf_value = all_time_place_maxpf.get(str(team))
            if (
                all_time_place_record_value is not None
                and all_time_place_pf_value is not None
                and all_time_place_maxpf_value is not None
            ):
                win_variance = (
                    all_time_place_record_value
                    - ((all_time_place_pf_value + all_time_place_maxpf_value) / 2)
                )
            else:
                win_variance = None
            # All-time luck = the AVERAGE of each season's luck (a season's
            # luck is the SUM of its weekly luck — team_year keeps that). Weekly
            # luck is ~zero-sum per week, but adversity is a persistent team
            # trait, so a straight sum over every week ever lets a chronically
            # healthy/injured team's luck pile up without bound and grow purely
            # with tenure (steven +7.1 / shmuel −6.9). Averaging seasons keeps
            # all-time luck on the same scale as a single season and is fair
            # across differing tenures.
            _luck_w = pd.to_numeric(g.get("Luck"), errors="coerce").fillna(0.0)
            _luck_yr = g.get("Year")
            avg_yearly_luck = (
                float(_luck_w.groupby(_luck_yr).sum().mean())
                if _luck_yr is not None and len(_luck_w) else 0.0
            )
            # Clutch index (improvement #9, team_all_time only): how a manager
            # performs in the WINNERS'-bracket playoffs vs the regular season,
            # all-time. Playoff = Semifinal / Final / 3rd Place; regular = the
            # "Week N" games. N/A for a team that never reached the playoffs
            # (no delta to take). Win? is numeric (1/0) in-build.
            _clutch_pf = None
            _clutch_wp = None
            _wkn = g.get("Week Name")
            if _wkn is not None:
                _is_po = _wkn.isin(("Semifinal", "Final", "3rd Place"))
                _is_rg = _wkn.astype(str).str.startswith("Week ")
                _pf_n = pd.to_numeric(g["PF"], errors="coerce")
                _w_n = pd.to_numeric(g["Win?"], errors="coerce")
                if _is_po.any() and _is_rg.any():
                    _clutch_pf = round(float(_pf_n[_is_po].mean() - _pf_n[_is_rg].mean()), 2)
                    _clutch_wp = round(float(_w_n[_is_po].mean() - _w_n[_is_rg].mean()), 4)
            row = {
                "Team": str(team),
                "All time win %": round((wins + 0.5 * ties) / gp, 4),
                "All time record": _record_str(wins, losses, ties),
                "Championships": championship_counts.get(str(team), 0),
                "Playoff PF minus regular-season PF": _clutch_pf,
                "Playoff win % minus regular-season win %": _clutch_wp,
                "Number of playoff appearances": playoff_appearance_counts.get(str(team), 0),
                "Number of championship appearances": champ_appearance_counts.get(str(team), 0),
                "Number of last place finishes": last_place_counts.get(str(team), 0),
                "Record & win % vs each team": "N/A",
                "Record & win % vs playoff teams": "N/A",
                "Record & win % vs non-playoff teams": "N/A",
                "Record & win % vs champions": "N/A",
                "Record & win % vs last place": "N/A",
                "Win Variance": team_year_win_variance.get(str(team)),
                "Draft Value": 0,
                "Number of first round picks made": 0,
                "Total number of picks made": 0,
                "Points": round(pf, 2),
                "Avg points": round(pf / gp, 2),
                "Points against": round(pa, 2),
                "Avg points against": round(pa / gp, 2),
                "Differential": round(diff, 2),
                "Avg differential": round(diff / gp, 2),
                "Max PF": round(maxpf_sum, 2),
                "Avg max PF": round(maxpf_avg, 2) if not math.isnan(maxpf_avg) else None,
                "Efficiency": round(pf / maxpf_sum, 4) if maxpf_sum else None,
                "Weeks of injuries": int(pd.to_numeric(g.get("Number of Injuries"), errors="coerce").fillna(0.0).sum()),
                "Weeks suspensions": int(pd.to_numeric(g.get("Number of suspensions"), errors="coerce").fillna(0.0).sum()),
                "Hardship": float(pd.to_numeric(g.get("Hardship"), errors="coerce").fillna(0.0).sum()),
                "Starter-adjusted Hardship": round(float(pd.to_numeric(g.get("Starter-adjusted Hardship"), errors="coerce").fillna(0.0).sum()), 4),
                "Offseason starter turnover": 0,
                "Inseason starter turnover": 0,
                "Offseason roster turnover": 0,
                "Inseason roster turnover": 0,
                # Average of per-season-final Tanking values, not sum.
                # Per-season Tanking is already an expanding-mean tank
                # score; summing seasons creates an unbounded number
                # with no clear interpretation.
                "Tanking": float(pd.to_numeric(g.get("Tanking"), errors="coerce").dropna().mean() or 0.0),
                "Avg yearly luck": round(avg_yearly_luck, 4),
            }
            # NOTE: Unique-player position counts for team_all are added
            # later (line ~7977 "team_unique_player_counts") via direct
            # team_all["Team"].map() assignment — no need to seed here.
            rows.append(row)
        team_all = pd.DataFrame(rows)

        # 3-year roster retention rate (improvement #16) — team_all_time = the
        # AVERAGE of a team's measurable yearly retention rates (N/A if none).
        try:
            if not team_all.empty and "Team" in team_all.columns:
                _ret_by_team: Dict[str, List[float]] = {}
                for (_tm, _yr), _r in retention_3yr_by_ty.items():
                    _ret_by_team.setdefault(str(_tm), []).append(_r)
                team_all["3-year roster retention rate"] = [
                    (round(sum(_ret_by_team[str(t)]) / len(_ret_by_team[str(t)]), 4)
                     if str(t) in _ret_by_team else None)
                    for t in team_all["Team"]
                ]
        except Exception as e:
            _log_exc(debug, "team_all_retention", e)

        # Rebuild all win % and record columns for team-all-time
        try:
            playoff_pairs = {
                (year, team)
                for year, teams in playoff_teams_by_season.items()
                for team in teams
            }
            champ_pairs = {
                (year, champ)
                for year, champ in champion_by_season.items()
                if champ
            }
            last_place_pairs = {
                (year, lastp)
                for year, lastp in last_place_by_season.items()
                if lastp
            }
            non_playoff_pairs = {
                (year, team)
                for year, teams in teams_by_season.items()
                for team in teams
                if team not in playoff_teams_by_season.get(year, set())
            }
            for idx, row in team_all.iterrows():
                team = str(row["Team"])
                wlt = _wlt_for_team(games_df, team)
                team_all.at[idx, "All time record"] = _record_str(*wlt)
                team_all.at[idx, "All time win %"] = _win_pct(wlt)
                # All-time bracket records: regular season / winners' playoffs
                # (Semifinal+Final) / toilet bowl (Toilet Semis+Final). Win % is
                # N/A for a bracket never reached.
                for _kind, _lbl in (("regular", "Regular season"), ("playoff", "Playoff"), ("toilet", "Toilet bowl")):
                    _b = _wlt_bracket(games_df, team, _kind)
                    team_all.at[idx, f"{_lbl} record"] = _record_str(*_b)
                    team_all.at[idx, f"{_lbl} win %"] = _win_pct_or_na(_b)
                # The two consolation games on their own (record only — small N).
                team_all.at[idx, "Third place game record"] = _record_str(*_wlt_bracket(games_df, team, "third_place"))
                team_all.at[idx, "Toilet losers game record"] = _record_str(*_wlt_bracket(games_df, team, "toilet_losers"))

                opp_list = [t for t in teams if t != team]
                pieces = []
                played_winpcts: List[Tuple[str, float]] = []  # (opp, win%) for opps actually played
                for opp in opp_list:
                    wlt_opp = _wlt_for_team(games_df, team, opps={opp})
                    pieces.append(f"{opp}: {_record_str(*wlt_opp)} ({_win_pct(wlt_opp)})")
                    team_all.at[idx, f"Record vs {opp}"] = _record_str(*wlt_opp)
                    team_all.at[idx, f"Win % vs {opp}"] = _win_pct(wlt_opp)
                    if sum(wlt_opp) > 0:  # at least one game vs this opponent
                        played_winpcts.append((opp, _win_pct(wlt_opp)))
                team_all.at[idx, "Record & win % vs each team"] = "; ".join(pieces) if pieces else "N/A"

                # Item 13: highest / lowest Win % vs a single opponent (only
                # opponents actually played), plus the opponent name for each.
                if played_winpcts:
                    _hi = max(played_winpcts, key=lambda kv: kv[1])
                    _lo = min(played_winpcts, key=lambda kv: kv[1])
                    team_all.at[idx, "Highest Win % vs a team"] = _hi[1]
                    team_all.at[idx, "Team for highest Win %"] = _hi[0]
                    team_all.at[idx, "Lowest Win % vs a team"] = _lo[1]
                    team_all.at[idx, "Team for lowest Win %"] = _lo[0]

            for idx, row in team_all.iterrows():
                team = str(row["Team"])
                wlt_play = _wlt_for_team_pairs(games_df, team, playoff_pairs)
                wlt_non = _wlt_for_team_pairs(games_df, team, non_playoff_pairs)
                wlt_ch = _wlt_for_team_pairs(games_df, team, champ_pairs)
                wlt_last = _wlt_for_team_pairs(games_df, team, last_place_pairs)

                team_all.at[idx, "Record & win % vs playoff teams"] = f"{_record_str(*wlt_play)} ({_win_pct(wlt_play)})" if playoff_pairs else "N/A"
                team_all.at[idx, "Record & win % vs non-playoff teams"] = f"{_record_str(*wlt_non)} ({_win_pct(wlt_non)})" if non_playoff_pairs else "N/A"
                team_all.at[idx, "Record & win % vs champions"] = f"{_record_str(*wlt_ch)} ({_win_pct(wlt_ch)})" if champ_pairs else "N/A"
                team_all.at[idx, "Record & win % vs last place"] = f"{_record_str(*wlt_last)} ({_win_pct(wlt_last)})" if last_place_pairs else "N/A"
        except Exception as e:
            _log_exc(debug, "team_all_rebuild_records", e)

        # --------------------------
        # Fill missing Team-all-time columns from team-week (flags, roster composition, etc.)
        # --------------------------
        try:
            agg_all = tw.groupby("Team", as_index=False).agg(
                **{
                    "Times Brosenzweig": ("Brosenzweig", "sum"),
                    "Times Sisenzweig": ("Sisenzweig", "sum"),
                    "Times Highest score?": ("Highest score?", "sum"),
                    "Times Lowest score?": ("Lowest score?", "sum"),
                    "Times Narrowest victory?": ("Narrowest victory?", "sum"),
                    "Times Largest blowout?": ("Largest blowout?", "sum"),
                    "Times Most efficient?": ("Most efficient?", "sum"),
                    "Times Least efficient?": ("Least efficient?", "sum"),
                    "Times Top half of league?": ("Top half of league?", "sum"),
                    "Times One-man army?": ("One-man army?", "sum"),
                    "Times Most bench points?": ("Most bench points?", "sum"),
                    "Times Most injured?": ("Most injured?", "sum"),
                    "Most number of players started from same NFL team": ("Most number of players started from same NFL team", "max"),
                    "Most number of players rostered from same NFL team": ("Most number of players rostered from same NFL team", "max"),
                    "Most number of QBs started from same NFL team": ("Most number of QBs started from same NFL team", "max"),
                    "Most number of QBs rostered from same NFL team": ("Most number of QBs rostered from same NFL team", "max"),
                    "Most number of RBs started from same NFL team": ("Most number of RBs started from same NFL team", "max"),
                    "Most number of RBs rostered from same NFL team": ("Most number of RBs rostered from same NFL team", "max"),
                    "Most number of WR started from same NFL team": ("Most number of WR started from same NFL team", "max"),
                    "Most number of WR rostered from same NFL team": ("Most number of WR rostered from same NFL team", "max"),
                    "Most number of TE started from same NFL team": ("Most number of TE started from same NFL team", "max"),
                    "Most number of TE rostered from same NFL team": ("Most number of TE rostered from same NFL team", "max"),
                    "Number of NFL teams among starting players": ("Number of NFL teams among starting players", "max"),
                    "Number of NFL teams among rostered players": ("Number of NFL teams among rostered players", "max"),
                    "Number of rookies started": ("Number of rookies started", "sum"),
                    "Number of rookies rostered": ("Number of rookies rostered", "sum"),
                    "Player average age": ("Player average age", "mean"),
                    "Team age including picks": ("Team age including picks", "mean"),
                    "Difference between highest and lowest starters": ("Difference between highest and lowest starters", "max"),
                    "Combined matchup score": ("Combined matchup score", "max"),
                    "Number of donuts": ("Number of donuts", "sum"),
                    "Number of starter donuts": ("Number of starter donuts", "sum"),
                    "Number of players under 10": ("Number of players under 10", "sum"),
                    "Number of starters under 10": ("Number of starters under 10", "sum"),
                    "Number of players over 20": ("Number of players over 20", "sum"),
                    "Number of starters over 20": ("Number of starters over 20", "sum"),
                    "Number of players over 30": ("Number of players over 30", "sum"),
                    "Number of starters over 30": ("Number of starters over 30", "sum"),
                    "Number of players over 40": ("Number of players over 40", "sum"),
                    "Number of starters over 40": ("Number of starters over 40", "sum"),
                    "Number of players over 50": ("Number of players over 50", "sum"),
                    "Number of starters over 50": ("Number of starters over 50", "sum"),
                    "Number of cuffs rostered": ("Number of cuffs rostered", "sum"),
                    "Number of cuffs started": ("Number of cuffs started", "sum"),
                    "Weeks of starter injuries": ("Number of starter injuries", "sum"),
                    "Weeks of starter suspensions": ("Number of starter suspensions", "sum"),
                }
            )
            team_all = team_all.merge(agg_all, how="left", on="Team")
        except Exception as e:
            _log_exc(debug, "team_all_aggregate_fill", e)

        # Season-based streaks (the season grain is the natural "weekly" unit
        # for these, so they live on team_year, not team_week): consecutive
        # seasons making the playoffs / finishing >= .500. Terminal-encoded
        # like the weekly streaks — only the final season of a run carries the
        # length, intermediate seasons read "In Progress", broken/none reads 0,
        # and a not-yet-complete season reads "N/A".
        try:
            if not team_year.empty and {"Team", "Year"}.issubset(team_year.columns):
                ty = team_year.copy()
                ty["__yr"] = pd.to_numeric(ty["Year"], errors="coerce")
                _wpct = pd.to_numeric(ty.get("Win %"), errors="coerce")
                pa_out: Dict[Any, Any] = {}
                ws_out: Dict[Any, Any] = {}
                for team, g in ty.sort_values(["Team", "__yr"]).groupby("Team"):
                    g = g.reset_index()  # 'index' col holds the team_year row index
                    pa_run = []  # (row_index, value_or_None_if_incomplete)
                    ws_run = []
                    pa = ws = 0
                    for _, r in g.iterrows():
                        ri = r["index"]
                        yr = r["__yr"]
                        if pd.notna(yr) and int(yr) in playoff_teams_by_season:
                            made = str(team) in playoff_teams_by_season.get(int(yr), set())
                            pa = pa + 1 if made else 0
                            pa_run.append((ri, pa))
                        else:
                            pa_run.append((ri, None))
                        wp = r.get("Win %")
                        wp = pd.to_numeric(pd.Series([wp]), errors="coerce").iloc[0]
                        if pd.notna(wp):
                            ws = ws + 1 if float(wp) >= 0.5 else 0
                            ws_run.append((ri, ws))
                        else:
                            ws_run.append((ri, None))
                    for run, out in ((pa_run, pa_out), (ws_run, ws_out)):
                        for k in range(len(run)):
                            ri, c = run[k]
                            if c is None:
                                out[ri] = "N/A"
                            elif c == 0:
                                out[ri] = 0
                            elif k + 1 < len(run) and run[k + 1][1] == c + 1:
                                out[ri] = "In Progress"
                            else:
                                out[ri] = c
                team_year["Playoff appearance streak"] = [pa_out.get(i, "N/A") for i in team_year.index]
                team_year["Winning season streak"] = [ws_out.get(i, "N/A") for i in team_year.index]
        except Exception as e:
            _log_exc(debug, "season_based_streaks", e)

        # Roll up team_year-only fields into team_all_time. These were all
        # hardcoded to 0 (Draft Value, picks made, transactions, trades,
        # FAAB, turnover, future cap) because the previous agg_all merge
        # only pulled from team_week which doesn't have them.
        try:
            if not team_year.empty:
                ty_for_all = team_year.groupby("Team", as_index=False).agg(
                    **{
                        "Draft Value": ("Draft Value", "sum"),
                        "Number of first round picks made": ("Number of first round picks made", "sum"),
                        "Total number of picks made": ("Total number of picks made", "sum"),
                        "Number of transactions": ("Number of transactions", "sum"),
                        "Offseason trades": ("Offseason trades", "sum"),
                        "Inseason trades": ("Inseason trades", "sum"),
                        "Total trades": ("Total trades", "sum"),
                        "Amount of FAAB spent": ("Amount of FAAB spent", "sum"),
                        # Per-season AVERAGE at all-time (item 4), not a sum —
                        # "how much does this team typically turn over per year".
                        "Offseason starter turnover": ("Offseason starter turnover", "mean"),
                        "Inseason starter turnover": ("Inseason starter turnover", "mean"),
                        "Offseason roster turnover": ("Offseason roster turnover", "mean"),
                        "Inseason roster turnover": ("Inseason roster turnover", "mean"),
                        "Average weekly starter turnover": ("Average weekly starter turnover", "mean"),
                        "Average weekly roster turnover": ("Average weekly roster turnover", "mean"),
                        "Future draft capital": ("Future draft capital", "sum"),
                    }
                )
                team_all = team_all.drop(
                    columns=[c for c in ty_for_all.columns if c != "Team" and c in team_all.columns],
                    errors="ignore",
                ).merge(ty_for_all, on="Team", how="left")
        except Exception as e:
            _log_exc(debug, "team_all_from_team_year", e)

        # All-time Luck = Σ of all weekly Luck (no win% multiplier — see the
        # team_year note above and plan/LUCK_REWORK.md).

        # Override cuff counts with UNIQUE player counts (item 9).
        try:
            if not team_all.empty and unique_cuffs_by_team_all:
                for _col in ["Number of cuffs rostered", "Number of cuffs started"]:
                    if _col in team_all.columns:
                        team_all[_col] = [
                            int(unique_cuffs_by_team_all.get((str(t),), {}).get(_col, 0))
                            for t in team_all["Team"]
                        ]
        except Exception as e:
            _log_exc(debug, "team_all_unique_cuffs", e)

        # Top-up rollup: trades / transactions / FAAB from preseason seasons
        # that didn't produce a team_year row (e.g. 2026 before NFL Week 1).
        # We read trades_rows / transactions_rows directly and add only the
        # delta — counts already attributed to a team_year stay attributed.
        try:
            ty_years = set(team_year["Year"].astype(int).unique().tolist()) if not team_year.empty else set()
            # Trades not represented in team_year
            extra_trades_by_team: Dict[str, int] = defaultdict(int)
            for tr_row in trades_rows:
                try:
                    dt = tr_row.get("Date")
                    if not dt:
                        continue
                    yr = int(str(dt)[:4])
                    if yr in ty_years:
                        continue
                    tm = tr_row.get("Team")
                    if tm:
                        extra_trades_by_team[str(tm)] += 1
                except Exception:
                    continue
            extra_tx_by_team: Dict[str, int] = defaultdict(int)
            extra_faab_by_team: Dict[str, float] = defaultdict(float)
            for tx_row in transactions_rows:
                try:
                    dt = tx_row.get("Date")
                    if not dt:
                        continue
                    yr = int(str(dt)[:4])
                    if yr in ty_years:
                        continue
                    tm = tx_row.get("Team")
                    if tm:
                        extra_tx_by_team[str(tm)] += 1
                        try:
                            faab = float(tx_row.get("Faab") or 0.0)
                        except Exception:
                            faab = 0.0
                        extra_faab_by_team[str(tm)] += faab
                except Exception:
                    continue

            if extra_trades_by_team or extra_tx_by_team or extra_faab_by_team:
                for idx, row in team_all.iterrows():
                    tm = str(row.get("Team") or "")
                    if tm in extra_tx_by_team:
                        cur = pd.to_numeric(team_all.at[idx, "Number of transactions"], errors="coerce")
                        team_all.at[idx, "Number of transactions"] = int((0 if pd.isna(cur) else cur) + extra_tx_by_team[tm])
                    if tm in extra_faab_by_team:
                        cur = pd.to_numeric(team_all.at[idx, "Amount of FAAB spent"], errors="coerce")
                        team_all.at[idx, "Amount of FAAB spent"] = round(float((0 if pd.isna(cur) else cur) + extra_faab_by_team[tm]), 2)
        except Exception as e:
            _log_exc(debug, "team_all_preseason_topup", e)

        # team_all_time Offseason/Inseason/Total trades: count distinct trades
        # directly across ALL seasons (incl. preseason seasons with no team_year
        # row), authoritative over the team_year sum.
        try:
            if not team_all.empty:
                _tt_all: Dict[str, Dict[str, set]] = defaultdict(lambda: {"off": set(), "in": set(), "tot": set()})
                for (tm, _sy), buckets in _team_trade_dates_split.items():
                    for _bk in ("off", "in", "tot"):
                        _tt_all[tm][_bk] |= buckets[_bk]
                for _col, _bk in [("Offseason trades", "off"), ("Inseason trades", "in"), ("Total trades", "tot")]:
                    team_all[_col] = [int(len(_tt_all.get(str(t), {}).get(_bk, set()))) for t in team_all["Team"]]
        except Exception as e:
            _log_exc(debug, "team_all_trades_split", e)

        # Unique-player positional counts for team-year and team-all-time (by Player ID)
        try:
            required_cols = {"Team", "Year", "Player ID", "Position", "Starter/Bench"}
            if not pw.empty and required_cols.issubset(set(pw.columns)):
                pw_counts = pw[list(required_cols)].copy()
                pw_counts = pw_counts.dropna(subset=["Player ID", "Team", "Year"])
                pw_counts["Player ID"] = pw_counts["Player ID"].astype(str)
                pw_counts["Position"] = pw_counts["Position"].astype(str).str.upper()
                pw_counts["Starter/Bench"] = pw_counts["Starter/Bench"].astype(str)
                pw_counts = pw_counts[pw_counts["Position"].isin(["QB", "RB", "WR", "TE"])]

                year_rostered = (
                    pw_counts.groupby(["Team", "Year", "Position"])["Player ID"]
                    .nunique()
                    .unstack("Position")
                    .fillna(0)
                    .reset_index()
                )
                year_started = (
                    pw_counts[pw_counts["Starter/Bench"].str.lower().eq("starter")]
                    .groupby(["Team", "Year", "Position"])["Player ID"]
                    .nunique()
                    .unstack("Position")
                    .fillna(0)
                    .reset_index()
                )

                started_rename = {pos: f"Number of {pos} started" for pos in ["QB", "RB", "WR", "TE"] if pos in year_started.columns}
                rostered_rename = {pos: f"Number of {pos} rostered" for pos in ["QB", "RB", "WR", "TE"] if pos in year_rostered.columns}
                year_started = year_started.rename(columns=started_rename)
                year_rostered = year_rostered.rename(columns=rostered_rename)

                if not team_year.empty:
                    team_year = team_year.merge(year_started, on=["Team", "Year"], how="left")
                    team_year = team_year.merge(year_rostered, on=["Team", "Year"], how="left")
                    for pos in ["QB", "RB", "WR", "TE"]:
                        team_year[f"Number of {pos} started"] = (
                            pd.to_numeric(team_year.get(f"Number of {pos} started"), errors="coerce")
                            .fillna(0)
                            .astype(int)
                        )
                        team_year[f"Number of {pos} rostered"] = (
                            pd.to_numeric(team_year.get(f"Number of {pos} rostered"), errors="coerce")
                            .fillna(0)
                            .astype(int)
                        )

                all_rostered = (
                    pw_counts.groupby(["Team", "Position"])["Player ID"]
                    .nunique()
                    .unstack("Position")
                    .fillna(0)
                )
                all_started = (
                    pw_counts[pw_counts["Starter/Bench"].str.lower().eq("starter")]
                    .groupby(["Team", "Position"])["Player ID"]
                    .nunique()
                    .unstack("Position")
                    .fillna(0)
                )

                if not team_all.empty:
                    for pos in ["QB", "RB", "WR", "TE"]:
                        if pos in all_started.columns:
                            team_all[f"Number of {pos} started"] = (
                                team_all["Team"].map(all_started[pos]).fillna(0).astype(int)
                            )
                        else:
                            team_all[f"Number of {pos} started"] = 0
                        if pos in all_rostered.columns:
                            team_all[f"Number of {pos} rostered"] = (
                                team_all["Team"].map(all_rostered[pos]).fillna(0).astype(int)
                            )
                        else:
                            team_all[f"Number of {pos} rostered"] = 0

                if "Rookie?" in pw.columns:
                    rookies = pw[["Team", "Year", "Player ID", "Starter/Bench", "Rookie?"]].copy()
                    rookies = rookies.dropna(subset=["Player ID", "Team", "Year"])
                    rookies["Player ID"] = rookies["Player ID"].astype(str)
                    rookies["Starter/Bench"] = rookies["Starter/Bench"].astype(str)
                    rookies["Rookie?"] = pd.to_numeric(rookies["Rookie?"], errors="coerce").fillna(0).astype(int)
                    rookies = rookies[rookies["Rookie?"] == 1]

                    year_rookies_rostered = (
                        rookies.groupby(["Team", "Year"])["Player ID"]
                        .nunique()
                        .reset_index()
                    )
                    year_rookies_started = (
                        rookies[rookies["Starter/Bench"].str.lower().eq("starter")]
                        .groupby(["Team", "Year"])["Player ID"]
                        .nunique()
                        .reset_index()
                    )

                    if not team_year.empty:
                        team_year = team_year.drop(
                            columns=["Number of rookies started", "Number of rookies rostered"],
                            errors="ignore",
                        )
                        team_year = team_year.merge(
                            year_rookies_started.rename(columns={"Player ID": "Number of rookies started"}),
                            on=["Team", "Year"],
                            how="left",
                        )
                        team_year = team_year.merge(
                            year_rookies_rostered.rename(columns={"Player ID": "Number of rookies rostered"}),
                            on=["Team", "Year"],
                            how="left",
                        )
                        team_year["Number of rookies started"] = (
                            pd.to_numeric(team_year.get("Number of rookies started"), errors="coerce")
                            .fillna(0)
                            .astype(int)
                        )
                        team_year["Number of rookies rostered"] = (
                            pd.to_numeric(team_year.get("Number of rookies rostered"), errors="coerce")
                            .fillna(0)
                            .astype(int)
                        )

                    all_rookies_rostered = rookies.groupby("Team")["Player ID"].nunique()
                    all_rookies_started = (
                        rookies[rookies["Starter/Bench"].str.lower().eq("starter")]
                        .groupby("Team")["Player ID"]
                        .nunique()
                    )

                    if not team_all.empty:
                        team_all["Number of rookies started"] = (
                            team_all["Team"].map(all_rookies_started).fillna(0).astype(int)
                        )
                        team_all["Number of rookies rostered"] = (
                            team_all["Team"].map(all_rookies_rostered).fillna(0).astype(int)
                        )
        except Exception as e:
            _log_exc(debug, "team_unique_player_counts", e)

        # vs-category records (all-time)
        try:
            playoff_pairs = {
                (year, team)
                for year, teams in playoff_teams_by_season.items()
                for team in teams
            }
            champ_pairs = {
                (year, champ)
                for year, champ in champion_by_season.items()
                if champ
            }
            last_place_pairs = {
                (year, lastp)
                for year, lastp in last_place_by_season.items()
                if lastp
            }
            non_playoff_pairs = {
                (year, team)
                for year, teams in teams_by_season.items()
                for team in teams
                if team not in playoff_teams_by_season.get(year, set())
            }

            extra = []
            for team in team_all["Team"].astype(str).tolist():
                wlt_play = _wlt_for_team_pairs(games_df, team, playoff_pairs)
                wlt_non = _wlt_for_team_pairs(games_df, team, non_playoff_pairs)
                wlt_ch = _wlt_for_team_pairs(games_df, team, champ_pairs)
                wlt_last = _wlt_for_team_pairs(games_df, team, last_place_pairs)

                extra.append({
                    "Team": team,
                    "Record vs playoff teams": _record_str(*wlt_play) if playoff_pairs else "N/A",
                    "Win % vs playoff teams": _win_pct(wlt_play) if playoff_pairs else None,
                    "Record vs non-playoff teams": _record_str(*wlt_non) if non_playoff_pairs else "N/A",
                    "Win % vs non-playoff teams": _win_pct(wlt_non) if non_playoff_pairs else None,
                    "Record vs champions": _record_str(*wlt_ch) if champ_pairs else "N/A",
                    "Win % vs champions": _win_pct(wlt_ch) if champ_pairs else None,
                    "Record vs last place": _record_str(*wlt_last) if last_place_pairs else "N/A",
                    "Win % vs last place": _win_pct(wlt_last) if last_place_pairs else None,
                })
            extra = pd.DataFrame(extra)
            team_all = team_all.drop(columns=[c for c in extra.columns if c in team_all.columns and c!="Team"], errors="ignore")
            team_all = team_all.merge(extra, how="left", on="Team")
        except Exception as e:
            _log_exc(debug, "team_all_vs_records", e)

    # League rollups
    # League-wide unique extras (Phase 5B item 2): rookies and NFL-team counts
    # at year / all-time must be DISTINCT across the whole period, not summed
    # weekly (rookies) or the weekly max (NFL teams). Built from player_week.
    def _league_unique_extras(pw_df: pd.DataFrame, group_cols: List[str]) -> Dict[Tuple, Dict[str, int]]:
        out: Dict[Tuple, Dict[str, int]] = {}
        if pw_df.empty or "Player ID" not in pw_df.columns:
            return out
        df = pw_df.copy()
        df["_starter"] = df.get("Starter/Bench", "").astype(str).str.lower() == "starter"
        df["_rookie"] = df.get("Rookie?", "").astype(str).str.lower().isin(["true", "1", "yes"])
        df["_nfl"] = df.get("NFL team", "").astype(str).str.strip()
        for col in group_cols:
            if col == "Year":
                df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64").map(
                    lambda x: int(x) if pd.notna(x) else None
                )
            else:
                df[col] = df[col].astype(object)
        df = df.dropna(subset=group_cols)

        def _key(k):
            return k if isinstance(k, tuple) else (k,)

        rk = df[df["_rookie"] & df["Player ID"].notna()]
        for k, v in rk.groupby(group_cols)["Player ID"].nunique().items():
            out.setdefault(_key(k), {})["Number of rookies rostered"] = int(v)
        for k, v in rk[rk["_starter"]].groupby(group_cols)["Player ID"].nunique().items():
            out.setdefault(_key(k), {})["Number of rookies started"] = int(v)
        nf = df[df["_nfl"].ne("") & df["_nfl"].str.lower().ne("nan")]
        for k, v in nf.groupby(group_cols)["_nfl"].nunique().items():
            out.setdefault(_key(k), {})["Number of NFL teams among rostered players"] = int(v)
        for k, v in nf[nf["_starter"]].groupby(group_cols)["_nfl"].nunique().items():
            out.setdefault(_key(k), {})["Number of NFL teams among starting players"] = int(v)
        return out

    _LX_KEYS = ["Number of rookies started", "Number of rookies rostered",
                "Number of NFL teams among starting players", "Number of NFL teams among rostered players"]
    league_extra_by_year: Dict[Tuple, Dict[str, int]] = {}
    league_extra_all: Dict[str, int] = {}
    if not pw.empty:
        try:
            league_extra_by_year = _league_unique_extras(pw, ["Year"])
            league_extra_all = _league_unique_extras(pw.assign(_all="all"), ["_all"]).get(("all",), {})
        except Exception as e:
            _log_exc(debug, "league_unique_extras", e)

    # Distinct trade events per period (Phase 5B item 1): each trade has a unique
    # created timestamp shared by all participating teams' rows, so a trade is
    # counted ONCE league-wide instead of once per team (the old team-sum
    # double-counted 2-team trades and triple-counted 3-team trades).
    _trade_dates_by_year: Dict[int, set] = defaultdict(set)
    _trade_dates_by_yw: Dict[Tuple[int, int], set] = defaultdict(set)
    _trade_dates_all: set = set()

    def _trade_wk(dt_str, season):
        # Phase 5C item 9: an offseason trade rolls into Wk 1's WEEKLY bucket
        # only if within 7 days before kickoff; deeper-offseason trades get 0
        # (no weekly bucket). The league season/all-time totals are counted
        # from distinct trade dates per Season, so they still include these.
        if not dt_str:
            return 1
        try:
            d = datetime.fromisoformat(str(dt_str).replace("Z", "+00:00")).date()
            ss = date(int(season), 9, 7)
            if d < ss:
                return 1 if (ss - d).days <= 7 else 0
            return max(1, min(17, (d - ss).days // 7 + 1))
        except Exception:
            return 1

    for _tr in trades_rows:
        _d = _tr.get("Date")
        _sy = _to_int(_tr.get("Season"), None)
        if _d and _sy is not None:
            _trade_dates_by_year[int(_sy)].add(str(_d))
            _trade_dates_by_yw[(int(_sy), int(_trade_wk(_d, int(_sy))))].add(str(_d))
            _trade_dates_all.add(str(_d))

    # Highest / lowest single STARTER score, league-wide (Phase 5B item 5).
    _star_hi_yw: Dict[Tuple[int, int], float] = {}
    _star_lo_yw: Dict[Tuple[int, int], float] = {}
    _star_hi_y: Dict[int, float] = {}
    _star_lo_y: Dict[int, float] = {}
    _star_hi_all = None
    _star_lo_all = None
    if not pw.empty and "Starter/Bench" in pw.columns:
        try:
            _pst = pw[pw["Starter/Bench"].astype(str).str.lower() == "starter"].copy()
            _pst["_pts"] = pd.to_numeric(_pst.get("Points"), errors="coerce")
            _pst["_y"] = pd.to_numeric(_pst.get("Year"), errors="coerce")
            _pst["_w"] = pd.to_numeric(_pst.get("Week"), errors="coerce")
            _pst = _pst.dropna(subset=["_pts", "_y", "_w"])
            for (yy, ww), gg in _pst.groupby(["_y", "_w"]):
                _star_hi_yw[(int(yy), int(ww))] = float(gg["_pts"].max())
                _star_lo_yw[(int(yy), int(ww))] = float(gg["_pts"].min())
            for yy, gg in _pst.groupby("_y"):
                _star_hi_y[int(yy)] = float(gg["_pts"].max())
                _star_lo_y[int(yy)] = float(gg["_pts"].min())
            _star_hi_all = float(_pst["_pts"].max())
            _star_lo_all = float(_pst["_pts"].min())
        except Exception as e:
            _log_exc(debug, "league_starter_hi_lo", e)

    league_week = pd.DataFrame()
    league_year = pd.DataFrame()
    league_all = pd.DataFrame()
    if not tw.empty:
        g_week = tw.copy()
        g_week["PF"] = pd.to_numeric(g_week["PF"], errors="coerce").fillna(0.0)
        g_week["Margin"] = pd.to_numeric(g_week["Margin"], errors="coerce")
        g_week["Efficiency"] = pd.to_numeric(g_week["Efficiency"], errors="coerce")
        g_week["Max PF"] = pd.to_numeric(g_week["Max PF"], errors="coerce").fillna(0.0)

        rows = []
        for (yr, wk), g in g_week.groupby(["Year", "Week"]):
            margin_abs = g["Margin"].abs()
            rows.append({
                "Year": int(yr),
                "Week": int(wk),
                "PF": float(g["PF"].sum()),
                "PF Range": float(g["PF"].max() - g["PF"].min()) if not g.empty else 0.0,
                "Avg margin": (float(g.loc[pd.to_numeric(g.get("Win?"), errors="coerce") == 1, "Margin"].mean()) if (pd.to_numeric(g.get("Win?"), errors="coerce") == 1).any() else None),
                "Margin range": (float(g.loc[pd.to_numeric(g.get("Win?"), errors="coerce") == 1, "Margin"].max() - g.loc[pd.to_numeric(g.get("Win?"), errors="coerce") == 1, "Margin"].min()) if (pd.to_numeric(g.get("Win?"), errors="coerce") == 1).any() else None),
                "Number of games within 10": int((margin_abs <= 10).sum() / 2),
                "Number of games within 5": int((margin_abs <= 5).sum() / 2),
                "Max PF": float(g["Max PF"].sum()),
                "Efficiency": float(g["Efficiency"].mean()) if g["Efficiency"].notna().any() else None,
                "Number of Injuries": int(pd.to_numeric(g.get("Number of Injuries"), errors="coerce").fillna(0.0).sum()),
                "Number of suspensions": int(pd.to_numeric(g.get("Number of suspensions"), errors="coerce").fillna(0.0).sum()),
                "Number of players on bye": int(pd.to_numeric(g.get("Number of players on bye"), errors="coerce").fillna(0.0).sum()),
                # League weekly starter turnover = league-wide TOTAL (sum of
                # every team's turnover that week), not the average.
                "Starter turnover from previous week": float(pd.to_numeric(g.get("Starter turnover from previous week"), errors="coerce").fillna(0.0).sum()),
                "UPST": int(pd.to_numeric(g.get("UPST"), errors="coerce").fillna(0.0).sum()),
                "Hardship": float(pd.to_numeric(g.get("Hardship"), errors="coerce").fillna(0.0).sum()),
                "Starter-adjusted Hardship": round(float(pd.to_numeric(g.get("Starter-adjusted Hardship"), errors="coerce").fillna(0.0).sum()), 4),
                # League-week Tanking = mean across teams. Per-team
                # Tanking is already a normalized score; summing 8
                # teams' scores would be misleading.
                "Tanking": float(pd.to_numeric(g.get("Tanking"), errors="coerce").dropna().mean() or 0.0),
                "Luck": float(pd.to_numeric(g.get("Luck"), errors="coerce").fillna(0.0).sum()),
                "Increase in points from previous week": float(pd.to_numeric(g.get("Increase in points from previous week"), errors="coerce").fillna(0.0).sum()),
                "Number of QB started": int(pd.to_numeric(g.get("Number of QB started"), errors="coerce").fillna(0.0).sum()),
                "Number of WR started": int(pd.to_numeric(g.get("Number of WR started"), errors="coerce").fillna(0.0).sum()),
                "Number of RB started": int(pd.to_numeric(g.get("Number of RB started"), errors="coerce").fillna(0.0).sum()),
                "Number of TE started": int(pd.to_numeric(g.get("Number of TE started"), errors="coerce").fillna(0.0).sum()),
                "Number of QB rostered": int(pd.to_numeric(g.get("Number of QB rostered"), errors="coerce").fillna(0.0).sum()),
                "Number of WR rostered": int(pd.to_numeric(g.get("Number of WR rostered"), errors="coerce").fillna(0.0).sum()),
                "Number of RB rostered": int(pd.to_numeric(g.get("Number of RB rostered"), errors="coerce").fillna(0.0).sum()),
                "Number of TE rostered": int(pd.to_numeric(g.get("Number of TE rostered"), errors="coerce").fillna(0.0).sum()),
                "Number of transactions": int(pd.to_numeric(g.get("Number of transactions"), errors="coerce").fillna(0.0).sum()),
                # Distinct trade events this week (Phase 5B item 1), not the
                # per-team sum.
                "Number of trades": int(len(_trade_dates_by_yw.get((int(yr), int(wk)), set()))),
                "Highest starter score": _star_hi_yw.get((int(yr), int(wk))),
                "Lowest starter score": _star_lo_yw.get((int(yr), int(wk))),
                "Amount of FAAB spent": float(pd.to_numeric(g.get("Amount of FAAB spent"), errors="coerce").fillna(0.0).sum()),
            })
        league_week = pd.DataFrame(rows)

        # Attach Week Name (custom week naming) if available
        try:
            if 'week_name_global' in locals() and (not week_name_global.empty):
                league_week = league_week.merge(week_name_global, on=["Year","Week"], how="left")
        except Exception as e:
            _log_exc(debug, "league_week_week_name", e)


        # Fill additional league-week columns from team-week (schema completeness)
        try:
            agg_lw = g_week.groupby(["Year","Week"], as_index=False).agg(
                **{
                    "Most number of players started from same NFL team": ("Most number of players started from same NFL team", "max"),
                    "Most number of players rostered from same NFL team": ("Most number of players rostered from same NFL team", "max"),
                    "Most number of QBs started from same NFL team": ("Most number of QBs started from same NFL team", "max"),
                    "Most number of QBs rostered from same NFL team": ("Most number of QBs rostered from same NFL team", "max"),
                    "Most number of RBs started from same NFL team": ("Most number of RBs started from same NFL team", "max"),
                    "Most number of RBs rostered from same NFL team": ("Most number of RBs rostered from same NFL team", "max"),
                    "Most number of WR started from same NFL team": ("Most number of WR started from same NFL team", "max"),
                    "Most number of WR rostered from same NFL team": ("Most number of WR rostered from same NFL team", "max"),
                    "Most number of TE started from same NFL team": ("Most number of TE started from same NFL team", "max"),
                    "Most number of TE rostered from same NFL team": ("Most number of TE rostered from same NFL team", "max"),
                    "Number of NFL teams among starting players": ("Number of NFL teams among starting players", "max"),
                    "Number of NFL teams among rostered players": ("Number of NFL teams among rostered players", "max"),
                    "Number of rookies started": ("Number of rookies started", "sum"),
                    "Number of rookies rostered": ("Number of rookies rostered", "sum"),
                    "Player average age": ("Player average age", "mean"),
                    "Team age including picks": ("Team age including picks", "mean"),
                    "Difference between highest and lowest starters": ("Difference between highest and lowest starters", "max"),
                    "Number of donuts": ("Number of donuts", "sum"),
                    "Number of starting donuts": ("Number of starter donuts", "sum"),
                    "Number of players under 10": ("Number of players under 10", "sum"),
                    "Number of players over 20": ("Number of players over 20", "sum"),
                    "Number of players over 30": ("Number of players over 30", "sum"),
                    "Number of players over 40": ("Number of players over 40", "sum"),
                    "Number of players over 50": ("Number of players over 50", "sum"),
                    "Number of cuffs rostered": ("Number of cuffs rostered", "sum"),
                    "Number of cuffs started": ("Number of cuffs started", "sum"),
                    "Startup draft players remaining": ("Startup draft players remaining", "max"),
                }
            )
            league_week = league_week.merge(agg_lw, how="left", on=["Year","Week"])
            league_week["Amount of FAAB spent"] = pd.to_numeric(league_week.get("Amount of FAAB spent"), errors="coerce").fillna(0.0)
            # League range = highest − lowest starter (item-5 disambiguation;
            # the agg gave the max single-team spread).
            if {"Highest starter score", "Lowest starter score"} <= set(league_week.columns):
                league_week["Difference between highest and lowest starters"] = (
                    pd.to_numeric(league_week["Highest starter score"], errors="coerce")
                    - pd.to_numeric(league_week["Lowest starter score"], errors="coerce")
                )
        except Exception as e:
            _log_exc(debug, "league_week_fill_extra", e)

        rows = []
        for yr, g in g_week.groupby("Year"):
            margin_abs = g["Margin"].abs()
            win_mask = pd.to_numeric(g.get("Win?"), errors="coerce") == 1
            rows.append({
                "Year": int(yr),
                "(smallest) Playoff tiebreaker": "N/A",
                "PF": float(g["PF"].sum()),
                "Avg PF": float(g["PF"].mean()) if g["PF"].notna().any() else None,
                "PF Range": float(g["PF"].max() - g["PF"].min()) if g["PF"].notna().any() else None,
                "Avg margin": float(g.loc[win_mask, "Margin"].mean()) if win_mask.any() else None,
                "Margin range": float(g.loc[win_mask, "Margin"].max() - g.loc[win_mask, "Margin"].min()) if win_mask.any() else None,
                "Number of games within 10": int((margin_abs <= 10).sum() / 2),
                "Number of games within 5": int((margin_abs <= 5).sum() / 2),
                "Max PF": float(g["Max PF"].sum()),
                "Avg max PF": float(g["Max PF"].mean()) if g["Max PF"].notna().any() else None,
                "Efficiency": float(g["Efficiency"].mean()) if g["Efficiency"].notna().any() else None,
                "Number of weeks missed due to injury": int(pd.to_numeric(g.get("Number of Injuries"), errors="coerce").fillna(0.0).sum()),
                "Number of weeks missed due to suspensions": int(pd.to_numeric(g.get("Number of suspensions"), errors="coerce").fillna(0.0).sum()),
                # League-year turnover = sum across all teams from team_year
                # (team_year is computed before league_year). Inseason starter
                # turnover was previously averaged from team_week, which is a
                # different thing; switch to team_year sum for consistency.
                "Inseason starter turnover": 0,  # filled from team_year below
                "Offseason starter turnover": 0,  # filled from team_year below
                "Inseason roster turnover": 0,    # filled from team_year below
                "Offseason roster turnover": 0,   # filled from team_year below
                "UPST": int(pd.to_numeric(g.get("UPST"), errors="coerce").fillna(0.0).sum()),
                # League-year Tanking = mean of weekly league Tanking.
                "Tanking": float(pd.to_numeric(g.get("Tanking"), errors="coerce").dropna().mean() or 0.0),
                "Luck": float(pd.to_numeric(g.get("Luck"), errors="coerce").fillna(0.0).sum()),
                "Increase in points from previous week": float(pd.to_numeric(g.get("Increase in points from previous week"), errors="coerce").fillna(0.0).sum()),
                # Unique-player position counts (Phase 1B, item 5): distinct
                # league-wide players started / rostered this year.
                **{
                    f"Number of {_pos} {_kind}": int(unique_pos_by_year.get((int(yr),), {}).get(f"Number of {_pos} {_kind}", 0))
                    for _pos in ["QB", "WR", "RB", "TE"]
                    for _kind in ["started", "rostered"]
                },
                "Amount of FAAB spent": float(pd.to_numeric(g.get("Amount of FAAB spent"), errors="coerce").fillna(0.0).sum()),
            })
        league_year = pd.DataFrame(rows)

        # Fill additional league-year columns from team-week
        try:
            agg_ly = g_week.groupby(["Year"], as_index=False).agg(
                **{
                    "Most number of players started from same NFL team": ("Most number of players started from same NFL team", "max"),
                    "Most number of players rostered from same NFL team": ("Most number of players rostered from same NFL team", "max"),
                    "Most number of QBs started from same NFL team": ("Most number of QBs started from same NFL team", "max"),
                    "Most number of QBs rostered from same NFL team": ("Most number of QBs rostered from same NFL team", "max"),
                    "Most number of RBs started from same NFL team": ("Most number of RBs started from same NFL team", "max"),
                    "Most number of RBs rostered from same NFL team": ("Most number of RBs rostered from same NFL team", "max"),
                    "Most number of WR started from same NFL team": ("Most number of WR started from same NFL team", "max"),
                    "Most number of WR rostered from same NFL team": ("Most number of WR rostered from same NFL team", "max"),
                    "Most number of TE started from same NFL team": ("Most number of TE started from same NFL team", "max"),
                    "Most number of TE rostered from same NFL team": ("Most number of TE rostered from same NFL team", "max"),
                    "Number of NFL teams among starting players": ("Number of NFL teams among starting players", "max"),
                    "Number of NFL teams among rostered players": ("Number of NFL teams among rostered players", "max"),
                    "Number of rookies started": ("Number of rookies started", "sum"),
                    "Number of rookies rostered": ("Number of rookies rostered", "sum"),
                    "Player average age": ("Player average age", "mean"),
                    "Team age including picks": ("Team age including picks", "mean"),
                    "Difference between highest and lowest starters": ("Difference between highest and lowest starters", "max"),
                    "Number of donuts": ("Number of donuts", "sum"),
                    "Number of starting donuts": ("Number of starter donuts", "sum"),
                    "Number of players under 10": ("Number of players under 10", "sum"),
                    "Number of players over 20": ("Number of players over 20", "sum"),
                    "Number of players over 30": ("Number of players over 30", "sum"),
                    "Number of players over 40": ("Number of players over 40", "sum"),
                    "Number of players over 50": ("Number of players over 50", "sum"),
                    "Number of cuffs rostered": ("Number of cuffs rostered", "sum"),
                    "Number of cuffs started": ("Number of cuffs started", "sum"),
                    "Startup draft players remaining": ("Startup draft players remaining", "max"),
                }
            )
            league_year = league_year.merge(agg_ly, how="left", on=["Year"])
            league_year["Amount of FAAB spent"] = pd.to_numeric(league_year.get("Amount of FAAB spent"), errors="coerce").fillna(0.0)
        except Exception as e:
            _log_exc(debug, "league_year_fill_extra", e)

        # Roll up turnover + transactions/trades from team_year into league_year.
        # These were hardcoded to 0 even though the team_year aggregation produces
        # real values per team.
        try:
            if not team_year.empty:
                ty_agg = team_year.groupby("Year", as_index=False).agg(
                    **{
                        "Offseason starter turnover": ("Offseason starter turnover", "sum"),
                        "Inseason starter turnover": ("Inseason starter turnover", "sum"),
                        "Offseason roster turnover": ("Offseason roster turnover", "sum"),
                        "Inseason roster turnover": ("Inseason roster turnover", "sum"),
                        "_ty_tx": ("Number of transactions", "sum"),
                        "_ty_tr": ("Number of trades", "sum"),
                    }
                )
                league_year = league_year.drop(
                    columns=[
                        "Offseason starter turnover",
                        "Inseason starter turnover",
                        "Offseason roster turnover",
                        "Inseason roster turnover",
                    ],
                    errors="ignore",
                ).merge(ty_agg, on="Year", how="left")
                # Number of transactions / trades roll up from team_year too —
                # league_year had these at 0 before despite team_year being correct.
                if "_ty_tx" in league_year.columns:
                    league_year["Number of transactions"] = pd.to_numeric(league_year["_ty_tx"], errors="coerce").fillna(0).astype(int)
                if "_ty_tr" in league_year.columns:
                    league_year["Number of trades"] = pd.to_numeric(league_year["_ty_tr"], errors="coerce").fillna(0).astype(int)
                league_year.drop(columns=["_ty_tx", "_ty_tr"], inplace=True, errors="ignore")
        except Exception as e:
            _log_exc(debug, "league_year_team_year_rollup", e)

        # UNIQUE cuff players league-wide per year (item 9), overriding the sum.
        try:
            if not league_year.empty and unique_cuffs_by_year and "Year" in league_year.columns:
                for _col in ["Number of cuffs rostered", "Number of cuffs started"]:
                    if _col in league_year.columns:
                        league_year[_col] = [
                            int(unique_cuffs_by_year.get((int(y),), {}).get(_col, 0)) if pd.notna(y) else 0
                            for y in pd.to_numeric(league_year["Year"], errors="coerce")
                        ]
        except Exception as e:
            _log_exc(debug, "league_year_unique_cuffs", e)

        # League-wide unique rookies / NFL-teams per year + distinct trades
        # (Phase 5B items 1, 2), overriding the summed / max placeholders.
        try:
            if not league_year.empty and "Year" in league_year.columns:
                _yrs = pd.to_numeric(league_year["Year"], errors="coerce")
                for _col in _LX_KEYS:
                    if _col in league_year.columns:
                        league_year[_col] = [
                            int(league_extra_by_year.get((int(y),), {}).get(_col, 0)) if pd.notna(y) else 0
                            for y in _yrs
                        ]
                for _tc, _bk in [("Offseason trades", "off"), ("Inseason trades", "in"), ("Total trades", "tot")]:
                    league_year[_tc] = [
                        int(len(_league_trade_dates_split.get(int(y), {}).get(_bk, set()))) if pd.notna(y) else 0
                        for y in _yrs
                    ]
                league_year["Highest starter score"] = [_star_hi_y.get(int(y)) if pd.notna(y) else None for y in _yrs]
                league_year["Lowest starter score"] = [_star_lo_y.get(int(y)) if pd.notna(y) else None for y in _yrs]
                # League range = highest − lowest starter (item-5 disambiguation).
                league_year["Difference between highest and lowest starters"] = [
                    (float(_star_hi_y[int(y)] - _star_lo_y[int(y)])
                     if (pd.notna(y) and int(y) in _star_hi_y and int(y) in _star_lo_y) else None)
                    for y in _yrs
                ]
        except Exception as e:
            _log_exc(debug, "league_year_unique_extras", e)

        league_all = pd.DataFrame([{
            "PF": float(pd.to_numeric(g_week["PF"], errors="coerce").fillna(0.0).sum()),
            "Avg PF": float(pd.to_numeric(g_week["PF"], errors="coerce").fillna(0.0).mean()),
            "PF Range": float(g_week["PF"].max() - g_week["PF"].min()) if g_week["PF"].notna().any() else None,
            "Avg margin": (float(g_week.loc[pd.to_numeric(g_week.get("Win?"), errors="coerce") == 1, "Margin"].mean()) if (pd.to_numeric(g_week.get("Win?"), errors="coerce") == 1).any() else None),
            "Margin range": (float(g_week.loc[pd.to_numeric(g_week.get("Win?"), errors="coerce") == 1, "Margin"].max() - g_week.loc[pd.to_numeric(g_week.get("Win?"), errors="coerce") == 1, "Margin"].min()) if (pd.to_numeric(g_week.get("Win?"), errors="coerce") == 1).any() else None),
            "Number of games within 10": int((g_week["Margin"].abs() <= 10).sum() / 2),
            "Number of games within 5": int((g_week["Margin"].abs() <= 5).sum() / 2),
            "Max PF": float(pd.to_numeric(g_week["Max PF"], errors="coerce").fillna(0.0).sum()),
            "Avg max PF": float(pd.to_numeric(g_week["Max PF"], errors="coerce").fillna(0.0).mean()),
            "Efficiency": float(pd.to_numeric(g_week["Efficiency"], errors="coerce").dropna().mean()) if g_week["Efficiency"].notna().any() else None,
            "Number of weeks missed due to injury": int(pd.to_numeric(g_week.get("Number of Injuries"), errors="coerce").fillna(0.0).sum()),
            "Number of weeks missed due to suspensions": int(pd.to_numeric(g_week.get("Number of suspensions"), errors="coerce").fillna(0.0).sum()),
            "UPST": int(pd.to_numeric(g_week.get("UPST"), errors="coerce").fillna(0.0).sum()),
            # League-all-time Tanking = mean across all weeks.
            "Tanking": float(pd.to_numeric(g_week.get("Tanking"), errors="coerce").dropna().mean() or 0.0),
            "Luck": float(pd.to_numeric(g_week.get("Luck"), errors="coerce").fillna(0.0).sum()),
            # Unique-player position counts (Phase 1B, item 5): distinct
            # players started / rostered league-wide across all years.
            **{
                f"Number of {_pos} {_kind}": int(unique_pos_league_all.get(f"Number of {_pos} {_kind}", 0))
                for _pos in ["QB", "WR", "RB", "TE"]
                for _kind in ["started", "rostered"]
            },
            "Number of transactions": int(pd.to_numeric(g_week.get("Number of transactions"), errors="coerce").fillna(0.0).sum()),
            "Number of trades": int(pd.to_numeric(g_week.get("Number of trades"), errors="coerce").fillna(0.0).sum()),
            "Amount of FAAB spent": float(pd.to_numeric(g_week.get("Amount of FAAB spent"), errors="coerce").fillna(0.0).sum()),
            "Most number of players started from same NFL team": float(pd.to_numeric(g_week.get("Most number of players started from same NFL team"), errors="coerce").fillna(0.0).max()),
            "Most number of players rostered from same NFL team": float(pd.to_numeric(g_week.get("Most number of players rostered from same NFL team"), errors="coerce").fillna(0.0).max()),
            "Most number of QBs started from same NFL team": float(pd.to_numeric(g_week.get("Most number of QBs started from same NFL team"), errors="coerce").fillna(0.0).max()),
        }])

        # Fill additional league-all-time columns from team-week
        try:
            league_all["Most number of players started from same NFL team"] = float(pd.to_numeric(g_week.get("Most number of players started from same NFL team"), errors="coerce").max())
            league_all["Most number of players rostered from same NFL team"] = float(pd.to_numeric(g_week.get("Most number of players rostered from same NFL team"), errors="coerce").max())
            league_all["Most number of QBs started from same NFL team"] = float(pd.to_numeric(g_week.get("Most number of QBs started from same NFL team"), errors="coerce").max())
            league_all["Most number of QBs rostered from same NFL team"] = float(pd.to_numeric(g_week.get("Most number of QBs rostered from same NFL team"), errors="coerce").max())
            league_all["Most number of RBs started from same NFL team"] = float(pd.to_numeric(g_week.get("Most number of RBs started from same NFL team"), errors="coerce").max())
            league_all["Most number of RBs rostered from same NFL team"] = float(pd.to_numeric(g_week.get("Most number of RBs rostered from same NFL team"), errors="coerce").max())
            league_all["Most number of WR started from same NFL team"] = float(pd.to_numeric(g_week.get("Most number of WR started from same NFL team"), errors="coerce").max())
            league_all["Most number of WR rostered from same NFL team"] = float(pd.to_numeric(g_week.get("Most number of WR rostered from same NFL team"), errors="coerce").max())
            league_all["Most number of TE started from same NFL team"] = float(pd.to_numeric(g_week.get("Most number of TE started from same NFL team"), errors="coerce").max())
            league_all["Most number of TE rostered from same NFL team"] = float(pd.to_numeric(g_week.get("Most number of TE rostered from same NFL team"), errors="coerce").max())
            # League-wide UNIQUE across all time (Phase 5B item 2): distinct
            # rookie players and distinct NFL teams, not the weekly max / sum.
            league_all["Number of NFL teams among starting players"] = int(league_extra_all.get("Number of NFL teams among starting players", 0))
            league_all["Number of NFL teams among rostered players"] = int(league_extra_all.get("Number of NFL teams among rostered players", 0))
            league_all["Number of rookies started"] = int(league_extra_all.get("Number of rookies started", 0))
            league_all["Number of rookies rostered"] = int(league_extra_all.get("Number of rookies rostered", 0))
            league_all["Player average age"] = float(pd.to_numeric(g_week.get("Player average age"), errors="coerce").mean())
            league_all["Team age including picks"] = float(pd.to_numeric(g_week.get("Team age including picks"), errors="coerce").mean())
            # League "Difference between highest and lowest starters" = the
            # league-wide RANGE (highest starter − lowest starter), so it
            # reconciles with the Highest/Lowest starter score columns (Phase
            # 5C item-5 disambiguation; was the max single-team spread).
            league_all["Highest starter score"] = _star_hi_all
            league_all["Lowest starter score"] = _star_lo_all
            league_all["Difference between highest and lowest starters"] = (
                float(_star_hi_all - _star_lo_all) if (_star_hi_all is not None and _star_lo_all is not None) else None
            )
            league_all["Number of donuts"] = float(pd.to_numeric(g_week.get("Number of donuts"), errors="coerce").sum())
            league_all["Number of starting donuts"] = float(pd.to_numeric(g_week.get("Number of starter donuts"), errors="coerce").sum())
            league_all["Number of players under 10"] = float(pd.to_numeric(g_week.get("Number of players under 10"), errors="coerce").sum())
            league_all["Number of players over 20"] = float(pd.to_numeric(g_week.get("Number of players over 20"), errors="coerce").sum())
            league_all["Number of players over 30"] = float(pd.to_numeric(g_week.get("Number of players over 30"), errors="coerce").sum())
            league_all["Number of players over 40"] = float(pd.to_numeric(g_week.get("Number of players over 40"), errors="coerce").sum())
            league_all["Number of players over 50"] = float(pd.to_numeric(g_week.get("Number of players over 50"), errors="coerce").sum())
            # UNIQUE cuff players league-wide all-time (item 9), overriding the
            # player-week sum.
            league_all["Number of cuffs rostered"] = int(unique_cuffs_league_all.get("Number of cuffs rostered", 0))
            league_all["Number of cuffs started"] = int(unique_cuffs_league_all.get("Number of cuffs started", 0))
            league_all["Startup draft players remaining"] = float(pd.to_numeric(g_week.get("Startup draft players remaining"), errors="coerce").max())
            league_all["Amount of FAAB spent"] = float(pd.to_numeric(g_week.get("Amount of FAAB spent"), errors="coerce").fillna(0.0).sum())
            # Distinct trades all-time (Phase 5B item 1).
            # Offseason / Inseason / Total trades all-time = sum across seasons
            # (each trade lives in exactly one season).
            league_all["Offseason trades"] = int(sum(len(v["off"]) for v in _league_trade_dates_split.values()))
            league_all["Inseason trades"] = int(sum(len(v["in"]) for v in _league_trade_dates_split.values()))
            league_all["Total trades"] = int(sum(len(v["tot"]) for v in _league_trade_dates_split.values()))
        except Exception as e:
            _log_exc(debug, "league_all_fill_extra", e)


    # --------------------------
    # Write outputs (schema contract)
    # --------------------------
    
    # --------------------------
    # Transactions / Trades: link columns + tanking (best-effort)
    # --------------------------
    # Tanking-delta (Phase 6E): the marginal change in a team's Tanking score
    # caused by a single transaction/trade, holding everything else constant.
    #   Tanking = (1/6)t1(PF)+(1/6)t2(MaxPF)+(1/6)t3(age)+(1/6)t4(curr picks)
    #             +(1/9)t5(future cap)
    # PF / MaxPF (t1,t2) and the rest of the roster are held fixed, so only the
    # age term and (for trades) future draft capital move:
    #     ΔTanking = (1/6)·Δt3 + (1/9)·Δfuture_cap
    #   Δt3         = −(avg_age_post − avg_age_pre)/(L_age − 21)
    #   avg_age      = "Team age including picks" (entities = rostered players +
    #                  future picks). avg_age_pre = A (the team's roster age that
    #                  week), N = that week's entity count, and
    #                  avg_age_post = (N·A − Σsent_age + Σrecv_age)/(N − k_sent + k_recv).
    #   Δfuture_cap  = round-weighted future picks received − sent (trades only;
    #                  waiver/FA transactions never move picks). Current-year
    #                  rookie picks aren't tradeable mid-season, so t4 ≈ 0.
    # Net sign: getting younger and/or richer in picks ⇒ positive (more
    # tanking); dealing picks/youth for win-now talent ⇒ negative.
    _TANK_A: Dict[Tuple[str, int, int], float] = {}
    _TANK_N: Dict[Tuple[str, int, int], float] = {}
    _TANK_LAGE: Dict[int, float] = {}
    try:
        _agecol = "Team age including picks"
        if not tw.empty and {"Team", "Year", "Week"}.issubset(tw.columns) and _agecol in tw.columns:
            _has_n = "_RosterNInclPicks" in tw.columns
            for _, _r in tw.dropna(subset=["Team", "Year", "Week"]).iterrows():
                try:
                    _k = (str(_r["Team"]), int(_r["Year"]), int(_r["Week"]))
                except Exception:
                    continue
                _a = pd.to_numeric(pd.Series([_r.get(_agecol)]), errors="coerce").iloc[0]
                if pd.notna(_a):
                    _TANK_A[_k] = float(_a)
                if _has_n:
                    _n = pd.to_numeric(pd.Series([_r.get("_RosterNInclPicks")]), errors="coerce").iloc[0]
                    if pd.notna(_n) and _n > 0:
                        _TANK_N[_k] = float(_n)
            # League average roster age (incl picks) per season → baseline for
            # the age term, matching the tank formula's L_age denominator.
            _la = tw.copy()
            _la["_AGE"] = pd.to_numeric(_la[_agecol], errors="coerce")
            for _yr, _gg in _la.dropna(subset=["_AGE"]).groupby("Year"):
                try:
                    _TANK_LAGE[int(_yr)] = float(_gg["_AGE"].mean())
                except Exception:
                    continue
    except Exception as e:
        _log_exc(debug, "tanking_delta_maps", e)

    def _tanking_delta_row(team, season_i, wk_i, recv_sum, recv_n, sent_sum, sent_n, fcap_delta) -> float:
        """Marginal ΔTanking for one transaction/trade (see block header)."""
        try:
            si = int(season_i); wi = int(wk_i)
            A = _TANK_A.get((str(team), si, wi))
            N = _TANK_N.get((str(team), si, wi))
            if A is None or N is None:
                for wk_try in (wi - 1, wi + 1):  # adjacent-week fallback
                    if A is None:
                        A = _TANK_A.get((str(team), si, wk_try))
                    if N is None:
                        N = _TANK_N.get((str(team), si, wk_try))
            dt3 = 0.0
            if A is not None and N is not None:
                count_post = float(N) - float(sent_n) + float(recv_n)
                if count_post > 0:
                    avg_post = (float(N) * float(A) - float(sent_sum) + float(recv_sum)) / count_post
                    B = _TANK_LAGE.get(si)
                    if B is not None:
                        B = float(B) - 21.0
                        if abs(B) > 1e-9:
                            dt3 = -(avg_post - float(A)) / B
            return round((1.0 / 6.0) * dt3 + (1.0 / 9.0) * float(fcap_delta or 0.0), 4)
        except Exception:
            return 0.0

    try:
        # normalize date
        if not tx.empty and "Date" in tx.columns:
            # format='ISO8601' parses each value independently against
            # any ISO-8601 variant rather than inferring one format
            # across the column. Without it, pandas locks onto the
            # first row's format (Sleeper's microsecond-precision
            # space-separated form) and silently coerces any other
            # ISO variant to NaT — which dropped the manual Puka row
            # whose Date was the cleaner 'YYYY-MM-DDTHH:MM:SS+ZZ:ZZ'.
            tx["Date"] = pd.to_datetime(tx["Date"], errors="coerce", utc=True, format="ISO8601")
            # Drop rows that couldn't be anchored to a real timestamp. The
            # orphan-drop guard in the source loop only catches the case
            # where both created_dt and created_date are None; values that
            # stringify but fail pd.to_datetime (e.g., 'None'/'NaT'/garbled
            # epochs) still produce NaT here. Those rows can't be ordered
            # for link-prev/next or year-tanking joins, and would otherwise
            # land in transactions.csv with Date='N/A'.
            tx = tx[tx["Date"].notna()].reset_index(drop=True)
            tx = tx.sort_values(["Team","Date"]).reset_index(drop=True)
            # link columns as row numbers (1-indexed within each team).
            # 'Link to previous' on row 1 is NaN, otherwise points at row N-1.
            # 'Link to next' on the last row of each team is NaN, otherwise
            # points at row N+1. The previous formula used shift(-1)+2 which
            # was off-by-one (row 1 got linked to row 3 instead of row 2).
            tx["Link to previous transaction"] = tx.groupby("Team").cumcount().replace(0, np.nan)
            tx["Link to next transaction"] = tx.groupby("Team").cumcount() + 2
            tx.loc[tx.groupby("Team").tail(1).index, "Link to next transaction"] = np.nan

            # Tanking (Phase 6E) — marginal ΔTanking from THIS transaction,
            # holding all else constant (see the tanking-delta block header).
            # Waiver/FA transactions only shift the roster's age term; the
            # week the move landed in supplies the team's roster age (A) and
            # entity count (N). Week derived from Date: NFL Week 1 ~Sept 7,
            # week N starts 7*(N-1) days later; pre-Wk1 floors to 1, post-Wk17
            # caps at 17.
            if not tw.empty and "Season" in tx.columns:
                def _week_for_tx(dt_obj, season_val) -> int:
                    try:
                        d = dt_obj.date() if hasattr(dt_obj, "date") else dt_obj
                        season_start = date(int(season_val), 9, 7)
                        if d < season_start:
                            return 1
                        wk = (d - season_start).days // 7 + 1
                        return max(1, min(17, int(wk)))
                    except Exception:
                        return 1

                def _tcol(name):
                    return (pd.to_numeric(tx[name], errors="coerce").fillna(0.0)
                            if name in tx.columns else pd.Series(0.0, index=tx.index))
                _rs = _tcol("_tank_recv_age_sum"); _rn = _tcol("_tank_recv_n")
                _ss = _tcol("_tank_sent_age_sum"); _sn = _tcol("_tank_sent_n")
                _fc = _tcol("_tank_fcap_delta")
                vals: List[float] = []
                for i, (t, s, d) in enumerate(zip(tx["Team"], tx["Season"], tx["Date"])):
                    try:
                        season_i = int(s)
                    except Exception:
                        vals.append(0.0); continue
                    wk_i = _week_for_tx(d, season_i)
                    vals.append(_tanking_delta_row(
                        t, season_i, wk_i,
                        _rs.iloc[i], _rn.iloc[i], _ss.iloc[i], _sn.iloc[i], _fc.iloc[i],
                    ))
                tx["Tanking"] = vals
        else:
            if "Tanking" in tx.columns:
                tx["Tanking"] = pd.to_numeric(tx["Tanking"], errors="coerce").fillna(0.0)
    except Exception as e:
        _log_exc(debug, "transactions_links_tanking", e)

    try:
        if not tr.empty and "Date" in tr.columns:
            tr["Date"] = pd.to_datetime(tr["Date"], errors="coerce", utc=True, format="ISO8601")
            # Same guard as transactions: drop rows we can't anchor in time.
            tr = tr[tr["Date"].notna()].reset_index(drop=True)
            tr = tr.sort_values(["Team","Date"]).reset_index(drop=True)
            # Per-asset link columns ("Link to next/previous transaction per
            # asset") are filled later in the player-chain block (Phase 7B),
            # which reuses the same cross-tx/trade chains as the transaction
            # links. The old per-team cumcount links were removed there.

            # Trades (Phase 6E): marginal ΔTanking from the picks/players that
            # changed hands — same delta model as transactions, but trades also
            # move the future-capital term.
            if not tw.empty and "Season" in tr.columns:
                def _week_for_tr(dt_obj, season_val) -> int:
                    try:
                        d = dt_obj.date() if hasattr(dt_obj, "date") else dt_obj
                        season_start = date(int(season_val), 9, 7)
                        if d < season_start:
                            return 1
                        wk = (d - season_start).days // 7 + 1
                        return max(1, min(17, int(wk)))
                    except Exception:
                        return 1

                def _trcol(name):
                    return (pd.to_numeric(tr[name], errors="coerce").fillna(0.0)
                            if name in tr.columns else pd.Series(0.0, index=tr.index))
                _rs = _trcol("_tank_recv_age_sum"); _rn = _trcol("_tank_recv_n")
                _ss = _trcol("_tank_sent_age_sum"); _sn = _trcol("_tank_sent_n")
                _fc = _trcol("_tank_fcap_delta")
                vals: List[float] = []
                for i, (t, s, d) in enumerate(zip(tr["Team"], tr["Season"], tr["Date"])):
                    try:
                        season_i = int(s)
                    except Exception:
                        vals.append(0.0); continue
                    wk_i = _week_for_tr(d, season_i)
                    vals.append(_tanking_delta_row(
                        t, season_i, wk_i,
                        _rs.iloc[i], _rn.iloc[i], _ss.iloc[i], _sn.iloc[i], _fc.iloc[i],
                    ))
                tr["Tanking"] = vals
    except Exception as e:
        _log_exc(debug, "trades_links_tanking", e)

    # Player-chain links (Phase 6D): split the transaction links into chains
    # that follow the ADDED player and the DROPPED player across every later /
    # earlier event involving them — transactions AND trades. References are
    # "#N" = transactions.csv row N, "T#N" = trades.csv row N (1-indexed, final
    # sorted order). Both tx and tr are fully sorted/indexed by this point.
    try:
        if not tx.empty:
            # A real player name — excludes blanks and every "absent" sentinel
            # the data uses. Critically, pure-drop rows carry Player Added as
            # NaN/None (rendered "N/A" only at write time), so str() yields
            # "nan"/"None"; without filtering those, every no-add row collapses
            # into one phantom chains["nan"] bucket and links to each other.
            def _real_player(_v):
                _s = str(_v).strip()
                return bool(_s) and _s.lower() not in ("nan", "none", "n/a") and _s != "0.0"

            def _is_player_asset(_a):
                _s = str(_a).strip()
                # Exclude pick labels (YYYY ...) and FAAB ("$N FAAB").
                return _real_player(_a) and not re.match(r"^\d{4}\b", _s) and not _s.endswith("FAAB")

            chains: Dict[str, List[Tuple[Any, str]]] = defaultdict(list)
            for _i in tx.index:
                _ref = f"#{int(_i) + 1}"; _d = tx.at[_i, "Date"]
                _add = str(tx.at[_i, "Player Added"]).strip(); _drop = str(tx.at[_i, "Player Dropped"]).strip()
                if _real_player(_add):
                    chains[_add].append((_d, _ref))
                if _real_player(_drop):
                    chains[_drop].append((_d, _ref))
            if not tr.empty:
                for _j in tr.index:
                    _ref = f"T#{int(_j) + 1}"; _d = tr.at[_j, "Date"]
                    for _col in ("Assets received", "Assets sent"):
                        for _a in str(tr.at[_j, _col] or "").split(";"):
                            if _is_player_asset(_a):
                                chains[_a.strip()].append((_d, _ref))
            _nat = pd.Timestamp.min.tz_localize("UTC")
            for _p in chains:
                chains[_p].sort(key=lambda e: (e[0] if pd.notna(e[0]) else _nat))

            # Map each ref to its EVENT. Every transaction is its own event; all
            # rows of one trade (a 3-team trade emits several) share the trade's
            # _tx_id. A player's next/previous link then skips the OTHER rows of
            # the same trade it's already on and lands on the next DISTINCT
            # transaction/trade that involves them.
            _ref_event: Dict[str, str] = {f"#{int(_i) + 1}": f"X{int(_i) + 1}" for _i in tx.index}
            if not tr.empty and "_tx_id" in tr.columns:
                for _j in tr.index:
                    _ref_event[f"T#{int(_j) + 1}"] = f"E{tr.at[_j, '_tx_id']}"

            def _neighbors(player, this_ref):
                ch = chains.get(player, [])
                _pos = next((_k for _k in range(len(ch)) if ch[_k][1] == this_ref), None)
                if _pos is None:
                    return None, None
                _ev = _ref_event.get(this_ref)
                _nxt = next((ch[_k][1] for _k in range(_pos + 1, len(ch))
                             if _ref_event.get(ch[_k][1]) != _ev), None)
                _prv = next((ch[_k][1] for _k in range(_pos - 1, -1, -1)
                             if _ref_event.get(ch[_k][1]) != _ev), None)
                return _nxt, _prv

            # Pick chains: follow a draft pick to the next / previous TRADE that
            # moved it, keyed by its canonical identity (year, round, original
            # owner) — NOT the display label, whose projected slot can drift for
            # an un-drafted future pick. Built from the RECEIVED side only so the
            # two mirror rows of a single trade event don't link to each other.
            # Deliberately does NOT bridge a pick to the player drafted with it.
            pick_chains: Dict[Tuple[int, int, str], List[Tuple[Any, str]]] = defaultdict(list)
            # Bug #6: a pick's HOME row on the picks sheet, by canonical key. Used
            # as the fallback link when a pick has no earlier trade (its first
            # trade's "previous") so the cell links to the pick's origin instead
            # of dead "N/A". Populated in the draft-row bridging loop below.
            pick_home_phref: Dict[Tuple[int, int, str], str] = {}
            if "_recv_pick_meta" in tr.columns:
                for _j in tr.index:
                    _ref = f"T#{int(_j) + 1}"; _d = tr.at[_j, "Date"]
                    _pm = tr.at[_j, "_recv_pick_meta"]
                    for _m in (_pm if isinstance(_pm, list) else []):
                        try:
                            _key = (int(_m[0]), int(_m[1]), str(_m[2]))
                        except Exception:
                            continue
                        pick_chains[_key].append((_d, _ref))
                for _k in pick_chains:
                    pick_chains[_k].sort(key=lambda e: (e[0] if pd.notna(e[0]) else _nat))

            # Draft-row bridging: connect the pick chain and the player chain
            # through the pick_history DRAFT row, without crossing them. A
            # pick's chain TERMINATES at its draft row (its last trade's "next"
            # -> the draft); the drafted player's chain STARTS at the same draft
            # row (their first event's "previous" -> the draft). Reference
            # "PH#N" = picks.csv row N (ph keeps its build order through
            # output, so the index is stable). Anchor date = late August of the
            # draft year: after the offseason pick trades, before the rookie
            # season's events, so it sorts last for the pick and first for the
            # player.
            if not ph.empty and {"Year", "Number", "Original Team", "Player Picked"}.issubset(set(ph.columns)):
                for _pi in ph.index:
                    _phref = f"PH#{int(_pi) + 1}"
                    _ym = re.match(r"\s*(\d{4})", str(ph.at[_pi, "Year"]))
                    _nm = re.match(r"\s*(\d+)\.", str(ph.at[_pi, "Number"]))
                    if not _ym or not _nm:
                        continue
                    _yr = int(_ym.group(1)); _rd = int(_nm.group(1))
                    _orig = str(ph.at[_pi, "Original Team"]).strip()
                    # Anchor each PH# RELATIVE to the chain it joins, not a fixed
                    # calendar date (which mis-orders the 2021 startup/vet draft
                    # and draft-day pick trades). For a pick the draft is its
                    # TERMINAL, so anchor just after its last trade; for a player
                    # the draft is the START, so anchor just before their first
                    # event. Fallback to late-August when the chain is empty
                    # (no other event references it then anyway). The date is
                    # only used for sort order — the "PH#N" ref is identical.
                    _fallback = pd.Timestamp(year=_yr, month=8, day=28, tz="UTC")
                    _pk = (_yr, _rd, _orig)
                    # The 2021 startup/vet draft shares its (year, round, orig)
                    # canonical key with the 2021 ROOKIE draft, so adding a vet
                    # PH# terminal here would pollute the rookie pick's chain
                    # (its "previous" would land on the vet draft row). Vet picks
                    # were never traded — they have no pick chain — so skip the
                    # terminal for them. Their drafted player's chain start is
                    # still anchored below.
                    _is_vet = "(vet)" in str(ph.at[_pi, "Year"]).lower()
                    if not _is_vet:
                        _pdates = [e[0] for e in pick_chains.get(_pk, []) if pd.notna(e[0])]
                        _term = (max(_pdates) + pd.Timedelta(days=1)) if _pdates else _fallback
                        pick_chains[_pk].append((_term, _phref))  # pick terminal
                        pick_home_phref[_pk] = _phref  # pick's home row (Bug #6)
                    _pl = str(ph.at[_pi, "Player Picked"]).strip()
                    if _real_player(_pl):
                        _edates = [e[0] for e in chains.get(_pl, []) if pd.notna(e[0])]
                        _start = (min(_edates) - pd.Timedelta(days=1)) if _edates else _fallback
                        chains[_pl].append((_start, _phref))  # player chain start
                _ksort = lambda e: (e[0] if pd.notna(e[0]) else _nat)
                for _k in pick_chains:
                    pick_chains[_k].sort(key=_ksort)
                for _p in chains:
                    chains[_p].sort(key=_ksort)

            def _pick_neighbors(key, this_ref):
                ch = pick_chains.get(key, [])
                for _k in range(len(ch)):
                    if ch[_k][1] == this_ref:
                        return (ch[_k + 1][1] if _k + 1 < len(ch) else None,
                                ch[_k - 1][1] if _k > 0 else None)
                return None, None

            a_next, a_prev, d_next, d_prev = [], [], [], []
            for _i in tx.index:
                _ref = f"#{int(_i) + 1}"
                _add = str(tx.at[_i, "Player Added"]).strip(); _drop = str(tx.at[_i, "Player Dropped"]).strip()
                if _real_player(_add):
                    _n, _p = _neighbors(_add, _ref); a_next.append(_n); a_prev.append(_p)
                else:
                    a_next.append(None); a_prev.append(None)
                if _real_player(_drop):
                    _n, _p = _neighbors(_drop, _ref); d_next.append(_n); d_prev.append(_p)
                else:
                    d_next.append(None); d_prev.append(None)
            tx["Link to next transaction (added player)"] = a_next
            tx["Link to previous transaction (added player)"] = a_prev
            tx["Link to next transaction (dropped player)"] = d_next
            tx["Link to previous transaction (dropped player)"] = d_prev

            # Per-asset trade links (Phase 7B + pick chains): for each asset
            # RECEIVED in a trade, follow it to its next / previous event as a
            # ';'-joined list aligned 1:1 with "Assets received". Players resolve
            # through the player chain (transaction "#N" / trade "T#N"); draft
            # picks resolve through the pick chain to the next/prev TRADE that
            # moved that pick (canonical id from _recv_pick_meta, consumed in the
            # same order the picks appear in "Assets received"). FAAB → "N/A".
            if not tr.empty:
                tr_next_pa, tr_prev_pa = [], []
                for _j in tr.index:
                    _ref = f"T#{int(_j) + 1}"
                    _nexts, _prevs = [], []
                    _pm = tr.at[_j, "_recv_pick_meta"] if "_recv_pick_meta" in tr.columns else None
                    _pm = _pm if isinstance(_pm, list) else []
                    _pi = 0
                    for _a in str(tr.at[_j, "Assets received"] or "").split(";"):
                        _a = _a.strip()
                        if not _a or _a.upper() == "N/A":
                            continue
                        if _is_player_asset(_a):
                            _n, _p = _neighbors(_a, _ref)
                            _nexts.append(_n or "N/A"); _prevs.append(_p or "N/A")
                        elif re.match(r"^\d{4}\b", _a):  # draft pick
                            _key = None
                            if _pi < len(_pm):
                                try:
                                    _key = (int(_pm[_pi][0]), int(_pm[_pi][1]), str(_pm[_pi][2]))
                                except Exception:
                                    _key = None
                            _pi += 1
                            if _key is not None:
                                _n, _p = _pick_neighbors(_key, _ref)
                                # Bug #6: with no earlier/later trade, fall back to
                                # the pick's own picks-sheet home row so the cell
                                # is never a dead "N/A" (FAAB is the only exception).
                                _home = pick_home_phref.get(_key)
                                _nexts.append(_n or _home or "N/A")
                                _prevs.append(_p or _home or "N/A")
                            else:
                                _nexts.append("N/A"); _prevs.append("N/A")
                        else:  # FAAB / other — not chainable
                            _nexts.append("N/A"); _prevs.append("N/A")
                    tr_next_pa.append("; ".join(_nexts) if _nexts else None)
                    tr_prev_pa.append("; ".join(_prevs) if _prevs else None)
                tr["Link to next transaction per asset"] = tr_next_pa
                tr["Link to previous transaction per asset"] = tr_prev_pa

            # Phase 8F: picks links. Bridge both chains through the draft row:
            #   Link to next transaction      = the drafted PLAYER's first event
            #                                   AFTER the draft (player chain,
            #                                   whose start is this PH# row)
            #   Link to previous transaction  = the PICK's last trade BEFORE the
            #                                   draft (pick chain, whose terminal
            #                                   is this PH# row)
            if not ph.empty and {"Year", "Number", "Original Team", "Player Picked"}.issubset(set(ph.columns)):
                _ph_next, _ph_prev = [], []
                for _pi in ph.index:
                    _phref = f"PH#{int(_pi) + 1}"
                    _pl = str(ph.at[_pi, "Player Picked"]).strip()
                    _ym = re.match(r"\s*(\d{4})", str(ph.at[_pi, "Year"]))
                    _nm = re.match(r"\s*(\d+)\.", str(ph.at[_pi, "Number"]))
                    _nx = _pv = None
                    if _ym and _nm:
                        _pk = (int(_ym.group(1)), int(_nm.group(1)),
                               str(ph.at[_pi, "Original Team"]).strip())
                        # previous = the pick's last trade before the draft
                        _, _pv = _pick_neighbors(_pk, _phref)
                    # next = the drafted player's first event after the draft
                    if _real_player(_pl):
                        _nx, _ = _neighbors(_pl, _phref)
                    _ph_next.append(_nx)
                    _ph_prev.append(_pv)
                ph["Link to next transaction"] = _ph_next
                ph["Link to previous transaction"] = _ph_prev
    except Exception as e:
        _log_exc(debug, "transactions_player_chain_links", e)

    # i5 (#15): build the full Sleeper-style asset history for hover-comments.
    # Every event on its own line, with the FULL deal/transaction detail; the
    # draft is its own event; a player's history begins with the history of the
    # pick they were drafted at.
    try:
        def _hd(_s):  # YYYY-MM-DD in US Eastern, matching the displayed sheets
            try:
                return pd.to_datetime(_s, utc=True).tz_convert("America/New_York").strftime("%Y-%m-%d")
            except Exception:
                return str(_s or "")[:10]
        def _cl(_s):
            return str(_s or "").strip()
        def _present(_s):
            return _cl(_s) and _cl(_s).lower() not in ("n/a", "nan")
        def _hts(_s):  # full Eastern timestamp — sort key so same-day events order by TIME
            try:
                return pd.to_datetime(_s, utc=True).tz_convert("America/New_York").strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                return str(_s or "")[:19]

        # Pick-trade hops keyed by canonical (year, round, original-owner name),
        # carrying (sort_ts, receiving team, full deal text), matched off each
        # trade's receiving side.
        _ph_hops: Dict[Tuple[int, int, str], List[Tuple[str, str, str]]] = defaultdict(list)
        for _tr in trades_rows:
            _tm = _cl(_tr.get("Team")); _d = _hd(_tr.get("Date")); _ts = _hts(_tr.get("Date"))
            _recv = _cl(_tr.get("Assets received")); _sent = _cl(_tr.get("Assets sent"))
            _deal = f"{_d}: pick traded to {_tm} ({_tm} got {_recv}; sent {_sent})"
            for _m in (_tr.get("_recv_pick_meta") or []):
                try:
                    _k = (int(_m[0]), int(_m[1]), str(_m[2]))
                except Exception:
                    continue
                _ph_hops[_k].append((_ts, _tm, _deal))
        _ph_trade_cols = [c for c in ph.columns
                          if str(c).startswith("Trade ") and str(c)[6:].isdigit()]

        # Per-player events (full detail), by Sleeper player_id, keyed by sort_ts.
        _pl_events: Dict[str, List[Tuple[str, str]]] = defaultdict(list)
        for _tx in transactions_rows:
            _tm = _cl(_tx.get("Team")); _d = _hd(_tx.get("Date")); _ts = _hts(_tx.get("Date"))
            _added = _cl(_tx.get("Player Added")); _dropped = _cl(_tx.get("Player Dropped"))
            _faab = _cl(_tx.get("Faab"))
            _how = f"waiver ${_faab}" if _present(_faab) else "free agent"
            if _tx.get("_added_pid"):
                _ex = f"; dropped {_dropped}" if _present(_dropped) else ""
                _pl_events[str(_tx["_added_pid"])].append((_ts, f"{_d}: added by {_tm} ({_how}{_ex})"))
            if _tx.get("_dropped_pid"):
                _ex = f" (added {_added})" if _present(_added) else ""
                _pl_events[str(_tx["_dropped_pid"])].append((_ts, f"{_d}: dropped by {_tm}{_ex}"))
        for _tr in trades_rows:
            _tm = _cl(_tr.get("Team")); _d = _hd(_tr.get("Date")); _ts = _hts(_tr.get("Date"))
            _recv = _cl(_tr.get("Assets received")); _sent = _cl(_tr.get("Assets sent"))
            _deal = f"{_d}: traded to {_tm} ({_tm} got {_recv}; sent {_sent})"
            for _pid in (_tr.get("_recv_player_ids") or []):
                if _pid:
                    _pl_events[str(_pid)].append((_ts, _deal))

        def _draft_anchor(_y) -> str:
            """Date (YYYY-MM-DD) the season's draft actually happened, so the draft
            EVENT sorts after the pick's pre-draft trades and before the drafted
            player's post-draft moves. Falls back to Aug 31 for not-yet-drafted
            seasons. Fixes draft-then-traded-same-period mis-ordering (Higgins)."""
            _dds = draft_dates_by_season.get(int(_y))
            if _dds:
                return max(_dds).strftime("%Y-%m-%d")
            return f"{int(_y)}-08-31"

        def _pick_hist_lines(_pi) -> Tuple[List[Tuple[str, str]], int]:
            """(sorted lines, number_of_trades) for the pick at ph row index _pi:
            original owner -> commissioner moves -> recorded trades -> draft. The
            trade count includes off-platform commissioner moves (real trades
            Sleeper missed beyond its 3-yr window) but NOT the toilet/FAAB award."""
            _yr_disp = str(ph.at[_pi, "Year"])
            _ym = re.match(r"\s*(\d{4})", _yr_disp)
            if not _ym:
                return [], 0
            _yr = int(_ym.group(1)); _num = _cl(ph.at[_pi, "Number"])
            _orig = _cl(ph.at[_pi, "Original Team"])
            _final = (_cl(ph.at[_pi, "Final Team"]) if "Final Team" in ph.columns
                      else (_cl(ph.at[_pi, "Team"]) if "Team" in ph.columns else _orig))
            _ply = _cl(ph.at[_pi, "Player Picked"])
            _drafted = _ply.lower() not in ("", "unknown", "n/a", "nan", "none")
            _drafted_txt = "drafted " + (f"{_ply} " if _drafted else "") + f"({_num})"
            _draft_key = f"{_draft_anchor(_yr)} 12:00:00"          # draft (noon)
            # Commissioner moves sort SECOND, immediately after the "originally
            # …'s pick" header: they're off-platform moves that predate Sleeper's
            # recorded trades, so anywhere later reads out of order (user flag).
            _move_key = "0000-00-00 00:00:01"
            # The "originally …'s pick" header pins to the top of every comment
            # (sort key 0000) so all histories start the same way, ahead of even a
            # player's pre-draft free-agent stints.
            _is_209 = _num == "2.09"
            _is_5xx = bool(re.match(r"\s*5\.", _num))
            if "(vet)" in _yr_disp.lower():
                # The startup (vet) draft is the league's first event — anchor it
                # at the start of the year so it sorts ahead of every later trade.
                return [("0000-00-00 00:00:00", f"{_yr} {_num} — originally {_orig}'s pick"),
                        (f"{_yr}-01-01 00:00:00", f"{_yr} startup (vet) draft: {_final} {_drafted_txt}")], 0
            if _is_209 or _is_5xx:
                # 2.09 toilet reward / 5.xx FAAB buy: an off-platform AWARD, not a
                # trade. The 2.09 originates with the prior toilet-bracket winner
                # and is then commissioner-assigned to the team that drafts; the
                # 5.xx is bought by and stays with one team. Either way the award
                # itself does NOT count toward Number of trades.
                _hdr = f"{_yr} {_num} — originally {_orig}'s pick" + (" (toilet-bowl reward)" if _is_209 else "")
                _lines = [("0000-00-00 00:00:00", _hdr)]
                if _final and _final != _orig and _present(_final):
                    _lines.append((_move_key, f"{_yr}: Commissioner moved to {_final}"))
                _lines.append((_draft_key, f"{_yr} draft: {_final} {_drafted_txt}"))
                return _lines, 0
            _lines: List[Tuple[str, str]] = [("0000-00-00 00:00:00", f"{_yr} {_num} — originally {_orig}'s pick")]
            _nm = re.match(r"\s*(\d+)\.", _num)
            _covered: set = set()
            _ntr = 0
            if _nm:
                _key = (_yr, int(_nm.group(1)), _orig)
                _anchor_d = _draft_anchor(_yr)
                _same_day = 0
                for _ts, _recv, _line in sorted(_ph_hops.get(_key, [])):
                    _k = _ts
                    if _drafted and str(_ts)[:10] == _anchor_d:
                        # A pick can only be traded BEFORE it's used, so a trade
                        # on draft DAY always precedes the draft even when its
                        # clock time is after the noon draft anchor (e.g. the
                        # 2025 draft + its day-of pick trades). Clamp to 11:MM,
                        # preserving the trades' real order.
                        _same_day += 1
                        _k = f"{_anchor_d} 11:{_same_day:02d}:00"
                    _lines.append((_k, _line))
                    _covered.add(_recv); _ntr += 1
            # Untracked chain hops (no recorded Sleeper trade) = commissioner-applied
            # off-platform trades, beyond Sleeper's 3-yr window. They ARE trades, so
            # they count toward Number of trades and show in the history.
            _cm_n = 0
            for _tc in _ph_trade_cols:
                _own = _cl(ph.at[_pi, _tc])
                if _own and _own.lower() not in ("n/a", "nan") and _own != _orig and _own not in _covered:
                    _cm_n += 1
                    _lines.append((f"0000-00-00 00:00:{_cm_n:02d}", f"{_yr}: Commissioner moved to {_own}"))
                    _covered.add(_own); _ntr += 1
            # A future pick (not yet drafted) gets no draft line — the lineage
            # just runs to the latest move.
            if _drafted:
                _lines.append((_draft_key, f"{_yr} Draft: {_final} {_drafted_txt}"))
            return _lines, _ntr

        # (#change 4) Future picks: the draft ORDER isn't finalized, so the slot
        # on the pick's OWN row is just a guess off the owner's roster position.
        # Blank it to "R.??" (the Original Team column already identifies it;
        # embedded references elsewhere use "{yr} {round}(team)" via pick_lookup).
        # Dynamic — a drafted pick keeps its real slot.
        if not ph.empty and "Player Picked" in ph.columns:
            for _pi in ph.index:
                if _cl(ph.at[_pi, "Player Picked"]).lower() in ("", "unknown", "n/a", "nan", "none"):
                    _mr = re.match(r"\s*(\d+)", str(ph.at[_pi, "Number"]))
                    if _mr:
                        ph.at[_pi, "Number"] = f"{int(_mr.group(1))}.??"

        # Build each pick's lineage BEFORE dropping the Trade N columns it reads.
        _pick_lineage: Dict[int, List[Tuple[str, str]]] = {}
        _pick_ntrades: Dict[int, int] = {}
        _pid_to_pick_idx: Dict[str, int] = {}
        if not ph.empty:
            for _pi in ph.index:
                _lines, _nt = _pick_hist_lines(_pi)
                _pick_lineage[int(_pi)] = _lines
                _pick_ntrades[int(_pi)] = _nt
                _ppid = ph.at[_pi, "_player_id"] if "_player_id" in ph.columns else None
                if _ppid:
                    _pid_to_pick_idx.setdefault(str(_ppid), int(_pi))

        # Player history = the lineage of the pick they were drafted at (vet
        # included) + every later add / drop / trade, sorted by full timestamp.
        for _pid, _meta in pid_meta.items():
            _name = _cl(_meta.get("full_name"))
            if not _name:
                continue
            _seed = _pick_lineage.get(_pid_to_pick_idx.get(str(_pid), -1), [])
            _all = _seed + _pl_events.get(str(_pid), [])
            if _all:
                player_history_text[_name] = "\n".join(_l for _, _l in sorted(_all))

        # Pick comment = the drafted player's FULL history (it shouldn't stop at
        # the draft); an Unknown / not-yet-drafted pick falls back to its lineage.
        if not ph.empty:
            for _pi in ph.index:
                _ppid = ph.at[_pi, "_player_id"] if "_player_id" in ph.columns else None
                _pname = _cl(pid_meta.get(str(_ppid), {}).get("full_name")) if _ppid else ""
                if _pname and _pname in player_history_text:
                    pick_history_text[int(_pi)] = player_history_text[_pname]
                elif _pick_lineage.get(int(_pi)):
                    pick_history_text[int(_pi)] = "\n".join(_l for _, _l in sorted(_pick_lineage[int(_pi)]))
            # Add "Number of trades" and DROP the now-redundant Trade N chain cols.
            ph["Number of trades"] = [int(_pick_ntrades.get(int(_pi), 0)) for _pi in ph.index]
            if _ph_trade_cols:
                ph.drop(columns=_ph_trade_cols, inplace=True, errors="ignore")
        _log(debug, f"[{_now_iso()}] INFO built asset history: {len(pick_history_text)} picks, {len(player_history_text)} players")
    except Exception as e:
        _log_exc(debug, "asset_history_text", e)

    # Known-player validation report (best-effort): checks core columns against public expectations.
    try:
        validation = _known_player_column_errors(repo_root, pw)
        val_path = repo_root / "exports" / "raw" / "known_player_column_errors.csv"
        val_path.parent.mkdir(parents=True, exist_ok=True)
        validation.to_csv(val_path, index=False)
        if not validation.empty:
            _log(debug, f"[{_now_iso()}] WARN known-player validation mismatches: {len(validation)}")
    except Exception as e:
        _log_exc(debug, "known_player_validation", e)

    # ------------------------------------------------------------------
    # Standardize output row order across every sheet. Was previously
    # left to upstream-specific sorts (groupby default, sort-by-Sleeper-
    # Player-ID, etc), which were stable for identical inputs but
    # produced large row-position diffs whenever a new player joined
    # the data set or a Sleeper roster snapshot changed.
    #
    # Sort spec per sheet (none changes the cell data — only row order):
    #   player_week      Player(alpha)  Year, Week
    #   player_year      Player(alpha)  Year
    #   player_all_time  Player(alpha)
    #   team_week        Team(alpha)    Year, Week
    #   team_year        Team(alpha)    Year
    #   team_all_time    Team(alpha)
    #   league_week      Year, Week
    #   league_year      Year
    #   league_all_time  single row; no sort
    #   transactions     Team, Date     (already sorted upstream; preserved
    #                                    so Link to next/previous row
    #                                    references stay valid)
    #   trades           Team, Date     (same)
    #   pick_history     custom order from rebuild; preserved
    #
    # case-insensitive `key` lambda + kind="mergesort" for stability.
    def _ci_key(s):
        return s.astype(str).str.lower() if s.name in {"Player", "Team"} else s

    def _safe_sort(df, by):
        if df is None or len(df) == 0:
            return df
        keys = [k for k in by if k in df.columns]
        if not keys:
            return df
        try:
            return df.sort_values(by=keys, key=_ci_key, kind="mergesort").reset_index(drop=True)
        except Exception as e:
            _log_exc(debug, f"standardize_sort_{','.join(keys)}", e)
            return df

    pw           = _safe_sort(pw,           ["Player", "Year", "Week"])
    player_year  = _safe_sort(player_year,  ["Player", "Year"])
    player_all   = _safe_sort(player_all,   ["Player"])
    tw           = _safe_sort(tw,           ["Team", "Year", "Week"])
    team_year    = _safe_sort(team_year,    ["Team", "Year"])
    team_all     = _safe_sort(team_all,     ["Team"])
    league_week  = _safe_sort(league_week,  ["Year", "Week"])
    league_year  = _safe_sort(league_year,  ["Year"])
    # transactions / trades / pick_history left as-is (already sorted
    # by Team+Date upstream so the within-team row-index links remain
    # valid; pick_history uses its custom Year/Number ordering).

    # ------------------------------------------------------------------
    # Item 4: "O-Score" on picks / transactions / trades.
    # For each sheet, take 4 stats, convert each to its PERCENTILE (0-100)
    # across that sheet's rows, and average the four. The KTC component is the
    # percentile of the MOST RECENT populated KTC value/diff on the row. O-Score
    # is N/A unless all four components are present (per user). For picks, the
    # 2021 vet/startup draft is excluded entirely — from the stat AND from every
    # percentile pool — and its rows are N/A.
    # ------------------------------------------------------------------
    def _add_oscore(df: pd.DataFrame, stat_cols: List[str], ktc_cols: List[str],
                    exclude_mask: Optional[pd.Series] = None,
                    droppable: Tuple[str, ...] = ()) -> None:
        if df is None or df.empty:
            return
        # Most-recent populated KTC value (ktc_cols ordered latest → earliest).
        _mr = pd.Series(np.nan, index=df.index)
        for _c in ktc_cols:
            if _c in df.columns:
                _v = pd.to_numeric(df[_c], errors="coerce")
                _mr = _mr.where(_mr.notna(), _v)
        _series: List[pd.Series] = []
        for _sc in stat_cols:
            _s = _mr.copy() if _sc == "__MOST_RECENT_KTC__" else pd.to_numeric(df.get(_sc), errors="coerce")
            if exclude_mask is not None:
                _s = _s.where(~exclude_mask)
            _series.append(_s)
        # Percentile per stat: normal 'average' tie handling, EXCEPT a value of
        # exactly 0 (no production — e.g. the ~87% of adds that never started →
        # % of starts = 0) is pushed to the BOTTOM of its tie via the 'min' rank
        # instead of the middle. Non-zero ties keep the average (middle) rule.
        def _pct(_s):
            _avg = _s.rank(pct=True, method="average") * 100.0
            _min = _s.rank(pct=True, method="min") * 100.0
            return _avg.where(_s != 0, _min)
        _pcts = pd.concat([_pct(_s) for _s in _series], axis=1)
        # Average the AVAILABLE components, but a 'droppable' component (e.g. the
        # addition value of a player never rostered for a full week) may be
        # missing — every REQUIRED component must be present or the O-Score is
        # N/A. So picks with no on-team addition value still score off the other
        # three, while a missing required signal (retired/untracked → no KTC,
        # pure drop, one-sided untracked trade) → N/A.
        _req = [_i for _i, _sc in enumerate(stat_cols) if _sc not in droppable]
        _req_present = _pcts.iloc[:, _req].notna().all(axis=1)
        _osc = _pcts.mean(axis=1).where(_req_present)
        df["O-Score"] = _osc.round(1)

    try:
        _add_oscore(
            ph,
            ["Avg points added", "Pick-adjusted Difference in Player addition value", "__MOST_RECENT_KTC__",
             "Pick-adjusted Difference in Avg career PPG adjusted by position"],
            # KTC component is the PICK-ADJUSTED KTC difference (most-recent
            # populated), so a pick's market value is judged vs its draft-slot
            # window, not in absolute terms (points added stays absolute).
            ["Pick-adjusted Difference in KTC 4 years after draft day",
             "Pick-adjusted Difference in KTC 3 years after draft day",
             "Pick-adjusted Difference in KTC 2 years after draft day",
             "Pick-adjusted Difference in KTC 1 year after draft day",
             "Pick-adjusted Difference in KTC at end of rookie year",
             "Pick-adjusted Difference in KTC on draft day"],
            exclude_mask=(ph["Year"].astype(str).str.contains("vet") if (not ph.empty and "Year" in ph.columns) else None),
            droppable=("Pick-adjusted Difference in Player addition value",),
        )
        _add_oscore(
            tx,
            ["Avg net points", "Player addition value", "__MOST_RECENT_KTC__",
             "% of starts made while rostered"],
            ["KTC value of player added 2 years later", "KTC value of player added 1 year later",
             "KTC value of player added at end of season", "KTC value of player added at deal time"],
            # '% of starts made while rostered' is N/A for adds never rostered
            # a full week (weeks_played==0) — drop it so those adds still score
            # off net points / addition value (0) / KTC instead of going N/A.
            # Pure drops stay N/A (no added player -> no KTC) as intended.
            droppable=("% of starts made while rostered",),
        )
        _add_oscore(
            tr,
            ["Avg net points", "Trade addition value", "__MOST_RECENT_KTC__", "Trade impact score"],
            ["KTC value difference 2 years later", "KTC value difference 1 year later",
             "KTC value difference at end of season", "KTC value difference at deal time"],
        )
    except Exception as e:
        _log_exc(debug, "oscore_item4", e)

    # ----- O-Score for PURE DROPS (drop-only transactions) -----
    # Scored in a SEPARATE universe: percentile ranks computed only against the
    # other pure drops, averaged over the available components, then DIVIDED BY
    # TWO so the ceiling is 50 (a drop can never out-score a good add).
    # Components (higher = better):
    #   1) most recent populated Net KTC value (2yr -> 1yr -> end-of-season ->
    #      deal time), with 0 filled in when the row has no populated net KTC
    #      at all (untracked dropped player)
    #   2) Dropped avg points   (negated post-drop PPG; 0 = never played again)
    #   3) Dropped total points (negated post-drop total over the same window)
    #   4) Player addition value (the composite already on the row)
    # No zero-to-bottom tie rule here: a 0 in the dropped-points columns is the
    # BEST outcome (the player never played again), not "no production".
    # Placed BEFORE the manager-skill aggregation so these scores feed the
    # Transaction skill at 1/3 weight (see _tx_pure below); every NON-pure-drop
    # O-Score is left exactly as computed above.
    _tx_pure = None    # boolean mask of pure-drop tx rows, reused by Transaction skill
    try:
        if tx is not None and not tx.empty and "O-Score" in tx.columns:
            def _blank(_c):
                # True where the cell is empty / N/A — robust to "", None, NaN,
                # "N/A", "nan". Missing column -> treat every row as blank.
                if _c not in tx.columns:
                    return pd.Series(True, index=tx.index)
                _s = tx[_c].astype("string").str.strip().str.lower()
                return _s.isna() | _s.isin(["", "n/a", "nan", "none"])
            # A pure drop is an add-nobody / drop-somebody row. Use the internal
            # sleeper-id columns as the source of truth (present pre-output) and
            # fall back to the display names, so this can't be fooled by however
            # the empty add side happens to be rendered on a given build.
            _add_blank = _blank("_added_pid") if "_added_pid" in tx.columns else _blank("Player Added")
            _drop_set = ~(_blank("_dropped_pid") if "_dropped_pid" in tx.columns else _blank("Player Dropped"))
            _pure = (_add_blank & _drop_set).fillna(False)
            _tx_pure = _pure
            _n_pure = int(_pure.sum())
            _log(debug, f"[{_now_iso()}] INFO pure-drop O-Score: {_n_pure} drop-only txns detected "
                        f"(cols: _added_pid={'_added_pid' in tx.columns}, Player Added={'Player Added' in tx.columns})")
            if _n_pure:
                _mr = pd.Series(np.nan, index=tx.index)
                for _c in ["Net KTC value 2 years later", "Net KTC value 1 year later",
                           "Net KTC value at end of season", "Net KTC value at deal time"]:
                    _v = pd.to_numeric(tx.get(_c), errors="coerce")
                    _mr = _mr.where(_mr.notna(), _v)
                _comps = [
                    _mr.fillna(0.0),
                    pd.to_numeric(tx.get("Dropped avg points"), errors="coerce"),
                    pd.to_numeric(tx.get("Dropped total points"), errors="coerce"),
                    pd.to_numeric(tx.get("Player addition value"), errors="coerce"),
                ]
                _pcts = pd.concat(
                    [(_s.where(_pure)).rank(pct=True, method="average") * 100.0 for _s in _comps],
                    axis=1)
                _osc_drop = (_pcts.mean(axis=1) / 2.0).round(1)
                tx.loc[_pure, "O-Score"] = _osc_drop[_pure]
                _log(debug, f"[{_now_iso()}] INFO pure-drop O-Score: scored {_n_pure} drop-only transactions (ceiling 50)")
    except Exception as e:
        _log_exc(debug, "oscore_pure_drops", e)

    # ----- Manager skill: O-Score aggregates (drafting / trading / transaction) -----
    # Per (Team, Year) and per Team all-time, the SAMPLE-SIZE-SHRUNK mean O-Score
    # of the picks the team MADE, the trades it was in, and the transactions it
    # made. Shrinkage toward the league-neutral 50 via (Σw·o + K·50)/(Σw + K),
    # K=5: a manager with 2 great moves can't out-rank one with 25 solid ones,
    # and inactive managers sit near neutral instead of being over-rewarded.
    # N/A for a (team, year) with no events of that type. Picks use Final Team
    # (the drafter); trades/transactions use Team + Season; vet picks carry an
    # N/A O-Score and drop out of the mean automatically. Transaction skill now
    # INCLUDES pure-drop txns, but each counts 1/3 as much as an add/swap (weight
    # 1/3 vs 1) so a flurry of drops can't dominate the score.
    try:
        _SKILL_K, _SKILL_PRIOR = 5.0, 50.0

        def _skill_maps(_df, _team_col, _year_col, _weights=None):
            """-> ({(team, year): skill}, {team: skill}) of weighted shrunk-mean
            O-Scores. _weights: optional per-row weight Series (default 1.0)."""
            if _df is None or _df.empty or "O-Score" not in _df.columns or _team_col not in _df.columns:
                return {}, {}
            _wt = (pd.to_numeric(_weights, errors="coerce").reindex(_df.index).fillna(1.0)
                   if _weights is not None else pd.Series(1.0, index=_df.index))
            _w = pd.DataFrame({
                "_tm": _df[_team_col].astype(str),
                "_o": pd.to_numeric(_df["O-Score"], errors="coerce"),
                "_yr": _df[_year_col].astype(str).str.extract(r"(\d{4})")[0] if _year_col in _df.columns else None,
                "_wt": _wt.astype(float),
            }).dropna(subset=["_o"])
            _w["_wo"] = _w["_wt"] * _w["_o"]
            def _shrink(g):
                # weighted shrinkage: (Σ w·o + K·prior) / (Σ w + K)
                return ((g["_wo"] + _SKILL_K * _SKILL_PRIOR) / (g["_wt"] + _SKILL_K)).round(1)
            _ga = _w.groupby("_tm")[["_wo", "_wt"]].sum()
            all_map = _shrink(_ga).to_dict()
            year_map = {}
            if _w["_yr"].notna().any():
                _gy = _w.dropna(subset=["_yr"]).groupby(["_tm", "_yr"])[["_wo", "_wt"]].sum()
                for (tm, yr), v in _shrink(_gy).items():
                    year_map[(tm, int(yr))] = v
            return year_map, all_map

        # Transactions: pure drops (drop-only rows) count 1/3 as much as adds/swaps.
        _tx_wt = None
        if _tx_pure is not None and tx is not None and not tx.empty:
            _tx_wt = pd.Series(1.0, index=tx.index)
            _tx_wt[_tx_pure.reindex(tx.index).fillna(False)] = 1.0 / 3.0

        _skill_specs = [
            ("Drafting skill", ph, "Final Team", "Year", None),
            ("Trading skill", tr, "Team", "Season", None),
            ("Transaction skill", tx, "Team", "Season", _tx_wt),
        ]
        for _name, _df, _tc, _yc, _wts in _skill_specs:
            _ymap, _amap = _skill_maps(_df, _tc, _yc, _wts)
            if not team_year.empty and {"Team", "Year"}.issubset(team_year.columns):
                _yrs = pd.to_numeric(team_year["Year"], errors="coerce")
                team_year[_name] = [
                    (_ymap.get((str(t), int(y))) if pd.notna(y) else None)
                    for t, y in zip(team_year["Team"], _yrs)
                ]
            if not team_all.empty and "Team" in team_all.columns:
                team_all[_name] = [_amap.get(str(t)) for t in team_all["Team"]]
    except Exception as e:
        _log_exc(debug, "manager_skill", e)

    # ----- PR B: All-play win % + Losses from hardship (team_year + all_time) -----
    # All-play = score the team vs EVERY other team each week (schedule-luck-free):
    # win % = (Σ teams with strictly lower PF) / (Σ other teams). Losses from
    # hardship = count of team_week 'Loss from hardship?' flags.
    try:
        _ay_w, _ay_g = defaultdict(float), defaultdict(float)   # (team, year)
        _at_w, _at_g = defaultdict(float), defaultdict(float)   # team all-time
        _lh_y, _lh_t = defaultdict(int), defaultdict(int)       # losses-from-hardship
        if not tw.empty and {"Team", "Year", "Week", "PF"}.issubset(tw.columns):
            _t = tw.copy()
            _t["_pf"] = pd.to_numeric(_t["PF"], errors="coerce")
            _t["_yr"] = pd.to_numeric(_t["Year"], errors="coerce")
            _t = _t.dropna(subset=["_pf", "_yr"])
            _lhflag = (_t["Loss from hardship?"].astype(str).str.lower().isin(["true", "1", "yes"])
                       if "Loss from hardship?" in _t.columns else pd.Series(False, index=_t.index))
            for (_yr, _wk), _grp in _t.groupby(["_yr", "Week"]):
                _pfs = _grp["_pf"].tolist()
                _n = len(_grp)
                for _i2, _r in _grp.iterrows():
                    _tm, _y = str(_r["Team"]), int(_yr)
                    _w = float(sum(1 for v in _pfs if v < _r["_pf"]))
                    _ay_w[(_tm, _y)] += _w; _ay_g[(_tm, _y)] += float(_n - 1)
                    _at_w[_tm] += _w; _at_g[_tm] += float(_n - 1)
                    if bool(_lhflag.loc[_i2]):
                        _lh_y[(_tm, _y)] += 1; _lh_t[_tm] += 1
        def _ap_diff(_ap_list, _winpct_series):
            _wp = pd.to_numeric(_winpct_series, errors="coerce")
            return [(round(a - w, 4) if (a is not None and pd.notna(w)) else None)
                    for a, w in zip(_ap_list, _wp)]
        if not team_year.empty and {"Team", "Year"}.issubset(team_year.columns):
            _yrs = pd.to_numeric(team_year["Year"], errors="coerce")
            _ap = [(round(_ay_w[(str(t), int(y))] / _ay_g[(str(t), int(y))], 4)
                    if pd.notna(y) and _ay_g.get((str(t), int(y)), 0) > 0 else None)
                   for t, y in zip(team_year["Team"], _yrs)]
            team_year["All-play win %"] = _ap
            team_year["All-play win % minus Win %"] = _ap_diff(_ap, team_year.get("Win %"))
            team_year["Losses from hardship"] = pd.array(
                [(int(_lh_y.get((str(t), int(y)), 0))
                  if pd.notna(y) and _ay_g.get((str(t), int(y)), 0) > 0 else None)
                 for t, y in zip(team_year["Team"], _yrs)], dtype="Int64")
        if not team_all.empty and "Team" in team_all.columns:
            _apa = [(round(_at_w[str(t)] / _at_g[str(t)], 4) if _at_g.get(str(t), 0) > 0 else None)
                    for t in team_all["Team"]]
            team_all["All-play win %"] = _apa
            team_all["All-play win % minus Win %"] = _ap_diff(_apa, team_all.get("All time win %"))
            team_all["Losses from hardship"] = pd.array(
                [(int(_lh_t.get(str(t), 0)) if _at_g.get(str(t), 0) > 0 else None)
                 for t in team_all["Team"]], dtype="Int64")
    except Exception as e:
        _log_exc(debug, "allplay_losses_hardship", e)

    # Convert every timestamp column in the dataset to US Eastern (DST-aware)
    # for display — done LAST, after all date-based logic and sorting, so it's
    # purely cosmetic. These three are the only time-of-day columns; all other
    # date columns are date-only or numeric.
    try:
        if not tx.empty and "Date" in tx.columns:
            tx["Date"] = _to_eastern_display(tx["Date"])
        if not tx.empty and "Date dropped/traded" in tx.columns:
            tx["Date dropped/traded"] = _to_eastern_display(tx["Date dropped/traded"])
        if not tr.empty and "Date" in tr.columns:
            tr["Date"] = _to_eastern_display(tr["Date"])
    except Exception as e:
        _log_exc(debug, "eastern_time_convert", e)

    # Display rename: picks 'Final Team' -> 'Team'. Kept as 'Final Team'
    # internally (all reconstruction / lookup logic above uses that name);
    # renamed here, after every consumer has run, so only the output column
    # changes. The plan/catalog list it as 'Team' and order it right after
    # 'Player Picked' (which the picks freeze pins, alongside Year/Number).
    try:
        if not ph.empty and "Final Team" in ph.columns and "Team" not in ph.columns:
            ph = ph.rename(columns={"Final Team": "Team"})
    except Exception as e:
        _log_exc(debug, "picks_final_team_rename", e)

    context = {
        "player_week": pw,
        "player_year": player_year,
        "player_all_time": player_all,
        "team_week": tw,
        "team_year": team_year,
        "team_all_time": team_all,
        "league_week": league_week,
        "league_year": league_year,
        "league_all_time": league_all,
        "transactions": tx,
        "trades": tr,
        "pick_history": ph,
    }
    tables = [
        (doc.FILE_NAME, doc.build_output(context), doc.PLAN_KEY)
        for doc in DOCUMENT_MODULES
    ]
    write_outputs(tables)

    # Phase 12 #45: build-time data-quality log. Run the same sanity checks the
    # committed test (tests/test_sanity_ranges.py) uses over the just-written
    # CSVs and record a one-line summary + every ERROR/WARN into build_debug.log,
    # so range anomalies / N/A-vs-0 regressions surface in every build's log
    # without a manual audit pass.
    try:
        from lotg_support.sanity import collect_findings, summarize
        _sf = collect_findings(repo_root / "exports")
        _log(debug, f"[{_now_iso()}] INFO {summarize(_sf)}")
        for _f in _sf:
            _log(debug, f"[{_now_iso()}] {_f.severity} sanity {_f.sheet}.{_f.column}: {_f.detail}")
    except Exception as e:
        _log_exc(debug, "data_quality_sanity_log", e)


def main() -> None:
    repo_root = Path(__file__).resolve().parent.parent

    try:
        cfg = yaml.safe_load((repo_root / "config/league.yaml").read_text())
        league_id = str(cfg["league_id"])
        min_season = cfg.get("min_season")
        max_season = cfg.get("max_season")

        mode = str(os.environ.get("LOTG_MODE", "both")).lower().strip()
        if mode not in {"snapshot", "build", "both"}:
            mode = "both"

        if mode in {"snapshot", "both"}:
            from lotg_support.snapshot import snapshot_all
            try:
                snapshot_all(repo_root, league_id=league_id, min_season=min_season, max_season=max_season)
            except Exception as e:
                _fatal_log(repo_root, "snapshot_all", e)
                raise

        if mode in {"build", "both"}:
            try:
                build_all(repo_root)
            except Exception as e:
                _fatal_log(repo_root, "build_all", e)
                raise
    except Exception as e:
        _fatal_log(repo_root, "main", e)
        raise


if __name__ == "__main__":
    try:
        main()
    except Exception:
        raise
